"""School OS and smart timetable routes extracted from the legacy monolith.

The public URLs and business logic are preserved.  Definitions are patched back
into samtm_platform so older route functions that resolve late-bound helpers keep
the same behaviour they had in one file.

V19.8: parallel guruh xonasi yozilmagan bo'lsa ham jadval yaratiladi; xona
keyin frontendda "Xona yo'q" holatida tahrirlanadi. Yakuniy generator repair
bosqichi sinf ichki oknosini yopadi, kunlik soatlarni tenglashtiradi va J/T
hamda texnologiyani imkon qadar 3–6-darsga xavfsiz almashtiradi.
"""
try:
    from . import samtm_platform as _platform
    from .samtm_platform import *
except ImportError:  # Railway working directory may be backend/
    import samtm_platform as _platform
    from samtm_platform import *

import copy as _samtm_copy
import time as _samtm_time
import threading as _samtm_threading
import base64 as _v237_base64
import binascii as _v237_binascii
import io as _v237_io
import math as _v237_math
import unicodedata as _v237_unicodedata
import zipfile as _v237_zipfile

# V22.0 exact solver alohida modulda saqlanadi. Modul importi xavfsiz, ammo
# jadval endpointi OR-Tools o'rnatilmagan muhitda eski greedy generatorga
# yashirincha qaytmaydi: bitta kanonik generator — exact CP-SAT. Shu orqali
# deploydagi yetishmagan dependency darhol aniq ko'rinadi va eski/qisman
# natija yangi generatorniki deb saqlanib qolmaydi.
try:
    from .samtm_exact_timetable import (
        DefaultTimetableAdapter as _V220DefaultTimetableAdapter,
        EXACT_SOLVER_RELEASE as _V230_EXACT_MODULE_RELEASE,
        ORTOOLS_AVAILABLE as _V216_ORTOOLS_AVAILABLE,
        candidate_hard_violations as _v220_candidate_hard_violations,
        solve_exact_timetable as _v216_solve_exact,
        validate_timetable_placements as _v218_validate_placements,
    )
except ImportError:  # Railway working directory may be backend/
    try:
        from samtm_exact_timetable import (
            DefaultTimetableAdapter as _V220DefaultTimetableAdapter,
            EXACT_SOLVER_RELEASE as _V230_EXACT_MODULE_RELEASE,
            ORTOOLS_AVAILABLE as _V216_ORTOOLS_AVAILABLE,
            candidate_hard_violations as _v220_candidate_hard_violations,
            solve_exact_timetable as _v216_solve_exact,
            validate_timetable_placements as _v218_validate_placements,
        )
    except ImportError as _v216_exact_import_error:
        _V216_ORTOOLS_AVAILABLE = False
        _V220DefaultTimetableAdapter = None
        _V230_EXACT_MODULE_RELEASE = "SAMTM-EXACT-MODULE-UNAVAILABLE"
        _v220_candidate_hard_violations = None
        _v216_solve_exact = None
        _v218_validate_placements = None
        _V216_EXACT_IMPORT_ERROR = str(_v216_exact_import_error)
    else:
        _V216_EXACT_IMPORT_ERROR = None
else:
    _V216_EXACT_IMPORT_ERROR = None

# V23 runtime va engine alohida kichik modullarda saqlanadi. Railway main.py
# backendni top-level modul sifatida yuklagani uchun relative va absolute import
# yo'llarining ikkalasi ham qo'llab-quvvatlanadi.
try:
    from .samtm_schedule_runtime import (
        Progress as _V230Progress,
        RUNTIME_RELEASE as _V234_RUNTIME_RELEASE,
        RuntimePolicy as _V230RuntimePolicy,
        ScheduleRuntime as _V230ScheduleRuntime,
        Stage as _V230Stage,
    )
    from . import samtm_timetable_engine as _timetable_engine
except ImportError:
    from samtm_schedule_runtime import (
        Progress as _V230Progress,
        RUNTIME_RELEASE as _V234_RUNTIME_RELEASE,
        RuntimePolicy as _V230RuntimePolicy,
        ScheduleRuntime as _V230ScheduleRuntime,
        Stage as _V230Stage,
    )
    import samtm_timetable_engine as _timetable_engine

SAMTM_TIMETABLE_ENGINE_RELEASE = getattr(
    _timetable_engine, "ENGINE_RELEASE", "SAMTM-TIMETABLE-ENGINE-V23"
)
_timetable_filter_reasons = _timetable_engine.filter_reasons
_timetable_internal_policy = _timetable_engine.internal_policy
_timetable_mode_config = _timetable_engine.mode_config
_timetable_public_modes = _timetable_engine.public_modes
_timetable_stage_label = _timetable_engine.stage_label

# ``from samtm_platform import *`` Python qoidasiga ko'ra nomi ``_`` bilan
# boshlanadigan yordamchilarni import qilmaydi. Maktab kodi esa eski monolitdagi
# shu ichki yordamchilardan ham foydalanadi. Dunder metama'lumotlarni tegmasdan,
# platformadagi barcha oddiy va private nomlarni lokal namespace'ga ulaymiz.
# setdefault maktab modulining o'z ta'riflarini keyin xavfsiz ustun qo'yadi.
for _platform_name, _platform_value in vars(_platform).items():
    if not _platform_name.startswith("__"):
        globals().setdefault(_platform_name, _platform_value)

_V19_IMPORTED_NAMES = set(globals())

# V19.8 public compatibility belgisi ``main.py`` va amaldagi frontendda
# qattiq tekshiriladi. V23.7 funksiyalari package revision/app.version orqali
# ajratiladi; bu public qiymatni o'zgartirish Railway worker bootini to'xtatadi.
SAMTM_SCHOOL_RELEASE = "samtm-school-workspace-link-v19.8"
SAMTM_JADVAL_RELEASE = "JADVAL-ONE-V23.6-LIVE-REATTACH-ALL-TEACHERS"
# Eski frontend aynan V22.0 satrini qattiq tekshiradi. Public compatibility
# qiymati o'zgarmaydi; real algoritm versiyasi alohida qaytariladi.
SAMTM_EXACT_JADVAL_RELEASE = "SAMTM-EXACT-CP-SAT-V22.0"
SAMTM_EXACT_INTERNAL_RELEASE = "SAMTM-EXACT-CP-SAT-V23.6-DUAL-SHIFT-WEEK-ROUND-ROBIN"
# ``main.py`` REV55 ni ham qat'iy tekshiradi. Bu ikkinchi public handshake
# o'zgarmaydi; yangi imkoniyatlar alohida feature revision bilan belgilanadi.
SAMTM_SCHOOL_PACKAGE_REVISION = "multi-school-access-2month-rev55"
SAMTM_SCHOOL_FEATURE_REVISION = "teacher-grid-partial-group-class-v23.7.2"
_platform.SAMTM_RELEASE = SAMTM_SCHOOL_RELEASE
_platform.SAMTM_PACKAGE_REVISION = SAMTM_SCHOOL_PACKAGE_REVISION
_platform.SAMTM_FEATURE_REVISION = SAMTM_SCHOOL_FEATURE_REVISION
try:
    # FastAPI public app version ham eski main/deploy handshake bilan mos qoladi.
    # Real feature build yuqoridagi SAMTM_SCHOOL_FEATURE_REVISION da ko'rsatilgan.
    app.version = "19.8"
except Exception:
    pass


def _v220_generation_budget_seconds():
    """Texnik xavfsizlik chegarasi; oddiy ishni faqat foydalanuvchi to‘xtatadi."""
    try:
        value = float(os.getenv("SAMTM_JADVAL_GENERATION_BUDGET_SECONDS", "604800"))
    except (TypeError, ValueError):
        value = 604800.0
    return max(3600.0, min(2592000.0, value))

def _sinf_guruh_soni_normalizatsiya(usul, guruh_soni):
    """Guruh usuliga mos 1–4 oralig'idagi haqiqiy guruh sonini qaytaradi.

    Bu helper eski monolitda maktab modulidan oldin ta'riflangan edi. Modulga
    ajratishda ta'rifi tushib qolib, yuklama matritsasida NameError bergan.
    """
    usul = (usul or "none").strip().lower()
    if usul == "none":
        return 1
    if usul == "gender":
        return 2
    try:
        soni = int(guruh_soni or 2)
    except (TypeError, ValueError) as error:
        raise HTTPException(
            status_code=400,
            detail="Guruh soni raqam bo'lishi kerak",
        ) from error
    if soni not in (2, 3, 4):
        raise HTTPException(
            status_code=400,
            detail="Guruh soni 2, 3 yoki 4 bo'lishi kerak",
        )
    return soni

# ═══════════════════════════════════════════════════════════
# V18.45 — AQILLI MAKTAB BOSH SAHIFASI / O'QITUVCHI BUGUNI / YUKLAMA
# Kundalikning o'rnini bosmaydi: bu yerda ichki monitoring, yordamchi
# signal va rejalashtirish ishlaydi.
# ═══════════════════════════════════════════════════════════

def _v1845_smart_school_tables(cur):
    cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS haftalik_dars_soati INTEGER")
    _rejalashtirish_jadvallari(cur)
    cur.execute("""CREATE TABLE IF NOT EXISTS dars_monitoring_baholari(
        id BIGSERIAL PRIMARY KEY,
        maktab_id INTEGER NOT NULL REFERENCES maktablar(id) ON DELETE CASCADE,
        sinf_id INTEGER NOT NULL REFERENCES maktab_sinflari(id) ON DELETE CASCADE,
        oquvchi_user_id BIGINT NOT NULL REFERENCES users(user_id) ON DELETE CASCADE,
        oqituvchi_user_id BIGINT NOT NULL REFERENCES users(user_id) ON DELETE CASCADE,
        sana DATE NOT NULL DEFAULT CURRENT_DATE,
        fan TEXT NOT NULL,
        mavzu TEXT,
        foiz INTEGER NOT NULL,
        izoh TEXT,
        yaratilgan_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
        CHECK (foiz BETWEEN 0 AND 100)
    )""")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_monitoring_bola_fan_sana ON dars_monitoring_baholari(oquvchi_user_id, fan, sana DESC)")
    cur.execute("""CREATE TABLE IF NOT EXISTS aqlli_holatlar(
        id BIGSERIAL PRIMARY KEY,
        maktab_id INTEGER NOT NULL REFERENCES maktablar(id) ON DELETE CASCADE,
        oquvchi_user_id BIGINT NOT NULL REFERENCES users(user_id) ON DELETE CASCADE,
        turi TEXT NOT NULL,
        daraja INTEGER NOT NULL DEFAULT 1,
        sarlavha TEXT NOT NULL,
        tavsif TEXT,
        masul_rol TEXT,
        holat TEXT NOT NULL DEFAULT 'ochiq',
        manba TEXT,
        manba_sana DATE,
        yaratilgan_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
        yangilangan_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
    )""")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_aqlli_holat_maktab_holat ON aqlli_holatlar(maktab_id, holat, daraja DESC)")


def _v1845_joriy_chorak(sana):
    from datetime import date
    if isinstance(sana, str):
        sana = date.fromisoformat(sana)
    y, m = sana.year, sana.month
    if m in (9, 10):
        return date(y, 9, 1), date(y, 10, 31), 1
    if m in (11, 12):
        return date(y, 11, 1), date(y, 12, 31), 2
    if m in (1, 2, 3):
        return date(y, 1, 1), date(y, 3, 31), 3
    if m in (4, 5):
        return date(y, 4, 1), date(y, 5, 31), 4
    return date(y, 6, 1), date(y, 8, 31), 0


def _v1845_davomat_holat_yangila(cur, maktab_id, bola_id, sana):
    """Jarima qo'llamaydi va tashqi tashkilotga avtomatik shaxsiy ma'lumot yubormaydi.
    Faqat maktab ichida kuzatuv darajasini oshiradi; tashqi hamkor kerak bo'lsa
    vakolatli xodim holatni alohida ko'rib chiqadi."""
    _v1845_smart_school_tables(cur)
    bosh, oxir, chorak = _v1845_joriy_chorak(sana)
    cur.execute("""
        SELECT COUNT(*) AS son FROM davomat
        WHERE user_id=%s AND holat='kelmadi' AND sana BETWEEN %s AND %s
    """, (bola_id, bosh, oxir))
    son = int((cur.fetchone() or {}).get("son") or 0)
    if son <= 0:
        return None
    if son == 1:
        daraja, masul, sarlavha = 1, "sinf_rahbari", "Birinchi sababsiz davomat signali"
    elif son == 2:
        daraja, masul, sarlavha = 2, "psixolog_va_manaviyat", "Takroriy davomat — maktab ichki kuzatuvi"
    elif son == 3:
        daraja, masul, sarlavha = 3, "direktor", "Davomat bo'yicha rahbariyat ko'rigi kerak"
    else:
        daraja, masul, sarlavha = 4, "vakolatli_hamkor_korigi", "Surunkali davomat — vakolatli ko'rib chiqish tavsiyasi"
    cur.execute("""
        SELECT id FROM aqlli_holatlar
        WHERE maktab_id=%s AND oquvchi_user_id=%s AND turi='davomat'
          AND holat='ochiq' AND manba_sana BETWEEN %s AND %s
        ORDER BY id DESC LIMIT 1
    """, (maktab_id, bola_id, bosh, oxir))
    bor = cur.fetchone()
    tavsif = f"{chorak or '-'}-chorak davrida {son} marta sababsiz kelmagan. Avtomatik jarima qo'llanmaydi; mas'ul vaziyatni ko'rib chiqadi."
    if bor:
        cur.execute("""
            UPDATE aqlli_holatlar SET daraja=%s,sarlavha=%s,tavsif=%s,masul_rol=%s,
                manba_sana=%s,yangilangan_at=NOW() WHERE id=%s
        """, (daraja, sarlavha, tavsif, masul, sana, bor["id"]))
        return bor["id"]
    cur.execute("""
        INSERT INTO aqlli_holatlar(maktab_id,oquvchi_user_id,turi,daraja,sarlavha,tavsif,masul_rol,manba,manba_sana)
        VALUES(%s,%s,'davomat',%s,%s,%s,%s,'davomat',%s) RETURNING id
    """, (maktab_id, bola_id, daraja, sarlavha, tavsif, masul, sana))
    return cur.fetchone()["id"]


class V1845MonitoringYozuvi(BaseModel):
    oquvchi_user_id: int
    foiz: int
    izoh: Optional[str] = None


class V1845MonitoringSorov(BaseModel):
    token: str
    sinf_id: int
    fan: str
    mavzu: Optional[str] = None
    sana: Optional[str] = None
    yozuvlar: list[V1845MonitoringYozuvi]


@app.post("/api/oqituvchi/kunlik_monitoring")
def v1845_kunlik_monitoring(sorov: V1845MonitoringSorov):
    teacher_id = _jwt_tekshir(sorov.token)
    conn = _db(); cur = conn.cursor()
    _v1845_smart_school_tables(cur); _xodim_sinf_birikmalari_jadvali(cur)
    cur.execute("SELECT maktab_id FROM maktab_sinflari WHERE id=%s", (sorov.sinf_id,))
    sinf = cur.fetchone()
    if not sinf:
        cur.close(); conn.close(); raise HTTPException(status_code=404, detail="Sinf topilmadi")
    ruxsat = _maktab_boshqaruvchi_mi(cur, teacher_id, sinf["maktab_id"])
    if not ruxsat:
        cur.execute("""
            SELECT 1 FROM maktab_dars_birikmalari
            WHERE user_id=%s AND sinf_id=%s AND LOWER(TRIM(fan_nomi))=LOWER(TRIM(%s)) LIMIT 1
        """, (teacher_id, sorov.sinf_id, sorov.fan))
        ruxsat = cur.fetchone() is not None
    if not ruxsat:
        cur.close(); conn.close(); raise HTTPException(status_code=403, detail="Bu sinf/fan monitoringini kiritishga ruxsat yo'q")
    sana = sorov.sana or datetime.now().date().isoformat()
    saqlandi = 0
    for y in sorov.yozuvlar:
        if not 0 <= int(y.foiz) <= 100:
            cur.close(); conn.close(); raise HTTPException(status_code=400, detail="Monitoring foizi 0..100 bo'lishi kerak")
        cur.execute("""
            INSERT INTO dars_monitoring_baholari(maktab_id,sinf_id,oquvchi_user_id,oqituvchi_user_id,sana,fan,mavzu,foiz,izoh)
            VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (sinf["maktab_id"], sorov.sinf_id, y.oquvchi_user_id, teacher_id, sana, sorov.fan.strip(), sorov.mavzu, int(y.foiz), y.izoh))
        saqlandi += 1
    conn.commit(); cur.close(); conn.close()
    return {"holat": "saqlandi", "soni": saqlandi}


@app.get("/api/oqituvchi/bugun")
def v1845_oqituvchi_bugun(token: str):
    teacher_id = _jwt_tekshir(token)
    from zoneinfo import ZoneInfo
    hozir = datetime.now(ZoneInfo("Asia/Tashkent"))
    kun = hozir.isoweekday()
    joriy_hafta_turi = "toq" if int(hozir.isocalendar().week) % 2 else "juft"
    conn = _db(); cur = conn.cursor()
    _v1845_smart_school_tables(cur); _xodim_sinf_birikmalari_jadvali(cur)
    pass  # V19: DDL moved to startup migration.
    pass  # V19: DDL moved to startup migration.
    cur.execute("SELECT full_name,maktab_id,haftalik_dars_soati,kundalik_baho_eslatmasi FROM users WHERE user_id=%s", (teacher_id,))
    teacher = cur.fetchone()
    if not teacher:
        cur.close(); conn.close(); raise HTTPException(status_code=404, detail="O'qituvchi topilmadi")
    darslar = []
    hafta_darsi = 0
    cur.execute("SELECT to_regclass('public.aqlli_jadval_urinishlari_v2') AS table_name")
    smart_bor = bool((cur.fetchone() or {}).get("table_name"))
    smart_run = None
    if smart_bor and teacher.get("maktab_id"):
        cur.execute("""SELECT id FROM aqlli_jadval_urinishlari_v2
                       WHERE maktab_id=%s AND holat='tasdiqlangan' ORDER BY id DESC LIMIT 1""",
                    (teacher["maktab_id"],))
        smart_run = cur.fetchone()
    if smart_run:
        cur.execute("""SELECT e.id,e.dars_raqami,e.fan_nomi AS fan,
                              COALESCE(r.nomi,e.xona_matni) AS xona,
                              e.boshlanish_vaqti,e.tugash_vaqti,
                              s.id AS sinf_id,s.sinf,s.harf,e.guruh_kaliti,e.smena,e.hafta_turi
                       FROM aqlli_jadval_slotlari_v2 e
                       JOIN maktab_sinflari s ON s.id=e.sinf_id
                       LEFT JOIN aqlli_xonalar_v2 r ON r.id=e.xona_id
                       WHERE e.urinish_id=%s AND e.hafta_kuni=%s AND e.oqituvchi_user_id=%s
                         AND e.hafta_turi IN ('har_hafta',%s)
                       ORDER BY e.smena,e.dars_raqami,s.sinf::int,s.harf,e.guruh_kaliti""",
                    (smart_run["id"], kun, teacher_id, joriy_hafta_turi))
        darslar = cur.fetchall()
        cur.execute("""SELECT COUNT(*) AS hafta_darsi FROM aqlli_jadval_slotlari_v2
                       WHERE urinish_id=%s AND oqituvchi_user_id=%s
                         AND hafta_turi IN ('har_hafta',%s)""",
                    (smart_run["id"], teacher_id, joriy_hafta_turi))
        hafta_darsi = int((cur.fetchone() or {}).get("hafta_darsi") or 0)
    else:
        cur.execute("""
            SELECT DISTINCT j.id,j.dars_raqami,j.fan,j.xona,j.boshlanish_vaqti,j.tugash_vaqti,
                   s.id AS sinf_id,s.sinf,s.harf,j.guruh_kaliti,s.sinf::int AS sinf_sort
            FROM dars_jadvali j
            JOIN maktab_sinflari s ON s.id=j.sinf_id
            LEFT JOIN maktab_dars_birikmalari b
              ON b.sinf_id=j.sinf_id AND LOWER(TRIM(b.fan_nomi))=LOWER(TRIM(j.fan)) AND b.user_id=%s
            WHERE j.kun=%s AND (j.oqituvchi_user_id=%s OR (j.oqituvchi_user_id IS NULL AND b.user_id=%s))
            ORDER BY j.dars_raqami,sinf_sort,s.harf
        """, (teacher_id, kun, teacher_id, teacher_id))
        darslar = cur.fetchall()
        cur.execute("""
            SELECT COUNT(DISTINCT j.id) AS hafta_darsi FROM dars_jadvali j
            LEFT JOIN maktab_dars_birikmalari b
              ON b.sinf_id=j.sinf_id AND LOWER(TRIM(b.fan_nomi))=LOWER(TRIM(j.fan)) AND b.user_id=%s
            WHERE j.oqituvchi_user_id=%s OR (j.oqituvchi_user_id IS NULL AND b.user_id=%s)
        """, (teacher_id, teacher_id, teacher_id))
        hafta_darsi = int((cur.fetchone() or {}).get("hafta_darsi") or 0)
    kundalik_eslatma_yoqilgan = bool(teacher.get("kundalik_baho_eslatmasi"))
    cur.close(); conn.close()
    kundalik_eslatma = bool(kundalik_eslatma_yoqilgan and darslar and hozir.hour >= 15)
    return {
        "sana": hozir.date().isoformat(), "kun": kun, "hafta_turi": joriy_hafta_turi,
        "oqituvchi": teacher["full_name"],
        "darslar": darslar, "haftalik_reja_soati": teacher["haftalik_dars_soati"],
        "jadvaldagi_haftalik_soat": hafta_darsi,
        "kundalik_baho_eslatmasi_yoqilgan": kundalik_eslatma_yoqilgan,
        "kundalik_baho_eslatma": kundalik_eslatma,
        "kundalik_baho_eslatma_matni": "Bugungi darslaringiz bo'yicha Kundalikda baholarni kiritishni unutmadingizmi?" if kundalik_eslatma else None,
    }


class V1846KundalikEslatmaSozlama(BaseModel):
    yoqilgan: bool


@app.put("/api/oqituvchi/kundalik-baho-eslatmasi")
def v1846_kundalik_baho_eslatmasi(sorov: V1846KundalikEslatmaSozlama, token: str):
    teacher_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor()
    pass  # V19: DDL moved to startup migration.
    cur.execute(
        "UPDATE users SET kundalik_baho_eslatmasi=%s WHERE user_id=%s RETURNING user_id",
        (bool(sorov.yoqilgan), teacher_id),
    )
    if not cur.fetchone():
        conn.rollback(); cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="O'qituvchi topilmadi")
    conn.commit(); cur.close(); conn.close()
    return {"holat": "saqlandi", "yoqilgan": bool(sorov.yoqilgan)}


@app.get("/api/maktab/yuklama_xulosasi")
def v1845_yuklama_xulosasi(token: str, maktab_id: int):
    actor_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor()
    _v1845_smart_school_tables(cur)
    # V18.57: GET endpoint ma'lumotni o'qiydi, dublikat tozalash kabi yozuv amali bajarmaydi.
    if not _maktab_boshqaruvchi_mi(cur, actor_id, maktab_id):
        cur.close(); conn.close(); raise HTTPException(status_code=403, detail="Faqat maktab rahbariyati yoki admin ko'ra oladi")
    cur.execute("SELECT to_regclass('public.aqlli_jadval_urinishlari_v2') AS table_name")
    smart_bor = bool((cur.fetchone() or {}).get("table_name"))
    smart_run_id = None
    if smart_bor:
        cur.execute("SELECT id FROM aqlli_jadval_urinishlari_v2 WHERE maktab_id=%s AND holat='tasdiqlangan' ORDER BY id DESC LIMIT 1", (maktab_id,))
        run = cur.fetchone(); smart_run_id = run["id"] if run else None
    if smart_run_id:
        cur.execute("""
            SELECT u.user_id,u.full_name,u.fanlari,u.haftalik_dars_soati,
                   COUNT(e.id) AS jadvaldagi_soat
            FROM users u
            LEFT JOIN aqlli_jadval_slotlari_v2 e
              ON e.oqituvchi_user_id=u.user_id AND e.urinish_id=%s
            WHERE u.maktab_id=%s AND u.lavozim IS NOT NULL
            GROUP BY u.user_id,u.full_name,u.fanlari,u.haftalik_dars_soati
            ORDER BY u.full_name
        """, (smart_run_id, maktab_id))
    else:
        cur.execute("""
            SELECT u.user_id,u.full_name,u.fanlari,u.haftalik_dars_soati,
                   COUNT(j.id) AS jadvaldagi_soat
            FROM users u
            LEFT JOIN dars_jadvali j ON j.oqituvchi_user_id=u.user_id
            WHERE u.maktab_id=%s AND u.lavozim IS NOT NULL
            GROUP BY u.user_id,u.full_name,u.fanlari,u.haftalik_dars_soati
            ORDER BY u.full_name
        """, (maktab_id,))
    xodimlar = cur.fetchall()
    for x in xodimlar:
        reja = x.get("haftalik_dars_soati")
        amaldagi = int(x.get("jadvaldagi_soat") or 0)
        x["farq"] = None if reja is None else int(reja) - amaldagi
        x["holat"] = "kiritilmagan" if reja is None else ("ortiqcha" if amaldagi > int(reja) else "toliq" if amaldagi == int(reja) else "yetishmaydi")
    cur.close(); conn.close()
    return {"xodimlar": xodimlar}


@app.get("/api/maktab/aqlli_holatlar")
def v1845_aqlli_holatlar(token: str, maktab_id: int, holat: str = "ochiq"):
    actor_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor(); _v1845_smart_school_tables(cur)
    if not _maktab_boshqaruvchi_mi(cur, actor_id, maktab_id):
        cur.close(); conn.close(); raise HTTPException(status_code=403, detail="Faqat maktab rahbariyati yoki admin ko'ra oladi")
    cur.execute("""
        SELECT h.*,u.full_name FROM aqlli_holatlar h
        JOIN users u ON u.user_id=h.oquvchi_user_id
        WHERE h.maktab_id=%s AND h.holat=%s
        ORDER BY h.daraja DESC,h.yangilangan_at DESC LIMIT 100
    """, (maktab_id, holat))
    natija=cur.fetchall(); cur.close(); conn.close(); return {"holatlar": natija}


# V18.45: eski davomat endpointidan keyin holat yaratilishi uchun route-level wrapper emas,
# balki mavjud endpoint ishlagach davomat ma'lumotidan signalni keyingi so'rovda dashboard/holatlar
# orqali hisoblash mumkin. Quyidagi endpoint kerak bo'lsa qo'lda qayta hisoblaydi.
@app.post("/api/maktab/aqlli_holatlarni_yangila")
def v1845_holatlarni_yangila(token: str, maktab_id: int):
    actor_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor(); _davomat_jadvali(cur); _v1845_smart_school_tables(cur)
    if not _maktab_boshqaruvchi_mi(cur, actor_id, maktab_id):
        cur.close(); conn.close(); raise HTTPException(status_code=403, detail="Faqat maktab rahbariyati yoki admin yangilay oladi")
    cur.execute("""
        SELECT DISTINCT d.user_id,d.sana FROM davomat d
        JOIN maktab_sinflari s ON s.id=d.sinf_id
        WHERE s.maktab_id=%s AND d.holat='kelmadi' AND d.sana >= CURRENT_DATE - INTERVAL '120 days'
        ORDER BY d.sana
    """, (maktab_id,))
    son=0
    for r in cur.fetchall():
        _v1845_davomat_holat_yangila(cur,maktab_id,r["user_id"],r["sana"]); son += 1
    conn.commit(); cur.close(); conn.close(); return {"holat":"yangilandi","yozuvlar":son}

# ========================= V18.45 END =========================


# ═══════════════════════════════════════════════════════════
# V18.47 — ADMIN UCHUN "ROL SIFATIDA KO'RISH" (READ-ONLY PREVIEW)
# Admin maktabni turli rollar ko'zidan ko'radi, lekin bu endpointlar
# hech qanday ma'lumotni o'zgartirmaydi va boshqa foydalanuvchi tokenini bermaydi.
# ═══════════════════════════════════════════════════════════

_V1847_PREVIEW_ROLLAR = {
    "maktab_admin": "Maktab admini",
    "direktor": "Maktab direktori",
    "zavuch": "O'quv ishlari bo'yicha direktor o'rinbosari",
    "manaviyatchi": "Ma'naviy-ma'rifiy ishlar bo'yicha direktor o'rinbosari",
    "fan_oqituvchisi": "Fan o'qituvchisi",
    "sinf_rahbari": "Sinf rahbari",
    "oquvchi": "O'quvchi",
    "ota_ona": "Ota-ona",
    "psixolog": "Psixolog",
    "hamshira": "Hamshira",
}


def _v1847_preview_common(cur, maktab_id: int):
    _maktab_jadvali(cur)
    _maktab_sinflari_jadvali(cur)
    _sinf_azolari_jadvali(cur)
    _davomat_jadvali(cur)
    _v1845_smart_school_tables(cur)
    _xodim_sinf_birikmalari_jadvali(cur)
    # V18.57: rol ko'rish ham read-only; avtomatik baza tozalash bu yerda ishlamaydi.

    cur.execute("""
        SELECT id, nomi, maktab_raqami, viloyat, tuman, smena_soni
        FROM maktablar WHERE id=%s AND archived_at IS NULL
    """, (maktab_id,))
    maktab = cur.fetchone()
    if not maktab:
        raise HTTPException(status_code=404, detail="Maktab topilmadi")

    cur.execute("""
        SELECT s.id,s.sinf,s.harf,s.rahbar_user_id,
               COALESCE(u.full_name,'') AS rahbar_ismi,
               COUNT(a.user_id) AS oquvchi_soni
        FROM maktab_sinflari s
        LEFT JOIN users u ON u.user_id=s.rahbar_user_id
        LEFT JOIN maktab_sinf_azolari a ON a.sinf_id=s.id
        WHERE s.maktab_id=%s
        GROUP BY s.id,s.sinf,s.harf,s.rahbar_user_id,u.full_name
        ORDER BY s.sinf::int,s.harf
    """, (maktab_id,))
    sinflar = cur.fetchall()

    cur.execute("""
        SELECT user_id,full_name,lavozim,fanlari,haftalik_dars_soati
        FROM users
        WHERE maktab_id=%s AND lavozim IS NOT NULL
        ORDER BY
          CASE lavozim
            WHEN 'direktor' THEN 1
            WHEN 'zam_direktor_uquv' THEN 2
            WHEN 'zam_direktor_tarbiya' THEN 3
            WHEN 'psixolog' THEN 4
            WHEN 'hamshira' THEN 5
            WHEN 'fan_oqituvchisi' THEN 6
            ELSE 20
          END,
          full_name
    """, (maktab_id,))
    xodimlar = cur.fetchall()

    cur.execute("""
        SELECT COUNT(DISTINCT a.user_id) AS son
        FROM maktab_sinf_azolari a
        JOIN maktab_sinflari s ON s.id=a.sinf_id
        WHERE s.maktab_id=%s
    """, (maktab_id,))
    oquvchi_soni = int((cur.fetchone() or {}).get("son") or 0)

    cur.execute("""
        SELECT COUNT(*) AS son FROM aqlli_holatlar
        WHERE maktab_id=%s AND holat='ochiq'
    """, (maktab_id,))
    ochiq_holatlar = int((cur.fetchone() or {}).get("son") or 0)

    cur.execute("""
        SELECT COUNT(*) AS son
        FROM davomat d
        JOIN maktab_sinflari s ON s.id=d.sinf_id
        WHERE s.maktab_id=%s AND d.sana=CURRENT_DATE AND d.holat='kelmadi'
    """, (maktab_id,))
    bugun_kelmagan = int((cur.fetchone() or {}).get("son") or 0)

    return maktab, sinflar, xodimlar, {
        "oquvchi_soni": oquvchi_soni,
        "sinf_soni": len(sinflar),
        "xodim_soni": len(xodimlar),
        "ochiq_holatlar": ochiq_holatlar,
        "bugun_kelmagan": bugun_kelmagan,
    }


@app.get("/api/admin/maktab_korish_katalogi")
def v1847_maktab_korish_katalogi(token: str, maktab_id: int):
    _admin_tekshir(token)
    conn = _db(); cur = conn.cursor()
    try:
        maktab, sinflar, xodimlar, xulosa = _v1847_preview_common(cur, maktab_id)

        cur.execute("""
            SELECT a.sinf_id,u.user_id,u.full_name
            FROM maktab_sinf_azolari a
            JOIN users u ON u.user_id=a.user_id
            JOIN maktab_sinflari s ON s.id=a.sinf_id
            WHERE s.maktab_id=%s
            ORDER BY a.sinf_id,u.full_name
        """, (maktab_id,))
        oquvchilar = cur.fetchall()

        return {
            "read_only": True,
            "maktab": maktab,
            "xulosa": xulosa,
            "rollar": [{"kalit": k, "nom": v} for k, v in _V1847_PREVIEW_ROLLAR.items()],
            "sinflar": sinflar,
            "xodimlar": xodimlar,
            "oquvchilar": oquvchilar,
        }
    finally:
        cur.close(); conn.close()


@app.get("/api/admin/maktab_rol_korish")
def v1847_maktab_rol_korish(
    token: str,
    maktab_id: int,
    rol: str,
    user_id: Optional[int] = None,
    sinf_id: Optional[int] = None,
    oquvchi_id: Optional[int] = None,
):
    _admin_tekshir(token)
    rol = str(rol or "").strip().lower()
    if rol not in _V1847_PREVIEW_ROLLAR:
        raise HTTPException(status_code=400, detail="Noma'lum ko'rish roli")

    conn = _db(); cur = conn.cursor()
    try:
        maktab, sinflar, xodimlar, xulosa = _v1847_preview_common(cur, maktab_id)

        response = {
            "read_only": True,
            "rol": rol,
            "rol_nomi": _V1847_PREVIEW_ROLLAR[rol],
            "maktab": maktab,
            "xulosa": xulosa,
            "tanlangan": {},
            "kartalar": [],
            "bolimlar": [],
            "ogohlantirish": "ADMIN KO'RISH REJIMI — bu oynada hech qanday ma'lumot o'zgartirilmaydi.",
        }

        def card(label, value, tone="blue"):
            response["kartalar"].append({"label": label, "value": value, "tone": tone})

        def section(title, subtitle="", items=None, empty_text=None, kind="list"):
            response["bolimlar"].append({
                "title": title, "subtitle": subtitle, "items": items or [],
                "empty_text": empty_text, "kind": kind,
            })

        # Maktab admini / direktor
        if rol in {"maktab_admin", "direktor"}:
            card("O'quvchilar", xulosa["oquvchi_soni"], "blue")
            card("Sinflar", xulosa["sinf_soni"], "teal")
            card("Xodimlar", xulosa["xodim_soni"], "green")
            card("Bugun kelmagan", xulosa["bugun_kelmagan"], "amber" if xulosa["bugun_kelmagan"] else "green")
            card("Ochiq aqlli holat", xulosa["ochiq_holatlar"], "red" if xulosa["ochiq_holatlar"] else "green")
            section(
                "Maktab boshqaruv markazi",
                "Rahbariyat bir qarashda ko'radigan asosiy holatlar.",
                [
                    {"title": "Dars jadvali va yuklama", "detail": "O'qituvchi yuklamasi, bo'sh darslar va konfliktlar."},
                    {"title": "Davomat signallari", "detail": f"Bugun {xulosa['bugun_kelmagan']} ta kelmagan yozuvi bor."},
                    {"title": "Aqlli holatlar", "detail": f"{xulosa['ochiq_holatlar']} ta ochiq kuzatuv holati."},
                    {"title": "Sinf va xodimlar", "detail": f"{xulosa['sinf_soni']} sinf · {xulosa['xodim_soni']} xodim."},
                ],
            )

        # Zavuch
        elif rol == "zavuch":
            cur.execute("""
                SELECT u.user_id,u.full_name,u.fanlari,u.haftalik_dars_soati,
                       COUNT(j.id) AS jadvaldagi_soat
                FROM users u
                LEFT JOIN dars_jadvali j ON j.oqituvchi_user_id=u.user_id
                WHERE u.maktab_id=%s AND u.lavozim IS NOT NULL
                GROUP BY u.user_id,u.full_name,u.fanlari,u.haftalik_dars_soati
                ORDER BY ABS(COALESCE(u.haftalik_dars_soati,0)-COUNT(j.id)) DESC,u.full_name
                LIMIT 12
            """, (maktab_id,))
            yuklama = cur.fetchall()
            card("Sinflar", xulosa["sinf_soni"], "teal")
            card("O'qituvchilar", sum(1 for x in xodimlar if x.get("fanlari")), "blue")
            card("Bugun kelmagan", xulosa["bugun_kelmagan"], "amber")
            section("O'qituvchi yuklamasi", "Jadval va belgilangan haftalik soatni solishtirish.", [
                {"title": x["full_name"], "detail": f"{x.get('fanlari') or 'Fan ko‘rsatilmagan'} · reja {x.get('haftalik_dars_soati') if x.get('haftalik_dars_soati') is not None else '—'} · jadval {int(x.get('jadvaldagi_soat') or 0)}"}
                for x in yuklama
            ], "Yuklama ma'lumoti yo'q")

        # Ma'naviyatchi
        elif rol == "manaviyatchi":
            cur.execute("""
                SELECT h.daraja,h.sarlavha,h.tavsif,u.full_name
                FROM aqlli_holatlar h JOIN users u ON u.user_id=h.oquvchi_user_id
                WHERE h.maktab_id=%s AND h.holat='ochiq'
                ORDER BY h.daraja DESC,h.yangilangan_at DESC LIMIT 12
            """, (maktab_id,))
            holatlar = cur.fetchall()
            card("Bugun kelmagan", xulosa["bugun_kelmagan"], "amber")
            card("Ochiq holatlar", xulosa["ochiq_holatlar"], "red" if xulosa["ochiq_holatlar"] else "green")
            card("Sinflar", xulosa["sinf_soni"], "teal")
            section("E'tibor talab qiladigan o'quvchilar", "Davomat va tarbiyaviy kuzatuvlar.", [
                {"title": x["full_name"], "detail": f"Daraja {x['daraja']} · {x['sarlavha']}"}
                for x in holatlar
            ], "Hozircha ochiq holat yo'q")

        # Psixolog
        elif rol == "psixolog":
            cur.execute("""
                SELECT h.daraja,h.sarlavha,h.tavsif,u.full_name
                FROM aqlli_holatlar h JOIN users u ON u.user_id=h.oquvchi_user_id
                WHERE h.maktab_id=%s AND h.holat='ochiq'
                  AND (h.masul_rol ILIKE '%%psixolog%%' OR h.daraja>=2)
                ORDER BY h.daraja DESC,h.yangilangan_at DESC LIMIT 12
            """, (maktab_id,))
            holatlar = cur.fetchall()
            card("Ko'rib chiqiladigan holat", len(holatlar), "amber" if holatlar else "green")
            card("Jami o'quvchi", xulosa["oquvchi_soni"], "blue")
            section("Psixolog ish navbati", "Faqat psixologga tegishli yoki yordam talab qiladigan holatlar.", [
                {"title": x["full_name"], "detail": f"Daraja {x['daraja']} · {x['sarlavha']}"}
                for x in holatlar
            ], "Psixolog uchun ochiq signal yo'q")

        # Hamshira — sog'liq yozuvlari alohida modul sifatida keyin ulanadi.
        elif rol == "hamshira":
            card("Jami o'quvchi", xulosa["oquvchi_soni"], "blue")
            card("Sinflar", xulosa["sinf_soni"], "teal")
            section(
                "Hamshira ish maydoni",
                "Sog'liq ma'lumotlari maxfiy bo'lgani uchun admin preview faqat interfeys tuzilishini ko'rsatadi.",
                [
                    {"title": "Bugungi murojaatlar", "detail": "Sog'liq moduli ulanganidan keyin shu yerda chiqadi."},
                    {"title": "Dori/allergiya ogohlantirishlari", "detail": "Faqat vakolatli hamshiraga ko'rinadigan yopiq blok bo'ladi."},
                    {"title": "Tibbiy ko'rik rejalari", "detail": "Sinf kesimida rejalashtiriladi."},
                ],
            )

        # Fan o'qituvchisi
        elif rol == "fan_oqituvchisi":
            if user_id is None:
                raise HTTPException(status_code=400, detail="Fan o'qituvchisini tanlang")
            cur.execute("""
                SELECT user_id,full_name,fanlari,haftalik_dars_soati
                FROM users WHERE user_id=%s AND maktab_id=%s
            """, (user_id, maktab_id))
            teacher = cur.fetchone()
            if not teacher:
                raise HTTPException(status_code=404, detail="Tanlangan o'qituvchi topilmadi")
            response["tanlangan"] = teacher
            cur.execute("""
                SELECT DISTINCT s.id AS sinf_id,s.sinf,s.harf,b.fan_nomi,s.sinf::int AS sinf_sort
                FROM maktab_dars_birikmalari b
                JOIN maktab_sinflari s ON s.id=b.sinf_id
                WHERE b.user_id=%s AND b.maktab_id=%s
                ORDER BY sinf_sort,s.harf,b.fan_nomi
            """, (user_id, maktab_id))
            birikmalar = cur.fetchall()
            cur.execute("""
                SELECT j.dars_raqami,j.fan,j.xona,j.boshlanish_vaqti,s.sinf,s.harf
                FROM dars_jadvali j JOIN maktab_sinflari s ON s.id=j.sinf_id
                WHERE j.oqituvchi_user_id=%s AND j.kun=EXTRACT(ISODOW FROM CURRENT_DATE)::int
                ORDER BY j.dars_raqami
            """, (user_id,))
            bugun = cur.fetchall()
            card("Bugungi dars", len(bugun), "teal")
            card("Biriktirilgan sinf/fan", len(birikmalar), "blue")
            card("Haftalik yuklama", teacher.get("haftalik_dars_soati") if teacher.get("haftalik_dars_soati") is not None else "—", "green")
            section("Bugungi darslarim", teacher["full_name"], [
                {"title": f"{x['dars_raqami']}-dars · {x['fan']}", "detail": f"{x['sinf']}-{x['harf']}" + (f" · {x['boshlanish_vaqti']}" if x.get("boshlanish_vaqti") else "")}
                for x in bugun
            ], "Bugun dars topilmadi")
            section("Mening sinf/fanlarim", "Monitoring va to'garak uchun asosiy ish maydoni.", [
                {"title": f"{x['sinf']}-{x['harf']}", "detail": x["fan_nomi"]} for x in birikmalar
            ], "Sinf/fan birikmasi topilmadi")

        # Sinf rahbari
        elif rol == "sinf_rahbari":
            if sinf_id is None:
                raise HTTPException(status_code=400, detail="Sinfni tanlang")
            cur.execute("""
                SELECT s.id,s.sinf,s.harf,s.rahbar_user_id,u.full_name AS rahbar_ismi
                FROM maktab_sinflari s LEFT JOIN users u ON u.user_id=s.rahbar_user_id
                WHERE s.id=%s AND s.maktab_id=%s
            """, (sinf_id, maktab_id))
            sinf = cur.fetchone()
            if not sinf:
                raise HTTPException(status_code=404, detail="Sinf topilmadi")
            response["tanlangan"] = sinf
            cur.execute("""
                SELECT u.user_id,u.full_name,
                       COUNT(*) FILTER (WHERE d.holat='kelmadi' AND d.sana>=CURRENT_DATE-INTERVAL '30 days') AS kelmadi30,
                       COUNT(*) FILTER (WHERE d.holat='kechikdi' AND d.sana>=CURRENT_DATE-INTERVAL '30 days') AS kechikdi30
                FROM maktab_sinf_azolari a JOIN users u ON u.user_id=a.user_id
                LEFT JOIN davomat d ON d.user_id=u.user_id AND d.sinf_id=a.sinf_id
                WHERE a.sinf_id=%s
                GROUP BY u.user_id,u.full_name ORDER BY u.full_name
            """, (sinf_id,))
            bolalar = cur.fetchall()
            card("O'quvchilar", len(bolalar), "blue")
            card("Rahbar", sinf.get("rahbar_ismi") or "Biriktirilmagan", "teal")
            section(f"{sinf['sinf']}-{sinf['harf']} sinfim", "Sinf rahbari faqat o'z sinfiga kerakli ishlarni ko'radi.", [
                {"title": x["full_name"], "detail": f"30 kun: {int(x.get('kelmadi30') or 0)} kelmadi · {int(x.get('kechikdi30') or 0)} kechikdi"}
                for x in bolalar
            ], "Sinfda o'quvchi yo'q")

        # O'quvchi / ota-ona
        elif rol in {"oquvchi", "ota_ona"}:
            if sinf_id is None:
                raise HTTPException(status_code=400, detail="Avval sinfni tanlang")
            if oquvchi_id is None:
                cur.execute("""
                    SELECT u.user_id FROM maktab_sinf_azolari a
                    JOIN users u ON u.user_id=a.user_id
                    WHERE a.sinf_id=%s ORDER BY u.full_name LIMIT 1
                """, (sinf_id,))
                first = cur.fetchone()
                oquvchi_id = first["user_id"] if first else None
            if oquvchi_id is None:
                raise HTTPException(status_code=404, detail="Bu sinfda o'quvchi topilmadi")
            cur.execute("""
                SELECT u.user_id,u.full_name,s.id AS sinf_id,s.sinf,s.harf
                FROM users u JOIN maktab_sinf_azolari a ON a.user_id=u.user_id
                JOIN maktab_sinflari s ON s.id=a.sinf_id
                WHERE u.user_id=%s AND s.id=%s AND s.maktab_id=%s
            """, (oquvchi_id, sinf_id, maktab_id))
            bola = cur.fetchone()
            if not bola:
                raise HTTPException(status_code=404, detail="O'quvchi tanlangan sinfda topilmadi")
            response["tanlangan"] = bola
            cur.execute("""
                SELECT fan,ROUND(AVG(foiz)) AS foiz,COUNT(*) AS urinish
                FROM dars_monitoring_baholari
                WHERE oquvchi_user_id=%s
                GROUP BY fan ORDER BY AVG(foiz) ASC NULLS LAST LIMIT 10
            """, (oquvchi_id,))
            fanlar = cur.fetchall()
            cur.execute("""
                SELECT COUNT(*) FILTER (WHERE holat='keldi') AS keldi,
                       COUNT(*) FILTER (WHERE holat='kelmadi') AS kelmadi,
                       COUNT(*) FILTER (WHERE holat='kechikdi') AS kechikdi
                FROM davomat
                WHERE user_id=%s AND sana>=CURRENT_DATE-INTERVAL '30 days'
            """, (oquvchi_id,))
            davomat = cur.fetchone() or {}
            card("30 kunda kelgan", int(davomat.get("keldi") or 0), "green")
            card("Kelmagan", int(davomat.get("kelmadi") or 0), "amber")
            card("Kechikkan", int(davomat.get("kechikdi") or 0), "amber")
            if fanlar:
                ortacha = round(sum(float(x.get("foiz") or 0) for x in fanlar) / len(fanlar))
                card("Monitoring o'rtacha", f"{ortacha}%", "blue")
            title = "Mening bilim xaritam" if rol == "oquvchi" else "Farzandimning bilim xaritasi"
            section(title, f"{bola['full_name']} · {bola['sinf']}-{bola['harf']}", [
                {"title": x["fan"], "detail": f"{int(round(float(x.get('foiz') or 0)))}% · {int(x.get('urinish') or 0)} ta monitoring"}
                for x in fanlar
            ], "Hali monitoring natijasi yo'q")
            section(
                "Tavsiya etiladigan yo'l",
                "Tizim mavjud monitoringdan kelib chiqib qaysi fan/mavzuga e'tibor kerakligini ko'rsatadi.",
                [
                    {"title": "Bo'shliqni yopish", "detail": "Past natijali fan va mavzularga qisqa tushuntirish + mashqlar."},
                    {"title": "Bilimni ushlab turish", "detail": "Yaxshi o'zlashtirilgan mavzularga qisqa takrorlash."},
                    {"title": "To'garak va qo'shimcha dars", "detail": "Mos kurslar mavjud bo'lsa shu yerda tavsiya qilinadi."},
                ],
            )

        return response
    finally:
        cur.close(); conn.close()

# ========================= V18.47 END =========================


# ═══════════════════════════════════════════════════════════
# V18.50 — XODIM DUBLIKATLARINI AVTOMATIK TUZATISH
# Faqat Excel importidan yaratilgan manfiy user_id xodimlarga tegadi.
# Haqiqiy Google/Telegram foydalanuvchi hisoblari avtomatik birlashtirilmaydi.
# Eng yangi import yozuvi (eng kichik/manfiy user_id) saqlanadi.
# Eski dublikatlar xavfsiz arxivlanadi, bog'lanishlar yangi yozuvga ko'chiriladi.
# ═══════════════════════════════════════════════════════════

def _v1850_xodim_dublikatlarini_tozala(cur, maktab_id: int):
    _xodim_kod_jadvali(cur)
    _maktab_sinflari_jadvali(cur)
    _xodim_sinf_birikmalari_jadvali(cur)
    _rejalashtirish_jadvallari(cur)
    _xodim_davomati_jadvali(cur)
    _v1845_smart_school_tables(cur)

    cur.execute("""
        SELECT
            LOWER(REGEXP_REPLACE(TRIM(full_name), '\\s+', ' ', 'g')) AS ism_kalit,
            COALESCE(lavozim,'') AS lavozim_kalit,
            ARRAY_AGG(user_id ORDER BY user_id ASC) AS ids
        FROM users
        WHERE maktab_id=%s
          AND user_id < 0
          AND lavozim IS NOT NULL
          AND TRIM(COALESCE(full_name,'')) <> ''
        GROUP BY 1,2
        HAVING COUNT(*) > 1
    """, (maktab_id,))
    guruhlar = cur.fetchall()

    tozalangan = 0
    for g in guruhlar:
        ids = list(g.get("ids") or [])
        if len(ids) < 2:
            continue

        # Import har safar yanada kichik manfiy ID beradi, shu sabab MIN = eng yangi.
        asosiy = ids[0]
        dublikatlar = ids[1:]

        for eski in dublikatlar:
            # Eng to'liq ma'lumotlarni asosiy yozuvga saqlab qolamiz.
            cur.execute("""
                UPDATE users AS yangi
                SET fanlari = COALESCE(NULLIF(yangi.fanlari,''), eski.fanlari),
                    oqitadigan_sinflari = COALESCE(NULLIF(yangi.oqitadigan_sinflari,''), eski.oqitadigan_sinflari),
                    ish_staji = COALESCE(yangi.ish_staji, eski.ish_staji),
                    toifasi = COALESCE(NULLIF(yangi.toifasi,''), eski.toifasi),
                    haftalik_dars_soati = COALESCE(yangi.haftalik_dars_soati, eski.haftalik_dars_soati)
                FROM users AS eski
                WHERE yangi.user_id=%s AND eski.user_id=%s
            """, (asosiy, eski))

            # Sinf rahbarligi.
            cur.execute(
                "UPDATE maktab_sinflari SET rahbar_user_id=%s WHERE maktab_id=%s AND rahbar_user_id=%s",
                (asosiy, maktab_id, eski),
            )

            # Xodim-sinf bog'lanishlarini conflict-siz ko'chiramiz.
            cur.execute("""
                INSERT INTO maktab_xodim_sinflari(maktab_id,user_id,sinf_id,fanlari)
                SELECT maktab_id,%s,sinf_id,fanlari
                FROM maktab_xodim_sinflari
                WHERE maktab_id=%s AND user_id=%s
                ON CONFLICT(user_id,sinf_id) DO UPDATE
                SET fanlari=COALESCE(EXCLUDED.fanlari,maktab_xodim_sinflari.fanlari)
            """, (asosiy, maktab_id, eski))
            cur.execute(
                "DELETE FROM maktab_xodim_sinflari WHERE maktab_id=%s AND user_id=%s",
                (maktab_id, eski),
            )

            # Aniq sinf-fan-guruh birikmalarini ko'chiramiz.
            cur.execute("""
                INSERT INTO maktab_dars_birikmalari(maktab_id,user_id,sinf_id,fan_nomi,guruh_kaliti)
                SELECT maktab_id,%s,sinf_id,fan_nomi,guruh_kaliti
                FROM maktab_dars_birikmalari
                WHERE maktab_id=%s AND user_id=%s
                ON CONFLICT(user_id,sinf_id,fan_nomi,guruh_kaliti) DO NOTHING
            """, (asosiy, maktab_id, eski))
            cur.execute(
                "DELETE FROM maktab_dars_birikmalari WHERE maktab_id=%s AND user_id=%s",
                (maktab_id, eski),
            )

            # Jadvaldagi darslar.
            cur.execute(
                "UPDATE dars_jadvali SET oqituvchi_user_id=%s WHERE oqituvchi_user_id=%s",
                (asosiy, eski),
            )

            # Monitoring yozuvlari (o'qituvchi sifatida).
            cur.execute(
                "UPDATE dars_monitoring_baholari SET oqituvchi_user_id=%s WHERE maktab_id=%s AND oqituvchi_user_id=%s",
                (asosiy, maktab_id, eski),
            )

            # Xodim davomatida bir kunlik conflict bo'lishi mumkin — avval merge.
            cur.execute("""
                INSERT INTO xodim_davomati(maktab_id,user_id,sana,holat,izoh,belgilagan_user_id,belgilangan_at)
                SELECT maktab_id,%s,sana,holat,izoh,belgilagan_user_id,belgilangan_at
                FROM xodim_davomati
                WHERE maktab_id=%s AND user_id=%s
                ON CONFLICT(maktab_id,user_id,sana) DO NOTHING
            """, (asosiy, maktab_id, eski))
            cur.execute(
                "DELETE FROM xodim_davomati WHERE maktab_id=%s AND user_id=%s",
                (maktab_id, eski),
            )

            # Direktor ko'rsatkichi eski duplicatega qaragan bo'lsa.
            cur.execute(
                "UPDATE maktablar SET direktor_user_id=%s WHERE id=%s AND direktor_user_id=%s",
                (asosiy, maktab_id, eski),
            )

            # Eski duplicate kirish kodini bekor qilamiz.
            cur.execute("DELETE FROM xodim_kod WHERE user_id=%s", (eski,))

            # FK sabab foydalanuvchini qattiq DELETE qilmaymiz.
            # Maktabdan chiqarib, lavozimini bo'shatamiz — dashboard/importda qayta chiqmaydi.
            cur.execute("""
                UPDATE users
                SET maktab_id=NULL,
                    lavozim=NULL,
                    fanlari=NULL,
                    oqitadigan_sinflari=NULL,
                    haftalik_dars_soati=NULL,
                    full_name=full_name || ' [dublikat arxiv]'
                WHERE user_id=%s AND user_id<0
            """, (eski,))
            tozalangan += 1

    return {"tozalangan": tozalangan, "guruhlar": len(guruhlar)}


@app.post("/api/admin/maktab_xodim_dublikatlarini_tozala")
def v1850_maktab_xodim_dublikatlarini_tozala(token: str, maktab_id: int):
    _admin_tekshir(token)
    conn = _db(); cur = conn.cursor()
    try:
        natija = _v1850_xodim_dublikatlarini_tozala(cur, maktab_id)
        conn.commit()
        return {"holat": "tozalandi", **natija}
    except Exception:
        conn.rollback()
        raise
    finally:
        cur.close(); conn.close()

# ========================= V18.50 END =========================

# ═══════════════════════════════════════════════════════════
# V18.52 — AQILLI DARS JADVALI 2.0
# Kalendar, smena, o'qituvchi vaqti/yuklamasi, xonalar, parallel guruhlar,
# draft→tasdiq, aniq diagnostika va mavzu rejasini real sanalarga taqsimlash.
# ═══════════════════════════════════════════════════════════

import random as _v1852_random
from collections import defaultdict as _v1852_defaultdict, Counter as _v1852_Counter

_V1852_TABLES_READY = False
_V1852_TABLES_LOCK = threading.Lock()
_V1852_DAM_TURLARI = {"dam", "bayram", "tatil", "qoshimcha_dam"}
_V1852_OQISH_TURLARI = {"oqish", "qoshimcha_oqish"}
_V1852_HOLATLAR = {"taxminiy", "tasdiqlangan"}
_V1852_KUN_TURLARI = _V1852_DAM_TURLARI | _V1852_OQISH_TURLARI
_V1852_VAQT_TURLARI = {"band", "afzal_bosh", "metod_kuni"}
_V1852_HAFTA = {1: "Dushanba", 2: "Seshanba", 3: "Chorshanba", 4: "Payshanba", 5: "Juma", 6: "Shanba", 7: "Yakshanba"}


def _v1855_sinf_raqami(value):
    """`1`, `1-sinf`, `01` kabi qiymatlardan sinf raqamini xavfsiz oladi."""
    match = re.search(r"\d+", str(value or ""))
    if not match:
        return None
    try:
        return int(match.group(0))
    except (TypeError, ValueError):
        return None


def _v1855_boshlangich_sinf(class_row):
    grade = _v1855_sinf_raqami((class_row or {}).get("sinf"))
    return grade is not None and 1 <= grade <= 4


def _v1856_seed_default_class_day_rules(cur, maktab_id: int):
    """Birinchi ishga tushishda eski 1–4/Shanba qoidasini preset sifatida saqlaydi.

    Marker yozilgach admin uni o'chirsa, keyingi reload'da qayta paydo bo'lmaydi.
    """
    cur.execute("SELECT 1 FROM aqlli_sinf_kun_blok_seed_v2 WHERE maktab_id=%s", (maktab_id,))
    if cur.fetchone():
        return
    cur.execute("SELECT COUNT(*) AS son FROM aqlli_sinf_kun_bloklari_v2 WHERE maktab_id=%s", (maktab_id,))
    existing = int((cur.fetchone() or {}).get("son") or 0)
    if existing == 0:
        for grade in (1, 2, 3, 4):
            cur.execute("""INSERT INTO aqlli_sinf_kun_bloklari_v2(
                maktab_id,sinf_daraja,hafta_kuni,izoh,faol)
                VALUES(%s,%s,6,%s,TRUE)""",
                (maktab_id, grade, "Boshlang'ich parallel uchun Shanba kuni dars yo'q"))
    cur.execute("""INSERT INTO aqlli_sinf_kun_blok_seed_v2(maktab_id,versiya)
                   VALUES(%s,1) ON CONFLICT(maktab_id) DO NOTHING""", (maktab_id,))


def _v1856_class_day_rule_rows(cur, maktab_id: int):
    _v1856_seed_default_class_day_rules(cur, maktab_id)
    cur.execute("""SELECT b.*,s.sinf,s.harf
                   FROM aqlli_sinf_kun_bloklari_v2 b
                   LEFT JOIN maktab_sinflari s ON s.id=b.sinf_id
                   WHERE b.maktab_id=%s AND b.faol=TRUE
                   ORDER BY COALESCE(b.sinf_daraja,999),s.sinf::int NULLS LAST,s.harf,b.hafta_kuni""",
                (maktab_id,))
    rows = cur.fetchall()
    for row in rows:
        day_name = _V1852_HAFTA.get(int(row["hafta_kuni"]), str(row["hafta_kuni"]))
        if row.get("sinf_id") is not None:
            row["qamrov"] = "sinf"
            row["yorliq"] = f"{row.get('sinf','')}-{row.get('harf','')} · {day_name}"
        else:
            row["qamrov"] = "parallel"
            row["yorliq"] = f"Barcha {row.get('sinf_daraja')}-sinflar · {day_name}"
    return rows


def _v1856_class_day_rule_map(rows):
    exact = set()
    grades = set()
    for row in rows or []:
        day = int(row.get("hafta_kuni") or 0)
        if row.get("sinf_id") is not None:
            exact.add((int(row["sinf_id"]), day))
        elif row.get("sinf_daraja") is not None:
            grades.add((int(row["sinf_daraja"]), day))
    return {"exact": exact, "grades": grades, "rows": rows or []}


def _v1856_class_day_block_reason(class_row, day: int, rule_map):
    class_id = int((class_row or {}).get("id") or 0)
    grade = _v1855_sinf_raqami((class_row or {}).get("sinf"))
    day_name = _V1852_HAFTA.get(int(day), str(day))
    if (class_id, int(day)) in rule_map.get("exact", set()):
        return f"{(class_row or {}).get('sinf','')}-{(class_row or {}).get('harf','')} uchun {day_name} kuni dars bloklangan"
    if grade is not None and (grade, int(day)) in rule_map.get("grades", set()):
        return f"Barcha {grade}-sinflar uchun {day_name} kuni dars bloklangan"
    return None


def _v1856_schedule_block_violations(cur, maktab_id: int, run_id: int | None = None):
    _v1856_seed_default_class_day_rules(cur, maktab_id)
    if run_id is None:
        run = _v1852_active_run(cur, maktab_id)
        if not run:
            return []
        run_id = int(run["id"])
    cur.execute("""SELECT s.id AS sinf_id,s.sinf,s.harf,e.hafta_kuni,
                          COUNT(DISTINCT e.id) AS dars_soni
                   FROM aqlli_jadval_slotlari_v2 e
                   JOIN maktab_sinflari s ON s.id=e.sinf_id
                   JOIN aqlli_sinf_kun_bloklari_v2 b
                     ON b.maktab_id=e.maktab_id AND b.faol=TRUE
                    AND b.hafta_kuni=e.hafta_kuni
                    AND (b.sinf_id=e.sinf_id OR
                         (b.sinf_id IS NULL AND b.sinf_daraja=
                          NULLIF(REGEXP_REPLACE(COALESCE(s.sinf,''),'[^0-9]','','g'),'')::int))
                   WHERE e.maktab_id=%s AND e.urinish_id=%s
                   GROUP BY s.id,s.sinf,s.harf,e.hafta_kuni
                   ORDER BY s.sinf::int,s.harf,e.hafta_kuni""", (maktab_id, run_id))
    rows = cur.fetchall()
    for row in rows:
        row["kun_nomi"] = _V1852_HAFTA.get(int(row["hafta_kuni"]), str(row["hafta_kuni"]))
    return rows


def _v1852_create_tables(cur):
    global _V1852_TABLES_READY
    _rejalashtirish_jadvallari(cur)
    _maktab_sinflari_jadvali(cur)
    _xodim_sinf_birikmalari_jadvali(cur)
    _maktab_fanlari_jadvali(cur)
    cur.execute(
        "ALTER TABLE maktablar ADD COLUMN IF NOT EXISTS alifbo_turi "
        "TEXT NOT NULL DEFAULT 'latin_xalqaro'"
    )
    cur.execute("ALTER TABLE maktab_sinflari ADD COLUMN IF NOT EXISTS smena INTEGER DEFAULT 1")
    cur.execute(
        "ALTER TABLE maktab_sinflari ADD COLUMN IF NOT EXISTS "
        "talim_tili TEXT NOT NULL DEFAULT 'uz'"
    )
    cur.execute("ALTER TABLE maktab_sinflari ADD COLUMN IF NOT EXISTS bino TEXT")
    cur.execute("ALTER TABLE maktab_sinflari ADD COLUMN IF NOT EXISTS xona TEXT")
    # Yangi maktab ustasi xona nomini matn sifatida ham saqlaydi, lekin
    # generator uchun asosiy identifikator katalogdagi aniq xona ID sidir.
    # Ustunlar plain BIGINT: eski bazalarda noma'lum tarixiy qiymat bo'lsa
    # migratsiya FK sabab to'xtab qolmasin; yangi yozuvlar route ichida qat'iy
    # katalog tekshiruvidan o'tadi.
    cur.execute("ALTER TABLE maktab_sinflari ADD COLUMN IF NOT EXISTS bino_id BIGINT")
    cur.execute("ALTER TABLE maktab_sinflari ADD COLUMN IF NOT EXISTS xona_id BIGINT")
    cur.execute("ALTER TABLE maktab_sinflari ADD COLUMN IF NOT EXISTS psixolog_user_id BIGINT")
    cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS haftalik_dars_soati INTEGER")
    cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS jadval_raqami INTEGER")
    cur.execute("CREATE UNIQUE INDEX IF NOT EXISTS uq_users_maktab_jadval_raqami ON users(maktab_id,jadval_raqami) WHERE jadval_raqami IS NOT NULL")

    cur.execute("""CREATE TABLE IF NOT EXISTS aqlli_oquv_yillari_v2(
        id BIGSERIAL PRIMARY KEY,
        maktab_id INTEGER NOT NULL REFERENCES maktablar(id) ON DELETE CASCADE,
        nomi TEXT NOT NULL,
        boshlanish DATE NOT NULL,
        tugash DATE NOT NULL,
        hafta_kunlari INTEGER NOT NULL DEFAULT 6 CHECK(hafta_kunlari IN (5,6)),
        faol BOOLEAN NOT NULL DEFAULT TRUE,
        yaratgan_user_id BIGINT,
        yangilangan_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
        UNIQUE(maktab_id,nomi)
    )""")
    cur.execute("""CREATE TABLE IF NOT EXISTS aqlli_choraklar_v2(
        id BIGSERIAL PRIMARY KEY,
        oquv_yili_id BIGINT NOT NULL REFERENCES aqlli_oquv_yillari_v2(id) ON DELETE CASCADE,
        chorak INTEGER NOT NULL CHECK(chorak BETWEEN 1 AND 4),
        boshlanish DATE NOT NULL,
        tugash DATE NOT NULL,
        holat TEXT NOT NULL DEFAULT 'taxminiy' CHECK(holat IN ('taxminiy','tasdiqlangan')),
        UNIQUE(oquv_yili_id,chorak)
    )""")
    cur.execute("""CREATE TABLE IF NOT EXISTS aqlli_kalendar_kunlari_v2(
        id BIGSERIAL PRIMARY KEY,
        maktab_id INTEGER NOT NULL REFERENCES maktablar(id) ON DELETE CASCADE,
        sana DATE NOT NULL,
        turi TEXT NOT NULL CHECK(turi IN ('oqish','dam','bayram','tatil','qoshimcha_dam','qoshimcha_oqish')),
        nomi TEXT,
        holat TEXT NOT NULL DEFAULT 'taxminiy' CHECK(holat IN ('taxminiy','tasdiqlangan')),
        izoh TEXT,
        yaratgan_user_id BIGINT,
        yangilangan_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
        UNIQUE(maktab_id,sana)
    )""")
    cur.execute("""CREATE TABLE IF NOT EXISTS aqlli_smena_sozlamalari_v2(
        maktab_id INTEGER NOT NULL REFERENCES maktablar(id) ON DELETE CASCADE,
        smena INTEGER NOT NULL CHECK(smena IN (1,2)),
        dars_soni INTEGER NOT NULL DEFAULT 7 CHECK(dars_soni BETWEEN 1 AND 12),
        boshlanish_vaqti TIME NOT NULL,
        dars_daqiqa INTEGER NOT NULL DEFAULT 45 CHECK(dars_daqiqa BETWEEN 20 AND 90),
        tanaffus_daqiqa INTEGER NOT NULL DEFAULT 5 CHECK(tanaffus_daqiqa BETWEEN 0 AND 40),
        katta_tanaffus_darsdan_keyin INTEGER CHECK(katta_tanaffus_darsdan_keyin BETWEEN 1 AND 11),
        katta_tanaffus_daqiqa INTEGER NOT NULL DEFAULT 15 CHECK(katta_tanaffus_daqiqa BETWEEN 0 AND 60),
        PRIMARY KEY(maktab_id,smena)
    )""")
    cur.execute("""CREATE TABLE IF NOT EXISTS aqlli_xonalar_v2(
        id BIGSERIAL PRIMARY KEY,
        maktab_id INTEGER NOT NULL REFERENCES maktablar(id) ON DELETE CASCADE,
        nomi TEXT NOT NULL,
        turi TEXT NOT NULL DEFAULT 'oddiy',
        sigim INTEGER,
        faol BOOLEAN NOT NULL DEFAULT TRUE,
        UNIQUE(maktab_id,nomi)
    )""")
    cur.execute("ALTER TABLE aqlli_xonalar_v2 ADD COLUMN IF NOT EXISTS manba_xona_id BIGINT")
    cur.execute("ALTER TABLE aqlli_xonalar_v2 ADD COLUMN IF NOT EXISTS bino_id BIGINT")
    cur.execute("ALTER TABLE aqlli_xonalar_v2 ADD COLUMN IF NOT EXISTS xona_raqami TEXT")
    cur.execute("ALTER TABLE aqlli_xonalar_v2 ADD COLUMN IF NOT EXISTS qavat INTEGER")
    cur.execute("ALTER TABLE aqlli_xonalar_v2 ADD COLUMN IF NOT EXISTS darsga_yaroqli BOOLEAN NOT NULL DEFAULT TRUE")
    cur.execute("UPDATE aqlli_xonalar_v2 SET turi='reserve' WHERE turi='maxsus'")
    cur.execute("UPDATE aqlli_xonalar_v2 SET turi='classroom' WHERE turi='oddiy'")
    cur.execute("""CREATE TABLE IF NOT EXISTS aqlli_oqituvchi_vaqti_v2(
        id BIGSERIAL PRIMARY KEY,
        maktab_id INTEGER NOT NULL REFERENCES maktablar(id) ON DELETE CASCADE,
        user_id BIGINT NOT NULL REFERENCES users(user_id) ON DELETE CASCADE,
        hafta_kuni INTEGER NOT NULL CHECK(hafta_kuni BETWEEN 1 AND 7),
        smena INTEGER NOT NULL DEFAULT 0 CHECK(smena BETWEEN 0 AND 2),
        dars_raqami INTEGER NOT NULL DEFAULT 0 CHECK(dars_raqami BETWEEN 0 AND 12),
        turi TEXT NOT NULL CHECK(turi IN ('band','afzal_bosh','metod_kuni')),
        qattiq BOOLEAN NOT NULL DEFAULT TRUE,
        izoh TEXT,
        UNIQUE(maktab_id,user_id,hafta_kuni,smena,dars_raqami,turi)
    )""")
    cur.execute("""CREATE TABLE IF NOT EXISTS aqlli_oqituvchi_qoidalari_v2(
        maktab_id INTEGER NOT NULL REFERENCES maktablar(id) ON DELETE CASCADE,
        user_id BIGINT NOT NULL REFERENCES users(user_id) ON DELETE CASCADE,
        kunlik_max INTEGER NOT NULL DEFAULT 6 CHECK(kunlik_max BETWEEN 1 AND 12),
        ketma_ket_max INTEGER NOT NULL DEFAULT 4 CHECK(ketma_ket_max BETWEEN 1 AND 12),
        okno_max INTEGER NOT NULL DEFAULT 1 CHECK(okno_max BETWEEN 0 AND 6),
        afzal_smena INTEGER NOT NULL DEFAULT 0 CHECK(afzal_smena BETWEEN 0 AND 2),
        eng_erta_dars INTEGER NOT NULL DEFAULT 1 CHECK(eng_erta_dars BETWEEN 1 AND 12),
        eng_kech_dars INTEGER NOT NULL DEFAULT 12 CHECK(eng_kech_dars BETWEEN 1 AND 12),
        PRIMARY KEY(maktab_id,user_id)
    )""")
    cur.execute("""CREATE TABLE IF NOT EXISTS aqlli_sinf_fan_yuklamalari_v2(
        id BIGSERIAL PRIMARY KEY,
        maktab_id INTEGER NOT NULL REFERENCES maktablar(id) ON DELETE CASCADE,
        sinf_id INTEGER NOT NULL REFERENCES maktab_sinflari(id) ON DELETE CASCADE,
        fan_nomi TEXT NOT NULL,
        haftalik_soat INTEGER NOT NULL DEFAULT 0 CHECK(haftalik_soat BETWEEN 0 AND 20),
        kunlik_max INTEGER NOT NULL DEFAULT 1 CHECK(kunlik_max BETWEEN 1 AND 4),
        ketma_ket_mumkin BOOLEAN NOT NULL DEFAULT FALSE,
        afzal_oxirgi_dars INTEGER NOT NULL DEFAULT 5 CHECK(afzal_oxirgi_dars BETWEEN 1 AND 12),
        asosiy_oqituvchi_user_id BIGINT REFERENCES users(user_id),
        xona_id BIGINT REFERENCES aqlli_xonalar_v2(id),
        nazorat_soni INTEGER NOT NULL DEFAULT 0 CHECK(nazorat_soni BETWEEN 0 AND 10),
        nazoratdan_keyin_tahlil BOOLEAN NOT NULL DEFAULT TRUE,
        mustahkamlash_soni INTEGER NOT NULL DEFAULT 0 CHECK(mustahkamlash_soni BETWEEN 0 AND 20),
        ogirlik INTEGER NOT NULL DEFAULT 2 CHECK(ogirlik BETWEEN 1 AND 3),
        UNIQUE(maktab_id,sinf_id,fan_nomi)
    )""")
    cur.execute("""CREATE TABLE IF NOT EXISTS aqlli_guruh_sozlamalari_v2(
        maktab_id INTEGER NOT NULL REFERENCES maktablar(id) ON DELETE CASCADE,
        sinf_id INTEGER NOT NULL REFERENCES maktab_sinflari(id) ON DELETE CASCADE,
        fan_nomi TEXT NOT NULL,
        guruh_kaliti TEXT NOT NULL,
        oqituvchi_user_id BIGINT REFERENCES users(user_id),
        xona_id BIGINT REFERENCES aqlli_xonalar_v2(id),
        PRIMARY KEY(maktab_id,sinf_id,fan_nomi,guruh_kaliti)
    )""")
    cur.execute("""CREATE TABLE IF NOT EXISTS aqlli_jadval_urinishlari_v2(
        id BIGSERIAL PRIMARY KEY,
        maktab_id INTEGER NOT NULL REFERENCES maktablar(id) ON DELETE CASCADE,
        holat TEXT NOT NULL DEFAULT 'draft' CHECK(holat IN ('draft','tasdiqlangan','almashtirilgan','bekor')),
        yaratgan_user_id BIGINT,
        yaratilgan_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
        tasdiqlangan_at TIMESTAMPTZ,
        sifat INTEGER NOT NULL DEFAULT 0,
        joylashtirildi INTEGER NOT NULL DEFAULT 0,
        joylashtirilmadi INTEGER NOT NULL DEFAULT 0,
        diagnostika JSONB NOT NULL DEFAULT '{}'::jsonb,
        sozlamalar JSONB NOT NULL DEFAULT '{}'::jsonb
    )""")
    cur.execute("""CREATE TABLE IF NOT EXISTS aqlli_jadval_jarayoni_v2243(
        maktab_id INTEGER PRIMARY KEY REFERENCES maktablar(id) ON DELETE CASCADE,
        qidiruv_nonce BIGINT,
        jadval_raqami BIGINT NOT NULL,
        yaxshilanish INTEGER NOT NULL DEFAULT 0,
        foiz INTEGER NOT NULL DEFAULT 0 CHECK(foiz BETWEEN 0 AND 100),
        bosqich TEXT NOT NULL,
        xabar TEXT NOT NULL,
        yangilangan_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
    )""")
    cur.execute("""ALTER TABLE aqlli_jadval_jarayoni_v2243
                   ADD COLUMN IF NOT EXISTS toxtatish_soraldi BOOLEAN NOT NULL DEFAULT FALSE""")
    # Client yuborgan millisekund/nonce bir nechta brauzer yoki qayta yuborilgan
    # HTTP so'rovida takrorlanishi mumkin. Jarayon lease'i faqat PostgreSQL
    # ajratadigan, butun baza bo'yicha monoton va qayta ishlatilmaydigan qiymat.
    # Yuqori START eski client nonce'lari bilan tasodifiy ustma-ust tushishni ham
    # amalda yo'q qiladi; qidiruv_nonce public maydoni compatibility uchun qoladi.
    cur.execute("""CREATE SEQUENCE IF NOT EXISTS aqlli_jadval_qidiruv_lease_v237
                   AS BIGINT START WITH 1000000000000000
                   MINVALUE 1000000000000000
                   MAXVALUE 9007199254740991 NO CYCLE""")
    cur.execute("""ALTER SEQUENCE aqlli_jadval_qidiruv_lease_v237
                   MINVALUE 1000000000000000
                   MAXVALUE 9007199254740991 NO CYCLE""")
    cur.execute("""CREATE TABLE IF NOT EXISTS aqlli_jadval_revisionlari_v2258(
        maktab_id INTEGER NOT NULL REFERENCES maktablar(id) ON DELETE CASCADE,
        urinish_id BIGINT NOT NULL REFERENCES aqlli_jadval_urinishlari_v2(id) ON DELETE CASCADE,
        revision INTEGER NOT NULL CHECK(revision >= 0),
        yaratilgan_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
        sifat INTEGER NOT NULL DEFAULT 0,
        bosqich TEXT NOT NULL,
        metrics JSONB NOT NULL DEFAULT '{}'::jsonb,
        slotlar JSONB NOT NULL DEFAULT '[]'::jsonb,
        PRIMARY KEY(urinish_id,revision)
    )""")
    cur.execute("""ALTER TABLE aqlli_jadval_revisionlari_v2258
                   ADD COLUMN IF NOT EXISTS yaxshilanish JSONB NOT NULL DEFAULT '{}'::jsonb""")
    cur.execute("""CREATE TABLE IF NOT EXISTS aqlli_jadval_slotlari_v2(
        id BIGSERIAL PRIMARY KEY,
        urinish_id BIGINT NOT NULL REFERENCES aqlli_jadval_urinishlari_v2(id) ON DELETE CASCADE,
        maktab_id INTEGER NOT NULL REFERENCES maktablar(id) ON DELETE CASCADE,
        sinf_id INTEGER NOT NULL REFERENCES maktab_sinflari(id) ON DELETE CASCADE,
        hafta_kuni INTEGER NOT NULL CHECK(hafta_kuni BETWEEN 1 AND 7),
        smena INTEGER NOT NULL CHECK(smena IN (1,2)),
        dars_raqami INTEGER NOT NULL CHECK(dars_raqami BETWEEN 1 AND 12),
        fan_nomi TEXT NOT NULL,
        oqituvchi_user_id BIGINT REFERENCES users(user_id),
        guruh_kaliti TEXT NOT NULL DEFAULT 'whole',
        xona_id BIGINT REFERENCES aqlli_xonalar_v2(id),
        xona_matni TEXT,
        boshlanish_vaqti TEXT,
        tugash_vaqti TEXT,
        yuklama_id BIGINT REFERENCES aqlli_sinf_fan_yuklamalari_v2(id),
        takror_raqami INTEGER NOT NULL DEFAULT 1,
        UNIQUE(urinish_id,sinf_id,hafta_kuni,smena,dars_raqami,guruh_kaliti)
    )""")
    # 0,5 soat — yarimta dars emas: bitta slotda toq/juft haftalar
    # almashadi. Ikki aylanish qatori bir xil kun va dars raqamida
    # qonuniy saqlanishi uchun hafta turi unikal kalitga kiradi.
    cur.execute("""ALTER TABLE aqlli_jadval_slotlari_v2
                   ADD COLUMN IF NOT EXISTS hafta_turi TEXT NOT NULL DEFAULT 'har_hafta'""")
    cur.execute("""DO $$
        DECLARE constraint_name TEXT;
        BEGIN
          FOR constraint_name IN
            SELECT c.conname
            FROM pg_constraint c
            JOIN pg_class t ON t.oid=c.conrelid
            WHERE t.relname='aqlli_jadval_slotlari_v2'
              AND c.contype='u'
              AND pg_get_constraintdef(c.oid) ILIKE '%urinish_id%sinf_id%hafta_kuni%smena%dars_raqami%guruh_kaliti%'
          LOOP
            EXECUTE format('ALTER TABLE aqlli_jadval_slotlari_v2 DROP CONSTRAINT %I', constraint_name);
          END LOOP;
        END $$""")
    cur.execute("""CREATE UNIQUE INDEX IF NOT EXISTS uq_aqlli_slot_rotation_v2
                   ON aqlli_jadval_slotlari_v2(
                     urinish_id,sinf_id,hafta_kuni,smena,dars_raqami,guruh_kaliti,hafta_turi
                   )""")
    cur.execute("""CREATE TABLE IF NOT EXISTS aqlli_mavzu_rejalari_v2(
        id BIGSERIAL PRIMARY KEY,
        maktab_id INTEGER NOT NULL REFERENCES maktablar(id) ON DELETE CASCADE,
        sinf_id INTEGER NOT NULL REFERENCES maktab_sinflari(id) ON DELETE CASCADE,
        fan_nomi TEXT NOT NULL,
        chorak INTEGER NOT NULL CHECK(chorak BETWEEN 1 AND 4),
        tartib INTEGER NOT NULL,
        mavzu TEXT NOT NULL,
        soat INTEGER NOT NULL DEFAULT 1 CHECK(soat BETWEEN 1 AND 10),
        turi TEXT NOT NULL DEFAULT 'mavzu' CHECK(turi IN ('mavzu','nazorat','xato_tahlil','mustahkamlash','masala')),
        topic_code TEXT,
        manba TEXT NOT NULL DEFAULT 'qolda' CHECK(manba IN ('qolda','dts')),
        UNIQUE(maktab_id,sinf_id,fan_nomi,chorak,tartib)
    )""")
    cur.execute("""CREATE TABLE IF NOT EXISTS aqlli_mavzu_taqvimi_v2(
        id BIGSERIAL PRIMARY KEY,
        maktab_id INTEGER NOT NULL REFERENCES maktablar(id) ON DELETE CASCADE,
        urinish_id BIGINT REFERENCES aqlli_jadval_urinishlari_v2(id) ON DELETE SET NULL,
        slot_id BIGINT REFERENCES aqlli_jadval_slotlari_v2(id) ON DELETE SET NULL,
        sinf_id INTEGER NOT NULL REFERENCES maktab_sinflari(id) ON DELETE CASCADE,
        fan_nomi TEXT NOT NULL,
        chorak INTEGER NOT NULL CHECK(chorak BETWEEN 1 AND 4),
        sana DATE NOT NULL,
        hafta_kuni INTEGER NOT NULL,
        smena INTEGER NOT NULL,
        dars_raqami INTEGER NOT NULL,
        tartib INTEGER NOT NULL,
        mavzu TEXT NOT NULL,
        turi TEXT NOT NULL DEFAULT 'mavzu',
        manba TEXT NOT NULL DEFAULT 'avto' CHECK(manba IN ('avto','oqituvchi')),
        qulflangan BOOLEAN NOT NULL DEFAULT FALSE,
        oqituvchi_user_id BIGINT,
        UNIQUE(maktab_id,sinf_id,fan_nomi,chorak,sana,smena,dars_raqami)
    )""")
    cur.execute("""CREATE TABLE IF NOT EXISTS aqlli_jadval_ozgartirish_sorovlari_v2(
        id BIGSERIAL PRIMARY KEY,
        maktab_id INTEGER NOT NULL REFERENCES maktablar(id) ON DELETE CASCADE,
        slot_id BIGINT NOT NULL REFERENCES aqlli_jadval_slotlari_v2(id) ON DELETE CASCADE,
        soragan_user_id BIGINT NOT NULL REFERENCES users(user_id),
        yangi_hafta_kuni INTEGER NOT NULL,
        yangi_smena INTEGER NOT NULL,
        yangi_dars_raqami INTEGER NOT NULL,
        izoh TEXT,
        holat TEXT NOT NULL DEFAULT 'kutilmoqda' CHECK(holat IN ('kutilmoqda','tasdiqlandi','rad_etildi')),
        korib_chiqgan_user_id BIGINT,
        yaratilgan_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
        yangilangan_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
    )""")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_aqlli_kalendar_v2 ON aqlli_kalendar_kunlari_v2(maktab_id,sana)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_aqlli_slot_teacher_v2 ON aqlli_jadval_slotlari_v2(urinish_id,oqituvchi_user_id,hafta_kuni,dars_raqami)")
    cur.execute("""CREATE TABLE IF NOT EXISTS aqlli_sinf_kun_bloklari_v2(
        id BIGSERIAL PRIMARY KEY,
        maktab_id INTEGER NOT NULL REFERENCES maktablar(id) ON DELETE CASCADE,
        sinf_id INTEGER REFERENCES maktab_sinflari(id) ON DELETE CASCADE,
        sinf_daraja INTEGER,
        hafta_kuni INTEGER NOT NULL CHECK(hafta_kuni BETWEEN 1 AND 6),
        izoh TEXT,
        faol BOOLEAN NOT NULL DEFAULT TRUE,
        yaratgan_user_id BIGINT,
        yangilangan_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
        CHECK ((sinf_id IS NOT NULL AND sinf_daraja IS NULL) OR
               (sinf_id IS NULL AND sinf_daraja BETWEEN 1 AND 11))
    )""")
    cur.execute("""CREATE UNIQUE INDEX IF NOT EXISTS uq_aqlli_sinf_kun_blok_exact_v2
                   ON aqlli_sinf_kun_bloklari_v2(maktab_id,sinf_id,hafta_kuni)
                   WHERE sinf_id IS NOT NULL""")
    cur.execute("""CREATE UNIQUE INDEX IF NOT EXISTS uq_aqlli_sinf_kun_blok_parallel_v2
                   ON aqlli_sinf_kun_bloklari_v2(maktab_id,sinf_daraja,hafta_kuni)
                   WHERE sinf_daraja IS NOT NULL""")
    cur.execute("""CREATE TABLE IF NOT EXISTS aqlli_sinf_kun_blok_seed_v2(
        maktab_id INTEGER PRIMARY KEY REFERENCES maktablar(id) ON DELETE CASCADE,
        versiya INTEGER NOT NULL DEFAULT 1,
        yaratilgan_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
    )""")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_aqlli_slot_class_v2 ON aqlli_jadval_slotlari_v2(urinish_id,sinf_id,hafta_kuni,dars_raqami)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_aqlli_topic_calendar_v2 ON aqlli_mavzu_taqvimi_v2(maktab_id,sinf_id,fan_nomi,chorak,sana)")
    cur.execute("""CREATE TABLE IF NOT EXISTS aqlli_sinf_soati_qoidalari_v2(
        id BIGSERIAL PRIMARY KEY,
        maktab_id INTEGER NOT NULL REFERENCES maktablar(id) ON DELETE CASCADE,
        sinf_id INTEGER NOT NULL REFERENCES maktab_sinflari(id) ON DELETE CASCADE,
        hafta_kuni INTEGER NOT NULL CHECK(hafta_kuni BETWEEN 1 AND 6),
        dars_raqami INTEGER NOT NULL CHECK(dars_raqami BETWEEN 1 AND 12),
        faol BOOLEAN NOT NULL DEFAULT TRUE,
        yaratgan_user_id BIGINT,
        yangilangan_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
        UNIQUE(maktab_id,sinf_id)
    )""")
    cur.execute("ALTER TABLE aqlli_sinf_soati_qoidalari_v2 ADD COLUMN IF NOT EXISTS fan_nomi TEXT NOT NULL DEFAULT 'KELAJAK SOATI'")
    cur.execute("ALTER TABLE aqlli_sinf_soati_qoidalari_v2 ADD COLUMN IF NOT EXISTS haftalik_soat INTEGER NOT NULL DEFAULT 1")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_aqlli_sinf_soati_v2 ON aqlli_sinf_soati_qoidalari_v2(maktab_id,hafta_kuni,dars_raqami)")
    _V1852_TABLES_READY = True


def _v1852_tables(cur):
    if _V1852_TABLES_READY:
        return
    with _V1852_TABLES_LOCK:
        if not _V1852_TABLES_READY:
            _v1852_create_tables(cur)


@app.on_event("startup")
def _v1852_startup_tables():
    global _V1852_TABLES_READY
    try:
        conn = _db(); cur = conn.cursor()
        _v1852_create_tables(cur)
        conn.commit(); cur.close(); conn.close()
        _V1852_TABLES_READY = True
    except Exception as exc:
        _V1852_TABLES_READY = False
        print(f"[V18.52 jadval jadvallari] {exc}")


def _v1852_manager(cur, user_id: int, maktab_id: int) -> bool:
    # _maktab_boshqaruvchi_mi umumiy adminni ham tekshiradi.
    return _maktab_boshqaruvchi_mi(cur, user_id, maktab_id)


def _v1852_staff(cur, user_id: int, maktab_id: int) -> bool:
    return _v1852_manager(cur, user_id, maktab_id) or _maktab_xodimi_mi(cur, user_id, maktab_id)


def _v1852_teacher_subject_allowed(cur, user_id: int, maktab_id: int, sinf_id: int, fan: str) -> bool:
    if _v1852_manager(cur, user_id, maktab_id):
        return True
    cur.execute("""SELECT 1 FROM maktab_dars_birikmalari
                   WHERE maktab_id=%s AND user_id=%s AND sinf_id=%s
                     AND LOWER(TRIM(fan_nomi))=LOWER(TRIM(%s)) LIMIT 1""",
                (maktab_id, user_id, sinf_id, fan))
    return cur.fetchone() is not None


def _v1852_time_str(value) -> str:
    return value.strftime("%H:%M") if hasattr(value, "strftime") else str(value or "")[:5]


def _v1852_shift_slots(row: dict) -> list[dict]:
    start_text = _v1852_time_str(row.get("boshlanish_vaqti") or "08:00")
    current = datetime.strptime(start_text, "%H:%M")
    count = int(row.get("dars_soni") or 7)
    lesson = int(row.get("dars_daqiqa") or 45)
    normal_break = int(row.get("tanaffus_daqiqa") or 5)
    big_after = row.get("katta_tanaffus_darsdan_keyin")
    big_break = int(row.get("katta_tanaffus_daqiqa") or 15)
    result = []
    for number in range(1, count + 1):
        end = current + timedelta(minutes=lesson)
        result.append({"dars_raqami": number, "boshlanish": current.strftime("%H:%M"), "tugash": end.strftime("%H:%M")})
        if number < count:
            pause = big_break if big_after and number == int(big_after) else normal_break
            current = end + timedelta(minutes=pause)
    return result


def _v1852_default_shifts(cur, maktab_id: int, shift_count=None):
    """Maktab e'lon qilgan smenalar uchungina standart vaqtlarni yaratadi.

    ``shift_count`` berilmagan eski chaqiruvlar maktabning ``smena_soni``
    qiymatini bazadan oladi. Avval yaratilgan ortiqcha smena qatorlari ataylab
    o'chirilmaydi: tarixiy sozlamani yo'qotmasdan, faqat yangi noto'g'ri qator
    qo'shilishining oldi olinadi.
    """
    if shift_count is None:
        cur.execute(
            "SELECT COALESCE(smena_soni,1) AS smena_soni FROM maktablar WHERE id=%s",
            (maktab_id,),
        )
        school_row = cur.fetchone()
        if isinstance(school_row, dict):
            shift_count = school_row.get("smena_soni")
        elif school_row:
            shift_count = school_row[0]
    try:
        shift_count = int(shift_count or 1)
    except (TypeError, ValueError):
        shift_count = 1
    if shift_count not in (1, 2):
        shift_count = 1

    defaults = [
        (1, 7, "08:00", 45, 5, 3, 15),
        (2, 7, "13:30", 45, 5, 3, 15),
    ]
    for smena, dars_soni, start, lesson, pause, big_after, big_pause in defaults[:shift_count]:
        cur.execute("""INSERT INTO aqlli_smena_sozlamalari_v2(
            maktab_id,smena,dars_soni,boshlanish_vaqti,dars_daqiqa,tanaffus_daqiqa,
            katta_tanaffus_darsdan_keyin,katta_tanaffus_daqiqa)
            VALUES(%s,%s,%s,%s,%s,%s,%s,%s) ON CONFLICT(maktab_id,smena) DO NOTHING""",
            (maktab_id, smena, dars_soni, start, lesson, pause, big_after, big_pause))


def _v1852_active_year(cur, maktab_id: int):
    cur.execute("SELECT * FROM aqlli_oquv_yillari_v2 WHERE maktab_id=%s AND faol=TRUE ORDER BY id DESC LIMIT 1", (maktab_id,))
    return cur.fetchone()


def _v1890_generation_year(cur, maktab_id: int):
    """Jadval uchun o'quv yilini majburiy qilmaydigan ichki vaqt oralig'i."""
    year = _v1852_active_year(cur, maktab_id)
    if year:
        return year
    today = date.today()
    start_year = today.year if today.month >= 7 else today.year - 1
    start = date(start_year, 9, 1)
    end = date(start_year + 1, 5, 31)
    name = f"__JADVAL_VAQTINCHA_{start_year}_{start_year + 1}__"
    cur.execute("""INSERT INTO aqlli_oquv_yillari_v2(
        maktab_id,nomi,boshlanish,tugash,hafta_kunlari,faol,yangilangan_at)
        VALUES(%s,%s,%s,%s,6,TRUE,NOW())
        ON CONFLICT(maktab_id,nomi) DO UPDATE SET faol=TRUE,yangilangan_at=NOW()
        RETURNING *""", (maktab_id, name, start, end))
    year = cur.fetchone()
    quarters = [
        (1, start, date(start_year, 10, 31)),
        (2, date(start_year, 11, 1), date(start_year, 12, 31)),
        (3, date(start_year + 1, 1, 1), date(start_year + 1, 3, 31)),
        (4, date(start_year + 1, 4, 1), end),
    ]
    for number, q_start, q_end in quarters:
        cur.execute("""INSERT INTO aqlli_choraklar_v2(
            oquv_yili_id,chorak,boshlanish,tugash,holat)
            VALUES(%s,%s,%s,%s,'taxminiy')
            ON CONFLICT(oquv_yili_id,chorak) DO NOTHING""",
            (year["id"], number, q_start, q_end))
    return year


def _v1852_active_run(cur, maktab_id: int):
    cur.execute("SELECT * FROM aqlli_jadval_urinishlari_v2 WHERE maktab_id=%s AND holat='tasdiqlangan' ORDER BY id DESC LIMIT 1", (maktab_id,))
    return cur.fetchone()


def _v1852_is_school_day(cur, maktab_id: int, sana: date, weekdays: int) -> bool:
    cur.execute("SELECT turi FROM aqlli_kalendar_kunlari_v2 WHERE maktab_id=%s AND sana=%s", (maktab_id, sana))
    row = cur.fetchone()
    if row:
        return row["turi"] in _V1852_OQISH_TURLARI
    return sana.isoweekday() <= weekdays


def _v1852_quarter_for_date(cur, year_id: int, sana: date):
    cur.execute("SELECT chorak FROM aqlli_choraklar_v2 WHERE oquv_yili_id=%s AND %s BETWEEN boshlanish AND tugash", (year_id, sana))
    row = cur.fetchone()
    return int(row["chorak"]) if row else None


def _v1852_validate_quarters(start: date, end: date, quarters: list[dict]):
    parsed = []
    seen = set()
    for item in quarters:
        number = int(item.get("chorak") or 0)
        if number not in (1, 2, 3, 4) or number in seen:
            raise HTTPException(status_code=400, detail="Choraklar 1, 2, 3, 4 bo'lib takrorlanmasligi kerak")
        seen.add(number)
        try:
            q_start = date.fromisoformat(str(item.get("boshlanish")))
            q_end = date.fromisoformat(str(item.get("tugash")))
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"{number}-chorak sanasi noto'g'ri") from exc
        status = str(item.get("holat") or "taxminiy")
        if status not in _V1852_HOLATLAR or q_end < q_start:
            raise HTTPException(status_code=400, detail=f"{number}-chorak sanasi/holatini tekshiring")
        if q_start < start or q_end > end:
            raise HTTPException(status_code=400, detail=f"{number}-chorak o'quv yili chegarasidan chiqdi")
        parsed.append((number, q_start, q_end, status))
    if seen != {1, 2, 3, 4}:
        raise HTTPException(status_code=400, detail="Barcha 4 chorakni kiriting")
    parsed.sort(key=lambda x: x[0])
    for current, following in zip(parsed, parsed[1:]):
        if following[1] <= current[2]:
            raise HTTPException(status_code=400, detail=f"{current[0]} va {following[0]}-chorak sanalari ustma-ust tushdi")
    return parsed


def _v1852_previous_weekday(value: date, max_weekday: int = 6):
    result = value
    while result.isoweekday() > max_weekday:
        result -= timedelta(days=1)
    return result


def _v1852_next_weekday(value: date, max_weekday: int = 6):
    result = value
    while result.isoweekday() > max_weekday:
        result += timedelta(days=1)
    return result


def _v1852_suggest_calendar(start: date, end: date, weekdays: int):
    max_day = 5 if weekdays == 5 else 6
    year = start.year
    q1_end = _v1852_previous_weekday(date(year, 10, 31), max_day)
    q2_start = _v1852_next_weekday(q1_end + timedelta(days=7), max_day)
    q2_end = _v1852_previous_weekday(date(year, 12, 27), max_day)
    q3_start = _v1852_next_weekday(date(year + 1, 1, 11), max_day)
    q3_end = _v1852_previous_weekday(date(year + 1, 3, 20), max_day)
    q4_start = _v1852_next_weekday(q3_end + timedelta(days=8), max_day)
    boundaries = [
        (1, start, min(q1_end, end)),
        (2, max(q2_start, start), min(q2_end, end)),
        (3, max(q3_start, start), min(q3_end, end)),
        (4, max(q4_start, start), end),
    ]
    # Juda qisqa yoki noodatiy o'quv yili bo'lsa teng bo'lib taqsimlaymiz.
    if any(qe < qs for _, qs, qe in boundaries):
        total = max(4, (end - start).days + 1)
        step = total // 4
        boundaries = []
        cursor = start
        for number in range(1, 5):
            q_end = end if number == 4 else min(end, cursor + timedelta(days=step - 1))
            boundaries.append((number, cursor, q_end))
            cursor = q_end + timedelta(days=1)
    holidays = []
    common = [
        (1, 1, "Yangi yil"), (3, 8, "Xotin-qizlar kuni"), (3, 21, "Navro'z"),
        (5, 9, "Xotira va qadrlash kuni"), (9, 1, "Mustaqillik kuni"),
        (10, 1, "O'qituvchi va murabbiylar kuni"), (12, 8, "Konstitutsiya kuni"),
    ]
    for yy in range(start.year, end.year + 1):
        for month, day, name in common:
            try:
                d = date(yy, month, day)
            except ValueError:
                continue
            if start <= d <= end:
                holidays.append({"sana": d.isoformat(), "turi": "bayram", "nomi": name, "holat": "taxminiy"})
    return {
        "choraklar": [
            {"chorak": n, "boshlanish": qs.isoformat(), "tugash": qe.isoformat(), "holat": "taxminiy"}
            for n, qs, qe in boundaries
        ],
        "maxsus_kunlar": holidays,
        "ogohlantirish": "Bu sanalar faqat taxminiy tavsiya. Rasmiy kalendar/Kundalik bilan tekshirib tasdiqlang.",
    }


class V1852CalendarSuggest(BaseModel):
    boshlanish: date
    tugash: date
    hafta_kunlari: int = 6


@app.post("/api/maktab/aqlli_jadval/v2/kalendar_tavsiya")
def v1852_calendar_suggest(sorov: V1852CalendarSuggest, token: str):
    _jwt_tekshir(token)
    if sorov.hafta_kunlari not in (5, 6) or sorov.tugash <= sorov.boshlanish:
        raise HTTPException(status_code=400, detail="Sanalarni va 5/6 kunlik haftani tekshiring")
    return _v1852_suggest_calendar(sorov.boshlanish, sorov.tugash, sorov.hafta_kunlari)


class V1852CalendarSave(BaseModel):
    maktab_id: int
    nomi: str
    boshlanish: date
    tugash: date
    hafta_kunlari: int = 6
    choraklar: list[dict]
    smenalar: list[dict] = []


@app.put("/api/maktab/aqlli_jadval/v2/kalendar")
def v1852_calendar_save(sorov: V1852CalendarSave, token: str):
    user_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor()
    try:
        _v192_tables(cur)
        if not _v1852_manager(cur, user_id, sorov.maktab_id):
            raise HTTPException(status_code=403, detail="Kalendarni faqat maktab rahbariyati boshqaradi")
        if sorov.hafta_kunlari not in (5, 6) or sorov.tugash <= sorov.boshlanish:
            raise HTTPException(status_code=400, detail="O'quv yili sanalarini tekshiring")
        quarters = _v1852_validate_quarters(sorov.boshlanish, sorov.tugash, sorov.choraklar)
        cur.execute("UPDATE aqlli_oquv_yillari_v2 SET faol=FALSE WHERE maktab_id=%s AND nomi<>%s", (sorov.maktab_id, sorov.nomi.strip()))
        cur.execute("""INSERT INTO aqlli_oquv_yillari_v2(
            maktab_id,nomi,boshlanish,tugash,hafta_kunlari,faol,yaratgan_user_id,yangilangan_at)
            VALUES(%s,%s,%s,%s,%s,TRUE,%s,NOW())
            ON CONFLICT(maktab_id,nomi) DO UPDATE SET boshlanish=EXCLUDED.boshlanish,
              tugash=EXCLUDED.tugash,hafta_kunlari=EXCLUDED.hafta_kunlari,
              faol=TRUE,yaratgan_user_id=EXCLUDED.yaratgan_user_id,yangilangan_at=NOW()
            RETURNING id""", (sorov.maktab_id, sorov.nomi.strip(), sorov.boshlanish,
                                sorov.tugash, sorov.hafta_kunlari, user_id))
        year_id = cur.fetchone()["id"]
        cur.execute("DELETE FROM aqlli_choraklar_v2 WHERE oquv_yili_id=%s", (year_id,))
        psycopg2.extras.execute_values(cur,
            "INSERT INTO aqlli_choraklar_v2(oquv_yili_id,chorak,boshlanish,tugash,holat) VALUES %s",
            [(year_id, number, start, end, status) for number, start, end, status in quarters])
        _v1852_default_shifts(cur, sorov.maktab_id)
        for shift in sorov.smenalar:
            smena = int(shift.get("smena") or 0)
            if smena not in (1, 2):
                raise HTTPException(status_code=400, detail="Smena 1 yoki 2 bo'lishi kerak")
            cur.execute("""INSERT INTO aqlli_smena_sozlamalari_v2(
                maktab_id,smena,dars_soni,boshlanish_vaqti,dars_daqiqa,tanaffus_daqiqa,
                katta_tanaffus_darsdan_keyin,katta_tanaffus_daqiqa)
                VALUES(%s,%s,%s,%s,%s,%s,%s,%s)
                ON CONFLICT(maktab_id,smena) DO UPDATE SET dars_soni=EXCLUDED.dars_soni,
                  boshlanish_vaqti=EXCLUDED.boshlanish_vaqti,dars_daqiqa=EXCLUDED.dars_daqiqa,
                  tanaffus_daqiqa=EXCLUDED.tanaffus_daqiqa,
                  katta_tanaffus_darsdan_keyin=EXCLUDED.katta_tanaffus_darsdan_keyin,
                  katta_tanaffus_daqiqa=EXCLUDED.katta_tanaffus_daqiqa""",
                (sorov.maktab_id, smena, int(shift.get("dars_soni") or 7),
                 str(shift.get("boshlanish_vaqti") or ("08:00" if smena == 1 else "13:30")),
                 int(shift.get("dars_daqiqa") or 45), int(shift.get("tanaffus_daqiqa") or 5),
                 int(shift.get("katta_tanaffus_darsdan_keyin") or 3),
                 int(shift.get("katta_tanaffus_daqiqa") or 15)))
        conn.commit()
        return {"holat": "saqlandi", "oquv_yili_id": year_id}
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


def _v1852_rebuild_all_topic_calendars(cur, maktab_id: int):
    cur.execute("""SELECT DISTINCT maktab_id,sinf_id,fan_nomi,chorak
                   FROM aqlli_mavzu_rejalari_v2 WHERE maktab_id=%s""", (maktab_id,))
    groups = cur.fetchall()
    rebuilt = 0
    warnings = []
    for group in groups:
        result = _v1852_distribute_topics_one(cur, maktab_id, group["sinf_id"], group["fan_nomi"], group["chorak"])
        rebuilt += 1
        warnings.extend(result.get("ogohlantirishlar") or [])
    return {"qayta_taqsimlandi": rebuilt, "ogohlantirishlar": warnings[:50]}


class V1852SpecialDay(BaseModel):
    maktab_id: int
    sana: date
    turi: str
    nomi: Optional[str] = None
    holat: str = "tasdiqlangan"
    izoh: Optional[str] = None


@app.put("/api/maktab/aqlli_jadval/v2/maxsus_kun")
def v1852_special_day_save(sorov: V1852SpecialDay, token: str):
    user_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor()
    try:
        _v1852_tables(cur)
        if not _v1852_manager(cur, user_id, sorov.maktab_id):
            raise HTTPException(status_code=403, detail="Maxsus kunni faqat maktab rahbariyati belgilaydi")
        if sorov.turi not in _V1852_KUN_TURLARI or sorov.holat not in _V1852_HOLATLAR:
            raise HTTPException(status_code=400, detail="Kun turi yoki holati noto'g'ri")
        cur.execute("""INSERT INTO aqlli_kalendar_kunlari_v2(
            maktab_id,sana,turi,nomi,holat,izoh,yaratgan_user_id,yangilangan_at)
            VALUES(%s,%s,%s,%s,%s,%s,%s,NOW())
            ON CONFLICT(maktab_id,sana) DO UPDATE SET turi=EXCLUDED.turi,nomi=EXCLUDED.nomi,
              holat=EXCLUDED.holat,izoh=EXCLUDED.izoh,yaratgan_user_id=EXCLUDED.yaratgan_user_id,
              yangilangan_at=NOW()""", (sorov.maktab_id, sorov.sana, sorov.turi,
                                          sorov.nomi, sorov.holat, sorov.izoh, user_id))
        rebuild = _v1852_rebuild_all_topic_calendars(cur, sorov.maktab_id)
        conn.commit()
        return {"holat": "saqlandi", **rebuild}
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


@app.delete("/api/maktab/aqlli_jadval/v2/maxsus_kun")
def v1852_special_day_delete(token: str, maktab_id: int, sana: date):
    user_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor()
    try:
        _v1852_tables(cur)
        if not _v1852_manager(cur, user_id, maktab_id):
            raise HTTPException(status_code=403, detail="Ruxsat yo'q")
        cur.execute("DELETE FROM aqlli_kalendar_kunlari_v2 WHERE maktab_id=%s AND sana=%s", (maktab_id, sana))
        rebuild = _v1852_rebuild_all_topic_calendars(cur, maktab_id)
        conn.commit(); return {"holat": "o'chirildi", **rebuild}
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


def _v205_room_normalized_name(value):
    """Xona nomini generator, katalog va saqlangan slotlar uchun birxillashtiradi."""
    return re.sub(r"\s+", " ", str(value or "")).strip().casefold()


def _v205_split_room_names(value):
    """Vergul/nuqtali vergul/yangi qatordagi xonalarni alohida nomga ajratadi."""
    result = []
    seen = set()
    for raw in re.split(r"[;,\n]+", str(value or "")):
        name = re.sub(r"\s+", " ", raw).strip()
        key = _v205_room_normalized_name(name)
        if key and key not in seen:
            seen.add(key)
            result.append(name)
    return result


def _v205_persisted_room_key(room_id, room_text, catalog_name_to_id=None):
    """Saqlangan id yoki eski matndan bitta kanonik xona resursini qaytaradi."""
    if room_id is not None:
        return f"room:{int(room_id)}"
    normalized = _v205_room_normalized_name(room_text)
    if not normalized:
        return None
    mapped_id = (catalog_name_to_id or {}).get(normalized)
    return f"room:{int(mapped_id)}" if mapped_id is not None else f"text:{normalized}"


class V1852Room(BaseModel):
    maktab_id: int
    nomi: str
    turi: str = "classroom"
    sigim: Optional[int] = None


@app.post("/api/maktab/aqlli_jadval/v2/xona")
def v1852_room_save(sorov: V1852Room, token: str):
    user_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor()
    try:
        _v1852_tables(cur)
        if not _v1852_manager(cur, user_id, sorov.maktab_id):
            raise HTTPException(status_code=403, detail="Xonani faqat maktab rahbariyati qo'shadi")
        names = _v205_split_room_names(sorov.nomi)
        if not names:
            raise HTTPException(status_code=400, detail="Xona nomini kiriting")
        if len(names) > 100:
            raise HTTPException(status_code=400, detail="Bir amalda 100 tadan ortiq xona qo'shilmaydi")
        room_type = str(sorov.turi or "classroom").strip().lower()
        room_type = {"oddiy": "classroom", "maxsus": "reserve"}.get(room_type, room_type)
        if room_type not in {"classroom", "reserve", "sport", "non_teaching"}:
            raise HTTPException(status_code=400, detail="Xona turi noto'g'ri")
        teaching_enabled = room_type != "non_teaching"
        saved_rooms = []
        for name in names:
            cur.execute("""INSERT INTO aqlli_xonalar_v2(
                             maktab_id,nomi,turi,sigim,faol,darsga_yaroqli)
                           VALUES(%s,%s,%s,%s,TRUE,%s)
                           ON CONFLICT(maktab_id,nomi) DO UPDATE SET turi=EXCLUDED.turi,
                             sigim=EXCLUDED.sigim,faol=TRUE,
                             darsga_yaroqli=EXCLUDED.darsga_yaroqli
                           RETURNING id,nomi""",
                        (sorov.maktab_id, name, room_type, sorov.sigim, teaching_enabled))
            saved_rooms.append(dict(cur.fetchone()))
        room_ids = [int(room["id"]) for room in saved_rooms]
        conn.commit(); return {
            "holat": "saqlandi",
            # Eski frontend bitta xona qo'shganda shu maydonni kutadi.
            "xona_id": room_ids[0],
            "xona_idlar": room_ids,
            "xonalar": saved_rooms,
            "xona_soni": len(saved_rooms),
        }
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


class V1852TeacherRules(BaseModel):
    kunlik_max: int = 6
    ketma_ket_max: int = 4
    okno_max: int = 1
    afzal_smena: int = 0
    eng_erta_dars: int = 1
    eng_kech_dars: int = 12


class V1852TeacherAvailability(BaseModel):
    maktab_id: int
    user_id: int
    qoidalar: V1852TeacherRules
    vaqtlar: list[dict] = []


@app.put("/api/maktab/aqlli_jadval/v2/oqituvchi_vaqti")
def v1852_teacher_availability_save(sorov: V1852TeacherAvailability, token: str):
    actor_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor()
    try:
        _v1852_tables(cur)
        if actor_id != sorov.user_id and not _v1852_manager(cur, actor_id, sorov.maktab_id):
            raise HTTPException(status_code=403, detail="Faqat o'qituvchining o'zi yoki rahbariyat o'zgartira oladi")
        cur.execute("SELECT 1 FROM users WHERE user_id=%s AND maktab_id=%s", (sorov.user_id, sorov.maktab_id))
        if not cur.fetchone():
            raise HTTPException(status_code=404, detail="O'qituvchi shu maktabda topilmadi")
        rules = sorov.qoidalar
        if rules.eng_kech_dars < rules.eng_erta_dars:
            raise HTTPException(status_code=400, detail="Eng kech dars eng erta darsdan oldin bo'lmaydi")
        cur.execute("""INSERT INTO aqlli_oqituvchi_qoidalari_v2(
            maktab_id,user_id,kunlik_max,ketma_ket_max,okno_max,afzal_smena,eng_erta_dars,eng_kech_dars)
            VALUES(%s,%s,%s,%s,%s,%s,%s,%s)
            ON CONFLICT(maktab_id,user_id) DO UPDATE SET kunlik_max=EXCLUDED.kunlik_max,
              ketma_ket_max=EXCLUDED.ketma_ket_max,okno_max=EXCLUDED.okno_max,
              afzal_smena=EXCLUDED.afzal_smena,eng_erta_dars=EXCLUDED.eng_erta_dars,
              eng_kech_dars=EXCLUDED.eng_kech_dars""",
            (sorov.maktab_id, sorov.user_id, rules.kunlik_max, rules.ketma_ket_max,
             rules.okno_max, rules.afzal_smena, rules.eng_erta_dars, rules.eng_kech_dars))
        cur.execute("DELETE FROM aqlli_oqituvchi_vaqti_v2 WHERE maktab_id=%s AND user_id=%s", (sorov.maktab_id, sorov.user_id))
        rows = []
        for item in sorov.vaqtlar:
            day = int(item.get("hafta_kuni") or 0)
            shift = int(item.get("smena") or 0)
            period = int(item.get("dars_raqami") or 0)
            kind = str(item.get("turi") or "")
            hard = bool(item.get("qattiq", True))
            if day not in range(1, 8) or shift not in (0, 1, 2) or period not in range(0, 13) or kind not in _V1852_VAQT_TURLARI:
                raise HTTPException(status_code=400, detail="O'qituvchi vaqtlaridan biri noto'g'ri")
            if kind == "metod_kuni":
                shift, period = 0, 0
                hard = True
            rows.append((sorov.maktab_id, sorov.user_id, day, shift, period, kind, hard, item.get("izoh")))
        if rows:
            psycopg2.extras.execute_values(cur,
                """INSERT INTO aqlli_oqituvchi_vaqti_v2(
                    maktab_id,user_id,hafta_kuni,smena,dars_raqami,turi,qattiq,izoh) VALUES %s""", rows)
        conn.commit(); return {"holat": "saqlandi", "vaqt_soni": len(rows)}
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


@app.get("/api/maktab/aqlli_jadval/v2/mening_vaqtim")
def v1852_my_availability(token: str):
    user_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor()
    try:
        _v1852_tables(cur)
        cur.execute("SELECT maktab_id,full_name FROM users WHERE user_id=%s", (user_id,))
        user = cur.fetchone()
        if not user or not user.get("maktab_id"):
            return {"maktab_id": None, "oqituvchi": None, "qoidalar": None, "vaqtlar": []}
        cur.execute("SELECT * FROM aqlli_oqituvchi_qoidalari_v2 WHERE maktab_id=%s AND user_id=%s", (user["maktab_id"], user_id))
        rules = cur.fetchone()
        cur.execute("SELECT * FROM aqlli_oqituvchi_vaqti_v2 WHERE maktab_id=%s AND user_id=%s ORDER BY hafta_kuni,smena,dars_raqami", (user["maktab_id"], user_id))
        times = cur.fetchall()
        return {"maktab_id": user["maktab_id"], "oqituvchi": user["full_name"], "qoidalar": rules, "vaqtlar": times}
    finally:
        cur.close(); conn.close()


class V1852LoadItem(BaseModel):
    fan_nomi: str
    haftalik_soat: float
    kunlik_max: int = 1
    ketma_ket_mumkin: bool = False
    afzal_oxirgi_dars: int = 5
    asosiy_oqituvchi_user_id: Optional[int] = None
    xona_id: Optional[int] = None
    nazorat_soni: int = 0
    nazoratdan_keyin_tahlil: bool = True
    mustahkamlash_soni: int = 0
    ogirlik: int = 2


class V1852ClassLoads(BaseModel):
    maktab_id: int
    sinf_id: int
    fanlar: list[V1852LoadItem]


@app.put("/api/maktab/aqlli_jadval/v2/fan_soatlari")
def v1852_loads_save(sorov: V1852ClassLoads, token: str):
    user_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor()
    try:
        _v1852_tables(cur)
        if not _v1852_manager(cur, user_id, sorov.maktab_id):
            raise HTTPException(status_code=403, detail="Fan soatini faqat rahbariyat belgilaydi")
        cur.execute("SELECT 1 FROM maktab_sinflari WHERE id=%s AND maktab_id=%s", (sorov.sinf_id, sorov.maktab_id))
        if not cur.fetchone():
            raise HTTPException(status_code=404, detail="Sinf topilmadi")
        received = set()
        for item in sorov.fanlar:
            fan = re.sub(r"\s+", " ", item.fan_nomi or "").strip()
            if not fan:
                continue
            key = fan.casefold()
            if key in received:
                raise HTTPException(status_code=400, detail=f"'{fan}' ikki marta yuborilgan")
            received.add(key)
            cur.execute("""INSERT INTO aqlli_sinf_fan_yuklamalari_v2(
                maktab_id,sinf_id,fan_nomi,haftalik_soat,kunlik_max,ketma_ket_mumkin,
                afzal_oxirgi_dars,asosiy_oqituvchi_user_id,xona_id,nazorat_soni,
                nazoratdan_keyin_tahlil,mustahkamlash_soni,ogirlik)
                VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ON CONFLICT(maktab_id,sinf_id,fan_nomi) DO UPDATE SET
                  haftalik_soat=EXCLUDED.haftalik_soat,kunlik_max=EXCLUDED.kunlik_max,
                  ketma_ket_mumkin=EXCLUDED.ketma_ket_mumkin,
                  afzal_oxirgi_dars=EXCLUDED.afzal_oxirgi_dars,
                  asosiy_oqituvchi_user_id=EXCLUDED.asosiy_oqituvchi_user_id,
                  xona_id=EXCLUDED.xona_id,nazorat_soni=EXCLUDED.nazorat_soni,
                  nazoratdan_keyin_tahlil=EXCLUDED.nazoratdan_keyin_tahlil,
                  mustahkamlash_soni=EXCLUDED.mustahkamlash_soni,ogirlik=EXCLUDED.ogirlik""",
                (sorov.maktab_id, sorov.sinf_id, fan, item.haftalik_soat,
                 item.kunlik_max, item.ketma_ket_mumkin, item.afzal_oxirgi_dars,
                 item.asosiy_oqituvchi_user_id, item.xona_id, item.nazorat_soni,
                 item.nazoratdan_keyin_tahlil, item.mustahkamlash_soni, item.ogirlik))

            # Fan-soat oynasidagi o'zgarish exact DARS_BIRIKMALARI manbasiga ham yoziladi.
            cur.execute("""UPDATE maktab_dars_birikmalari
                           SET haftalik_soat=%s,kunlik_max=%s,manba='fan_soatlari_oynasi'
                           WHERE maktab_id=%s AND sinf_id=%s
                             AND LOWER(TRIM(fan_nomi))=LOWER(TRIM(%s))""",
                        (item.haftalik_soat, item.kunlik_max,
                         sorov.maktab_id, sorov.sinf_id, fan))
            if int(cur.rowcount or 0) == 0 and float(item.haftalik_soat or 0) > 0:
                if not item.asosiy_oqituvchi_user_id:
                    raise HTTPException(
                        status_code=400,
                        detail=f"{fan}: avval aniq o'qituvchini biriktiring",
                    )
                cur.execute("""INSERT INTO maktab_dars_birikmalari(
                                maktab_id,user_id,sinf_id,fan_nomi,guruh_kaliti,
                                haftalik_soat,kunlik_max,manba)
                               VALUES(%s,%s,%s,%s,'whole',%s,%s,'fan_soatlari_oynasi')
                               ON CONFLICT(user_id,sinf_id,fan_nomi,guruh_kaliti)
                               DO UPDATE SET haftalik_soat=EXCLUDED.haftalik_soat,
                                             kunlik_max=EXCLUDED.kunlik_max,
                                             manba=EXCLUDED.manba""",
                            (sorov.maktab_id, item.asosiy_oqituvchi_user_id,
                             sorov.sinf_id, fan, item.haftalik_soat, item.kunlik_max))

        manba_mosligi = _v1875_rebuild_schedule_sources(
            cur, sorov.maktab_id, cancel_drafts=True, reason="fan_soatlari_oynasi"
        )
        if manba_mosligi.get("xatolar"):
            raise HTTPException(
                status_code=400,
                detail="Fan-soat manbasi xatolari: " + "; ".join(manba_mosligi["xatolar"][:12]),
            )
        conn.commit()
        return {"holat": "saqlandi", "fan_soni": len(received), "moslik": manba_mosligi}
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


class V1852GroupSettingItem(BaseModel):
    fan_nomi: str
    guruh_kaliti: str
    oqituvchi_user_id: Optional[int] = None
    xona_id: Optional[int] = None


class V1852GroupSettings(BaseModel):
    maktab_id: int
    sinf_id: int
    guruhlar: list[V1852GroupSettingItem]
    # Backward-safe default: keyingi guruh turi oldingi 1/2 yoki O‘g‘il/Qiz
    # sozlamalarini o‘chirib yubormaydi. To‘liq almashtirish faqat UI aniq
    # tasdiqlab FALSE yuborgandagina bajariladi.
    oldingi_tizimlarni_saqlash: bool = True


@app.put("/api/maktab/aqlli_jadval/v2/guruh_sozlamalari")
def v1852_group_settings_save(sorov: V1852GroupSettings, token: str):
    user_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor()
    try:
        _v1852_tables(cur)
        if not _v1852_manager(cur, user_id, sorov.maktab_id):
            raise HTTPException(status_code=403, detail="Parallel guruhlarni faqat rahbariyat sozlaydi")
        if not sorov.oldingi_tizimlarni_saqlash:
            cur.execute(
                "DELETE FROM aqlli_guruh_sozlamalari_v2 "
                "WHERE maktab_id=%s AND sinf_id=%s",
                (sorov.maktab_id, sorov.sinf_id),
            )
        rows = []
        for item in sorov.guruhlar:
            fan = re.sub(r"\s+", " ", item.fan_nomi or "").strip()
            group_key = str(item.guruh_kaliti or "").strip()
            if not fan or not group_key or group_key == "whole":
                continue
            rows.append((sorov.maktab_id, sorov.sinf_id, fan, group_key, item.oqituvchi_user_id, item.xona_id))
        if rows:
            psycopg2.extras.execute_values(cur, """INSERT INTO aqlli_guruh_sozlamalari_v2(
                maktab_id,sinf_id,fan_nomi,guruh_kaliti,oqituvchi_user_id,xona_id)
                VALUES %s
                ON CONFLICT(maktab_id,sinf_id,fan_nomi,guruh_kaliti)
                DO UPDATE SET
                  oqituvchi_user_id=EXCLUDED.oqituvchi_user_id,
                  xona_id=EXCLUDED.xona_id""", rows)
        cur.execute(
            "SELECT COUNT(*) AS soni FROM aqlli_guruh_sozlamalari_v2 "
            "WHERE maktab_id=%s AND sinf_id=%s",
            (sorov.maktab_id, sorov.sinf_id),
        )
        active_count = int((cur.fetchone() or {}).get("soni") or 0)
        conn.commit(); return {
            "holat": "saqlandi",
            "guruh_soni": len(rows),
            "faol_guruh_soni": active_count,
            "oldingi_tizimlar_saqlandi": bool(sorov.oldingi_tizimlarni_saqlash),
        }
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()



# V18.59 — shablonda tanlangan fanlarni bir joyga yig'ish.
# Eski importlarda fan users.fanlari, maktab_xodim_sinflari yoki
# maktab_dars_birikmalari jadvallaridan faqat bittasida qolgan bo'lishi mumkin.
def _v1859_fanlarni_ajrat(value):
    text = str(value or "").replace("\\n", "\n")
    result = []
    seen = set()
    for part in re.split(r"[;\n,]+", text):
        fan = re.sub(r"\s+", " ", str(part or "")).strip()
        if not fan:
            continue
        key = _xodim_excel_sarlavha_kaliti(fan)
        if key and key not in seen:
            seen.add(key)
            result.append(fan)
    return result


def _v1859_sinf_sort_key(label):
    match = re.match(r"^\s*(\d+)\s*[-–]?\s*(.*)$", str(label or ""))
    if not match:
        return (999, str(label or ""))
    return (int(match.group(1)), match.group(2).casefold())


def _v1859_effective_teachers(cur, maktab_id: int, user_ids=None, include_numbered=False):
    cur.execute("""SELECT u.user_id,u.full_name,u.lavozim,u.fanlari,
                          u.oqitadigan_sinflari,u.haftalik_dars_soati,u.jadval_raqami,
                          to_jsonb(u)->>'mutaxassisligi' AS mutaxassisligi,
                          NULLIF(to_jsonb(u)->>'haftalik_maqsad_soat','')::NUMERIC(5,1)
                              AS haftalik_maqsad_soat,
                          NULLIF(to_jsonb(u)->>'tugilgan_sana','') AS tugilgan_sana,
                          CASE
                              WHEN COALESCE(to_jsonb(u)->>'ish_staji','') ~ '^\\d+$'
                              THEN (to_jsonb(u)->>'ish_staji')::INTEGER
                              ELSE NULL
                          END AS ish_staji,
                          NULLIF(to_jsonb(u)->>'toifasi','') AS toifasi
                   FROM users u
                   WHERE u.maktab_id=%s AND u.lavozim IS NOT NULL
                   ORDER BY u.full_name,u.user_id""", (maktab_id,))
    rows = [dict(row) for row in cur.fetchall()]
    if user_ids:
        wanted = {int(x) for x in user_ids}
        rows = [row for row in rows if int(row["user_id"]) in wanted]
    by_id = {int(row["user_id"]): row for row in rows}
    fan_maps = {}
    class_maps = {}
    source_maps = {}
    assignment_counts = {}
    for uid, row in by_id.items():
        fan_maps[uid] = {}
        class_maps[uid] = set()
        source_maps[uid] = set()
        assignment_counts[uid] = 0
        for fan in _v1859_fanlarni_ajrat(row.get("fanlari")):
            fan_maps[uid][_xodim_excel_sarlavha_kaliti(fan)] = fan
            source_maps[uid].add("XODIMLAR shabloni")
        for class_name in _xodim_sinf_royxatini_ajrat(row.get("oqitadigan_sinflari")):
            class_maps[uid].add(class_name)

    if not by_id:
        return []

    cur.execute("""SELECT b.user_id,b.sinf_id,b.fan_nomi,b.guruh_kaliti,
                          s.sinf,s.harf
                   FROM maktab_dars_birikmalari b
                   JOIN maktab_sinflari s ON s.id=b.sinf_id
                   WHERE b.maktab_id=%s
                   ORDER BY b.user_id,s.sinf::int,s.harf,b.fan_nomi""", (maktab_id,))
    for link in cur.fetchall():
        uid = int(link["user_id"])
        if uid not in by_id:
            continue
        fan = re.sub(r"\s+", " ", str(link.get("fan_nomi") or "")).strip()
        if fan:
            fan_maps[uid][_xodim_excel_sarlavha_kaliti(fan)] = fan
            source_maps[uid].add("DARS_BIRIKMALARI")
        class_maps[uid].add(f"{link['sinf']}-{link['harf']}")
        assignment_counts[uid] += 1

    cur.execute("""SELECT x.user_id,x.sinf_id,x.fanlari,s.sinf,s.harf
                   FROM maktab_xodim_sinflari x
                   JOIN maktab_sinflari s ON s.id=x.sinf_id
                   WHERE x.maktab_id=%s
                   ORDER BY x.user_id,s.sinf::int,s.harf""", (maktab_id,))
    for link in cur.fetchall():
        uid = int(link["user_id"])
        if uid not in by_id:
            continue
        for fan in _v1859_fanlarni_ajrat(link.get("fanlari")):
            fan_maps[uid][_xodim_excel_sarlavha_kaliti(fan)] = fan
            source_maps[uid].add("Sinf–fan birikmasi")
        class_maps[uid].add(f"{link['sinf']}-{link['harf']}")

    # Faqat shu maktabning haqiqiy dars jarayoniga biriktirilgan xodimlari
    # jadval sozlamasida ko'rinadi. Eski importdan qolgan "Akaunt 1" kabi,
    # maktab_id yozilgan bo'lsa-da birorta fan/sinf/rahbarlikka ulanmagan
    # foydalanuvchi boshqa maktab jadvaliga aralashmaydi.
    cur.execute(
        "SELECT DISTINCT rahbar_user_id AS user_id FROM maktab_sinflari "
        "WHERE maktab_id=%s AND rahbar_user_id IS NOT NULL",
        (maktab_id,),
    )
    class_head_ids = {int(row["user_id"]) for row in cur.fetchall()}

    result = []
    for uid, row in by_id.items():
        subjects = sorted(fan_maps[uid].values(), key=lambda x: x.casefold())
        classes = sorted(class_maps[uid], key=_v1859_sinf_sort_key)
        row["fanlar_royxati"] = subjects
        row["fanlari"] = "\n".join(subjects) or None
        row["sinflar_royxati"] = classes
        row["sinflari"] = "; ".join(classes) or None
        row["fan_manbalari"] = sorted(source_maps[uid])
        row["dars_birikma_soni"] = int(assignment_counts[uid])
        # Formada tanlangan, lekin hali sinf qatori berilmagan fanlar ham
        # mutaxassislik profilidan qaytadi. Canonical dars yuklamasi baribir
        # maktab_dars_birikmalari bo'lib qoladi.
        row["otadigan_fanlari"] = _v1859_fanlarni_ajrat(
            row.get("mutaxassisligi")
        )
        row["dars_beruvchi"] = bool(
            subjects or classes or assignment_counts[uid] or uid in class_head_ids
        )
        row["fan_holati"] = "aniqlandi" if subjects else "fan_topilmadi"
        # V22.49: F.I.Sh. bilan oldindan kiritilgan, hali skeletga fan biriktirilmagan
        # o'qituvchi ham ro'yxatda qoladi. Uning jadval_raqami keyingi skelet tanlovining kaliti.
        if row["dars_beruvchi"] or (include_numbered and row.get("jadval_raqami") is not None):
            result.append(row)
    return result


def _v1852_setup_payload(cur, maktab_id: int):
    _v1852_default_shifts(cur, maktab_id)
    year = _v1852_active_year(cur, maktab_id)
    quarters = []
    if year:
        cur.execute("SELECT * FROM aqlli_choraklar_v2 WHERE oquv_yili_id=%s ORDER BY chorak", (year["id"],))
        quarters = cur.fetchall()
    cur.execute("SELECT * FROM aqlli_kalendar_kunlari_v2 WHERE maktab_id=%s ORDER BY sana", (maktab_id,))
    special_days = cur.fetchall()
    cur.execute("SELECT * FROM aqlli_smena_sozlamalari_v2 WHERE maktab_id=%s ORDER BY smena", (maktab_id,))
    shifts = cur.fetchall()
    for shift in shifts:
        shift["slotlar"] = _v1852_shift_slots(shift)
    cur.execute("""SELECT s.id,s.sinf,s.harf,COALESCE(s.smena,1) AS smena,s.bino,s.xona,
                          COALESCE(s.talim_tili,'uz') AS talim_tili,
                          s.rahbar_user_id,COALESCE(u.full_name,'') AS rahbar_ismi
                   FROM maktab_sinflari s
                   LEFT JOIN users u ON u.user_id=s.rahbar_user_id
                   WHERE s.maktab_id=%s ORDER BY s.id""", (maktab_id,))
    classes = cur.fetchall()
    cur.execute(
        "SELECT COALESCE(alifbo_turi,'latin_xalqaro') AS alifbo_turi "
        "FROM maktablar WHERE id=%s",
        (maktab_id,),
    )
    alphabet_type = (cur.fetchone() or {}).get("alifbo_turi") or "latin_xalqaro"
    classes.sort(key=lambda row: _v237_class_sort_key(row, alphabet_type))
    alphabet = _V237_CLASS_ALPHABETS.get(
        alphabet_type, _V237_CLASS_ALPHABETS["latin_xalqaro"]
    )
    alphabet_indexes = {
        _v237_parallel_label_key(label): index for index, label in enumerate(alphabet)
    }
    for class_row in classes:
        try:
            label_key = _v237_parallel_label_key(class_row.get("harf"))
        except ValueError:
            label_key = str(class_row.get("harf") or "").casefold()
        class_row["sinf_nomi"] = f"{class_row.get('sinf','')}-{class_row.get('harf','')}"
        class_row["alifbo_turi"] = alphabet_type
        class_row["alifbo_tartibi"] = alphabet_indexes.get(label_key)
    cur.execute("""SELECT q.*,s.sinf,s.harf,COALESCE(s.smena,1) AS smena,
                          s.rahbar_user_id,COALESCE(u.full_name,'') AS rahbar_ismi
                   FROM aqlli_sinf_soati_qoidalari_v2 q
                   JOIN maktab_sinflari s ON s.id=q.sinf_id
                   LEFT JOIN users u ON u.user_id=s.rahbar_user_id
                   WHERE q.maktab_id=%s AND q.faol=TRUE
                   ORDER BY s.sinf::int,s.harf""", (maktab_id,))
    class_hours = cur.fetchall()
    class_hour_counts = _v1852_Counter()
    for row in class_hours:
        if row.get("rahbar_user_id") is not None:
            class_hour_counts[int(row["rahbar_user_id"])] += int(row.get("haftalik_soat") or 1)
    plan = _v193_plan_payload(cur, maktab_id, classes)
    _v2249_ensure_teacher_numbers(cur, maktab_id)
    teachers = _v1859_effective_teachers(cur, maktab_id)
    for teacher in teachers:
        extra = int(class_hour_counts.get(int(teacher["user_id"]), 0))
        teacher["sinf_soati_soni"] = extra
        base = teacher.get("haftalik_dars_soati")
        teacher["haftalik_reja_jami"] = (
            round(float(base) + extra, 1)
            if base is not None else (extra or None)
        )
    cur.execute("SELECT fan_nomi FROM maktab_fanlari WHERE maktab_id=%s ORDER BY fan_nomi", (maktab_id,))
    subjects = [r["fan_nomi"] for r in cur.fetchall()]
    cur.execute("SELECT * FROM aqlli_xonalar_v2 WHERE maktab_id=%s AND faol=TRUE AND darsga_yaroqli=TRUE ORDER BY nomi", (maktab_id,))
    rooms = cur.fetchall()
    cur.execute("SELECT * FROM aqlli_oqituvchi_qoidalari_v2 WHERE maktab_id=%s ORDER BY user_id", (maktab_id,))
    teacher_rules = cur.fetchall()
    cur.execute("SELECT * FROM aqlli_oqituvchi_vaqti_v2 WHERE maktab_id=%s ORDER BY user_id,hafta_kuni,smena,dars_raqami", (maktab_id,))
    teacher_times = cur.fetchall()
    cur.execute("SELECT * FROM aqlli_sinf_fan_yuklamalari_v2 WHERE maktab_id=%s ORDER BY sinf_id,fan_nomi", (maktab_id,))
    loads = cur.fetchall()
    cur.execute("SELECT * FROM aqlli_guruh_sozlamalari_v2 WHERE maktab_id=%s ORDER BY sinf_id,fan_nomi,guruh_kaliti", (maktab_id,))
    group_settings = cur.fetchall()
    cur.execute("""SELECT b.sinf_id,b.user_id,b.fan_nomi,b.guruh_kaliti,
                          b.haftalik_soat,b.kunlik_max,b.manba,u.full_name,
                          s.sinf,s.harf,(s.sinf || '-' || s.harf) AS sinf_nomi
                   FROM maktab_dars_birikmalari b
                   JOIN users u ON u.user_id=b.user_id
                   JOIN maktab_sinflari s ON s.id=b.sinf_id
                   WHERE b.maktab_id=%s
                   ORDER BY s.sinf::int,s.harf,b.fan_nomi,b.guruh_kaliti,u.full_name""", (maktab_id,))
    assignments = cur.fetchall()
    cur.execute("""SELECT id,holat,yaratilgan_at,tasdiqlangan_at,sifat,joylashtirildi,joylashtirilmadi,diagnostika,
                          COALESCE((diagnostika->>'generator_rejimi')::int,1) AS generator_rejimi,
                          COALESCE(diagnostika->>'yumshatish_rejimi','strict') AS yumshatish_rejimi
                   FROM aqlli_jadval_urinishlari_v2 WHERE maktab_id=%s ORDER BY id DESC LIMIT 4""", (maktab_id,))
    runs = cur.fetchall()
    class_day_blocks = _v1856_class_day_rule_rows(cur, maktab_id)
    return {
        "oquv_yili": year, "choraklar": quarters, "maxsus_kunlar": special_days,
        "alifbo_turi": alphabet_type,
        "smenalar": shifts, "sinflar": classes, "oqituvchilar": teachers,
        "fanlar": subjects, "xonalar": rooms, "oqituvchi_qoidalari": teacher_rules,
        "oqituvchi_vaqtlari": teacher_times, "fan_soatlari": loads,
        "guruh_sozlamalari": group_settings, "birikmalar": assignments, "urinishlar": runs,
        "sinf_kun_bloklari": class_day_blocks, "sinf_soatlari": class_hours,
        "avtomatik_qoidalar": {
            "sinf_kun_bloklari": class_day_blocks,
            "izoh": "Rahbariyat xohlagan parallel yoki aniq sinf uchun dars bo'lmaydigan hafta kunini belgilaydi",
        },
    }


class V1856ClassDayBlocks(BaseModel):
    maktab_id: int
    qamrov: str = "parallel"
    sinf_darajalari: list[int] = []
    sinf_idlar: list[int] = []
    hafta_kunlari: list[int] = []
    izoh: Optional[str] = None


@app.put("/api/maktab/aqlli_jadval/v2/sinf_kun_bloklari")
def v1856_class_day_blocks_save(sorov: V1856ClassDayBlocks, token: str):
    user_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor()
    try:
        _v1852_tables(cur)
        if not _v1852_manager(cur, user_id, sorov.maktab_id):
            raise HTTPException(status_code=403, detail="Sinf kunlarini faqat maktab rahbariyati boshqaradi")
        _v1856_seed_default_class_day_rules(cur, sorov.maktab_id)
        qamrov = str(sorov.qamrov or "parallel").strip().lower()
        days = sorted(set(int(x) for x in sorov.hafta_kunlari))
        if not days or any(day not in range(1, 7) for day in days):
            raise HTTPException(status_code=400, detail="Dushanba–Shanbadan bitta kunni tanlang")
        if len(days) != 1:
            raise HTTPException(
                status_code=400,
                detail="Har bir qoidada faqat bitta hafta kuni bo'lishi kerak. Yana boshqa kun kerak bo'lsa, uni alohida qoida qilib saqlang.",
            )
        created = 0
        if qamrov == "parallel":
            grades = sorted(set(int(x) for x in sorov.sinf_darajalari))
            if not grades or any(grade not in range(1, 12) for grade in grades):
                raise HTTPException(status_code=400, detail="Kamida bitta 1–11-sinf parallelini tanlang")
            cur.execute("""SELECT DISTINCT NULLIF(REGEXP_REPLACE(COALESCE(sinf,''),'[^0-9]','','g'),'')::int AS grade
                           FROM maktab_sinflari WHERE maktab_id=%s""", (sorov.maktab_id,))
            available = {int(r["grade"]) for r in cur.fetchall() if r.get("grade") is not None}
            missing = [grade for grade in grades if grade not in available]
            if missing:
                raise HTTPException(status_code=400, detail=f"Maktabda bu parallellar yo'q: {', '.join(map(str, missing))}-sinf")
            for grade in grades:
                for day in days:
                    cur.execute("""SELECT id FROM aqlli_sinf_kun_bloklari_v2
                                   WHERE maktab_id=%s AND sinf_id IS NULL AND sinf_daraja=%s AND hafta_kuni=%s""",
                                (sorov.maktab_id, grade, day))
                    row = cur.fetchone()
                    if row:
                        cur.execute("""UPDATE aqlli_sinf_kun_bloklari_v2 SET faol=TRUE,izoh=%s,
                                       yaratgan_user_id=%s,yangilangan_at=NOW() WHERE id=%s""",
                                    (sorov.izoh, user_id, row["id"]))
                    else:
                        cur.execute("""INSERT INTO aqlli_sinf_kun_bloklari_v2(
                            maktab_id,sinf_daraja,hafta_kuni,izoh,faol,yaratgan_user_id)
                            VALUES(%s,%s,%s,%s,TRUE,%s)""",
                            (sorov.maktab_id, grade, day, sorov.izoh, user_id))
                        created += 1
        elif qamrov == "sinf":
            class_ids = sorted(set(int(x) for x in sorov.sinf_idlar))
            if not class_ids:
                raise HTTPException(status_code=400, detail="Kamida bitta aniq sinfni tanlang")
            cur.execute("SELECT id FROM maktab_sinflari WHERE maktab_id=%s AND id=ANY(%s)",
                        (sorov.maktab_id, class_ids))
            found = {int(r["id"]) for r in cur.fetchall()}
            missing = [cid for cid in class_ids if cid not in found]
            if missing:
                raise HTTPException(status_code=400, detail="Tanlangan sinflardan biri bu maktabga tegishli emas")
            for class_id in class_ids:
                for day in days:
                    cur.execute("""SELECT id FROM aqlli_sinf_kun_bloklari_v2
                                   WHERE maktab_id=%s AND sinf_id=%s AND hafta_kuni=%s""",
                                (sorov.maktab_id, class_id, day))
                    row = cur.fetchone()
                    if row:
                        cur.execute("""UPDATE aqlli_sinf_kun_bloklari_v2 SET faol=TRUE,izoh=%s,
                                       yaratgan_user_id=%s,yangilangan_at=NOW() WHERE id=%s""",
                                    (sorov.izoh, user_id, row["id"]))
                    else:
                        cur.execute("""INSERT INTO aqlli_sinf_kun_bloklari_v2(
                            maktab_id,sinf_id,hafta_kuni,izoh,faol,yaratgan_user_id)
                            VALUES(%s,%s,%s,%s,TRUE,%s)""",
                            (sorov.maktab_id, class_id, day, sorov.izoh, user_id))
                        created += 1
        else:
            raise HTTPException(status_code=400, detail="Qamrov parallel yoki sinf bo'lishi kerak")
        rules = _v1856_class_day_rule_rows(cur, sorov.maktab_id)
        violations = _v1856_schedule_block_violations(cur, sorov.maktab_id)
        conn.commit()
        return {
            "holat": "saqlandi", "yangi_qoida_soni": created,
            "qoidalar": rules, "faol_jadvaldagi_zid_darslar": violations,
            "zid_dars_soni": sum(int(x.get("dars_soni") or 0) for x in violations),
        }
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


@app.delete("/api/maktab/aqlli_jadval/v2/sinf_kun_bloki")
def v1856_class_day_block_delete(token: str, maktab_id: int, blok_id: int):
    user_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor()
    try:
        _v1852_tables(cur)
        if not _v1852_manager(cur, user_id, maktab_id):
            raise HTTPException(status_code=403, detail="Sinf kunlarini faqat maktab rahbariyati boshqaradi")
        cur.execute("DELETE FROM aqlli_sinf_kun_bloklari_v2 WHERE id=%s AND maktab_id=%s RETURNING id",
                    (blok_id, maktab_id))
        if not cur.fetchone():
            raise HTTPException(status_code=404, detail="Qoida topilmadi")
        rules = _v1856_class_day_rule_rows(cur, maktab_id)
        conn.commit()
        return {"holat": "o'chirildi", "qoidalar": rules}
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


@app.get("/api/maktab/aqlli_jadval/v2/sozlamalar")
def v1852_setup(token: str, maktab_id: int):
    user_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor()
    try:
        _v1852_tables(cur)
        if not _v1852_staff(cur, user_id, maktab_id):
            raise HTTPException(status_code=403, detail="Bu maktab jadvalini ko'rishga ruxsat yo'q")
        manager = _v1852_manager(cur, user_id, maktab_id)
        if manager:
            # Yangi yoki eski maktabdan qat'i nazar barcha faol sinfga bir
            # martadan SINF SOATI avtomatik mavjud bo'ladi. Qo'lda tanlangan
            # kun/dars ON CONFLICT sabab o'zgarmaydi.
            _v199_ensure_class_hour_rules(cur, maktab_id, actor_id=user_id)
        payload = _v1852_setup_payload(cur, maktab_id)
        payload["joriy_user_id"] = user_id
        payload["boshqaruvchi"] = manager
        if not manager:
            own_assignments = [row for row in payload["birikmalar"] if int(row["user_id"]) == int(user_id)]
            allowed_pairs = {(int(row["sinf_id"]), str(row["fan_nomi"]).strip().casefold()) for row in own_assignments}
            allowed_class_ids = {pair[0] for pair in allowed_pairs}
            payload["oqituvchilar"] = [row for row in payload["oqituvchilar"] if int(row["user_id"]) == int(user_id)]
            payload["oqituvchi_qoidalari"] = [row for row in payload["oqituvchi_qoidalari"] if int(row["user_id"]) == int(user_id)]
            payload["oqituvchi_vaqtlari"] = [row for row in payload["oqituvchi_vaqtlari"] if int(row["user_id"]) == int(user_id)]
            payload["birikmalar"] = own_assignments
            payload["sinflar"] = [row for row in payload["sinflar"] if int(row["id"]) in allowed_class_ids]
            payload["fan_soatlari"] = [row for row in payload["fan_soatlari"] if (int(row["sinf_id"]), str(row["fan_nomi"]).strip().casefold()) in allowed_pairs]
            payload["guruh_sozlamalari"] = [row for row in payload["guruh_sozlamalari"] if (int(row["sinf_id"]), str(row["fan_nomi"]).strip().casefold()) in allowed_pairs]
            payload["urinishlar"] = [row for row in payload["urinishlar"] if row["holat"] == "tasdiqlangan"][:1]
        conn.commit()
        return payload
    finally:
        cur.close(); conn.close()


def _v1852_teacher_rules_map(rows):
    result = {}
    for row in rows:
        result[int(row["user_id"])] = {
            "kunlik_max": int(row.get("kunlik_max") or 6),
            "ketma_ket_max": int(row.get("ketma_ket_max") or 4),
            "okno_max": int(row.get("okno_max") or 1),
            "afzal_smena": int(row.get("afzal_smena") or 0),
            "eng_erta_dars": int(row.get("eng_erta_dars") or 1),
            "eng_kech_dars": int(row.get("eng_kech_dars") or 12),
        }
    return result


def _v1852_availability_maps(rows):
    hard = set(); soft = set(); method_hard = set(); method_soft = set()
    for row in rows:
        key = (int(row["user_id"]), int(row["hafta_kuni"]), int(row.get("smena") or 0), int(row.get("dars_raqami") or 0))
        if row["turi"] == "metod_kuni":
            # REV52: METOD KUNI tavsiya emas, to'liq darsdan xoli kun.
            # Eski bazada qattiq=False saqlangan yoki rasmiy/avtomatik yozuv
            # bo'lsa ham generator bu kunga birorta dars joylamaydi.
            method_hard.add((key[0], key[1]))
        elif row["turi"] == "band":
            (hard if row.get("qattiq") else soft).add(key)
        else:
            soft.add(key)
    return hard, soft, method_hard, method_soft


def _v1852_blocked(hard, teacher, day, shift, period):
    return any(key in hard for key in (
        (teacher, day, 0, 0), (teacher, day, shift, 0),
        (teacher, day, 0, period), (teacher, day, shift, period),
    ))






def _v1852_gap_count(periods: set[int]) -> int:
    return 0 if len(periods) < 2 else max(periods) - min(periods) + 1 - len(periods)


def _v205_annotate_class_home_room(class_row, catalog_rows):
    """Sinfning matnli uy xonasini katalogdagi ayni jismoniy xona bilan bog'laydi.

    Eski maktablarda ``maktab_sinflari.xona`` matn bo'lib qolgan. Katalogda
    shu xona topilsa generator ham, keyin saqlanadigan slot ham bitta
    ``room:<id>`` resursidan foydalanadi. Bir nechta xona yozilgan yoki xona
    darsga yaroqsiz bo'lsa, uni bitta soxta xona deb qabul qilmaymiz.
    """
    row = dict(class_row)
    raw_room = re.sub(r"\s+", " ", str(row.get("xona") or "")).strip()
    building = re.sub(r"\s+", " ", str(row.get("bino") or "")).strip()
    display_text = " ".join(part for part in (building, raw_room) if part)
    row["_home_room_id"] = None
    row["_home_room_text"] = None
    row["_home_room_key"] = None
    row["_home_room_invalid"] = None

    # V20.9 yaratish route'i sinf uy xonasining kanonik ID sini saqlaydi.
    # Matn keyin tahrirlangan yoki ikki binoda bir xil xona raqami bo'lsa ham
    # ID birinchi manba bo'lib qoladi. Eski sinflar uchun pastdagi bino+xona
    # matnidan topish fallbacki saqlanadi.
    stored_room_id = row.get("xona_id")
    if stored_room_id is not None:
        try:
            stored_room_id = int(stored_room_id)
        except (TypeError, ValueError):
            stored_room_id = None
    if stored_room_id is not None:
        matched_by_id = next(
            (
                room for room in catalog_rows
                if int(room.get("id") or 0) == stored_room_id
            ),
            None,
        )
        if matched_by_id is None:
            # Eski katalog qayta yaratilganda sinfda avvalgi xona ID si qolishi
            # mumkin. Bu holat barcha sinflarni "Xona yo'q" ogohlantirishiga
            # aylantirmasin: pastda mavjud xona matni bo'yicha qayta bog'laymiz.
            # Matn ham bo'lmasa xona shunchaki ko'rsatilmaydi.
            stored_room_id = None
        if matched_by_id is not None and (
            not bool(matched_by_id.get("faol")) or not bool(
            matched_by_id.get("darsga_yaroqli")
            )
        ):
            row["_home_room_invalid"] = "sinf xonasi dars o'tishga yaroqsiz"
            return row
        if matched_by_id is not None:
            row["_home_room_id"] = stored_room_id
            row["_home_room_text"] = str(
                matched_by_id.get("nomi") or display_text or raw_room
            ).strip()
            row["_home_room_key"] = f"room:{stored_room_id}"
            return row
    if not raw_room:
        return row

    room_parts = _v205_split_room_names(raw_room)
    if len(room_parts) != 1:
        row["_home_room_invalid"] = "sinfga bitta aniq xona tanlanmagan"
        return row

    candidate_keys = []
    for value in (display_text, raw_room):
        key = _v205_room_normalized_name(value)
        if key and key not in candidate_keys:
            candidate_keys.append(key)
    by_name = _v1852_defaultdict(list)
    for room in catalog_rows:
        key = _v205_room_normalized_name(room.get("nomi"))
        if key:
            by_name[key].append(room)

    matched = None
    for key in candidate_keys:
        rows = by_name.get(key, [])
        if rows:
            matched = next(
                (
                    room for room in rows
                    if bool(room.get("faol")) and bool(room.get("darsga_yaroqli"))
                ),
                rows[0],
            )
            break
    if matched is not None:
        if not bool(matched.get("faol")) or not bool(matched.get("darsga_yaroqli")):
            row["_home_room_invalid"] = "sinf xonasi dars o'tishga yaroqsiz"
            return row
        room_id = int(matched["id"])
        row["_home_room_id"] = room_id
        row["_home_room_text"] = str(matched.get("nomi") or display_text).strip()
        row["_home_room_key"] = f"room:{room_id}"
        return row

    # Katalogga hali kiritilmagan eski, bitta aniq xona matni ham resursdir.
    # Uni id bilan emas, barqaror kanonik matn kaliti bilan kuzatamiz.
    normalized_display = _v205_room_normalized_name(display_text)
    if normalized_display:
        row["_home_room_text"] = display_text
        row["_home_room_key"] = f"text:{normalized_display}"
    return row


def _v220_room_source_errors(classes, catalog_rows, group_settings):
    """Return deterministic room-source errors before the exact model runs.

    One physical home room may be shared by one class from each shift because
    their real lesson times do not overlap.  Only duplicates inside the same
    shift are a source error.
    The exact solver then quite correctly locks both classes to the same room
    and may only be able to report a generic global ``INFEASIBLE``.  Surface
    the violated source invariant here instead.

    Parallel groups are checked with the same room fallback as the exact
    adapter: the alphabetically first group stays in the class home room when
    it has no explicit room; later room-less groups remain ``Xona yo'q``.
    Invalid/inactive stored room IDs are deliberately ignored because
    generation already converts those IDs to ``None`` rather than blocking a
    timetable.
    """
    errors = []
    occupied_home_rooms = {}
    for class_id, class_row in sorted(classes.items()):
        room_key = class_row.get("_home_room_key")
        if not room_key:
            continue
        label = f"{class_row.get('sinf', '')}-{class_row.get('harf', '')}"
        shift = 2 if int(class_row.get("smena") or 1) == 2 else 1
        occupied_key = (shift, str(room_key))
        previous = occupied_home_rooms.get(occupied_key)
        if previous:
            errors.append(
                "Uy xona bir smenada takror biriktirilgan: "
                f"{class_row.get('_home_room_text') or room_key}. Bitta "
                f"xona {shift}-smenada faqat bitta sinfga beriladi "
                f"({previous} va {label})."
            )
        else:
            occupied_home_rooms[occupied_key] = label

    valid_room_ids = {
        int(row["id"])
        for row in catalog_rows
        if row.get("id") is not None
        and bool(row.get("faol"))
        and bool(row.get("darsga_yaroqli"))
    }
    room_labels = {
        int(row["id"]): str(row.get("nomi") or row["id"])
        for row in catalog_rows
        if row.get("id") is not None
    }
    grouped = _v1852_defaultdict(list)
    for setting in group_settings:
        grouped[(
            int(setting["sinf_id"]),
            _v1875_subject_key(setting.get("fan_nomi")),
        )].append(setting)

    for (class_id, _subject_key), rows in sorted(grouped.items()):
        if len(rows) < 2:
            continue
        class_row = classes.get(class_id, {})
        class_label = f"{class_row.get('sinf', '')}-{class_row.get('harf', '')}"
        subject = str(rows[0].get("fan_nomi") or "").strip()
        used = {}
        for index, row in enumerate(sorted(
            rows, key=lambda item: str(item.get("guruh_kaliti") or "")
        )):
            raw_room_id = row.get("xona_id")
            room_key = None
            room_label = None
            if raw_room_id is not None:
                room_id = int(raw_room_id)
                if room_id in valid_room_ids:
                    room_key = f"room:{room_id}"
                    room_label = room_labels.get(room_id, str(room_id))
            elif index == 0:
                room_key = class_row.get("_home_room_key")
                room_label = class_row.get("_home_room_text")
            if not room_key:
                continue
            group_label = str(row.get("guruh_kaliti") or index + 1)
            previous_group = used.get(str(room_key))
            if previous_group:
                errors.append(
                    f"{class_label} / {subject}: {previous_group} va "
                    f"{group_label} parallel guruhlarga bitta xona "
                    f"({room_label or room_key}) biriktirilgan."
                )
            else:
                used[str(room_key)] = group_label
    return errors


def _v1852_prepare_generation(cur, maktab_id: int):
    year = _v1890_generation_year(cur, maktab_id)
    _v1852_default_shifts(cur, maktab_id)
    cur.execute("SELECT * FROM aqlli_smena_sozlamalari_v2 WHERE maktab_id=%s", (maktab_id,))
    shift_rows = cur.fetchall()
    shifts = {int(row["smena"]): {**row, "slotlar": _v1852_shift_slots(row)} for row in shift_rows}
    cur.execute("""SELECT id,sinf,harf,COALESCE(smena,1) AS smena,bino,xona,
                          bino_id,xona_id,rahbar_user_id
                   FROM maktab_sinflari WHERE maktab_id=%s ORDER BY sinf::int,harf""", (maktab_id,))
    classes = {int(row["id"]): dict(row) for row in cur.fetchall()}
    cur.execute("SELECT * FROM aqlli_sinf_fan_yuklamalari_v2 WHERE maktab_id=%s AND haftalik_soat>0 ORDER BY sinf_id,fan_nomi", (maktab_id,))
    loads = [dict(row) for row in cur.fetchall()]
    if not loads:
        raise HTTPException(status_code=400, detail="Avval sinflar uchun haftalik fan soatlarini kiriting")
    cur.execute("""SELECT b.sinf_id,b.user_id,b.fan_nomi,b.guruh_kaliti,b.haftalik_soat,
                          u.full_name,u.haftalik_dars_soati
                   FROM maktab_dars_birikmalari b JOIN users u ON u.user_id=b.user_id
                   WHERE b.maktab_id=%s
                   ORDER BY b.sinf_id,b.fan_nomi,b.guruh_kaliti,u.full_name""", (maktab_id,))
    assignments = cur.fetchall()
    cur.execute("SELECT * FROM aqlli_guruh_sozlamalari_v2 WHERE maktab_id=%s", (maktab_id,))
    group_settings = [dict(row) for row in cur.fetchall()]
    cur.execute("SELECT * FROM aqlli_oqituvchi_qoidalari_v2 WHERE maktab_id=%s", (maktab_id,))
    rules_rows = cur.fetchall()
    cur.execute("SELECT * FROM aqlli_oqituvchi_vaqti_v2 WHERE maktab_id=%s", (maktab_id,))
    availability_rows = cur.fetchall()
    cur.execute("""SELECT u.user_id,u.full_name,u.haftalik_dars_soati
                   FROM users u
                   WHERE u.maktab_id=%s AND u.lavozim IS NOT NULL
                     AND (
                       EXISTS(SELECT 1 FROM maktab_dars_birikmalari b
                              WHERE b.maktab_id=%s AND b.user_id=u.user_id)
                       OR EXISTS(SELECT 1 FROM maktab_xodim_sinflari x
                                 WHERE x.maktab_id=%s AND x.user_id=u.user_id)
                       OR EXISTS(SELECT 1 FROM maktab_sinflari s
                                 WHERE s.maktab_id=%s AND s.rahbar_user_id=u.user_id)
                     )""", (maktab_id, maktab_id, maktab_id, maktab_id))
    teachers = {int(row["user_id"]): row for row in cur.fetchall()}
    # Yaroqsiz/inaktiv xona nomini oddiy matnli uy xonasi deb qayta tiriltirib
    # yubormaslik uchun identifikatsiyada barcha katalog qatorlari ko'riladi.
    cur.execute("SELECT * FROM aqlli_xonalar_v2 WHERE maktab_id=%s", (maktab_id,))
    all_room_rows = [dict(row) for row in cur.fetchall()]
    rooms = {
        int(row["id"]): row for row in all_room_rows
        if bool(row.get("faol")) and bool(row.get("darsga_yaroqli"))
    }
    classes = {
        class_id: _v205_annotate_class_home_room(class_row, all_room_rows)
        for class_id, class_row in classes.items()
    }
    valid_room_ids = set(rooms)
    # O'chirilgan yoki darsga yaroqsiz eski xona birikmasi jadvalni
    # to'xtatmaydi: xona yo'q bo'lib, keyin rahbariyat qo'lda tuzatadi.
    for load in loads:
        if load.get("xona_id") is not None and int(load["xona_id"]) not in valid_room_ids:
            load["xona_id"] = None
    for setting in group_settings:
        if setting.get("xona_id") is not None and int(setting["xona_id"]) not in valid_room_ids:
            setting["xona_id"] = None
    return year, shifts, classes, loads, assignments, group_settings, rules_rows, availability_rows, teachers, rooms


def _v1852_build_jobs_base(classes, loads, assignments, group_settings, teachers):
    assignment_map = _v1852_defaultdict(list)
    for row in assignments:
        assignment_map[(int(row["sinf_id"]), str(row["fan_nomi"]).strip().casefold())].append(row)
    group_map = {}
    for row in group_settings:
        group_map[(int(row["sinf_id"]), str(row["fan_nomi"]).strip().casefold(), str(row["guruh_kaliti"]))] = row
    jobs = []
    warnings = []
    for load in loads:
        class_id = int(load["sinf_id"])
        if class_id not in classes:
            continue
        fan = str(load["fan_nomi"]).strip()
        rows = assignment_map.get((class_id, fan.casefold()), [])
        non_whole = [r for r in rows if str(r.get("guruh_kaliti") or "whole") != "whole"]
        whole = [r for r in rows if str(r.get("guruh_kaliti") or "whole") == "whole"]
        fixed_groups = []
        teacher_options = []
        if non_whole:
            seen_groups = set()
            for row in non_whole:
                group_key = str(row.get("guruh_kaliti") or "whole")
                if group_key in seen_groups:
                    continue
                seen_groups.add(group_key)
                setting = group_map.get((class_id, fan.casefold(), group_key), {})
                teacher_id = setting.get("oqituvchi_user_id") or row.get("user_id")
                fixed_groups.append({
                    "guruh_kaliti": group_key,
                    "teacher": int(teacher_id) if teacher_id is not None else None,
                    "xona_id": int(setting["xona_id"]) if setting.get("xona_id") else None,
                })
            if any(g["teacher"] is None for g in fixed_groups):
                warnings.append(f"{classes[class_id]['sinf']}-{classes[class_id]['harf']} {fan}: guruh o'qituvchisi to'liq biriktirilmagan")
        else:
            primary = load.get("asosiy_oqituvchi_user_id")
            if primary is not None:
                teacher_options = [int(primary)]
            else:
                teacher_options = list(dict.fromkeys(int(r["user_id"]) for r in whole if r.get("user_id") is not None))
            if not teacher_options:
                warnings.append(f"{classes[class_id]['sinf']}-{classes[class_id]['harf']} {fan}: o'qituvchi biriktirilmagan")
        weekly_hours = max(0.0, float(load.get("haftalik_soat") or 0))
        whole_occurrences = int(math.floor(weekly_hours))
        has_half_rotation = abs(weekly_hours - whole_occurrences - 0.5) < 1e-9
        occurrence_count = whole_occurrences + (1 if has_half_rotation else 0)
        for occurrence in range(1, occurrence_count + 1):
            rotation_weight = 0.5 if has_half_rotation and occurrence == occurrence_count else 1.0
            jobs.append({
                "job_id": f"{load['id']}:{occurrence}", "load_id": int(load["id"]),
                "sinf_id": class_id, "fan": fan, "occurrence": occurrence,
                "smena": int(classes[class_id].get("smena") or 1),
                # V22.26 yagona kontrakt: fan bir sinfga bir kunda 2
                # soatgacha joylashishi mumkin, lekin hech qachon 3 emas.
                # Bu ruxsat sig'im uchun; optimizator imkon bo'lsa tarqatadi.
                "daily_max": 2,
                "consecutive_allowed": bool(load.get("ketma_ket_mumkin")),
                "preferred_last": int(load.get("afzal_oxirgi_dars") or 5),
                "weight": int(load.get("ogirlik") or 2),
                "room_id": int(load["xona_id"]) if load.get("xona_id") else None,
                "groups": fixed_groups, "teacher_options": teacher_options,
                "rotation_weight": rotation_weight,
                "difficulty": (100 if fixed_groups else 0) + (50 if load.get("xona_id") else 0) + weekly_hours,
            })
    return jobs, warnings


def _v1852_candidate_reasons_base(job, day, period, selected_teachers, room_keys, state, context):
    reasons = []
    # Xona — jadval yaratishni to'xtatadigan shart emas. Parallel guruhning
    # alohida xonasi hali yozilmagan bo'lsa slot baribir yaratiladi; frontend
    # uni "Xona yo'q" deb ko'rsatadi va keyin qo'lda to'ldirish mumkin.
    non_null_teachers = [teacher for teacher in selected_teachers if teacher is not None]
    if len(non_null_teachers) != len(set(non_null_teachers)):
        reasons.append("bir o'qituvchi ikki parallel guruhga biriktirilgan")
    non_null_rooms = [key for key in room_keys if key]
    if len(non_null_rooms) != len(set(non_null_rooms)):
        reasons.append("parallel guruhlar bir xil xonaga biriktirilgan")
    class_slot = (job["sinf_id"], day, job["smena"], period)
    if class_slot in state["class_busy"]:
        reasons.append("sinf band")
    if state["subject_daily"].get((job["sinf_id"], job["fan"].casefold(), day), 0) >= job["daily_max"]:
        reasons.append("fan kunlik maksimumga yetgan")
    for teacher in selected_teachers:
        if teacher is None:
            reasons.append("o'qituvchi biriktirilmagan")
            continue
        rules = context["rules"].get(teacher, context["default_rules"])
        if period < rules["eng_erta_dars"] or period > rules["eng_kech_dars"]:
            reasons.append("o'qituvchi ruxsat etgan dars oralig'idan tashqari")
        if (teacher, day) in context["method_hard"]:
            reasons.append("o'qituvchining metod kuni")
        if _v1852_blocked(context["hard"], teacher, day, job["smena"], period):
            reasons.append("o'qituvchi bu vaqtda band")
        if (teacher, day, job["smena"], period) in state["teacher_busy"]:
            reasons.append("o'qituvchi boshqa darsda")
        cap = context["teacher_caps"].get(teacher)
        if cap is not None and state["teacher_week"].get(teacher, 0) >= cap:
            reasons.append("o'qituvchining haftalik yuklamasi to'lgan")
        if state["teacher_daily"].get((teacher, day), 0) >= rules["kunlik_max"]:
            reasons.append("o'qituvchining kunlik maksimumi to'lgan")
        # Ketma-ket dars soni qulaylik talabi, majburiy blok emas. Masalan,
        # metod kuni bor 30 soatli o'qituvchi qolgan 5 kunda 6 tadan dars
        # o'tishi mumkin; 4 ta ketma-ket limitini qattiq qo'llash uning 6 ta
        # darsini asossiz ravishda jadvaldan chiqarib yuborar edi. Me'yordan
        # oshish pastdagi ballashda jazolanadi, lekin dars tashlab ketilmaydi.
    for key in room_keys:
        if key and (key, day, job["smena"], period) in state["room_busy"]:
            reasons.append("xona band")
    return list(dict.fromkeys(reasons))


def _v199_base_choose_teacher(job, day, period, state, context):
    if job["groups"]:
        return [g.get("teacher") for g in job["groups"]]
    options = job["teacher_options"]
    if not options:
        return [None]
    valid = []
    for teacher in options:
        reasons = _v1852_candidate_reasons(job, day, period, [teacher], [], state, context)
        if not reasons:
            valid.append((state["teacher_week"].get(teacher, 0), state["teacher_daily"].get((teacher, day), 0), teacher))
    return [min(valid)[2]] if valid else [options[0]]


def _v1852_room_keys(job, selected_teachers, classes):
    class_row = classes[job["sinf_id"]]
    home_room_key = class_row.get("_home_room_key")

    def member_keys(source_job):
        if source_job.get("groups"):
            # Birinchi guruh sinf xonasida qoladi. Keyingi guruhga aniq xona
            # yozilmagan bo'lsa None qaytadi va natijada "Xona yo'q" bo'ladi.
            return [
                f"room:{group['xona_id']}"
                if group.get("xona_id")
                else (home_room_key if group_index == 0 else None)
                for group_index, group in enumerate(source_job["groups"])
            ]
        if source_job.get("room_id"):
            return [f"room:{source_job['room_id']}"]
        return [home_room_key]

    rotation_members = job.get("rotation_members") or []
    if rotation_members:
        # A/B a'zolarining xonasi kombinatsiyalangan jobda o'chib ketadi.
        # Fazali room_busy hali yo'q, shuning uchun barcha haqiqiy a'zo
        # xonalarining konservativ birlashmasi yashirin kolliziyani oldini oladi.
        combined = []
        seen = set()
        for member in rotation_members:
            keys = member_keys(member)
            non_null = [key for key in keys if key]
            if len(non_null) != len(set(non_null)):
                # Bitta fazaning parallel guruhlari ayni xonani olgan: bazaviy
                # tekshiruv takror kalit orqali aniq xato qaytarsin.
                return keys
            for key in keys:
                if key and key not in seen:
                    seen.add(key)
                    combined.append(key)
        return combined or [None]

    if job.get("groups"):
        # Guruh xonasi yozilmagan bo'lsa barcha guruhga bir xil sinf xonasi
        # kalitini berish mumkin emas: bu ularni soxta "bir xil xona" ziddiyatiga
        # tushirib, butun guruhli fanni jadvaldan chiqarib yuborar edi. Birinchi
        # guruh sinf xonasida qoladi, keyingi xonasiz guruhlar esa jadvalga
        # room=None bilan tushadi va frontendda "Xona yo'q" deb ko'rinadi.
        return member_keys(job)
    return member_keys(job)




def _v1852_place_job_base(job, day, period, teachers, room_keys, state, context):
    state["class_busy"].add((job["sinf_id"], day, job["smena"], period))
    state["subject_daily"][(job["sinf_id"], job["fan"].casefold(), day)] += 1
    state["class_subject_periods"][(job["sinf_id"], job["fan"].casefold(), day)].add(period)
    for teacher in teachers:
        if teacher is None:
            continue
        state["teacher_busy"].add((teacher, day, job["smena"], period))
        state["teacher_week"][teacher] += 1
        state["teacher_daily"][(teacher, day)] += 1
        state["teacher_periods"][(teacher, day, job["smena"])].add(period)
    for key in room_keys:
        if key:
            state["room_busy"].add((key, day, job["smena"], period))
    state["placements"].append({"job": job, "day": day, "period": period, "teachers": teachers, "room_keys": room_keys})


def _v1852_new_schedule_state():
    return {
        "class_busy": set(), "teacher_busy": set(), "room_busy": set(),
        "subject_daily": _v1852_defaultdict(int),
        "class_subject_periods": _v1852_defaultdict(set),
        "teacher_week": _v1852_defaultdict(int),
        "teacher_daily": _v1852_defaultdict(int),
        "teacher_periods": _v1852_defaultdict(set), "placements": [],
    }


def _v1852_job_teacher_ids(job):
    result = set()
    members = job.get("rotation_members") or []
    source_jobs = members or [job]
    for source in source_jobs:
        groups = source.get("groups") or []
        if groups:
            for group in groups:
                if group.get("teacher") is not None:
                    result.add(int(group["teacher"]))
        else:
            for teacher in source.get("teacher_options") or []:
                if teacher is not None:
                    result.add(int(teacher))
    return result


_V213_CORE_PERIOD6_LIMIT = 2
_V213_CORE_PERIOD6_REASON = (
    "asosiy og'ir fan 6-darsga haftada ko'pi bilan 2 kun qo'yiladi"
)






_V205_SCHEDULE_RULES = [
    "Tasdiqlangan fan–sinf–guruh yuklamasidagi har bir soat aynan bir marta joylashtiriladi; ortiqcha yoki tushib qolgan darsga ruxsat yo'q.",
    "Bitta sinf bir vaqtda faqat bitta dars paketida bo'ladi; guruhli fanlar esa shu katak ichida parallel ko'rsatiladi.",
    "Bitta o'qituvchi bir xil real vaqtda, smenasi boshqa bo'lsa ham, ikki sinfga qo'yilmaydi.",
    "Guruhli fanlarning barcha guruhlari uchun sinf va tegishli o'qituvchilar bir vaqtda bo'sh bo'lishi shart.",
    "O'qituvchining qizil kuni/soati va qattiq BAND vaqti barcha darslar, jumladan KELAJAK SOATI uchun ham qat'iy yopiq.",
    "Metod kuni oddiy darslar uchun yopiq. Qat'iy model yechimsizligini isbotlasa, faqat 1–4-sinf o'qituvchisining Shanbadan boshqa metod kunidagi ko'pi bilan 2 aniq katak qizil/BANDni ochmaydigan ikkinchi exact model bilan ishlatilishi mumkin; qo'llangan kataklar hisobotda saqlanadi.",
    "KELAJAK SOATI tanlangan kun va dars raqamiga qat'iy joylashtiriladi; boshqa vaqtga surilmaydi.",
    "Sinf faqat o'z smenasida va shu smena uchun yaratilgan mavjud dars raqamlariga joylashtiriladi.",
    "Sinfning qizil/yopiq kuni va 1–4-sinf uchun Shanba taqiqi buzilmaydi.",
    "Har bir fan sinf-fan yuklamasidagi kunlik maksimumdan oshmaydi; kunlik max 1 bo'lsa shu fan bir kunda ikkinchi marta qo'yilmaydi.",
    "1–4-sinfda 6-dars qo'yilmaydi; boshlang'ich sinfning akademik va yengil fan limitlari saqlanadi.",
    "O'qituvchining eng erta–eng kech darsi, kunlik maksimumi va haftalik yuklama chegarasi tekshiriladi.",
    "Bir xona bir vaqtda ikki darsga berilmaydi; xona yozilmagan bo'lsa dars 'Xona yo'q' bilan yaratilishi mumkin.",
    "Sinf jadvalida kun boshidan yoki darslar orasida bo'sh okno qoldirilmaydi.",
    "Sinfning kunlik yuklamasi imkon qadar barqaror taqsimlanadi; keskin 3/5/6 kabi notekis kun saqlanmaydi.",
    "Matematika, algebra, geometriya, ona tili, adabiyot, fizika, kimyo va biologiya kabi asosiy fanlar 1–5-darslarga ustuvor joylashtiriladi; boshqa legal katak qolmasa 6-dars ishlatiladi, lekin bir sinfda haftasiga ko'pi bilan 2 kun.",
    "Jismoniy tarbiya va texnologiya 1-darsga qo'yilmaydi; zarur bo'lsa haftada bir marta yonma-yon juft dars bo'lishi mumkin.",
    "Metod kunidan tashqari ish kataklarining 20 foizidan ko'pi qizil/yopiq bo'lgan o'qituvchilar avval joylashtiriladi; oddiy darslari faqat qolgan yashil kataklarga ixcham va oknosiz yig'iladi.",
    "Ikki smenada ishlaydigan o'qituvchining real vaqt ustma-ustligi qat'iy taqiqlanadi; uzoq kutish keskin jazolanadigan qulaylik mezoni bo'lib, to'liq xavfsiz jadvalni bekor qilmaydi.",
    "Barcha qat'iy darslar sig'masa yarim jadval saqlanmaydi; aniq sig'magan dars va uni to'sgan barcha qoidalar qaytariladi.",
]




def _v212_hygiene_is_soft(reason):
    """Gigiyena hisobotidagi pedagogik tavsiya va qattiq xatoni ajratadi.

    Sinf oynasi, Shanba, maksimal dars soni va smena chegarasi qat'iy qoladi.
    Faqat to'liq jadval uchun zaruratda ishlatilishi mumkin bo'lgan 5-darsdagi
    og'ir fan hamda 5 akademik darsli kunlar soni ogohlantirish bo'ladi.
    """
    normalized = (
        str(reason or "").casefold()
        .replace("‘", "'").replace("’", "'").replace("`", "'")
    )
    return any(marker in normalized for marker in (
        "5-darsda og'ir fan",
        "5 akademik darsli kunlar",
    ))




def _v1852_open_candidates(job, state, context, *, exact):
    """Bitta aniq katakni hard-safe post-processing uchun tekshiradi."""
    rejected = _v1852_Counter()
    day, period = (int(value) for value in exact)

    shift = context["shifts"].get(job["smena"])
    if not shift:
        return [], _v1852_Counter({"smena sozlanmagan": 1})
    if day not in range(1, int(context["weekdays"]) + 1):
        rejected.update(["aniq katak kuni o'qish haftasidan tashqari"])
        return [], rejected

    fixed_day = int(job.get("fixed_day") or 0)
    fixed_period = int(job.get("fixed_period") or 0)
    if fixed_day and day != fixed_day:
        rejected.update(["belgilangan sinf soati kuni boshqa"])
        return [], rejected
    if fixed_period and period != fixed_period:
        rejected.update(["belgilangan sinf soati dars raqami boshqa"])
        return [], rejected

    class_row = context["classes"].get(job["sinf_id"], {})
    blocked_reason = _v1856_class_day_block_reason(
        class_row, day, context.get("class_day_blocks", {})
    )
    if blocked_reason:
        rejected.update([blocked_reason])
        return [], rejected

    available_periods = {
        int(slot["dars_raqami"]) for slot in shift.get("slotlar", [])
    }
    if period not in available_periods:
        rejected.update(["aniq dars raqami smenada mavjud emas"])
        return [], rejected

    teachers = _v1852_choose_teacher(job, day, period, state, context)
    room_keys = _v1852_room_keys(job, teachers, context["classes"])
    reasons = _v1852_candidate_reasons(
        job, day, period, teachers, room_keys, state, context
    )
    if reasons:
        rejected.update(reasons)
        return [], rejected
    return [(0.0, day, period, teachers, room_keys)], rejected


def _v1852_rebuild_schedule_state(placements, context):
    state = _v1852_new_schedule_state()
    for placement in placements:
        _v1852_place_job(
            placement["job"], placement["day"], placement["period"],
            placement["teachers"], placement["room_keys"], state, context,
        )
        if placement.get("method_exceptions"):
            state["placements"][-1]["method_exceptions"] = [
                tuple(int(value) for value in token)
                for token in placement.get("method_exceptions") or []
            ]
    return state




def _v209_class_gap_failure_detail(state, classes):
    by_class_day = _v1852_defaultdict(set)
    for placement in state.get("placements", []):
        by_class_day[(
            int((placement.get("job") or {}).get("sinf_id") or 0),
            int(placement.get("day") or 0),
        )].add(int(placement.get("period") or 0))
    problems = []
    for (class_id, day), periods in sorted(by_class_day.items()):
        clean_periods = sorted(period for period in periods if period > 0)
        if not clean_periods:
            continue
        missing = [
            period for period in range(1, max(clean_periods) + 1)
            if period not in periods
        ]
        if not missing:
            continue
        class_row = classes.get(class_id, {})
        class_name = f"{class_row.get('sinf','')}-{class_row.get('harf','')}"
        day_name = _V1852_HAFTA.get(day, str(day))
        missing_text = ", ".join(f"{period}-dars" for period in missing)
        problems.append({
            "raqam": len(problems) + 1,
            "sinf_id": class_id,
            "sinf": class_name,
            "fan": "Sinf jadvalidagi ichki bo‘shliq",
            "kun": day,
            "kun_nomi": day_name,
            "bosh_darslar": missing,
            "oqituvchilar": [],
            "sabablar": [{
                "sabab": "sinf ichki oknosi",
                "izoh": (
                    f"{class_name}: {day_name} kuni {missing_text} bo‘sh, "
                    f"lekin undan keyin {max(clean_periods)}-darsgacha mashg‘ulot bor."
                ),
                "yechim": (
                    "Sinfning kunlik sig‘imi yoki qattiq o‘qituvchi vaqtlarini "
                    "tahrirlang; generator bo‘shliq bilan draft saqlamaydi."
                ),
            }],
        })
    return {
        "code": "JADVALGA_SIGMADI",
        "asl_code": "SINF_ICHKI_OKNOSI",
        "message": (
            "Sinf jadvalida ichki bo‘sh dars qoldi. Xavfsiz bo‘lmagan draft saqlanmadi."
        ),
        "muammolar": problems,
        "joylashmagan": len(problems),
        "generator_rejimi": 1,
    }




def _v209_integrity_failure_detail(errors):
    problems = []
    for error in errors:
        text = str(error)
        problems.append({
            "raqam": len(problems) + 1,
            "sinf": "Yakuniy tekshiruv",
            "fan": "Jadval mosligi",
            "oqituvchilar": [],
            "sabablar": [{
                "sabab": "yakuniy moslik xatosi",
                "izoh": text,
                "yechim": (
                    "Ko‘rsatilgan sinf, o‘qituvchi, xona yoki yuklama "
                    "sozlamasini tahrirlab yagona generatorni qayta ishga tushiring."
                ),
            }],
        })
    return {
        "code": "JADVALGA_SIGMADI",
        "asl_code": "JADVAL_MOSLIGI_XATOSI",
        "message": "Yakuniy tekshiruv xato topdi; draft to‘liq rollback qilindi.",
        "muammolar": problems,
        "joylashmagan": len(problems),
        "generator_rejimi": 1,
    }




def _v216_exact_job_key(job):
    """Exact natijada bir job yo'qolmagan/takrorlanmaganini tekshirish kaliti."""
    job = job or {}
    if job.get("job_id") is not None:
        return ("job_id", str(job.get("job_id")))
    return (
        "legacy",
        str(job.get("load_id") or ""),
        str(job.get("sinf_id") or ""),
        str(job.get("fan") or ""),
        str(job.get("occurrence") or ""),
        str(job.get("hafta_turi") or "har_hafta"),
    )


def _v216_exact_complete_state(result, jobs, context):
    """Faqat barcha job aynan bir marta qaytgan exact natijani state qiladi."""
    placements = list((result or {}).get("placements") or [])
    expected = _v1852_Counter(_v216_exact_job_key(job) for job in jobs)
    actual = _v1852_Counter(
        _v216_exact_job_key((placement or {}).get("job") or {})
        for placement in placements
    )
    if len(placements) != len(jobs) or actual != expected:
        raise ValueError(
            "Exact solver to'liq deb qaytardi, ammo joblar soni yoki "
            "identifikatorlari manba bilan teng emas"
        )
    for placement in placements:
        if not isinstance(placement, dict):
            raise ValueError("Exact placement obyekt bo'lishi kerak")
        if not int(placement.get("day") or 0) or not int(
            placement.get("period") or 0
        ):
            raise ValueError("Exact placement kuni yoki dars raqami yo'q")
        placement["teachers"] = list(placement.get("teachers") or [])
        placement["room_keys"] = list(placement.get("room_keys") or [])
    # Modul state qaytarsa ham kanonik indekslarni shu faylning faol builderi
    # bilan qayta quramiz. Keyingi qulaylik va SQL validator bir xil formatni
    # ko'radi; solver qaytargan begona/yetishmagan indeksga ishonilmaydi.
    return _v1852_rebuild_schedule_state(placements, context)


def _v216_method_day_recommendations(raw_rows, context, teachers):
    """Strict muvaffaqiyatsiz bo'lsa ko'pi bilan 2 ta qo'lda istisno tavsiyasi.

    Tavsiya hech qachon contextni mutatsiya qilmaydi. Qizil/BAND bilan ham
    yopilgan katak chiqarib tashlanadi: metod kuni istisnosi qizil vaqtni ochib
    yubormaydi. Administrator keyin o'zi tanlamaguncha hech narsa qo'llanmaydi.
    """
    result = []
    seen = set()
    method_hard = set((context or {}).get("method_hard") or set())
    hard = set((context or {}).get("hard") or set())
    shifts = (context or {}).get("shifts") or {}
    for raw in list(raw_rows or []):
        if len(result) >= 2:
            break
        raw = dict(raw or {})
        recommendation_kind = str(
            raw.get("turi") or raw.get("kind") or "metod_kuni"
        ).casefold()
        if recommendation_kind not in {"metod_kuni", "method_day"}:
            continue
        # Exact analyzer complete relaxed solve bilan isbotlamagan satr UIga
        # chiqmaydi. Maydon bo'lmagan eski adapter natijasi pastdagi method/hard
        # va qizil tekshiruvlardan baribir o'tishi shart.
        if raw.get("isbotlangan") is False or raw.get("proven") is False:
            continue
        try:
            teacher_id = int(
                raw.get("oqituvchi_id")
                or raw.get("teacher_id")
                or raw.get("user_id")
                or 0
            )
            day = int(raw.get("kun") or raw.get("day") or 0)
            period = int(
                raw.get("dars")
                or raw.get("dars_raqami")
                or raw.get("period")
                or 0
            )
        except (TypeError, ValueError):
            continue
        if not teacher_id or not day or not period:
            continue
        if (teacher_id, day) not in method_hard:
            continue
        requested_shift = raw.get("smena") or raw.get("shift")
        try:
            requested_shift = int(requested_shift) if requested_shift else None
        except (TypeError, ValueError):
            requested_shift = None
        shift_candidates = (
            [requested_shift] if requested_shift in shifts else sorted(shifts)
        )
        safe_shift = next((
            int(shift)
            for shift in shift_candidates
            if not _v1852_blocked(
                hard, teacher_id, day, int(shift), period
            )
        ), None)
        if safe_shift is None:
            # Metod qatorini yumshatish baribir qizil/BAND sabab yordam bermaydi.
            continue
        signature = (teacher_id, day, safe_shift, period)
        if signature in seen:
            continue
        seen.add(signature)
        slot = next((
            item for item in (shifts.get(safe_shift) or {}).get("slotlar", [])
            if int(item.get("dars_raqami") or 0) == period
        ), None)
        time_label = None
        if slot:
            start = str(slot.get("boshlanish") or "")[:5]
            end = str(slot.get("tugash") or "")[:5]
            if start or end:
                time_label = f"{start}–{end}".strip("–")
        teacher_name = str(
            raw.get("oqituvchi")
            or raw.get("teacher_name")
            or (teachers.get(teacher_id) or {}).get("full_name")
            or teacher_id
        )
        lessons = raw.get("joylashadigan_darslar")
        if lessons is None:
            lessons = raw.get("affected_lessons") or raw.get("jobs") or []
        if not isinstance(lessons, list):
            lessons = [lessons] if lessons else []
        result.append({
            "raqam": len(result) + 1,
            "oqituvchi_id": teacher_id,
            "oqituvchi": teacher_name,
            "kun": day,
            "kun_nomi": _V1852_HAFTA.get(day, str(day)),
            "smena": safe_shift,
            "dars": period,
            "vaqt": time_label,
            "sabab": str(
                raw.get("sabab")
                or raw.get("reason")
                or "Shu metod kuni katagini ochish strict modelga qo'shimcha legal variant beradi."
            ),
            "joylashadigan_darslar": lessons,
            "kamayadigan_oynalar": int(
                raw.get("kamayadigan_oynalar")
                or raw.get("reduced_gaps")
                or raw.get("kamayish")
                or 0
            ),
            "qizil_buzilmaydi": True,
            "avtomatik_qollanmagan": True,
            "amal": (
                f"{_V1852_HAFTA.get(day, day)} kuni {teacher_name} uchun "
                f"faqat {safe_shift}-smena {period}-darsni metod kuni "
                "istisnosi sifatida qo'lda ruxsat bering; qizil/BAND "
                "vaqtlarini o'zgartirmang."
            ),
        })
    return result


def _v217_exact_failure_problems(exact_result, classes, teachers, status=None):
    """Exact diagnostikani frontend tushunadigan, hech qachon bo'sh bo'lmagan ro'yxatga aylantiradi."""
    diagnostics = dict((exact_result or {}).get("diagnostics") or {})
    classes = classes or {}
    teachers = teachers or {}
    problems = []

    def class_name(class_id):
        row = classes.get(int(class_id or 0), {}) or {}
        grade = str(row.get("sinf") or "").strip()
        letter = str(row.get("harf") or "").strip()
        return f"{grade}-{letter}".strip("-") or str(class_id or "Noma'lum")

    def teacher_rows(teacher_ids):
        rows = []
        for teacher_id in teacher_ids or []:
            try:
                teacher_id = int(teacher_id)
            except (TypeError, ValueError):
                continue
            rows.append({
                "user_id": teacher_id,
                "full_name": str(
                    (teachers.get(teacher_id) or {}).get("full_name")
                    or teacher_id
                ),
            })
        return rows

    for item in list(diagnostics.get("empty_domains") or []):
        item = dict(item or {})
        class_id = item.get("sinf_id")
        label = class_name(class_id)
        fixed_day = int(item.get("fixed_day") or 0)
        fixed_period = int(item.get("fixed_period") or 0)
        fixed_text = (
            f" Dars {_V1852_HAFTA.get(fixed_day, fixed_day)} kuni "
            f"{fixed_period}-darsga qat'iy biriktirilgan."
            if fixed_day and fixed_period else ""
        )
        problem_teachers = teacher_rows(item.get("teacher_ids") or [])
        problems.append({
            "raqam": len(problems) + 1,
            "sarlavha": f"{label} sinf · {item.get('fan') or 'Dars'}",
            "sinf": label,
            "sinf_id": class_id,
            "fan": item.get("fan") or "Dars",
            "smena": int(item.get("smena") or 1),
            "takror_raqami": int(item.get("takror_raqami") or 1),
            "oqituvchilar": problem_teachers,
            "sabablar": [{
                "sabab": "legal katak qolmagan",
                "izoh": (
                    "Bu dars uchun smena, sinf kuni, o'qituvchi qizil/BAND, "
                    "metod kuni va ruxsat etilgan dars oralig'idan o'tadigan "
                    "bitta ham katak qolmagan." + fixed_text
                ),
                "yechim": (
                    "Agar bu qat'iy Kelajak/Sinf soati bo'lsa uning tanlangan "
                    "kun-soatini tekshiring. Oddiy fan bo'lsa ko'rsatilgan "
                    "o'qituvchining metod kuni va BAND vaqtlarini tekshiring."
                ),
            }],
        })

    for conflict in list(diagnostics.get("hard_conflicts") or []):
        conflict = dict(conflict or {})
        kind = str(conflict.get("kind") or "global_resource_cycle")
        teacher_id = conflict.get("teacher_id")
        if kind == "teacher_capacity":
            title = f"O'qituvchi sig'imi · {conflict.get('teacher_name') or teacher_id}"
            label = "O'qituvchi sig'imi"
            fan = str(conflict.get("teacher_name") or teacher_id or "O'qituvchi")
            linked_teachers = teacher_rows([teacher_id])
        elif kind == "class_capacity":
            label = str(conflict.get("class_name") or class_name(conflict.get("class_id")))
            title = f"{label} sinf · haftalik sig'im"
            fan = "Haftalik sig'im"
            linked_teachers = []
        elif kind == "fixed_collision":
            label = str(conflict.get("class_name") or "Qat'iy darslar")
            title = f"Qat'iy vaqt to'qnashuvi · {conflict.get('resource') or label}"
            fan = "Qat'iy vaqt to'qnashuvi"
            linked_teachers = []
        else:
            label = "Butun jadval"
            title = "Butun jadval · global resurs ziddiyati"
            fan = "Global resurs ziddiyati"
            linked_teachers = []
        problems.append({
            "raqam": len(problems) + 1,
            "sarlavha": title,
            "sinf": label,
            "sinf_id": conflict.get("class_id"),
            "fan": fan,
            "oqituvchilar": linked_teachers,
            "sabablar": [{
                "sabab": kind,
                "izoh": str(
                    conflict.get("message")
                    or "Qattiq resurs cheklovlari bir vaqtda ziddiyat qilgan."
                ),
                "yechim": str(
                    conflict.get("solution")
                    or "O'qituvchi qat'iy vaqtlarini va oldindan qotirilgan darslarni tekshiring."
                ),
            }],
            "isbot": {
                key: conflict.get(key)
                for key in (
                    "phase", "required_lessons", "available_lessons",
                    "shortage", "method_days", "day", "period",
                )
                if conflict.get(key) is not None
            },
        })

    if not problems:
        normalized_status = str(
            status or (exact_result or {}).get("status") or "UNKNOWN"
        ).upper()
        proven = normalized_status == "INFEASIBLE"
        model_invalid = normalized_status == "MODEL_INVALID"
        technical_detail = "; ".join(
            str(value) for value in diagnostics.get("validation_errors") or []
        ) or str(diagnostics.get("exception") or "Model texnik tekshiruvdan o'tmadi.")
        problems.append({
            "raqam": 1,
            "sarlavha": (
                "Butun jadval · global resurs ziddiyati"
                if proven else
                "Exact model · texnik tekshiruv xatosi"
                if model_invalid else
                "Butun jadval · qidiruv xulosasi yakunlanmadi"
            ),
            "sinf": "Butun jadval",
            "fan": (
                "Global resurs ziddiyati" if proven else
                "Model tekshiruvi" if model_invalid else
                "Qidiruv diagnostikasi"
            ),
            "oqituvchilar": [],
            "sabablar": [{
                "sabab": (
                    "global qattiq ziddiyat" if proven else
                    "model texnik xatosi" if model_invalid else
                    "global sabab ajratilmadi"
                ),
                "izoh": ((
                    "Har bir darsda alohida legal katak bo'lishi mumkin, ammo "
                    "ularni sinf, o'qituvchi va xona real-vaqt "
                    "to'qnashuvlarisiz birgalikda tanlab bo'lmadi."
                ) if proven else technical_detail if model_invalid else (
                    "Qidiruv vaqt chegarasida tugadi; bu jadval imkonsiz "
                    "degan isbot emas va taxminiy qattiq ziddiyat ko'rsatilmaydi."
                )),
                "yechim": (
                    "Metod-kuni bo'yicha isbotlangan tavsiya chiqsa faqat o'sha "
                    "katakni ko'rib chiqing; aks holda qayta qidirish mumkin."
                ),
            }],
        })
    return problems[:100]


def _v216_exact_failure_detail(
    status, exact_result, recommendations, classes=None, teachers=None
):
    """CP-SAT statusini foydalanuvchiga yolg'onsiz, kanonik ko'rinishda beradi."""
    normalized = str(status or "UNKNOWN").upper()
    proof_complete = normalized == "INFEASIBLE"
    if normalized == "INFEASIBLE":
        code = "QAT_IY_QOIDALARDA_YECHIM_YOQ"
        message = (
            "Exact solver qizil/BAND, metod kuni, qat'iy dars, smena hamda "
            "sinf/o'qituvchi/xona real-vaqt xavfsizlik qoidalarini saqlagan "
            "holda to'liq jadval topilmasligini isbotladi. Quyida aniq "
            "sig'im yoki ziddiyat ko'rsatilgan; yarim draft saqlanmadi."
        )
        retry = False
    elif normalized == "MODEL_INVALID":
        code = "EXACT_MODEL_INVALID"
        message = (
            "Exact jadval modeli texnik tekshiruvdan o'tmadi. Legacy natija "
            "bilan yashirilmadi va hech qanday draft saqlanmadi."
        )
        retry = False
    else:
        normalized = "UNKNOWN"
        code = "HISOBLASH_VAQTI_TUGADI"
        message = (
            "Exact qidiruv vaqt chegarasida to'liq jadval topmadi, lekin "
            "imkonsizlikni ham isbotlamadi. Qisman natija saqlanmadi."
        )
        retry = True
    problems = _v217_exact_failure_problems(
        exact_result, classes or {}, teachers or {}, normalized
    )
    exact_diagnostics = dict((exact_result or {}).get("diagnostics") or {})
    method_analysis = dict(
        exact_diagnostics.get("metod_kuni_tahlili") or {}
    )
    method_status = str(method_analysis.get("status") or "NOT_RUN").upper()
    if recommendations:
        method_explanation = (
            "Alohida relaxed exact tekshiruv qizil/BAND vaqtlarini saqlagan "
            "holda quyidagi 1–2 metod katagi yordam berishini to'liq jadval "
            "bilan tasdiqladi; istisno avtomatik qo'llanmadi."
        )
    elif method_status == "UNKNOWN":
        method_explanation = (
            "Metod-kuni bo'yicha alohida tekshiruv vaqt chegarasida xulosa "
            "bermadi. Bo'sh tavsiya 'yordam bermaydi' degan isbot emas."
        )
    elif method_status == "ERROR":
        method_explanation = (
            "Metod-kuni diagnostikasida texnik xato qaytdi; shu sabab taxminiy "
            "katak tavsiya qilinmadi."
        )
    elif method_status == "INFEASIBLE":
        method_explanation = (
            "Qizil/BANDni saqlab ko'pi bilan ikki metod katagini ochadigan "
            "alohida model ham to'liq jadval topolmasligini isbotladi."
        )
    elif method_status == "NOT_NEEDED":
        method_explanation = "Qattiq metod kuni belgilanmagan; metod istisnosi kerak emas."
    else:
        method_explanation = (
            "Qizil/BAND vaqtlarini saqlab 1–2 metod katagi bilan to'liq yechim "
            "tasdiqlanmadi; taxminiy istisno berilmadi."
        )
    try:
        calculation_seconds = round(float(
            (exact_result or {}).get("wall_time_seconds")
            or exact_diagnostics.get("solver_wall_time_seconds")
            or 0.0
        ), 3)
    except (TypeError, ValueError):
        calculation_seconds = 0.0
    try:
        method_seconds = round(float(
            method_analysis.get("wall_time_seconds") or 0.0
        ), 3)
    except (TypeError, ValueError):
        method_seconds = 0.0
    return {
        "code": code,
        "solver_status": normalized,
        "proof_complete": proof_complete,
        "message": message,
        "qayta_urinish_mumkin": retry,
        "muammolar": problems,
        "joylashtirilmagan": len(problems),
        "joylashtirilgan_qisman_natija_saqlanmadi": True,
        "metod_kuni_istisno_tavsiyalari": list(recommendations or [])[:2],
        "metod_kuni_tavsiya_izohi": method_explanation,
        "metod_kuni_tahlili": method_analysis,
        "avtomatik_yumshatish": False,
        "hisoblash_soniya": calculation_seconds,
        "kandidat_soni": int(exact_diagnostics.get("candidates") or 0),
        "model_ozgaruvchi_soni": int(
            exact_diagnostics.get("model_variables") or 0
        ),
        "model_cheklov_soni": int(
            exact_diagnostics.get("model_constraints") or 0
        ),
        "solver_conflicts": int(exact_diagnostics.get("conflicts") or 0),
        "solver_branches": int(exact_diagnostics.get("branches") or 0),
        "metod_qidiruv_holati": method_status,
        "metod_qidiruv_soniya": method_seconds,
        "diagnostika": exact_diagnostics,
    }


def _v236_nonce_matches(stored_nonce, expected_nonce):
    """PostgreSQL ``IS NOT DISTINCT FROM`` bilan bir xil nonce taqqoslash."""
    if stored_nonce is None or expected_nonce is None:
        return stored_nonce is None and expected_nonce is None
    try:
        return int(stored_nonce) == int(expected_nonce)
    except (TypeError, ValueError):
        return str(stored_nonce) == str(expected_nonce)


def _v2243_progress_write(
    maktab_id, qidiruv_nonce, jadval_raqami, yaxshilanish, foiz, bosqich, xabar
):
    """Faqat aynan shu server lease + run hali joriy bo'lsa progress yozadi.

    Stop signali tushganidan keyin heartbeat yoki oddiy progress bir bayt ham
    yangilanmaydi. READY bilan Stop poygasida esa shu UPDATE linearizatsiya
    nuqtasi bo'ladi: Stop oldin commit qilgan bo'lsa DB holati ``toxtatildi``
    bo'lib qoladi; READY oldin commit qilsa keyingi Stop terminal qatorga tegmaydi.
    """
    progress_conn = None
    progress_cur = None
    written = False
    try:
        progress_conn = _db()
        progress_cur = progress_conn.cursor()
        requested_stage = str(bosqich or "hisoblash")
        requested_message = str(xabar or "Jadval hisoblanmoqda")
        revision = max(0, int(yaxshilanish or 0))
        stopped_number = str(int(jadval_raqami)) + (
            f".{revision}" if revision else ""
        )
        stopped_message = (
            f"To‘xtatildi. Eng yaxshi Jadval #{stopped_number} saqlandi "
            "va ochish mumkin."
        )
        cancelled_error_message = (
            "To‘xtatildi. 100% checkpoint saqlangan bo‘lsa uni ochish mumkin; "
            "aks holda oldingi jadval o‘zgarmadi."
        )
        progress_cur.execute(
            """UPDATE aqlli_jadval_jarayoni_v2243
                  SET yaxshilanish=%s,
                      foiz=GREATEST(foiz,%s),
                      bosqich=CASE
                          WHEN toxtatish_soraldi=TRUE
                           AND %s IN ('tayyor','xato','toxtatildi')
                          THEN 'toxtatildi'
                          ELSE %s
                      END,
                      xabar=CASE
                          WHEN toxtatish_soraldi=TRUE
                           AND %s='tayyor'
                          THEN %s
                          WHEN toxtatish_soraldi=TRUE
                           AND %s='xato'
                          THEN %s
                          ELSE %s
                      END,
                      yangilangan_at=NOW()
                WHERE maktab_id=%s
                  AND qidiruv_nonce IS NOT DISTINCT FROM %s
                  AND jadval_raqami=%s
                  AND bosqich NOT IN ('tayyor','xato','toxtatildi')
                  AND (
                      toxtatish_soraldi=FALSE
                      OR %s IN ('tayyor','xato','toxtatildi')
                  )
                RETURNING bosqich""",
            (
                revision,
                max(0, min(100, int(foiz or 0))),
                requested_stage,
                requested_stage,
                requested_stage,
                stopped_message,
                requested_stage,
                cancelled_error_message,
                requested_message,
                int(maktab_id),
                int(qidiruv_nonce) if qidiruv_nonce is not None else None,
                int(jadval_raqami),
                requested_stage,
            ),
        )
        written = int(progress_cur.rowcount or 0) > 0
        progress_conn.commit()
    except Exception as progress_error:
        if progress_conn is not None:
            try:
                progress_conn.rollback()
            except Exception:
                pass
        print(f"[JADVAL-PROGRESS-V22.43] yozilmadi: {progress_error}", flush=True)
    finally:
        if progress_cur is not None:
            progress_cur.close()
        if progress_conn is not None:
            progress_conn.close()
    return bool(written)


def _v2244_cancel_requested(maktab_id, qidiruv_nonce=None, run_id=None):
    """Aynan joriy server lease + run uchun Stop signalini o‘qiydi.

    DB tekshiruvi ishlamasa fail-closed: solver eski yoki nazoratsiz lease bilan
    davom etmaydi. Canonical start workerga hech qachon None/0 lease bermaydi.
    """
    # Stop endpoint va generator bir workerda bo'lsa DB qayta o'qilishi
    # vaqtincha xato qilgan paytda ham lokal signal darhol ishlaydi. Boshqa
    # workerda bo'lsa PostgreSQL qatori umumiy signal bo'lib qoladi.
    local_cancelled = globals().get("_V2244_LOCAL_CANCELLED")
    local_lock = globals().get("_V2244_BACKGROUND_LOCK")
    if local_cancelled is not None and local_lock is not None:
        with local_lock:
            if qidiruv_nonce is None:
                if any(
                    int(key[0]) == int(maktab_id)
                    for key in local_cancelled
                ):
                    return True
            elif run_id is None:
                if any(
                    int(key[0]) == int(maktab_id)
                    and int(key[1]) == int(qidiruv_nonce)
                    for key in local_cancelled
                ):
                    return True
            elif (
                int(maktab_id), int(qidiruv_nonce), int(run_id)
            ) in local_cancelled:
                return True
    cancel_conn = None
    cancel_cur = None
    try:
        cancel_conn = _db(); cancel_cur = cancel_conn.cursor()
        cancel_cur.execute(
            "SELECT qidiruv_nonce,jadval_raqami,bosqich,toxtatish_soraldi FROM aqlli_jadval_jarayoni_v2243 WHERE maktab_id=%s",
            (int(maktab_id),),
        )
        row = cancel_cur.fetchone()
        if not row:
            return True
        # Qatordagi nonce o'zgargan bo'lsa bu worker endi eski: yangi qidiruv
        # progressini yoki checkpointini bosmasligi uchun darhol to'xtaydi.
        if not _v236_nonce_matches(
            row.get("qidiruv_nonce"), qidiruv_nonce
        ):
            return True
        if run_id is not None and int(row.get("jadval_raqami") or 0) != int(run_id):
            return True
        if row.get("bosqich") in {"tayyor", "xato", "toxtatildi"}:
            return True
        return bool(row.get("toxtatish_soraldi"))
    except Exception:
        return True
    finally:
        if cancel_cur is not None: cancel_cur.close()
        if cancel_conn is not None: cancel_conn.close()


def _v236_generation_expiry_stage(
    bosqich, toxtatish_soraldi, heartbeat_age
):
    """Hung worker uchun terminal holatni sof va sinovbop aniqlaydi.

    Stop so'rovi foydalanuvchi kutayotgan interaktiv amal bo'lgani uchun uning
    grace muddati qisqa. Oddiy worker esa vaqtinchalik DB/CPU kechikishida xato
    deb belgilanmasligi uchun avvalgi 90 soniyalik chegarani saqlaydi.
    """
    if str(bosqich or "") in {"tayyor", "xato", "toxtatildi"}:
        return None
    age = max(0.0, float(heartbeat_age or 0.0))
    if bool(toxtatish_soraldi):
        return "toxtatildi" if age >= 15.0 else None
    return "xato" if age >= 90.0 else None


@app.get("/api/maktab/aqlli_jadval/v3/jarayon")
def v2243_generation_progress(token: str, maktab_id: int, qidiruv_nonce: Optional[int] = None):
    user_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor()
    try:
        _v1852_tables(cur)
        if not _v1852_manager(cur, user_id, maktab_id):
            raise HTTPException(status_code=403, detail="Jadval jarayonini ko‘rishga ruxsat yo‘q")
        cur.execute(
            """SELECT p.*,
                      EXTRACT(EPOCH FROM (NOW()-p.yangilangan_at)) AS yangilanish_yoshi,
                      EXISTS(
                          SELECT 1
                            FROM aqlli_jadval_urinishlari_v2 r
                           WHERE r.id=p.jadval_raqami
                             AND r.maktab_id=p.maktab_id
                             AND r.joylashtirilmadi=0
                      ) AS toliq_jadval_saqlandi
                 FROM aqlli_jadval_jarayoni_v2243 p
                WHERE p.maktab_id=%s""",
            (maktab_id,),
        )
        row = cur.fetchone()
        if not row or (qidiruv_nonce is not None and int(row.get("qidiruv_nonce") or 0) != int(qidiruv_nonce)):
            conn.commit()
            return {"faol": False}
        payload = dict(row)
        terminal = {"tayyor", "xato", "toxtatildi"}
        heartbeat_age = max(0.0, float(
            payload.get("yangilanish_yoshi") or 0
        ))
        expiry_stage = _v236_generation_expiry_stage(
            payload.get("bosqich"),
            payload.get("toxtatish_soraldi"),
            heartbeat_age,
        )
        stale = expiry_stage is not None
        if expiry_stage == "toxtatildi":
            saved_number = str(payload.get("jadval_raqami") or "")
            saved_revision = int(payload.get("yaxshilanish") or 0)
            if saved_number and saved_number != "0" and saved_revision > 0:
                saved_number += f".{saved_revision}"
            stale_message = (
                f"To‘xtatildi. Jadval #{saved_number} saqlandi va ochish mumkin."
                if bool(payload.get("toliq_jadval_saqlandi")) else
                "To‘xtatildi. Hali 100% to‘liq jadval saqlanmagan; "
                "oldingi jadval o‘chirilmagan."
            )
            cur.execute(
                """UPDATE aqlli_jadval_jarayoni_v2243
                      SET bosqich='toxtatildi',foiz=100,xabar=%s,
                          yangilangan_at=NOW()
                    WHERE maktab_id=%s
                      AND qidiruv_nonce IS NOT DISTINCT FROM %s
                      AND jadval_raqami=%s
                      AND toxtatish_soraldi=TRUE
                      AND yangilangan_at <= NOW() - INTERVAL '15 seconds'
                      AND bosqich NOT IN ('tayyor','xato','toxtatildi')""",
                (
                    stale_message, maktab_id,
                    payload.get("qidiruv_nonce"),
                    int(payload.get("jadval_raqami") or 0),
                ),
            )
            if int(cur.rowcount or 0) > 0:
                payload.update({
                    "bosqich": "toxtatildi", "foiz": 100,
                    "xabar": stale_message,
                })
            else:
                stale = False
        elif expiry_stage == "xato":
            stale_message = (
                "Generator workerining yurak urishi 90 soniyadan oshdi. "
                "Jarayon faol emas; oldingi 100% jadval saqlangan bo'lsa u "
                "o'chirilmagan. Yangi qidiruvni boshlash mumkin."
            )
            cur.execute(
                """UPDATE aqlli_jadval_jarayoni_v2243
                      SET bosqich='xato',foiz=100,xabar=%s,
                          toxtatish_soraldi=FALSE,yangilangan_at=NOW()
                    WHERE maktab_id=%s
                      AND qidiruv_nonce IS NOT DISTINCT FROM %s
                      AND jadval_raqami=%s
                      AND toxtatish_soraldi=FALSE
                      AND yangilangan_at <= NOW() - INTERVAL '90 seconds'
                      AND bosqich NOT IN ('tayyor','xato','toxtatildi')""",
                (
                    stale_message, maktab_id,
                    payload.get("qidiruv_nonce"),
                    int(payload.get("jadval_raqami") or 0),
                ),
            )
            if int(cur.rowcount or 0) > 0:
                payload.update({
                    "bosqich": "xato", "foiz": 100,
                    "xabar": stale_message, "toxtatish_soraldi": False,
                })
            else:
                stale = False
        conn.commit()
        payload["faol"] = payload.get("bosqich") not in terminal
        payload["eskirgan"] = bool(stale)
        payload["yangilanish_yoshi_soniya"] = heartbeat_age
        payload["ko_rinish_raqami"] = (
            f"{payload['jadval_raqami']}.{payload['yaxshilanish']}"
            if int(payload.get("yaxshilanish") or 0) > 0
            else str(payload["jadval_raqami"])
        )
        return payload
    finally:
        cur.close(); conn.close()


class V1852Generate(BaseModel):
    maktab_id: int
    urinishlar_soni: int = 1
    # Eski frontend client nonce yuborishi mumkin, lekin server uni lease deb
    # qabul qilmaydi. Canonical start PostgreSQL sequence'dan yangi qidiruv_nonce
    # yaratadi va javobda qaytaradi; polling/Stop aynan shu qiymatni ishlatadi.
    qidiruv_nonce: Optional[int] = None
    # Eski frontendlar yuborishi mumkin. Yagona generator bu qiymatni
    # e'tiborsiz qoldiradi va har doim bir xil kuchli siyosatni ishlatadi.
    generator_rejimi: int = 1


@app.post("/api/maktab/aqlli_jadval/v3/toxtatish")
def v2244_stop_generation(token: str, maktab_id: int, qidiruv_nonce: Optional[int] = None):
    user_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor()
    try:
        _v1852_tables(cur)
        if not _v1852_manager(cur, user_id, maktab_id):
            raise HTTPException(status_code=403, detail="Jadval jarayonini to‘xtatishga ruxsat yo‘q")
        cur.execute(
            """UPDATE aqlli_jadval_jarayoni_v2243
               SET toxtatish_soraldi=TRUE,bosqich='toxtatish_soraldi',
                   xabar='Foydalanuvchi to‘xtatishni so‘radi. Eng yaxshi to‘liq natija xavfsiz yakunlanmoqda.',
                   yangilangan_at=NOW()
               WHERE maktab_id=%s
                 AND qidiruv_nonce IS NOT DISTINCT FROM %s
                 AND bosqich NOT IN ('tayyor','xato','toxtatildi')
               RETURNING jadval_raqami,qidiruv_nonce""",
            (maktab_id, qidiruv_nonce),
        )
        row = cur.fetchone(); conn.commit()
        if row:
            job_key = (
                int(maktab_id), int(row.get("qidiruv_nonce") or 0),
                int(row.get("jadval_raqami") or 0),
            )
            with _V2244_BACKGROUND_LOCK:
                # Faqat shu process exact lease/run workeriga egalik qilsa
                # local signal qo'yiladi. Boshqa Gunicorn workeri DB flagni
                # ko'radi; begona local marker yangi threadga yopishmaydi.
                if job_key in _V2244_BACKGROUND_JOBS:
                    _V2244_LOCAL_CANCELLED.add(job_key)
        return {
            "qabul_qilindi": bool(row),
            "jadval_raqami": row.get("jadval_raqami") if row else None,
            "qidiruv_nonce": row.get("qidiruv_nonce") if row else None,
        }
    finally:
        cur.close(); conn.close()



def _v2253_saved_run_placements(cur, run_id, jobs, context):
    """Rebuild one already-saved complete draft into the in-memory optimizer state.

    V22.53 does not throw away a good #57 merely because the user presses the
    generator again.  When the source hash is unchanged, the stored full draft
    is the incumbent.  We identify every canonical job by load_id + occurrence
    (and A/B member phase), then restore the exact day/period and teacher set.
    """
    cur.execute(
        """SELECT sinf_id,hafta_kuni,smena,dars_raqami,fan_nomi,
                  oqituvchi_user_id,guruh_kaliti,xona_id,xona_matni,
                  yuklama_id,takror_raqami,hafta_turi
           FROM aqlli_jadval_slotlari_v2
           WHERE urinish_id=%s
           ORDER BY hafta_kuni,smena,dars_raqami,sinf_id,guruh_kaliti,hafta_turi,id""",
        (int(run_id),),
    )
    slots = [dict(row) for row in cur.fetchall()]
    if not slots:
        return None

    by_load = _v1852_defaultdict(list)
    class_hour_rows = _v1852_defaultdict(list)
    for row in slots:
        if row.get("yuklama_id") is None:
            class_hour_rows[
                (
                    int(row.get("sinf_id") or 0),
                    _v1875_subject_key(row.get("fan_nomi")),
                    int(row.get("takror_raqami") or 1),
                )
            ].append(row)
        else:
            by_load[
                (
                    int(row.get("yuklama_id") or 0),
                    int(row.get("takror_raqami") or 1),
                    str(row.get("hafta_turi") or "har_hafta"),
                )
            ].append(row)

    def position(rows):
        values = {
            (int(row.get("hafta_kuni") or 0), int(row.get("dars_raqami") or 0))
            for row in rows
        }
        return next(iter(values)) if len(values) == 1 else None

    placements = []
    for job in jobs:
        selected_rows = []
        members = job.get("rotation_members") or []
        if members:
            positions = set()
            for member in members:
                phase = str(member.get("hafta_turi") or "toq")
                rows = by_load.get(
                    (
                        int(member.get("load_id") or 0),
                        int(member.get("occurrence") or 1),
                        phase,
                    ),
                    [],
                )
                if not rows:
                    return None
                pos = position(rows)
                if not pos:
                    return None
                positions.add(pos)
                selected_rows.extend(rows)
            if len(positions) != 1:
                return None
            day, period = next(iter(positions))
            teachers = [
                teacher for teacher in _v1852_choose_teacher(
                    job, day, period, _v1852_new_schedule_state(), context
                )
                if teacher is not None
            ]
        elif job.get("is_class_hour"):
            rows = class_hour_rows.get(
                (
                    int(job.get("sinf_id") or 0),
                    _v1875_subject_key(job.get("fan")),
                    int(job.get("occurrence") or 1),
                ),
                [],
            )
            if not rows:
                return None
            pos = position(rows)
            if not pos:
                return None
            day, period = pos
            selected_rows = list(rows)
            teacher = next(
                (
                    int(row["oqituvchi_user_id"])
                    for row in rows
                    if row.get("oqituvchi_user_id") is not None
                ),
                None,
            )
            if teacher is None:
                options = [
                    int(value) for value in job.get("teacher_options") or []
                    if value is not None
                ]
                teacher = options[0] if options else None
            teachers = [teacher] if teacher is not None else []
        else:
            rows = by_load.get(
                (
                    int(job.get("load_id") or 0),
                    int(job.get("occurrence") or 1),
                    "har_hafta",
                ),
                [],
            )
            # A half-hour job that was not paired can be stored as TOQ/JUFT.
            if not rows and float(job.get("rotation_weight") or 1.0) < 1.0:
                rows = (
                    by_load.get(
                        (
                            int(job.get("load_id") or 0),
                            int(job.get("occurrence") or 1),
                            "toq",
                        ),
                        [],
                    )
                    or by_load.get(
                        (
                            int(job.get("load_id") or 0),
                            int(job.get("occurrence") or 1),
                            "juft",
                        ),
                        [],
                    )
                )
            if not rows:
                return None
            pos = position(rows)
            if not pos:
                return None
            day, period = pos
            selected_rows = list(rows)
            if job.get("groups"):
                row_by_group = {
                    str(row.get("guruh_kaliti") or "whole"): row
                    for row in rows
                }
                teachers = []
                for group in job.get("groups") or []:
                    row = row_by_group.get(str(group.get("guruh_kaliti") or "whole"))
                    teacher = (
                        int(row["oqituvchi_user_id"])
                        if row and row.get("oqituvchi_user_id") is not None
                        else group.get("teacher")
                    )
                    if teacher is not None:
                        teachers.append(int(teacher))
            else:
                teacher = next(
                    (
                        int(row["oqituvchi_user_id"])
                        for row in rows
                        if row.get("oqituvchi_user_id") is not None
                    ),
                    None,
                )
                if teacher is None:
                    options = [
                        int(value) for value in job.get("teacher_options") or []
                        if value is not None
                    ]
                    teacher = options[0] if options else None
                teachers = [teacher] if teacher is not None else []

        room_keys = _v1852_room_keys(job, teachers, context.get("classes") or {})
        placements.append(
            {
                "job": job,
                "day": int(day),
                "period": int(period),
                "teachers": list(teachers),
                "room_keys": list(room_keys),
                "penalty": 0,
            }
        )

    if len(placements) != len(jobs):
        return None
    return placements


def _v2253_entry_rows_from_state(
    state, run_id, maktab_id, classes, rooms, shifts
):
    """Serialize an optimized incumbent back to the existing run id."""
    shift_slot_map = {
        (s, int(slot["dars_raqami"])): slot
        for s, row in shifts.items()
        for slot in row["slotlar"]
    }
    rows = []
    for placement in state.get("placements") or []:
        job = placement["job"]
        day = int(placement["day"])
        period = int(placement["period"])
        selected_teachers = list(placement.get("teachers") or [])
        time_slot = shift_slot_map[(int(job["smena"]), period)]

        if job.get("rotation_members"):
            class_row = classes[job["sinf_id"]]
            home_room_id = class_row.get("_home_room_id")
            home_room_text = class_row.get("_home_room_text")
            for member in job["rotation_members"]:
                phase = member.get("hafta_turi") or "toq"
                member_teachers = _v199_rotation_member_teachers(member)
                if member.get("groups"):
                    for group_index, (group, teacher) in enumerate(
                        zip(member["groups"], member_teachers)
                    ):
                        room_id = group.get("xona_id")
                        if room_id:
                            room_text = rooms.get(int(room_id), {}).get("nomi")
                        elif group_index == 0:
                            room_id = home_room_id
                            room_text = home_room_text
                        else:
                            room_text = None
                        rows.append(
                            (
                                run_id, maktab_id, member["sinf_id"], day,
                                member["smena"], period, member["fan"], teacher,
                                group["guruh_kaliti"], room_id, room_text,
                                time_slot["boshlanish"], time_slot["tugash"],
                                member["load_id"], member["occurrence"], phase,
                            )
                        )
                else:
                    teacher = member_teachers[0] if member_teachers else None
                    room_id = member.get("room_id")
                    if room_id:
                        room_text = rooms.get(int(room_id), {}).get("nomi")
                    else:
                        room_id = home_room_id
                        room_text = home_room_text
                    rows.append(
                        (
                            run_id, maktab_id, member["sinf_id"], day,
                            member["smena"], period, member["fan"], teacher,
                            "whole", room_id, room_text,
                            time_slot["boshlanish"], time_slot["tugash"],
                            member["load_id"], member["occurrence"], phase,
                        )
                    )
        elif job.get("groups"):
            class_row = classes[job["sinf_id"]]
            home_room_id = class_row.get("_home_room_id")
            home_room_text = class_row.get("_home_room_text")
            for group_index, (group, teacher) in enumerate(
                zip(job["groups"], selected_teachers)
            ):
                room_id = group.get("xona_id")
                if room_id:
                    room_text = rooms.get(int(room_id), {}).get("nomi")
                elif group_index == 0:
                    room_id = home_room_id
                    room_text = home_room_text
                else:
                    room_text = None
                rows.append(
                    (
                        run_id, maktab_id, job["sinf_id"], day, job["smena"],
                        period, job["fan"], teacher, group["guruh_kaliti"],
                        room_id, room_text, time_slot["boshlanish"],
                        time_slot["tugash"], job["load_id"], job["occurrence"],
                        "har_hafta",
                    )
                )
        else:
            teacher = selected_teachers[0] if selected_teachers else None
            room_id = job.get("room_id")
            class_row = classes[job["sinf_id"]]
            if room_id:
                room_text = rooms.get(int(room_id), {}).get("nomi")
            else:
                room_id = class_row.get("_home_room_id")
                room_text = class_row.get("_home_room_text")
            rows.append(
                (
                    run_id, maktab_id, job["sinf_id"], day, job["smena"],
                    period, job["fan"], teacher, "whole", room_id, room_text,
                    time_slot["boshlanish"], time_slot["tugash"],
                    job["load_id"], job["occurrence"], "har_hafta",
                )
            )
    return rows


def _v2258_json_class_day_signature(state):
    """Tuple kalitli ichki signature'ni JSONB uchun xavfsiz yozuvlarga aylantiradi."""
    return [
        {
            "sinf_id": int(class_id),
            "hafta_kuni": int(day),
            "dars_soni": int(count),
        }
        for (class_id, day), count in sorted(
            _v226_frozen_class_day_signature(state).items()
        )
    ]


def _v2258_save_complete_checkpoint(
    *, maktab_id, user_id, run_id, revision, source_hash, state, jobs,
    context, classes, rooms, shifts, year, qidiruv_nonce, stage,
    improvement=None,
):
    """100% incumbentni optimizatordan OLDIN va har accepted swapdan keyin saqlaydi."""
    if len(state.get("placements") or []) != len(jobs):
        return False
    conn = _db(); cur = conn.cursor()
    try:
        _v1852_tables(cur)
        cur.execute(
            "SELECT pg_advisory_xact_lock(%s)",
            (1900000000 + int(maktab_id),),
        )
        # Advisory lock yangi Boshlash tranzaksiyasi bilan poygani yopadi;
        # qatorning server lease + run jufti esa aynan shu worker hali joriy
        # egasi ekanini isbotlaydi. Qonuniy checkpoint transient try-lock
        # sabab rad etilmaydi, eski worker yangi run ustiga slot yozolmaydi.
        cur.execute(
            """SELECT qidiruv_nonce,jadval_raqami,bosqich,toxtatish_soraldi
                 FROM aqlli_jadval_jarayoni_v2243
                WHERE maktab_id=%s
                FOR UPDATE""",
            (int(maktab_id),),
        )
        lease = cur.fetchone() or {}
        if (
            not lease
            or not _v236_nonce_matches(
                lease.get("qidiruv_nonce"), qidiruv_nonce
            )
            or int(lease.get("jadval_raqami") or 0) != int(run_id)
            or lease.get("bosqich") in {"tayyor", "xato", "toxtatildi"}
        ):
            conn.rollback()
            return False
        # Stop bilan bir paytda kelgan birinchi 100% checkpoint va yakuniy
        # promotion saqlanishi mumkin. Ammo Stopdan keyin yangi revision
        # qabul qilinmaydi.
        if (
            bool(lease.get("toxtatish_soraldi"))
            and int(revision or 0) > 0
            and str(stage or "") != "yakuniy"
        ):
            conn.rollback()
            return False
        if _v1875_source_fingerprint(cur, maktab_id) != source_hash:
            conn.rollback()
            return False
        metrics = _v196_attempt_metrics(state, context)
        quality = max(0, min(100, round(
            100
            - int(metrics.get("oqituvchi_bitta_darsli_kun") or 0) * 0.8
            - int(metrics.get("oqituvchi_ichki_okno") or 0) * 0.25
            - int(metrics.get("ikki_smenali_4soatdan_uzoq") or 0) * 2.0
        )))
        diagnostics = {
            "v2258_checkpoint": True,
            "v2258_revision": int(revision or 0),
            "v2258_stage": str(stage),
            "v2258_metrics": metrics,
            # JSON object kaliti tuple bo'la olmaydi. Ro'yxat ko'rinishi
            # checkpoint saqlashni TypeError bilan yiqilishidan himoya qiladi.
            "v2258_frozen_class_day_counts": _v2258_json_class_day_signature(state),
            "solver_status": "FEASIBLE_VALIDATED_CHECKPOINT",
            "tasdiqlash_mumkin": True,
            "yaxshilanish": dict(improvement or {}),
        }
        settings = {
            "hafta_kunlari": int(context.get("weekdays") or 0),
            "oquv_yili_id": year.get("id"),
            "manba_hash": source_hash,
            "v2258_revision": int(revision or 0),
        }
        cur.execute(
            """INSERT INTO aqlli_jadval_urinishlari_v2(
                   id,maktab_id,holat,yaratgan_user_id,sifat,joylashtirildi,
                   joylashtirilmadi,diagnostika,sozlamalar)
               VALUES(%s,%s,'draft',%s,%s,%s,0,%s::jsonb,%s::jsonb)
               ON CONFLICT(id) DO UPDATE SET
                   sifat=EXCLUDED.sifat,joylashtirildi=EXCLUDED.joylashtirildi,
                   joylashtirilmadi=0,diagnostika=EXCLUDED.diagnostika,
                   sozlamalar=EXCLUDED.sozlamalar""",
            (
                int(run_id), int(maktab_id), int(user_id), int(quality),
                len(jobs), json.dumps(diagnostics, ensure_ascii=False, default=str),
                json.dumps(settings, ensure_ascii=False, default=str),
            ),
        )
        cur.execute(
            "DELETE FROM aqlli_jadval_slotlari_v2 WHERE urinish_id=%s",
            (int(run_id),),
        )
        rows = _v2253_entry_rows_from_state(
            state, run_id, maktab_id, classes, rooms, shifts
        )
        if rows:
            psycopg2.extras.execute_values(
                cur,
                """INSERT INTO aqlli_jadval_slotlari_v2(
                    urinish_id,maktab_id,sinf_id,hafta_kuni,smena,dars_raqami,
                    fan_nomi,oqituvchi_user_id,guruh_kaliti,xona_id,xona_matni,
                    boshlanish_vaqti,tugash_vaqti,yuklama_id,takror_raqami,
                    hafta_turi) VALUES %s""",
                rows, page_size=1000,
            )
        report = _v1875_schedule_integrity_report(cur, maktab_id, run_id)
        if report.get("xatolar") or not report.get("tayyor"):
            conn.rollback()
            return False
        diagnostics["jadval_mosligi"] = report
        slot_keys = (
            "urinish_id", "maktab_id", "sinf_id", "hafta_kuni", "smena",
            "dars_raqami", "fan_nomi", "oqituvchi_user_id", "guruh_kaliti",
            "xona_id", "xona_matni", "boshlanish_vaqti", "tugash_vaqti",
            "yuklama_id", "takror_raqami", "hafta_turi",
        )
        snapshot = [dict(zip(slot_keys, row)) for row in rows]
        cur.execute(
            """INSERT INTO aqlli_jadval_revisionlari_v2258(
                   maktab_id,urinish_id,revision,sifat,bosqich,metrics,slotlar,yaxshilanish)
               VALUES(%s,%s,%s,%s,%s,%s::jsonb,%s::jsonb,%s::jsonb)
               ON CONFLICT(urinish_id,revision) DO UPDATE SET
                   yaratilgan_at=NOW(),sifat=EXCLUDED.sifat,
                   bosqich=EXCLUDED.bosqich,metrics=EXCLUDED.metrics,
                   slotlar=EXCLUDED.slotlar,yaxshilanish=EXCLUDED.yaxshilanish""",
            (
                int(maktab_id), int(run_id), int(revision or 0), int(quality),
                str(stage),
                json.dumps(metrics, ensure_ascii=False, default=str),
                json.dumps(snapshot, ensure_ascii=False, default=str),
                json.dumps(dict(improvement or {}), ensure_ascii=False, default=str),
            ),
        )
        cur.execute(
            "UPDATE aqlli_jadval_urinishlari_v2 SET diagnostika=%s::jsonb WHERE id=%s",
            (json.dumps(diagnostics, ensure_ascii=False, default=str), int(run_id)),
        )
        conn.commit()
        _v2243_progress_write(
            maktab_id, qidiruv_nonce, run_id, revision,
            55 if not revision else min(94, 60 + int(revision)),
            "toliq_saqlandi" if not revision else "revision_saqlandi",
            (
                f"Jadval #{run_id} bazaga saqlandi: {len(jobs)}/{len(jobs)} dars. "
                "Uni hozir ochish mumkin; o‘qituvchilar yaxshilanmoqda."
                if not revision else
                f"Jadval #{run_id}.{revision} saqlandi. Oxirgi tekshirilgan variantni ochish mumkin."
            ),
        )
        return True
    except Exception as checkpoint_error:
        conn.rollback()
        print(f"[JADVAL-CHECKPOINT-V22.58] saqlanmadi: {checkpoint_error}", flush=True)
        return False
    finally:
        cur.close(); conn.close()


def _v237_generate_claimed(
    sorov: V1852Generate, token: str, *, claimed_run_id: int,
    qidiruv_lease: int,
):
    """Canonical claim ajratgan bitta lease/run ichida solverni bajaradi.

    Bu private implementation hech qachon o'zi run yoki client nonce claim
    qilmaydi. Public v2 va v3 endpointlar faqat canonical background starterga
    kiradi, shuning uchun boshqa public yo'l solverni claimsiz ishga tushira
    olmaydi.
    """
    user_id = _jwt_tekshir(token)
    run_id = int(claimed_run_id)
    active_lease = int(qidiruv_lease)
    conn = _db()
    cur = conn.cursor()
    try:
        _v1852_tables(cur)
        if not _v1852_manager(cur, user_id, sorov.maktab_id):
            raise HTTPException(
                status_code=403,
                detail="Jadvalni faqat maktab rahbariyati yaratadi",
            )
        # Claim tranzaksiyasi tugashi yoki qonuniy checkpoint tugashini kutadi;
        # transient pg_try natijasi sabab haqiqiy worker rad etilmaydi.
        cur.execute(
            "SELECT pg_advisory_xact_lock(%s)",
            (1900000000 + int(sorov.maktab_id),),
        )
        cur.execute(
            """SELECT qidiruv_nonce,jadval_raqami,bosqich,toxtatish_soraldi
                 FROM aqlli_jadval_jarayoni_v2243
                WHERE maktab_id=%s
                FOR UPDATE""",
            (int(sorov.maktab_id),),
        )
        claimed = cur.fetchone() or {}
        if (
            not claimed
            or not _v236_nonce_matches(
                claimed.get("qidiruv_nonce"), active_lease
            )
            or int(claimed.get("jadval_raqami") or 0) != run_id
            or claimed.get("bosqich") in {"tayyor", "xato", "toxtatildi"}
        ):
            conn.rollback()
            raise RuntimeError("Generator lease/run claim eskirgan")
        if bool(claimed.get("toxtatish_soraldi")):
            conn.rollback()
            raise RuntimeError("Generator boshlanishidan oldin to‘xtatildi")

        # Lease qatorining FOR UPDATE va advisory locklari faqat yuqoridagi
        # qisqa ownership tekshiruvi uchun. Manbalarni qayta qurish, preflight
        # va job/context tayyorlash vaqtida progress qatorini qulflab turmaymiz:
        # boshqa worker Stop flagini darhol yoza olishi shart.
        conn.commit()

        def abort_if_cancelled(phase):
            if _v2244_cancel_requested(
                sorov.maktab_id, active_lease, run_id
            ):
                # Shu preparation tranzaksiyasidagi yarim o'zgarishlarni ham
                # saqlamaymiz. Background wrapper joriy exact lease/run uchun
                # terminal ``toxtatildi`` progressini yozadi.
                conn.rollback()
                raise RuntimeError(
                    f"Generator {phase} bosqichida to‘xtatildi"
                )

        abort_if_cancelled("tayyorlash boshlanishi")

        _v199_ensure_class_hour_rules(cur, sorov.maktab_id, actor_id=user_id)
        if "_v1876_group_review_report" in globals():
            group_review = _v1876_group_review_report(cur, sorov.maktab_id)
            if not group_review.get("tayyor"):
                raise HTTPException(
                    status_code=409,
                    detail="Avval guruh o‘qituvchilarini tasdiqlang: "
                    + "; ".join(group_review.get("xatolar", [])[:10]),
                )

        abort_if_cancelled("sinf soatlarini tekshirish")

        sync = _v1875_rebuild_schedule_sources(
            cur, sorov.maktab_id, cancel_drafts=False,
            reason="jadval_yaratish_v23",
        )
        if sync.get("xatolar"):
            raise HTTPException(
                status_code=409,
                detail="Jadval manbasi mos emas: "
                + "; ".join(sync["xatolar"][:12]),
            )
        abort_if_cancelled("manbalarni qayta qurish")

        preflight = _v1875_preflight_report(cur, sorov.maktab_id)
        if not preflight.get("tayyor"):
            raise HTTPException(
                status_code=409,
                detail="Jadval tekshiruvdan o‘tmadi: "
                + "; ".join(preflight.get("xatolar", [])[:12]),
            )
        source_hash = str(preflight.get("manba_hash") or "")
        abort_if_cancelled("preflight")

        (
            year, shifts, classes, loads, assignments, group_settings,
            rules_rows, availability_rows, teachers, rooms,
        ) = _v1852_prepare_generation(cur, sorov.maktab_id)
        jobs, warnings = _v1852_build_jobs(
            classes, loads, assignments, group_settings, teachers
        )
        class_hour_rules = _v1866_class_hour_rule_rows(cur, sorov.maktab_id)
        class_hour_jobs, class_hour_warnings = _v1866_build_class_hour_jobs(
            classes, class_hour_rules
        )
        jobs = class_hour_jobs + jobs
        warnings.extend(class_hour_warnings)
        abort_if_cancelled("dars ishlarini tayyorlash")

        rules = _v1852_teacher_rules_map(rules_rows)
        hard, soft, method_hard, method_soft = _v1852_availability_maps(
            availability_rows
        )
        class_day_blocks = _v1856_class_day_rule_map(
            _v1856_class_day_rule_rows(cur, sorov.maktab_id)
        )
        class_hour_counts = _v1852_Counter(
            int(job["teacher_options"][0])
            for job in class_hour_jobs if job.get("teacher_options")
        )
        caps = {}
        for teacher_id, teacher in teachers.items():
            base = teacher.get("haftalik_dars_soati")
            caps[teacher_id] = (
                float(base) + int(class_hour_counts.get(int(teacher_id), 0))
                if base is not None else None
            )
        context = {
            "weekdays": int(year["hafta_kunlari"]),
            "shifts": shifts,
            "classes": classes,
            "teachers": teachers,
            "rules": rules,
            "default_rules": {
                "kunlik_max": 6, "ketma_ket_max": 4, "okno_max": 1,
                "afzal_smena": 0, "eng_erta_dars": 1, "eng_kech_dars": 12,
            },
            "hard": hard,
            "soft": soft,
            "method_hard": method_hard,
            "method_soft": method_soft,
            "teacher_caps": caps,
            "class_day_blocks": class_day_blocks,
            "availability_rows": [dict(row) for row in availability_rows],
            "max_subject_repeat_days": 2,
            "practical_repeat_day_limit": 1,
            "core_period6_day_limit": 2,
            "practical_min_period": 2,
            # Metod kuni ham qizil/BAND kabi avtomatik ochilmaydi. Administrator
            # avval manba qoidalarini aniq tuzatadi, generator esa qat'iy qoladi.
            "allow_fixed_class_hour_method_exception": False,
            # Birinchi exact jadvalning o'zida sinf kunlari teng: masalan
            # 30 soat 6/6/6/6/6, 29 soat 6/6/6/6/5. 2/3/4/5/6 kabi
            # notekis taqsimot keyingi bosqichga qotirib qo'yilmaydi.
            "hard_balance_class_days": True,
            "v207_policy_stage": "strict",
            "v203_emergency_repeat_days": 0,
            "v207_requested_mode": 1,
            "v208_mode_config": dict(_timetable_mode_config()),
        }
        jobs, fractional_warnings = _v220_repair_fractional_pairs(jobs, context)
        warnings.extend(fractional_warnings)
        context["v196_teacher_demand"] = _v196_teacher_demand(jobs)
        context["v196_teacher_shift_demand"] = _v196_teacher_shift_demand(jobs)
        context["v196_class_distribution"] = _v196_class_distribution(jobs, context)
        abort_if_cancelled("solver kontekstini tayyorlash")

        # Har bir yangi bosish yangi asosiy jadval ochadi. Eski #57/#60 draft
        # yangi qidiruvga incumbent bo'lib aralashmaydi; faqat shu run ichidagi
        # accepted yaxshilanishlar #NN.1, #NN.2 ko'rinishida davom etadi.
        initial_state = None
        starting_revision = 0
        # Preparation tugagan paytdagi yakuniy atomar CAS: faqat aynan shu
        # server lease + run hanuz faol, Stop so'ralmagan va nonterminal bo'lsa
        # runtimega o'tishga ruxsat beradi. Preparation paytidagi Stop yoki
        # yangi claim bu UPDATEni 0 qatorga aylantiradi.
        cur.execute(
            """UPDATE aqlli_jadval_jarayoni_v2243
                  SET yaxshilanish=0,foiz=GREATEST(foiz,8),
                      bosqich='tayyorlash',xabar=%s,
                      yangilangan_at=NOW()
                WHERE maktab_id=%s
                  AND qidiruv_nonce IS NOT DISTINCT FROM %s
                  AND jadval_raqami=%s
                  AND toxtatish_soraldi=FALSE
                  AND bosqich NOT IN ('tayyor','xato','toxtatildi')
                RETURNING qidiruv_nonce,jadval_raqami""",
            (
                f"Jadval #{run_id}: {len(jobs)} ta dars tayyorlandi.",
                int(sorov.maktab_id), active_lease, run_id,
            ),
        )
        if not cur.fetchone():
            conn.rollback()
            raise HTTPException(
                status_code=409,
                detail="Bu workerning qidiruv nonce lease'i eskirgan; yangi jadval jarayoni saqlandi",
            )
        conn.commit()
    finally:
        cur.close()
        conn.close()

    try:
        seed = int(source_hash[:16], 16)
    except (TypeError, ValueError):
        seed = int(sorov.maktab_id) * 1_000_003 + len(jobs)
    seed ^= active_lease & ((1 << 63) - 1)

    def write_progress(progress: _V230Progress):
        _v2243_progress_write(
            sorov.maktab_id, active_lease, progress.run_id,
            progress.revision, progress.percent, progress.stage.value,
            progress.message,
        )

    def solve(*, jobs, context, seed, max_seconds, quality_seconds,
              cancel_requested):
        if not (_V216_ORTOOLS_AVAILABLE and callable(_v216_solve_exact)):
            return {
                "complete": False,
                "status": "MODEL_INVALID",
                "message": "OR-Tools o‘rnatilmagan",
            }
        exact_context = dict(context)
        exact_context.update({
            "exact_feasibility_only": True,
            "exact_stop_after_first_solution": True,
            # Birinchi 100% yechim runtime'ga darhol qaytadi va revision-0
            # checkpoint sifatida saqlanadi. Sifat/ustoz oynosi faqat shundan
            # keyingi revision bosqichida yaxshilanadi.
            "exact_quality_after_feasible": False,
            "exact_quality_seconds": 0.0,
            "exact_analyze_method_relaxation": True,
            "exact_analyze_method_on_unknown": False,
            # Tahlil tavsiya berishi mumkin, lekin metod kunini avtomatik
            # yumshatib natija sifatida qabul qilish qat'iyan o'chirilgan.
            "exact_apply_bounded_method_fallback": False,
            "exact_relaxation_seconds": min(10.0, float(max_seconds)),
            "method_exception_primary_only": True,
            "exact_num_workers": max(1, min(4, int(os.getenv("SAMTM_EXACT_NUM_WORKERS", "4")))),
            "exact_cancel_requested": cancel_requested,
        })
        raw = dict(_v216_solve_exact(
            jobs, exact_context, candidate_builder=None,
            state_builder=_v1852_rebuild_schedule_state,
            seed=seed, max_seconds=max_seconds,
        ) or {})
        if not raw.get("complete"):
            return raw
        try:
            state = _v216_exact_complete_state(raw, jobs, exact_context)
        except Exception as error:
            return {
                "complete": False, "status": "MODEL_INVALID",
                "message": str(error),
            }
        return {
            "complete": True,
            "status": raw.get("status"),
            "state": state,
            "diagnostics": dict(raw.get("diagnostics") or {}),
        }

    def validate(rows):
        if not callable(_v218_validate_placements):
            return ["Hard-validator yuklanmagan"]
        return _v218_validate_placements(
            jobs, rows, context, allow_method_exceptions=False
        )

    def persist(*, run_id, revision, state, diagnostics, terminal):
        return _v2258_save_complete_checkpoint(
            maktab_id=sorov.maktab_id,
            user_id=user_id,
            run_id=run_id,
            revision=revision,
            source_hash=source_hash,
            state=state,
            jobs=jobs,
            context=context,
            classes=classes,
            rooms=rooms,
            shifts=shifts,
            year=year,
            qidiruv_nonce=active_lease,
            stage="yakuniy" if terminal else (
                "hard_feasible" if not revision else "teacher_accepted"
            ),
            improvement=(
                dict(diagnostics or {}).get("yaxshilanish")
                or dict(diagnostics or {})
            ) if int(revision or 0) > 0 else {},
        )

    def improve(*, state, context, deadline, cancel_requested, accepted, seed):
        improve_context = dict(context)
        improve_context["v206_deadline"] = deadline
        improve_context["v2244_cancel_requested"] = cancel_requested
        improve_context["v226_frozen_class_day_counts"] = dict(
            _v226_frozen_class_day_signature(state)
        )
        last_scan_progress_at = [0.0]

        def on_teacher_scan(pass_no, target_index, target_total, teacher_id,
                            trial_count, swap_count):
            now = _samtm_time.monotonic()
            if (
                target_index < target_total
                and now - last_scan_progress_at[0] < 0.35
            ):
                return
            last_scan_progress_at[0] = now
            teacher_name = str(
                (teachers.get(int(teacher_id)) or {}).get("full_name")
                or f"O‘qituvchi {teacher_id}"
            )
            # Foiz faqat saqlangan revision bilan oshadi; tekshiruv tartibi
            # xabardagi N/Jami orqali ko'rinadi. Shunda keyingi revision
            # saqlanganda progress foizi orqaga sakramaydi.
            # Keep teacher-scan progress monotonic with accepted revisions.
            # Accepted revision callbacks may already publish 94%, so the next
            # scan must never visually move the same run backwards to 93%.
            percent = min(94, 60 + int(swap_count))
            _v2243_progress_write(
                sorov.maktab_id, active_lease, run_id,
                int(starting_revision) + int(swap_count), percent,
                "oqituvchi_yaxshilanmoqda",
                f"Jadval #{run_id}: {int(target_index)}/{int(target_total)} — "
                f"{teacher_name} oknolari tekshirilmoqda; "
                f"{int(trial_count)} ta xavfsiz variant ko‘rildi "
                f"({int(pass_no)}-aylanish).",
            )

        improve_context["v236_teacher_scan_callback"] = on_teacher_scan

        def on_accept(swap_no, teacher_id, before_score, after_score,
                      checkpoint_state=None):
            if checkpoint_state is not None:
                teacher_name = str(
                    (teachers.get(int(teacher_id)) or {}).get("full_name")
                    or f"O‘qituvchi {teacher_id}"
                )
                before_values = list(before_score or ())
                after_values = list(after_score or ())
                value = lambda values, index: int(values[index]) if len(values) > index else 0
                summary_parts = []
                before_wait = value(before_values, 1)
                after_wait = value(after_values, 1)
                before_gap_count = value(before_values, 2)
                after_gap_count = value(after_values, 2)
                before_internal = value(before_values, 3)
                after_internal = value(after_values, 3)
                before_overstay = value(before_values, 5)
                after_overstay = value(after_values, 5)
                before_one = value(before_values, 6)
                after_one = value(after_values, 6)
                before_days = value(before_values, 9)
                after_days = value(after_values, 9)
                if after_wait < before_wait:
                    summary_parts.append(
                        f"haqiqiy okno {before_wait} daqiqadan {after_wait} daqiqaga qisqardi"
                    )
                if after_gap_count < before_gap_count:
                    summary_parts.append(
                        f"okno soni {before_gap_count} tadan {after_gap_count} taga kamaydi"
                    )
                if after_internal < before_internal:
                    summary_parts.append(
                        f"smena ichki oknosi {before_internal} tadan {after_internal} taga kamaydi"
                    )
                if after_overstay < before_overstay:
                    summary_parts.append(
                        f"ortiqcha qolish {before_overstay} daqiqadan {after_overstay} daqiqaga qisqardi"
                    )
                if after_one < before_one:
                    summary_parts.append(
                        f"1 darsli kun {before_one} tadan {after_one} taga kamaydi"
                    )
                if after_days < before_days:
                    summary_parts.append(
                        f"ish kuni {before_days} tadan {after_days} taga kamaydi"
                    )
                if not summary_parts:
                    summary_parts.append("darslar bir-biriga yaqinroq joylashtirildi")
                revision_details = {
                    "swap": int(swap_no),
                    "teacher_id": int(teacher_id),
                    "oqituvchi": teacher_name,
                    "before_score": before_score,
                    "after_score": after_score,
                    "sinf_kun_soatlari_qotirilgan": True,
                    "xulosa": teacher_name + ": " + "; ".join(summary_parts),
                }
                accepted_result = bool(accepted(
                    checkpoint_state, revision_details
                ))
                if not accepted_result:
                    improve_context["v2253_last_rejection_reason"] = str(
                        revision_details.get("runtime_rejection_reason")
                        or "runtime_rejected"
                    )
                return accepted_result
            return False

        improve_context["v2253_improvement_callback"] = on_accept
        rng = _v1852_random.Random(seed)
        return _v196_optimize_teacher_windows(
            state, improve_context, rng, max_swaps=None
        )

    runtime = _V230ScheduleRuntime(
        solve=solve,
        validate=validate,
        persist=persist,
        improve=improve,
        write_progress=write_progress,
        cancel_requested=lambda: _v2244_cancel_requested(
            sorov.maktab_id, active_lease, run_id
        ),
        policy=_V230RuntimePolicy(
            solve_seconds=max(15.0, min(45.0, _v220_generation_budget_seconds())),
            post_feasible_quality_seconds=2.5,
            # Sun'iy 45/60 soniyalik kesish yo'q. Barcha o'qituvchilar
            # round-robin tekshiriladi; faqat imkon tugashi yoki To'xtatish
            # signali jarayonni yakunlaydi.
            improve_seconds=0.0,
            cancel_poll_seconds=0.25,
            retry_unknown_until_stopped=True,
        ),
    )
    result = runtime.run(
        run_id=run_id,
        jobs=jobs,
        context=context,
        seed=seed,
        initial_state=initial_state,
        starting_revision=starting_revision,
    )
    return {
        "holat": result.stage.value,
        "urinish_id": result.run_id,
        "jadval_raqami": result.run_id,
        "yaxshilanish": result.revision,
        "jami_soat": len(jobs),
        "joylashtirildi": len(jobs) if result.complete else 0,
        "joylashtirilmadi": 0 if result.complete else len(jobs),
        "diagnostika": result.diagnostics,
        "solver_status": "FEASIBLE_VALIDATED" if result.complete else "FAILED",
        "tasdiqlash_mumkin": bool(result.complete),
        "foydalanuvchi_toxtatdi": result.stage == _V230Stage.STOPPED,
        "xabar": result.message,
        "ogohlantirishlar": warnings,
    }


_V2244_BACKGROUND_JOBS = {}
_V2244_BACKGROUND_LOCK = _samtm_threading.Lock()
_V2244_LOCAL_CANCELLED = set()


def _v237_claim_generation(cur, maktab_id: int):
    """Bitta maktab uchun globally unique lease + run ID ni atomar ajratadi."""
    school_id = int(maktab_id)
    cur.execute(
        "SELECT pg_advisory_xact_lock(%s)",
        (1900000000 + school_id,),
    )

    def current_row():
        cur.execute(
            """SELECT bosqich,qidiruv_nonce,jadval_raqami,toxtatish_soraldi,
                      EXTRACT(EPOCH FROM (NOW()-yangilangan_at)) AS yangilanish_yoshi
                 FROM aqlli_jadval_jarayoni_v2243
                WHERE maktab_id=%s""",
            (school_id,),
        )
        return cur.fetchone() or {}

    current = current_row()
    current_expiry = _v236_generation_expiry_stage(
        current.get("bosqich"),
        current.get("toxtatish_soraldi"),
        current.get("yangilanish_yoshi"),
    )
    if current_expiry == "toxtatildi":
        cur.execute(
            """UPDATE aqlli_jadval_jarayoni_v2243
                  SET bosqich='toxtatildi',foiz=100,
                      xabar='To‘xtatish so‘rovi yakunlandi. Saqlangan 100% jadval o‘chirilmagan.',
                      yangilangan_at=NOW()
                WHERE maktab_id=%s
                  AND qidiruv_nonce IS NOT DISTINCT FROM %s
                  AND jadval_raqami=%s
                  AND toxtatish_soraldi=TRUE
                  AND yangilangan_at <= NOW() - INTERVAL '15 seconds'
                  AND bosqich NOT IN ('tayyor','xato','toxtatildi')""",
            (
                school_id, current.get("qidiruv_nonce"),
                int(current.get("jadval_raqami") or 0),
            ),
        )
        # CAS yutdimi yoki boshqa worker qatorni o'zgartirdimi — qaror doim
        # yangi snapshotdan olinadi.
        current = current_row()
    elif current_expiry == "xato":
        cur.execute(
            """UPDATE aqlli_jadval_jarayoni_v2243
                  SET bosqich='xato',foiz=100,
                      xabar='Generator workerining yurak urishi 90 soniyadan oshdi. Yangi qidiruvni boshlash mumkin.',
                      toxtatish_soraldi=FALSE,yangilangan_at=NOW()
                WHERE maktab_id=%s
                  AND qidiruv_nonce IS NOT DISTINCT FROM %s
                  AND jadval_raqami=%s
                  AND toxtatish_soraldi=FALSE
                  AND yangilangan_at <= NOW() - INTERVAL '90 seconds'
                  AND bosqich NOT IN ('tayyor','xato','toxtatildi')""",
            (
                school_id, current.get("qidiruv_nonce"),
                int(current.get("jadval_raqami") or 0),
            ),
        )
        # Oddiy stale worker ham Stop kabi CAS + re-read bilan expire qilinadi;
        # eski snapshot yangi heartbeat ustidan yangi claim ocholmaydi.
        current = current_row()

    current_stage = current.get("bosqich")
    current_age = float(current.get("yangilanish_yoshi") or 0.0)
    if (
        current_stage not in {None, "tayyor", "xato", "toxtatildi"}
        and current_age < 90.0
    ):
        raise HTTPException(
            status_code=409,
            detail="Bu maktab uchun jadval allaqachon yaratilmoqda",
        )
    # Agar stale CAS boshqa heartbeat sabab yutqazgan bo'lsa current_row()
    # nonterminal/fresh qaytaradi va yuqoridagi gate 409 beradi. Har qanday
    # qolgan nonterminal qatorga fail-closed munosabat qilamiz.
    if current_stage not in {None, "tayyor", "xato", "toxtatildi"}:
        raise HTTPException(
            status_code=409,
            detail="Oldingi jadval workeri hali lease egasi",
        )

    cur.execute(
        "SELECT nextval('aqlli_jadval_qidiruv_lease_v237') AS lease"
    )
    lease = int(cur.fetchone()["lease"])
    cur.execute(
        "SELECT nextval(pg_get_serial_sequence('aqlli_jadval_urinishlari_v2','id')) AS id"
    )
    run_id = int(cur.fetchone()["id"])
    cur.execute(
        """INSERT INTO aqlli_jadval_jarayoni_v2243(
               maktab_id,qidiruv_nonce,jadval_raqami,yaxshilanish,foiz,
               bosqich,xabar,toxtatish_soraldi,yangilangan_at)
           VALUES(%s,%s,%s,0,1,'navbatda',%s,FALSE,NOW())
           ON CONFLICT(maktab_id) DO UPDATE SET
               qidiruv_nonce=EXCLUDED.qidiruv_nonce,
               jadval_raqami=EXCLUDED.jadval_raqami,
               yaxshilanish=0,foiz=1,bosqich='navbatda',
               xabar=EXCLUDED.xabar,toxtatish_soraldi=FALSE,
               yangilangan_at=NOW()
           RETURNING qidiruv_nonce,jadval_raqami""",
        (
            school_id, lease, run_id,
            f"Jadval #{run_id} yaratish navbatga olindi; to‘xtatish tugmasi faol.",
        ),
    )
    claimed = cur.fetchone() or {}
    if (
        int(claimed.get("qidiruv_nonce") or 0) != lease
        or int(claimed.get("jadval_raqami") or 0) != run_id
    ):
        raise RuntimeError("Canonical jadval claim yozilmadi")
    return lease, run_id


@app.post("/api/maktab/aqlli_jadval/v2/yaratish")
def v1852_generate(sorov: V1852Generate, token: str):
    """Legacy v2 ham claimsiz solver emas, aynan v3 background yo'lidir."""
    return v2244_start_generation(sorov, token)


@app.post("/api/maktab/aqlli_jadval/v3/boshlash")
def v2244_start_generation(sorov: V1852Generate, token: str):
    """Canonical DB claimdan keyin uzun hisoblashni backgroundda boshlaydi."""
    user_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor()
    try:
        _v1852_tables(cur)
        if not _v1852_manager(cur, user_id, sorov.maktab_id):
            raise HTTPException(status_code=403, detail="Jadvalni faqat maktab rahbariyati yaratadi")
        active_lease, run_id = _v237_claim_generation(
            cur, sorov.maktab_id
        )
        conn.commit()
    finally:
        cur.close(); conn.close()

    job_key = (int(sorov.maktab_id), active_lease, run_id)
    with _V2244_BACKGROUND_LOCK:
        _V2244_LOCAL_CANCELLED.discard(job_key)
    heartbeat_done = _samtm_threading.Event()

    def keep_heartbeat():
        while not heartbeat_done.wait(10.0):
            heartbeat_conn = None
            heartbeat_cur = None
            try:
                heartbeat_conn = _db(); heartbeat_cur = heartbeat_conn.cursor()
                heartbeat_cur.execute(
                    """UPDATE aqlli_jadval_jarayoni_v2243 SET yangilangan_at=NOW()
                       WHERE maktab_id=%s
                         AND qidiruv_nonce IS NOT DISTINCT FROM %s
                         AND jadval_raqami=%s
                         AND toxtatish_soraldi=FALSE
                         AND bosqich NOT IN ('tayyor','xato','toxtatildi')""",
                    (sorov.maktab_id, active_lease, run_id),
                )
                refreshed = int(heartbeat_cur.rowcount or 0) > 0
                heartbeat_conn.commit()
                # 0 qator: Stop so'ralgan, nonce lease boshqa workerga o'tgan
                # yoki jarayon terminal. Bu worker boshqa heartbeat yozmaydi.
                if not refreshed:
                    return
            except Exception:
                if heartbeat_conn is not None: heartbeat_conn.rollback()
            finally:
                if heartbeat_cur is not None: heartbeat_cur.close()
                if heartbeat_conn is not None: heartbeat_conn.close()

    def run_background():
        heartbeat = _samtm_threading.Thread(target=keep_heartbeat, daemon=True)
        heartbeat.start()
        try:
            _v237_generate_claimed(
                sorov, token, claimed_run_id=run_id,
                qidiruv_lease=active_lease,
            )
        except Exception as error:
            print(f"[JADVAL-BACKGROUND-V22.55] yakunlandi: {error}", flush=True)
            # Frontend cheksiz 2–15% da qolib ketmasin. Generator ichida
            # xato bo'lsa ham joriy progress qatori terminal "xato" holatiga
            # o'tadi va foydalanuvchi aniq sababni ko'radi.
            error_conn = None
            error_cur = None
            try:
                error_conn = _db(); error_cur = error_conn.cursor()
                _v1852_tables(error_cur)
                error_cur.execute(
                    """SELECT jadval_raqami,yaxshilanish
                       FROM aqlli_jadval_jarayoni_v2243
                       WHERE maktab_id=%s
                         AND qidiruv_nonce IS NOT DISTINCT FROM %s
                         AND jadval_raqami=%s""",
                    (sorov.maktab_id, active_lease, run_id),
                )
                row = error_cur.fetchone() or {}
                if row.get("jadval_raqami") is not None:
                    was_cancelled = _v2244_cancel_requested(
                        sorov.maktab_id, active_lease, run_id
                    )
                    error_cur.execute(
                        "SELECT 1 AS bor FROM aqlli_jadval_urinishlari_v2 WHERE id=%s AND joylashtirilmadi=0",
                        (int(row["jadval_raqami"]),),
                    )
                    complete_saved = bool(error_cur.fetchone())
                    _v2243_progress_write(
                        sorov.maktab_id, active_lease,
                        int(row["jadval_raqami"]),
                        int(row.get("yaxshilanish") or 0),
                        100, "toxtatildi" if was_cancelled else "xato",
                        (
                            f"To‘xtatildi. Jadval #{row['jadval_raqami']}"
                            f"{'.' + str(row.get('yaxshilanish')) if int(row.get('yaxshilanish') or 0) else ''} "
                            "saqlandi va ochish mumkin."
                            if was_cancelled and complete_saved else
                            "To‘xtatildi. Hali 100% to‘liq jadval saqlanmagan; oldingi jadval o‘chirilmagan."
                            if was_cancelled else
                            f"Jadval hisoblash to'xtadi: {str(error)[:700]}"
                        ),
                    )
            except Exception as progress_error:
                print(f"[JADVAL-BACKGROUND-V22.55] xato holati yozilmadi: {progress_error}", flush=True)
            finally:
                if error_cur is not None: error_cur.close()
                if error_conn is not None: error_conn.close()
        finally:
            heartbeat_done.set()
            with _V2244_BACKGROUND_LOCK:
                # Eski thread ayni key ostida yangi ownership yozuvini o'chira
                # olmaydi. Sequence lease sabab key qayta ishlatilmaydi, lekin
                # identity check cleanupni ham qat'iy qiladi.
                if _V2244_BACKGROUND_JOBS.get(job_key) is _samtm_threading.current_thread():
                    _V2244_BACKGROUND_JOBS.pop(job_key, None)
                    _V2244_LOCAL_CANCELLED.discard(job_key)

    worker = _samtm_threading.Thread(target=run_background, daemon=True)
    with _V2244_BACKGROUND_LOCK:
        _V2244_BACKGROUND_JOBS[job_key] = worker
    try:
        worker.start()
    except Exception as start_error:
        with _V2244_BACKGROUND_LOCK:
            if _V2244_BACKGROUND_JOBS.get(job_key) is worker:
                _V2244_BACKGROUND_JOBS.pop(job_key, None)
                _V2244_LOCAL_CANCELLED.discard(job_key)
        _v2243_progress_write(
            sorov.maktab_id, active_lease, run_id, 0, 100, "xato",
            f"Jadval background workeri boshlanmadi: {str(start_error)[:500]}",
        )
        raise HTTPException(
            status_code=500,
            detail="Jadval background workerini boshlash imkoni bo‘lmadi",
        )
    return {
        "qabul_qilindi": True,
        "qidiruv_nonce": active_lease,
        "jadval_raqami": run_id,
        "xabar": "Jadval backendda mustaqil yaratila boshladi. Oynani yopish hisoblashni to‘xtatmaydi.",
    }


@app.get("/api/maktab/aqlli_jadval/v3/revisionlar")
def v2258_revision_list(token: str, urinish_id: int):
    user_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor()
    try:
        _v1852_tables(cur)
        cur.execute("SELECT maktab_id FROM aqlli_jadval_urinishlari_v2 WHERE id=%s", (urinish_id,))
        run = cur.fetchone()
        if not run or not _v1852_staff(cur, user_id, run["maktab_id"]):
            raise HTTPException(status_code=404, detail="Jadval topilmadi")
        cur.execute(
            """SELECT revision,yaratilgan_at,sifat,bosqich,metrics,yaxshilanish
               FROM aqlli_jadval_revisionlari_v2258
               WHERE urinish_id=%s ORDER BY revision""",
            (urinish_id,),
        )
        return {"urinish_id": urinish_id, "revisionlar": cur.fetchall()}
    finally:
        cur.close(); conn.close()


@app.get("/api/maktab/aqlli_jadval/v2/urinish")
def v1852_run_detail(token: str, urinish_id: int, yaxshilanish: Optional[int] = None):
    user_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor()
    try:
        _v1852_tables(cur)
        cur.execute("SELECT * FROM aqlli_jadval_urinishlari_v2 WHERE id=%s", (urinish_id,))
        run = cur.fetchone()
        if not run:
            raise HTTPException(status_code=404, detail="Jadval urinishi topilmadi")
        if not _v1852_staff(cur, user_id, run["maktab_id"]):
            raise HTTPException(status_code=403, detail="Ruxsat yo'q")
        if yaxshilanish is not None:
            cur.execute(
                """SELECT revision,yaratilgan_at,sifat,bosqich,metrics,slotlar,yaxshilanish
                   FROM aqlli_jadval_revisionlari_v2258
                   WHERE urinish_id=%s AND revision=%s""",
                (urinish_id, int(yaxshilanish)),
            )
            revision_row = cur.fetchone()
            if not revision_row:
                raise HTTPException(status_code=404, detail="Jadval revisioni topilmadi")
            entries = list(revision_row.get("slotlar") or [])
            class_ids = sorted({int(row["sinf_id"]) for row in entries if row.get("sinf_id")})
            teacher_ids = sorted({int(row["oqituvchi_user_id"]) for row in entries if row.get("oqituvchi_user_id")})
            room_ids = sorted({int(row["xona_id"]) for row in entries if row.get("xona_id")})
            class_map = {}
            teacher_map = {}
            room_map = {}
            if class_ids:
                cur.execute("SELECT id,sinf,harf FROM maktab_sinflari WHERE id=ANY(%s)", (class_ids,))
                class_map = {int(row["id"]): row for row in cur.fetchall()}
            if teacher_ids:
                cur.execute("SELECT user_id,full_name FROM users WHERE user_id=ANY(%s)", (teacher_ids,))
                teacher_map = {int(row["user_id"]): row["full_name"] for row in cur.fetchall()}
            if room_ids:
                cur.execute("SELECT id,nomi FROM aqlli_xonalar_v2 WHERE id=ANY(%s)", (room_ids,))
                room_map = {int(row["id"]): row["nomi"] for row in cur.fetchall()}
            for entry in entries:
                cls = class_map.get(int(entry.get("sinf_id") or 0), {})
                entry["sinf"] = cls.get("sinf")
                entry["harf"] = cls.get("harf")
                entry["oqituvchi_ismi"] = teacher_map.get(int(entry.get("oqituvchi_user_id") or 0))
                entry["xona_nomi"] = room_map.get(int(entry.get("xona_id") or 0)) or entry.get("xona_matni")
            run = dict(run)
            run["yaxshilanish"] = int(revision_row["revision"])
            run["yaratilgan_at"] = revision_row["yaratilgan_at"]
            run["sifat"] = revision_row["sifat"]
            run["diagnostika"] = {
                **dict(run.get("diagnostika") or {}),
                "v2258_metrics": revision_row.get("metrics") or {},
                "yaxshilanish": revision_row.get("yaxshilanish") or {},
            }
        else:
            # Asosiy #NN har accepted revisionda eng yaxshi slotlarga
            # almashtiriladi. Uni qayta ochganda "birinchi jadval" deb noto'g'ri
            # ko'rsatmaslik uchun bazadagi haqiqiy final revisionni ham beramiz.
            run = dict(run)
            run_diagnostics = dict(run.get("diagnostika") or {})
            run["yaxshilanish"] = int(
                run_diagnostics.get("v2258_revision") or 0
            )
            cur.execute("""SELECT e.*,s.sinf,s.harf,u.full_name AS oqituvchi_ismi,r.nomi AS xona_nomi
                       FROM aqlli_jadval_slotlari_v2 e
                       JOIN maktab_sinflari s ON s.id=e.sinf_id
                       LEFT JOIN users u ON u.user_id=e.oqituvchi_user_id
                       LEFT JOIN aqlli_xonalar_v2 r ON r.id=e.xona_id
                       WHERE e.urinish_id=%s ORDER BY s.sinf::int,s.harf,e.hafta_kuni,e.smena,e.dars_raqami,e.guruh_kaliti""", (urinish_id,))
            entries = cur.fetchall()
        current_week_type = "toq" if datetime.now().isocalendar().week % 2 else "juft"
        return {"urinish": run, "slotlar": entries, "joriy_hafta_turi": current_week_type}
    finally:
        cur.close(); conn.close()


class V1852Approve(BaseModel):
    urinish_id: int
    majburan: bool = False


@app.post("/api/maktab/aqlli_jadval/v2/tasdiqlash")
def v1852_approve(sorov: V1852Approve, token: str):
    user_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor()
    try:
        _v1852_tables(cur)
        cur.execute("SELECT * FROM aqlli_jadval_urinishlari_v2 WHERE id=%s FOR UPDATE", (sorov.urinish_id,))
        run = cur.fetchone()
        if not run:
            raise HTTPException(status_code=404, detail="Jadval urinishi topilmadi")
        if not _v1852_manager(cur, user_id, run["maktab_id"]):
            raise HTTPException(status_code=403, detail="Faqat rahbariyat tasdiqlaydi")
        if "_v1876_group_review_report" in globals():
            group_review = _v1876_group_review_report(cur, run["maktab_id"])
            if not group_review.get("tayyor"):
                raise HTTPException(
                    status_code=409,
                    detail=(
                        "Guruh va o'qituvchi taqsimoti o'zgargan yoki tasdiqlanmagan. "
                        "Taqsimotni qayta tasdiqlab, yangi draft yarating"
                    ),
                )
        if run["holat"] != "draft":
            raise HTTPException(status_code=409, detail="Faqat draft jadval tasdiqlanadi")

        exact_moslik = _v1875_schedule_integrity_report(
            cur, run["maktab_id"], run["id"]
        )
        if int(run.get("joylashtirilmadi") or 0) > 0 or not exact_moslik.get("tayyor"):
            details = "; ".join(exact_moslik.get("xatolar", [])[:12])
            if int(run.get("joylashtirilmadi") or 0) > 0:
                details = (
                    f"{int(run.get('joylashtirilmadi') or 0)} ta dars joylashtirilmagan"
                    + (f"; {details}" if details else "")
                )
            raise HTTPException(
                status_code=409,
                detail=(
                    "Jadval reja bilan 100% mos emas. Majburan tasdiqlash mumkin emas: "
                    + (details or "sinf, fan yoki o'qituvchi soatlarida farq bor")
                ),
            )

        violations = _v1856_schedule_block_violations(cur, run["maktab_id"], run["id"])
        if violations:
            details = "; ".join(
                f"{row['sinf']}-{row['harf']} · {row['kun_nomi']} ({row['dars_soni']} ta)"
                for row in violations[:8]
            )
            raise HTTPException(
                status_code=409,
                detail=f"Draftda sinf-kun qoidalariga zid darslar bor: {details}. Yangi draft yarating",
            )
        sinf_soati_xatolari = _v1866_class_hour_violations(cur, run["maktab_id"], run["id"])
        if sinf_soati_xatolari:
            details = "; ".join(row["izoh"] for row in sinf_soati_xatolari[:8])
            raise HTTPException(
                status_code=409,
                detail=f"Sinf soati qoidasi bajarilmagan: {details}. 3-bosqichdagi qoidani yoki o‘qituvchi vaqtini tuzatib yangi draft yarating",
            )
        gigiyena_xatolari = _v1874_schedule_hygiene_violations(
            cur, run["maktab_id"], run["id"]
        )
        qattiq_gigiyena_xatolari = [
            row for row in gigiyena_xatolari
            if not _v212_hygiene_is_soft(row.get("sabab"))
        ]
        if qattiq_gigiyena_xatolari:
            details = "; ".join(
                f"{row['sinf']}: {row['sabab']}"
                for row in qattiq_gigiyena_xatolari[:10]
            )
            raise HTTPException(
                status_code=409,
                detail=(
                    "Draft sanitariya-gigiyena qoidalariga zid: "
                    f"{details}. Yangi draft yarating"
                ),
            )
        if int(run["joylashtirilmadi"] or 0) > 0:
            raise HTTPException(
                status_code=409,
                detail="Joylashtirilmagan darslar bor. Jadval to'liq bo'lmaguncha tasdiqlanmaydi",
            )
        cur.execute("UPDATE aqlli_jadval_urinishlari_v2 SET holat='almashtirilgan' WHERE maktab_id=%s AND holat='tasdiqlangan'", (run["maktab_id"],))
        cur.execute("UPDATE aqlli_jadval_urinishlari_v2 SET holat='tasdiqlangan',tasdiqlangan_at=NOW() WHERE id=%s", (run["id"],))
        rebuild = _v1852_rebuild_all_topic_calendars(cur, run["maktab_id"])
        conn.commit()
        return {"holat": "tasdiqlandi", "urinish_id": run["id"],
                "moslik": exact_moslik, **rebuild}
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


@app.get("/api/maktab/aqlli_jadval/v2/faol")
def v1852_active_schedule(token: str, maktab_id: int, sinf_id: Optional[int] = None, oqituvchi_user_id: Optional[int] = None):
    user_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor()
    try:
        _v1852_tables(cur)
        if not _v1852_staff(cur, user_id, maktab_id):
            raise HTTPException(status_code=403, detail="Ruxsat yo'q")
        run = _v1852_active_run(cur, maktab_id)
        if not run:
            return {"urinish": None, "slotlar": []}
        where = ["e.urinish_id=%s"]; args = [run["id"]]
        if sinf_id is not None:
            where.append("e.sinf_id=%s"); args.append(sinf_id)
        if oqituvchi_user_id is not None:
            where.append("e.oqituvchi_user_id=%s"); args.append(oqituvchi_user_id)
        cur.execute("""SELECT e.*,s.sinf,s.harf,u.full_name AS oqituvchi_ismi,
                              COALESCE(r.nomi,e.xona_matni) AS xona_nomi
                       FROM aqlli_jadval_slotlari_v2 e
                       JOIN maktab_sinflari s ON s.id=e.sinf_id
                       LEFT JOIN users u ON u.user_id=e.oqituvchi_user_id
                       LEFT JOIN aqlli_xonalar_v2 r ON r.id=e.xona_id
                       WHERE """ + " AND ".join(where) +
                    " ORDER BY e.hafta_kuni,e.smena,e.dars_raqami,s.sinf::int,s.harf,e.guruh_kaliti", tuple(args))
        current_week_type = "toq" if datetime.now().isocalendar().week % 2 else "juft"
        return {"urinish": run, "slotlar": cur.fetchall(), "joriy_hafta_turi": current_week_type}
    finally:
        cur.close(); conn.close()


class V1852ChangeRequest(BaseModel):
    slot_id: int
    yangi_hafta_kuni: int
    yangi_smena: int
    yangi_dars_raqami: int
    izoh: Optional[str] = None


@app.post("/api/maktab/aqlli_jadval/v2/ozgartirish_sorovi")
def v1852_change_request(sorov: V1852ChangeRequest, token: str):
    user_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor()
    try:
        _v1852_tables(cur)
        cur.execute("""SELECT e.*,r.holat,s.sinf AS sinf_daraja,s.harf AS sinf_harfi
                       FROM aqlli_jadval_slotlari_v2 e
                       JOIN aqlli_jadval_urinishlari_v2 r ON r.id=e.urinish_id
                       JOIN maktab_sinflari s ON s.id=e.sinf_id
                       WHERE e.id=%s""", (sorov.slot_id,))
        slot = cur.fetchone()
        if not slot or slot["holat"] != "tasdiqlangan":
            raise HTTPException(status_code=404, detail="Faol dars topilmadi")
        if str(slot.get("fan_nomi") or "").strip().casefold() == "sinf soati":
            raise HTTPException(
                status_code=409,
                detail="Sinf soati qo‘lda ko‘chirilmaydi. 3-bosqichda kun yoki dars raqamini o‘zgartirib yangi draft yarating",
            )
        if user_id != slot.get("oqituvchi_user_id") and not _v1852_manager(cur, user_id, slot["maktab_id"]):
            raise HTTPException(status_code=403, detail="Faqat shu dars o'qituvchisi yoki rahbariyat so'rov beradi")
        if sorov.yangi_hafta_kuni not in range(1, 7):
            raise HTTPException(status_code=400, detail="Hafta kuni Dushanbadan Shanbagacha bo'lishi kerak")
        rule_map = _v1856_class_day_rule_map(_v1856_class_day_rule_rows(cur, slot["maktab_id"]))
        blocked_reason = _v1856_class_day_block_reason(
            {"id": slot["sinf_id"], "sinf": slot.get("sinf_daraja"), "harf": slot.get("sinf_harfi")},
            sorov.yangi_hafta_kuni, rule_map,
        )
        if blocked_reason:
            raise HTTPException(status_code=400, detail=f"Darsni ko'chirish mumkin emas: {blocked_reason}")
        cur.execute("""INSERT INTO aqlli_jadval_ozgartirish_sorovlari_v2(
            maktab_id,slot_id,soragan_user_id,yangi_hafta_kuni,yangi_smena,yangi_dars_raqami,izoh)
            VALUES(%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
            (slot["maktab_id"], slot["id"], user_id, sorov.yangi_hafta_kuni,
             sorov.yangi_smena, sorov.yangi_dars_raqami, sorov.izoh))
        request_id = cur.fetchone()["id"]
        conn.commit(); return {"holat": "yuborildi", "sorov_id": request_id}
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


def _v1852_topic_permission(cur, user_id: int, maktab_id: int, sinf_id: int, fan: str):
    if not _v1852_teacher_subject_allowed(cur, user_id, maktab_id, sinf_id, fan):
        raise HTTPException(status_code=403, detail="Faqat shu fan o'qituvchisi yoki rahbariyat mavzu rejasini o'zgartiradi")


@app.post("/api/maktab/aqlli_jadval/v2/mavzu_reja/dts_import")
def v1852_import_dts_topics(token: str, maktab_id: int, sinf_id: int, fan: str, chorak: int):
    user_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor()
    try:
        _v1852_tables(cur)
        _v1852_topic_permission(cur, user_id, maktab_id, sinf_id, fan)
        cur.execute("SELECT sinf FROM maktab_sinflari WHERE id=%s AND maktab_id=%s", (sinf_id, maktab_id))
        cls = cur.fetchone()
        if not cls:
            raise HTTPException(status_code=404, detail="Sinf topilmadi")
        cur.execute("""SELECT MIN(topic_code) AS topic_code,
                              COALESCE(NULLIF(mavzu_name,''),NULLIF(kichik_name,''),NULLIF(bolim_name,''),bob_name) AS mavzu
                       FROM dts_tree
                       WHERE grade=%s AND UPPER(TRIM(subject_name))=UPPER(TRIM(%s))
                         AND is_deleted=FALSE AND COALESCE(NULLIF(quarter,''),'1')=%s
                       GROUP BY COALESCE(NULLIF(mavzu_name,''),NULLIF(kichik_name,''),NULLIF(bolim_name,''),bob_name)
                       HAVING COALESCE(NULLIF(mavzu_name,''),NULLIF(kichik_name,''),NULLIF(bolim_name,''),bob_name) IS NOT NULL
                       ORDER BY MIN(topic_code)""", (str(cls["sinf"]), fan, str(chorak)))
        topics = [{"mavzu": row["mavzu"], "soat": 1, "turi": "mavzu", "topic_code": row["topic_code"], "manba": "dts"} for row in cur.fetchall()]
        return {"mavzular": topics, "sinf": cls["sinf"], "fan": fan, "chorak": chorak}
    finally:
        cur.close(); conn.close()


class V1852TopicItem(BaseModel):
    mavzu: str
    soat: int = 1
    turi: str = "mavzu"
    topic_code: Optional[str] = None
    manba: str = "qolda"


class V1852TopicPlan(BaseModel):
    maktab_id: int
    sinf_id: int
    fan_nomi: str
    chorak: int
    mavzular: list[V1852TopicItem]


@app.put("/api/maktab/aqlli_jadval/v2/mavzu_reja")
def v1852_topic_plan_save(sorov: V1852TopicPlan, token: str):
    user_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor()
    try:
        _v1852_tables(cur)
        _v1852_topic_permission(cur, user_id, sorov.maktab_id, sorov.sinf_id, sorov.fan_nomi)
        if sorov.chorak not in (1, 2, 3, 4):
            raise HTTPException(status_code=400, detail="Chorak 1–4 bo'lishi kerak")
        rows = []
        for index, item in enumerate(sorov.mavzular, 1):
            title = re.sub(r"\s+", " ", item.mavzu or "").strip()
            if not title:
                raise HTTPException(status_code=400, detail=f"{index}-mavzu nomi bo'sh")
            if item.turi not in {"mavzu", "nazorat", "xato_tahlil", "mustahkamlash", "masala"}:
                raise HTTPException(status_code=400, detail=f"{index}-mavzu turi noto'g'ri")
            rows.append((sorov.maktab_id, sorov.sinf_id, sorov.fan_nomi.strip(), sorov.chorak,
                         index, title, item.soat, item.turi, item.topic_code, item.manba if item.manba in {"qolda", "dts"} else "qolda"))
        cur.execute("DELETE FROM aqlli_mavzu_rejalari_v2 WHERE maktab_id=%s AND sinf_id=%s AND LOWER(fan_nomi)=LOWER(%s) AND chorak=%s",
                    (sorov.maktab_id, sorov.sinf_id, sorov.fan_nomi, sorov.chorak))
        if rows:
            psycopg2.extras.execute_values(cur, """INSERT INTO aqlli_mavzu_rejalari_v2(
                maktab_id,sinf_id,fan_nomi,chorak,tartib,mavzu,soat,turi,topic_code,manba) VALUES %s""", rows)
        conn.commit(); return {"holat": "saqlandi", "mavzu_soni": len(rows), "jami_soat": sum(int(r[6]) for r in rows)}
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


def _v1852_topic_sequence(plan_rows, available_count, load_row):
    base = []
    for row in plan_rows:
        for _ in range(max(1, int(row.get("soat") or 1))):
            base.append({"mavzu": row["mavzu"], "turi": row["turi"], "topic_code": row.get("topic_code")})
    if available_count <= len(base):
        return base[:available_count], [], base[available_count:]
    extra_count = available_count - len(base)
    controls = min(int(load_row.get("nazorat_soni") or 0), extra_count)
    analyses = controls if bool(load_row.get("nazoratdan_keyin_tahlil")) else 0
    if controls + analyses > extra_count:
        controls = extra_count // (2 if analyses else 1)
        analyses = controls if bool(load_row.get("nazoratdan_keyin_tahlil")) else 0
    reserve = []
    for number in range(1, controls + 1):
        reserve.append({"mavzu": f"Nazorat ishi {number}", "turi": "nazorat", "topic_code": None})
        if analyses:
            reserve.append({"mavzu": f"Nazorat ishi {number} xatolarini tahlil qilish", "turi": "xato_tahlil", "topic_code": None})
    reinforcement = min(int(load_row.get("mustahkamlash_soni") or 0), extra_count - len(reserve))
    for _ in range(reinforcement):
        reserve.append({"mavzu": "Mavzularni mustahkamlash", "turi": "mustahkamlash", "topic_code": None})
    toggle = 0
    while len(reserve) < extra_count:
        reserve.append({
            "mavzu": "Masalalar yechish va amaliy mashqlar" if toggle % 2 == 0 else "Chorak mavzularini mustahkamlash",
            "turi": "masala" if toggle % 2 == 0 else "mustahkamlash", "topic_code": None,
        })
        toggle += 1
    # Nazoratlarni chorak bo'ylab tarqatamiz, hammasini oxiriga tiqmaymiz.
    sequence = list(base)
    if controls and base:
        pairs = []
        cursor = 0
        while cursor < len(reserve):
            pair = [reserve[cursor]]
            cursor += 1
            if cursor < len(reserve) and reserve[cursor]["turi"] == "xato_tahlil":
                pair.append(reserve[cursor]); cursor += 1
            pairs.append(pair)
        result = []
        chunk = max(1, len(base) // (len(pairs) + 1))
        index = 0
        for pair in pairs:
            next_index = min(len(base), index + chunk)
            result.extend(base[index:next_index]); result.extend(pair); index = next_index
        result.extend(base[index:]); sequence = result
        while len(sequence) < available_count and cursor < len(reserve):
            sequence.append(reserve[cursor]); cursor += 1
    else:
        sequence.extend(reserve)
    sequence = sequence[:available_count]
    return sequence, reserve, []


def _v1852_schedule_occurrences(cur, maktab_id: int, sinf_id: int, fan: str, quarter_row, year_row, run_id: int):
    cur.execute("""SELECT MIN(id) AS id,hafta_kuni,smena,dars_raqami,hafta_turi,
                          MIN(oqituvchi_user_id) AS oqituvchi_user_id
                   FROM aqlli_jadval_slotlari_v2
                   WHERE urinish_id=%s AND sinf_id=%s AND LOWER(fan_nomi)=LOWER(%s)
                   GROUP BY hafta_kuni,smena,dars_raqami,hafta_turi
                   ORDER BY hafta_kuni,smena,dars_raqami,hafta_turi""",
                (run_id, sinf_id, fan))
    patterns = cur.fetchall()
    cur.execute("SELECT id,sinf,harf FROM maktab_sinflari WHERE id=%s AND maktab_id=%s", (sinf_id, maktab_id))
    class_row = cur.fetchone() or {"id": sinf_id}
    rule_map = _v1856_class_day_rule_map(_v1856_class_day_rule_rows(cur, maktab_id))
    occurrences = []
    current = quarter_row["boshlanish"]
    while current <= quarter_row["tugash"]:
        if _v1852_is_school_day(cur, maktab_id, current, int(year_row["hafta_kunlari"])):
            for pattern in patterns:
                if current.isoweekday() == int(pattern["hafta_kuni"]):
                    if _v1856_class_day_block_reason(class_row, current.isoweekday(), rule_map):
                        continue
                    phase = str(pattern.get("hafta_turi") or "har_hafta")
                    current_phase = "toq" if int(current.isocalendar().week) % 2 else "juft"
                    if phase not in {"har_hafta", current_phase}:
                        continue
                    occurrences.append({"sana": current, **pattern})
        current += timedelta(days=1)
    occurrences.sort(key=lambda x: (x["sana"], x["smena"], x["dars_raqami"]))
    return occurrences


def _v1852_distribute_topics_one(cur, maktab_id: int, sinf_id: int, fan: str, chorak: int):
    year = _v1852_active_year(cur, maktab_id)
    run = _v1852_active_run(cur, maktab_id)
    if not year or not run:
        return {"taqsimlandi": 0, "ogohlantirishlar": [f"{fan}: faol o'quv yili yoki tasdiqlangan jadval yo'q"]}
    cur.execute("SELECT * FROM aqlli_choraklar_v2 WHERE oquv_yili_id=%s AND chorak=%s", (year["id"], chorak))
    quarter = cur.fetchone()
    if not quarter:
        return {"taqsimlandi": 0, "ogohlantirishlar": [f"{chorak}-chorak sanasi topilmadi"]}
    occurrences = _v1852_schedule_occurrences(cur, maktab_id, sinf_id, fan, quarter, year, run["id"])
    if not occurrences:
        return {"taqsimlandi": 0, "ogohlantirishlar": [f"{fan}: tasdiqlangan jadvalda dars topilmadi"]}
    cur.execute("""SELECT * FROM aqlli_mavzu_rejalari_v2 WHERE maktab_id=%s AND sinf_id=%s
                   AND LOWER(fan_nomi)=LOWER(%s) AND chorak=%s ORDER BY tartib""",
                (maktab_id, sinf_id, fan, chorak))
    plan = cur.fetchall()
    if not plan:
        return {"taqsimlandi": 0, "ogohlantirishlar": [f"{fan}: mavzu rejasi kiritilmagan"]}
    cur.execute("""SELECT * FROM aqlli_sinf_fan_yuklamalari_v2 WHERE maktab_id=%s AND sinf_id=%s
                   AND LOWER(fan_nomi)=LOWER(%s)""", (maktab_id, sinf_id, fan))
    load = cur.fetchone() or {}
    cur.execute("""SELECT * FROM aqlli_mavzu_taqvimi_v2 WHERE maktab_id=%s AND sinf_id=%s
                   AND LOWER(fan_nomi)=LOWER(%s) AND chorak=%s AND qulflangan=TRUE""",
                (maktab_id, sinf_id, fan, chorak))
    locked = cur.fetchall()
    occupied = {(r["sana"], int(r["smena"]), int(r["dars_raqami"])) for r in locked}
    valid_keys = {(o["sana"], int(o["smena"]), int(o["dars_raqami"])) for o in occurrences}
    warnings = []
    for row in locked:
        key = (row["sana"], int(row["smena"]), int(row["dars_raqami"]))
        if key not in valid_keys:
            warnings.append(f"Qulflangan '{row['mavzu']}' {row['sana']} sanada jadvalga mos emas; o'qituvchi tekshirsin")
    free_occurrences = [o for o in occurrences if (o["sana"], int(o["smena"]), int(o["dars_raqami"])) not in occupied]
    sequence, added, overflow = _v1852_topic_sequence(plan, len(free_occurrences), load)
    cur.execute("""DELETE FROM aqlli_mavzu_taqvimi_v2 WHERE maktab_id=%s AND sinf_id=%s
                   AND LOWER(fan_nomi)=LOWER(%s) AND chorak=%s AND qulflangan=FALSE""",
                (maktab_id, sinf_id, fan, chorak))
    rows = []
    for order, (occurrence, topic) in enumerate(zip(free_occurrences, sequence), 1):
        rows.append((maktab_id, run["id"], occurrence["id"], sinf_id, fan, chorak,
                     occurrence["sana"], occurrence["hafta_kuni"], occurrence["smena"],
                     occurrence["dars_raqami"], order, topic["mavzu"], topic["turi"],
                     "avto", False, occurrence.get("oqituvchi_user_id")))
    if rows:
        psycopg2.extras.execute_values(cur, """INSERT INTO aqlli_mavzu_taqvimi_v2(
            maktab_id,urinish_id,slot_id,sinf_id,fan_nomi,chorak,sana,hafta_kuni,smena,
            dars_raqami,tartib,mavzu,turi,manba,qulflangan,oqituvchi_user_id) VALUES %s""", rows)
    if overflow:
        warnings.append(f"{fan}: {len(overflow)} soatlik mavzu chorakdagi darslarga sig'madi")
    return {"taqsimlandi": len(rows), "avto_qoshildi": added, "sigmagan": overflow, "ogohlantirishlar": warnings}


@app.post("/api/maktab/aqlli_jadval/v2/mavzularni_taqsimlash")
def v1852_distribute_topics(token: str, maktab_id: int, sinf_id: int, fan: str, chorak: int):
    user_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor()
    try:
        _v1852_tables(cur)
        _v1852_topic_permission(cur, user_id, maktab_id, sinf_id, fan)
        result = _v1852_distribute_topics_one(cur, maktab_id, sinf_id, fan, chorak)
        conn.commit(); return {"holat": "taqsimlandi", **result}
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


@app.get("/api/maktab/aqlli_jadval/v2/mavzu_reja")
def v1852_topic_plan_get(token: str, maktab_id: int, sinf_id: int, fan: str, chorak: int):
    user_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor()
    try:
        _v1852_tables(cur)
        if not _v1852_staff(cur, user_id, maktab_id):
            raise HTTPException(status_code=403, detail="Ruxsat yo'q")
        cur.execute("""SELECT * FROM aqlli_mavzu_rejalari_v2 WHERE maktab_id=%s AND sinf_id=%s
                       AND LOWER(fan_nomi)=LOWER(%s) AND chorak=%s ORDER BY tartib""",
                    (maktab_id, sinf_id, fan, chorak))
        plan = cur.fetchall()
        cur.execute("""SELECT * FROM aqlli_mavzu_taqvimi_v2 WHERE maktab_id=%s AND sinf_id=%s
                       AND LOWER(fan_nomi)=LOWER(%s) AND chorak=%s ORDER BY sana,smena,dars_raqami""",
                    (maktab_id, sinf_id, fan, chorak))
        calendar = cur.fetchall()
        return {"mavzular": plan, "taqvim": calendar}
    finally:
        cur.close(); conn.close()


class V1852TopicCalendarEdit(BaseModel):
    maktab_id: int
    taqvim_id: int
    mavzu: Optional[str] = None
    sana: Optional[date] = None


@app.patch("/api/maktab/aqlli_jadval/v2/mavzu_taqvimi")
def v1852_topic_calendar_edit(sorov: V1852TopicCalendarEdit, token: str):
    user_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor()
    try:
        _v1852_tables(cur)
        cur.execute("SELECT * FROM aqlli_mavzu_taqvimi_v2 WHERE id=%s AND maktab_id=%s", (sorov.taqvim_id, sorov.maktab_id))
        row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Mavzu taqvimi yozuvi topilmadi")
        _v1852_topic_permission(cur, user_id, sorov.maktab_id, row["sinf_id"], row["fan_nomi"])
        new_date = sorov.sana or row["sana"]
        cur.execute("SELECT id,sinf,harf FROM maktab_sinflari WHERE id=%s AND maktab_id=%s", (row["sinf_id"], sorov.maktab_id))
        class_row = cur.fetchone() or {"id": row["sinf_id"]}
        rule_map = _v1856_class_day_rule_map(_v1856_class_day_rule_rows(cur, sorov.maktab_id))
        blocked_reason = _v1856_class_day_block_reason(class_row, new_date.isoweekday(), rule_map)
        if blocked_reason:
            raise HTTPException(status_code=400, detail=f"Mavzuni bu sanaga ko'chirish mumkin emas: {blocked_reason}")
        year = _v1852_active_year(cur, sorov.maktab_id)
        run = _v1852_active_run(cur, sorov.maktab_id)
        if not year or not run or not _v1852_is_school_day(cur, sorov.maktab_id, new_date, int(year["hafta_kunlari"])):
            raise HTTPException(status_code=400, detail="Tanlangan sana o'qish kuni emas")
        cur.execute("SELECT boshlanish,tugash FROM aqlli_choraklar_v2 WHERE oquv_yili_id=%s AND chorak=%s", (year["id"], row["chorak"]))
        quarter = cur.fetchone()
        if not quarter or not (quarter["boshlanish"] <= new_date <= quarter["tugash"]):
            raise HTTPException(status_code=400, detail="Tanlangan sana shu chorak chegarasida emas")
        cur.execute("""SELECT MIN(id) AS id,hafta_kuni,smena,dars_raqami,MIN(oqituvchi_user_id) AS oqituvchi_user_id
                       FROM aqlli_jadval_slotlari_v2 WHERE urinish_id=%s AND sinf_id=%s
                         AND LOWER(fan_nomi)=LOWER(%s) AND hafta_kuni=%s
                       GROUP BY hafta_kuni,smena,dars_raqami ORDER BY smena,dars_raqami LIMIT 1""",
                    (run["id"], row["sinf_id"], row["fan_nomi"], new_date.isoweekday()))
        slot = cur.fetchone()
        if not slot:
            raise HTTPException(status_code=400, detail="Bu sananing hafta kunida ushbu fan darsi yo'q")
        cur.execute("""SELECT id FROM aqlli_mavzu_taqvimi_v2 WHERE maktab_id=%s AND sinf_id=%s
                       AND LOWER(fan_nomi)=LOWER(%s) AND chorak=%s AND sana=%s
                       AND smena=%s AND dars_raqami=%s AND id<>%s LIMIT 1""",
                    (sorov.maktab_id, row["sinf_id"], row["fan_nomi"], row["chorak"],
                     new_date, slot["smena"], slot["dars_raqami"], row["id"]))
        if cur.fetchone():
            raise HTTPException(status_code=409, detail="Bu sana va dars vaqtida boshqa mavzu allaqachon turibdi")
        title = re.sub(r"\s+", " ", sorov.mavzu or row["mavzu"]).strip()
        cur.execute("""UPDATE aqlli_mavzu_taqvimi_v2 SET sana=%s,hafta_kuni=%s,smena=%s,
                       dars_raqami=%s,slot_id=%s,mavzu=%s,manba='oqituvchi',qulflangan=TRUE,
                       oqituvchi_user_id=%s WHERE id=%s""",
                    (new_date, new_date.isoweekday(), slot["smena"], slot["dars_raqami"],
                     slot["id"], title, slot.get("oqituvchi_user_id"), row["id"]))
        conn.commit(); return {"holat": "yangilandi", "qulflangan": True}
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


@app.post("/api/maktab/aqlli_jadval/v2/mavzu_taqvimi/qulfni_och")
def v1852_topic_unlock(token: str, maktab_id: int, taqvim_id: int):
    user_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor()
    try:
        _v1852_tables(cur)
        cur.execute("SELECT * FROM aqlli_mavzu_taqvimi_v2 WHERE id=%s AND maktab_id=%s", (taqvim_id, maktab_id))
        row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Yozuv topilmadi")
        _v1852_topic_permission(cur, user_id, maktab_id, row["sinf_id"], row["fan_nomi"])
        cur.execute("UPDATE aqlli_mavzu_taqvimi_v2 SET qulflangan=FALSE,manba='avto' WHERE id=%s", (taqvim_id,))
        result = _v1852_distribute_topics_one(cur, maktab_id, row["sinf_id"], row["fan_nomi"], row["chorak"])
        conn.commit(); return {"holat": "avtomatik_taqsimotga_qaytdi", **result}
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()

# ========================= V18.52 END =========================

# ═══════════════════════════════════════════════════════════
# V18.54 — KASBIY RIVOJLANISH KUNI + OMMAVIY O'QITUVCHI VAQTI
# Aniq hafta kuni respublika bo'yicha bitta qilib belgilanmagan; tizim
# fan guruhlari va maktab haftasidan kelib chiqib TAXMINIY tavsiya beradi.
# ═══════════════════════════════════════════════════════════

class V1854BulkTeacherAvailability(BaseModel):
    maktab_id: int
    user_ids: list[int]
    qoidalar: Optional[V1852TeacherRules] = None
    vaqtlar: list[dict] = []
    rejim: str = "almashtirish"  # almashtirish | ustiga_qoshish


def _v1854_validate_time_items(items):
    rows = []
    for item in items:
        day = int(item.get("hafta_kuni") or 0)
        shift = int(item.get("smena") or 0)
        period = int(item.get("dars_raqami") or 0)
        kind = str(item.get("turi") or "")
        hard = bool(item.get("qattiq", True))
        if day not in range(1, 8) or shift not in (0, 1, 2) or period not in range(0, 13) or kind not in _V1852_VAQT_TURLARI:
            raise HTTPException(status_code=400, detail="Ommaviy vaqt sozlamalaridan biri noto'g'ri")
        if kind == "metod_kuni":
            shift, period = 0, 0
            hard = True
        rows.append((day, shift, period, kind, hard, item.get("izoh")))
    return rows


@app.put("/api/maktab/aqlli_jadval/v2/oqituvchi_vaqti_bulk")
def v1854_teacher_availability_bulk(sorov: V1854BulkTeacherAvailability, token: str):
    actor_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor()
    try:
        _v1852_tables(cur)
        if not _v1852_manager(cur, actor_id, sorov.maktab_id):
            raise HTTPException(status_code=403, detail="Ommaviy sozlamani faqat maktab rahbariyati qo'llaydi")
        user_ids = list(dict.fromkeys(int(x) for x in sorov.user_ids if int(x)))
        if not user_ids:
            raise HTTPException(status_code=400, detail="Kamida bitta o'qituvchini tanlang")
        if sorov.rejim not in ("almashtirish", "ustiga_qoshish"):
            raise HTTPException(status_code=400, detail="Ommaviy qo'llash rejimi noto'g'ri")
        cur.execute("SELECT user_id FROM users WHERE maktab_id=%s AND user_id=ANY(%s)", (sorov.maktab_id, user_ids))
        topilgan = {int(r["user_id"]) for r in cur.fetchall()}
        yoq = [x for x in user_ids if x not in topilgan]
        if yoq:
            raise HTTPException(status_code=400, detail=f"Maktabda topilmagan xodim IDlari: {yoq[:10]}")
        rows = _v1854_validate_time_items(sorov.vaqtlar)
        for uid in user_ids:
            if sorov.qoidalar is not None:
                r = sorov.qoidalar
                if r.eng_kech_dars < r.eng_erta_dars:
                    raise HTTPException(status_code=400, detail="Eng kech dars eng erta darsdan oldin bo'lmaydi")
                cur.execute("""INSERT INTO aqlli_oqituvchi_qoidalari_v2(
                    maktab_id,user_id,kunlik_max,ketma_ket_max,okno_max,afzal_smena,eng_erta_dars,eng_kech_dars)
                    VALUES(%s,%s,%s,%s,%s,%s,%s,%s)
                    ON CONFLICT(maktab_id,user_id) DO UPDATE SET
                      kunlik_max=EXCLUDED.kunlik_max,ketma_ket_max=EXCLUDED.ketma_ket_max,
                      okno_max=EXCLUDED.okno_max,afzal_smena=EXCLUDED.afzal_smena,
                      eng_erta_dars=EXCLUDED.eng_erta_dars,eng_kech_dars=EXCLUDED.eng_kech_dars""",
                    (sorov.maktab_id,uid,r.kunlik_max,r.ketma_ket_max,r.okno_max,r.afzal_smena,r.eng_erta_dars,r.eng_kech_dars))
            if sorov.rejim == "almashtirish":
                cur.execute("DELETE FROM aqlli_oqituvchi_vaqti_v2 WHERE maktab_id=%s AND user_id=%s", (sorov.maktab_id,uid))
            for day, shift, period, kind, hard, note in rows:
                cur.execute("""INSERT INTO aqlli_oqituvchi_vaqti_v2(
                    maktab_id,user_id,hafta_kuni,smena,dars_raqami,turi,qattiq,izoh)
                    VALUES(%s,%s,%s,%s,%s,%s,%s,%s)
                    ON CONFLICT(maktab_id,user_id,hafta_kuni,smena,dars_raqami,turi) DO UPDATE SET
                      qattiq=EXCLUDED.qattiq,izoh=EXCLUDED.izoh""",
                    (sorov.maktab_id,uid,day,shift,period,kind,hard,note))
        conn.commit()
        return {"holat":"saqlandi","o'qituvchi_soni":len(user_ids),"vaqt_soni":len(rows)}
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


class V1854MethodSuggestion(BaseModel):
    maktab_id: int
    user_ids: list[int] = []
    saqlash: bool = False
    qattiq: bool = False
    almashtirish: bool = False


def _v1854_subject_group(fanlari):
    text = " ".join(str(fanlari or "").lower().replace("\\n", " ").split())
    if any(k in text for k in ("boshlang", "maktabgacha")):
        return "Boshlang'ich ta'lim"
    if any(k in text for k in ("algebra", "geometri", "matemat", "informat", "fizika", "astronom")):
        return "Aniq fanlar"
    if any(k in text for k in ("biolog", "kimyo", "geograf", "tabiiy", "ekolog")):
        return "Tabiiy fanlar"
    if any(k in text for k in ("ingliz", "rus tili", "nemis", "fransuz", "ona tili", "adabiyot")):
        return "Tillar va adabiyot"
    if any(k in text for k in ("tarix", "huquq", "tarbiya", "iqtisod", "ma'nav")):
        return "Ijtimoiy fanlar"
    if any(k in text for k in ("musiqa", "tasvir", "chizmach", "texnolog", "jismoniy")):
        return "Amaliy fanlar"
    return "Boshqa fanlar"


@app.post("/api/maktab/aqlli_jadval/v2/metod_kun_tavsiya")
def v1854_method_day_suggest(sorov: V1854MethodSuggestion, token: str):
    actor_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor()
    try:
        _v1852_tables(cur)
        if not _v1852_manager(cur, actor_id, sorov.maktab_id):
            raise HTTPException(status_code=403, detail="Kasbiy rivojlanish kunini faqat maktab rahbariyati tavsiya qiladi")
        year = _v1852_active_year(cur, sorov.maktab_id)
        weekdays = int((year or {}).get("hafta_kunlari") or 6)
        ids = list(dict.fromkeys(int(x) for x in sorov.user_ids if int(x)))
        teachers = _v1859_effective_teachers(cur, sorov.maktab_id, ids or None)
        teachers = [row for row in teachers if row.get("dars_beruvchi")]
        groups = {}
        for t in teachers:
            groups.setdefault(_v1854_subject_group(t.get("fanlari")), []).append(t)
        day_counts = {d:0 for d in range(1,weekdays+1)}
        group_days = {}
        for group in sorted(groups, key=lambda g: (-len(groups[g]), g)):
            day = min(day_counts, key=lambda d: (day_counts[d], d))
            group_days[group] = day
            day_counts[day] += len(groups[group])
        existing = set()
        cur.execute("SELECT user_id FROM aqlli_oqituvchi_vaqti_v2 WHERE maktab_id=%s AND turi='metod_kuni'", (sorov.maktab_id,))
        existing = {int(r["user_id"]) for r in cur.fetchall()}
        suggestions = []
        saved = 0
        for group, rows in groups.items():
            day = group_days[group]
            for t in rows:
                uid = int(t["user_id"])
                suggestions.append({
                    "user_id":uid,"full_name":t["full_name"],"fan_guruhi":group,
                    "fanlari":t.get("fanlar_royxati") or [],
                    "sinflari":t.get("sinflar_royxati") or [],
                    "hafta_kuni":day,"kun_nomi":_V1852_HAFTA.get(day,str(day)),
                    "holat":"taxminiy","mavjud":uid in existing,
                })
                if sorov.saqlash:
                    if uid in existing and not sorov.almashtirish:
                        continue
                    if sorov.almashtirish:
                        cur.execute("DELETE FROM aqlli_oqituvchi_vaqti_v2 WHERE maktab_id=%s AND user_id=%s AND turi='metod_kuni'", (sorov.maktab_id,uid))
                    cur.execute("""INSERT INTO aqlli_oqituvchi_vaqti_v2(
                        maktab_id,user_id,hafta_kuni,smena,dars_raqami,turi,qattiq,izoh)
                        VALUES(%s,%s,%s,0,0,'metod_kuni',%s,%s)
                        ON CONFLICT(maktab_id,user_id,hafta_kuni,smena,dars_raqami,turi) DO UPDATE SET
                          qattiq=EXCLUDED.qattiq,izoh=EXCLUDED.izoh""",
                        (sorov.maktab_id,uid,day,sorov.qattiq,
                         "V18.54 taxminiy kasbiy rivojlanish kuni; tuman/tayanch maktab jadvali bilan tasdiqlang"))
                    saved += 1
        if sorov.saqlash:
            conn.commit()
        return {
            "tavsiyalar":suggestions,"saqlandi":saved,
            "ogohlantirish":"Kasbiy rivojlanish kuni rasmiy tizimda mavjud, ammo aniq hafta kuni hudud va tayanch maktab jadvali bilan belgilanadi. Bu faqat taxminiy, tahrirlanadigan tavsiya.",
        }
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


@app.get("/api/maktab/aqlli_jadval/v2/oqituvchi_fan_hisoboti")
def v1859_teacher_subject_report(token: str, maktab_id: int):
    user_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor()
    try:
        _v1852_tables(cur)
        if not _v1852_staff(cur, user_id, maktab_id):
            raise HTTPException(status_code=403, detail="Bu maktab xodimlarini ko'rishga ruxsat yo'q")
        teachers = _v1859_effective_teachers(cur, maktab_id)
        cur.execute("""SELECT user_id,hafta_kuni,qattiq,izoh
                       FROM aqlli_oqituvchi_vaqti_v2
                       WHERE maktab_id=%s AND turi='metod_kuni'
                       ORDER BY user_id,hafta_kuni""", (maktab_id,))
        methods = {}
        for row in cur.fetchall():
            methods.setdefault(int(row["user_id"]), []).append({
                "hafta_kuni": int(row["hafta_kuni"]),
                "kun_nomi": _V1852_HAFTA.get(int(row["hafta_kuni"]), str(row["hafta_kuni"])),
                "qattiq": bool(row["qattiq"]),
                "izoh": row.get("izoh"),
            })
        for teacher in teachers:
            teacher["metod_kunlari"] = methods.get(int(teacher["user_id"]), [])
        return {
            "xodimlar": teachers,
            "jami": len(teachers),
            "fanli": sum(1 for row in teachers if row.get("fanlar_royxati")),
            "fansiz": sum(1 for row in teachers if not row.get("fanlar_royxati")),
            "izoh": "Fanlar XODIMLAR va DARS_BIRIKMALARI shablonlaridagi import ma'lumotlaridan birlashtirildi.",
        }
    finally:
        cur.close(); conn.close()


# ========================= V18.54 END =========================



# ═══════════════════════════════════════════════════════════
# V18.56 — MOSLASHUVCHAN SINF-KUN QOIDALARI
# Rahbariyat xohlagan parallel yoki aniq sinfga xohlagan hafta kunini bir tugma bilan bloklaydi.
# ═══════════════════════════════════════════════════════════

# ═══════════════════════════════════════════════════════════
# V18.57 — MAKTAB BOSH SAHIFASI UCHUN XAVFSIZ READ API
# Bir qo'shimcha jadval buzilsa ham butun bosh sahifa yiqilmaydi.
# GET so'rovlari dublikat tozalash yoki migratsiya kabi yozuv amali bajarmaydi.
# ═══════════════════════════════════════════════════════════

def _v1857_table_exists(cur, table_name: str) -> bool:
    cur.execute("SELECT to_regclass(%s) AS table_name", (f"public.{table_name}",))
    return bool((cur.fetchone() or {}).get("table_name"))


def _v1857_columns(cur, table_name: str) -> set[str]:
    cur.execute(
        "SELECT column_name FROM information_schema.columns "
        "WHERE table_schema='public' AND table_name=%s",
        (table_name,),
    )
    return {str(row["column_name"]) for row in cur.fetchall()}


def _v1857_has_columns(cur, table_name: str, required: set[str]) -> bool:
    return _v1857_table_exists(cur, table_name) and required.issubset(_v1857_columns(cur, table_name))


def _v1857_normal_name(value) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).casefold()


@app.get("/api/maktab/dashboard_xavfsiz")
def v1857_safe_school_dashboard(token: str, maktab_id: int):
    user_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor()
    warnings = []
    try:
        if not _maktab_boshqaruvchi_mi(cur, user_id, maktab_id):
            raise HTTPException(status_code=403, detail="Faqat maktab rahbariyati yoki admin ko'ra oladi")

        maktab_cols = _v1857_columns(cur, "maktablar")
        payment_expr = "pulli" if "pulli" in maktab_cols else "FALSE AS pulli"
        fee_expr = "oylik_tolov" if "oylik_tolov" in maktab_cols else "NULL::INTEGER AS oylik_tolov"
        shift_count_expr = (
            "COALESCE(smena_soni,1) AS smena_soni"
            if "smena_soni" in maktab_cols else "1 AS smena_soni"
        )
        alphabet_expr = (
            "COALESCE(alifbo_turi,'latin_xalqaro') AS alifbo_turi"
            if "alifbo_turi" in maktab_cols
            else "'latin_xalqaro'::TEXT AS alifbo_turi"
        )
        cur.execute(
            f"SELECT nomi,{payment_expr},{fee_expr},{shift_count_expr},{alphabet_expr} "
            "FROM maktablar WHERE id=%s",
            (maktab_id,),
        )
        maktab = cur.fetchone()
        if not maktab:
            raise HTTPException(status_code=404, detail="Maktab topilmadi")

        sinf_cols = _v1857_columns(cur, "maktab_sinflari")
        psixolog_bor = "psixolog_user_id" in sinf_cols
        psixolog_select = "COALESCE(p.full_name,'') AS psixolog_ismi" if psixolog_bor else "''::TEXT AS psixolog_ismi"
        psixolog_join = "LEFT JOIN users p ON p.user_id=s.psixolog_user_id" if psixolog_bor else ""
        smena_select = "COALESCE(s.smena,1) AS smena" if "smena" in sinf_cols else "1 AS smena"
        cur.execute(f"""
            SELECT s.id,s.sinf,s.harf,{smena_select},
                   COALESCE(u.full_name,'') AS rahbar_ismi,
                   {psixolog_select},
                   (SELECT COUNT(*) FROM maktab_sinf_azolari a WHERE a.sinf_id=s.id) AS oquvchi_soni
            FROM maktab_sinflari s
            LEFT JOIN users u ON u.user_id=s.rahbar_user_id
            {psixolog_join}
            WHERE s.maktab_id=%s
            ORDER BY NULLIF(REGEXP_REPLACE(COALESCE(s.sinf,''),'[^0-9]','','g'),'')::int NULLS LAST,
                     s.harf,s.id
        """, (maktab_id,))
        sinflar = cur.fetchall()
        class_map = {int(row["id"]): row for row in sinflar}

        for row in sinflar:
            row["tolagan_soni"] = 0
            row["bugun_kelgan_soni"] = 0
            row["bugun_belgilangan_soni"] = 0
            row["davomat_kun_7"] = 0
            row["ortacha_bilim"] = None

        # Davomat optional modul.
        if _v1857_has_columns(cur, "davomat", {"sinf_id", "user_id", "sana", "holat"}):
            cur.execute("""
                SELECT d.sinf_id,
                       COUNT(DISTINCT d.user_id) FILTER (WHERE d.holat='keldi') AS kelgan,
                       COUNT(DISTINCT d.user_id) AS belgilangan,
                       COUNT(DISTINCT d.sana) FILTER (WHERE d.sana>=CURRENT_DATE-INTERVAL '7 days') AS kun_7
                FROM davomat d
                JOIN maktab_sinflari s ON s.id=d.sinf_id
                WHERE s.maktab_id=%s AND (d.sana=CURRENT_DATE OR d.sana>=CURRENT_DATE-INTERVAL '7 days')
                GROUP BY d.sinf_id
            """, (maktab_id,))
            for item in cur.fetchall():
                row = class_map.get(int(item["sinf_id"]))
                if row:
                    row["bugun_kelgan_soni"] = int(item.get("kelgan") or 0)
                    row["bugun_belgilangan_soni"] = int(item.get("belgilangan") or 0)
                    row["davomat_kun_7"] = int(item.get("kun_7") or 0)
        else:
            warnings.append("Davomat jadvali hali tayyor emas; davomat ko'rsatkichlari vaqtincha 0 ko'rsatildi.")

        # To'lov faqat pulli maktabda va kerakli ustunlar mavjud bo'lsa hisoblanadi.
        if bool(maktab.get("pulli")):
            required = {"user_id", "maktab_id", "oy", "tolangan_summa"}
            if _v1857_has_columns(cur, "tolovlar", required):
                current_month = datetime.now().strftime("%Y-%m")
                required_sum = int(maktab.get("oylik_tolov") or 0)
                cur.execute("""
                    SELECT a.sinf_id,COUNT(DISTINCT t.user_id) AS tolagan
                    FROM maktab_sinf_azolari a
                    JOIN maktab_sinflari s ON s.id=a.sinf_id
                    JOIN tolovlar t ON t.user_id=a.user_id AND t.maktab_id=s.maktab_id
                    WHERE s.maktab_id=%s AND t.oy=%s AND t.tolangan_summa>=%s
                    GROUP BY a.sinf_id
                """, (maktab_id, current_month, required_sum))
                for item in cur.fetchall():
                    row = class_map.get(int(item["sinf_id"]))
                    if row:
                        row["tolagan_soni"] = int(item.get("tolagan") or 0)
            else:
                warnings.append("To'lov moduli eski sxemada; bosh sahifa to'lovsiz xavfsiz yuklandi.")

        # Bilim ko'rsatkichi optional.
        if _v1857_has_columns(cur, "learned_topics", {"user_id", "score"}):
            cur.execute("""
                SELECT a.sinf_id,ROUND(AVG(lt.score)) AS ortacha_bilim
                FROM maktab_sinf_azolari a
                JOIN maktab_sinflari s ON s.id=a.sinf_id
                LEFT JOIN learned_topics lt ON lt.user_id=a.user_id
                WHERE s.maktab_id=%s
                GROUP BY a.sinf_id
            """, (maktab_id,))
            for item in cur.fetchall():
                row = class_map.get(int(item["sinf_id"]))
                if row:
                    row["ortacha_bilim"] = item.get("ortacha_bilim")
        else:
            warnings.append("Bilim analitikasi jadvali topilmadi; sinf reytingi vaqtincha yashirildi.")

        muammoli_oquvchilar = []
        if _v1857_has_columns(cur, "davomat", {"sinf_id", "user_id", "sana", "holat"}):
            cur.execute("""
                SELECT u.user_id,u.full_name,s.sinf,s.harf,
                       COUNT(*) FILTER (WHERE d.holat='kelmadi') AS songi_hafta_kelmagan
                FROM maktab_sinf_azolari a
                JOIN users u ON u.user_id=a.user_id
                JOIN maktab_sinflari s ON s.id=a.sinf_id
                LEFT JOIN davomat d ON d.user_id=a.user_id AND d.sinf_id=a.sinf_id
                                    AND d.sana>=CURRENT_DATE-INTERVAL '7 days'
                WHERE s.maktab_id=%s
                GROUP BY u.user_id,u.full_name,s.sinf,s.harf
                HAVING COUNT(*) FILTER (WHERE d.holat='kelmadi')>=2
                ORDER BY songi_hafta_kelmagan DESC LIMIT 20
            """, (maktab_id,))
            muammoli_oquvchilar = cur.fetchall()

        xodim_bugun = {"jami": 0, "keldi": 0}
        cur.execute("SELECT COUNT(*) AS jami FROM users WHERE maktab_id=%s AND lavozim IS NOT NULL", (maktab_id,))
        xodim_bugun["jami"] = int((cur.fetchone() or {}).get("jami") or 0)
        if _v1857_has_columns(cur, "xodim_davomati", {"maktab_id", "user_id", "sana", "holat"}):
            cur.execute("""
                SELECT COUNT(DISTINCT user_id) AS keldi FROM xodim_davomati
                WHERE maktab_id=%s AND sana=CURRENT_DATE AND holat='keldi'
            """, (maktab_id,))
            xodim_bugun["keldi"] = int((cur.fetchone() or {}).get("keldi") or 0)

        jami_oquvchi = sum(int(row.get("oquvchi_soni") or 0) for row in sinflar)
        jami_tolagan = sum(int(row.get("tolagan_soni") or 0) for row in sinflar)
        jami_kelgan = sum(int(row.get("bugun_kelgan_soni") or 0) for row in sinflar)
        jami_belgilangan = sum(int(row.get("bugun_belgilangan_soni") or 0) for row in sinflar)
        sinflar_belgilamagan = sum(
            1 for row in sinflar
            if int(row.get("oquvchi_soni") or 0)>0 and int(row.get("bugun_belgilangan_soni") or 0)==0
        )
        rated = [row for row in sinflar if row.get("ortacha_bilim") is not None]
        rated.sort(key=lambda row: float(row.get("ortacha_bilim") or 0), reverse=True)

        return {
            "maktab_nomi": maktab["nomi"],
            "smena_soni": int(maktab.get("smena_soni") or 1),
            "alifbo_turi": maktab.get("alifbo_turi") or "latin_xalqaro",
            "pulli": bool(maktab.get("pulli")),
            "oylik_tolov": maktab.get("oylik_tolov"),
            "sinflar": sinflar,
            "tolov_xulosasi": ({
                "jami_oquvchi": jami_oquvchi,
                "tolagan": jami_tolagan,
                "qarzdor": max(0, jami_oquvchi-jami_tolagan),
            } if bool(maktab.get("pulli")) else None),
            "bugungi_davomat": {
                "jami_oquvchi": jami_oquvchi,
                "kelgan": jami_kelgan,
                "belgilangan": jami_belgilangan,
                "sinflar_belgilamagan": sinflar_belgilamagan,
            },
            "xodim_bugun": xodim_bugun,
            "muammoli_oquvchilar": muammoli_oquvchilar,
            "eng_yaxshi_sinf": rated[0] if rated else None,
            "etibor_kerak_sinf": rated[-1] if len(rated)>1 else None,
            "diagnostika_ogohlantirishlari": warnings,
        }
    finally:
        cur.close(); conn.close()


@app.get("/api/maktab/yuklama_xulosasi_xavfsiz")
def v1857_safe_workload(token: str, maktab_id: int):
    actor_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor()
    warnings = []
    try:
        if not _maktab_boshqaruvchi_mi(cur, actor_id, maktab_id):
            raise HTTPException(status_code=403, detail="Faqat maktab rahbariyati yoki admin ko'ra oladi")

        user_cols = _v1857_columns(cur, "users")
        load_expr = "haftalik_dars_soati" if "haftalik_dars_soati" in user_cols else "NULL::INTEGER AS haftalik_dars_soati"
        cur.execute(f"""
            SELECT user_id,full_name,lavozim,fanlari,{load_expr}
            FROM users WHERE maktab_id=%s AND lavozim IS NOT NULL ORDER BY full_name,user_id
        """, (maktab_id,))
        people = cur.fetchall()
        ids = [int(row["user_id"]) for row in people]
        counts = {uid: 0 for uid in ids}
        assigned_counts = {uid: 0 for uid in ids}
        if _v1857_has_columns(
            cur, "maktab_dars_birikmalari", {"maktab_id", "user_id", "haftalik_soat"}
        ):
            cur.execute("""SELECT user_id,COALESCE(SUM(haftalik_soat),0) AS soat
                           FROM maktab_dars_birikmalari
                           WHERE maktab_id=%s GROUP BY user_id""", (maktab_id,))
            for assignment in cur.fetchall():
                uid = int(assignment["user_id"])
                if uid in assigned_counts:
                    assigned_counts[uid] = float(assignment.get("soat") or 0)
        psixolog_class_counts = {}
        class_hour_counts = {}
        if _v1857_has_columns(cur, "aqlli_sinf_soati_qoidalari_v2", {"maktab_id", "sinf_id", "faol"}) and            _v1857_has_columns(cur, "maktab_sinflari", {"id", "rahbar_user_id"}):
            cur.execute("""SELECT s.rahbar_user_id AS user_id,COUNT(*) AS son
                           FROM aqlli_sinf_soati_qoidalari_v2 q
                           JOIN maktab_sinflari s ON s.id=q.sinf_id
                           WHERE q.maktab_id=%s AND q.faol=TRUE AND s.rahbar_user_id IS NOT NULL
                           GROUP BY s.rahbar_user_id""", (maktab_id,))
            class_hour_counts = {
                int(row["user_id"]): int(row.get("son") or 0) for row in cur.fetchall()
            }
        if "psixolog_user_id" in _v1857_columns(cur, "maktab_sinflari"):
            cur.execute("""
                SELECT psixolog_user_id AS user_id,COUNT(*) AS son
                FROM maktab_sinflari
                WHERE maktab_id=%s AND psixolog_user_id IS NOT NULL
                GROUP BY psixolog_user_id
            """, (maktab_id,))
            psixolog_class_counts = {
                int(row["user_id"]): int(row.get("son") or 0) for row in cur.fetchall()
            }

        smart_ok = _v1857_has_columns(
            cur, "aqlli_jadval_urinishlari_v2", {"id", "maktab_id", "holat"}
        ) and _v1857_has_columns(
            cur, "aqlli_jadval_slotlari_v2", {"urinish_id", "oqituvchi_user_id"}
        )
        active_run_id = None
        if smart_ok:
            cur.execute("""SELECT id FROM aqlli_jadval_urinishlari_v2
                           WHERE maktab_id=%s AND holat='tasdiqlangan'
                           ORDER BY id DESC LIMIT 1""", (maktab_id,))
            active = cur.fetchone(); active_run_id = active["id"] if active else None
        if active_run_id:
            cur.execute("""SELECT oqituvchi_user_id,COUNT(*) AS son
                           FROM aqlli_jadval_slotlari_v2
                           WHERE urinish_id=%s AND oqituvchi_user_id IS NOT NULL
                           GROUP BY oqituvchi_user_id""", (active_run_id,))
            for row in cur.fetchall():
                counts[int(row["oqituvchi_user_id"])] = int(row.get("son") or 0)
        elif _v1857_has_columns(cur, "dars_jadvali", {"id", "oqituvchi_user_id"}):
            cur.execute("""SELECT oqituvchi_user_id,COUNT(id) AS son FROM dars_jadvali
                           WHERE oqituvchi_user_id IS NOT NULL GROUP BY oqituvchi_user_id""")
            for row in cur.fetchall():
                uid = int(row["oqituvchi_user_id"])
                if uid in counts:
                    counts[uid] = int(row.get("son") or 0)
        else:
            warnings.append("Jadval jadvali topilmadi; amaldagi yuklama 0 deb ko'rsatildi.")

        # Import dublikatlari ekranda qayta takrorlanmasin; bu READ-ONLY birlashtirish.
        grouped = {}
        for person in people:
            uid = int(person["user_id"])
            if uid < 0:
                key = ("import", _v1857_normal_name(person.get("full_name")), str(person.get("lavozim") or ""))
            else:
                key = ("user", uid)
            current = grouped.get(key)
            if current is None:
                current = dict(person)
                current["birlashtirilgan_idlar"] = []
                current["jadvaldagi_soat"] = 0
                current["biriktirilgan_soat"] = 0
                current["psixolog_sinf_soni"] = 0
                current["sinf_soati_soni"] = 0
                grouped[key] = current
            current["birlashtirilgan_idlar"].append(uid)
            current["jadvaldagi_soat"] += int(counts.get(uid, 0))
            current["biriktirilgan_soat"] += float(assigned_counts.get(uid, 0))
            current["psixolog_sinf_soni"] += int(psixolog_class_counts.get(uid, 0))
            current["sinf_soati_soni"] += int(class_hour_counts.get(uid, 0))
            if current.get("haftalik_dars_soati") is None and person.get("haftalik_dars_soati") is not None:
                current["haftalik_dars_soati"] = person.get("haftalik_dars_soati")
            if not current.get("fanlari") and person.get("fanlari"):
                current["fanlari"] = person.get("fanlari")
            if uid < int(current.get("user_id") or uid):
                current["user_id"] = uid

        result = list(grouped.values())
        result.sort(key=lambda row: _v1857_normal_name(row.get("full_name")))
        for row in result:
            base_plan = row.get("haftalik_dars_soati")
            extra = int(row.get("sinf_soati_soni") or 0)
            plan = (
                round(float(base_plan) + extra, 1)
                if base_plan is not None else (extra or None)
            )
            row["haftalik_reja_jami"] = plan
            actual = float(
                row.get("jadvaldagi_soat") if active_run_id
                else row.get("biriktirilgan_soat") or 0
            )
            actual = round(actual, 1)
            row["amaldagi_soat"] = actual
            row["hisob_manbasi"] = "tasdiqlangan_jadval" if active_run_id else "oqituvchi_yuklamasi"
            difference = None if plan is None else round(float(plan) - actual, 1)
            row["farq"] = difference
            row["holat"] = (
                "kiritilmagan" if plan is None else
                "ortiqcha" if difference < -1e-9 else
                "toliq" if abs(difference) <= 1e-9 else "yetishmaydi"
            )
        return {"xodimlar": result, "diagnostika_ogohlantirishlari": warnings}
    finally:
        cur.close(); conn.close()


@app.get("/api/maktab/aqlli_holatlar_xavfsiz")
def v1857_safe_smart_cases(token: str, maktab_id: int, holat: str = "ochiq"):
    actor_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor()
    try:
        if not _maktab_boshqaruvchi_mi(cur, actor_id, maktab_id):
            raise HTTPException(status_code=403, detail="Faqat maktab rahbariyati yoki admin ko'ra oladi")
        required = {"id", "maktab_id", "oquvchi_user_id", "daraja", "sarlavha", "holat", "yangilangan_at"}
        if not _v1857_has_columns(cur, "aqlli_holatlar", required):
            return {"holatlar": [], "diagnostika_ogohlantirishlari": ["Aqlli holatlar moduli hali tayyor emas."]}
        cur.execute("""
            SELECT h.*,COALESCE(u.full_name,'Noma''lum o''quvchi') AS full_name
            FROM aqlli_holatlar h LEFT JOIN users u ON u.user_id=h.oquvchi_user_id
            WHERE h.maktab_id=%s AND h.holat=%s
            ORDER BY h.daraja DESC,h.yangilangan_at DESC LIMIT 100
        """, (maktab_id, holat))
        return {"holatlar": cur.fetchall(), "diagnostika_ogohlantirishlari": []}
    finally:
        cur.close(); conn.close()

# ========================= V18.57 END =========================

# ========================= V18.60 END =========================

# ═══════════════════════════════════════════════════════════
# V18.66 — METOD KUNINI OMMAVIY TOZALASH + QAT'IY SINF SOATI
# ═══════════════════════════════════════════════════════════

class V1866MethodDayClear(BaseModel):
    maktab_id: int
    user_ids: list[int]
    hafta_kuni: Optional[int] = None


@app.post("/api/maktab/aqlli_jadval/v2/metod_kunlarini_tozalash")
def v1866_method_days_clear(sorov: V1866MethodDayClear, token: str):
    actor_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor()
    try:
        _v1852_tables(cur)
        if not _v1852_manager(cur, actor_id, sorov.maktab_id):
            raise HTTPException(status_code=403, detail="Metod kunlarini faqat maktab rahbariyati ommaviy tozalaydi")
        ids = list(dict.fromkeys(int(x) for x in sorov.user_ids if int(x)))
        if not ids:
            raise HTTPException(status_code=400, detail="Kamida bitta o‘qituvchini tanlang")
        if sorov.hafta_kuni is not None and int(sorov.hafta_kuni) not in range(1, 7):
            raise HTTPException(status_code=400, detail="Hafta kuni Dushanba–Shanba oralig‘ida bo‘lishi kerak")
        cur.execute("SELECT user_id FROM users WHERE maktab_id=%s AND user_id=ANY(%s)", (sorov.maktab_id, ids))
        found = {int(row["user_id"]) for row in cur.fetchall()}
        missing = [uid for uid in ids if uid not in found]
        if missing:
            raise HTTPException(status_code=400, detail=f"Maktabda topilmagan xodim IDlari: {missing[:10]}")
        where = ["maktab_id=%s", "user_id=ANY(%s)", "turi='metod_kuni'"]
        args = [sorov.maktab_id, ids]
        if sorov.hafta_kuni is not None:
            where.append("hafta_kuni=%s")
            args.append(int(sorov.hafta_kuni))
        cur.execute("DELETE FROM aqlli_oqituvchi_vaqti_v2 WHERE " + " AND ".join(where), tuple(args))
        deleted = int(cur.rowcount or 0)
        conn.commit()
        return {
            "holat": "tozalandi", "o'qituvchi_soni": len(ids),
            "o'chirilgan_metod_kuni": deleted, "hafta_kuni": sorov.hafta_kuni,
        }
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


class V1866ClassHourBulk(BaseModel):
    maktab_id: int
    qamrov: str = "parallel"  # parallel | aniq
    sinf_darajalari: list[int] = []
    sinf_idlar: list[int] = []
    hafta_kuni: int
    dars_raqami: int
    fan_nomi: str = "KELAJAK SOATI"
    haftalik_soat: int = 1


def _v1866_class_hour_rule_rows(cur, maktab_id: int):
    cur.execute("""SELECT q.*,s.sinf,s.harf,COALESCE(s.smena,1) AS smena,
                          s.rahbar_user_id,COALESCE(u.full_name,'') AS rahbar_ismi
                   FROM aqlli_sinf_soati_qoidalari_v2 q
                   JOIN maktab_sinflari s ON s.id=q.sinf_id
                   LEFT JOIN users u ON u.user_id=s.rahbar_user_id
                   WHERE q.maktab_id=%s AND q.faol=TRUE
                   ORDER BY s.sinf::int,s.harf""", (maktab_id,))
    return cur.fetchall()


def _v199_ensure_class_hour_rules(cur, maktab_id: int, class_ids=None, actor_id=None):
    """Har sinf uchun yagona, takrorlanmaydigan SINF SOATI qoidasini yaratadi.

    Sinf soati fan yuklamasiga yozilmaydi: jadval generatorida u alohida bitta
    sessiya sifatida hisoblanadi. Shu sabab o'quv reja jami ham, o'qituvchi
    yuklamasi ham bir soatga oshadi, lekin fan soatlari ikki marta sanalmaydi.
    Mavjud qo'lda sozlangan kun/dars o'zgartirilmaydi.
    """
    _v1852_tables(cur)
    year = _v1852_active_year(cur, maktab_id)
    weekdays = max(1, min(6, int((year or {}).get("hafta_kunlari") or 6)))
    default_day = min(5, weekdays)  # odatda Juma
    args = [int(maktab_id)]
    where = ["maktab_id=%s"]
    if class_ids is not None:
        ids = sorted({int(value) for value in class_ids if value is not None})
        if not ids:
            return {"yaratildi": 0, "jami": 0, "hafta_kuni": default_day, "dars_raqami": 1}
        where.append("id=ANY(%s)")
        args.append(ids)
    cur.execute(
        "SELECT id FROM maktab_sinflari WHERE " + " AND ".join(where) + " ORDER BY id",
        tuple(args),
    )
    target_ids = [int(row["id"]) for row in cur.fetchall()]
    created = 0
    for class_id in target_ids:
        cur.execute(
            """INSERT INTO aqlli_sinf_soati_qoidalari_v2(
                   maktab_id,sinf_id,hafta_kuni,dars_raqami,faol,
                   yaratgan_user_id,yangilangan_at)
               VALUES(%s,%s,%s,1,TRUE,%s,NOW())
               ON CONFLICT(maktab_id,sinf_id) DO UPDATE SET
                 faol=TRUE,
                 fan_nomi=COALESCE(NULLIF(TRIM(aqlli_sinf_soati_qoidalari_v2.fan_nomi),''),'KELAJAK SOATI'),
                 haftalik_soat=GREATEST(1,COALESCE(aqlli_sinf_soati_qoidalari_v2.haftalik_soat,1)),
                 yangilangan_at=NOW()""",
            (maktab_id, class_id, default_day, actor_id),
        )
        created += int(cur.rowcount or 0)
    return {
        "yaratildi": created,
        "jami": len(target_ids),
        "hafta_kuni": default_day,
        "dars_raqami": 1,
    }


def _v1866_target_classes(cur, sorov: V1866ClassHourBulk):
    qamrov = str(sorov.qamrov or "parallel").strip().lower()
    if qamrov == "parallel":
        grades = sorted(set(int(x) for x in sorov.sinf_darajalari))
        if not grades or any(g not in range(1, 12) for g in grades):
            raise HTTPException(status_code=400, detail="Kamida bitta 1–11-sinf parallelini tanlang")
        cur.execute("""SELECT s.id,s.sinf,s.harf,COALESCE(s.smena,1) AS smena,s.rahbar_user_id,
                              COALESCE(u.full_name,'') AS rahbar_ismi
                       FROM maktab_sinflari s LEFT JOIN users u ON u.user_id=s.rahbar_user_id
                       WHERE s.maktab_id=%s
                         AND NULLIF(REGEXP_REPLACE(COALESCE(s.sinf,''),'[^0-9]','','g'),'')::int=ANY(%s)
                       ORDER BY s.sinf::int,s.harf""", (sorov.maktab_id, grades))
    elif qamrov == "aniq":
        ids = sorted(set(int(x) for x in sorov.sinf_idlar if int(x)))
        if not ids:
            raise HTTPException(status_code=400, detail="Kamida bitta aniq sinfni tanlang")
        cur.execute("""SELECT s.id,s.sinf,s.harf,COALESCE(s.smena,1) AS smena,s.rahbar_user_id,
                              COALESCE(u.full_name,'') AS rahbar_ismi
                       FROM maktab_sinflari s LEFT JOIN users u ON u.user_id=s.rahbar_user_id
                       WHERE s.maktab_id=%s AND s.id=ANY(%s)
                       ORDER BY s.sinf::int,s.harf""", (sorov.maktab_id, ids))
    else:
        raise HTTPException(status_code=400, detail="Qamrov parallel yoki aniq bo‘lishi kerak")
    rows = cur.fetchall()
    if not rows:
        raise HTTPException(status_code=404, detail="Tanlangan sinflar topilmadi")
    return rows


def _v1866_shift_interval_map(shift_rows):
    """Smena+dars raqamini haqiqiy boshlanish/tugash daqiqasiga aylantiradi."""
    result = {}
    for raw_row in shift_rows:
        row = dict(raw_row)
        shift = int(row.get("smena") or 0)
        for slot in _v1852_shift_slots(row):
            try:
                start_hour, start_minute = str(slot["boshlanish"]).split(":", 1)
                end_hour, end_minute = str(slot["tugash"]).split(":", 1)
                result[(shift, int(slot["dars_raqami"]))] = (
                    int(start_hour) * 60 + int(start_minute),
                    int(end_hour) * 60 + int(end_minute),
                )
            except (KeyError, TypeError, ValueError):
                continue
    return result


def _v1866_intervals_overlap(first, second):
    return bool(
        first and second
        and int(first[0]) < int(second[1])
        and int(second[0]) < int(first[1])
    )


def _v220_max_nonoverlapping_interval_count(intervals):
    """Haqiqiy vaqtda sig'adigan eng ko'p dars sonini qaytaradi."""
    normalized = set()
    for interval in intervals or ():
        try:
            start, end = int(interval[0]), int(interval[1])
        except (IndexError, TypeError, ValueError):
            continue
        if end > start:
            normalized.add((start, end))

    count = 0
    last_end = None
    # Eng erta tugaydigan intervalni tanlash maksimal sonni beradi. Ikkinchi
    # kalit teng tugashlarda natijani deterministik saqlaydi.
    for start, end in sorted(normalized, key=lambda item: (item[1], item[0])):
        if last_end is None or start >= last_end:
            count += 1
            last_end = end
    return count


@app.put("/api/maktab/aqlli_jadval/v2/sinf_soati_bulk")
def v1866_class_hour_bulk_save(sorov: V1866ClassHourBulk, token: str):
    actor_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor()
    try:
        _v1852_tables(cur)
        if not _v1852_manager(cur, actor_id, sorov.maktab_id):
            raise HTTPException(status_code=403, detail="Sinf soatini faqat maktab rahbariyati belgilaydi")
        year = _v1852_active_year(cur, sorov.maktab_id)
        weekdays = int((year or {}).get("hafta_kunlari") or 6)
        if int(sorov.hafta_kuni) not in range(1, weekdays + 1):
            raise HTTPException(status_code=400, detail=f"O‘qish haftasi {weekdays} kun. Mos hafta kunini tanlang")
        if int(sorov.dars_raqami) not in range(1, 13):
            raise HTTPException(status_code=400, detail="Dars raqami 1–12 oralig‘ida bo‘lishi kerak")
        fan_nomi = _v192_clean_subject(sorov.fan_nomi) or "KELAJAK SOATI"
        haftalik_soat = int(sorov.haftalik_soat)
        if haftalik_soat not in range(1, 6):
            raise HTTPException(status_code=400, detail="Kelajak soati haftasiga 1–5 soat bo‘lishi mumkin")
        classes = _v1866_target_classes(cur, sorov)
        cur.execute("SELECT * FROM aqlli_smena_sozlamalari_v2 WHERE maktab_id=%s", (sorov.maktab_id,))
        shift_rows = [dict(row) for row in cur.fetchall()]
        shift_limits = {
            int(row["smena"]): int(row.get("dars_soni") or 0)
            for row in shift_rows
        }
        shift_intervals = _v1866_shift_interval_map(shift_rows)
        class_day_map = _v1856_class_day_rule_map(_v1856_class_day_rule_rows(cur, sorov.maktab_id))
        cur.execute(
            "SELECT * FROM aqlli_oqituvchi_vaqti_v2 WHERE maktab_id=%s",
            (sorov.maktab_id,),
        )
        hard, _soft, _method_hard, _method_soft = (
            _v1852_availability_maps(cur.fetchall())
        )
        target_ids = {int(row["id"]) for row in classes}
        cur.execute("""SELECT q.sinf_id,q.hafta_kuni,q.dars_raqami,COALESCE(s.smena,1) AS smena,s.rahbar_user_id
                       FROM aqlli_sinf_soati_qoidalari_v2 q
                       JOIN maktab_sinflari s ON s.id=q.sinf_id
                       WHERE q.maktab_id=%s AND q.faol=TRUE""", (sorov.maktab_id,))
        # Bir rahbarning ikki smenadagi dars raqami boshqacha bo'lsa ham,
        # haqiqiy soatlari ustma-ust kelishi mumkin. Shu sabab faqat
        # (smena,dars_raqami) tengligini emas, vaqt oralig'ini tekshiramiz.
        occupied = []
        for row in cur.fetchall():
            if int(row["sinf_id"]) in target_ids or row.get("rahbar_user_id") is None:
                continue
            occupied.append({
                "rahbar_user_id": int(row["rahbar_user_id"]),
                "hafta_kuni": int(row["hafta_kuni"]),
                "smena": int(row["smena"]),
                "dars_raqami": int(row["dars_raqami"]),
                "sinf_id": int(row["sinf_id"]),
            })
        saved = 0
        skipped = []
        for cls in classes:
            label = f"{cls['sinf']}-{cls['harf']}"
            leader = cls.get("rahbar_user_id")
            shift = int(cls.get("smena") or 1)
            if int(sorov.dars_raqami) > int(shift_limits.get(shift) or 0):
                skipped.append({"sinf": label, "sabab": f"{shift}-smenada {sorov.dars_raqami}-dars mavjud emas"})
                continue
            blocked = _v1856_class_day_block_reason(cls, int(sorov.hafta_kuni), class_day_map)
            if blocked:
                skipped.append({"sinf": label, "sabab": blocked})
                continue
            if leader is not None:
                if _v1852_blocked(
                    hard,
                    int(leader),
                    int(sorov.hafta_kuni),
                    shift,
                    int(sorov.dars_raqami),
                ):
                    skipped.append({
                        "sinf": label,
                        "sabab": (
                            "sinf rahbarining tanlangan vaqti qizil/BAND; "
                            "KELAJAK SOATI ham qizil vaqtga qo'yilmaydi"
                        ),
                    })
                    continue
                target_interval = shift_intervals.get(
                    (shift, int(sorov.dars_raqami))
                )
                conflict = next((
                    row for row in occupied
                    if int(row["rahbar_user_id"]) == int(leader)
                    and int(row["hafta_kuni"]) == int(sorov.hafta_kuni)
                    and _v1866_intervals_overlap(
                        target_interval,
                        shift_intervals.get((
                            int(row["smena"]), int(row["dars_raqami"])
                        )),
                    )
                ), None)
                if conflict:
                    skipped.append({
                        "sinf": label,
                        "sabab": (
                            "sinf rahbarining boshqa qat'iy KELAJAK SOATI "
                            "haqiqiy vaqt bilan ustma-ust"
                        ),
                    })
                    continue
            cur.execute("""INSERT INTO aqlli_sinf_soati_qoidalari_v2(
                            maktab_id,sinf_id,hafta_kuni,dars_raqami,faol,yaratgan_user_id,
                            fan_nomi,haftalik_soat,yangilangan_at)
                           VALUES(%s,%s,%s,%s,TRUE,%s,%s,%s,NOW())
                           ON CONFLICT(maktab_id,sinf_id) DO UPDATE SET
                             hafta_kuni=EXCLUDED.hafta_kuni,dars_raqami=EXCLUDED.dars_raqami,
                             faol=TRUE,yaratgan_user_id=EXCLUDED.yaratgan_user_id,
                             fan_nomi=EXCLUDED.fan_nomi,haftalik_soat=EXCLUDED.haftalik_soat,
                             yangilangan_at=NOW()""",
                        (sorov.maktab_id, cls["id"], sorov.hafta_kuni, sorov.dars_raqami,
                         actor_id, fan_nomi, haftalik_soat))
            if leader is not None:
                occupied.append({
                    "rahbar_user_id": int(leader),
                    "hafta_kuni": int(sorov.hafta_kuni),
                    "smena": shift,
                    "dars_raqami": int(sorov.dars_raqami),
                    "sinf_id": int(cls["id"]),
                })
            saved += 1
        if saved == 0:
            detail = "; ".join(f"{row['sinf']}: {row['sabab']}" for row in skipped[:8]) or "mos sinf topilmadi"
            raise HTTPException(status_code=400, detail=f"Sinf soati saqlanmadi — {detail}")
        conn.commit()
        return {"holat": "saqlandi", "saqlandi": saved, "otkazib_yuborildi": skipped}
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


@app.delete("/api/maktab/aqlli_jadval/v2/sinf_soati")
def v1866_class_hour_delete(token: str, maktab_id: int, sinf_id: int):
    actor_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor()
    try:
        _v1852_tables(cur)
        if not _v1852_manager(cur, actor_id, maktab_id):
            raise HTTPException(status_code=403, detail="Sinf soatini faqat maktab rahbariyati o‘chiradi")
        raise HTTPException(
            status_code=409,
            detail="KELAJAK SOATI majburiy: uni o‘chirib bo‘lmaydi, faqat kuni, vaqti va nomini tahrirlang",
        )
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


def _v1866_build_class_hour_jobs(classes, rules):
    jobs = []
    warnings = []
    for row in rules:
        class_id = int(row["sinf_id"])
        cls = classes.get(class_id)
        if not cls:
            continue
        leader = cls.get("rahbar_user_id")
        if leader is None:
            warnings.append(
                f"{cls['sinf']}-{cls['harf']} KELAJAK SOATI qat’iy katakka "
                "o‘qituvchisiz qo‘yildi; sinf rahbarini keyin biriktiring"
            )
        weekly = max(1, min(5, int(row.get("haftalik_soat") or 1)))
        for occurrence in range(1, weekly + 1):
            jobs.append({
                "job_id": f"sinf-soati:{class_id}:{occurrence}", "load_id": None,
                "sinf_id": class_id,
                "fan": _v192_clean_subject(row.get("fan_nomi")) or "KELAJAK SOATI",
                "occurrence": occurrence,
                "smena": int(cls.get("smena") or 1), "daily_max": 1,
                "consecutive_allowed": False, "preferred_last": int(row["dars_raqami"]),
                "weight": 1, "room_id": None, "groups": [],
                "teacher_options": [int(leader)] if leader is not None else [],
                "difficulty": 100000 - occurrence,
                "fixed_day": int(row["hafta_kuni"]) if occurrence == 1 else None,
                "fixed_period": int(row["dars_raqami"]) if occurrence == 1 else None,
                "is_class_hour": True,
            })
    return jobs, warnings


def _v1866_class_hour_violations(cur, maktab_id: int, run_id: int):
    rules = _v1866_class_hour_rule_rows(cur, maktab_id)
    errors = []
    for row in rules:
        cur.execute("""SELECT 1 FROM aqlli_jadval_slotlari_v2
                       WHERE urinish_id=%s AND sinf_id=%s AND hafta_kuni=%s AND smena=%s
                         AND dars_raqami=%s
                         AND (%s IS NULL OR oqituvchi_user_id=%s) LIMIT 1""",
                    (run_id, row["sinf_id"], row["hafta_kuni"], row["smena"],
                     row["dars_raqami"], row.get("rahbar_user_id"), row.get("rahbar_user_id")))
        if not cur.fetchone():
            errors.append({
                "sinf_id": row["sinf_id"],
                "izoh": f"{row['sinf']}-{row['harf']} · {_V1852_HAFTA.get(int(row['hafta_kuni']), row['hafta_kuni'])} · {row['dars_raqami']}-dars",
            })
    return errors

# ========================= V18.66 END =========================



# ═══════════════════════════════════════════════════════════
# V18.68 — O'QITUVCHI VAQT MATRITSASI
# ═══════════════════════════════════════════════════════════

class V1868TeacherMatrixItem(BaseModel):
    user_id: int
    qoidalar: Optional[V1852TeacherRules] = None
    vaqtlar: list[dict] = []


class V1868TeacherMatrixSave(BaseModel):
    maktab_id: int
    oqituvchilar: list[V1868TeacherMatrixItem]


@app.put("/api/maktab/aqlli_jadval/v2/oqituvchi_vaqt_matritsasi")
def v1868_teacher_time_matrix_save(sorov: V1868TeacherMatrixSave, token: str):
    actor_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor()
    try:
        _v1852_tables(cur)
        items_by_user = {}
        for item in sorov.oqituvchilar:
            uid = int(item.user_id)
            if uid:
                items_by_user[uid] = item
        if not items_by_user:
            raise HTTPException(status_code=400, detail="Saqlash uchun o'qituvchi tanlanmagan")
        if len(items_by_user) > 300:
            raise HTTPException(status_code=400, detail="Bir so'rovda 300 tadan ortiq xodim saqlanmaydi")

        manager = _v1852_manager(cur, actor_id, sorov.maktab_id)
        if not manager and (len(items_by_user) != 1 or actor_id not in items_by_user):
            raise HTTPException(
                status_code=403,
                detail="O'qituvchi faqat o'z vaqtini, rahbariyat esa barcha o'qituvchilarni o'zgartira oladi",
            )

        user_ids = list(items_by_user)
        cur.execute(
            "SELECT user_id FROM users WHERE maktab_id=%s AND user_id=ANY(%s)",
            (sorov.maktab_id, user_ids),
        )
        found = {int(row["user_id"]) for row in cur.fetchall()}
        missing = [uid for uid in user_ids if uid not in found]
        if missing:
            raise HTTPException(status_code=400, detail=f"Maktabda topilmagan xodimlar: {missing[:10]}")

        inserted_count = 0
        for uid, item in items_by_user.items():
            if item.qoidalar is not None:
                rules = item.qoidalar
                if rules.eng_kech_dars < rules.eng_erta_dars:
                    raise HTTPException(
                        status_code=400,
                        detail=f"{uid}: eng kech dars eng erta darsdan oldin bo'lmaydi",
                    )
                cur.execute(
                    """INSERT INTO aqlli_oqituvchi_qoidalari_v2(
                        maktab_id,user_id,kunlik_max,ketma_ket_max,okno_max,
                        afzal_smena,eng_erta_dars,eng_kech_dars)
                       VALUES(%s,%s,%s,%s,%s,%s,%s,%s)
                       ON CONFLICT(maktab_id,user_id) DO UPDATE SET
                         kunlik_max=EXCLUDED.kunlik_max,
                         ketma_ket_max=EXCLUDED.ketma_ket_max,
                         okno_max=EXCLUDED.okno_max,
                         afzal_smena=EXCLUDED.afzal_smena,
                         eng_erta_dars=EXCLUDED.eng_erta_dars,
                         eng_kech_dars=EXCLUDED.eng_kech_dars""",
                    (
                        sorov.maktab_id, uid, rules.kunlik_max, rules.ketma_ket_max,
                        rules.okno_max, rules.afzal_smena, rules.eng_erta_dars,
                        rules.eng_kech_dars,
                    ),
                )

            validated = _v1854_validate_time_items(item.vaqtlar)
            unique_rows = {}
            for day, row_shift, period, kind, hard, note in validated:
                unique_rows[(day, row_shift, period, kind)] = (
                    day, row_shift, period, kind, hard, note
                )

            cur.execute(
                "DELETE FROM aqlli_oqituvchi_vaqti_v2 WHERE maktab_id=%s AND user_id=%s",
                (sorov.maktab_id, uid),
            )
            rows = [
                (sorov.maktab_id, uid, day, row_shift, period, kind, hard, note)
                for day, row_shift, period, kind, hard, note in unique_rows.values()
            ]
            if rows:
                psycopg2.extras.execute_values(
                    cur,
                    """INSERT INTO aqlli_oqituvchi_vaqti_v2(
                        maktab_id,user_id,hafta_kuni,smena,dars_raqami,turi,qattiq,izoh)
                       VALUES %s""",
                    rows,
                )
                inserted_count += len(rows)

        conn.commit()
        return {
            "holat": "saqlandi",
            "oqituvchi_soni": len(items_by_user),
            "vaqt_soni": inserted_count,
        }
    except Exception:
        conn.rollback()
        raise
    finally:
        cur.close(); conn.close()

# ========================= V18.68 END =========================

# ═══════════════════════════════════════════════════════════
# V18.71 — AVTO METOD KUNI FAQAT ADMIN YOQQANDA ISHLAYDI
# Maktab fan → kun qoidalarini bir marta saqlaydi. Avto O'CHIQ bo'lsa
# tizim o'zi hech qanday metod kuni topmaydi va eski avto belgilarni olib tashlaydi.
# ═══════════════════════════════════════════════════════════

_V1871_AUTO_METHOD_PREFIX = "V18.71 AUTO METOD:"
_V1871_OLD_AUTO_PREFIX = "V18.54 taxminiy kasbiy rivojlanish kuni"


def _v1871_method_subject_key(value):
    text = str(value or "").casefold()
    for old in ("‘", "’", "`", "ʼ", "ʻ"):
        text = text.replace(old, "'")
    return " ".join(text.split())


def _v1871_auto_method_tables(cur):
    cur.execute("""CREATE TABLE IF NOT EXISTS aqlli_metod_avto_sozlamalari_v2(
        maktab_id INTEGER PRIMARY KEY REFERENCES maktablar(id) ON DELETE CASCADE,
        yoqilgan BOOLEAN NOT NULL DEFAULT FALSE,
        yangilagan_user_id BIGINT,
        yangilangan_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
    )""")
    cur.execute("""CREATE TABLE IF NOT EXISTS aqlli_metod_fan_qoidalari_v2(
        id BIGSERIAL PRIMARY KEY,
        maktab_id INTEGER NOT NULL REFERENCES maktablar(id) ON DELETE CASCADE,
        fan_nomi TEXT NOT NULL,
        fan_kaliti TEXT NOT NULL,
        hafta_kuni INTEGER NOT NULL CHECK(hafta_kuni BETWEEN 1 AND 7),
        qattiq BOOLEAN NOT NULL DEFAULT TRUE,
        tartib INTEGER NOT NULL DEFAULT 0,
        UNIQUE(maktab_id, fan_kaliti)
    )""")
    cur.execute(
        "CREATE INDEX IF NOT EXISTS idx_metod_fan_qoidalari_v2 "
        "ON aqlli_metod_fan_qoidalari_v2(maktab_id,tartib,id)"
    )


def _v1871_auto_method_where():
    return "(COALESCE(izoh,'') LIKE %s OR COALESCE(izoh,'') LIKE %s)"


def _v1871_method_rules(cur, maktab_id: int):
    cur.execute(
        """SELECT id,fan_nomi,fan_kaliti,hafta_kuni,qattiq,tartib
           FROM aqlli_metod_fan_qoidalari_v2
           WHERE maktab_id=%s ORDER BY tartib,id""",
        (maktab_id,),
    )
    return cur.fetchall()


def _v1871_method_report(cur, maktab_id: int, rules=None):
    rules = list(rules if rules is not None else _v1871_method_rules(cur, maktab_id))
    teachers = [
        row for row in _v1859_effective_teachers(cur, maktab_id)
        if row.get("dars_beruvchi")
    ]
    result = []
    for rule in rules:
        matched = []
        for teacher in teachers:
            subjects = teacher.get("fanlar_royxati") or []
            if any(
                _v1871_method_subject_key(subject) == rule["fan_kaliti"]
                for subject in subjects
            ):
                matched.append({
                    "user_id": int(teacher["user_id"]),
                    "full_name": teacher["full_name"],
                })
        result.append({
            "id": rule.get("id"),
            "fan_nomi": rule["fan_nomi"],
            "hafta_kuni": int(rule["hafta_kuni"]),
            "kun_nomi": _V1852_HAFTA.get(
                int(rule["hafta_kuni"]), str(rule["hafta_kuni"])
            ),
            "qattiq": bool(rule["qattiq"]),
            "oqituvchi_soni": len(matched),
            "oqituvchilar": matched[:50],
        })
    return result


class V1871AutoMethodRule(BaseModel):
    fan_nomi: str
    hafta_kuni: int
    qattiq: bool = True


class V1871AutoMethodSettings(BaseModel):
    maktab_id: int
    yoqilgan: bool = False
    qoidalar: list[V1871AutoMethodRule] = []


@app.on_event("startup")
def _v1871_startup_disable_old_auto_methods():
    """Eski V18.54 taxminiy avtomatik metod kunlarini bir marta tozalaydi.

    Qo'lda belgilangan metod kunlariga tegmaydi. V18.71 dagi yangi avto
    qoidalar faqat administrator YOQ qilgandan keyin yaratiladi.
    """
    try:
        conn = _db()
        cur = conn.cursor()
        _v1871_auto_method_tables(cur)
        cur.execute(
            """DELETE FROM aqlli_oqituvchi_vaqti_v2
               WHERE turi='metod_kuni' AND COALESCE(izoh,'') LIKE %s""",
            (_V1871_OLD_AUTO_PREFIX + "%",),
        )
        conn.commit()
        cur.close()
        conn.close()
    except Exception as exc:
        print(f"[V18.71 eski avto metod tozalash] {exc}")


@app.get("/api/maktab/aqlli_jadval/v2/metod_avto_sozlama")
def v1871_auto_method_get(token: str, maktab_id: int):
    actor_id = _jwt_tekshir(token)
    conn = _db()
    cur = conn.cursor()
    try:
        _v1852_tables(cur)
        _v1871_auto_method_tables(cur)
        if not _v1852_staff(cur, actor_id, maktab_id):
            raise HTTPException(
                status_code=403,
                detail="Bu maktab metod kuni sozlamasini ko'rishga ruxsat yo'q",
            )
        cur.execute(
            "SELECT yoqilgan FROM aqlli_metod_avto_sozlamalari_v2 "
            "WHERE maktab_id=%s",
            (maktab_id,),
        )
        row = cur.fetchone()
        rules = _v1871_method_rules(cur, maktab_id)
        return {
            "yoqilgan": bool((row or {}).get("yoqilgan")),
            "qoidalar": [
                {
                    "id": rule.get("id"),
                    "fan_nomi": rule["fan_nomi"],
                    "hafta_kuni": int(rule["hafta_kuni"]),
                    "qattiq": bool(rule["qattiq"]),
                }
                for rule in rules
            ],
            "hisobot": _v1871_method_report(cur, maktab_id, rules),
            "izoh": (
                "Avto O'CHIQ bo'lsa tizim metod kunini o'zi aniqlamaydi. "
                "Avto YOQ bo'lsa faqat shu fan→kun qoidalari ishlaydi."
            ),
        }
    finally:
        cur.close()
        conn.close()


@app.put("/api/maktab/aqlli_jadval/v2/metod_avto_sozlama")
def v1871_auto_method_save(sorov: V1871AutoMethodSettings, token: str):
    actor_id = _jwt_tekshir(token)
    conn = _db()
    cur = conn.cursor()
    try:
        _v1852_tables(cur)
        _v1871_auto_method_tables(cur)
        if not _v1852_manager(cur, actor_id, sorov.maktab_id):
            raise HTTPException(
                status_code=403,
                detail="Avto metod kunini faqat maktab rahbariyati boshqaradi",
            )

        year = _v1852_active_year(cur, sorov.maktab_id)
        weekdays = int((year or {}).get("hafta_kunlari") or 6)

        # Faqat maktabdagi aniq fan nomlari. TARIX/TARBIYA kabi yaqin
        # so'zlar hech qachon bir-biriga aralashmaydi.
        teachers = [
            row for row in _v1859_effective_teachers(cur, sorov.maktab_id)
            if row.get("dars_beruvchi")
        ]
        canonical = {}
        for teacher in teachers:
            for subject in teacher.get("fanlar_royxati") or []:
                key = _v1871_method_subject_key(subject)
                if key:
                    canonical.setdefault(key, str(subject).strip())

        normalized_rules = []
        seen = set()
        for index, item in enumerate(sorov.qoidalar):
            key = _v1871_method_subject_key(item.fan_nomi)
            if not key or key in seen:
                continue
            if key not in canonical:
                raise HTTPException(
                    status_code=400,
                    detail=(
                        "Maktab o'qituvchilarida aniq fan topilmadi: "
                        f"{item.fan_nomi}"
                    ),
                )
            day = int(item.hafta_kuni)
            if day < 1 or day > weekdays:
                raise HTTPException(
                    status_code=400,
                    detail=f"{canonical[key]} uchun hafta kuni noto'g'ri",
                )
            seen.add(key)
            normalized_rules.append({
                "fan_nomi": canonical[key],
                "fan_kaliti": key,
                "hafta_kuni": day,
                "qattiq": bool(item.qattiq),
                "tartib": index,
            })

        if sorov.yoqilgan and not normalized_rules:
            raise HTTPException(
                status_code=400,
                detail=(
                    "Avto metod kunini yoqish uchun kamida bitta "
                    "fan→kun qoidasi kerak"
                ),
            )

        cur.execute(
            "DELETE FROM aqlli_metod_fan_qoidalari_v2 WHERE maktab_id=%s",
            (sorov.maktab_id,),
        )
        for rule in normalized_rules:
            cur.execute(
                """INSERT INTO aqlli_metod_fan_qoidalari_v2(
                    maktab_id,fan_nomi,fan_kaliti,hafta_kuni,qattiq,tartib)
                   VALUES(%s,%s,%s,%s,%s,%s)""",
                (
                    sorov.maktab_id,
                    rule["fan_nomi"],
                    rule["fan_kaliti"],
                    rule["hafta_kuni"],
                    rule["qattiq"],
                    rule["tartib"],
                ),
            )

        cur.execute(
            """INSERT INTO aqlli_metod_avto_sozlamalari_v2(
                maktab_id,yoqilgan,yangilagan_user_id,yangilangan_at)
               VALUES(%s,%s,%s,NOW())
               ON CONFLICT(maktab_id) DO UPDATE SET
                 yoqilgan=EXCLUDED.yoqilgan,
                 yangilagan_user_id=EXCLUDED.yangilagan_user_id,
                 yangilangan_at=NOW()""",
            (sorov.maktab_id, bool(sorov.yoqilgan), actor_id),
        )

        # Eski va yangi avtomatik qatorlar qayta quriladi.
        # Qo'lda belgilangan metod kunlariga tegilmaydi.
        auto_where = _v1871_auto_method_where()
        cur.execute(
            f"""DELETE FROM aqlli_oqituvchi_vaqti_v2
                WHERE maktab_id=%s AND turi='metod_kuni'
                  AND {auto_where}""",
            (
                sorov.maktab_id,
                _V1871_AUTO_METHOD_PREFIX + "%",
                _V1871_OLD_AUTO_PREFIX + "%",
            ),
        )

        applied = 0
        skipped_manual = []
        conflicts = []

        if sorov.yoqilgan:
            cur.execute(
                f"""SELECT DISTINCT user_id
                    FROM aqlli_oqituvchi_vaqti_v2
                    WHERE maktab_id=%s AND turi='metod_kuni'
                      AND NOT {auto_where}""",
                (
                    sorov.maktab_id,
                    _V1871_AUTO_METHOD_PREFIX + "%",
                    _V1871_OLD_AUTO_PREFIX + "%",
                ),
            )
            manual_users = {int(row["user_id"]) for row in cur.fetchall()}

            for teacher in teachers:
                uid = int(teacher["user_id"])
                subject_keys = {
                    _v1871_method_subject_key(subject)
                    for subject in (teacher.get("fanlar_royxati") or [])
                }
                matched = [
                    rule
                    for rule in normalized_rules
                    if rule["fan_kaliti"] in subject_keys
                ]
                if not matched:
                    continue

                if uid in manual_users:
                    skipped_manual.append({
                        "user_id": uid,
                        "full_name": teacher["full_name"],
                    })
                    continue

                chosen = matched[0]
                days = sorted({rule["hafta_kuni"] for rule in matched})
                if len(days) > 1:
                    conflicts.append({
                        "user_id": uid,
                        "full_name": teacher["full_name"],
                        "fanlar": [rule["fan_nomi"] for rule in matched],
                        "kunlar": [
                            _V1852_HAFTA.get(day, str(day))
                            for day in days
                        ],
                        "tanlangan": _V1852_HAFTA.get(
                            chosen["hafta_kuni"],
                            str(chosen["hafta_kuni"]),
                        ),
                    })

                note = (
                    f"{_V1871_AUTO_METHOD_PREFIX} {chosen['fan_nomi']} | "
                    f"{_V1852_HAFTA.get(chosen['hafta_kuni'], chosen['hafta_kuni'])}"
                )
                cur.execute(
                    """INSERT INTO aqlli_oqituvchi_vaqti_v2(
                        maktab_id,user_id,hafta_kuni,smena,dars_raqami,
                        turi,qattiq,izoh)
                       VALUES(%s,%s,%s,0,0,'metod_kuni',%s,%s)
                       ON CONFLICT(
                         maktab_id,user_id,hafta_kuni,smena,dars_raqami,turi
                       ) DO UPDATE SET
                         qattiq=EXCLUDED.qattiq,
                         izoh=EXCLUDED.izoh""",
                    (
                        sorov.maktab_id,
                        uid,
                        chosen["hafta_kuni"],
                        chosen["qattiq"],
                        note,
                    ),
                )
                applied += 1

        conn.commit()
        rules = _v1871_method_rules(cur, sorov.maktab_id)
        return {
            "holat": "saqlandi",
            "yoqilgan": bool(sorov.yoqilgan),
            "qoida_soni": len(rules),
            "avto_belgilangan": applied,
            "qolda_metod_borligi_uchun_otkazildi": skipped_manual,
            "bir_necha_fan_kuni_ziddiyati": conflicts,
            "hisobot": _v1871_method_report(
                cur, sorov.maktab_id, rules
            ),
        }
    except Exception:
        conn.rollback()
        raise
    finally:
        cur.close()
        conn.close()

# ========================= V18.71 END =========================

# ═══════════════════════════════════════════════════════════
# V18.73 — RASMIY METOD KUNLARI PRESETI + PSIXOLOG ISTISNOSI
# O'zA 12.12.2024 dagi fan guruhlari asosida birinchi ochilishda
# qattiq metod kunlari avtomatik qo'llanadi. Keyin rahbariyat
# o'qituvchi vaqt matritsasida bittalab tuzatishi mumkin.
# ═══════════════════════════════════════════════════════════

_V1873_METHOD_PREFIX = "V18.73 RASMIY METOD:"
_V1873_SOURCE_URL = "https://uza.uz/cn/posts/kasbiy-rivojlanish-kuni-va-kasbiy-rivojlanish-soati-joriy-etildi_667022"
_V1873_DAY_LABELS = {
    1: "Dushanba", 2: "Seshanba", 3: "Chorshanba",
    4: "Payshanba", 5: "Juma", 6: "Shanba",
}
_V1873_OFFICIAL_DISPLAY = {
    1: ["Tarix", "Davlat va huquq asoslari", "Tarbiya"],
    2: ["Ona tili", "Adabiyot", "O‘zbek tili", "Rus tili"],
    3: ["Fizika", "Astronomiya", "Kimyo", "Biologiya", "Geografiya", "Iqtisodiyot", "Tabiiy fan"],
    4: ["Matematika", "Algebra", "Geometriya", "Informatika", "Axborot texnologiyalari"],
    5: ["Ingliz tili", "Nemis tili", "Fransuz tili", "Boshqa xorijiy tillar"],
    6: ["Boshlang‘ich ta’lim", "Tasviriy san’at", "Chizmachilik", "Musiqa", "Texnologiya", "Jismoniy tarbiya", "Chaqiruvga qadar boshlang‘ich tayyorgarlik"],
}


def _v1873_norm(value):
    text = str(value or "").casefold()
    for old in ("‘", "’", "`", "ʼ", "ʻ"):
        text = text.replace(old, "'")
    text = re.sub(r"[^0-9a-zа-яёўқғҳў' ]+", " ", text, flags=re.IGNORECASE)
    return " ".join(text.split())


def _v1873_subject_day(subject, central_days=None):
    key = _v1873_norm(subject)
    if not key:
        return None
    central_day = (central_days or {}).get(_v1875_subject_key(subject))
    if central_day:
        return int(central_day)

    # Shanba amaliy fanlari avval tekshiriladi: "jismoniy tarbiya"
    # oddiy "tarbiya" bilan Dushanbaga tushib ketmasin.
    if any(token in key for token in (
        "jismoniy tarbiya", "tasviriy san", "chizmachilik",
        "musiqa", "chaqiruvga qadar", "boshlang'ich harbiy",
    )):
        return 6
    if "texnologiya" in key and "axborot" not in key and "informatika" not in key:
        return 6

    # Xorijiy tillar. Ona/O'zbek/Rus tili Seshanbada qoladi.
    if any(token in key for token in ("ingliz tili", "nemis tili", "fransuz tili", "xorijiy til")):
        return 5

    # Aniq fanlar.
    if any(token in key for token in ("matematika", "algebra", "geometriya", "informatika", "axborot texnolog")):
        return 4

    # Tabiiy fanlar va iqtisodiyot.
    if any(token in key for token in (
        "fizika", "astronomiya", "kimyo", "biologiya", "geografiya",
        "iqtisod", "tadbirkor", "tabiiy fan", "science",
    )):
        return 3

    # Filologiya.
    if any(token in key for token in ("ona tili", "adabiyot", "o'zbek tili", "rus tili", "o'qish savodxonligi")):
        return 2

    # Ijtimoiy fanlar.
    if "tarix" in key or "davlat va huquq" in key or key == "tarbiya":
        return 1

    # "Boshqa xorijiy tillar": yuqoridagi ichki tillar chiqarib tashlangach
    # qolgan "... tili" fanlari Juma bo'ladi.
    if key.endswith(" tili"):
        return 5
    return None


def _v1873_tables(cur):
    _v1871_auto_method_tables(cur)
    cur.execute(
        "ALTER TABLE aqlli_metod_avto_sozlamalari_v2 "
        "ADD COLUMN IF NOT EXISTS rasmiy_preset_v1873 BOOLEAN NOT NULL DEFAULT FALSE"
    )
    cur.execute(
        "ALTER TABLE aqlli_metod_avto_sozlamalari_v2 "
        "ADD COLUMN IF NOT EXISTS rasmiy_manba TEXT"
    )


def _v1873_primary_teacher_ids(cur, maktab_id):
    cur.execute("""
        SELECT DISTINCT rahbar_user_id AS user_id
        FROM maktab_sinflari
        WHERE maktab_id=%s AND rahbar_user_id IS NOT NULL
          AND NULLIF(REGEXP_REPLACE(COALESCE(sinf,''),'[^0-9]','','g'),'')::int BETWEEN 1 AND 4
    """, (maktab_id,))
    return {int(row["user_id"]) for row in cur.fetchall()}


def _v1873_assignments(cur, maktab_id):
    central_days_by_language = {}
    for language in _V238_INSTRUCTION_LANGUAGES:
        central_days_by_language[language] = {
            _v1875_subject_key(row["fan_nomi"]): int(row["metod_kuni"])
            for row in _v201_central_rows(
                cur,
                language,
                required_sections=("metod",),
            )
            if row.get("metod_kuni")
        }
    # Har fan qatori o'zi biriktirilgan sinf tilidan metod kunini oladi.
    # Bitta o'qituvchi turli tillarda dars bersa, ziddiyat odatdagi hisobotda
    # ko'rsatiladi; UZ metod kuni RU/EN faniga jim qo'llanmaydi.
    cur.execute("""SELECT b.user_id,b.fan_nomi,
                          COALESCE(s.talim_tili,'uz') AS talim_tili
                     FROM maktab_dars_birikmalari b
                     JOIN maktab_sinflari s ON s.id=b.sinf_id
                    WHERE b.maktab_id=%s""", (maktab_id,))
    subject_languages = {}
    for row in cur.fetchall():
        key = (int(row["user_id"]), _v1875_subject_key(row["fan_nomi"]))
        subject_languages.setdefault(key, set()).add(
            _v238_normalize_instruction_language(row.get("talim_tili"))
        )
    # Legacy importlarda canonical dars birikmasi hali yaratilmagan, ammo
    # xodim–sinf fan bog'lanishi mavjud bo'lishi mumkin. Bunday o'qituvchini
    # UZga majburan tushirmay, bog'langan sinfning haqiqiy tilini olamiz.
    cur.execute("""SELECT x.user_id,x.fanlari,
                          COALESCE(s.talim_tili,'uz') AS talim_tili
                     FROM maktab_xodim_sinflari x
                     JOIN maktab_sinflari s ON s.id=x.sinf_id
                    WHERE x.maktab_id=%s""", (maktab_id,))
    for row in cur.fetchall():
        language = _v238_normalize_instruction_language(row.get("talim_tili"))
        for subject in _v1859_fanlarni_ajrat(row.get("fanlari")):
            key = (int(row["user_id"]), _v1875_subject_key(subject))
            subject_languages.setdefault(key, set()).add(language)
    primary_ids = _v1873_primary_teacher_ids(cur, maktab_id)
    teachers = [
        row for row in _v1859_effective_teachers(cur, maktab_id)
        if row.get("dars_beruvchi") and str(row.get("lavozim") or "") != "psixolog"
    ]
    assignments = []
    conflicts = []
    for teacher in teachers:
        uid = int(teacher["user_id"])
        class_grades = []
        for class_name in teacher.get("sinflar_royxati") or []:
            match = re.match(r"^\s*(\d+)", str(class_name))
            if match:
                class_grades.append(int(match.group(1)))
        subject_keys = [_v1873_norm(subject) for subject in (teacher.get("fanlar_royxati") or [])]
        primary_subject_hits = sum(
            1 for key in subject_keys
            if any(token in key for token in (
                "ona tili", "o'qish savodxonligi", "matematika",
                "tabiiy fan", "boshlang'ich ta'lim"
            ))
        )
        primary_teacher = (
            uid in primary_ids
            or any("boshlang'ich ta'lim" in key for key in subject_keys)
            or (class_grades and max(class_grades) <= 4 and primary_subject_hits >= 2)
        )
        if primary_teacher:
            assignments.append({
                "user_id": uid,
                "full_name": teacher["full_name"],
                "hafta_kuni": 6,
                "asos": "Boshlang‘ich ta’lim",
                "fanlar": teacher.get("fanlar_royxati") or [],
            })
            continue

        by_day = {}
        for subject in teacher.get("fanlar_royxati") or []:
            languages = subject_languages.get(
                (uid, _v1875_subject_key(subject)), {"uz"}
            )
            days_for_subject = {}
            for language in sorted(languages):
                day = _v1873_subject_day(
                    subject, central_days_by_language.get(language, {})
                )
                if day:
                    days_for_subject.setdefault(day, []).append(language)
            # Bir fan UZ/RU/ENda ayni metod kuniga tushsa u ovozni ikki marta
            # oshirmaydi. Faqat haqiqatan boshqa kun bo'lsa til belgisi bilan
            # alohida nomzod bo'lib qoladi.
            for day, day_languages in sorted(days_for_subject.items()):
                label = str(subject)
                if len(days_for_subject) > 1:
                    label = f"{label} [{'/'.join(value.upper() for value in day_languages)}]"
                by_day.setdefault(day, []).append(label)
        if not by_day:
            continue

        # Bir nechta guruhga tushgan o'qituvchida eng ko'p fan tushgan kun;
        # teng bo'lsa haftadagi oldingi kun tanlanadi. Hisobotda ziddiyat chiqadi.
        chosen_day = sorted(by_day, key=lambda d: (-len(by_day[d]), d))[0]
        if len(by_day) > 1:
            conflicts.append({
                "user_id": uid,
                "full_name": teacher["full_name"],
                "kunlar": [
                    {"hafta_kuni": day, "kun_nomi": _V1873_DAY_LABELS[day], "fanlar": by_day[day]}
                    for day in sorted(by_day)
                ],
                "tanlangan_kun": chosen_day,
                "tanlangan_kun_nomi": _V1873_DAY_LABELS[chosen_day],
            })
        assignments.append({
            "user_id": uid,
            "full_name": teacher["full_name"],
            "hafta_kuni": chosen_day,
            "asos": ", ".join(by_day[chosen_day]),
            "fanlar": teacher.get("fanlar_royxati") or [],
        })
    return assignments, conflicts


def _v1873_apply_official(cur, maktab_id, replace_existing=True):
    assignments, conflicts = _v1873_assignments(cur, maktab_id)
    user_ids = [row["user_id"] for row in assignments]

    # Avval eski V18.71/V18.73 avtomatik qatorlar tozalanadi.
    cur.execute("""
        DELETE FROM aqlli_oqituvchi_vaqti_v2
        WHERE maktab_id=%s AND turi='metod_kuni'
          AND (COALESCE(izoh,'') LIKE 'V18.71 AUTO METOD:%%'
               OR COALESCE(izoh,'') LIKE 'V18.73 RASMIY METOD:%%')
    """, (maktab_id,))

    if replace_existing and user_ids:
        # Birinchi rasmiy qo'llashda shu o'qituvchilarning noto'g'ri qo'lda
        # kiritilgan metod kunlari ham almashtiriladi. Psixolog bu ro'yxatda yo'q.
        cur.execute("""
            DELETE FROM aqlli_oqituvchi_vaqti_v2
            WHERE maktab_id=%s AND turi='metod_kuni' AND user_id=ANY(%s)
        """, (maktab_id, user_ids))

    rows = []
    for item in assignments:
        note = (
            f"{_V1873_METHOD_PREFIX} {item['asos']} | "
            f"{_V1873_DAY_LABELS[item['hafta_kuni']]}"
        )
        rows.append((
            maktab_id, item["user_id"], item["hafta_kuni"], 0, 0,
            "metod_kuni", True, note,
        ))
    if rows:
        psycopg2.extras.execute_values(cur, """
            INSERT INTO aqlli_oqituvchi_vaqti_v2(
                maktab_id,user_id,hafta_kuni,smena,dars_raqami,turi,qattiq,izoh)
            VALUES %s
            ON CONFLICT(maktab_id,user_id,hafta_kuni,smena,dars_raqami,turi)
            DO UPDATE SET qattiq=EXCLUDED.qattiq,izoh=EXCLUDED.izoh
        """, rows)
    return assignments, conflicts


def _v1873_report(assignments, conflicts):
    grouped = []
    for day in range(1, 7):
        teachers = [row for row in assignments if int(row["hafta_kuni"]) == day]
        grouped.append({
            "hafta_kuni": day,
            "kun_nomi": _V1873_DAY_LABELS[day],
            "fanlar": _V1873_OFFICIAL_DISPLAY[day],
            "oqituvchi_soni": len(teachers),
            "oqituvchilar": [
                {"user_id": row["user_id"], "full_name": row["full_name"], "asos": row["asos"]}
                for row in teachers[:80]
            ],
        })
    return {
        "kunlar": grouped,
        "jami_oqituvchi": len(assignments),
        "ziddiyatlar": conflicts,
        "ziddiyat_soni": len(conflicts),
        "psixolog_izohi": "Psixologga metod kuni avtomatik belgilanmaydi.",
        "manba": _V1873_SOURCE_URL,
    }


class V1873OfficialMethodSettings(BaseModel):
    maktab_id: int
    yoqilgan: bool = True
    qayta_qollash: bool = False


@app.get("/api/maktab/aqlli_jadval/v3/metod_rasmiy_sozlama")
def v1873_official_method_get(token: str, maktab_id: int):
    actor_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor()
    try:
        _v1852_tables(cur)
        _v1873_tables(cur)
        if not _v1852_staff(cur, actor_id, maktab_id):
            raise HTTPException(status_code=403, detail="Metod kuni sozlamasini ko'rishga ruxsat yo'q")

        cur.execute("""
            SELECT yoqilgan,rasmiy_preset_v1873
            FROM aqlli_metod_avto_sozlamalari_v2 WHERE maktab_id=%s
        """, (maktab_id,))
        setting = cur.fetchone()
        first_apply = False

        if _v1852_manager(cur, actor_id, maktab_id) and not bool((setting or {}).get("rasmiy_preset_v1873")):
            assignments, conflicts = _v1873_apply_official(cur, maktab_id, replace_existing=True)
            cur.execute("""
                INSERT INTO aqlli_metod_avto_sozlamalari_v2(
                    maktab_id,yoqilgan,yangilagan_user_id,yangilangan_at,
                    rasmiy_preset_v1873,rasmiy_manba)
                VALUES(%s,TRUE,%s,NOW(),TRUE,%s)
                ON CONFLICT(maktab_id) DO UPDATE SET
                    yoqilgan=TRUE,
                    yangilagan_user_id=EXCLUDED.yangilagan_user_id,
                    yangilangan_at=NOW(),
                    rasmiy_preset_v1873=TRUE,
                    rasmiy_manba=EXCLUDED.rasmiy_manba
            """, (maktab_id, actor_id, _V1873_SOURCE_URL))
            conn.commit()
            first_apply = True
            enabled = True
        else:
            assignments, conflicts = _v1873_assignments(cur, maktab_id)
            enabled = bool((setting or {}).get("yoqilgan"))

        return {
            "yoqilgan": enabled,
            "birinchi_marta_qollandi": first_apply,
            **_v1873_report(assignments, conflicts),
        }
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


@app.put("/api/maktab/aqlli_jadval/v3/metod_rasmiy_sozlama")
def v1873_official_method_save(sorov: V1873OfficialMethodSettings, token: str):
    actor_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor()
    try:
        _v1852_tables(cur)
        _v1873_tables(cur)
        if not _v1852_manager(cur, actor_id, sorov.maktab_id):
            raise HTTPException(status_code=403, detail="Rasmiy metod kunlarini faqat maktab rahbariyati boshqaradi")
        # Rasmiy metod sozlamasi o'qituvchi bo'yicha global qo'llanadi. Shuning
        # uchun maktabdagi barcha ta'lim tillarida markaziy metod override deb
        # belgilanadi; faqat UZni belgilash RU/ENni keyin bosib ketishi mumkin.
        for language in _V238_INSTRUCTION_LANGUAGES:
            _v201_mark_school_override(
                cur, sorov.maktab_id, "metod_kunlari", actor_id, language
            )

        assignments, conflicts = _v1873_assignments(cur, sorov.maktab_id)
        if sorov.yoqilgan:
            assignments, conflicts = _v1873_apply_official(
                cur, sorov.maktab_id, replace_existing=True
            )
        else:
            cur.execute("""
                DELETE FROM aqlli_oqituvchi_vaqti_v2
                WHERE maktab_id=%s AND turi='metod_kuni'
                  AND COALESCE(izoh,'') LIKE %s
            """, (sorov.maktab_id, _V1873_METHOD_PREFIX + "%"))

        cur.execute("""
            INSERT INTO aqlli_metod_avto_sozlamalari_v2(
                maktab_id,yoqilgan,yangilagan_user_id,yangilangan_at,
                rasmiy_preset_v1873,rasmiy_manba)
            VALUES(%s,%s,%s,NOW(),TRUE,%s)
            ON CONFLICT(maktab_id) DO UPDATE SET
                yoqilgan=EXCLUDED.yoqilgan,
                yangilagan_user_id=EXCLUDED.yangilagan_user_id,
                yangilangan_at=NOW(),
                rasmiy_preset_v1873=TRUE,
                rasmiy_manba=EXCLUDED.rasmiy_manba
        """, (sorov.maktab_id, bool(sorov.yoqilgan), actor_id, _V1873_SOURCE_URL))
        conn.commit()
        return {
            "holat": "saqlandi",
            "yoqilgan": bool(sorov.yoqilgan),
            "qayta_qollandi": bool(sorov.yoqilgan and sorov.qayta_qollash),
            **_v1873_report(assignments, conflicts),
        }
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()

# ========================= V18.73 END =========================

# ═══════════════════════════════════════════════════════════
# V18.74 — SANQvaN ASOSIDAGI DARS JADVALI MANTIG'I
# 1–4-sinf kunlik yuklama, 5–11-sinf 6 dars limiti,
# asosiy fanlarni 2–3/2–4-darslarga, yengil fanlarni kechroq,
# jismoniy tarbiyani oxirgi darslarga joylashtirish.
# ═══════════════════════════════════════════════════════════

_v1874_base_build_jobs = _v1852_build_jobs_base
_v1874_base_candidate_reasons = _v1852_candidate_reasons_base
_v1874_base_place_job = _v1852_place_job_base


def _v1874_subject_key(value):
    text = unicodedata.normalize("NFKC", str(value or "")).casefold()
    for old in ("‘", "’", "`", "ʼ", "ʻ", "'", "\""):
        text = text.replace(old, "")
    text = re.sub(r"[^a-z0-9а-яёқғҳў]+", " ", text, flags=re.IGNORECASE)
    return " ".join(text.split())


def _v1874_grade(class_row):
    match = re.search(r"\d+", str((class_row or {}).get("sinf") or ""))
    return int(match.group()) if match else 0


def _v1874_contains(key, *phrases):
    return any(_v1874_subject_key(phrase) in key for phrase in phrases)


def _v1874_subject_profile(fan, grade):
    key = _v1874_subject_key(fan)
    is_class_hour = (
        key == "sinf soati"
        or "sinf soati" in key
        or key == "kelajak soati"
        or "kelajak soati" in key
    )
    is_physical = _v1874_contains(key, "jismoniy tarbiya", "jismoniy madaniyat")
    is_technology = (
        _v1874_contains(key, "texnologiya", "mehnat")
        and not _v1874_contains(key, "axborot texnologiyalari")
    )
    is_art = _v1874_contains(key, "tasviriy sanat", "chizmachilik", "rasm")
    is_music = _v1874_contains(key, "musiqa")
    is_upbringing = _v1874_contains(key, "tarbiya") and not is_physical
    is_military = _v1874_contains(key, "chaqiruvga qadar")
    is_math = _v1874_contains(key, "matematika", "algebra", "geometriya")
    is_native_language = _v1874_contains(key, "ona tili", "ozbek tili")
    is_reading = _v1874_contains(key, "oqish savodxonligi", "oqish")
    is_language = _v1874_contains(
        key, "ona tili", "ozbek tili", "rus tili", "ingliz tili",
        "nemis tili", "fransuz tili", "xorijiy til"
    )
    is_natural = _v1874_contains(
        key, "tabiiy fan", "tabiat", "informatika", "axborot texnologiyalari",
        "fizika", "kimyo", "biologiya", "geografiya", "astronomiya",
        "iqtisodiyot", "iqtisodiy bilim"
    )
    is_history = _v1874_contains(key, "tarix", "davlat va huquq", "huquq asoslari")
    is_literature = _v1874_contains(key, "adabiyot")
    is_core_science = _v1874_contains(
        key, "fizika", "kimyo", "biologiya", "tabiiy fan", "tabiat"
    )

    primary_light = bool(
        is_class_hour or is_physical or is_technology or is_art or
        is_music or is_upbringing or is_military
    )
    primary_core = bool(
        grade <= 4 and not primary_light and
        (is_math or is_language or is_reading or is_natural or is_literature)
    )

    if grade <= 4:
        if is_math:
            difficulty = 8
        elif is_language:
            difficulty = 7
        elif is_natural:
            difficulty = 6
        elif is_literature:
            difficulty = 5
        elif is_reading:
            difficulty = 4
        elif is_art or is_music:
            difficulty = 3
        elif is_technology or is_upbringing or is_class_hour:
            difficulty = 2
        elif is_physical:
            difficulty = 1
        else:
            difficulty = 4
    else:
        if _v1874_contains(key, "kimyo"):
            difficulty = 13
        elif _v1874_contains(key, "geometriya"):
            difficulty = 12
        elif _v1874_contains(key, "fizika"):
            difficulty = 12
        elif _v1874_contains(key, "algebra"):
            difficulty = 10
        elif _v1874_contains(key, "matematika"):
            difficulty = 11
        elif is_language:
            difficulty = 10
        elif is_natural:
            difficulty = 8
        elif is_history:
            difficulty = 8
        elif is_literature:
            difficulty = 6
        elif is_technology or is_art:
            difficulty = 3
        elif is_music or is_physical or is_upbringing or is_class_hour:
            difficulty = 2
        else:
            difficulty = 6

    light = bool(primary_light or (grade >= 5 and (
        is_physical or is_technology or is_art or is_music or
        is_upbringing or is_class_hour or is_military
    )))
    heavy = bool(not light and difficulty >= 7)
    written_heavy = bool(not light and (
        is_math or is_language or is_reading or is_natural or
        is_history or is_literature
    ))
    # Pedagogik asosiy fanlar: aynan o'quvchining ertalabki yuqori diqqatini
    # talab qiladigan yozma/tahliliy fanlar. Xorijiy til, informatika, tarix va
    # geografiya muhim, lekin bu qatlamga avtomatik kiritilmaydi — aks holda
    # deyarli barcha fan "asosiy" bo'lib, ustuvorlikning ma'nosi yo'qoladi.
    core_priority = bool(
        primary_core or is_math or is_native_language or is_literature
        or is_core_science or is_reading
    )

    return {
        "key": key,
        "academic": not is_class_hour,
        "class_hour": is_class_hour,
        "physical": is_physical,
        "technology": is_technology,
        "light": light,
        "primary_light": primary_light,
        "primary_core": primary_core,
        "heavy": heavy,
        "written_heavy": written_heavy,
        "math": is_math,
        "native_language": is_native_language,
        "core_science": is_core_science,
        "core_priority": core_priority,
        "difficulty": int(difficulty),
    }


def _v1874_subject_period_penalty(profile, grade, period, max_period=None):
    """Fan uchun tavsiya etilgan dars oralig'ini yumshoq ustuvorlikka aylantiradi.

    Oraliq qattiq blok emas: o'qituvchi, xona yoki boshqa majburiy qoida sabab
    mos joy qolmasa generator istisno slotdan foydalanishi mumkin.
    """
    grade = int(grade or 0)
    period = int(period or 0)
    max_period = int(max_period or _v1874_max_total_periods(grade))
    if profile.get("physical"):
        # Jismoniy tarbiya 3–6-darslarda afzal; 1–2 faqat zarur istisno.
        return {1: 320, 2: 190, 3: 25, 4: -12, 5: -24, 6: -28}.get(
            period, 20 + abs(period - min(6, max_period)) * 8
        )
    if profile.get("technology"):
        # Texnologiya amaliy fan: 1–2-dars faqat boshqa yechim qolmaganda,
        # odatda 4–6-darslarda (imkon bo'lsa J/T dan keyin) joylashadi.
        return {1: 330, 2: 220, 3: 48, 4: -12, 5: -25, 6: -30}.get(
            period, 35 + abs(period - min(6, max_period)) * 8
        )
    if profile.get("math"):
        # Matematika 1–5 oralig'ida qoladi, 2–4 eng samarali vaqt.
        return {1: 2, 2: -10, 3: -12, 4: -9, 5: 5, 6: 70}.get(
            period, 90 + max(0, period - 6) * 15
        )
    return 0


def _v1874_max_total_periods(grade):
    return 5 if 1 <= int(grade or 0) <= 4 else 6


def _v1874_fifth_day_limit(grade):
    if int(grade or 0) == 1:
        return 2
    if 2 <= int(grade or 0) <= 4:
        return 4
    return 0


def _v1874_weekly_max(grade):
    grade = int(grade or 0)
    if grade == 1:
        return 22
    if 2 <= grade <= 4:
        return 24
    if grade == 5:
        return 32
    if grade == 6:
        return 33
    if grade == 7:
        return 35
    if 8 <= grade <= 11:
        return 36
    return None


def _v1874_profile_for_job(job, context):
    profile = job.get("v1874_profile")
    if profile:
        return profile
    class_row = context.get("classes", {}).get(job.get("sinf_id"), {})
    grade = int(job.get("v1874_grade") or _v1874_grade(class_row))
    profile = _v1874_subject_profile(job.get("fan"), grade)
    if job.get("is_class_hour"):
        # Fan nomi administrator tomonidan o'zgartirilishi mumkin. Shuning
        # uchun KELAJAK SOATI profilini matndan taxmin qilmay, ish turidan
        # qat'iy belgilaymiz: u yengil va akademik bo'lmagan maxsus sessiya.
        profile = {
            **profile,
            "academic": False,
            "class_hour": True,
            "physical": False,
            "technology": False,
            "light": True,
            "primary_light": True,
            "primary_core": False,
            "heavy": False,
            "written_heavy": False,
            "math": False,
            "native_language": False,
            "core_science": False,
            "core_priority": False,
            "difficulty": 2,
        }
    job["v1874_grade"] = grade
    job["v1874_profile"] = profile
    return profile


def _v1874_build_jobs(classes, loads, assignments, group_settings, teachers):
    jobs, warnings = _v1874_base_build_jobs(
        classes, loads, assignments, group_settings, teachers
    )
    # Bir fan soati ikki o'qituvchi orasida 3+1 kabi taqsimlangan bo'lsa,
    # generator ham aynan shu kvotani saqlaydi; shunchaki tasodifiy o'qituvchi tanlamaydi.
    whole_quotas = _v1852_defaultdict(list)
    for assignment in assignments:
        if _v1875_group_key(assignment.get("guruh_kaliti")) != "whole":
            continue
        teacher_id = assignment.get("user_id")
        if teacher_id is None:
            continue
        key = (
            int(assignment["sinf_id"]),
            str(assignment.get("fan_nomi") or "").strip().casefold(),
        )
        whole_quotas[key].extend(
            [int(teacher_id)] * max(0, int(math.ceil(float(assignment.get("haftalik_soat") or 0))))
        )
    class_hours = _v1852_Counter()
    for job in jobs:
        if not job.get("groups"):
            quota = whole_quotas.get(
                (int(job["sinf_id"]), str(job.get("fan") or "").strip().casefold()), []
            )
            occurrence = max(1, int(job.get("occurrence") or 1))
            if occurrence <= len(quota):
                job["teacher_options"] = [quota[occurrence - 1]]
        class_row = classes.get(job["sinf_id"], {})
        grade = _v1874_grade(class_row)
        profile = _v1874_subject_profile(job["fan"], grade)
        job["v1874_grade"] = grade
        job["v1874_profile"] = profile
        job["difficulty"] = int(job.get("difficulty") or 0) + profile["difficulty"] * 3
        if profile["heavy"]:
            job["weight"] = max(int(job.get("weight") or 1), 3)
        elif profile["light"]:
            job["weight"] = 1
        if grade <= 4 and profile["primary_core"]:
            job["preferred_last"] = min(int(job.get("preferred_last") or 4), 4)
        elif profile["physical"] or profile["light"]:
            job["preferred_last"] = _v1874_max_total_periods(grade)
        elif profile["heavy"]:
            job["preferred_last"] = min(int(job.get("preferred_last") or 4), 4)
    # Har bir 0,5 qator boshqa 0,5 qator bilan bitta dars slotiga juftlanadi:
    # birinchisi toq, ikkinchisi juft haftada o'tadi. 1,5 soat esa bitta
    # har-haftalik dars + bitta aylanish darsi bo'ladi.
    regular_jobs = [job for job in jobs if float(job.get("rotation_weight") or 1) >= 1]
    half_by_class = _v1852_defaultdict(list)
    for job in jobs:
        if float(job.get("rotation_weight") or 1) >= 1:
            continue
        signature = (
            int(job["sinf_id"]), int(job.get("smena") or 1),
            tuple(group.get("guruh_kaliti") for group in job.get("groups") or []),
        )
        half_by_class[signature].append(job)
    rotation_pairs = 0
    for signature, half_jobs in sorted(half_by_class.items(), key=lambda item: item[0]):
        # Bir katakda almashadigan 0,5 fanlar vaqt talabi bo'yicha bir-biriga
        # yaqin bo'lsin: og'ir fan og'ir fan bilan, amaliy/yengil fan esa
        # shunga o'xshash fan bilan juftlanadi. Shunda A/B slotning ikkala
        # haftasi ham bir xil qulay dars vaqtiga tushadi.
        ordered_halves = sorted(
            half_jobs,
            key=lambda item: (
                bool((item.get("v1874_profile") or {}).get("physical")),
                bool((item.get("v1874_profile") or {}).get("light")),
                int((item.get("v1874_profile") or {}).get("difficulty") or 0),
                int(item.get("preferred_last") or 5),
                str(item.get("fan") or "").casefold(),
                int(item.get("load_id") or 0),
            ),
        )
        for index in range(0, len(ordered_halves), 2):
            members = ordered_halves[index:index + 2]
            phases = ("toq", "juft")
            rotation_members = [
                {**member, "hafta_turi": phases[member_index]}
                for member_index, member in enumerate(members)
            ]
            first = members[0]
            combined = {
                **first,
                "job_id": "rotation:" + ":".join(str(member.get("job_id")) for member in members),
                "fan": " / ".join(str(member.get("fan") or "") for member in members),
                "groups": [],
                "teacher_options": [],
                "room_id": None,
                "rotation_weight": 1.0,
                "rotation_members": rotation_members,
                "difficulty": max(float(member.get("difficulty") or 0) for member in members) + 25,
                "weight": max(int(member.get("weight") or 1) for member in members),
                "preferred_last": min(int(member.get("preferred_last") or 5) for member in members),
            }
            regular_jobs.append(combined)
            rotation_pairs += 1
    jobs = regular_jobs
    for job in jobs:
        class_hours[job["sinf_id"]] += 1

    for class_id, total in sorted(class_hours.items()):
        class_row = classes.get(class_id, {})
        grade = _v1874_grade(class_row)
        weekly_max = _v1874_weekly_max(grade)
        if weekly_max is not None and total > weekly_max:
            warnings.append(
                f"{class_row.get('sinf','')}-{class_row.get('harf','')}: "
                f"{total} soat yuklama gigiyenik haftalik maksimum {weekly_max} soatdan oshgan"
            )

    warnings.append(
        "SanQvaN profili faol: 1–4-sinfda 6-dars qo‘yilmaydi; "
        "5–11-sinfda majburiy darslar 6 tadan oshmaydi; asosiy fanlar ertaroq, "
        "yengil va jismoniy tarbiya darslari kechroq joylashtiriladi"
    )
    if rotation_pairs:
        warnings.append(
            f"A/B hafta aylanishi faol: {rotation_pairs} ta slotda 0,5 soatli fanlar "
            "toq va juft haftalarda navbat bilan o'tadi"
        )
    return jobs, warnings


def _v199_rotation_member_teachers(member):
    if member.get("groups"):
        return [group.get("teacher") for group in member.get("groups") or []]
    options = member.get("teacher_options") or []
    return [options[0] if options else None]


def _v1852_choose_teacher(job, day, period, state, context):
    members = job.get("rotation_members") or []
    if not members:
        return _v199_base_choose_teacher(job, day, period, state, context)
    selected = []
    for member in members:
        for teacher in _v199_rotation_member_teachers(member):
            if teacher not in selected:
                selected.append(teacher)
    return selected or [None]


def _v1874_state_maps(state):
    total = state.setdefault("class_daily_total", _v1852_defaultdict(int))
    academic = state.setdefault("class_daily_academic", _v1852_defaultdict(int))
    fifth_days = state.setdefault("class_fifth_academic_days", _v1852_defaultdict(set))
    difficulty = state.setdefault("class_day_difficulty", _v1852_defaultdict(int))
    period_jobs = state.setdefault("class_period_jobs", _v1852_defaultdict(dict))
    return total, academic, fifth_days, difficulty, period_jobs


def _v1874_candidate_reasons(job, day, period, selected_teachers, room_keys, state, context):
    reasons = list(_v1874_base_candidate_reasons(
        job, day, period, selected_teachers, room_keys, state, context
    ))
    # REV52: SINIF SOATI ham metod kuniga tushmaydi. Agar sinf soati qoidasi
    # aynan metod kuniga qotirilgan bo'lsa, diagnostika to'qnashuvni ko'rsatadi;
    # metod kuni esa buzilmaydi.
    profile = _v1874_profile_for_job(job, context)
    grade = int(job.get("v1874_grade") or 0)
    total_map, academic_map, fifth_map, _, _ = _v1874_state_maps(state)
    class_day = (job["sinf_id"], day)
    total_count = int(total_map.get(class_day, 0))
    academic_count = int(academic_map.get(class_day, 0))
    max_total = _v1874_max_total_periods(grade)

    if 1 <= grade <= 4 and int(day) == 6:
        reasons.append("1–4-sinf uchun Shanba o‘qish kuni emas")
    if int(period) > max_total:
        reasons.append(f"{grade}-sinf uchun {max_total}-darsdan keyin majburiy dars qo‘yilmaydi")
    if total_count >= max_total:
        reasons.append(f"sinfning kunlik jami {max_total} ta mashg‘ulot limiti to‘lgan")

    if 1 <= grade <= 4:
        if profile["academic"]:
            if state["subject_daily"].get(
                (job["sinf_id"], job["fan"].casefold(), day), 0
            ) >= 1:
                reasons.append("boshlang‘ich sinfda bir fan shu kuni takror qo‘yilmaydi")
            if academic_count >= 5:
                reasons.append("boshlang‘ich sinfning akademik kunlik limiti to‘lgan")
            if academic_count == 4:
                fifth_days = fifth_map.get(job["sinf_id"], set())
                fifth_limit = _v1874_fifth_day_limit(grade)
                if day not in fifth_days and len(fifth_days) >= fifth_limit:
                    reasons.append(
                        f"{grade}-sinfda 5 akademik darsli kunlar soni {fifth_limit} tadan oshmaydi"
                    )
                if not profile["primary_light"]:
                    reasons.append("boshlang‘ichda 5-akademik dars faqat yengil fan bo‘lishi kerak")
        if int(period) == 5 and profile["academic"] and not profile["primary_light"]:
            reasons.append("5-darsga matematika, til, o‘qish yoki boshqa og‘ir fan qo‘yilmaydi")

    return list(dict.fromkeys(reasons))




def _v1874_place_job(job, day, period, teachers, room_keys, state, context):
    _v1874_base_place_job(job, day, period, teachers, room_keys, state, context)
    rotation_members = job.get("rotation_members") or []
    if rotation_members:
        # Bazaviy joylashtiruvchi har bir o'qituvchiga 1 soat qo'shadi.
        # A/B aylanishida esa o'qituvchi har faza uchun 0,5 yuklama oladi.
        contributions = _v1852_defaultdict(float)
        for member in rotation_members:
            for teacher in set(_v199_rotation_member_teachers(member)):
                if teacher is not None:
                    contributions[int(teacher)] += 0.5
        for teacher, contribution in contributions.items():
            adjustment = max(0.0, 1.0 - contribution)
            state["teacher_week"][teacher] -= adjustment
            state["teacher_daily"][(teacher, day)] -= adjustment
    profile = _v1874_profile_for_job(job, context)
    total_map, academic_map, fifth_map, difficulty_map, period_jobs = _v1874_state_maps(state)
    class_day = (job["sinf_id"], day)
    total_map[class_day] += 1
    if profile["academic"]:
        academic_map[class_day] += 1
        if academic_map[class_day] == 5:
            fifth_map[job["sinf_id"]].add(day)
    difficulty_map[class_day] += int(profile["difficulty"])
    period_jobs[class_day][int(period)] = job


def _v1874_schedule_hygiene_violations(cur, maktab_id: int, run_id: int):
    cur.execute(
        """SELECT DISTINCT e.sinf_id,s.sinf,s.harf,e.hafta_kuni,e.dars_raqami,e.fan_nomi,
                          s.sinf::int AS sinf_sort
           FROM aqlli_jadval_slotlari_v2 e
           JOIN maktab_sinflari s ON s.id=e.sinf_id
           WHERE e.maktab_id=%s AND e.urinish_id=%s
           ORDER BY sinf_sort,s.harf,e.hafta_kuni,e.dars_raqami,e.fan_nomi""",
        (maktab_id, run_id),
    )
    rows = cur.fetchall()
    by_class_day = _v1852_defaultdict(list)
    class_names = {}
    for row in rows:
        class_id = int(row["sinf_id"])
        grade_match = re.search(r"\d+", str(row.get("sinf") or ""))
        grade = int(grade_match.group()) if grade_match else 0
        class_names[class_id] = (f"{row.get('sinf','')}-{row.get('harf','')}", grade)
        by_class_day[(class_id, int(row["hafta_kuni"]))].append(row)

    violations = []
    fifth_days = _v1852_defaultdict(set)
    core_period6_days = _v1852_defaultdict(set)
    for (class_id, day), day_rows in by_class_day.items():
        class_name, grade = class_names[class_id]
        period_fans = _v1852_defaultdict(set)
        for row in day_rows:
            period_fans[int(row["dars_raqami"])].add(str(row["fan_nomi"]))
        total_sessions = len(period_fans)
        max_total = _v1874_max_total_periods(grade)
        if period_fans:
            last_period = max(period_fans)
            missing_periods = [
                period for period in range(1, last_period + 1)
                if period not in period_fans
            ]
            if missing_periods:
                missing_text = ", ".join(str(period) for period in missing_periods)
                violations.append({
                    "sinf": class_name,
                    "sabab": (
                        f"{_V1852_HAFTA.get(day, day)} kuni darslar orasida "
                        f"bo'sh okno bor: {missing_text}-dars"
                    ),
                })
        if 1 <= grade <= 4 and day == 6:
            violations.append({"sinf": class_name, "sabab": "Shanba kuni boshlang‘ich sinf darsi bor"})
        if total_sessions > max_total:
            violations.append({
                "sinf": class_name,
                "sabab": f"{_V1852_HAFTA.get(day, day)} kuni {total_sessions} ta mashg‘ulot; maksimum {max_total}",
            })
        if period_fans and max(period_fans) > max_total:
            violations.append({
                "sinf": class_name,
                "sabab": f"{_V1852_HAFTA.get(day, day)} kuni {max(period_fans)}-dars gigiyenik limitdan tashqari",
            })

        academic_periods = 0
        subject_periods = _v1852_defaultdict(set)
        for period, fans in period_fans.items():
            has_academic = False
            for fan in fans:
                profile = _v1874_subject_profile(fan, grade)
                if profile["academic"]:
                    has_academic = True
                    subject_periods[profile["key"]].add(period)
                if int(period) == 6 and profile.get("core_priority"):
                    core_period6_days[class_id].add(int(day))
                if 1 <= grade <= 4 and period == 5 and profile["academic"] and not profile["primary_light"]:
                    violations.append({
                        "sinf": class_name,
                        "sabab": f"{_V1852_HAFTA.get(day, day)} 5-darsda og‘ir fan: {fan}",
                    })
            if has_academic:
                academic_periods += 1
        if academic_periods >= 5 and 1 <= grade <= 4:
            fifth_days[class_id].add(day)
        # V22.36: ayni fan kuniga 2 marta bo'lishi qonuniy zaxira.
        # Phase-aware yakuniy validator 3-martani alohida qat'iy rad etadi;
        # gigiyena qatlami legal 2+1+1 / 2+2+1 taqsimotni qayta rad etmaydi.

    for class_id, days in fifth_days.items():
        class_name, grade = class_names[class_id]
        limit = _v1874_fifth_day_limit(grade)
        if len(days) > limit:
            violations.append({
                "sinf": class_name,
                "sabab": f"5 akademik darsli kunlar {len(days)} ta; maksimum {limit}",
            })

    for class_id, days in core_period6_days.items():
        if len(days) <= _V213_CORE_PERIOD6_LIMIT:
            continue
        class_name, _grade = class_names[class_id]
        violations.append({
            "sinf": class_name,
            "sabab": (
                f"{_V213_CORE_PERIOD6_REASON}: {len(days)} kun topildi"
            ),
        })

    # O‘qituvchining haftalik chegarasi ham tasdiqlash oldidan qayta tekshiriladi.
    cur.execute(
        """WITH teacher_sessions AS (
               SELECT DISTINCT
                      e.oqituvchi_user_id,e.sinf_id,e.hafta_kuni,
                      e.smena,e.dars_raqami,
                      COALESCE(e.hafta_turi,'har_hafta') AS hafta_turi,
                      EXISTS(
                          SELECT 1 FROM aqlli_sinf_soati_qoidalari_v2 q
                          WHERE q.maktab_id=e.maktab_id
                            AND q.sinf_id=e.sinf_id
                            AND q.faol=TRUE
                            AND q.hafta_kuni=e.hafta_kuni
                            AND q.dars_raqami=e.dars_raqami
                      ) AS sinf_soati
               FROM aqlli_jadval_slotlari_v2 e
               WHERE e.maktab_id=%s AND e.urinish_id=%s
                 AND e.oqituvchi_user_id IS NOT NULL
           )
           SELECT t.oqituvchi_user_id,u.full_name,u.haftalik_dars_soati,
                  SUM(CASE WHEN t.hafta_turi IN ('toq','juft')
                           THEN 0.5 ELSE 1.0 END) AS amaldagi,
                  SUM(CASE WHEN t.sinf_soati
                           THEN CASE WHEN t.hafta_turi IN ('toq','juft')
                                     THEN 0.5 ELSE 1.0 END
                           ELSE 0.0 END) AS sinf_soati
           FROM teacher_sessions t
           JOIN users u ON u.user_id=t.oqituvchi_user_id
           GROUP BY t.oqituvchi_user_id,u.full_name,u.haftalik_dars_soati""",
        (maktab_id, run_id),
    )
    for row in cur.fetchall():
        if row.get("haftalik_dars_soati") is None:
            continue
        cap = round(
            float(row["haftalik_dars_soati"])
            + float(row.get("sinf_soati") or 0),
            1,
        )
        actual = round(float(row.get("amaldagi") or 0), 1)
        if actual > cap + 1e-9:
            violations.append({
                "sinf": row["full_name"],
                "sabab": f"o‘qituvchi yuklamasi {actual} soat, ruxsat etilgan {cap} soatdan oshgan",
            })

    unique = []
    seen = set()
    for item in violations:
        key = (item["sinf"], item["sabab"])
        if key not in seen:
            seen.add(key)
            unique.append(item)
    return unique

# ========================= V18.74 END =========================

# ═══════════════════════════════════════════════════════════
# V18.75 — SHABLON → REJA → DRAFT → TASDIQ 100% MOSLIK
# O'qituvchi haftalik jami, o'qituvchi kun/vaqt cheklovi,
# sinf haftalik jami, fan haftalik soati va haqiqiy jadval bir xil bo'lmasa
# tasdiqlash mutlaqo rad etiladi.
# ═══════════════════════════════════════════════════════════


def _v1875_tables(cur):
    _v1852_tables(cur)
    cur.execute(
        "ALTER TABLE aqlli_sinf_fan_yuklamalari_v2 "
        "ADD COLUMN IF NOT EXISTS manba TEXT NOT NULL DEFAULT 'qolda'"
    )
    cur.execute(
        "ALTER TABLE aqlli_sinf_fan_yuklamalari_v2 "
        "ADD COLUMN IF NOT EXISTS manba_hash TEXT"
    )
    cur.execute("""CREATE TABLE IF NOT EXISTS aqlli_jadval_manba_holati_v2(
        maktab_id INTEGER PRIMARY KEY REFERENCES maktablar(id) ON DELETE CASCADE,
        manba_hash TEXT NOT NULL,
        sabab TEXT,
        yangilangan_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
    )""")


def _v1875_subject_key(value):
    if "_v1874_subject_key" in globals():
        return _v1874_subject_key(value)
    return _xodim_excel_sarlavha_kaliti(value)


def _v1875_group_key(value):
    key = str(value or "whole").strip()
    if key.casefold() in {"", "whole", "butun sinf", "butun_sinf"}:
        return "whole"
    return key


def _v1875_exact_assignment_model(cur, maktab_id: int):
    _xodim_sinf_birikmalari_jadvali(cur)
    cur.execute("""SELECT b.id,b.sinf_id,b.user_id,b.fan_nomi,b.guruh_kaliti,
                          b.haftalik_soat,b.kunlik_max,b.manba,
                          u.full_name,s.sinf,s.harf,COALESCE(s.smena,1) AS smena
                   FROM maktab_dars_birikmalari b
                   JOIN users u ON u.user_id=b.user_id
                   JOIN maktab_sinflari s ON s.id=b.sinf_id
                   WHERE b.maktab_id=%s
                   ORDER BY s.sinf::int,s.harf,b.fan_nomi,b.guruh_kaliti,u.full_name""",
                (maktab_id,))
    rows = [dict(row) for row in cur.fetchall()]
    pairs = {}
    errors = []
    warnings = []

    for row in rows:
        subject = re.sub(r"\s+", " ", str(row.get("fan_nomi") or "")).strip()
        subject_key = _v1875_subject_key(subject)
        group_key = _v1875_group_key(row.get("guruh_kaliti"))
        try:
            hours = float(row.get("haftalik_soat") or 0)
        except (TypeError, ValueError):
            hours = 0
        try:
            daily_max = int(row.get("kunlik_max") or 1)
        except (TypeError, ValueError):
            daily_max = 1
        class_label = f"{row.get('sinf','')}-{row.get('harf','')}"
        if not subject_key:
            errors.append(f"{class_label}: fan nomi bo'sh")
            continue
        if hours <= 0:
            errors.append(
                f"{class_label} / {subject} / {row.get('full_name')}: haftalik soat 0 yoki bo'sh"
            )
        if daily_max < 1 or daily_max > 4:
            errors.append(f"{class_label} / {subject}: kunlik max 1–4 bo'lishi kerak")
        key = (int(row["sinf_id"]), subject_key)
        pair = pairs.setdefault(key, {
            "sinf_id": int(row["sinf_id"]),
            "sinf": class_label,
            "smena": int(row.get("smena") or 1),
            "fan_nomi": subject,
            "fan_kaliti": subject_key,
            "rows": [],
        })
        pair["rows"].append({
            **row,
            "fan_nomi": subject,
            "fan_kaliti": subject_key,
            "guruh_kaliti": group_key,
            "haftalik_soat": hours,
            "kunlik_max": daily_max,
            "user_id": int(row["user_id"]),
        })

    teacher_hours = _v1852_Counter()
    class_hours = _v1852_Counter()
    valid_pairs = {}

    for key, pair in pairs.items():
        whole = [row for row in pair["rows"] if row["guruh_kaliti"] == "whole"]
        groups = [row for row in pair["rows"] if row["guruh_kaliti"] != "whole"]
        if whole and groups:
            errors.append(
                f"{pair['sinf']} / {pair['fan_nomi']}: butun sinf va guruh qatorlari birga yozilgan"
            )
            continue

        if whole:
            if len(whole) != 1:
                errors.append(
                    f"{pair['sinf']} / {pair['fan_nomi']}: butun sinf uchun takror qatorlar bor"
                )
                continue
            teachers = sorted({row["user_id"] for row in whole})
            hours_set = sorted({row["haftalik_soat"] for row in whole if row["haftalik_soat"] > 0})
            daily_set = sorted({row["kunlik_max"] for row in whole})
            if len(teachers) != 1:
                names = ", ".join(sorted({str(row.get("full_name") or row["user_id"]) for row in whole}))
                errors.append(
                    f"{pair['sinf']} / {pair['fan_nomi']}: butun sinf uchun bitta o'qituvchi kerak ({names})"
                )
                continue
            if len(hours_set) != 1:
                errors.append(
                    f"{pair['sinf']} / {pair['fan_nomi']}: o'qituvchi qatorlaridagi haftalik soatlar bir xil emas"
                )
                continue
            if len(daily_set) > 1:
                errors.append(
                    f"{pair['sinf']} / {pair['fan_nomi']}: kunlik maksimumlar bir xil emas"
                )
                continue
            hours = hours_set[0] if hours_set else 0
            teacher_id = teachers[0]
            pair.update({
                "turi": "whole",
                "haftalik_soat": hours,
                "kunlik_max": daily_set[0] if daily_set else 1,
                "asosiy_oqituvchi_user_id": teacher_id,
                "guruhlar": [],
                "oqituvchilar": [teacher_id],
            })
            teacher_hours[teacher_id] += hours
        else:
            by_group = {}
            for row in groups:
                by_group.setdefault(row["guruh_kaliti"], []).append(row)
            group_payload = []
            hours_set = set()
            daily_set = set()
            used_teachers = set()
            group_error = False
            for group_key, group_rows in sorted(by_group.items()):
                if len(group_rows) != 1:
                    errors.append(
                        f"{pair['sinf']} / {pair['fan_nomi']} / {group_key}: takror qatorlar bor"
                    )
                    group_error = True
                    continue
                group_teachers = sorted({row["user_id"] for row in group_rows})
                group_hours = sorted({row["haftalik_soat"] for row in group_rows if row["haftalik_soat"] > 0})
                group_daily = sorted({row["kunlik_max"] for row in group_rows})
                if len(group_teachers) != 1:
                    errors.append(
                        f"{pair['sinf']} / {pair['fan_nomi']} / {group_key}: bitta o'qituvchi kerak"
                    )
                    group_error = True
                    continue
                if len(group_hours) != 1:
                    errors.append(
                        f"{pair['sinf']} / {pair['fan_nomi']} / {group_key}: haftalik soat yagona emas"
                    )
                    group_error = True
                    continue
                teacher_id = group_teachers[0]
                if teacher_id in used_teachers:
                    errors.append(
                        f"{pair['sinf']} / {pair['fan_nomi']}: bir o'qituvchi ikki parallel guruhga biriktirilgan"
                    )
                    group_error = True
                used_teachers.add(teacher_id)
                hours = group_hours[0]
                hours_set.add(hours)
                daily_set.update(group_daily)
                group_payload.append({
                    "guruh_kaliti": group_key,
                    "oqituvchi_user_id": teacher_id,
                    "haftalik_soat": hours,
                    "kunlik_max": group_daily[0] if group_daily else 1,
                    "full_name": group_rows[0].get("full_name"),
                })
            if group_error:
                continue
            if len(hours_set) != 1:
                errors.append(
                    f"{pair['sinf']} / {pair['fan_nomi']}: parallel guruhlarning haftalik soati teng bo'lishi kerak"
                )
                continue
            if len(daily_set) > 1:
                errors.append(
                    f"{pair['sinf']} / {pair['fan_nomi']}: parallel guruhlarning kunlik maksimumi teng bo'lishi kerak"
                )
                continue
            if len(group_payload) < 2:
                warnings.append(
                    f"{pair['sinf']} / {pair['fan_nomi']}: guruhli fan uchun faqat {len(group_payload)} ta guruh topildi"
                )
            hours = next(iter(hours_set), 0)
            pair.update({
                "turi": "group",
                "haftalik_soat": hours,
                "kunlik_max": next(iter(daily_set), 1),
                "asosiy_oqituvchi_user_id": None,
                "guruhlar": group_payload,
                "oqituvchilar": [row["oqituvchi_user_id"] for row in group_payload],
            })
            for group in group_payload:
                teacher_hours[group["oqituvchi_user_id"]] += hours

        class_hours[pair["sinf_id"]] += float(pair.get("haftalik_soat") or 0)
        valid_pairs[key] = pair

    return {
        "rows": rows,
        "pairs": valid_pairs,
        "teacher_hours": dict(teacher_hours),
        "class_hours": dict(class_hours),
        "xatolar": list(dict.fromkeys(errors)),
        "ogohlantirishlar": list(dict.fromkeys(warnings)),
    }


def _v1875_source_fingerprint(cur, maktab_id: int):
    queries = [
        ("sinflar", """SELECT id,sinf,harf,COALESCE(smena,1) AS smena,
                                bino,xona,bino_id,xona_id,rahbar_user_id
                         FROM maktab_sinflari WHERE maktab_id=%s
                         ORDER BY id"""),
        ("birikmalar", """SELECT b.sinf_id,b.user_id,b.fan_nomi,b.guruh_kaliti,
                                  b.haftalik_soat,b.kunlik_max
                           FROM maktab_dars_birikmalari b
                           WHERE b.maktab_id=%s
                           ORDER BY b.sinf_id,b.fan_nomi,b.guruh_kaliti,b.user_id"""),
        ("yuklamalar", """SELECT sinf_id,fan_nomi,haftalik_soat,kunlik_max,
                                  ketma_ket_mumkin,afzal_oxirgi_dars,
                                  asosiy_oqituvchi_user_id,xona_id,ogirlik
                           FROM aqlli_sinf_fan_yuklamalari_v2
                           WHERE maktab_id=%s AND haftalik_soat>0
                           ORDER BY sinf_id,fan_nomi"""),
        ("vaqtlar", """SELECT user_id,hafta_kuni,smena,dars_raqami,turi,qattiq
                        FROM aqlli_oqituvchi_vaqti_v2 WHERE maktab_id=%s
                        ORDER BY user_id,hafta_kuni,smena,dars_raqami,turi"""),
        ("qoidalar", """SELECT user_id,kunlik_max,ketma_ket_max,okno_max,
                                  afzal_smena,eng_erta_dars,eng_kech_dars
                           FROM aqlli_oqituvchi_qoidalari_v2 WHERE maktab_id=%s
                           ORDER BY user_id"""),
        ("smenalar", """SELECT smena,dars_soni,boshlanish_vaqti,dars_daqiqa,
                                  tanaffus_daqiqa,katta_tanaffus_darsdan_keyin,
                                  katta_tanaffus_daqiqa
                           FROM aqlli_smena_sozlamalari_v2 WHERE maktab_id=%s
                           ORDER BY smena"""),
        ("sinf_kun", """SELECT sinf_daraja,sinf_id,hafta_kuni,faol
                           FROM aqlli_sinf_kun_bloklari_v2 WHERE maktab_id=%s
                           ORDER BY id"""),
        ("sinf_soati", """SELECT sinf_id,hafta_kuni,dars_raqami,fan_nomi,
                                    haftalik_soat,faol
                             FROM aqlli_sinf_soati_qoidalari_v2 WHERE maktab_id=%s
                             ORDER BY sinf_id"""),
        ("guruh_xonalari", """SELECT sinf_id,fan_nomi,guruh_kaliti,
                                        oqituvchi_user_id,xona_id
                                 FROM aqlli_guruh_sozlamalari_v2
                                 WHERE maktab_id=%s
                                 ORDER BY sinf_id,fan_nomi,guruh_kaliti"""),
        ("xonalar", """SELECT id,nomi,turi,bino_id,darsga_yaroqli,faol
                          FROM aqlli_xonalar_v2 WHERE maktab_id=%s
                          ORDER BY id"""),
    ]
    payload = {}
    for name, query in queries:
        cur.execute(query, (maktab_id,))
        payload[name] = [dict(row) for row in cur.fetchall()]
    raw = json.dumps(payload, ensure_ascii=False, sort_keys=True, default=str)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _v1875_rebuild_schedule_sources(cur, maktab_id: int, cancel_drafts=True, reason="sync"):
    _v1875_tables(cur)
    model = _v1875_exact_assignment_model(cur, maktab_id)
    if model["xatolar"]:
        return model

    # Ko'p tilli rejada tillardan bittasi draft bo'lsa global reja ham draft.
    # Bu holatda eski, tasdiqlangan yuklama va guruh sozlamalariga umuman
    # tegmaymiz: yangi qo'lda kiritilgan birikmalar keyingi to'liq tasdiqdan
    # so'ng UPSERT orqali sinxronlanadi.
    if _v193_approved_plan_map(cur, maktab_id) is None:
        draft_message = (
            "O'quv rejaning barcha faol tillari tasdiqlanmaguncha "
            "eski jadval manbasi o'zgartirilmaydi"
        )
        return {
            "tayyor": False,
            "fan_sinf_soni": 0,
            "oqituvchi_soni": len(model.get("teacher_hours") or {}),
            "sinflar": dict(model.get("class_hours") or {}),
            "oqituvchilar": dict(model.get("teacher_hours") or {}),
            "xatolar": (
                [draft_message]
                if reason in {"jadval_yaratish_v23", "moslik_tekshiruvi"}
                else []
            ),
            "ogohlantirishlar": list(dict.fromkeys([
                *(model.get("ogohlantirishlar") or []),
                draft_message,
            ])),
            "yuklamalar": [],
        }

    # Old pedagogik sozlamalarni fan nomi normallashgan kalit bilan saqlab qolamiz.
    cur.execute("SELECT * FROM aqlli_sinf_fan_yuklamalari_v2 WHERE maktab_id=%s", (maktab_id,))
    old_loads = [dict(row) for row in cur.fetchall()]
    old_map = {}
    for row in old_loads:
        old_map.setdefault(
            (int(row["sinf_id"]), _v1875_subject_key(row["fan_nomi"])), []
        ).append(row)

    cur.execute("SELECT * FROM aqlli_guruh_sozlamalari_v2 WHERE maktab_id=%s", (maktab_id,))
    old_groups = {
        (int(row["sinf_id"]), _v1875_subject_key(row["fan_nomi"]), str(row["guruh_kaliti"])): dict(row)
        for row in cur.fetchall()
    }

    # Faqat modeldagi aniq juftliklar UPSERT qilinadi. Maktab bo'yicha global
    # nolga tushirish/DELETE boshqa til yoki avvalgi guruh sxemasini yo'qotar edi.

    synced_pairs = []
    for key, pair in sorted(model["pairs"].items(), key=lambda item: (item[0][0], item[1]["fan_nomi"].casefold())):
        # Generator fan nomini oddiy casefold bilan bog'laydi. Shu sabab barcha exact
        # assignment qatorlarini bitta kanonik fan nomiga keltiramiz.
        for source_row in pair.get("rows", []):
            cur.execute(
                """UPDATE maktab_dars_birikmalari
                   SET fan_nomi=%s,guruh_kaliti=%s
                   WHERE id=%s""",
                (pair["fan_nomi"], source_row["guruh_kaliti"], source_row["id"]),
            )
        existing_candidates = old_map.get(key, [])
        existing = existing_candidates[0] if existing_candidates else {}
        fan_nomi = pair["fan_nomi"]
        defaults = {
            "ketma_ket_mumkin": bool(existing.get("ketma_ket_mumkin", False)),
            "afzal_oxirgi_dars": int(existing.get("afzal_oxirgi_dars") or 5),
            "xona_id": existing.get("xona_id"),
            "nazorat_soni": int(existing.get("nazorat_soni") or 0),
            "nazoratdan_keyin_tahlil": bool(existing.get("nazoratdan_keyin_tahlil", True)),
            "mustahkamlash_soni": int(existing.get("mustahkamlash_soni") or 0),
            "ogirlik": int(existing.get("ogirlik") or 2),
        }
        cur.execute("""INSERT INTO aqlli_sinf_fan_yuklamalari_v2(
                        maktab_id,sinf_id,fan_nomi,haftalik_soat,kunlik_max,
                        ketma_ket_mumkin,afzal_oxirgi_dars,
                        asosiy_oqituvchi_user_id,xona_id,nazorat_soni,
                        nazoratdan_keyin_tahlil,mustahkamlash_soni,ogirlik,
                        manba)
                       VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'DARS_BIRIKMALARI')
                       ON CONFLICT(maktab_id,sinf_id,fan_nomi) DO UPDATE SET
                         haftalik_soat=EXCLUDED.haftalik_soat,
                         kunlik_max=EXCLUDED.kunlik_max,
                         ketma_ket_mumkin=EXCLUDED.ketma_ket_mumkin,
                         afzal_oxirgi_dars=EXCLUDED.afzal_oxirgi_dars,
                         asosiy_oqituvchi_user_id=EXCLUDED.asosiy_oqituvchi_user_id,
                         xona_id=EXCLUDED.xona_id,
                         nazorat_soni=EXCLUDED.nazorat_soni,
                         nazoratdan_keyin_tahlil=EXCLUDED.nazoratdan_keyin_tahlil,
                         mustahkamlash_soni=EXCLUDED.mustahkamlash_soni,
                         ogirlik=EXCLUDED.ogirlik,
                         manba='DARS_BIRIKMALARI'""",
                    (
                        maktab_id, pair["sinf_id"], fan_nomi,
                        pair["haftalik_soat"], pair["kunlik_max"],
                        defaults["ketma_ket_mumkin"], defaults["afzal_oxirgi_dars"],
                        pair.get("asosiy_oqituvchi_user_id"), defaults["xona_id"],
                        defaults["nazorat_soni"], defaults["nazoratdan_keyin_tahlil"],
                        defaults["mustahkamlash_soni"], defaults["ogirlik"],
                    ))
        if pair["turi"] == "group":
            for group in pair["guruhlar"]:
                old = old_groups.get((pair["sinf_id"], pair["fan_kaliti"], group["guruh_kaliti"]), {})
                cur.execute("""INSERT INTO aqlli_guruh_sozlamalari_v2(
                                maktab_id,sinf_id,fan_nomi,guruh_kaliti,
                                oqituvchi_user_id,xona_id)
                               VALUES(%s,%s,%s,%s,%s,%s)
                               ON CONFLICT(maktab_id,sinf_id,fan_nomi,guruh_kaliti)
                               DO UPDATE SET oqituvchi_user_id=EXCLUDED.oqituvchi_user_id,
                                             xona_id=EXCLUDED.xona_id""",
                            (maktab_id, pair["sinf_id"], fan_nomi,
                             group["guruh_kaliti"], group["oqituvchi_user_id"],
                             old.get("xona_id")))
        synced_pairs.append({
            "sinf_id": pair["sinf_id"], "sinf": pair["sinf"],
            "fan": fan_nomi, "haftalik_soat": pair["haftalik_soat"],
            "turi": pair["turi"], "oqituvchilar": pair["oqituvchilar"],
        })

    # Faqat ayni modelda qatnashgan o'qituvchilar jami yangilanadi; aloqasiz
    # tarixiy/xizmat yozuvlari ommaviy nolga tushirilmaydi.
    for teacher_id, hours in model["teacher_hours"].items():
        cur.execute("UPDATE users SET haftalik_dars_soati=%s WHERE maktab_id=%s AND user_id=%s",
                    (round(float(hours), 1), maktab_id, int(teacher_id)))

    if cancel_drafts:
        cur.execute("""UPDATE aqlli_jadval_urinishlari_v2
                       SET holat='bekor'
                       WHERE maktab_id=%s AND holat='draft'""", (maktab_id,))

    source_hash = _v1875_source_fingerprint(cur, maktab_id)
    cur.execute("""INSERT INTO aqlli_jadval_manba_holati_v2(
                    maktab_id,manba_hash,sabab,yangilangan_at)
                   VALUES(%s,%s,%s,NOW())
                   ON CONFLICT(maktab_id) DO UPDATE SET
                     manba_hash=EXCLUDED.manba_hash,
                     sabab=EXCLUDED.sabab,
                     yangilangan_at=NOW()""",
                (maktab_id, source_hash, reason))

    return {
        "tayyor": True,
        "manba_hash": source_hash,
        "fan_sinf_soni": len(synced_pairs),
        "oqituvchi_soni": len(model["teacher_hours"]),
        "sinflar": dict(model["class_hours"]),
        "oqituvchilar": dict(model["teacher_hours"]),
        "xatolar": [],
        "ogohlantirishlar": model["ogohlantirishlar"],
        "yuklamalar": synced_pairs[:300],
    }


def _v1875_preflight_report(cur, maktab_id: int):
    _v1875_tables(cur)
    model = _v1875_exact_assignment_model(cur, maktab_id)
    errors = list(model["xatolar"])
    warnings = list(model["ogohlantirishlar"])
    auto_method_relaxation = bool(
        (_timetable_mode_config() or {}).get("method_day_relaxed", False)
    )

    year = _v1890_generation_year(cur, maktab_id)
    weekdays = int(year.get("hafta_kunlari") or 6)
    _v1852_default_shifts(cur, maktab_id)
    cur.execute("SELECT * FROM aqlli_smena_sozlamalari_v2 WHERE maktab_id=%s", (maktab_id,))
    shifts = {int(row["smena"]): dict(row) for row in cur.fetchall()}
    cur.execute("""SELECT id,sinf,harf,COALESCE(smena,1) AS smena,
                          bino,xona,bino_id,xona_id,rahbar_user_id
                   FROM maktab_sinflari WHERE maktab_id=%s ORDER BY sinf::int,harf""",
                (maktab_id,))
    classes = {int(row["id"]): dict(row) for row in cur.fetchall()}
    cur.execute("SELECT * FROM aqlli_sinf_fan_yuklamalari_v2 WHERE maktab_id=%s AND haftalik_soat>0",
                (maktab_id,))
    loads = [dict(row) for row in cur.fetchall()]
    load_map = {(int(row["sinf_id"]), _v1875_subject_key(row["fan_nomi"])): row for row in loads}

    # Imported/old data can bypass the current school wizard's global home
    # room uniqueness rule.  Detect it (and duplicate parallel-group rooms)
    # before CP-SAT so the administrator receives the exact source error
    # instead of a generic global INFEASIBLE result.
    cur.execute("SELECT * FROM aqlli_xonalar_v2 WHERE maktab_id=%s", (maktab_id,))
    room_rows = [dict(row) for row in cur.fetchall()]
    classes = {
        class_id: _v205_annotate_class_home_room(class_row, room_rows)
        for class_id, class_row in classes.items()
    }
    cur.execute(
        """SELECT * FROM aqlli_guruh_sozlamalari_v2
           WHERE maktab_id=%s ORDER BY sinf_id,fan_nomi,guruh_kaliti""",
        (maktab_id,),
    )
    group_room_rows = [dict(row) for row in cur.fetchall()]
    errors.extend(_v220_room_source_errors(classes, room_rows, group_room_rows))

    for key, pair in model["pairs"].items():
        load = load_map.get(key)
        if not load:
            errors.append(f"{pair['sinf']} / {pair['fan_nomi']}: Aqlli jadval yuklamasi yaratilmagan")
            continue
        if abs(float(load.get("haftalik_soat") or 0) - float(pair.get("haftalik_soat") or 0)) > 1e-9:
            errors.append(
                f"{pair['sinf']} / {pair['fan_nomi']}: shablon {pair['haftalik_soat']} soat, "
                f"jadval manbasi {load.get('haftalik_soat')} soat"
            )

    for key, load in load_map.items():
        if key not in model["pairs"]:
            errors.append(
                f"{classes.get(key[0], {}).get('sinf','')}-{classes.get(key[0], {}).get('harf','')} / "
                f"{load['fan_nomi']}: shablonda aniq o'qituvchi birikmasi yo'q"
            )

    class_day_blocks = _v1856_class_day_rule_map(_v1856_class_day_rule_rows(cur, maktab_id))
    class_hour_rules = _v1866_class_hour_rule_rows(cur, maktab_id)
    class_hour_by_class = {int(row["sinf_id"]): row for row in class_hour_rules}
    class_hour_by_teacher = _v1852_Counter()
    class_hour_rules_by_teacher = _v1852_defaultdict(list)
    for row in class_hour_rules:
        if row.get("rahbar_user_id") is not None:
            leader_id = int(row["rahbar_user_id"])
            class_hour_by_teacher[leader_id] += int(row.get("haftalik_soat") or 1)
            class_hour_rules_by_teacher[leader_id].append(row)

    # Qat'iy sinf-soati ikkita turli smenada boshqa dars raqamida tursa ham
    # haqiqiy soati ustma-ust kelishi mumkin. Generatorga yuborishdan oldin
    # rahbar, kun va aniq vaqt oralig'i bo'yicha tushunarli xato qaytaramiz.
    shift_intervals = _v1866_shift_interval_map(shifts.values())
    for teacher_id, fixed_rows in class_hour_rules_by_teacher.items():
        ordered_fixed = sorted(
            fixed_rows,
            key=lambda row: (
                int(row.get("hafta_kuni") or 0),
                int(row.get("smena") or 0),
                int(row.get("dars_raqami") or 0),
                int(row.get("sinf_id") or 0),
            ),
        )
        for index, first in enumerate(ordered_fixed):
            for second in ordered_fixed[index + 1:]:
                if int(first["hafta_kuni"]) != int(second["hafta_kuni"]):
                    continue
                first_interval = shift_intervals.get((
                    int(first["smena"]), int(first["dars_raqami"])
                ))
                second_interval = shift_intervals.get((
                    int(second["smena"]), int(second["dars_raqami"])
                ))
                if not _v1866_intervals_overlap(first_interval, second_interval):
                    continue
                leader_name = first.get("rahbar_ismi") or str(teacher_id)
                errors.append(
                    f"{leader_name}: {_V1852_HAFTA.get(int(first['hafta_kuni']), first['hafta_kuni'])} "
                    f"kuni {first['sinf']}-{first['harf']} ({first['smena']}-smena, "
                    f"{first['dars_raqami']}-dars) va {second['sinf']}-{second['harf']} "
                    f"({second['smena']}-smena, {second['dars_raqami']}-dars) qat'iy "
                    "KELAJAK SOATI haqiqiy vaqtda ustma-ust"
                )

    if not model["pairs"]:
        errors.append("DARS_BIRIKMALARI varag'idan birorta ham aniq sinf–fan–soat topilmadi")

    class_summary = []
    for class_id, cls in classes.items():
        grade = _v1874_grade(cls)
        shift = int(cls.get("smena") or 1)
        shift_periods = int(shifts.get(shift, {}).get("dars_soni") or 0)
        allowed_days = []
        for day in range(1, weekdays + 1):
            if not _v1856_class_day_block_reason(cls, day, class_day_blocks):
                allowed_days.append(day)
        fan_hours = float(model["class_hours"].get(class_id, 0))
        class_hour_rule = class_hour_by_class.get(class_id, {})
        # KELAJAK SOATI rahbar hali biriktirilmagan bo‘lsa ham sinfning
        # qat’iy jadval katagi hisoblanadi. Rahbar keyin tahrirda ulanadi.
        class_hour = int(class_hour_rule.get("haftalik_soat") or 0)
        if 1 <= grade <= 4:
            base_per_day = min(4, shift_periods)
            fifth_extra = min(_v1874_fifth_day_limit(grade), len(allowed_days)) if shift_periods >= 5 else 0
            academic_capacity = base_per_day * len(allowed_days) + fifth_extra
            physical_capacity = min(
                _v1874_max_total_periods(grade), shift_periods
            ) * len(allowed_days)
            # Boshlang'ichda akademik maksimumdan tashqari bo'sh 5-darsni
            # KELAJAK/SINF SOATI egallashi mumkin. U yangi fizik katak
            # yaratmaydi: 24 fan + 1 sinf soati 25 katakka sig'adi, ammo
            # 25 fan + 1 sinf soati sig'maydi.
            capacity = min(
                physical_capacity, academic_capacity + class_hour
            )
        else:
            academic_capacity = min(6, shift_periods) * len(allowed_days)
            physical_capacity = academic_capacity
            capacity = physical_capacity
        planned = fan_hours + class_hour
        if fan_hours <= 0:
            errors.append(f"{cls['sinf']}-{cls['harf']}: haftalik fan soatlari kiritilmagan")
        if shift not in shifts:
            errors.append(f"{cls['sinf']}-{cls['harf']}: {shift}-smena sozlanmagan")
        if fan_hours > academic_capacity:
            errors.append(
                f"{cls['sinf']}-{cls['harf']}: fanlar rejasi {fan_hours} soat, akademik sig'im {academic_capacity} soat"
            )
        if planned > capacity:
            errors.append(
                f"{cls['sinf']}-{cls['harf']}: fanlar+sinf soati {planned} ta sessiya, mavjud sig'im {capacity} ta"
            )
        weekly_max = _v1874_weekly_max(grade)
        if weekly_max is not None and fan_hours > weekly_max:
            errors.append(
                f"{cls['sinf']}-{cls['harf']}: fanlar rejasi {fan_hours} soat, me'yoriy maksimum {weekly_max} soat"
            )
        rule = class_hour_by_class.get(class_id)
        if rule:
            blocked = _v1856_class_day_block_reason(cls, int(rule["hafta_kuni"]), class_day_blocks)
            if blocked:
                errors.append(f"{cls['sinf']}-{cls['harf']} sinf soati: {blocked}")
            if int(rule["dars_raqami"]) > min(shift_periods, _v1874_max_total_periods(grade)):
                errors.append(
                    f"{cls['sinf']}-{cls['harf']} sinf soati: {rule['dars_raqami']}-dars sinf limitidan tashqari"
                )
        class_summary.append({
            "sinf_id": class_id, "sinf": f"{cls['sinf']}-{cls['harf']}",
            "smena": shift, "fan_soati": fan_hours, "sinf_soati": class_hour,
            "reja_jami": planned, "sigim": capacity,
            "farq": capacity - planned, "oqish_kunlari": len(allowed_days),
            "mos": planned <= capacity,
        })

    for key, pair in model["pairs"].items():
        cls = classes.get(pair["sinf_id"], {})
        grade = _v1874_grade(cls)
        allowed_days = sum(
            1 for day in range(1, weekdays + 1)
            if not _v1856_class_day_block_reason(cls, day, class_day_blocks)
        )
        daily_max = 2
        weekly_hours = float(pair.get("haftalik_soat") or 0)
        profile = _v1874_subject_profile(pair.get("fan_nomi"), grade)
        practical = bool(profile.get("physical") or profile.get("technology"))
        per_day_limit = 2
        repeat_day_limit = 1 if practical else 2
        # Exact kontrakt bilan aynan bir xil sig'im: har kuni avval bittadan,
        # manbada daily_max>1 bo'lsagina ko'pi bilan 1/2 ta kunda qo'shimcha
        # dars. Masalan 5 kun va daily_max=2 => oddiy fan 7, amaliy fan 6.
        extra_per_repeat_day = max(0, per_day_limit - 1)
        regular_capacity = allowed_days + (
            min(allowed_days, repeat_day_limit) * extra_per_repeat_day
        )
        # Preflight, exact model va yakuniy SQL validator bitta kontraktda:
        # kunlik_max=1 bo'lsa bir fan bir kunda faqat bir marta. Sig'im bundan
        # oshsa generatorga yuborib oxirida 108 ta xato chiqarish o'rniga shu
        # yerning o'zida bitta aniq manba xatosi qaytariladi.
        if weekly_hours > regular_capacity:
            errors.append(
                f"{pair['sinf']} / {pair['fan_nomi']}: {weekly_hours:g} soatni "
                f"{allowed_days} kunga kunlik max {daily_max} va takroriy "
                f"kun limiti {repeat_day_limit} bilan "
                "ushbu qoida bilan joylab bo'lmaydi; kunlik maksimumni yoki o'qish "
                "kunlarini ongli ravishda tahrirlang"
            )

    cur.execute("SELECT * FROM aqlli_oqituvchi_qoidalari_v2 WHERE maktab_id=%s", (maktab_id,))
    rules = _v1852_teacher_rules_map(cur.fetchall())
    cur.execute("SELECT * FROM aqlli_oqituvchi_vaqti_v2 WHERE maktab_id=%s", (maktab_id,))
    hard, soft, method_hard, method_soft = _v1852_availability_maps(cur.fetchall())
    cur.execute("SELECT user_id,full_name,haftalik_dars_soati FROM users WHERE maktab_id=%s", (maktab_id,))
    teacher_rows = {int(row["user_id"]): dict(row) for row in cur.fetchall()}

    # Pair-level common-domain preflight.  Old report class capacity and each
    # teacher capacity separately checked, but never asked whether the class
    # and every parallel-group teacher share the *same* legal day/period.  A
    # 5-hour subject could therefore show "0 xato" and become instant UNSAT
    # after a method day was intersected with the class calendar.
    default_pair_rules = {
        "kunlik_max": 6, "ketma_ket_max": 4, "okno_max": 1,
        "afzal_smena": 0, "eng_erta_dars": 1, "eng_kech_dars": 12,
    }
    for pair in model["pairs"].values():
        class_row = classes.get(int(pair["sinf_id"]), {})
        grade = _v1874_grade(class_row)
        shift = int(pair.get("smena") or class_row.get("smena") or 1)
        shift_periods = int(shifts.get(shift, {}).get("dars_soni") or 0)
        max_period = min(shift_periods, _v1874_max_total_periods(grade))
        profile = _v1874_subject_profile(pair.get("fan_nomi"), grade)
        daily_max = 2
        practical = bool(profile.get("physical") or profile.get("technology"))
        per_day_limit = 2
        repeat_day_limit = 1 if practical else 2
        first_period = 2 if (
            practical
        ) else 1
        teacher_ids = sorted({
            int(value) for value in pair.get("oqituvchilar") or []
            if value is not None
        })
        common_day_capacities = []
        for day in range(1, weekdays + 1):
            if _v1856_class_day_block_reason(
                class_row, day, class_day_blocks
            ):
                continue
            # Exact adapter currently keeps the seeded primary-school
            # Saturday policy hard as well; preflight must use the same rule.
            if 1 <= grade <= 4 and day == 6:
                continue
            if any((teacher_id, day) in method_hard for teacher_id in teacher_ids):
                continue
            common_periods = []
            for period in range(first_period, max_period + 1):
                legal_for_all = True
                for teacher_id in teacher_ids:
                    teacher_rule = rules.get(
                        teacher_id, default_pair_rules
                    ) or default_pair_rules
                    if (
                        period < int(teacher_rule.get("eng_erta_dars") or 1)
                        or period > int(teacher_rule.get("eng_kech_dars") or 12)
                        or _v1852_blocked(
                            hard, teacher_id, day, shift, period
                        )
                    ):
                        legal_for_all = False
                        break
                if legal_for_all:
                    common_periods.append(period)
            if common_periods:
                day_capacity = min(per_day_limit, len(common_periods))
                if practical and day_capacity >= 2:
                    has_adjacent_pair = any(
                        int(right) - int(left) == 1
                        for left, right in zip(
                            common_periods, common_periods[1:]
                        )
                    )
                    if not has_adjacent_pair:
                        day_capacity = 1
                common_day_capacities.append((day, day_capacity))

        common_days = [day for day, _capacity in common_day_capacities]
        repeat_extras = sorted(
            (
                max(0, int(day_capacity) - 1)
                for _day, day_capacity in common_day_capacities
            ),
            reverse=True,
        )
        common_capacity = len(common_days) + sum(
            repeat_extras[:repeat_day_limit]
        )
        weekly_sessions = float(pair.get("haftalik_soat") or 0)
        if weekly_sessions > common_capacity + 1e-9:
            teacher_names = ", ".join(
                str(teacher_rows.get(teacher_id, {}).get("full_name") or teacher_id)
                for teacher_id in teacher_ids
            ) or "o'qituvchi biriktirilmagan"
            strict_shortage = weekly_sessions - common_capacity
            bounded_primary_candidate = bool(
                auto_method_relaxation
                and 1 <= grade <= 4
                and strict_shortage <= 2 + 1e-9
                and any(
                    day != 6
                    and any(
                        (teacher_id, day) in method_hard
                        for teacher_id in teacher_ids
                    )
                    for day in range(1, weekdays + 1)
                )
            )
            message = (
                f"V22.40 DAILY-BALANCE · {pair['sinf']} / {pair['fan_nomi']}: sinf va "
                f"{teacher_names} uchun strict umumiy legal kun "
                f"{len(common_days)} ta; sig'im {common_capacity:g}, reja "
                f"{weekly_sessions:g}."
            )
            if bounded_primary_candidate:
                warnings.append(
                    message
                    + " Yagona exact generator qizil/BANDni ochmasdan "
                    "boshlang'ich o'qituvchining ko'pi bilan 2 ta aniq "
                    "metod-kuni katagi bilan to'liq yechimni tekshiradi."
                )
            else:
                errors.append(
                    message
                    + " Qizil yoki metod kunini taxminan ochmang; aynan "
                    "shu fan/o'qituvchi vaqtini tahrirlang."
                )
        elif weekly_sessions > len(common_days) and per_day_limit > 1:
            extra = weekly_sessions - len(common_days)
            minimum_repeat_days = int(math.ceil(
                extra / max(1, per_day_limit - 1)
            ))
            warnings.append(
                f"{pair['sinf']} / {pair['fan_nomi']}: mavjud umumiy kunlar "
                f"sabab kamida {minimum_repeat_days} kunda fan takrori kerak; "
                "generator avval boshqa fan bilan kunlararo almashtirishni "
                "sinaydi."
            )

    teacher_shifts = _v1852_defaultdict(set)
    for pair in model["pairs"].values():
        for teacher_id in pair.get("oqituvchilar", []):
            teacher_shifts[int(teacher_id)].add(int(pair.get("smena") or 1))
    for teacher_id, fixed_rows in class_hour_rules_by_teacher.items():
        for fixed_row in fixed_rows:
            teacher_shifts[int(teacher_id)].add(int(fixed_row.get("smena") or 1))

    teacher_summary = []
    all_teacher_ids = sorted(set(model["teacher_hours"]) | set(class_hour_by_teacher))
    primary_method_fallback_teachers = {
        int(teacher_id)
        for pair in model["pairs"].values()
        if 1 <= _v1874_grade(classes.get(int(pair["sinf_id"]), {})) <= 4
        for teacher_id in pair.get("oqituvchilar") or []
    }
    default_rules = {"kunlik_max": 6, "ketma_ket_max": 4, "okno_max": 1,
                     "afzal_smena": 0, "eng_erta_dars": 1, "eng_kech_dars": 12}
    for teacher_id in all_teacher_ids:
        row = teacher_rows.get(int(teacher_id), {"full_name": str(teacher_id)})
        base_plan = float(model["teacher_hours"].get(int(teacher_id), 0))
        class_hours = int(class_hour_by_teacher.get(int(teacher_id), 0))
        total_plan = base_plan + class_hours
        saved_base = row.get("haftalik_dars_soati")
        if saved_base is None or abs(float(saved_base) - base_plan) > 1e-9:
            errors.append(
                f"{row.get('full_name')}: shablon fan yuklamasi {base_plan} soat, xodim kartasida {saved_base if saved_base is not None else 'bo\'sh'}"
            )
        teacher_rules = rules.get(int(teacher_id), default_rules)
        shifts_used = teacher_shifts.get(int(teacher_id), set()) or set(shifts.keys())
        for fixed_row in class_hour_rules_by_teacher.get(int(teacher_id), []):
            fixed_period = int(fixed_row.get("dars_raqami") or 0)
            if (
                fixed_period < int(teacher_rules["eng_erta_dars"])
                or fixed_period > int(teacher_rules["eng_kech_dars"])
            ):
                errors.append(
                    f"{row.get('full_name')}: {fixed_row['sinf']}-{fixed_row['harf']} "
                    f"KELAJAK SOATI {fixed_period}-darsga qat'iy tanlangan, lekin "
                    f"o'qituvchi oralig'i {teacher_rules['eng_erta_dars']}–"
                    f"{teacher_rules['eng_kech_dars']}-dars"
                )
            fixed_day = int(fixed_row.get("hafta_kuni") or 0)
            fixed_shift = int(fixed_row.get("smena") or 1)
            if _v1852_blocked(
                hard,
                int(teacher_id),
                fixed_day,
                fixed_shift,
                fixed_period,
            ):
                errors.append(
                    f"{row.get('full_name')}: {fixed_row['sinf']}-{fixed_row['harf']} "
                    f"KELAJAK SOATI {_V1852_HAFTA.get(fixed_day, fixed_day)} "
                    f"kuni {fixed_period}-darsga tanlangan, ammo bu vaqt "
                    "o'qituvchida qizil/BAND. Qat’iy rejim bu katakni "
                    "avtomatik ochmaydi; vaqtni manbada tahrirlang"
                )
            if (
                not auto_method_relaxation
                and (int(teacher_id), fixed_day) in method_hard
            ):
                errors.append(
                    f"{row.get('full_name')}: {fixed_row['sinf']}-{fixed_row['harf']} "
                    f"KELAJAK SOATI {_V1852_HAFTA.get(fixed_day, fixed_day)} "
                    "metod kuniga tanlangan. Qat’iy rejim metod kunini "
                    "avtomatik ochmaydi; kun yoki darsni manbada tahrirlang"
                )
        capacity = 0
        fixed_exception_capacity = 0
        for day in range(1, weekdays + 1):
            open_intervals = []
            method_day = (int(teacher_id), day) in method_hard
            if not method_day:
                for shift in shifts_used:
                    count = int(shifts.get(int(shift), {}).get("dars_soni") or 0)
                    for period in range(1, count + 1):
                        if period < teacher_rules["eng_erta_dars"] or period > teacher_rules["eng_kech_dars"]:
                            continue
                        if _v1852_blocked(hard, int(teacher_id), day, int(shift), period):
                            continue
                        interval = shift_intervals.get((int(shift), period))
                        if interval:
                            open_intervals.append(interval)

            open_slots = _v220_max_nonoverlapping_interval_count(open_intervals)

            # Faqat siyosat aniq ruxsat bersagina oldindan tanlangan KELAJAK
            # SOATI metod kunidan istisno bo'lishi mumkin. V23 strict rejimida
            # bu qiymat 0 qoladi; qizil/BAND esa har doim yopiq.
            exceptional_fixed = 0
            if auto_method_relaxation:
                for fixed_row in class_hour_rules_by_teacher.get(int(teacher_id), []):
                    if int(fixed_row.get("hafta_kuni") or 0) != day:
                        continue
                    fixed_shift = int(fixed_row.get("smena") or 1)
                    fixed_period = int(fixed_row.get("dars_raqami") or 0)
                    if (
                        fixed_period < int(teacher_rules["eng_erta_dars"])
                        or fixed_period > int(teacher_rules["eng_kech_dars"])
                    ):
                        continue
                    if method_day and not _v1852_blocked(
                        hard, int(teacher_id), day, fixed_shift, fixed_period
                    ):
                        exceptional_fixed += 1
            daily_normal = min(open_slots, int(teacher_rules["kunlik_max"]))
            daily_total = min(
                open_slots + exceptional_fixed,
                int(teacher_rules["kunlik_max"]),
            )
            capacity += daily_total
            fixed_exception_capacity += max(0, daily_total - daily_normal)
        strict_shortage = max(0.0, total_plan - capacity)
        bounded_primary_candidate = bool(
            auto_method_relaxation
            and strict_shortage <= 2 + 1e-9
            and int(teacher_id) in primary_method_fallback_teachers
            and any(
                day != 6 and (int(teacher_id), day) in method_hard
                for day in range(1, weekdays + 1)
            )
        )
        if strict_shortage > 0:
            message = (
                f"{row.get('full_name')}: haftalik reja {total_plan:g} soat, "
                f"strict yashil vaqtlar va qat'iy KELAJAK SOATI bo'yicha "
                f"sig'im {capacity:g} soat"
            )
            if bounded_primary_candidate:
                warnings.append(
                    message
                    + "; exact generator boshlang'ich sinf uchun ko'pi bilan "
                    "2 ta aniq metod-kuni istisnosini qizil/BANDsiz tekshiradi."
                )
            else:
                errors.append(message)
        teacher_summary.append({
            "user_id": int(teacher_id), "full_name": row.get("full_name"),
            "fan_yuklama": base_plan, "sinf_soati": class_hours,
            "reja_jami": total_plan, "saqlangan_yuklama": saved_base,
            "qattiq_sigim": capacity,
            "kelajak_soati_istisno_sigimi": fixed_exception_capacity,
            "farq": capacity - total_plan,
            "strict_mos": total_plan <= capacity,
            "metod_istisnosi_bilan_tekshiriladi": bounded_primary_candidate,
            "mos": (
                (total_plan <= capacity or bounded_primary_candidate)
                and saved_base is not None
                and abs(float(saved_base) - base_plan) <= 1e-9
            ),
        })

    fan_summary = [
        {"sinf_id": pair["sinf_id"], "sinf": pair["sinf"],
         "fan": pair["fan_nomi"], "haftalik_soat": pair["haftalik_soat"],
         "kunlik_max": pair["kunlik_max"], "turi": pair["turi"],
         "oqituvchi_soni": len(pair.get("oqituvchilar", []))}
        for pair in model["pairs"].values()
    ]

    source_hash = _v1875_source_fingerprint(cur, maktab_id)
    unique_errors = list(dict.fromkeys(errors))
    unique_warnings = list(dict.fromkeys(warnings))
    return {
        "tayyor": not unique_errors,
        # Bu lokal manba/capacity tekshiruvi. U global jadval mavjudligini
        # isbotlamaydi; shuning uchun frontend endi uni soxta "100%" deb
        # ko'rsatmaydi.
        "solver_status": "INPUT_VALID" if not unique_errors else "INPUT_INVALID",
        "global_yechim_isbotlandi": False,
        "manba_hash": source_hash,
        "xatolar": unique_errors[:300],
        "ogohlantirishlar": unique_warnings[:300],
        "sinflar": class_summary,
        "oqituvchilar": teacher_summary,
        "fanlar": fan_summary,
        "xulosa": {
            "sinf_soni": len(class_summary),
            "oqituvchi_soni": len(teacher_summary),
            "fan_sinf_soni": len(fan_summary),
            "xato_soni": len(unique_errors),
        },
    }


def _v1875_schedule_integrity_report(cur, maktab_id: int, run_id: int):
    _v1875_tables(cur)
    model = _v1875_exact_assignment_model(cur, maktab_id)
    errors = list(model["xatolar"])
    warnings = list(model["ogohlantirishlar"])
    cur.execute("SELECT * FROM aqlli_jadval_urinishlari_v2 WHERE id=%s AND maktab_id=%s",
                (run_id, maktab_id))
    run = cur.fetchone()
    if not run:
        return {"tayyor": False, "xatolar": ["Draft topilmadi"],
                "ogohlantirishlar": [], "sinflar": [], "oqituvchilar": [], "fanlar": []}

    settings = run.get("sozlamalar") or {}
    approved_method_exceptions = set()
    if isinstance(settings, dict):
        for raw in settings.get("metod_kuni_istisnolari") or []:
            try:
                approved_method_exceptions.add((
                    int(raw.get("oqituvchi_id") or 0),
                    int(raw.get("kun") or 0),
                    int(raw.get("smena") or 0),
                    int(raw.get("dars") or 0),
                ))
            except (AttributeError, TypeError, ValueError):
                continue
    current_hash = _v1875_source_fingerprint(cur, maktab_id)
    stored_hash = settings.get("manba_hash") if isinstance(settings, dict) else None
    if stored_hash and stored_hash != current_hash:
        errors.append("Draft yaratilgandan keyin shablon, fan soati yoki vaqt qoidalari o'zgargan — yangi draft kerak")

    cur.execute("""SELECT e.*,s.sinf,s.harf,COALESCE(s.smena,1) AS sinf_smena,
                          u.full_name AS oqituvchi_ismi
                   FROM aqlli_jadval_slotlari_v2 e
                   JOIN maktab_sinflari s ON s.id=e.sinf_id
                   LEFT JOIN users u ON u.user_id=e.oqituvchi_user_id
                   WHERE e.maktab_id=%s AND e.urinish_id=%s
                   ORDER BY e.sinf_id,e.hafta_kuni,e.smena,e.dars_raqami,e.guruh_kaliti""",
                (maktab_id, run_id))
    slots = [dict(row) for row in cur.fetchall()]

    planned_subject = {
        key: float(pair["haftalik_soat"])
        for key, pair in model["pairs"].items()
    }
    scheduled_subject_sessions = _v1852_defaultdict(set)
    class_sessions = _v1852_defaultdict(set)
    class_hour_sessions = _v1852_defaultdict(set)
    teacher_sessions = _v1852_defaultdict(set)
    teacher_class_hour_sessions = _v1852_defaultdict(set)
    pair_occurrence_rows = _v1852_defaultdict(list)
    teacher_slot_map = _v1852_defaultdict(set)
    teacher_time_map = _v1852_defaultdict(list)
    room_time_map = _v1852_defaultdict(list)

    cur.execute("SELECT * FROM aqlli_oqituvchi_qoidalari_v2 WHERE maktab_id=%s", (maktab_id,))
    teacher_rules = _v1852_teacher_rules_map(cur.fetchall())
    cur.execute("SELECT * FROM aqlli_oqituvchi_vaqti_v2 WHERE maktab_id=%s", (maktab_id,))
    hard, soft, method_hard, method_soft = _v1852_availability_maps(cur.fetchall())
    class_day_blocks = _v1856_class_day_rule_map(_v1856_class_day_rule_rows(cur, maktab_id))
    cur.execute("SELECT id,sinf,harf,COALESCE(smena,1) AS smena FROM maktab_sinflari WHERE maktab_id=%s", (maktab_id,))
    classes = {int(row["id"]): dict(row) for row in cur.fetchall()}
    cur.execute("SELECT id,nomi FROM aqlli_xonalar_v2 WHERE maktab_id=%s", (maktab_id,))
    integrity_rooms = [dict(row) for row in cur.fetchall()]
    room_name_to_id = {
        _v205_room_normalized_name(row.get("nomi")): int(row["id"])
        for row in integrity_rooms if _v205_room_normalized_name(row.get("nomi"))
    }
    room_labels = {
        f"room:{int(row['id'])}": str(row.get("nomi") or row["id"])
        for row in integrity_rooms
    }
    cur.execute(
        "SELECT * FROM aqlli_smena_sozlamalari_v2 WHERE maktab_id=%s",
        (maktab_id,),
    )
    integrity_slot_intervals = {}
    for shift_row in cur.fetchall():
        shift_number = int(shift_row["smena"])
        for time_row in _v1852_shift_slots(shift_row):
            integrity_slot_intervals[(
                shift_number, int(time_row["dars_raqami"])
            )] = (
                _v196_clock_minutes(time_row.get("boshlanish")),
                _v196_clock_minutes(time_row.get("tugash")),
            )
    cur.execute("SELECT * FROM aqlli_sinf_fan_yuklamalari_v2 WHERE maktab_id=%s AND haftalik_soat>0", (maktab_id,))
    loads = {(int(row["sinf_id"]), _v1875_subject_key(row["fan_nomi"])): dict(row) for row in cur.fetchall()}
    integrity_class_hour_rules = _v1866_class_hour_rule_rows(cur, maktab_id)
    class_hour_subject_by_class = {
        int(row["sinf_id"]): _v1875_subject_key(
            row.get("fan_nomi") or "KELAJAK SOATI"
        )
        for row in integrity_class_hour_rules
    }
    class_hour_fixed_by_class = {
        int(row["sinf_id"]): (
            int(row.get("hafta_kuni") or 0),
            int(row.get("smena") or 1),
            int(row.get("dars_raqami") or 0),
        )
        for row in integrity_class_hour_rules
    }

    for slot in slots:
        class_id = int(slot["sinf_id"])
        day = int(slot["hafta_kuni"])
        shift = int(slot["smena"])
        period = int(slot["dars_raqami"])
        week_type = str(slot.get("hafta_turi") or "har_hafta")
        subject = str(slot.get("fan_nomi") or "").strip()
        subject_key = _v1875_subject_key(subject)
        session = (day, shift, period)
        class_sessions[class_id].add(session)
        is_class_hour = subject_key in {
            _v1875_subject_key("SINF SOATI"),
            _v1875_subject_key("KELAJAK SOATI"),
            class_hour_subject_by_class.get(class_id),
        }
        if is_class_hour:
            class_hour_sessions[class_id].add((*session, week_type))
        else:
            pair_key = (class_id, subject_key)
            scheduled_subject_sessions[pair_key].add((*session, week_type))
            pair_occurrence_rows[(pair_key, int(slot.get("takror_raqami") or 1))].append(slot)
            pair = model["pairs"].get(pair_key)
            if not pair:
                errors.append(f"{classes.get(class_id, {}).get('sinf','')}-{classes.get(class_id, {}).get('harf','')} / {subject}: reja manbasida yo'q dars")
            else:
                group_key = _v1875_group_key(slot.get("guruh_kaliti"))
                expected_teacher = None
                if pair["turi"] == "whole":
                    expected_teacher = pair.get("asosiy_oqituvchi_user_id")
                    if group_key != "whole":
                        errors.append(f"{pair['sinf']} / {pair['fan_nomi']}: butun sinf darsi {group_key} guruh bilan yozilgan")
                else:
                    expected = {g["guruh_kaliti"]: g["oqituvchi_user_id"] for g in pair["guruhlar"]}
                    expected_teacher = expected.get(group_key)
                    if group_key not in expected:
                        errors.append(f"{pair['sinf']} / {pair['fan_nomi']}: kutilmagan guruh {group_key}")
                actual_teacher = slot.get("oqituvchi_user_id")
                if expected_teacher is not None and int(actual_teacher or 0) != int(expected_teacher):
                    errors.append(
                        f"{pair['sinf']} / {pair['fan_nomi']} / {group_key}: reja o'qituvchisi {expected_teacher}, jadvalda {actual_teacher}"
                    )

        teacher_id = slot.get("oqituvchi_user_id")
        if teacher_id is not None:
            teacher_id = int(teacher_id)
            teacher_sessions[teacher_id].add((class_id, day, shift, period, week_type))
            if is_class_hour:
                teacher_class_hour_sessions[teacher_id].add(
                    (class_id, day, shift, period, week_type)
                )
            teacher_slot_map[(teacher_id, day, shift, period)].add((class_id, week_type))
            canonical_interval = integrity_slot_intervals.get((shift, period))
            teacher_time_map[(teacher_id, day)].append({
                "slot_id": int(slot.get("id") or 0),
                "sinf_id": class_id,
                "smena": shift,
                "dars_raqami": period,
                "hafta_turi": week_type,
                "boshlanish": (
                    canonical_interval[0] if canonical_interval
                    else _v196_clock_minutes(slot.get("boshlanish_vaqti"))
                ),
                "tugash": (
                    canonical_interval[1] if canonical_interval
                    else _v196_clock_minutes(slot.get("tugash_vaqti"))
                ),
            })
            # Faqat metod-kuni istisnosi fan nomiga emas, administrator
            # tanlagan qat'iy kun+smena+dars katagiga tegishli. Qizil/BAND
            # fixed Sinf soatida ham qat'iy yopiq qoladi.
            fixed_class_hour = bool(
                is_class_hour
                and class_hour_fixed_by_class.get(class_id) == session
            )
            approved_primary_method_exception = bool(
                (teacher_id, day, shift, period)
                in approved_method_exceptions
                and 1 <= _v1874_grade(classes.get(class_id, {})) <= 4
                and day != 6
            )
            if (
                (teacher_id, day) in method_hard
                and not fixed_class_hour
                and not approved_primary_method_exception
            ):
                errors.append(f"{slot.get('oqituvchi_ismi')}: {_V1852_HAFTA.get(day, day)} metod kuniga dars qo'yilgan")
            if _v1852_blocked(hard, teacher_id, day, shift, period):
                errors.append(f"{slot.get('oqituvchi_ismi')}: {_V1852_HAFTA.get(day, day)} {shift}-smena {period}-dars qattiq bloklangan")
        room_key = _v205_persisted_room_key(
            slot.get("xona_id"), slot.get("xona_matni"), room_name_to_id
        )
        if room_key:
            room_labels.setdefault(
                room_key, str(slot.get("xona_matni") or room_key)
            )
            canonical_interval = integrity_slot_intervals.get((shift, period))
            room_time_map[(room_key, day)].append({
                "slot_id": int(slot.get("id") or 0),
                "sinf_id": class_id,
                "smena": shift,
                "dars_raqami": period,
                "hafta_turi": week_type,
                "boshlanish": (
                    canonical_interval[0] if canonical_interval
                    else _v196_clock_minutes(slot.get("boshlanish_vaqti"))
                ),
                "tugash": (
                    canonical_interval[1] if canonical_interval
                    else _v196_clock_minutes(slot.get("tugash_vaqti"))
                ),
            })
        cls = classes.get(class_id, {})
        if shift != int(cls.get("smena") or 1):
            errors.append(f"{cls.get('sinf','')}-{cls.get('harf','')}: dars {shift}-smenaga tushgan, sinf smenasi {cls.get('smena')}")
        blocked = _v1856_class_day_block_reason(cls, day, class_day_blocks)
        if blocked:
            errors.append(f"{cls.get('sinf','')}-{cls.get('harf','')}: {blocked}")

    fan_summary = []
    all_pair_keys = set(planned_subject) | set(scheduled_subject_sessions)
    for key in sorted(all_pair_keys, key=lambda x: (x[0], x[1])):
        pair = model["pairs"].get(key)
        planned = float(planned_subject.get(key, 0))
        actual = round(sum(
            0.5 if session[3] in {"toq", "juft"} else 1.0
            for session in scheduled_subject_sessions.get(key, set())
        ), 1)
        class_label = pair["sinf"] if pair else f"sinf_id={key[0]}"
        fan = pair["fan_nomi"] if pair else key[1]
        if planned != actual:
            errors.append(f"{class_label} / {fan}: reja {planned} soat, jadval {actual} soat")
        fan_summary.append({"sinf": class_label, "fan": fan, "reja": planned,
                            "jadval": actual, "farq": actual - planned, "mos": planned == actual})

    for key, pair in model["pairs"].items():
        if pair["turi"] != "group":
            continue
        expected_groups = {g["guruh_kaliti"] for g in pair["guruhlar"]}
        for occurrence in range(1, int(math.ceil(float(pair["haftalik_soat"]))) + 1):
            rows = pair_occurrence_rows.get((key, occurrence), [])
            actual_groups = {_v1875_group_key(row.get("guruh_kaliti")) for row in rows}
            sessions = {(int(row["hafta_kuni"]), int(row["smena"]), int(row["dars_raqami"])) for row in rows}
            if actual_groups != expected_groups or len(sessions) != 1:
                errors.append(
                    f"{pair['sinf']} / {pair['fan_nomi']} / {occurrence}-takror: parallel guruhlar to'liq va bir vaqtda emas"
                )

    class_hour_rules = integrity_class_hour_rules
    class_hour_by_teacher = _v1852_Counter()
    for row in class_hour_rules:
        if row.get("rahbar_user_id") is not None:
            class_hour_by_teacher[int(row["rahbar_user_id"])] += int(row.get("haftalik_soat") or 1)
    class_hour_by_class = {int(row["sinf_id"]): row for row in class_hour_rules}
    class_summary = []
    all_class_ids = set(classes) | set(model["class_hours"]) | set(class_sessions)
    for class_id in sorted(all_class_ids):
        cls = classes.get(class_id, {})
        class_pairs = [pair for pair in model["pairs"].values() if int(pair["sinf_id"]) == class_id]
        full_sessions = sum(int(math.floor(float(pair.get("haftalik_soat") or 0))) for pair in class_pairs)
        half_sessions = sum(
            1 for pair in class_pairs
            if abs(float(pair.get("haftalik_soat") or 0) % 1 - 0.5) < 1e-9
        )
        base_planned = full_sessions + int(math.ceil(half_sessions / 2))
        class_hour_planned = int(class_hour_by_class.get(class_id, {}).get("haftalik_soat") or 0)
        planned = base_planned + class_hour_planned
        actual = len(class_sessions.get(class_id, set()))
        class_hour_actual = round(sum(
            0.5 if session[3] in {"toq", "juft"} else 1.0
            for session in class_hour_sessions.get(class_id, set())
        ), 1)
        subject_actual = round(float(actual) - float(class_hour_actual), 1)
        label = f"{cls.get('sinf','')}-{cls.get('harf','')}"
        if planned != actual:
            errors.append(
                f"{label}: {base_planned} soat fan yuklamasi + "
                f"{class_hour_planned} soat SINF SOATI = {planned} soat reja; "
                f"jadvalda {actual} soat"
            )
        class_summary.append({
            "sinf_id": class_id, "sinf": label,
            "fan_yuklama": base_planned,
            "sinf_soati_reja": class_hour_planned,
            "reja": planned,
            "fan_jadval": subject_actual,
            "sinf_soati_jadval": class_hour_actual,
            "jadval": actual, "farq": actual - planned,
            "mos": planned == actual,
            "tasdiq_matni": (
                f"{base_planned} soat fan yuklamasi + {class_hour_planned} soat "
                f"SINF SOATI = {planned}; jadvalda {actual} soat"
            ),
        })

    cur.execute("SELECT user_id,full_name,haftalik_dars_soati FROM users WHERE maktab_id=%s", (maktab_id,))
    teacher_rows = {int(row["user_id"]): dict(row) for row in cur.fetchall()}
    teacher_summary = []
    all_teacher_ids = set(model["teacher_hours"]) | set(class_hour_by_teacher) | set(teacher_sessions)
    for teacher_id in sorted(all_teacher_ids):
        row = teacher_rows.get(teacher_id, {"full_name": str(teacher_id)})
        base_planned = float(model["teacher_hours"].get(teacher_id, 0))
        class_hour_planned = int(class_hour_by_teacher.get(teacher_id, 0))
        planned = base_planned + class_hour_planned
        actual = round(sum(
            0.5 if session[4] in {"toq", "juft"} else 1.0
            for session in teacher_sessions.get(teacher_id, set())
        ), 1)
        class_hour_actual = round(sum(
            0.5 if session[4] in {"toq", "juft"} else 1.0
            for session in teacher_class_hour_sessions.get(teacher_id, set())
        ), 1)
        subject_actual = round(actual - class_hour_actual, 1)
        if planned != actual:
            errors.append(
                f"{row.get('full_name')}: {base_planned:g} soat fan yuklamasi + "
                f"{class_hour_planned} soat sinf rahbarligi = {planned:g} soat reja; "
                f"jadvalda {actual:g} soat"
            )
        teacher_summary.append({
            "user_id": teacher_id, "full_name": row.get("full_name"),
            "fan_yuklama": base_planned,
            "sinf_soati_reja": class_hour_planned,
            "reja": planned,
            "fan_jadval": subject_actual,
            "sinf_soati_jadval": class_hour_actual,
            "jadval": actual,
            "farq": actual - planned, "mos": planned == actual,
            "tasdiq_matni": (
                f"{base_planned:g} soat fan yuklamasi + {class_hour_planned} soat "
                f"sinf rahbarligi = {planned:g}; jadvalda {actual:g} soat"
            ),
        })

    for (teacher_id, day, shift, period), class_phases in teacher_slot_map.items():
        conflicts = {
            class_id for class_id, phase in class_phases
            if any(
                other_class != class_id
                and (phase == "har_hafta" or other_phase == "har_hafta" or phase == other_phase)
                for other_class, other_phase in class_phases
            )
        }
        if len(conflicts) > 1:
            errors.append(
                f"{teacher_rows.get(teacher_id, {}).get('full_name', teacher_id)}: "
                f"{_V1852_HAFTA.get(day, day)} {shift}-smena {period}-darsda {len(conflicts)} ta sinf"
            )
    # Dars raqami boshqa bo'lsa ham, 1- va 2-smena haqiqiy vaqtda ustma-ust
    # tushishi mumkin. Generator buni nomzod tanlashda tekshiradi; bu qatlam
    # esa saqlash oldidan phase-aware mustaqil himoya bo'lib xizmat qiladi.
    for (teacher_id, day), teacher_rows_for_day in teacher_time_map.items():
        for index, first in enumerate(teacher_rows_for_day):
            for second in teacher_rows_for_day[index + 1:]:
                first_phase = str(first.get("hafta_turi") or "har_hafta")
                second_phase = str(second.get("hafta_turi") or "har_hafta")
                phases_overlap = (
                    first_phase == "har_hafta"
                    or second_phase == "har_hafta"
                    or first_phase == second_phase
                )
                if not phases_overlap:
                    continue
                first_start, first_end = first.get("boshlanish"), first.get("tugash")
                second_start, second_end = second.get("boshlanish"), second.get("tugash")
                if None in (first_start, first_end, second_start, second_end):
                    overlaps = (
                        int(first["smena"]) == int(second["smena"])
                        and int(first["dars_raqami"]) == int(second["dars_raqami"])
                    )
                else:
                    overlaps = (
                        int(first_start) < int(second_end)
                        and int(second_start) < int(first_end)
                    )
                if not overlaps:
                    continue
                errors.append(
                    f"{teacher_rows.get(teacher_id, {}).get('full_name', teacher_id)}: "
                    f"{_V1852_HAFTA.get(day, day)} kuni haqiqiy vaqtda ikki dars "
                    f"({first['smena']}-smena {first['dars_raqami']}-dars va "
                    f"{second['smena']}-smena {second['dars_raqami']}-dars)"
                )
    for (room_key, day), room_rows in room_time_map.items():
        for index, first in enumerate(room_rows):
            for second in room_rows[index + 1:]:
                first_phase = str(first.get("hafta_turi") or "har_hafta")
                second_phase = str(second.get("hafta_turi") or "har_hafta")
                phases_overlap = (
                    first_phase == "har_hafta"
                    or second_phase == "har_hafta"
                    or first_phase == second_phase
                )
                if not phases_overlap:
                    continue
                first_start, first_end = first.get("boshlanish"), first.get("tugash")
                second_start, second_end = second.get("boshlanish"), second.get("tugash")
                if None in (first_start, first_end, second_start, second_end):
                    # Eski slotda vaqt matni yo'q bo'lsa, hech bo'lmasa ayni
                    # smena+dars raqamidagi kolliziyani o'tkazib yubormaymiz.
                    overlaps = (
                        int(first["smena"]) == int(second["smena"])
                        and int(first["dars_raqami"]) == int(second["dars_raqami"])
                    )
                else:
                    overlaps = (
                        int(first_start) < int(second_end)
                        and int(second_start) < int(first_end)
                    )
                if not overlaps:
                    continue
                room_label = room_labels.get(room_key, room_key)
                if int(first["smena"]) == int(second["smena"]):
                    errors.append(
                        f"Xona {room_label}: {_V1852_HAFTA.get(day, day)} "
                        f"{first['smena']}-smena {first['dars_raqami']}-darsda ikki dars"
                    )
                else:
                    errors.append(
                        f"Xona {room_label}: {_V1852_HAFTA.get(day, day)} "
                        "ikki smenadagi dars bilan real vaqtda ustma-ust"
                    )

    daily_teacher_sessions = _v1852_defaultdict(set)
    daily_subject_sessions = _v1852_defaultdict(set)
    for slot in slots:
        teacher_id = slot.get("oqituvchi_user_id")
        if teacher_id is not None:
            daily_teacher_sessions[(
                int(teacher_id), int(slot["hafta_kuni"])
            )].add((
                int(slot["smena"]), int(slot["dars_raqami"]),
                str(slot.get("hafta_turi") or "har_hafta"),
            ))
        subject_key = _v1875_subject_key(slot.get("fan_nomi"))
        if subject_key not in {
            _v1875_subject_key("SINF SOATI"),
            _v1875_subject_key("KELAJAK SOATI"),
            class_hour_subject_by_class.get(int(slot["sinf_id"])),
        }:
            daily_subject_sessions[(
                int(slot["sinf_id"]), subject_key, int(slot["hafta_kuni"])
            )].add((
                int(slot["smena"]), int(slot["dars_raqami"]),
                str(slot.get("hafta_turi") or "har_hafta"),
            ))
    for (teacher_id, day), sessions in daily_teacher_sessions.items():
        limit = int(teacher_rules.get(teacher_id, {"kunlik_max": 6})["kunlik_max"])
        for phase in ("toq", "juft"):
            count = len({
                (session[0], session[1]) for session in sessions
                if session[2] == "har_hafta" or session[2] == phase
            })
            if count > limit:
                errors.append(
                    f"{teacher_rows.get(teacher_id, {}).get('full_name', teacher_id)}: "
                    f"{_V1852_HAFTA.get(day, day)} {phase.upper()} haftada "
                    f"{count} dars, kunlik max {limit}"
                )
    subject_repeat_days = _v1852_defaultdict(list)
    for (class_id, subject_key, day), sessions in daily_subject_sessions.items():
        load = loads.get((class_id, subject_key))
        configured_daily = 2
        cls_for_profile = classes.get(class_id, {})
        grade = _v1874_grade(cls_for_profile)
        effective_daily = 2
        subject_profile = {}
        if load:
            subject_profile = _v1874_subject_profile(
                load.get("fan_nomi"), grade
            )
        practical = bool(
            subject_profile.get("physical") or subject_profile.get("technology")
        )
        explicit_practical_double = bool(practical and effective_daily > 1)
        allowed_daily = 2

        # TOQ/JUFT — ikkita real hafta fazasi. ``har_hafta`` ikkalasida ham
        # sanaladi; shu hisob exact solverning faza kontrakti bilan bir xil.
        for phase in ("toq", "juft"):
            phase_sessions = {
                session for session in sessions
                if session[2] == "har_hafta" or session[2] == phase
            }
            count = len(phase_sessions)
            if count > 1:
                subject_repeat_days[(
                    class_id, subject_key, phase, practical
                )].append(int(day))
            if explicit_practical_double and count > 1:
                periods = sorted({int(session[1]) for session in phase_sessions})
                if len(periods) != 2 or periods[1] - periods[0] != 1:
                    cls = classes.get(class_id, {})
                    errors.append(
                        f"{cls.get('sinf','')}-{cls.get('harf','')} / "
                        f"{load['fan_nomi']}: amaliy fan juftligi "
                        f"{_V1852_HAFTA.get(day, day)} kuni {phase.upper()} "
                        "haftada yonma-yon emas"
                    )
            if load and count > allowed_daily:
                cls = classes.get(class_id, {})
                errors.append(
                    f"{cls.get('sinf','')}-{cls.get('harf','')} / {load['fan_nomi']}: "
                    f"{_V1852_HAFTA.get(day, day)} {phase.upper()} haftada "
                    f"{count} marta, kunlik max {allowed_daily}"
                )
    for (
        class_id, subject_key, phase, practical
    ), repeat_days in subject_repeat_days.items():
        unique_days = sorted(set(int(day) for day in repeat_days))
        repeat_day_limit = 1 if practical else 2
        if len(unique_days) <= repeat_day_limit:
            continue
        load = loads.get((class_id, subject_key), {})
        cls = classes.get(class_id, {})
        errors.append(
            f"{cls.get('sinf','')}-{cls.get('harf','')} / "
            f"{load.get('fan_nomi') or subject_key}: bir kunda takror "
            f"{phase.upper()} haftada {len(unique_days)} kun; maksimum "
            f"{repeat_day_limit} kun"
        )

    hygiene = _v1874_schedule_hygiene_violations(cur, maktab_id, run_id)
    for item in hygiene:
        message = f"{item['sinf']}: {item['sabab']}"
        reason = str(item.get("sabab") or "")
        if _v212_hygiene_is_soft(reason):
            warnings.append("Pedagogik zaxira · " + message)
        else:
            errors.append(message)
    for item in _v1856_schedule_block_violations(cur, maktab_id, run_id):
        errors.append(f"{item['sinf']}-{item['harf']}: {item['kun_nomi']} bloklangan kunda dars bor")
    for item in _v1866_class_hour_violations(cur, maktab_id, run_id):
        errors.append(f"Sinf soati: {item['izoh']}")

    if int(run.get("joylashtirilmadi") or 0) > 0:
        errors.append(f"{int(run.get('joylashtirilmadi') or 0)} ta dars joylashtirilmagan")

    unique_errors = list(dict.fromkeys(errors))
    unique_warnings = list(dict.fromkeys(warnings))
    return {
        "tayyor": not unique_errors,
        "manba_hash": current_hash,
        "xatolar": unique_errors[:500],
        "ogohlantirishlar": unique_warnings[:300],
        "sinflar": class_summary,
        "oqituvchilar": teacher_summary,
        "fanlar": fan_summary,
        "xulosa": {
            "sinf_mos": sum(1 for row in class_summary if row["mos"]),
            "sinf_jami": len(class_summary),
            "oqituvchi_mos": sum(1 for row in teacher_summary if row["mos"]),
            "oqituvchi_jami": len(teacher_summary),
            "fan_mos": sum(1 for row in fan_summary if row["mos"]),
            "fan_jami": len(fan_summary),
            "xato_soni": len(unique_errors),
        },
    }


@app.post("/api/maktab/aqlli_jadval/v2/moslik")
def v1875_schedule_preflight(token: str, maktab_id: int):
    actor_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor()
    try:
        _v1875_tables(cur)
        if not _v1852_manager(cur, actor_id, maktab_id):
            raise HTTPException(status_code=403, detail="Moslik tekshiruvini faqat rahbariyat bajaradi")
        _v199_ensure_class_hour_rules(cur, maktab_id, actor_id=actor_id)
        if "_v1876_group_review_report" in globals():
            group_review = _v1876_group_review_report(cur, maktab_id)
            if not group_review.get("tayyor"):
                return {
                    "tayyor": False,
                    "xatolar": group_review.get("xatolar", []),
                    "ogohlantirishlar": group_review.get("ogohlantirishlar", []),
                    "guruh_tasdiqlash": group_review,
                    "sinflar": [], "oqituvchilar": [], "fanlar": [],
                    "xulosa": {
                        "xato_soni": len(group_review.get("xatolar", [])),
                        "guruh_tasdiqlanmagan": group_review.get("xulosa", {}).get("tasdiqlanmagan", 0),
                    },
                }
        sync = _v1875_rebuild_schedule_sources(cur, maktab_id, cancel_drafts=False, reason="moslik_tekshiruvi")
        if sync.get("xatolar"):
            conn.rollback()
            return {"tayyor": False, "xatolar": sync["xatolar"],
                    "ogohlantirishlar": sync.get("ogohlantirishlar", []),
                    "sinflar": [], "oqituvchilar": [], "fanlar": [],
                    "xulosa": {"xato_soni": len(sync["xatolar"])}}
        if not sync.get("tayyor"):
            conn.rollback()
            return {
                "tayyor": False,
                "xatolar": ["Barcha faol ta'lim tillari rejasi tasdiqlanmagan"],
                "ogohlantirishlar": sync.get("ogohlantirishlar", []),
                "sinflar": [], "oqituvchilar": [], "fanlar": [],
                "xulosa": {"xato_soni": 1},
            }
        report = _v1875_preflight_report(cur, maktab_id)
        conn.commit()
        return report
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()

# ========================= V18.75 END =========================


# ═══════════════════════════════════════════════════════════
# V18.76 — GURUHLI FANLARNI JADVALDAN OLDIN TASDIQLASH
# Sinf yaratishda saqlangan guruh tizimlari va Excel DARS_BIRIKMALARI
# bitta hisobotda solishtiriladi. Guruh→o'qituvchi taqsimoti to'liq
# tasdiqlanmaguncha moslik tekshiruvi, draft yaratish va tasdiqlash yopiq.
# ═══════════════════════════════════════════════════════════


def _v1876_tables(cur):
    _v1875_tables(cur)
    _sinf_kop_guruh_jadvallari(cur)
    cur.execute("""CREATE TABLE IF NOT EXISTS aqlli_fan_guruh_tasdiqlari_v2(
        maktab_id INTEGER NOT NULL REFERENCES maktablar(id) ON DELETE CASCADE,
        sinf_id INTEGER NOT NULL REFERENCES maktab_sinflari(id) ON DELETE CASCADE,
        fan_nomi TEXT NOT NULL,
        fan_kaliti TEXT NOT NULL,
        turi TEXT NOT NULL DEFAULT 'group' CHECK(turi IN ('whole','group')),
        tizim_id BIGINT REFERENCES maktab_sinf_guruh_tizimlari(id) ON DELETE SET NULL,
        manba_hash TEXT NOT NULL,
        tasdiqlangan BOOLEAN NOT NULL DEFAULT FALSE,
        tasdiqlagan_user_id BIGINT,
        yangilangan_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
        PRIMARY KEY(maktab_id,sinf_id,fan_kaliti)
    )""")
    cur.execute(
        "CREATE INDEX IF NOT EXISTS idx_fan_guruh_tasdiq_v2 "
        "ON aqlli_fan_guruh_tasdiqlari_v2(maktab_id,tasdiqlangan)"
    )


@app.on_event("startup")
def _v1876_startup_tables():
    try:
        conn = _db(); cur = conn.cursor()
        _v1876_tables(cur)
        conn.commit(); cur.close(); conn.close()
    except Exception as exc:
        print(f"[V18.76 guruh tasdiq jadvallari] {exc}", flush=True)


def _v1876_seed_legacy_group_systems(cur, maktab_id: int):
    """Eski guruhlash_usuli ustunidagi sozlamani ko'p tizimli jadvalga ko'chiradi."""
    cur.execute("""SELECT id,COALESCE(guruhlash_usuli,'none') AS guruhlash_usuli
                   FROM maktab_sinflari WHERE maktab_id=%s""", (maktab_id,))
    for cls in cur.fetchall():
        kind = str(cls.get("guruhlash_usuli") or "none").strip().lower()
        kind = {
            "1/2": "alphabet", "1_2": "alphabet", "numbered": "alphabet",
            "half": "alphabet", "ikki_guruh": "alphabet",
            "ogil_qiz": "gender", "o'g'il/qiz": "gender",
            "o‘g‘il/qiz": "gender", "boys_girls": "gender",
        }.get(kind, kind)
        if kind not in _SINF_GURUH_TIZIM_NOMLARI:
            continue
        cur.execute(
            "SELECT 1 FROM maktab_sinf_guruh_tizimlari "
            "WHERE sinf_id=%s AND faol=TRUE LIMIT 1",
            (cls["id"],),
        )
        if cur.fetchone():
            continue
        cur.execute("""INSERT INTO maktab_sinf_guruh_tizimlari(
                         sinf_id,turi,nomi,fanlar,faol,yangilangan_at)
                       VALUES(%s,%s,%s,ARRAY[]::TEXT[],TRUE,NOW())
                       ON CONFLICT(sinf_id,turi) DO UPDATE SET
                         nomi=EXCLUDED.nomi,faol=TRUE,yangilangan_at=NOW()""",
                    (cls["id"], kind, _SINF_GURUH_TIZIM_NOMLARI[kind]))


def _v1876_system_group_rows(cur, system):
    kind = str(system.get("turi") or "")
    if kind == "alphabet":
        guruh_soni = _sinf_guruh_soni_normalizatsiya("alphabet", system.get("guruh_soni"))
        return [
            {"guruh_kaliti": f"group_{raqam}", "guruh_nomi": f"{raqam}-guruh", "oquvchi_soni": 0}
            for raqam in range(1, guruh_soni + 1)
        ]
    if kind == "gender":
        return [
            {"guruh_kaliti": "boys", "guruh_nomi": "O'g'il bolalar", "oquvchi_soni": 0},
            {"guruh_kaliti": "girls", "guruh_nomi": "Qiz bolalar", "oquvchi_soni": 0},
        ]
    cur.execute("""SELECT guruh_kaliti,MIN(guruh_nomi) AS guruh_nomi,
                          COUNT(*) AS oquvchi_soni
                   FROM maktab_sinf_guruh_azolari
                   WHERE tizim_id=%s
                   GROUP BY guruh_kaliti
                   ORDER BY MIN(guruh_nomi),guruh_kaliti""", (system["id"],))
    return [dict(row) for row in cur.fetchall()]


def _v1876_group_system_catalog(cur, maktab_id: int):
    _sinf_kop_guruh_jadvallari(cur)
    _v1876_seed_legacy_group_systems(cur, maktab_id)
    cur.execute("""SELECT t.id,t.sinf_id,t.turi,t.nomi,t.fanlar,t.yangilangan_at,
                          s.sinf,s.harf,COALESCE(s.smena,1) AS smena,s.guruh_soni
                   FROM maktab_sinf_guruh_tizimlari t
                   JOIN maktab_sinflari s ON s.id=t.sinf_id
                   WHERE s.maktab_id=%s AND t.faol=TRUE
                   ORDER BY s.sinf::int,s.harf,
                     CASE t.turi WHEN 'gender' THEN 1 WHEN 'alphabet' THEN 2 ELSE 3 END,t.id""",
                (maktab_id,))
    result = []
    for raw in cur.fetchall():
        system = dict(raw)
        if system["turi"] in ("gender", "alphabet"):
            cur.execute(
                "SELECT COUNT(*) AS soni FROM maktab_sinf_guruh_azolari "
                "WHERE tizim_id=%s",
                (system["id"],),
            )
            if int((cur.fetchone() or {}).get("soni") or 0) == 0:
                _sinf_guruh_tizimini_taqsimla(cur, int(system["id"]))
        groups = _v1876_system_group_rows(cur, system)
        if system["turi"] in ("gender", "alphabet"):
            cur.execute("""SELECT guruh_kaliti,COUNT(*) AS soni
                           FROM maktab_sinf_guruh_azolari
                           WHERE tizim_id=%s GROUP BY guruh_kaliti""", (system["id"],))
            counts = {str(row["guruh_kaliti"]): int(row["soni"]) for row in cur.fetchall()}
            for group in groups:
                group["oquvchi_soni"] = counts.get(group["guruh_kaliti"], 0)
        system["fanlar"] = [
            re.sub(r"\s+", " ", str(value or "")).strip()
            for value in (system.get("fanlar") or [])
            if str(value or "").strip()
        ]
        system["fan_kalitlari"] = [_v1875_subject_key(value) for value in system["fanlar"]]
        system["guruhlar"] = groups
        result.append(system)
    return result


def _v1876_pair_hash(pair_rows, systems, subject_key):
    """Faqat shu sinf+fan taqsimotiga ta'sir qiladigan manbalarni hash qiladi."""
    payload = {
        "rows": [
            {
                "id": int(row.get("id") or 0),
                "user_id": int(row.get("user_id") or 0),
                "guruh": _v1875_group_key(row.get("guruh_kaliti")),
                "soat": round(float(row.get("haftalik_soat") or 0), 1),
                "kunlik": int(row.get("kunlik_max") or 1),
            }
            for row in sorted(
                pair_rows,
                key=lambda value: (int(value.get("id") or 0), int(value.get("user_id") or 0)),
            )
        ],
        "systems": [
            {
                "id": int(system["id"]),
                "turi": system["turi"],
                "shu_fanga_biriktirilgan": subject_key in (system.get("fan_kalitlari") or []),
                "guruhlar": [group["guruh_kaliti"] for group in system.get("guruhlar") or []],
            }
            for system in systems
        ],
    }
    raw = json.dumps(payload, ensure_ascii=False, sort_keys=True, default=str)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _v1876_pair_assignment_rows(cur, maktab_id: int):
    cur.execute("""SELECT b.id,b.sinf_id,b.user_id,b.fan_nomi,b.guruh_kaliti,
                          b.haftalik_soat,b.kunlik_max,b.manba,u.full_name,
                          s.sinf,s.harf,COALESCE(s.smena,1) AS smena,
                          s.bino,s.xona
                   FROM maktab_dars_birikmalari b
                   JOIN users u ON u.user_id=b.user_id
                   JOIN maktab_sinflari s ON s.id=b.sinf_id
                   WHERE b.maktab_id=%s
                   ORDER BY s.sinf::int,s.harf,b.fan_nomi,b.id""", (maktab_id,))
    pairs = {}
    for raw in cur.fetchall():
        row = dict(raw)
        subject = re.sub(r"\s+", " ", str(row.get("fan_nomi") or "")).strip()
        subject_key = _v1875_subject_key(subject)
        key = (int(row["sinf_id"]), subject_key)
        pair = pairs.setdefault(key, {
            "sinf_id": int(row["sinf_id"]),
            "sinf": f"{row['sinf']}-{row['harf']}",
            "sinf_daraja": int(row["sinf"]),
            "smena": int(row.get("smena") or 1),
            "bino": row.get("bino"),
            "xona": row.get("xona"),
            "fan_nomi": subject,
            "fan_kaliti": subject_key,
            "rows": [],
        })
        row["guruh_kaliti"] = _v1875_group_key(row.get("guruh_kaliti"))
        row["user_id"] = int(row["user_id"])
        pair["rows"].append(row)
    return pairs


def _v1876_teacher_candidates(cur, maktab_id: int):
    result = {}
    for teacher in _v1859_effective_teachers(cur, maktab_id):
        if not teacher.get("dars_beruvchi"):
            continue
        for subject in teacher.get("fanlar_royxati") or []:
            result.setdefault(_v1875_subject_key(subject), []).append({
                "user_id": int(teacher["user_id"]),
                "full_name": teacher["full_name"],
                "haftalik_dars_soati": teacher.get("haftalik_dars_soati"),
                "sinflar": teacher.get("sinflar_royxati") or [],
            })
    for key, teachers in result.items():
        unique = {int(teacher["user_id"]): teacher for teacher in teachers}
        result[key] = sorted(
            unique.values(),
            key=lambda teacher: (str(teacher["full_name"]).casefold(), teacher["user_id"]),
        )
    return result


def _v1876_load_map(cur, maktab_id: int):
    cur.execute("""SELECT sinf_id,fan_nomi,haftalik_soat,kunlik_max
                   FROM aqlli_sinf_fan_yuklamalari_v2
                   WHERE maktab_id=%s AND haftalik_soat>0""", (maktab_id,))
    return {
        (int(row["sinf_id"]), _v1875_subject_key(row["fan_nomi"])): dict(row)
        for row in cur.fetchall()
    }


def _v1876_group_review_report(cur, maktab_id: int):
    _v1876_tables(cur)
    systems = _v1876_group_system_catalog(cur, maktab_id)
    systems_by_class = _v1852_defaultdict(list)
    for system in systems:
        systems_by_class[int(system["sinf_id"])].append(system)

    pairs = _v1876_pair_assignment_rows(cur, maktab_id)
    # Guruh tizimiga fan biriktirilgan, ammo shablonda qator yo'q bo'lsa ham
    # tasdiqlash oynasida ko'rsatiladi.
    for system in systems:
        for subject in system.get("fanlar") or []:
            key = (int(system["sinf_id"]), _v1875_subject_key(subject))
            if key not in pairs:
                pairs[key] = {
                    "sinf_id": int(system["sinf_id"]),
                    "sinf": f"{system['sinf']}-{system['harf']}",
                    "sinf_daraja": int(system["sinf"]),
                    "smena": int(system.get("smena") or 1),
                    "fan_nomi": subject,
                    "fan_kaliti": _v1875_subject_key(subject),
                    "rows": [],
                }

    cur.execute(
        "SELECT * FROM aqlli_fan_guruh_tasdiqlari_v2 WHERE maktab_id=%s",
        (maktab_id,),
    )
    confirmations = {
        (int(row["sinf_id"]), str(row["fan_kaliti"])): dict(row)
        for row in cur.fetchall()
    }

    cur.execute(
        "SELECT * FROM aqlli_guruh_sozlamalari_v2 WHERE maktab_id=%s",
        (maktab_id,),
    )
    settings = _v1852_defaultdict(dict)
    for row in cur.fetchall():
        settings[(
            int(row["sinf_id"]),
            _v1875_subject_key(row["fan_nomi"]),
        )][str(row["guruh_kaliti"])] = dict(row)

    candidate_map = _v1876_teacher_candidates(cur, maktab_id)
    load_map = _v1876_load_map(cur, maktab_id)
    cur.execute(
        """SELECT id,nomi,turi,sigim FROM aqlli_xonalar_v2
           WHERE maktab_id=%s AND faol=TRUE AND darsga_yaroqli=TRUE
           ORDER BY CASE turi WHEN 'sport' THEN 1 WHEN 'reserve' THEN 2 ELSE 3 END,nomi""",
        (maktab_id,),
    )
    teaching_rooms = [dict(row) for row in cur.fetchall()]
    teaching_room_by_id = {int(row["id"]): row for row in teaching_rooms}
    review_pairs = []
    global_errors = []
    global_warnings = []

    for key, pair in sorted(
        pairs.items(),
        key=lambda item: (
            item[1]["sinf_daraja"],
            item[1]["sinf"],
            item[1]["fan_nomi"].casefold(),
        ),
    ):
        class_systems = systems_by_class.get(pair["sinf_id"], [])
        rows = pair["rows"]
        explicit = [row for row in rows if row["guruh_kaliti"] != "whole"]
        whole = [row for row in rows if row["guruh_kaliti"] == "whole"]

        distinct_teachers = []
        seen_teacher = set()
        for row in rows:
            if row["user_id"] not in seen_teacher:
                seen_teacher.add(row["user_id"])
                distinct_teachers.append(row)

        subject_systems = [
            system for system in class_systems
            if pair["fan_kaliti"] in (system.get("fan_kalitlari") or [])
        ]
        explicit_keys = {row["guruh_kaliti"] for row in explicit}
        matching_explicit = [
            system for system in class_systems
            if explicit_keys and explicit_keys.issubset({
                group["guruh_kaliti"] for group in system.get("guruhlar") or []
            })
        ]
        count_matching = [
            system for system in class_systems
            if len(system.get("guruhlar") or []) == len(distinct_teachers)
            and len(distinct_teachers) >= 2
        ]
        candidate_systems = []
        for system in subject_systems + matching_explicit + count_matching:
            if all(int(existing["id"]) != int(system["id"]) for existing in candidate_systems):
                candidate_systems.append(system)

        group_relevant = bool(explicit or len(distinct_teachers) > 1 or subject_systems)
        if not group_relevant:
            continue

        pair_hash = _v1876_pair_hash(rows, class_systems, pair["fan_kaliti"])
        confirmation = confirmations.get(key)
        valid_confirmation = bool(
            confirmation
            and confirmation.get("tasdiqlangan")
            and str(confirmation.get("manba_hash")) == pair_hash
        )
        mode = (
            str(confirmation.get("turi"))
            if valid_confirmation
            else "group"
        )

        selected_system_id = None
        if mode == "group":
            if valid_confirmation and confirmation.get("tizim_id"):
                selected_system_id = int(confirmation["tizim_id"])
            elif len(subject_systems) == 1:
                selected_system_id = int(subject_systems[0]["id"])
            elif len(candidate_systems) == 1:
                selected_system_id = int(candidate_systems[0]["id"])
        selected_system = next(
            (
                system for system in class_systems
                if int(system["id"]) == int(selected_system_id or 0)
            ),
            None,
        )

        errors = []
        warnings = []
        positive_hours = sorted({
            float(row.get("haftalik_soat") or 0)
            for row in rows if float(row.get("haftalik_soat") or 0) > 0
        })
        nonempty_daily = sorted({
            int(row.get("kunlik_max") or 1)
            for row in rows if row.get("kunlik_max") not in (None, "")
        })
        fallback_load = load_map.get(key) or {}
        if not positive_hours and float(fallback_load.get("haftalik_soat") or 0) > 0:
            positive_hours = [float(fallback_load["haftalik_soat"])]
            warnings.append("Haftalik soat oldingi sinf–fan yuklamasidan olindi")
        if not nonempty_daily and fallback_load:
            nonempty_daily = [int(fallback_load.get("kunlik_max") or 1)]

        weekly_hours = positive_hours[0] if len(positive_hours) == 1 else 0
        daily_max = nonempty_daily[0] if len(nonempty_daily) == 1 else 1
        if not rows:
            errors.append("Bu guruhli fan uchun shablonda o'qituvchi birikmasi yo'q")
        if len(positive_hours) != 1:
            errors.append("Guruhli fan uchun haftalik soat bitta aniq qiymat bo'lishi kerak")
        if len(nonempty_daily) > 1:
            errors.append("Guruhli fan qatorlaridagi kunlik maksimum teng bo'lishi kerak")
        if rows and any(float(row.get("haftalik_soat") or 0) <= 0 for row in rows) and weekly_hours:
            warnings.append(f"Soati bo'sh guruh qatorlari tasdiqda {weekly_hours} soatga tenglashtiriladi")
        if explicit and whole:
            errors.append("Butun sinf va guruh qatorlari birga yozilgan")

        group_payload = []
        if mode == "group":
            if not class_systems:
                errors.append("Sinfda hech qanday guruhlash tizimi yaratilmagan")
            elif not candidate_systems and not selected_system:
                errors.append("Bu fan uchun qaysi guruhlash tizimi ishlashi aniqlanmagan")
            elif len(candidate_systems) > 1 and not selected_system:
                warnings.append("Bir nechta guruhlash tizimi mos keldi — bittasini tanlang")

            if selected_system:
                system_groups = selected_system.get("guruhlar") or []
                saved = settings.get(key, {})
                home_candidates = {
                    _v205_room_normalized_name(value)
                    for value in (
                        " ".join(
                            part for part in (pair.get("bino"), pair.get("xona"))
                            if str(part or "").strip()
                        ),
                        pair.get("xona"),
                    )
                    if _v205_room_normalized_name(value)
                }
                home_room_id = next(
                    (
                        int(room["id"]) for room in teaching_rooms
                        if _v205_room_normalized_name(room.get("nomi")) in home_candidates
                    ),
                    None,
                )
                used_room_ids = set()
                sport_subject = any(
                    word in pair["fan_kaliti"]
                    for word in ("jismoniy", "sport", "fizkultura")
                )
                explicit_by_key = {row["guruh_kaliti"]: row for row in explicit}
                ordered_whole = sorted(
                    whole,
                    key=lambda row: (
                        int(row.get("id") or 0),
                        str(row.get("full_name") or "").casefold(),
                    ),
                )
                for index, group in enumerate(system_groups):
                    group_key = group["guruh_kaliti"]
                    source = explicit_by_key.get(group_key)
                    setting = saved.get(group_key) or {}
                    teacher_id = setting.get("oqituvchi_user_id") if valid_confirmation else None
                    if teacher_id is None and source:
                        teacher_id = source.get("user_id")
                    if teacher_id is None and index < len(ordered_whole):
                        teacher_id = ordered_whole[index].get("user_id")
                    teacher_name = next(
                        (
                            teacher["full_name"]
                            for teacher in candidate_map.get(pair["fan_kaliti"], [])
                            if int(teacher["user_id"]) == int(teacher_id or 0)
                        ),
                        source.get("full_name") if source else None,
                    )
                    # Birinchi guruh odatda sinf xonasida qoladi, ammo rahbar
                    # unga ham maxsus xona tanlagan bo'lsa bu tanlov yo'qolmaydi.
                    room_id = setting.get("xona_id")
                    if room_id is not None and int(room_id) not in teaching_room_by_id:
                        room_id = None
                    if room_id is None and index == 0 and home_room_id is not None:
                        # Birinchi guruhning bo'sh tanlovi "sinf xonasi" degani.
                        used_room_ids.add(int(home_room_id))
                    if room_id is None and index > 0:
                        preferred_types = ("sport",) if sport_subject else ("reserve", "classroom")
                        suggested_room = next(
                            (
                                room for room in teaching_rooms
                                if str(room.get("turi") or "classroom") in preferred_types
                                and int(room["id"]) not in used_room_ids
                            ),
                            None,
                        )
                        if suggested_room:
                            room_id = int(suggested_room["id"])
                        else:
                            warnings.append(
                                "Bo'linishga xona topilmadi — bir guruh sinf xonasida qoladi, qolganiga sport zal yoki zaxira xona kiriting"
                            )
                    if room_id is not None:
                        used_room_ids.add(int(room_id))
                    group_payload.append({
                        **group,
                        "oqituvchi_user_id": int(teacher_id) if teacher_id is not None else None,
                        "oqituvchi_ismi": teacher_name,
                        "xona_id": int(room_id) if room_id is not None else None,
                        "xona_nomi": teaching_room_by_id.get(int(room_id), {}).get("nomi") if room_id is not None else None,
                    })

                teacher_ids = [
                    group["oqituvchi_user_id"]
                    for group in group_payload
                    if group.get("oqituvchi_user_id") is not None
                ]
                if len(group_payload) < 2:
                    errors.append("Tanlangan guruhlash tizimida kamida 2 ta guruh bo'lishi kerak")
                if len(teacher_ids) != len(group_payload):
                    errors.append("Har bir guruhga bittadan o'qituvchi tanlang")
                if len(teacher_ids) != len(set(teacher_ids)):
                    errors.append("Parallel guruhlarga turli o'qituvchilar kerak")
                if rows and len(distinct_teachers) != len(group_payload) and not valid_confirmation:
                    warnings.append(
                        f"Shablonda {len(distinct_teachers)} ta o'qituvchi, "
                        f"tanlangan tizimda {len(group_payload)} ta guruh bor"
                    )
        else:
            # Butun sinf tasdiqlanganda source qatorida aynan bitta o'qituvchi bo'ladi.
            if len(whole) != 1:
                errors.append("Butun sinf uchun aynan bitta o'qituvchi tanlang")

        status = (
            "tasdiqlangan"
            if valid_confirmation and not errors
            else ("xato" if errors else "taklif")
        )
        if status != "tasdiqlangan":
            global_errors.append(
                f"{pair['sinf']} / {pair['fan_nomi']}: "
                + ("; ".join(errors) if errors else "guruh va o'qituvchilar tasdiqlanmagan")
            )
        global_warnings.extend(
            f"{pair['sinf']} / {pair['fan_nomi']}: {warning}"
            for warning in warnings
        )

        # Parallel guruh hisobining uchta alohida o'lchovi:
        #   1) sinf reja soati — guruhlar soniga ko'paymaydi;
        #   2) o'qituvchi-soat — har bir guruh o'qituvchisiga alohida yoziladi;
        #   3) jadval sloti — guruhlar bir vaqtda parallel turgani uchun reja soatiga teng.
        parallel_guruh_soni = (
            len(group_payload)
            if mode == "group" and group_payload
            else (
                len(distinct_teachers)
                if mode == "group" and len(distinct_teachers) >= 2
                else 1
            )
        )
        sinf_reja_soati = round(float(weekly_hours or 0), 1)
        jadval_parallel_slot_soni = round(float(weekly_hours or 0), 1)
        oqituvchi_soat_jami = round(
            float(weekly_hours or 0) * int(parallel_guruh_soni),
            1,
        )

        review_pairs.append({
            "sinf_id": pair["sinf_id"],
            "sinf": pair["sinf"],
            "sinf_daraja": pair["sinf_daraja"],
            "smena": pair["smena"],
            "fan_nomi": pair["fan_nomi"],
            "fan_kaliti": pair["fan_kaliti"],
            "haftalik_soat": weekly_hours,
            "sinf_reja_soati": sinf_reja_soati,
            "jadval_parallel_slot_soni": jadval_parallel_slot_soni,
            "parallel_guruh_soni": parallel_guruh_soni,
            "har_bir_guruh_oqituvchi_soati": round(float(weekly_hours or 0), 1),
            "oqituvchi_soat_jami": oqituvchi_soat_jami,
            "kunlik_max": daily_max,
            "status": status,
            "tasdiqlangan": status == "tasdiqlangan",
            "manba_hash": pair_hash,
            "turi": mode,
            "tizim_id": selected_system_id,
            "tizimlar": [
                {
                    "id": int(system["id"]),
                    "turi": system["turi"],
                    "nomi": system["nomi"],
                    "fan_biriktirilgan": pair["fan_kaliti"] in (system.get("fan_kalitlari") or []),
                    "guruhlar": system.get("guruhlar") or [],
                }
                for system in class_systems
            ],
            "guruhlar": group_payload,
            "asosiy_oqituvchi_user_id": whole[0]["user_id"] if len(whole) == 1 else None,
            "kandidat_oqituvchilar": candidate_map.get(pair["fan_kaliti"], []),
            "import_oqituvchilari": [
                {
                    "user_id": row["user_id"],
                    "full_name": row.get("full_name"),
                    "guruh_kaliti": row["guruh_kaliti"],
                    "haftalik_soat": row.get("haftalik_soat"),
                }
                for row in distinct_teachers
            ],
            "xatolar": errors,
            "ogohlantirishlar": warnings,
        })

    # Barcha sinflar ko'rinadi: guruhli fani yo'q sinf ham 'tizim yo'q/bor' deb
    # jadvaldan oldin tekshirilishi mumkin.
    cur.execute("""SELECT id,sinf,harf,COALESCE(smena,1) AS smena
                   FROM maktab_sinflari WHERE maktab_id=%s
                   ORDER BY sinf::int,harf""", (maktab_id,))
    class_map = {
        int(row["id"]): {
            "sinf_id": int(row["id"]),
            "sinf": f"{row['sinf']}-{row['harf']}",
            "sinf_daraja": int(row["sinf"]),
            "smena": int(row.get("smena") or 1),
            "tizimlar": [],
            "fanlar": [],
        }
        for row in cur.fetchall()
    }
    for system in systems:
        class_entry = class_map.setdefault(int(system["sinf_id"]), {
            "sinf_id": int(system["sinf_id"]),
            "sinf": f"{system['sinf']}-{system['harf']}",
            "sinf_daraja": int(system["sinf"]),
            "smena": int(system.get("smena") or 1),
            "tizimlar": [],
            "fanlar": [],
        })
        class_entry["tizimlar"].append({
            "id": int(system["id"]),
            "turi": system["turi"],
            "nomi": system["nomi"],
            "fanlar": system.get("fanlar") or [],
            "guruhlar": system.get("guruhlar") or [],
        })
    for pair in review_pairs:
        class_map[pair["sinf_id"]]["fanlar"].append(pair)

    confirmed = sum(1 for pair in review_pairs if pair["tasdiqlangan"])
    summary = {
        "sinf_soni": len(class_map),
        "guruh_tizimli_sinf_soni": sum(1 for row in class_map.values() if row["tizimlar"]),
        "guruhli_fan_soni": len(review_pairs),
        "tasdiqlangan": confirmed,
        "tasdiqlanmagan": len(review_pairs) - confirmed,
        "xato_soni": len(global_errors),
    }
    return {
        "tayyor": not global_errors,
        "xulosa": summary,
        "sinflar": sorted(
            class_map.values(),
            key=lambda row: (row["sinf_daraja"], row["sinf"]),
        ),
        "fanlar": review_pairs,
        "xonalar": teaching_rooms,
        "xatolar": list(dict.fromkeys(global_errors))[:300],
        "ogohlantirishlar": list(dict.fromkeys(global_warnings))[:300],
    }


class V1876GroupTeacher(BaseModel):
    guruh_kaliti: str
    oqituvchi_user_id: int
    xona_id: Optional[int] = None


class V1876GroupPairConfirm(BaseModel):
    sinf_id: int
    fan_nomi: str
    turi: str = "group"
    tizim_id: Optional[int] = None
    asosiy_oqituvchi_user_id: Optional[int] = None
    guruhlar: list[V1876GroupTeacher] = []


class V1876GroupConfirmBatch(BaseModel):
    maktab_id: int
    birikmalar: list[V1876GroupPairConfirm]


def _v1876_exact_subject_teachers(cur, maktab_id: int, subject_key: str):
    allowed = {}
    for teacher in _v1859_effective_teachers(cur, maktab_id):
        if any(
            _v1875_subject_key(subject) == subject_key
            for subject in (teacher.get("fanlar_royxati") or [])
        ):
            allowed[int(teacher["user_id"])] = teacher
    return allowed


def _v1876_pair_source_rows(cur, maktab_id: int, class_id: int, subject_key: str):
    cur.execute("""SELECT b.*,u.full_name FROM maktab_dars_birikmalari b
                   JOIN users u ON u.user_id=b.user_id
                   WHERE b.maktab_id=%s AND b.sinf_id=%s ORDER BY b.id""",
                (maktab_id, class_id))
    return [
        dict(row) for row in cur.fetchall()
        if _v1875_subject_key(row.get("fan_nomi")) == subject_key
    ]


def _v1876_delete_group_settings(cur, maktab_id: int, class_id: int, subject_key: str):
    cur.execute("""SELECT fan_nomi,guruh_kaliti FROM aqlli_guruh_sozlamalari_v2
                   WHERE maktab_id=%s AND sinf_id=%s""", (maktab_id, class_id))
    for row in cur.fetchall():
        if _v1875_subject_key(row.get("fan_nomi")) == subject_key:
            cur.execute("""DELETE FROM aqlli_guruh_sozlamalari_v2
                           WHERE maktab_id=%s AND sinf_id=%s
                             AND fan_nomi=%s AND guruh_kaliti=%s""",
                        (maktab_id, class_id, row["fan_nomi"], row["guruh_kaliti"]))


def _v1876_update_system_subjects(cur, class_id: int, subject: str, selected_system_id: Optional[int]):
    cur.execute("""SELECT id,fanlar FROM maktab_sinf_guruh_tizimlari
                   WHERE sinf_id=%s AND faol=TRUE ORDER BY id FOR UPDATE""", (class_id,))
    subject_key = _v1875_subject_key(subject)
    for row in cur.fetchall():
        subjects = [
            re.sub(r"\s+", " ", str(value or "")).strip()
            for value in (row.get("fanlar") or [])
            if str(value or "").strip()
        ]
        subjects = [value for value in subjects if _v1875_subject_key(value) != subject_key]
        if selected_system_id is not None and int(row["id"]) == int(selected_system_id):
            subjects.append(subject)
        cur.execute(
            "UPDATE maktab_sinf_guruh_tizimlari "
            "SET fanlar=%s,yangilangan_at=NOW() WHERE id=%s",
            (subjects, row["id"]),
        )


def _v1876_rebuild_teacher_class_links(cur, maktab_id: int, teacher_ids):
    """O'qituvchi almashganda eski sinf havolalarini exact birikmalardan qayta quradi."""
    teacher_ids = sorted({int(value) for value in teacher_ids if value is not None})
    if not teacher_ids:
        return
    cur.execute(
        "DELETE FROM maktab_xodim_sinflari "
        "WHERE maktab_id=%s AND user_id=ANY(%s)",
        (maktab_id, teacher_ids),
    )
    cur.execute("""SELECT b.user_id,b.sinf_id,b.fan_nomi,s.sinf,s.harf
                   FROM maktab_dars_birikmalari b
                   JOIN maktab_sinflari s ON s.id=b.sinf_id
                   WHERE b.maktab_id=%s AND b.user_id=ANY(%s)
                   ORDER BY b.user_id,s.sinf::int,s.harf,b.fan_nomi""",
                (maktab_id, teacher_ids))
    by_teacher_class = {}
    class_labels = _v1852_defaultdict(set)
    for row in cur.fetchall():
        key = (int(row["user_id"]), int(row["sinf_id"]))
        entry = by_teacher_class.setdefault(key, {})
        subject = re.sub(r"\s+", " ", str(row.get("fan_nomi") or "")).strip()
        entry[_v1875_subject_key(subject)] = subject
        class_labels[int(row["user_id"])].add(f"{row['sinf']}-{row['harf']}")
    for (teacher_id, class_id), subject_map in by_teacher_class.items():
        subjects = sorted(subject_map.values(), key=lambda value: value.casefold())
        cur.execute("""INSERT INTO maktab_xodim_sinflari(
                         maktab_id,user_id,sinf_id,fanlari)
                       VALUES(%s,%s,%s,%s)
                       ON CONFLICT(user_id,sinf_id) DO UPDATE SET
                         maktab_id=EXCLUDED.maktab_id,fanlari=EXCLUDED.fanlari""",
                    (maktab_id, teacher_id, class_id, "\n".join(subjects) or None))
    for teacher_id in teacher_ids:
        labels = sorted(class_labels.get(teacher_id, set()), key=_v1859_sinf_sort_key)
        cur.execute(
            "UPDATE users SET oqitadigan_sinflari=%s WHERE maktab_id=%s AND user_id=%s",
            ("; ".join(labels) or None, maktab_id, teacher_id),
        )


@app.get("/api/maktab/aqlli_jadval/v2/guruh_tasdiqlash")
def v1876_group_review(token: str, maktab_id: int):
    actor_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor()
    try:
        _v1876_tables(cur)
        if not _v1852_staff(cur, actor_id, maktab_id):
            raise HTTPException(status_code=403, detail="Guruh taqsimotini ko'rish vakolati yo'q")
        report = _v1876_group_review_report(cur, maktab_id)
        conn.commit()
        return report
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


@app.put("/api/maktab/aqlli_jadval/v2/guruh_tasdiqlash")
def v1876_group_confirm(sorov: V1876GroupConfirmBatch, token: str):
    actor_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor()
    try:
        _v1876_tables(cur)
        if not _v1852_manager(cur, actor_id, sorov.maktab_id):
            raise HTTPException(
                status_code=403,
                detail="Guruh o'qituvchilarini faqat rahbariyat tasdiqlaydi",
            )
        if not sorov.birikmalar:
            raise HTTPException(status_code=400, detail="Tasdiqlash uchun fan tanlanmagan")
        if len(sorov.birikmalar) > 300:
            raise HTTPException(status_code=400, detail="Bir amalda 300 tadan ortiq fan tasdiqlanmaydi")

        systems = _v1876_group_system_catalog(cur, sorov.maktab_id)
        systems_by_id = {int(system["id"]): system for system in systems}
        cur.execute(
            "SELECT id,sinf,harf FROM maktab_sinflari WHERE maktab_id=%s",
            (sorov.maktab_id,),
        )
        classes = {int(row["id"]): dict(row) for row in cur.fetchall()}
        cur.execute(
            """SELECT id FROM aqlli_xonalar_v2
               WHERE maktab_id=%s AND faol=TRUE AND darsga_yaroqli=TRUE""",
            (sorov.maktab_id,),
        )
        allowed_room_ids = {int(row["id"]) for row in cur.fetchall()}
        old_group_settings = {}
        cur.execute(
            "SELECT * FROM aqlli_guruh_sozlamalari_v2 WHERE maktab_id=%s",
            (sorov.maktab_id,),
        )
        for row in cur.fetchall():
            old_group_settings[(
                int(row["sinf_id"]),
                _v1875_subject_key(row["fan_nomi"]),
                str(row["guruh_kaliti"]),
            )] = dict(row)

        confirmed_count = 0
        all_affected_teacher_ids = set()

        for item in sorov.birikmalar:
            class_id = int(item.sinf_id)
            if class_id not in classes:
                raise HTTPException(status_code=400, detail=f"Sinf topilmadi: {class_id}")
            subject = re.sub(r"\s+", " ", str(item.fan_nomi or "")).strip()
            subject_key = _v1875_subject_key(subject)
            if not subject_key:
                raise HTTPException(status_code=400, detail="Fan nomi bo'sh")

            rows = _v1876_pair_source_rows(
                cur, sorov.maktab_id, class_id, subject_key
            )
            old_teacher_ids = {int(row["user_id"]) for row in rows}
            all_affected_teacher_ids.update(old_teacher_ids)
            hours_set = {
                float(row.get("haftalik_soat") or 0)
                for row in rows if float(row.get("haftalik_soat") or 0) > 0
            }
            daily_set = {
                int(row.get("kunlik_max") or 1)
                for row in rows if row.get("kunlik_max") not in (None, "")
            }
            if len(hours_set) != 1:
                cur.execute("""SELECT fan_nomi,haftalik_soat,kunlik_max
                               FROM aqlli_sinf_fan_yuklamalari_v2
                               WHERE maktab_id=%s AND sinf_id=%s""",
                            (sorov.maktab_id, class_id))
                matching_load = next(
                    (
                        dict(row) for row in cur.fetchall()
                        if _v1875_subject_key(row.get("fan_nomi")) == subject_key
                        and float(row.get("haftalik_soat") or 0) > 0
                    ),
                    None,
                )
                if matching_load:
                    hours_set = {float(matching_load["haftalik_soat"])}
                    daily_set = {int(matching_load.get("kunlik_max") or 1)}
            if len(hours_set) != 1:
                label = f"{classes[class_id]['sinf']}-{classes[class_id]['harf']}"
                raise HTTPException(
                    status_code=400,
                    detail=f"{label} / {subject}: haftalik soat aniq emas",
                )
            if len(daily_set) > 1:
                raise HTTPException(
                    status_code=400,
                    detail=f"{subject}: guruhlarning kunlik maksimumi teng emas",
                )
            weekly_hours = next(iter(hours_set))
            daily_max = next(iter(daily_set), 1)
            allowed_teachers = _v1876_exact_subject_teachers(
                cur, sorov.maktab_id, subject_key
            )
            row_ids = [int(row["id"]) for row in rows]

            mode = str(item.turi or "group").strip().lower()
            if mode not in {"whole", "group"}:
                raise HTTPException(status_code=400, detail="Turi whole yoki group bo'lishi kerak")

            if mode == "whole":
                teacher_id = int(item.asosiy_oqituvchi_user_id or 0)
                if teacher_id not in allowed_teachers:
                    raise HTTPException(
                        status_code=400,
                        detail=f"{subject}: tanlangan o'qituvchi bu fanga mos emas",
                    )
                if row_ids:
                    cur.execute(
                        "DELETE FROM maktab_dars_birikmalari WHERE id=ANY(%s)",
                        (row_ids,),
                    )
                cur.execute("""INSERT INTO maktab_dars_birikmalari(
                                maktab_id,user_id,sinf_id,fan_nomi,guruh_kaliti,
                                haftalik_soat,kunlik_max,manba)
                               VALUES(%s,%s,%s,%s,'whole',%s,%s,'guruh_tasdiq')""",
                            (
                                sorov.maktab_id, teacher_id, class_id, subject,
                                weekly_hours, daily_max,
                            ))
                _v1876_delete_group_settings(
                    cur, sorov.maktab_id, class_id, subject_key
                )
                _v1876_update_system_subjects(cur, class_id, subject, None)
                selected_system_id = None
                selected_teacher_ids = {teacher_id}
            else:
                system_id = int(item.tizim_id or 0)
                system = systems_by_id.get(system_id)
                if not system or int(system["sinf_id"]) != class_id:
                    raise HTTPException(
                        status_code=400,
                        detail=f"{subject}: tanlangan guruhlash tizimi bu sinfga tegishli emas",
                    )
                expected_groups = [
                    str(group["guruh_kaliti"])
                    for group in system.get("guruhlar") or []
                ]
                if len(expected_groups) < 2:
                    raise HTTPException(
                        status_code=400,
                        detail=f"{subject}: tanlangan tizimda kamida 2 ta guruh bo'lishi kerak",
                    )
                payload_map = {
                    str(group.guruh_kaliti): group for group in item.guruhlar
                }
                if set(payload_map) != set(expected_groups):
                    label = f"{classes[class_id]['sinf']}-{classes[class_id]['harf']}"
                    raise HTTPException(
                        status_code=400,
                        detail=(
                            f"{label} / {subject}: barcha guruhlar to'liq tanlanishi "
                            f"kerak ({', '.join(expected_groups)})"
                        ),
                    )
                teacher_ids = [
                    int(payload_map[key].oqituvchi_user_id)
                    for key in expected_groups
                ]
                if len(teacher_ids) != len(set(teacher_ids)):
                    raise HTTPException(
                        status_code=400,
                        detail=f"{subject}: parallel guruhlarga turli o'qituvchilar tanlang",
                    )
                for teacher_id in teacher_ids:
                    if teacher_id not in allowed_teachers:
                        raise HTTPException(
                            status_code=400,
                            detail=f"{subject}: {teacher_id} bu fanga mos o'qituvchi emas",
                        )
                selected_room_ids = []
                for group_key in expected_groups:
                    selected_room_id = payload_map[group_key].xona_id
                    if selected_room_id is None:
                        continue
                    selected_room_id = int(selected_room_id)
                    if selected_room_id not in allowed_room_ids:
                        raise HTTPException(
                            status_code=400,
                            detail=f"{subject}: tanlangan xona dars o'tishga yaroqli emas yoki faol emas",
                        )
                    selected_room_ids.append(selected_room_id)
                if len(selected_room_ids) != len(set(selected_room_ids)):
                    raise HTTPException(
                        status_code=400,
                        detail=f"{subject}: parallel guruhlarga bir xil qo'shimcha xona tanlangan",
                    )
                if row_ids:
                    cur.execute(
                        "DELETE FROM maktab_dars_birikmalari WHERE id=ANY(%s)",
                        (row_ids,),
                    )
                _v1876_delete_group_settings(
                    cur, sorov.maktab_id, class_id, subject_key
                )
                for group_index, group_key in enumerate(expected_groups):
                    group = payload_map[group_key]
                    teacher_id = int(group.oqituvchi_user_id)
                    cur.execute("""INSERT INTO maktab_dars_birikmalari(
                                    maktab_id,user_id,sinf_id,fan_nomi,guruh_kaliti,
                                    haftalik_soat,kunlik_max,manba)
                                   VALUES(%s,%s,%s,%s,%s,%s,%s,'guruh_tasdiq')""",
                                (
                                    sorov.maktab_id, teacher_id, class_id, subject,
                                    group_key, weekly_hours, daily_max,
                                ))
                    old = old_group_settings.get(
                        (class_id, subject_key, group_key), {}
                    )
                    room_id = (
                        group.xona_id
                        if group.xona_id is not None
                        else old.get("xona_id")
                    )
                    if room_id is not None and int(room_id) not in allowed_room_ids:
                        room_id = None
                    cur.execute("""INSERT INTO aqlli_guruh_sozlamalari_v2(
                                    maktab_id,sinf_id,fan_nomi,guruh_kaliti,
                                    oqituvchi_user_id,xona_id)
                                   VALUES(%s,%s,%s,%s,%s,%s)
                                   ON CONFLICT(
                                     maktab_id,sinf_id,fan_nomi,guruh_kaliti
                                   ) DO UPDATE SET
                                     oqituvchi_user_id=EXCLUDED.oqituvchi_user_id,
                                     xona_id=EXCLUDED.xona_id""",
                                (
                                    sorov.maktab_id, class_id, subject, group_key,
                                    teacher_id, room_id,
                                ))
                _v1876_update_system_subjects(
                    cur, class_id, subject, system_id
                )
                selected_system_id = system_id
                selected_teacher_ids = set(teacher_ids)

            all_affected_teacher_ids.update(selected_teacher_ids)
            # Tasdiq hashini o'zgargan aniq manbadan hisoblaymiz.
            fresh_rows = _v1876_pair_source_rows(
                cur, sorov.maktab_id, class_id, subject_key
            )
            fresh_systems = [
                system for system in _v1876_group_system_catalog(cur, sorov.maktab_id)
                if int(system["sinf_id"]) == class_id
            ]
            source_hash = _v1876_pair_hash(
                fresh_rows, fresh_systems, subject_key
            )
            cur.execute("""INSERT INTO aqlli_fan_guruh_tasdiqlari_v2(
                            maktab_id,sinf_id,fan_nomi,fan_kaliti,turi,tizim_id,
                            manba_hash,tasdiqlangan,tasdiqlagan_user_id,yangilangan_at)
                           VALUES(%s,%s,%s,%s,%s,%s,%s,TRUE,%s,NOW())
                           ON CONFLICT(maktab_id,sinf_id,fan_kaliti) DO UPDATE SET
                             fan_nomi=EXCLUDED.fan_nomi,turi=EXCLUDED.turi,
                             tizim_id=EXCLUDED.tizim_id,manba_hash=EXCLUDED.manba_hash,
                             tasdiqlangan=TRUE,
                             tasdiqlagan_user_id=EXCLUDED.tasdiqlagan_user_id,
                             yangilangan_at=NOW()""",
                        (
                            sorov.maktab_id, class_id, subject, subject_key, mode,
                            selected_system_id, source_hash, actor_id,
                        ))
            confirmed_count += 1

        _v1876_rebuild_teacher_class_links(
            cur, sorov.maktab_id, all_affected_teacher_ids
        )
        cur.execute("""UPDATE aqlli_jadval_urinishlari_v2 SET holat='bekor'
                       WHERE maktab_id=%s AND holat='draft'""",
                    (sorov.maktab_id,))

        # Tasdiqlarni bittalab saqlash mumkin. Qolgan guruhli fanlar hali
        # tasdiqlanmagan bo'lsa strict manba rebuild keyinga qoldiriladi.
        report = _v1876_group_review_report(cur, sorov.maktab_id)
        sync = {
            "tayyor": False,
            "kutilmoqda": report.get("xulosa", {}).get("tasdiqlanmagan", 0),
            "izoh": "Qolgan guruhli fanlar tasdiqlangandan keyin jadval manbasi sinxronlanadi",
        }
        if report.get("tayyor"):
            sync = _v1875_rebuild_schedule_sources(
                cur,
                sorov.maktab_id,
                cancel_drafts=False,
                reason="guruh_tasdiq",
            )
            if sync.get("xatolar"):
                raise HTTPException(
                    status_code=409,
                    detail=(
                        "Guruh tasdiqidan keyin yuklama xatosi: "
                        + "; ".join(sync["xatolar"][:12])
                    ),
                )
            report = _v1876_group_review_report(cur, sorov.maktab_id)

        conn.commit()
        return {
            "holat": "tasdiqlandi",
            "tasdiqlangan_soni": confirmed_count,
            "hisobot": report,
            "sinxronlash": sync,
        }
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()

# ========================= V18.76 END =========================

class V1876QuickGroupSystem(BaseModel):
    maktab_id: int
    sinf_id: int
    fan_nomi: str
    turi: str


@app.post("/api/maktab/aqlli_jadval/v2/guruh_tizimi_tez")
def v1876_quick_group_system(sorov: V1876QuickGroupSystem, token: str):
    actor_id = _jwt_tekshir(token)
    kind = str(sorov.turi or "").strip().lower()
    if kind not in {"alphabet", "gender"}:
        raise HTTPException(
            status_code=400,
            detail="Tez yaratishda faqat 1/2-guruh yoki O'g'il/Qiz tizimi tanlanadi",
        )
    subject = re.sub(r"\s+", " ", str(sorov.fan_nomi or "")).strip()
    if not subject:
        raise HTTPException(status_code=400, detail="Fan nomi bo'sh")

    conn = _db(); cur = conn.cursor()
    try:
        _v1876_tables(cur)
        # DDL tayyorlash transactionini canonical yuklama transactionidan
        # ajratamiz. Guruh yaratish va Step-3 qisman saqlash bir vaqtda kelsa
        # ham fan/guruh kalitlari bir-birining ustidan yozilmaydi.
        conn.commit()
        cur.execute(
            "SELECT maktab_id FROM maktab_sinflari WHERE id=%s",
            (sorov.sinf_id,),
        )
        cls = cur.fetchone()
        if not cls or int(cls["maktab_id"]) != int(sorov.maktab_id):
            raise HTTPException(status_code=404, detail="Sinf topilmadi")
        if not _maktab_sinf_boshqaruvchi_mi(
            cur, actor_id, sorov.maktab_id
        ) and not _v1852_manager(cur, actor_id, sorov.maktab_id):
            raise HTTPException(
                status_code=403,
                detail="Guruh tizimini faqat admin yoki o'quv ishlari zavuchi yaratadi",
            )
        cur.execute(
            "SELECT pg_advisory_xact_lock(%s)",
            (1922000000 + int(sorov.maktab_id),),
        )
        cur.execute(
            "SELECT pg_advisory_xact_lock(%s)",
            (1925000000 + int(sorov.maktab_id),),
        )

        cur.execute(
            "SELECT id,turi,fanlar,faol FROM maktab_sinf_guruh_tizimlari "
            "WHERE sinf_id=%s ORDER BY id FOR UPDATE",
            (sorov.sinf_id,),
        )
        all_systems = [dict(row) for row in cur.fetchall()]
        existing = next(
            (row for row in all_systems if str(row.get("turi")) == kind),
            None,
        )
        linked_elsewhere = [
            row for row in all_systems
            if str(row.get("turi")) != kind
            and bool(row.get("faol", True))
            and any(
                _v204_group_subject_matches(value, subject)
                for value in (row.get("fanlar") or [])
            )
        ]
        if linked_elsewhere:
            raise HTTPException(
                status_code=409,
                detail=(
                    "Bu fan boshqa guruh turiga biriktirilgan. Avval o'sha "
                    "fan yuklamasini tozalab, keyin guruh turini almashtiring"
                ),
            )
        subjects = [
            re.sub(r"\s+", " ", str(value or "")).strip()
            for value in ((existing or {}).get("fanlar") or [])
            if str(value or "").strip()
        ]
        if all(
            _v1875_subject_key(value) != _v1875_subject_key(subject)
            for value in subjects
        ):
            subjects.append(subject)

        cur.execute("""INSERT INTO maktab_sinf_guruh_tizimlari(
                         sinf_id,turi,nomi,fanlar,faol,yaratilgan_by,yangilangan_at)
                       VALUES(%s,%s,%s,%s,TRUE,%s,NOW())
                       ON CONFLICT(sinf_id,turi) DO UPDATE SET
                         nomi=EXCLUDED.nomi,
                         fanlar=EXCLUDED.fanlar,
                         faol=TRUE,
                         yangilangan_at=NOW()
                       RETURNING id""",
                    (
                        sorov.sinf_id,
                        kind,
                        _SINF_GURUH_TIZIM_NOMLARI[kind],
                        subjects,
                        actor_id,
                    ))
        system_id = int(cur.fetchone()["id"])
        cur.execute(
            "SELECT COUNT(*) AS soni FROM maktab_sinf_guruh_azolari "
            "WHERE tizim_id=%s",
            (system_id,),
        )
        if int((cur.fetchone() or {}).get("soni") or 0) == 0:
            _sinf_guruh_tizimini_taqsimla(cur, system_id)

        # Tezkor tugma faqat sinf–fan guruh sxemasini yaratadi/bog'laydi.
        # O'qituvchining ``whole`` yoki aniq guruh yuklamasi bu endpointda
        # UPDATE/DELETE qilinmaydi; foydalanuvchi uni setkada ongli tanlaydi.

        cur.execute(
            "SELECT turi FROM maktab_sinf_guruh_tizimlari "
            "WHERE sinf_id=%s AND faol=TRUE ORDER BY id",
            (sorov.sinf_id,),
        )
        active_types = [row["turi"] for row in cur.fetchall()]
        legacy = (
            active_types[0]
            if len(active_types) == 1
            else ("manual" if active_types else "none")
        )
        cur.execute(
            "UPDATE maktab_sinflari SET guruhlash_usuli=%s WHERE id=%s",
            (legacy, sorov.sinf_id),
        )
        cur.execute(
            """UPDATE aqlli_jadval_urinishlari_v2 SET holat='bekor'
                 WHERE maktab_id=%s AND holat='draft'""",
            (sorov.maktab_id,),
        )
        report = _v1876_group_review_report(cur, sorov.maktab_id)
        conn.commit()
        return {
            "holat": "saqlandi",
            "tizim_id": system_id,
            "turi": kind,
            "hisobot": report,
        }
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


class V238ClassGroupSystemsBatch(BaseModel):
    maktab_id: int
    sinf_idlar: list[int]
    tizimlar: list[str]


def _v238_normalize_group_system(value):
    key = re.sub(r"\s+", "", str(value or "")).strip().casefold()
    normalized = {
        "alphabet": "alphabet", "1/2": "alphabet", "1_2": "alphabet",
        "numbered": "alphabet", "ikki_guruh": "alphabet",
        "gender": "gender", "o'g'il/qiz": "gender", "o‘g‘il/qiz": "gender",
        "ogil_qiz": "gender", "boys_girls": "gender",
    }.get(key)
    if normalized is None:
        raise ValueError("Guruh turi faqat 1/2-guruh yoki O‘g‘il/Qiz bo‘lishi kerak")
    return normalized


@app.put("/api/maktab/aqlli_jadval/v3/sinf_guruh_tizimlari")
def v238_class_group_systems_add(sorov: V238ClassGroupSystemsBatch, token: str):
    """Tanlangan sinflarga guruh sxemalarini faqat qo‘shadi (merge/upsert).

    1/2-guruhdan keyin O‘g‘il/Qiz qo‘shilsa birinchi sxema, uning fanlari,
    a’zolari va o‘qituvchi/xona sozlamalari o‘chirilmaydi.
    """
    actor_id = _jwt_tekshir(token)
    try:
        class_ids = sorted({int(value) for value in sorov.sinf_idlar})
    except (TypeError, ValueError) as exc:
        raise HTTPException(status_code=400, detail="Sinf IDlari noto‘g‘ri") from exc
    if not class_ids:
        raise HTTPException(status_code=400, detail="Kamida bitta sinfni tanlang")
    try:
        kinds = []
        for value in sorov.tizimlar:
            kind = _v238_normalize_group_system(value)
            if kind not in kinds:
                kinds.append(kind)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    if not kinds:
        raise HTTPException(status_code=400, detail="Kamida bitta guruh turini tanlang")

    conn = _db(); cur = conn.cursor()
    try:
        _v1876_tables(cur)
        conn.commit()
        if not _v1852_manager(cur, actor_id, sorov.maktab_id):
            raise HTTPException(
                status_code=403,
                detail="Guruh tizimini faqat maktab rahbariyati sozlaydi",
            )
        cur.execute(
            "SELECT pg_advisory_xact_lock(%s)",
            (1926000000 + int(sorov.maktab_id),),
        )
        cur.execute(
            "SELECT id FROM maktab_sinflari "
            "WHERE maktab_id=%s AND id=ANY(%s) ORDER BY id FOR UPDATE",
            (sorov.maktab_id, class_ids),
        )
        found = {int(row["id"]) for row in cur.fetchall()}
        missing = sorted(set(class_ids) - found)
        if missing:
            raise HTTPException(
                status_code=404,
                detail="Bu maktabga tegishli bo‘lmagan sinf IDlari: "
                + ", ".join(map(str, missing)),
            )

        created = 0
        restored = 0
        for class_id in class_ids:
            for kind in kinds:
                cur.execute(
                    "SELECT id,faol FROM maktab_sinf_guruh_tizimlari "
                    "WHERE sinf_id=%s AND turi=%s FOR UPDATE",
                    (class_id, kind),
                )
                existing = cur.fetchone()
                if existing:
                    if not bool(existing.get("faol", True)):
                        restored += 1
                    cur.execute(
                        "UPDATE maktab_sinf_guruh_tizimlari "
                        "SET nomi=%s,faol=TRUE,yangilangan_at=NOW() WHERE id=%s",
                        (_SINF_GURUH_TIZIM_NOMLARI[kind], existing["id"]),
                    )
                    # Mavjud fanlar va a'zolar saqlanadi. Faqat tarixiy yarim
                    # qolgan, a'zosi umuman yo'q tizim qayta taqsimlanadi.
                    cur.execute(
                        "SELECT COUNT(*) AS soni FROM maktab_sinf_guruh_azolari "
                        "WHERE tizim_id=%s",
                        (existing["id"],),
                    )
                    if int((cur.fetchone() or {}).get("soni") or 0) == 0:
                        _sinf_guruh_tizimini_taqsimla(cur, int(existing["id"]))
                    continue
                cur.execute(
                    """INSERT INTO maktab_sinf_guruh_tizimlari(
                         sinf_id,turi,nomi,fanlar,faol,yaratilgan_by,yangilangan_at)
                       VALUES(%s,%s,%s,ARRAY[]::TEXT[],TRUE,%s,NOW())
                       RETURNING id""",
                    (class_id, kind, _SINF_GURUH_TIZIM_NOMLARI[kind], actor_id),
                )
                system_id = int(cur.fetchone()["id"])
                _sinf_guruh_tizimini_taqsimla(cur, system_id)
                created += 1

            cur.execute(
                "SELECT turi FROM maktab_sinf_guruh_tizimlari "
                "WHERE sinf_id=%s AND faol=TRUE ORDER BY id",
                (class_id,),
            )
            active_types = [row["turi"] for row in cur.fetchall()]
            legacy = (
                active_types[0] if len(active_types) == 1
                else ("manual" if active_types else "none")
            )
            cur.execute(
                "UPDATE maktab_sinflari SET guruhlash_usuli=%s WHERE id=%s",
                (legacy, class_id),
            )

        cur.execute(
            "UPDATE aqlli_jadval_urinishlari_v2 SET holat='bekor' "
            "WHERE maktab_id=%s AND holat='draft'",
            (sorov.maktab_id,),
        )
        cur.execute(
            """SELECT t.id,t.sinf_id,t.turi,t.nomi,t.fanlar,t.faol
                 FROM maktab_sinf_guruh_tizimlari t
                WHERE t.sinf_id=ANY(%s) AND t.faol=TRUE
                ORDER BY t.sinf_id,t.id""",
            (class_ids,),
        )
        active = [dict(row) for row in cur.fetchall()]
        conn.commit()
        return {
            "holat": "saqlandi",
            "qoshildi": created,
            "qayta_faollashtirildi": restored,
            "oldingi_tizimlar_saqlandi": True,
            "sinflar": class_ids,
            "tizimlar": active,
        }
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


# ========================= V18.76 QUICK/ADDITIVE GROUP SYSTEM END =========================


# ═══════════════════════════════════════════════════════════
# V19.2 — O'QITUVCHI-ASOSLI YUKLAMA + AQILLI ALMASHTIRISH
#
# Bitta kanonik qator:
#   o'qituvchi + fan + sinf + aniq guruh + haftalik soat.
#
# Excel import va saytdagi qo'lda kiritish aynan bir xil
# maktab_dars_birikmalari manbasiga yozadi. Shuning uchun bir o'qituvchi
# Fizika, Astronomiya va Iqtisoddan turli sinf/guruhlarga kirsa ham fanlar
# aralashmaydi. Sinf reja soati guruhlar soniga ko'paymaydi, o'qituvchi
# yuklamasi esa har bir guruh bo'yicha alohida yig'iladi.
# ═══════════════════════════════════════════════════════════


SAMTM_V19_2_AUTO_SWAP_DEFAULT = str(
    os.getenv("SAMTM_V19_2_AUTO_SWAP_DEFAULT", "true")
).strip().lower() not in {"0", "false", "no", "off"}
SAMTM_V19_2_SWAP_SUGGESTION_LIMIT = max(
    1, min(60, int(os.getenv("SAMTM_V19_2_SWAP_SUGGESTION_LIMIT", "24")))
)


def _v192_tables(cur):
    _v1876_tables(cur)
    cur.execute(
        "ALTER TABLE users ADD COLUMN IF NOT EXISTS tugilgan_sana DATE"
    )
    cur.execute(
        "ALTER TABLE users ADD COLUMN IF NOT EXISTS mutaxassisligi TEXT"
    )
    cur.execute(
        "ALTER TABLE users ADD COLUMN IF NOT EXISTS haftalik_maqsad_soat INTEGER"
    )
    cur.execute(
        "ALTER TABLE users ADD COLUMN IF NOT EXISTS jadval_raqami INTEGER"
    )
    cur.execute(
        "CREATE UNIQUE INDEX IF NOT EXISTS uq_users_maktab_jadval_raqami "
        "ON users(maktab_id,jadval_raqami) WHERE jadval_raqami IS NOT NULL"
    )
    cur.execute(
        "ALTER TABLE maktab_dars_birikmalari "
        "ADD COLUMN IF NOT EXISTS xona_id BIGINT REFERENCES aqlli_xonalar_v2(id) ON DELETE SET NULL"
    )
    cur.execute(
        "ALTER TABLE maktab_dars_birikmalari "
        "ADD COLUMN IF NOT EXISTS yangilangan_at TIMESTAMPTZ NOT NULL DEFAULT NOW()"
    )
    # O'quv rejada 0,5 va 1,5 soatli fanlar bor. O'qituvchi yuklamasi
    # ham aynan shu aniqlikda saqlanishi shart; aks holda 22/22 bo'lgan
    # yuklama ichidagi bitta 1,5 qator POST validatsiyasida yiqiladi.
    cur.execute("""ALTER TABLE maktab_dars_birikmalari
                   ALTER COLUMN haftalik_soat TYPE NUMERIC(5,1)
                   USING haftalik_soat::NUMERIC(5,1)""")
    cur.execute("""ALTER TABLE users
                   ALTER COLUMN haftalik_dars_soati TYPE NUMERIC(5,1)
                   USING haftalik_dars_soati::NUMERIC(5,1)""")
    cur.execute("""ALTER TABLE users
                   ALTER COLUMN haftalik_maqsad_soat TYPE NUMERIC(5,1)
                   USING haftalik_maqsad_soat::NUMERIC(5,1)""")
    cur.execute("""CREATE TABLE IF NOT EXISTS aqlli_jadval_boshqaruv_v19_2(
        maktab_id INTEGER PRIMARY KEY REFERENCES maktablar(id) ON DELETE CASCADE,
        avtomatik_tavsiya BOOLEAN NOT NULL DEFAULT TRUE,
        yangilagan_user_id BIGINT,
        yangilangan_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
    )""")
    cur.execute("""CREATE TABLE IF NOT EXISTS aqlli_oquv_reja_holati_v19_3(
        maktab_id INTEGER PRIMARY KEY REFERENCES maktablar(id) ON DELETE CASCADE,
        holat TEXT NOT NULL DEFAULT 'draft'
            CHECK(holat IN ('draft','tasdiqlangan')),
        versiya INTEGER NOT NULL DEFAULT 1,
        tasdiqlagan_user_id BIGINT,
        tasdiqlangan_at TIMESTAMPTZ,
        yangilangan_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
    )""")
    cur.execute("""CREATE TABLE IF NOT EXISTS aqlli_oquv_reja_til_holati_v238(
        maktab_id INTEGER NOT NULL REFERENCES maktablar(id) ON DELETE CASCADE,
        talim_tili TEXT NOT NULL,
        holat TEXT NOT NULL DEFAULT 'draft' CHECK(holat IN ('draft','tasdiqlangan')),
        versiya INTEGER NOT NULL DEFAULT 1,
        tasdiqlagan_user_id BIGINT,
        tasdiqlangan_at TIMESTAMPTZ,
        yangilangan_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
        PRIMARY KEY(maktab_id,talim_tili)
    )""")
    cur.execute("""CREATE TABLE IF NOT EXISTS aqlli_oquv_reja_qatorlari_v19_3(
        id BIGSERIAL PRIMARY KEY,
        maktab_id INTEGER NOT NULL REFERENCES maktablar(id) ON DELETE CASCADE,
        sinf_id INTEGER NOT NULL REFERENCES maktab_sinflari(id) ON DELETE CASCADE,
        fan_nomi TEXT NOT NULL,
        haftalik_soat NUMERIC(4,1) NOT NULL CHECK(haftalik_soat BETWEEN 0.5 AND 20),
        kunlik_max INTEGER NOT NULL DEFAULT 1 CHECK(kunlik_max BETWEEN 1 AND 4),
        manba TEXT NOT NULL DEFAULT 'qolda',
        yangilangan_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
        UNIQUE(maktab_id,sinf_id,fan_nomi)
    )""")
    # 2026–2027 tayanch rejada 0,5 va 1,5 soatli fanlar bor. Eski INTEGER
    # ustunni ma'lumotni yo'qotmasdan kasr soat saqlaydigan turga o'tkazamiz.
    cur.execute("""DO $$
        DECLARE constraint_name TEXT;
        BEGIN
          FOR constraint_name IN
            SELECT c.conname
            FROM pg_constraint c
            JOIN pg_class t ON t.oid=c.conrelid
            WHERE t.relname='aqlli_oquv_reja_qatorlari_v19_3'
              AND c.contype='c'
              AND pg_get_constraintdef(c.oid) ILIKE '%haftalik_soat%'
          LOOP
            EXECUTE format(
              'ALTER TABLE aqlli_oquv_reja_qatorlari_v19_3 DROP CONSTRAINT %I',
              constraint_name
            );
          END LOOP;
        END $$""")
    cur.execute("""ALTER TABLE aqlli_oquv_reja_qatorlari_v19_3
                   ALTER COLUMN haftalik_soat TYPE NUMERIC(4,1)
                   USING haftalik_soat::NUMERIC(4,1)""")
    cur.execute("""DO $$ BEGIN
        ALTER TABLE aqlli_oquv_reja_qatorlari_v19_3
          ADD CONSTRAINT aqlli_oquv_reja_v193_haftalik_soat_check
          CHECK(haftalik_soat BETWEEN 0.5 AND 20);
      EXCEPTION WHEN duplicate_object THEN NULL; END $$""")
    cur.execute("""ALTER TABLE aqlli_sinf_fan_yuklamalari_v2
                   ALTER COLUMN haftalik_soat TYPE NUMERIC(4,1)
                   USING haftalik_soat::NUMERIC(4,1)""")
    cur.execute(
        "CREATE INDEX IF NOT EXISTS idx_v193_plan_class "
        "ON aqlli_oquv_reja_qatorlari_v19_3(maktab_id,sinf_id)"
    )
    cur.execute("""CREATE TABLE IF NOT EXISTS aqlli_jadval_ozgarish_log_v19_2(
        id BIGSERIAL PRIMARY KEY,
        maktab_id INTEGER NOT NULL REFERENCES maktablar(id) ON DELETE CASCADE,
        urinish_id BIGINT NOT NULL REFERENCES aqlli_jadval_urinishlari_v2(id) ON DELETE CASCADE,
        manba_slot_id BIGINT,
        eski_hafta_kuni INTEGER NOT NULL,
        eski_dars_raqami INTEGER NOT NULL,
        yangi_hafta_kuni INTEGER NOT NULL,
        yangi_dars_raqami INTEGER NOT NULL,
        turi TEXT NOT NULL CHECK(turi IN ('kochirish','almashtirish')),
        bajargan_user_id BIGINT,
        yaratilgan_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
    )""")
    cur.execute(
        "CREATE INDEX IF NOT EXISTS idx_v192_change_log_run "
        "ON aqlli_jadval_ozgarish_log_v19_2(urinish_id,yaratilgan_at DESC)"
    )


@app.on_event("startup")
def _v192_startup_tables():
    try:
        conn = _db(); cur = conn.cursor()
        _v192_tables(cur)
        conn.commit(); cur.close(); conn.close()
        app.state.samtm_fractional_hours_ready = True
        app.state.samtm_fractional_hours_error = None
    except Exception as exc:
        app.state.samtm_fractional_hours_ready = False
        app.state.samtm_fractional_hours_error = type(exc).__name__
        print(f"[V19.8 0,5/1,5 soat migratsiyasi] {exc}", flush=True)


@app.get("/api/maktab/aqlli_jadval/v3/soat_imkoniyatlari")
def v197_fractional_hour_capabilities():
    """Frontend saqlashdan oldin aynan yangi backend ishlayotganini tekshiradi."""
    schema_ready = bool(
        getattr(app.state, "samtm_fractional_hours_ready", False)
    )
    return {
        # Eski V19.7 frontend aynan shu qiymatni tekshiradi. Platformaning
        # haqiqiy yangi versiyasi alohida qaytariladi — backendni birinchi
        # deploy qilganda foydalanuvchi V19.7 frontend bilan ham saqlay oladi.
        "release": "samtm-fractional-hours-ab-week-v19.7",
        "platform_release": SAMTM_SCHOOL_RELEASE,
        "jadval_release": SAMTM_JADVAL_RELEASE,
        "exact_jadval_release": SAMTM_EXACT_JADVAL_RELEASE,
        "exact_module_release": _V230_EXACT_MODULE_RELEASE,
        "exact_internal_release": SAMTM_EXACT_INTERNAL_RELEASE,
        "timetable_engine_release": SAMTM_TIMETABLE_ENGINE_RELEASE,
        "schedule_runtime_release": _V234_RUNTIME_RELEASE,
        "single_generator": True,
        "generator_soni": 1,
        "generator_turi": "yagona-exact-cp-sat",
        "generator_nomi": "Yagona kuchli generator",
        "exact_engine_ready": bool(_V216_ORTOOLS_AVAILABLE),
        "diagnostics_contract": "exact-failure-v21.9",
        "solver_pipeline": "hard-feasibility-first",
        "exact_engine": "google-ortools-cp-sat",
        "required_dependency": "ortools>=9.15,<9.16",
        "generation_budget_seconds": _v220_generation_budget_seconds(),
        "generator": _timetable_mode_config(),
        "generator_rejimlari": _timetable_public_modes(),
        "fractional_hours": schema_ready,
        "fraction_step": 0.5,
        "ab_week": schema_ready,
        "schema_ready": schema_ready,
        "example": {
            "geografiya": 1.5,
            "iqtisodiy_bilim_asoslari": 0.5,
            "ikki_haftalik_slotlar": [
                {"hafta": "toq", "geografiya": 2, "iqtisod": 0},
                {"hafta": "juft", "geografiya": 1, "iqtisod": 1},
            ],
        },
    }


SAMTM_SCHOOL_SETUP_RELEASE = "SAMTM-SCHOOL-SETUP-V23.7-XLSX-SHIFT-ALPHABET"

# Sinfning raqam qismi alohida ``maktab_sinflari.sinf`` ustunida 1..11 bo'lib
# qoladi. Quyidagi ketma-ketliklar faqat parallel yorlig'i (``harf``) uchun.
# Bu eski generatorlardagi ``sinf::int`` tartiblashini buzmasdan lotin/kirill
# tanlovini va ikki smenada uzluksiz davom etishni beradi.
_V237_CLASS_ALPHABETS = {
    "latin_xalqaro": tuple("ABCDEFGHIJKLMNOPQRSTUVWXYZ"),
    "uzbek_lotin": (
        "A", "B", "D", "E", "F", "G", "G‘", "H", "I", "J", "K", "L",
        "M", "N", "O", "O‘", "P", "Q", "R", "S", "T", "U", "V", "X",
        "Y", "Z", "Sh", "Ch", "Ng",
    ),
    "uzbek_kiril": (
        "А", "Б", "В", "Г", "Д", "Е", "Ё", "Ж", "З", "И", "Й", "К",
        "Л", "М", "Н", "О", "П", "Р", "С", "Т", "У", "Ф", "Х", "Ц",
        "Ч", "Ш", "Ъ", "Ь", "Э", "Ю", "Я", "Ў", "Қ", "Ғ", "Ҳ",
    ),
}


_V238_INSTRUCTION_LANGUAGES = {
    "uz": "O‘zbek tili",
    "ru": "Rus tili",
    "en": "Ingliz tili",
}
_V238_INSTRUCTION_LANGUAGE_ALIASES = {
    "uz": "uz", "uzb": "uz", "uzbek": "uz", "o'zbek": "uz",
    "o‘zbek": "uz", "ўзбек": "uz",
    "ru": "ru", "rus": "ru", "russian": "ru", "рус": "ru",
    "русский": "ru",
    "en": "en", "eng": "en", "english": "en", "ingliz": "en",
}


def _v238_normalize_instruction_language(value, default="uz"):
    """Sinf/o‘quv reja tilini bitta qat’iy DB kalitiga keltiradi."""
    raw = re.sub(r"\s+", " ", str(value or "")).strip().casefold()
    if not raw:
        raw = str(default or "uz").strip().casefold()
    normalized = _V238_INSTRUCTION_LANGUAGE_ALIASES.get(raw)
    if normalized is None:
        raise ValueError(
            "Ta’lim tili noto‘g‘ri. Faqat uz (o‘zbek), ru (rus) yoki en (ingliz) tanlanadi."
        )
    return normalized

_V237_APOSTROPHES = str.maketrans({
    "ʻ": "'", "ʼ": "'", "’": "'", "‘": "'", "`": "'", "´": "'",
    "ʹ": "'", "＇": "'",
})


def _v237_clean_parallel_label(value):
    label = re.sub(r"\s+", " ", str(value or "")).strip()
    if not 1 <= len(label) <= 40:
        raise ValueError("Sinf parallel nomi 1–40 ta belgi bo‘lishi kerak.")
    if any(_v237_unicodedata.category(char).startswith("C") for char in label):
        raise ValueError("Sinf parallel nomida boshqaruv/yashirin belgilar bo‘lishi mumkin emas.")
    return label


def _v237_parallel_label_key(value):
    label = _v237_clean_parallel_label(value).translate(_V237_APOSTROPHES)
    label = _v237_unicodedata.normalize("NFKC", label)
    return re.sub(r"\s+", " ", label).strip().casefold()


def _v237_class_sort_key(row, alphabet_type="latin_xalqaro"):
    try:
        grade = int(str((row or {}).get("sinf") or "").strip())
    except (TypeError, ValueError):
        grade = 999
    alphabet = _V237_CLASS_ALPHABETS.get(
        str(alphabet_type or "latin_xalqaro"),
        _V237_CLASS_ALPHABETS["latin_xalqaro"],
    )
    indexes = {_v237_parallel_label_key(label): index for index, label in enumerate(alphabet)}
    raw_label = str((row or {}).get("harf") or "")
    try:
        label_key = _v237_parallel_label_key(raw_label)
    except ValueError:
        label_key = raw_label.casefold()
    natural = tuple(
        (0, int(part)) if part.isdigit() else (1, part.casefold())
        for part in re.split(r"(\d+)", raw_label)
        if part != ""
    )
    return (grade, indexes.get(label_key, len(alphabet)), natural, int((row or {}).get("id") or 0))


class V237SchoolClassPlan(BaseModel):
    sinf: int
    birinchi_smena: int = 0
    ikkinchi_smena: int = 0
    talim_tili: str = "uz"


def _v237_materialize_class_plan(raw_plan, shift_count, alphabet_type):
    try:
        shift_count = int(shift_count or 1)
    except (TypeError, ValueError) as exc:
        raise ValueError("Smena soni 1 yoki 2 bo‘lishi kerak.") from exc
    if shift_count not in (1, 2):
        raise ValueError("Smena soni 1 yoki 2 bo‘lishi kerak.")
    alphabet_type = str(alphabet_type or "latin_xalqaro").strip().lower()
    alphabet = _V237_CLASS_ALPHABETS.get(alphabet_type)
    if alphabet is None:
        raise ValueError(
            "Alifbo turi latin_xalqaro, uzbek_lotin yoki uzbek_kiril bo‘lishi kerak."
        )
    sources = list(raw_plan or [])
    if not sources:
        return []
    seen_grade_languages = set()
    normalized_sources = []
    for source in sources:
        if isinstance(source, dict):
            item = dict(source)
        else:
            dump = getattr(source, "model_dump", None)
            item = dump() if callable(dump) else source.dict()
        try:
            grade = int(item.get("sinf"))
            first_count = int(item.get("birinchi_smena") or 0)
            second_count = int(item.get("ikkinchi_smena") or 0)
            instruction_language = _v238_normalize_instruction_language(
                item.get("talim_tili")
            )
        except (TypeError, ValueError) as exc:
            raise ValueError("Sinf va parallel sonlari butun raqam bo‘lishi kerak.") from exc
        if grade not in range(1, 12):
            raise ValueError("Sinf darajasi 1–11 oralig‘ida bo‘lishi kerak.")
        pair = (grade, instruction_language)
        if pair in seen_grade_languages:
            raise ValueError(
                f"{grade}-sinf / {_V238_INSTRUCTION_LANGUAGES[instruction_language]} "
                "rejasi ikki marta yuborilgan."
            )
        seen_grade_languages.add(pair)
        if first_count < 0 or second_count < 0:
            raise ValueError(f"{grade}-sinf parallel soni manfiy bo‘lishi mumkin emas.")
        if shift_count == 1 and second_count:
            raise ValueError(
                f"{grade}-sinf uchun 2-smena tanlangan, lekin maktab 1 smenali."
            )
        normalized_sources.append({
            "sinf": grade,
            "talim_tili": instruction_language,
            "birinchi_smena": first_count,
            "ikkinchi_smena": second_count,
        })

    materialized = []
    language_order = {"uz": 0, "ru": 1, "en": 2}
    grades = sorted({row["sinf"] for row in normalized_sources})
    for grade in grades:
        grade_rows = sorted(
            (row for row in normalized_sources if row["sinf"] == grade),
            key=lambda row: language_order[row["talim_tili"]],
        )
        total = sum(
            row["birinchi_smena"] + row["ikkinchi_smena"]
            for row in grade_rows
        )
        if total > len(alphabet):
            raise ValueError(
                f"{grade}-sinf uchun {total} ta parallel tanlangan; "
                f"{alphabet_type} alifbosida ko‘pi bilan {len(alphabet)} ta yorliq bor."
            )
        # Bir xil darajadagi UZ -> RU -> EN ketma-ketligi saqlanadi. Shu bilan
        # birga barcha 1-smena harflari tugagach 2-smena davom etadi; masalan
        # UZ(A), RU(B), EN(C), keyin 2-smena UZ(D)...
        next_index = 0
        for shift_key, shift in (("birinchi_smena", 1), ("ikkinchi_smena", 2)):
            for row in grade_rows:
                for _ in range(int(row[shift_key])):
                    label = alphabet[next_index]
                    materialized.append({
                        "sinf": grade,
                        "harf": label,
                        "smena": shift,
                        "name": f"{grade}-{label}",
                        "talim_tili": row["talim_tili"],
                        "alifbo_turi": alphabet_type,
                        "alifbo_tartibi": next_index,
                    })
                    next_index += 1
    if not materialized:
        raise ValueError("Sinf rejasida kamida bitta parallel sonini kiriting.")
    return sorted(materialized, key=lambda row: (row["sinf"], row["alifbo_tartibi"]))


def _v237_school_retry_config_mismatches(
    existing, shift_count, alphabet_type, region, district
):
    """Bir xil nomli tezkor retry faqat aynan o'sha maktab bo'lsa qabul qilinadi."""
    row = dict(existing or {})

    def normalized(value):
        return re.sub(r"\s+", " ", str(value or "")).strip().casefold()

    mismatches = []
    try:
        stored_shift_count = int(row.get("smena_soni") or 1)
    except (TypeError, ValueError):
        stored_shift_count = 1
    if stored_shift_count != int(shift_count or 1):
        mismatches.append("smena soni")
    if normalized(row.get("alifbo_turi") or "latin_xalqaro") != normalized(
        alphabet_type or "latin_xalqaro"
    ):
        mismatches.append("alifbo")
    if normalized(row.get("viloyat")) != normalized(region):
        mismatches.append("viloyat")
    if normalized(row.get("tuman")) != normalized(district):
        mismatches.append("tuman")
    return mismatches


def _v237_school_retry_class_plan_matches(expected_rows, actual_rows):
    """Retry faqat saqlangan va yuborilgan sinf rejalari aynan teng bo'lsa xavfsiz."""
    def keys(rows):
        return {
            (
                str(row["sinf"]),
                _v237_parallel_label_key(row["harf"]),
                int(row["smena"]),
                _v238_normalize_instruction_language(row.get("talim_tili")),
            )
            for row in (rows or [])
        }

    try:
        return keys(expected_rows) == keys(actual_rows)
    except (KeyError, TypeError, ValueError):
        return False


class V198SchoolWorkspaceLinkRequest(BaseModel):
    """V17 muassasasini eski maktab workspace'iga xavfsiz bog'lash so'rovi.

    Frontend tanlangan muassasa IDlarini yuboradi. Eski, xato holatdagi
    frontend umuman ID yubormasa ham joriy foydalanuvchining eng so'nggi
    V17 maktabi topilib tiklanadi.
    """

    organization_v17_id: Optional[int] = None
    context_id: Optional[int] = None
    # Mavjud maktablar ro'yxatidan kirilganda frontend haqiqiy legacy IDni
    # aynan ``maktab_id`` nomi bilan yuboradi. Avval bu maydon qabul
    # qilinmagani uchun mavjud maktab ham yangi maktab oqimiga tushib qolardi.
    maktab_id: Optional[int] = None
    # REV48: productiondagi eski frontend ``selected_id`` yuboradi. Uni
    # tashlab yubormaymiz: agar bu ID foydalanuvchiga tegishli haqiqiy
    # maktab bo'lsa, V17 jadvallariga kirmasdan bevosita shu maktab ochiladi.
    selected_id: Optional[int] = None
    existing_only: bool = False
    create_new: bool = False
    nomi: Optional[str] = None
    viloyat: Optional[str] = None
    tuman: Optional[str] = None
    smena_soni: int = 1
    alifbo_turi: str = "latin_xalqaro"
    sinf_rejasi: list[V237SchoolClassPlan] = []


def _v198_existing_school_for_user(
    cur,
    user_id: int,
    preferred_id: Optional[int] = None,
    preferred_name: Optional[str] = None,
):
    """Mavjud legacy maktabni xavfsiz va takror yaratmasdan qaytaradi.

    Tanlangan ro'yxat qatori haqiqiy ``maktab_id`` yuborsa avval shu ID
    tekshiriladi. ID yo'qolgan eski V17 yozuvlarida esa ayni nomdagi va joriy
    foydalanuvchiga tegishli maktab olinadi. Begona maktab faqat umumiy admin
    bo'lsa ochiladi.
    """
    cur.execute("SELECT 1 FROM admin_akkaunt WHERE uid=%s", (user_id,))
    is_admin = bool(cur.fetchone())

    candidate_id = 0
    try:
        candidate_id = int(preferred_id or 0)
    except (TypeError, ValueError):
        candidate_id = 0
    if candidate_id > 0:
        cur.execute(
            """SELECT m.id AS maktab_id,m.nomi AS maktab_nomi
                 FROM maktablar m
                WHERE m.id=%s
                  AND (
                    %s OR m.direktor_user_id=%s
                    OR EXISTS(
                      SELECT 1 FROM users u
                       WHERE u.user_id=%s AND u.maktab_id=m.id
                    )
                    OR EXISTS(
                      SELECT 1 FROM foydalanuvchi_muassasalari fm
                       WHERE fm.user_id=%s
                         AND fm.muassasa_turi='maktab'
                         AND fm.muassasa_id=m.id
                    )
                  )
                LIMIT 1""",
            (candidate_id, is_admin, user_id, user_id, user_id),
        )
        row = cur.fetchone()
        if row:
            return row

    school_name = re.sub(r"\s+", " ", str(preferred_name or "")).strip()
    if school_name:
        cur.execute(
            """SELECT m.id AS maktab_id,m.nomi AS maktab_nomi
                 FROM maktablar m
                WHERE lower(trim(m.nomi))=lower(trim(%s))
                  AND (
                    m.direktor_user_id=%s
                    OR EXISTS(
                      SELECT 1 FROM users u
                       WHERE u.user_id=%s AND u.maktab_id=m.id
                    )
                    OR EXISTS(
                      SELECT 1 FROM foydalanuvchi_muassasalari fm
                       WHERE fm.user_id=%s
                         AND fm.muassasa_turi='maktab'
                         AND fm.muassasa_id=m.id
                    )
                  )
                ORDER BY
                  CASE WHEN m.direktor_user_id=%s THEN 0 ELSE 1 END,
                  m.id DESC
                LIMIT 1""",
            (school_name, user_id, user_id, user_id, user_id),
        )
        row = cur.fetchone()
        if row:
            return row

    cur.execute(
        """SELECT m.id AS maktab_id,m.nomi AS maktab_nomi
             FROM users u JOIN maktablar m ON m.id=u.maktab_id
            WHERE u.user_id=%s LIMIT 1""",
        (user_id,),
    )
    row = cur.fetchone()
    if row:
        return row
    cur.execute(
        """SELECT m.id AS maktab_id,m.nomi AS maktab_nomi
             FROM foydalanuvchi_muassasalari fm
             JOIN maktablar m ON m.id=fm.muassasa_id
            WHERE fm.user_id=%s AND fm.muassasa_turi='maktab'
            ORDER BY fm.id DESC LIMIT 1""",
        (user_id,),
    )
    return cur.fetchone()


@app.post("/api/maktab/aqlli_jadval/v3/maktab_workspace_boglash")
def v198_link_school_workspace(
    sorov: V198SchoolWorkspaceLinkRequest,
    token: str,
):
    """Yangi yaratilgan V17 maktabga haqiqiy ``maktablar.id`` beradi.

    Oldingi frontend ``learning_contexts.id`` ni ``maktab_id`` deb yuborgan,
    holbuki maktab dashboardi faqat ``maktablar.id`` bilan ishlaydi. Ushbu
    endpoint organization row'ni bloklab, bir martalik legacy maktab yaratadi,
    external_id ni yozadi va direktor a'zoligini atomar saqlaydi. Bir necha
    marta chaqirilsa yangi maktab ko'paymaydi.
    """
    user_id = _jwt_tekshir(token)
    conn = _db()
    cur = conn.cursor()
    try:
        _maktab_jadvali(cur)
        _muassasa_jadvali(cur)

        # REV55: yangi maktabga o'tishdan oldin joriy eski maktabni
        # ko'p-muassasali jadvalga kafolatli yozamiz. Keyingi UPDATE faqat
        # faol workspace ko'rsatkichini almashtiradi; eski maktab va uning
        # ma'lumotlari hamda a'zoligi hech qachon o'chirilmaydi.
        cur.execute(
            "SELECT maktab_id,lavozim FROM users WHERE user_id=%s FOR UPDATE",
            (user_id,),
        )
        current_user = cur.fetchone()
        previous_school_id = (
            int(current_user["maktab_id"])
            if current_user and current_user.get("maktab_id")
            else None
        )
        if previous_school_id:
            cur.execute(
                """INSERT INTO foydalanuvchi_muassasalari(
                       user_id,muassasa_turi,muassasa_id,lavozim
                   ) VALUES(%s,'maktab',%s,%s)
                   ON CONFLICT(user_id,muassasa_turi,muassasa_id)
                   DO UPDATE SET lavozim=EXCLUDED.lavozim""",
                (
                    user_id,
                    previous_school_id,
                    str(current_user.get("lavozim") or "direktor"),
                ),
            )

        # Eski va yangi frontendlar bilan bir xil ishlaydi. Muhim jihat:
        # mavjud maktabni topish V17 schema/helper so'rovlaridan OLDIN
        # bajariladi. Shuning uchun organization_trials/learning_contexts
        # vaqtincha xato qilsa ham eski maktab yangi yaratish oynasiga
        # tushmaydi.
        preferred_school_id = sorov.maktab_id or sorov.selected_id
        requested_school = None
        if preferred_school_id is not None or sorov.existing_only:
            requested_school = _v198_existing_school_for_user(
                cur,
                user_id,
                preferred_id=preferred_school_id,
                preferred_name=sorov.nomi,
            )
        if requested_school:
            maktab_id = int(requested_school["maktab_id"])
            cur.execute(
                """UPDATE users
                      SET maktab_id=%s,
                          lavozim=COALESCE(NULLIF(lavozim,''),'direktor')
                    WHERE user_id=%s""",
                (maktab_id, user_id),
            )
            conn.commit()
            return {
                "holat": (
                    "mavjud_tiklandi"
                    if sorov.existing_only
                    else "mavjud_id_rev48"
                ),
                "maktab_id": maktab_id,
                "maktab_nomi": requested_school["maktab_nomi"],
                "legacy_yaratildi": False,
                "frontend_mosligi": (
                    "selected_id"
                    if sorov.maktab_id is None and sorov.selected_id is not None
                    else "maktab_id"
                ),
            }
        if sorov.existing_only:
            raise HTTPException(
                status_code=404,
                detail="Sizga tegishli mavjud maktab topilmadi.",
            )

        cur.execute(
            """SELECT to_regclass('public.organization_trials') AS trials,
                      to_regclass('public.learning_contexts') AS contexts,
                      to_regclass('public.context_memberships') AS memberships"""
        )
        tables = cur.fetchone() or {}
        organization = None
        explicit_new_school = bool(
            sorov.create_new
            and sorov.organization_v17_id is None
            and sorov.context_id is None
        )
        if tables.get("trials") and tables.get("contexts") and not explicit_new_school:
            access_parts = [
                "EXISTS(SELECT 1 FROM admin_akkaunt aa WHERE aa.uid=%s)",
                "o.creator_user_id=%s",
                "c.owner_user_id=%s",
            ]
            params = [user_id, user_id, user_id]
            if tables.get("memberships"):
                access_parts.append(
                    """EXISTS(
                         SELECT 1 FROM context_memberships cm
                          WHERE cm.context_id=o.context_id
                            AND cm.user_id=%s
                            AND cm.status='active'
                            AND cm.member_role IN ('owner','manager','director','administrator')
                       )"""
                )
                params.append(user_id)
            filters = [f"({' OR '.join(access_parts)})", "o.organization_type='school'"]
            if sorov.organization_v17_id is not None:
                filters.append("o.id=%s")
                params.append(int(sorov.organization_v17_id))
            if sorov.context_id is not None:
                filters.append("o.context_id=%s")
                params.append(int(sorov.context_id))
            cur.execute(
                f"""SELECT o.id AS organization_v17_id,o.context_id,
                           o.display_name,o.lifecycle_status,c.external_id
                      FROM organization_trials o
                      JOIN learning_contexts c ON c.id=o.context_id
                     WHERE {' AND '.join(filters)}
                     ORDER BY o.id DESC LIMIT 1
                     FOR UPDATE OF o,c""",
                tuple(params),
            )
            organization = cur.fetchone()

        if organization is None:
            if sorov.organization_v17_id is not None or sorov.context_id is not None:
                # V17 bog'lanish yozuvi yo'qolgan bo'lsa mavjud legacy maktabni
                # nomi yoki foydalanuvchining faol maktabi orqali tiklaymiz.
                # Bu holatda HECH QACHON yangi maktab yaratmaymiz.
                existing = _v198_existing_school_for_user(
                    cur,
                    user_id,
                    preferred_id=sorov.maktab_id or sorov.selected_id,
                    preferred_name=sorov.nomi,
                )
                if existing:
                    maktab_id = int(existing["maktab_id"])
                    cur.execute(
                        """UPDATE users
                              SET maktab_id=%s,
                                  lavozim=COALESCE(NULLIF(lavozim,''),'direktor')
                            WHERE user_id=%s""",
                        (maktab_id, user_id),
                    )
                    conn.commit()
                    return {
                        "holat": "mavjud_tiklandi",
                        "maktab_id": maktab_id,
                        "maktab_nomi": existing["maktab_nomi"],
                        "legacy_yaratildi": False,
                    }
                raise HTTPException(
                    status_code=404,
                    detail="Tanlangan maktab topilmadi yoki u sizga tegishli emas.",
                )
            if sorov.create_new:
                school_name = str(sorov.nomi or "").strip()
                if not school_name:
                    raise HTTPException(
                        status_code=400,
                        detail="Yangi maktab nomini kiriting.",
                    )
                if int(sorov.smena_soni or 1) not in (1, 2):
                    raise HTTPException(
                        status_code=400,
                        detail="Smena soni 1 yoki 2 bo‘lishi kerak.",
                    )
                try:
                    planned_classes = _v237_materialize_class_plan(
                        sorov.sinf_rejasi,
                        int(sorov.smena_soni or 1),
                        sorov.alifbo_turi,
                    )
                except ValueError as exc:
                    raise HTTPException(status_code=400, detail=str(exc)) from exc
                alphabet_type = str(sorov.alifbo_turi or "latin_xalqaro").strip().lower()
                _maktab_sinflari_jadvali(cur)
                _v1852_tables(cur)

                # Bir foydalanuvchining ikki marta tez bosishi ikki maktab
                # yaratmasin. Users qatori transaction tugaguncha bloklanadi;
                # ayni nomdagi juda yangi yozuv xavfsiz retry deb olinadi.
                cur.execute(
                    "SELECT user_id FROM users WHERE user_id=%s FOR UPDATE",
                    (user_id,),
                )
                if not cur.fetchone():
                    raise HTTPException(status_code=404, detail="Foydalanuvchi topilmadi.")
                cur.execute(
                    """SELECT id,nomi,COALESCE(smena_soni,1) AS smena_soni,
                              COALESCE(alifbo_turi,'latin_xalqaro') AS alifbo_turi,
                              COALESCE(viloyat,'') AS viloyat,
                              COALESCE(tuman,'') AS tuman
                         FROM maktablar
                        WHERE direktor_user_id=%s
                          AND lower(trim(nomi))=lower(trim(%s))
                          AND yaratilgan_at >= NOW() - INTERVAL '2 minutes'
                        ORDER BY id DESC LIMIT 1
                        FOR UPDATE""",
                    (user_id, school_name),
                )
                linked_school = cur.fetchone()
                created = linked_school is None
                if linked_school:
                    config_mismatches = _v237_school_retry_config_mismatches(
                        linked_school,
                        int(sorov.smena_soni or 1),
                        alphabet_type,
                        sorov.viloyat,
                        sorov.tuman,
                    )
                    if config_mismatches:
                        raise HTTPException(
                            status_code=409,
                            detail=(
                                "Aynan shu nomli maktab hozirgina boshqa sozlama bilan "
                                "yaratildi (" + ", ".join(config_mismatches) + "). "
                                "Mavjud maktabni oching yoki boshqa nom bilan yarating."
                            ),
                        )
                    maktab_id = int(linked_school["id"])
                    school_name = str(linked_school["nomi"])
                else:
                    cur.execute(
                        """INSERT INTO maktablar(
                               nomi,viloyat,tuman,smena_soni,direktor_user_id,
                               pulli,oylik_tolov,alifbo_turi
                           ) VALUES(%s,%s,%s,%s,%s,FALSE,NULL,%s)
                           RETURNING id""",
                        (
                            school_name,
                            str(sorov.viloyat or "").strip() or None,
                            str(sorov.tuman or "").strip() or None,
                            int(sorov.smena_soni or 1),
                            user_id,
                            alphabet_type,
                        ),
                    )
                    maktab_id = int(cur.fetchone()["id"])
                cur.execute(
                    """SELECT id,sinf,harf,COALESCE(smena,1) AS smena,
                              COALESCE(talim_tili,'uz') AS talim_tili
                         FROM maktab_sinflari WHERE maktab_id=%s
                         ORDER BY id FOR UPDATE""",
                    (maktab_id,),
                )
                current_classes = [dict(row) for row in cur.fetchall()]
                class_payload = []
                if not created and not _v237_school_retry_class_plan_matches(
                    planned_classes, current_classes
                ):
                    raise HTTPException(
                        status_code=409,
                        detail=(
                            "Aynan shu maktab hozirgina boshqa sinf rejasi bilan yaratildi. "
                            "Mavjud maktabni ochib sinflarni tahrirlang."
                        ),
                    )
                if current_classes:
                    class_payload = [
                        {
                            "id": int(row["id"]),
                            "name": f"{row['sinf']}-{row['harf']}",
                            "sinf": str(row["sinf"]),
                            "harf": row["harf"],
                            "smena": int(row["smena"]),
                            "talim_tili": _v238_normalize_instruction_language(
                                row.get("talim_tili")
                            ),
                        }
                        for row in sorted(
                            current_classes,
                            key=lambda item: _v237_class_sort_key(item, alphabet_type),
                        )
                    ]
                elif created and planned_classes:
                    cur.execute(
                        "UPDATE maktablar SET alifbo_turi=%s WHERE id=%s",
                        (alphabet_type, maktab_id),
                    )
                    for class_row in planned_classes:
                        cur.execute(
                            """INSERT INTO maktab_sinflari(
                                   maktab_id,sinf,harf,smena,talim_tili)
                               VALUES(%s,%s,%s,%s,%s) RETURNING id""",
                            (
                                maktab_id,
                                str(class_row["sinf"]),
                                class_row["harf"],
                                int(class_row["smena"]),
                                class_row["talim_tili"],
                            ),
                        )
                        class_id = int(cur.fetchone()["id"])
                        class_payload.append({
                            "id": class_id,
                            "name": class_row["name"],
                            "sinf": str(class_row["sinf"]),
                            "harf": class_row["harf"],
                            "smena": int(class_row["smena"]),
                            "talim_tili": class_row["talim_tili"],
                            "alifbo_tartibi": int(class_row["alifbo_tartibi"]),
                        })
                _v1852_default_shifts(
                    cur, maktab_id, int(sorov.smena_soni or 1)
                )
                cur.execute(
                    """INSERT INTO foydalanuvchi_muassasalari(
                           user_id,muassasa_turi,muassasa_id,lavozim
                       ) VALUES(%s,'maktab',%s,'direktor')
                       ON CONFLICT(user_id,muassasa_turi,muassasa_id)
                       DO UPDATE SET lavozim='direktor'""",
                    (user_id, maktab_id),
                )
                # Yangi yaratilgan/tanlangan maktab joriy faol maktab bo‘ladi.
                # COALESCE eski maktab ID sini saqlab qolib, yangi workspace'ni
                # yana eski maktabga qaytarayotgan edi.
                cur.execute(
                    """UPDATE users
                          SET maktab_id=%s,
                              lavozim=COALESCE(NULLIF(lavozim,''),'direktor')
                        WHERE user_id=%s""",
                    (maktab_id, user_id),
                )
                conn.commit()
                return {
                    "holat": "yaratildi" if created else "mavjud_retry",
                    "maktab_id": maktab_id,
                    "maktab_nomi": school_name,
                    "legacy_yaratildi": created,
                    "smena_soni": (
                        int(linked_school.get("smena_soni") or 1)
                        if linked_school else int(sorov.smena_soni or 1)
                    ),
                    "alifbo_turi": (
                        str(linked_school.get("alifbo_turi") or "latin_xalqaro")
                        if linked_school else alphabet_type
                    ),
                    "viloyat": (
                        linked_school.get("viloyat") if linked_school else
                        (str(sorov.viloyat or "").strip() or None)
                    ),
                    "tuman": (
                        linked_school.get("tuman") if linked_school else
                        (str(sorov.tuman or "").strip() or None)
                    ),
                    "sinflar": class_payload,
                }
            existing = _v198_existing_school_for_user(cur, user_id)
            if not existing:
                raise HTTPException(
                    status_code=404,
                    detail="Sizga tegishli maktab topilmadi. Maktabni qayta tanlang yoki yangisini yarating.",
                )
            conn.commit()
            return {
                "holat": "mavjud",
                "maktab_id": int(existing["maktab_id"]),
                "maktab_nomi": existing["maktab_nomi"],
                "legacy_yaratildi": False,
            }

        maktab_id = None
        if organization.get("external_id") is not None:
            try:
                candidate_id = int(organization["external_id"])
            except (TypeError, ValueError):
                candidate_id = 0
            if candidate_id > 0:
                cur.execute(
                    "SELECT id,nomi FROM maktablar WHERE id=%s",
                    (candidate_id,),
                )
                linked_school = cur.fetchone()
                if linked_school:
                    maktab_id = int(linked_school["id"])

        created = False
        school_name = str(organization.get("display_name") or "Yangi maktab").strip()
        if maktab_id is None:
            # V17 external_id yo'q bo'lsa ham avval ayni nomdagi mavjud
            # maktabni qidiramiz. Faqat rostdan ham topilmasa yangi yozuv
            # yaratiladi; eski maktabga kirishda dublikat paydo bo'lmaydi.
            existing = _v198_existing_school_for_user(
                cur,
                user_id,
                preferred_id=sorov.maktab_id,
                preferred_name=school_name,
            )
            if existing:
                maktab_id = int(existing["maktab_id"])
                school_name = existing["maktab_nomi"]
            else:
                cur.execute(
                    """INSERT INTO maktablar(
                           nomi,smena_soni,direktor_user_id,pulli,oylik_tolov
                       ) VALUES(%s,1,%s,FALSE,NULL) RETURNING id""",
                    (school_name, user_id),
                )
                maktab_id = int(cur.fetchone()["id"])
                created = True
            cur.execute(
                "UPDATE learning_contexts SET external_id=%s WHERE id=%s",
                (maktab_id, int(organization["context_id"])),
            )

        cur.execute(
            """INSERT INTO foydalanuvchi_muassasalari(
                   user_id,muassasa_turi,muassasa_id,lavozim
               ) VALUES(%s,'maktab',%s,'direktor')
               ON CONFLICT(user_id,muassasa_turi,muassasa_id)
               DO UPDATE SET lavozim='direktor'""",
            (user_id, maktab_id),
        )
        cur.execute(
            """UPDATE users
                  SET maktab_id=%s,
                      lavozim=COALESCE(NULLIF(lavozim,''),'direktor')
                WHERE user_id=%s""",
            (maktab_id, user_id),
        )
        cur.execute(
            """UPDATE maktablar
                  SET direktor_user_id=COALESCE(direktor_user_id,%s)
                WHERE id=%s""",
            (user_id, maktab_id),
        )
        # external_id avval mavjud bo'lsa ham normal integer qiymatga keladi.
        cur.execute(
            "UPDATE learning_contexts SET external_id=%s WHERE id=%s",
            (maktab_id, int(organization["context_id"])),
        )
        conn.commit()
        return {
            "holat": "boglandi",
            "maktab_id": maktab_id,
            "maktab_nomi": school_name,
            "organization_v17_id": int(organization["organization_v17_id"]),
            "context_id": int(organization["context_id"]),
            "legacy_yaratildi": created,
            "lifecycle_status": organization.get("lifecycle_status"),
        }
    except Exception:
        conn.rollback()
        raise
    finally:
        cur.close()
        conn.close()


class V237ClassCreateRequest(BaseModel):
    maktab_id: int
    sinf: int
    harf: str
    smena: int = 1
    xona_id: Optional[int] = None
    talim_tili: str = "uz"


class V237SchoolAlphabetRequest(BaseModel):
    maktab_id: int
    alifbo_turi: str


class V237ClassEditRequest(BaseModel):
    maktab_id: int
    sinf_id: int
    harf: str
    smena: int
    # Eski frontend bu maydonlarni yubormaydi. ``xona_yangilansin`` aniq
    # TRUE bo'lmaguncha mavjud xona tasodifan NULL bo'lib qolmaydi.
    xona_id: Optional[int] = None
    xona_yangilansin: bool = False
    # Eski frontend yubormasa mavjud sinf tili o‘zgarmaydi.
    talim_tili: Optional[str] = None


def _v237_resolve_class_room(cur, maktab_id: int, xona_id, smena: int,
                             exclude_sinf_id=None):
    """Uy xonasini tekshiradi va maktab_sinflari uchun kanonik qiymat beradi."""
    if xona_id is None:
        return {
            "xona_id": None, "bino_id": None,
            "xona": None, "bino": None,
        }
    try:
        room_id = int(xona_id)
    except (TypeError, ValueError) as exc:
        raise HTTPException(status_code=400, detail="Sinf xonasi noto'g'ri") from exc
    cur.execute(
        """SELECT r.id,r.nomi,r.turi,r.faol,r.darsga_yaroqli,
                  r.bino_id,r.xona_raqami,b.nomi AS bino_nomi
             FROM aqlli_xonalar_v2 r
             LEFT JOIN aqlli_binolar_v2 b ON b.id=r.bino_id
            WHERE r.id=%s AND r.maktab_id=%s FOR UPDATE OF r""",
        (room_id, int(maktab_id)),
    )
    room = cur.fetchone()
    if not room:
        raise HTTPException(status_code=404, detail="Tanlangan xona bu maktabda topilmadi")
    if not bool(room.get("faol", True)) or not bool(room.get("darsga_yaroqli", True)) \
            or str(room.get("turi") or "").strip().lower() == "non_teaching":
        raise HTTPException(status_code=409, detail="Tanlangan xona dars uchun faol emas")
    args = [int(maktab_id), int(smena), room_id]
    exclude_sql = ""
    if exclude_sinf_id is not None:
        exclude_sql = " AND id<>%s"
        args.append(int(exclude_sinf_id))
    cur.execute(
        """SELECT id,sinf,harf FROM maktab_sinflari
            WHERE maktab_id=%s AND COALESCE(smena,1)=%s AND xona_id=%s"""
        + exclude_sql + " LIMIT 1 FOR UPDATE",
        tuple(args),
    )
    occupied = cur.fetchone()
    if occupied:
        raise HTTPException(
            status_code=409,
            detail=(
                f"Bu xona {occupied['sinf']}-{occupied['harf']} sinfiga "
                f"{int(smena)}-smenada biriktirilgan"
            ),
        )
    room_text = str(room.get("xona_raqami") or room.get("nomi") or "").strip() or None
    return {
        "xona_id": room_id,
        "bino_id": int(room["bino_id"]) if room.get("bino_id") is not None else None,
        "xona": room_text,
        "bino": str(room.get("bino_nomi") or "").strip() or None,
    }


@app.post("/api/maktab/aqlli_jadval/v3/sinf_yaratish")
def v237_class_create(sorov: V237ClassCreateRequest, token: str):
    """Mavjud maktabga ID-bog'lanishlarni buzmasdan bitta yangi sinf qo'shadi."""
    actor_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor()
    try:
        # Xona/bino va o'quv reja jadvallari biznes qulfidan oldin tayyorlanadi.
        _v192_tables(cur)
        _v209_school_creation_tables(cur)
        conn.commit()
        if not _v1852_manager(cur, actor_id, sorov.maktab_id):
            raise HTTPException(status_code=403, detail="Sinfni faqat maktab rahbariyati qo'shadi")
        try:
            grade = int(sorov.sinf)
            label = _v237_clean_parallel_label(sorov.harf)
            label_key = _v237_parallel_label_key(label)
            instruction_language = _v238_normalize_instruction_language(
                sorov.talim_tili
            )
        except (TypeError, ValueError) as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        if grade not in range(1, 12):
            raise HTTPException(status_code=400, detail="Sinf darajasi 1–11 oralig'ida bo'lishi kerak")
        shift = int(sorov.smena)
        cur.execute("SELECT pg_advisory_xact_lock(%s)", (2373000000 + int(sorov.maktab_id),))
        cur.execute(
            "SELECT COALESCE(smena_soni,1) AS smena_soni FROM maktablar WHERE id=%s FOR UPDATE",
            (sorov.maktab_id,),
        )
        school = cur.fetchone()
        if not school:
            raise HTTPException(status_code=404, detail="Maktab topilmadi")
        shift_count = int(school.get("smena_soni") or 1)
        if shift not in (1, 2) or shift > shift_count:
            raise HTTPException(status_code=400, detail=f"Bu maktab {shift_count} smenali")
        cur.execute(
            "SELECT id,harf FROM maktab_sinflari WHERE maktab_id=%s AND sinf=%s FOR UPDATE",
            (sorov.maktab_id, str(grade)),
        )
        for row in cur.fetchall():
            try:
                duplicate = _v237_parallel_label_key(row.get("harf")) == label_key
            except ValueError:
                duplicate = False
            if duplicate:
                raise HTTPException(status_code=409, detail=f"{grade}-{label} sinfi allaqachon mavjud")
        room = _v237_resolve_class_room(
            cur, sorov.maktab_id, sorov.xona_id, shift
        )
        cur.execute(
            """INSERT INTO maktab_sinflari(
                   maktab_id,sinf,harf,smena,bino,xona,bino_id,xona_id,
                   talim_tili
               ) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
            (
                sorov.maktab_id, str(grade), label, shift,
                room["bino"], room["xona"], room["bino_id"], room["xona_id"],
                instruction_language,
            ),
        )
        class_id = int(cur.fetchone()["id"])
        _v199_ensure_class_hour_rules(cur, sorov.maktab_id, [class_id], actor_id)
        cur.execute(
            "SELECT id,sinf,harf,COALESCE(smena,1) AS smena,"
            "COALESCE(talim_tili,'uz') AS talim_tili FROM maktab_sinflari "
            "WHERE maktab_id=%s ORDER BY id",
            (sorov.maktab_id,),
        )
        _v193_ensure_plan(cur, sorov.maktab_id, [dict(row) for row in cur.fetchall()])
        cur.execute(
            "UPDATE aqlli_jadval_urinishlari_v2 SET holat='bekor' "
            "WHERE maktab_id=%s AND holat='draft'",
            (sorov.maktab_id,),
        )
        result = {
            "id": class_id, "sinf": str(grade), "harf": label,
            "name": f"{grade}-{label}", "smena": shift,
            "talim_tili": instruction_language, **room,
        }
        conn.commit()
        return {"holat": "yaratildi", "sinf": result, "oquv_reja_holati": "draft"}
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


@app.put("/api/maktab/aqlli_jadval/v3/sinf_alifbosi")
def v237_school_alphabet_save(sorov: V237SchoolAlphabetRequest, token: str):
    """Keyingi sinf takliflari/saralashi uchun alifboni saqlaydi; nomlarni o'zgartirmaydi."""
    actor_id = _jwt_tekshir(token)
    alphabet_type = str(sorov.alifbo_turi or "").strip().lower()
    if alphabet_type not in _V237_CLASS_ALPHABETS:
        raise HTTPException(status_code=400, detail="Sinf alifbosi noto'g'ri")
    conn = _db(); cur = conn.cursor()
    try:
        _v1852_tables(cur)
        if not _v1852_manager(cur, actor_id, sorov.maktab_id):
            raise HTTPException(status_code=403, detail="Sinf alifbosini faqat rahbariyat o'zgartiradi")
        cur.execute("SELECT pg_advisory_xact_lock(%s)", (2373000000 + int(sorov.maktab_id),))
        cur.execute(
            "UPDATE maktablar SET alifbo_turi=%s WHERE id=%s RETURNING id",
            (alphabet_type, sorov.maktab_id),
        )
        if not cur.fetchone():
            raise HTTPException(status_code=404, detail="Maktab topilmadi")
        conn.commit()
        return {"holat": "saqlandi", "alifbo_turi": alphabet_type}
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


@app.patch("/api/maktab/aqlli_jadval/v3/sinf_tahrirlash")
@app.put("/api/maktab/aqlli_jadval/v3/sinf_tahrirlash")
def v237_class_edit(sorov: V237ClassEditRequest, token: str):
    """Sinf ID sini saqlagan holda parallel yorlig'i va smenasini tahrirlaydi."""
    actor_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor()
    try:
        _v192_tables(cur)
        _v209_school_creation_tables(cur)
        conn.commit()
        if not _v1852_manager(cur, actor_id, sorov.maktab_id):
            raise HTTPException(
                status_code=403,
                detail="Sinf nomi va smenasini faqat maktab rahbariyati tahrirlaydi",
            )
        try:
            new_label = _v237_clean_parallel_label(sorov.harf)
            new_key = _v237_parallel_label_key(new_label)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        new_shift = int(sorov.smena)
        cur.execute(
            "SELECT pg_advisory_xact_lock(%s)",
            (2373000000 + int(sorov.maktab_id),),
        )
        cur.execute(
            "SELECT COALESCE(smena_soni,1) AS smena_soni FROM maktablar WHERE id=%s FOR UPDATE",
            (sorov.maktab_id,),
        )
        school = cur.fetchone()
        if not school:
            raise HTTPException(status_code=404, detail="Maktab topilmadi")
        shift_count = int(school.get("smena_soni") or 1)
        if new_shift not in (1, 2) or new_shift > shift_count:
            raise HTTPException(
                status_code=400,
                detail=f"Bu maktab {shift_count} smenali; {new_shift}-smena tanlab bo‘lmaydi",
            )
        cur.execute(
            """SELECT id,sinf,harf,COALESCE(smena,1) AS smena,
                      bino,xona,bino_id,xona_id,
                      COALESCE(talim_tili,'uz') AS talim_tili
                 FROM maktab_sinflari
                WHERE id=%s AND maktab_id=%s FOR UPDATE""",
            (sorov.sinf_id, sorov.maktab_id),
        )
        current = cur.fetchone()
        if not current:
            raise HTTPException(status_code=404, detail="Sinf topilmadi")
        # ``sinf`` raqam bo'lib qoladi. Faqat ayni darajadagi parallel yorlig'i
        # case/apostrof/bo'shliq normallashgan holda takror bo'lmasligi tekshiriladi.
        cur.execute(
            """SELECT id,harf FROM maktab_sinflari
                WHERE maktab_id=%s AND sinf=%s AND id<>%s FOR UPDATE""",
            (sorov.maktab_id, current["sinf"], sorov.sinf_id),
        )
        duplicate = None
        for row in cur.fetchall():
            try:
                existing_key = _v237_parallel_label_key(row.get("harf"))
            except ValueError:
                # Tarixiy buzilgan/bo'sh yorliq boshqa sog'lom sinfni
                # tahrirlashga to'sqinlik qilmasin; o'zi tahrirlanganda yangi
                # validatsiyadan baribir o'tadi.
                continue
            if existing_key == new_key:
                duplicate = row
                break
        if duplicate:
            raise HTTPException(
                status_code=409,
                detail=f"{current['sinf']}-{new_label} sinfi allaqachon mavjud",
            )
        try:
            new_language = (
                _v238_normalize_instruction_language(sorov.talim_tili)
                if sorov.talim_tili is not None
                else _v238_normalize_instruction_language(current.get("talim_tili"))
            )
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        changed_shift = int(current.get("smena") or 1) != new_shift
        changed_label = str(current.get("harf") or "") != new_label
        current_language = _v238_normalize_instruction_language(
            current.get("talim_tili")
        )
        changed_language = current_language != new_language
        if changed_language:
            # Tilni jim almashtirib eski fan/yuklama/guruh/jadvalni o‘chirish
            # mumkin emas. Bo‘sh sinfda ruxsat; ishlatilgan sinf uchun alohida
            # ko‘chirish oqimi kerakligi aniq 409 bilan bildiriladi.
            cur.execute(
                """SELECT
                     (SELECT COUNT(*) FROM aqlli_oquv_reja_qatorlari_v19_3
                       WHERE maktab_id=%s AND sinf_id=%s) AS reja,
                     (SELECT COUNT(*) FROM maktab_dars_birikmalari
                       WHERE maktab_id=%s AND sinf_id=%s) AS birikma,
                     (SELECT COUNT(*) FROM maktab_sinf_guruh_tizimlari
                       WHERE sinf_id=%s AND faol=TRUE) AS guruh,
                     (SELECT COUNT(*) FROM aqlli_jadval_slotlari_v2 e
                       JOIN aqlli_jadval_urinishlari_v2 u ON u.id=e.urinish_id
                      WHERE u.maktab_id=%s AND e.sinf_id=%s) AS jadval""",
                (
                    sorov.maktab_id, sorov.sinf_id,
                    sorov.maktab_id, sorov.sinf_id,
                    sorov.sinf_id,
                    sorov.maktab_id, sorov.sinf_id,
                ),
            )
            used = dict(cur.fetchone() or {})
            if any(int(used.get(key) or 0) > 0 for key in (
                "reja", "birikma", "guruh", "jadval"
            )):
                raise HTTPException(
                    status_code=409,
                    detail=(
                        "Bu sinfda o‘quv reja/yuklama/guruh yoki jadval ma’lumoti bor. "
                        "Ta’lim tilini almashtirish ma’lumotni o‘chirishi mumkin; yangi "
                        "sinfni to‘g‘ri tilda yarating yoki maxsus ko‘chirishni bajaring."
                    ),
                )
        room = {
            "xona_id": current.get("xona_id"),
            "bino_id": current.get("bino_id"),
            "xona": current.get("xona"),
            "bino": current.get("bino"),
        }
        if sorov.xona_yangilansin or (changed_shift and current.get("xona_id") is not None):
            room = _v237_resolve_class_room(
                cur,
                sorov.maktab_id,
                sorov.xona_id if sorov.xona_yangilansin else current.get("xona_id"),
                new_shift,
                exclude_sinf_id=sorov.sinf_id,
            )
        changed_room = any(
            (current.get(key) or None) != (room.get(key) or None)
            for key in ("xona_id", "bino_id", "xona", "bino")
        )
        if changed_shift or changed_label or changed_room or changed_language:
            cur.execute(
                """UPDATE maktab_sinflari
                      SET harf=%s,smena=%s,bino=%s,xona=%s,bino_id=%s,xona_id=%s,
                          talim_tili=%s
                    WHERE id=%s AND maktab_id=%s""",
                (
                    new_label, new_shift, room["bino"], room["xona"],
                    room["bino_id"], room["xona_id"],
                    new_language,
                    sorov.sinf_id, sorov.maktab_id,
                ),
            )
            # Draft eski smenaga qarab yaratilgan bo'lishi mumkin. Tasdiqlangan
            # jadval arxiv sifatida qoladi; yangi draft esa qayta yaratiladi.
            if changed_shift or changed_room or changed_language:
                cur.execute(
                    """UPDATE aqlli_jadval_urinishlari_v2 SET holat='bekor'
                        WHERE maktab_id=%s AND holat='draft'""",
                    (sorov.maktab_id,),
                )
            if changed_language:
                _v238_mark_plan_languages_draft(
                    cur, sorov.maktab_id, [current_language, new_language]
                )
                cur.execute("""UPDATE aqlli_oquv_reja_holati_v19_3
                                  SET holat='draft',tasdiqlagan_user_id=NULL,
                                      tasdiqlangan_at=NULL,yangilangan_at=NOW()
                                WHERE maktab_id=%s""", (sorov.maktab_id,))
        cur.execute(
            "SELECT COALESCE(alifbo_turi,'latin_xalqaro') AS alifbo_turi FROM maktablar WHERE id=%s",
            (sorov.maktab_id,),
        )
        alphabet_type = (cur.fetchone() or {}).get("alifbo_turi") or "latin_xalqaro"
        result = {
            "id": int(sorov.sinf_id),
            "sinf": str(current["sinf"]),
            "harf": new_label,
            "name": f"{current['sinf']}-{new_label}",
            "smena": new_shift,
            "talim_tili": new_language,
            **room,
            "alifbo_turi": alphabet_type,
            "tartib": list(_v237_class_sort_key({
                "id": sorov.sinf_id,
                "sinf": current["sinf"],
                "harf": new_label,
            }, alphabet_type)[:2]),
            "jadval_qayta_yaratish_kerak": (
                changed_shift or changed_room or changed_language
            ),
        }
        conn.commit()
        return {
            "holat": "saqlandi", "sinf": result,
            "ozgardi": changed_shift or changed_label or changed_room or changed_language,
        }
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


def _v192_clean_subject(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


# O'zbekiston Respublikasi Maktabgacha va maktab ta'limi vazirining
# 2026-yil 10-apreldagi 133-son buyrug'i, 1-ilova (o'zbek tilidagi maktablar).
# Bu ikki konstanta avvalgi monolitdan modulga ajratishda tushib qolgan edi.
# Ular modul importi paytida kerak: bo'lmasa Railway yangi deployni
# ``NameError: SAMTM_2026_2027_UZBEK_CURRICULUM`` bilan to'xtatadi.
SAMTM_2026_2027_UZBEK_CURRICULUM = (
    ("Ona tili", {1: 4, 2: 4, 3: 4, 4: 4, 5: 4, 6: 4, 7: 3, 8: 3, 9: 3, 10: 2, 11: 2}),
    ("O'qish savodxonligi", {1: 4, 2: 3, 3: 3, 4: 3}),
    ("Adabiyot", {5: 2, 6: 2, 7: 2, 8: 2, 9: 2, 10: 2, 11: 2}),
    ("Rus tili", {2: 2, 3: 2, 4: 2, 5: 2, 6: 2, 7: 2, 8: 2, 9: 2, 10: 2, 11: 2}),
    ("Chet tili", {1: 1, 2: 2, 3: 2, 4: 2, 5: 4, 6: 4, 7: 4, 8: 3, 9: 3, 10: 2, 11: 2}),
    ("Tarixdan hikoyalar", {5: 2}),
    ("Qadimgi dunyo tarixi", {6: 2}),
    ("O'zbekiston tarixi", {7: 2, 8: 2, 9: 2, 10: 1, 11: 1}),
    ("Jahon tarixi", {7: 1, 8: 1, 9: 1, 10: 1, 11: 1}),
    ("Davlat va huquq asoslari", {8: 1, 9: 1, 10: 1, 11: 1}),
    ("Tarbiya", {1: 1, 2: 1, 3: 1, 4: 1, 5: 1, 6: 1, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1}),
    ("Matematika", {1: 5, 2: 5, 3: 5, 4: 5, 5: 5, 6: 5, 7: 5}),
    ("Algebra", {8: 3, 9: 3, 10: 3, 11: 3}),
    ("Geometriya", {8: 2, 9: 2, 10: 2, 11: 2}),
    ("Informatika va axborot texnologiyalari", {1: 1, 2: 1, 3: 1, 4: 1, 5: 1, 6: 1, 7: 1, 8: 1, 9: 2, 10: 2, 11: 2}),
    ("Fizika", {7: 2, 8: 2, 9: 2, 10: 2, 11: 2}),
    ("Astronomiya", {11: 1}),
    ("Kimyo", {7: 2, 8: 2, 9: 2, 10: 2, 11: 2}),
    ("Biologiya", {7: 2, 8: 2, 9: 2, 10: 2, 11: 2}),
    ("Geografiya", {7: 2, 8: 1.5, 9: 1.5, 10: 2}),
    ("Iqtisodiy bilim asoslari", {8: 0.5, 9: 0.5}),
    ("Tadbirkorlik asoslari", {11: 1}),
    ("Tabiiy fan (Science)", {1: 1, 2: 1, 3: 1, 4: 1, 5: 2, 6: 3}),
    ("Musiqa madaniyati", {1: 1, 2: 1, 3: 1, 4: 1, 5: 1, 6: 1, 7: 1}),
    ("Tasviriy san'at", {1: 1, 2: 1, 3: 1, 4: 1, 5: 1, 6: 1, 7: 1}),
    ("Chizmachilik", {8: 1, 9: 1}),
    ("Texnologiya", {1: 1, 2: 1, 3: 1, 4: 1, 5: 2, 6: 2, 7: 2, 8: 1, 9: 1}),
    ("Jismoniy tarbiya", {1: 1, 2: 2, 3: 2, 4: 2, 5: 2, 6: 2, 7: 2, 8: 2, 9: 2, 10: 2, 11: 2}),
    ("Chaqiruvga qadar boshlang'ich tayyorgarlik", {10: 2, 11: 2}),
)

SAMTM_2026_2027_CURRICULUM_SOURCE = {
    "nomi": "2026-2027 tayanch o'quv reja",
    "buyruq": "MMTV 133-son",
    "sana": "2026-04-10",
    "ilova": "1-ilova",
    "talim_tili": "o'zbek",
    "sinf_jami": {
        1: 21, 2: 24, 3: 24, 4: 24, 5: 29, 6: 30,
        7: 35, 8: 33, 9: 34, 10: 31, 11: 31,
    },
}


_V238_RU_SUBJECT_NAMES = {
    "Ona tili": "Русский язык",
    "O'qish savodxonligi": "Литературное чтение",
    "Adabiyot": "Литература", "Rus tili": "Узбекский язык",
    "Chet tili": "Иностранный язык",
    "Tarixdan hikoyalar": "Рассказы по истории",
    "Qadimgi dunyo tarixi": "История древнего мира",
    "O'zbekiston tarixi": "История Узбекистана", "Jahon tarixi": "Всемирная история",
    "Davlat va huquq asoslari": "Основы государства и права",
    "Tarbiya": "Воспитание", "Matematika": "Математика",
    "Algebra": "Алгебра", "Geometriya": "Геометрия",
    "Informatika va axborot texnologiyalari": "Информатика и информационные технологии",
    "Fizika": "Физика", "Astronomiya": "Астрономия", "Kimyo": "Химия",
    "Biologiya": "Биология", "Geografiya": "География",
    "Iqtisodiy bilim asoslari": "Основы экономических знаний",
    "Tadbirkorlik asoslari": "Основы предпринимательства",
    "Tabiiy fan (Science)": "Естествознание (Science)",
    "Musiqa madaniyati": "Музыкальная культура", "Tasviriy san'at": "Изобразительное искусство",
    "Chizmachilik": "Черчение", "Texnologiya": "Технология",
    "Jismoniy tarbiya": "Физическая культура",
    "Chaqiruvga qadar boshlang'ich tayyorgarlik": "Начальная допризывная подготовка",
}
_V238_EN_SUBJECT_NAMES = {
    "Ona tili": "English Language", "O'qish savodxonligi": "Reading Literacy",
    "Adabiyot": "Literature", "Rus tili": "Russian Language",
    "Chet tili": "Foreign Language", "Tarixdan hikoyalar": "Stories from History",
    "Qadimgi dunyo tarixi": "Ancient World History",
    "O'zbekiston tarixi": "History of Uzbekistan", "Jahon tarixi": "World History",
    "Davlat va huquq asoslari": "Foundations of State and Law", "Tarbiya": "Character Education",
    "Matematika": "Mathematics", "Algebra": "Algebra", "Geometriya": "Geometry",
    "Informatika va axborot texnologiyalari": "Computer Science and Information Technology",
    "Fizika": "Physics", "Astronomiya": "Astronomy", "Kimyo": "Chemistry",
    "Biologiya": "Biology", "Geografiya": "Geography",
    "Iqtisodiy bilim asoslari": "Foundations of Economics",
    "Tadbirkorlik asoslari": "Foundations of Entrepreneurship",
    "Tabiiy fan (Science)": "Natural Science", "Musiqa madaniyati": "Music",
    "Tasviriy san'at": "Visual Arts", "Chizmachilik": "Technical Drawing",
    "Texnologiya": "Technology", "Jismoniy tarbiya": "Physical Education",
    "Chaqiruvga qadar boshlang'ich tayyorgarlik": "Pre-conscription Training",
}
SAMTM_2026_2027_RUSSIAN_CURRICULUM = tuple(
    (_V238_RU_SUBJECT_NAMES.get(subject, subject), hours)
    for subject, hours in SAMTM_2026_2027_UZBEK_CURRICULUM
)
SAMTM_2026_2027_ENGLISH_CURRICULUM = tuple(
    (_V238_EN_SUBJECT_NAMES.get(subject, subject), hours)
    for subject, hours in SAMTM_2026_2027_UZBEK_CURRICULUM
)
SAMTM_2026_2027_CURRICULA = {
    "uz": SAMTM_2026_2027_UZBEK_CURRICULUM,
    "ru": SAMTM_2026_2027_RUSSIAN_CURRICULUM,
    "en": SAMTM_2026_2027_ENGLISH_CURRICULUM,
}
SAMTM_2026_2027_CURRICULUM_SOURCES = {
    language: {**SAMTM_2026_2027_CURRICULUM_SOURCE, "talim_tili": language}
    for language in _V238_INSTRUCTION_LANGUAGES
}


SAMTM_V19_3_DEFAULT_CURRICULUM = SAMTM_2026_2027_UZBEK_CURRICULUM


def _v201_central_school_settings_tables(cur):
    """Admin boshqaradigan, yil nomiga qotirilmagan maktab andozasi."""
    cur.execute("""CREATE TABLE IF NOT EXISTS admin_maktab_andoza_versiyalari_v20_1(
        id BIGSERIAL PRIMARY KEY,
        nomi TEXT NOT NULL DEFAULT 'Amaldagi maktab andozasi',
        faol BOOLEAN NOT NULL DEFAULT TRUE,
        yangilagan_user_id BIGINT,
        yangilangan_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
    )""")
    cur.execute("""CREATE UNIQUE INDEX IF NOT EXISTS uq_admin_maktab_andoza_faol_v20_1
                   ON admin_maktab_andoza_versiyalari_v20_1(faol) WHERE faol=TRUE""")
    cur.execute("""CREATE TABLE IF NOT EXISTS admin_maktab_andoza_fanlari_v20_1(
        id BIGSERIAL PRIMARY KEY,
        versiya_id BIGINT NOT NULL REFERENCES admin_maktab_andoza_versiyalari_v20_1(id) ON DELETE CASCADE,
        talim_tili TEXT NOT NULL DEFAULT 'uz',
        sinf_darajasi INTEGER NOT NULL CHECK(sinf_darajasi BETWEEN 1 AND 11),
        fan_nomi TEXT NOT NULL,
        fan_kaliti TEXT NOT NULL,
        haftalik_soat NUMERIC(4,1) NOT NULL DEFAULT 0 CHECK(haftalik_soat BETWEEN 0 AND 20),
        metod_kuni INTEGER CHECK(metod_kuni BETWEEN 1 AND 7),
        kunlik_max INTEGER NOT NULL DEFAULT 1 CHECK(kunlik_max BETWEEN 1 AND 4),
        faol BOOLEAN NOT NULL DEFAULT TRUE,
        tartib INTEGER NOT NULL DEFAULT 0
    )""")
    cur.execute(
        "ALTER TABLE admin_maktab_andoza_fanlari_v20_1 "
        "ADD COLUMN IF NOT EXISTS talim_tili TEXT NOT NULL DEFAULT 'uz'"
    )
    # Eski unique(versiya,sinf,fan) RU/EN qatorlarini bloklaydi. Uni faqat
    # aniq shu jadvaldagi legacy unique bo‘lsa olib, til qatnashgan indexga
    # almashtiramiz; PK yoki boshqa constraintlarga tegmaymiz.
    cur.execute("""DO $$
        DECLARE constraint_name TEXT;
        BEGIN
          FOR constraint_name IN
            SELECT c.conname
              FROM pg_constraint c
              JOIN pg_class t ON t.oid=c.conrelid
             WHERE t.relname='admin_maktab_andoza_fanlari_v20_1'
               AND c.contype='u'
               AND pg_get_constraintdef(c.oid) ILIKE '%versiya_id%'
               AND pg_get_constraintdef(c.oid) ILIKE '%sinf_darajasi%'
               AND pg_get_constraintdef(c.oid) ILIKE '%fan_kaliti%'
               AND pg_get_constraintdef(c.oid) NOT ILIKE '%talim_tili%'
          LOOP
            EXECUTE format(
              'ALTER TABLE admin_maktab_andoza_fanlari_v20_1 DROP CONSTRAINT %I',
              constraint_name
            );
          END LOOP;
        END $$""")
    cur.execute("""CREATE UNIQUE INDEX IF NOT EXISTS uq_admin_andoza_fan_til_v238
                   ON admin_maktab_andoza_fanlari_v20_1(
                     versiya_id,talim_tili,sinf_darajasi,fan_kaliti)""")
    cur.execute("""CREATE TABLE IF NOT EXISTS maktab_andoza_override_v20_1(
        maktab_id INTEGER NOT NULL REFERENCES maktablar(id) ON DELETE CASCADE,
        talim_tili TEXT NOT NULL DEFAULT 'uz',
        bolim TEXT NOT NULL CHECK(bolim IN ('fanlar','oquv_reja','metod_kunlari')),
        alohida BOOLEAN NOT NULL DEFAULT FALSE,
        yangilagan_user_id BIGINT,
        yangilangan_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
        PRIMARY KEY(maktab_id,talim_tili,bolim)
    )""")
    cur.execute(
        "ALTER TABLE maktab_andoza_override_v20_1 "
        "ADD COLUMN IF NOT EXISTS talim_tili TEXT NOT NULL DEFAULT 'uz'"
    )
    cur.execute("""DO $$
        DECLARE constraint_name TEXT;
        BEGIN
          FOR constraint_name IN
            SELECT c.conname FROM pg_constraint c
            JOIN pg_class t ON t.oid=c.conrelid
            WHERE t.relname='maktab_andoza_override_v20_1'
              AND c.contype='p'
              AND pg_get_constraintdef(c.oid) NOT ILIKE '%talim_tili%'
          LOOP
            EXECUTE format(
              'ALTER TABLE maktab_andoza_override_v20_1 DROP CONSTRAINT %I',
              constraint_name
            );
          END LOOP;
        END $$""")
    cur.execute("""CREATE UNIQUE INDEX IF NOT EXISTS uq_maktab_andoza_override_til_v238
                   ON maktab_andoza_override_v20_1(
                     maktab_id,talim_tili,bolim)""")
    cur.execute("""CREATE TABLE IF NOT EXISTS admin_maktab_andoza_tasdiqlari_v20_2(
        versiya_id BIGINT NOT NULL REFERENCES admin_maktab_andoza_versiyalari_v20_1(id) ON DELETE CASCADE,
        talim_tili TEXT NOT NULL DEFAULT 'uz',
        bolim TEXT NOT NULL CHECK(bolim IN ('fanlar','yuklama','jadval','metod')),
        tasdiqlangan BOOLEAN NOT NULL DEFAULT FALSE,
        tasdiqlagan_user_id BIGINT,
        tasdiqlangan_at TIMESTAMPTZ,
        PRIMARY KEY(versiya_id,talim_tili,bolim)
    )""")
    cur.execute(
        "ALTER TABLE admin_maktab_andoza_tasdiqlari_v20_2 "
        "ADD COLUMN IF NOT EXISTS talim_tili TEXT NOT NULL DEFAULT 'uz'"
    )
    cur.execute("""DO $$
        DECLARE constraint_name TEXT;
        BEGIN
          FOR constraint_name IN
            SELECT c.conname FROM pg_constraint c
            JOIN pg_class t ON t.oid=c.conrelid
            WHERE t.relname='admin_maktab_andoza_tasdiqlari_v20_2'
              AND c.contype='p'
              AND pg_get_constraintdef(c.oid) NOT ILIKE '%talim_tili%'
          LOOP
            EXECUTE format(
              'ALTER TABLE admin_maktab_andoza_tasdiqlari_v20_2 DROP CONSTRAINT %I',
              constraint_name
            );
          END LOOP;
        END $$""")
    cur.execute("""CREATE UNIQUE INDEX IF NOT EXISTS uq_admin_andoza_tasdiq_til_v238
                   ON admin_maktab_andoza_tasdiqlari_v20_2(
                     versiya_id,talim_tili,bolim)""")
    cur.execute("SELECT id FROM admin_maktab_andoza_versiyalari_v20_1 WHERE faol=TRUE LIMIT 1")
    version = cur.fetchone()
    if version:
        version_id = int(version["id"])
    else:
        cur.execute("""INSERT INTO admin_maktab_andoza_versiyalari_v20_1(nomi,faol)
                       VALUES('Amaldagi maktab andozasi',TRUE)
                       ON CONFLICT DO NOTHING RETURNING id""")
        created_version = cur.fetchone()
        if created_version:
            version_id = int(created_version["id"])
        else:
            # Gunicorn workerlari startupni parallel bajarsa, boshqa worker
            # yaratgan faol versiyani kutib, aynan o'shani olamiz.
            cur.execute("""SELECT id FROM admin_maktab_andoza_versiyalari_v20_1
                            WHERE faol=TRUE LIMIT 1""")
            version_id = int(cur.fetchone()["id"])
    # Mavjud o‘zbek qatorlari o‘zgarmaydi. Yetishmagan RU/EN til tablari bir
    # marta ayni haftalik soatlar + aniq tarjima nomlari bilan seed qilinadi;
    # admin har tilni keyin mustaqil tahrirlaydi.
    for language, curriculum in SAMTM_2026_2027_CURRICULA.items():
        cur.execute(
            "SELECT 1 FROM admin_maktab_andoza_fanlari_v20_1 "
            "WHERE versiya_id=%s AND talim_tili=%s LIMIT 1",
            (version_id, language),
        )
        if cur.fetchone():
            continue
        order = 0
        for subject, grades in curriculum:
            for grade, hours in grades.items():
                cur.execute("""INSERT INTO admin_maktab_andoza_fanlari_v20_1(
                                versiya_id,talim_tili,sinf_darajasi,fan_nomi,
                                fan_kaliti,haftalik_soat,kunlik_max,tartib)
                               VALUES(%s,%s,%s,%s,%s,%s,1,%s)
                               ON CONFLICT DO NOTHING""",
                            (version_id, language, int(grade), subject,
                             _v1875_subject_key(subject), float(hours), order))
            order += 1
    for language in _V238_INSTRUCTION_LANGUAGES:
        for section in ("fanlar", "yuklama", "jadval", "metod"):
            cur.execute("""INSERT INTO admin_maktab_andoza_tasdiqlari_v20_2(
                            versiya_id,talim_tili,bolim,tasdiqlangan)
                           VALUES(%s,%s,%s,%s) ON CONFLICT DO NOTHING""",
                        (version_id, language, section, language == "uz"))
    # Oldingi build mexanik seed qilgan RU/EN qatorlar TRUE, lekin kim va
    # qachon tasdiqlagani NULL bo'lib qolgan. Faqat shunday texnik seedlar
    # draftga qaytariladi; haqiqiy admin tasdig'iga tegilmaydi.
    cur.execute("""UPDATE admin_maktab_andoza_tasdiqlari_v20_2
                      SET tasdiqlangan=FALSE
                    WHERE versiya_id=%s AND talim_tili IN ('ru','en')
                      AND tasdiqlangan=TRUE
                      AND tasdiqlagan_user_id IS NULL
                      AND tasdiqlangan_at IS NULL""", (version_id,))
    return version_id


def _v201_central_rows(
    cur,
    talim_tili="uz",
    include_unapproved=False,
    required_sections=None,
):
    version_id = _v201_central_school_settings_tables(cur)
    language = _v238_normalize_instruction_language(talim_tili)
    if not include_unapproved:
        sections = tuple(required_sections or ("fanlar", "yuklama", "jadval"))
        cur.execute("""SELECT bolim,tasdiqlangan
                         FROM admin_maktab_andoza_tasdiqlari_v20_2
                        WHERE versiya_id=%s AND talim_tili=%s
                          AND bolim=ANY(%s)""",
                    (version_id, language, list(sections)))
        approved = {
            row["bolim"] for row in cur.fetchall()
            if bool(row.get("tasdiqlangan"))
        }
        if not set(sections) <= approved:
            return []
    cur.execute("""SELECT talim_tili,sinf_darajasi,fan_nomi,haftalik_soat,metod_kuni,
                          kunlik_max,tartib
                     FROM admin_maktab_andoza_fanlari_v20_1
                    WHERE versiya_id=%s AND talim_tili=%s AND faol=TRUE
                    ORDER BY sinf_darajasi,tartib,fan_nomi""", (version_id, language))
    return [dict(row) for row in cur.fetchall()]


def _v201_central_curriculum(cur, talim_tili="uz"):
    grouped = {}
    daily_limits = {}
    names = {}
    for row in _v201_central_rows(cur, talim_tili):
        key = _v1875_subject_key(row["fan_nomi"])
        names[key] = row["fan_nomi"]
        grouped.setdefault(key, {})[int(row["sinf_darajasi"])] = float(row["haftalik_soat"])
        daily_limits.setdefault(key, {})[int(row["sinf_darajasi"])] = int(
            row.get("kunlik_max") or 1
        )
    return tuple(
        (names[key], grades, daily_limits.get(key, {}))
        for key, grades in grouped.items()
    )


def _v201_mark_school_override(
    cur, maktab_id: int, section: str, actor_id: int, talim_tili="uz"
):
    _v201_central_school_settings_tables(cur)
    language = _v238_normalize_instruction_language(talim_tili)
    cur.execute("""INSERT INTO maktab_andoza_override_v20_1(
                    maktab_id,talim_tili,bolim,alohida,
                    yangilagan_user_id,yangilangan_at)
                   VALUES(%s,%s,%s,TRUE,%s,NOW())
                   ON CONFLICT(maktab_id,talim_tili,bolim) DO UPDATE SET
                     alohida=TRUE,yangilagan_user_id=EXCLUDED.yangilagan_user_id,
                     yangilangan_at=NOW()""",
                (maktab_id, language, section, actor_id))


class V201CentralSubjectRow(BaseModel):
    sinf_darajasi: int
    fan_nomi: str
    haftalik_soat: float = 0
    metod_kuni: Optional[int] = None
    kunlik_max: int = 1


class V201CentralSchoolSettings(BaseModel):
    qatorlar: list[V201CentralSubjectRow]
    bolim: str = "fanlar"
    tasdiqlash: bool = True
    talim_tili: str = "uz"


@app.on_event("startup")
def _v201_central_school_settings_startup():
    try:
        conn = _db(); cur = conn.cursor()
        _v201_central_school_settings_tables(cur)
        conn.commit(); cur.close(); conn.close()
    except Exception as exc:
        print(f"[V20.1 markaziy maktab sozlamasi] {exc}", flush=True)


@app.get("/api/admin/maktab_markaziy_sozlamalari")
def v201_central_school_settings_get(token: str, talim_tili: str = "uz"):
    _admin_tekshir(token)
    try:
        language = _v238_normalize_instruction_language(talim_tili)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    conn = _db(); cur = conn.cursor()
    try:
        version_id = _v201_central_school_settings_tables(cur)
        rows = _v201_central_rows(cur, language, include_unapproved=True)
        by_language = {
            key: _v201_central_rows(cur, key, include_unapproved=True)
            for key in _V238_INSTRUCTION_LANGUAGES
        }
        cur.execute("""SELECT talim_tili,bolim,tasdiqlangan,tasdiqlangan_at
                         FROM admin_maktab_andoza_tasdiqlari_v20_2
                        WHERE versiya_id=%s AND talim_tili=%s""",
                    (version_id, language))
        selected_approval_rows = [dict(row) for row in cur.fetchall()]
        approvals = {row["bolim"]: {
            "tasdiqlangan": bool(row["tasdiqlangan"]),
            "tasdiqlangan_at": row["tasdiqlangan_at"],
        } for row in selected_approval_rows}
        cur.execute("""SELECT talim_tili,bolim,tasdiqlangan,tasdiqlangan_at
                         FROM admin_maktab_andoza_tasdiqlari_v20_2
                        WHERE versiya_id=%s""", (version_id,))
        approval_rows = [dict(row) for row in cur.fetchall()]
        return {
            "nomi": "Amaldagi maktab andozasi",
            "talim_tili": language,
            "tillar": [
                {"kalit": key, "nomi": label, "qator_soni": len(by_language[key])}
                for key, label in _V238_INSTRUCTION_LANGUAGES.items()
            ],
            "til_boyicha": by_language,
            "qatorlar": rows,
            "sinf_fanlari": {
                str(grade): [row["fan_nomi"] for row in rows if int(row["sinf_darajasi"]) == grade]
                for grade in range(1, 12)
            },
            "tasdiqlar": approvals,
            "tasdiqlar_til_boyicha": {
                key: {
                    row["bolim"]: {
                        "tasdiqlangan": bool(row["tasdiqlangan"]),
                        "tasdiqlangan_at": row["tasdiqlangan_at"],
                    }
                    for row in approval_rows if row["talim_tili"] == key
                }
                for key in _V238_INSTRUCTION_LANGUAGES
            },
        }
    finally:
        cur.close(); conn.close()


@app.put("/api/admin/maktab_markaziy_sozlamalari")
def v201_central_school_settings_save(sorov: V201CentralSchoolSettings, token: str):
    actor_id = _admin_tekshir(token)
    try:
        language = _v238_normalize_instruction_language(sorov.talim_tili)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    section = str(sorov.bolim or "").strip().lower()
    if section not in {"fanlar", "yuklama", "jadval", "metod"}:
        raise HTTPException(status_code=400, detail="Sozlama bo‘limi noto‘g‘ri")
    if section in {"yuklama", "jadval"}:
        # Amaldagi admin oynasi fan nomi, haftalik soat va kunlik maksimumni
        # bitta to'liq snapshotda yuboradi. Ularni alohida bo'lim sifatida
        # saqlash bitta live jadvalda qisman tasdiqlangan holat yaratadi.
        raise HTTPException(
            status_code=409,
            detail=(
                "Fanlar, haftalik yuklama va kunlik maksimum bitta andoza "
                "sifatida ‘fanlar’ bo‘limidan tasdiqlanadi"
            ),
        )
    if not sorov.tasdiqlash:
        # Bu jadval amaldagi markaziy andoza. Alohida draft jadvali yo'q ekan,
        # tasdiqlanmagan yozuvni shu yerga saqlash eski live qiymatni bosib
        # ketardi. Shuning uchun draftni jim jonlantirish o'rniga aniq bloklaymiz.
        raise HTTPException(
            status_code=409,
            detail="Markaziy andoza faqat ‘Saqlash va tasdiqlash’ orqali yangilanadi",
        )
    normalized = []
    seen = set()
    for index, item in enumerate(sorov.qatorlar):
        grade = int(item.sinf_darajasi)
        subject = _v192_clean_subject(item.fan_nomi)
        key = _v1875_subject_key(subject)
        hours = float(item.haftalik_soat)
        method_day = int(item.metod_kuni) if item.metod_kuni is not None else None
        daily_max = int(item.kunlik_max)
        if grade not in range(1, 12) or not key:
            raise HTTPException(status_code=400, detail="Sinf yoki fan nomi noto‘g‘ri")
        if (grade, key) in seen:
            raise HTTPException(status_code=400, detail=f"{grade}-sinf / {subject} takrorlangan")
        if not 0 <= hours <= 20 or daily_max not in range(1, 5):
            raise HTTPException(status_code=400, detail=f"{grade}-sinf / {subject} soati noto‘g‘ri")
        if method_day is not None and method_day not in range(1, 8):
            raise HTTPException(status_code=400, detail=f"{subject} metod kuni noto‘g‘ri")
        seen.add((grade, key))
        normalized.append((grade, subject, key, hours, method_day, daily_max, index))
    if not normalized:
        raise HTTPException(status_code=400, detail="Kamida bitta sinf–fan qatori kerak")
    conn = _db(); cur = conn.cursor()
    try:
        version_id = _v201_central_school_settings_tables(cur)
        if section == "metod":
            # Metod kuni alohida tasdiqlanadi, ammo fan/soat/maksimum bilan
            # bir jadvalda saqlanadi. Shu endpoint o'sha reja ustunlarini
            # qayta yozmasligi shart: yuborilgan snapshot aynan mavjud fanlar
            # to'plamiga mos bo'lsa, faqat ``metod_kuni`` yangilanadi.
            cur.execute("""SELECT sinf_darajasi,fan_kaliti
                             FROM admin_maktab_andoza_fanlari_v20_1
                            WHERE versiya_id=%s AND talim_tili=%s AND faol=TRUE""",
                        (version_id, language))
            existing_keys = {
                (int(row["sinf_darajasi"]), str(row["fan_kaliti"]))
                for row in cur.fetchall()
            }
            received_keys = {(row[0], row[2]) for row in normalized}
            if received_keys != existing_keys:
                raise HTTPException(
                    status_code=409,
                    detail=(
                        "Metod kunlari mavjud til andozasidagi barcha "
                        "sinf–fan qatorlari uchun to‘liq yuborilishi kerak"
                    ),
                )
            for grade, _subject, key, _hours, method_day, _daily, _order in normalized:
                cur.execute("""UPDATE admin_maktab_andoza_fanlari_v20_1
                                  SET metod_kuni=%s
                                WHERE versiya_id=%s AND talim_tili=%s
                                  AND sinf_darajasi=%s AND fan_kaliti=%s""",
                            (method_day, version_id, language, grade, key))
        else:
            cur.execute(
                "DELETE FROM admin_maktab_andoza_fanlari_v20_1 "
                "WHERE versiya_id=%s AND talim_tili=%s",
                (version_id, language),
            )
            for row in normalized:
                cur.execute("""INSERT INTO admin_maktab_andoza_fanlari_v20_1(
                                versiya_id,talim_tili,sinf_darajasi,fan_nomi,fan_kaliti,
                                haftalik_soat,metod_kuni,kunlik_max,tartib)
                               VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                            (version_id, language, *row))
        cur.execute("""UPDATE admin_maktab_andoza_versiyalari_v20_1
                          SET yangilagan_user_id=%s,yangilangan_at=NOW() WHERE id=%s""",
                    (actor_id, version_id))
        approval_sections = (
            ["fanlar", "yuklama", "jadval"]
            if section == "fanlar" else ["metod"]
        )
        cur.execute("""UPDATE admin_maktab_andoza_tasdiqlari_v20_2
                          SET tasdiqlangan=FALSE,tasdiqlagan_user_id=NULL,tasdiqlangan_at=NULL
                        WHERE versiya_id=%s AND talim_tili=%s
                          AND bolim=ANY(%s)""",
                    (version_id, language, approval_sections))
        if section == "fanlar":
            cur.execute("""UPDATE admin_maktab_andoza_tasdiqlari_v20_2
                              SET tasdiqlangan=FALSE,tasdiqlagan_user_id=NULL,tasdiqlangan_at=NULL
                            WHERE versiya_id=%s AND talim_tili=%s
                              AND bolim='metod'""",
                        (version_id, language))
        if sorov.tasdiqlash:
            cur.execute("""UPDATE admin_maktab_andoza_tasdiqlari_v20_2
                              SET tasdiqlangan=TRUE,tasdiqlagan_user_id=%s,tasdiqlangan_at=NOW()
                            WHERE versiya_id=%s AND talim_tili=%s
                              AND bolim=ANY(%s)""",
                        (actor_id, version_id, language, approval_sections))
        cur.execute("""SELECT bolim,tasdiqlangan
                         FROM admin_maktab_andoza_tasdiqlari_v20_2
                        WHERE versiya_id=%s AND talim_tili=%s
                          AND bolim IN ('fanlar','yuklama','jadval')""",
                    (version_id, language))
        plan_approvals = {
            row["bolim"] for row in cur.fetchall()
            if bool(row.get("tasdiqlangan"))
        }
        plan_ready = {"fanlar", "yuklama", "jadval"} <= plan_approvals
        # Markaziy o'zgarish faqat alohida sozlama qilmagan maktablarga
        # tarqaladi. Ularning keyingi ochilishida reja yangi andozadan qayta
        # quriladi; maxsus maktab sozlamasi hech qachon bosib ketilmaydi.
        cur.execute("SELECT id FROM maktablar ORDER BY id")
        schools = [int(row["id"]) for row in cur.fetchall()]
        if (
            not sorov.tasdiqlash
            or not plan_ready
            or section not in {"fanlar", "yuklama", "jadval"}
        ):
            # Markaziy draft yoki faqat metod-kun tahriri maktablarning
            # amaldagi rejalariga tarqalmaydi.
            schools = []
        fan_updated = plan_updated = method_updated = 0
        for school_id in schools:
            cur.execute("""SELECT 1 FROM maktab_sinflari
                            WHERE maktab_id=%s
                              AND COALESCE(talim_tili,'uz')=%s LIMIT 1""",
                        (school_id, language))
            if not cur.fetchone():
                continue
            cur.execute("""SELECT bolim FROM maktab_andoza_override_v20_1
                            WHERE maktab_id=%s AND talim_tili=%s
                              AND alohida=TRUE""", (school_id, language))
            overrides = {row["bolim"] for row in cur.fetchall()}
            if "fanlar" not in overrides:
                fan_updated += 1
            if "oquv_reja" not in overrides:
                cur.execute(
                    """DELETE FROM aqlli_oquv_reja_qatorlari_v19_3 q
                         USING maktab_sinflari s
                         WHERE q.maktab_id=%s AND q.sinf_id=s.id
                           AND s.maktab_id=%s
                           AND COALESCE(s.talim_tili,'uz')=%s""",
                    (school_id, school_id, language),
                )
                cur.execute("""UPDATE aqlli_oquv_reja_holati_v19_3
                                  SET holat='draft',versiya=versiya+1,
                                      tasdiqlagan_user_id=NULL,tasdiqlangan_at=NULL,
                                      yangilangan_at=NOW() WHERE maktab_id=%s""", (school_id,))
                cur.execute("""INSERT INTO aqlli_oquv_reja_til_holati_v238(
                                maktab_id,talim_tili,holat,versiya,yangilangan_at)
                               VALUES(%s,%s,'draft',1,NOW())
                               ON CONFLICT(maktab_id,talim_tili) DO UPDATE SET
                                 holat='draft',
                                 versiya=aqlli_oquv_reja_til_holati_v238.versiya+1,
                                 tasdiqlagan_user_id=NULL,tasdiqlangan_at=NULL,
                                 yangilangan_at=NOW()""", (school_id, language))
                cur.execute("""UPDATE aqlli_jadval_urinishlari_v2 SET holat='bekor'
                                WHERE maktab_id=%s AND holat='draft'""", (school_id,))
                plan_updated += 1
            if "metod_kunlari" not in overrides:
                # Metod qoidasi hozir fan nomiga bog‘langan, sinf tiliga emas.
                # Bitta til tabini saqlashda boshqa tillarning metod kunlarini
                # ommaviy DELETE qilish xavfli; admin uni alohida tasdiqlaydi.
                pass
        conn.commit()
        return {"holat": "tasdiqlandi" if sorov.tasdiqlash else "saqlandi",
                "bolim": section, "talim_tili": language,
                "til_rejasi_faol": plan_ready,
                "qator_soni": len(normalized),
                "maktablar": {"fanlar": fan_updated, "oquv_reja": plan_updated,
                              "metod_kunlari": method_updated},
                "izoh": "Alohida override qilmagan maktablar bu andozani avtomatik oladi."}
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


def _v193_grade_number(value):
    match = re.search(r"\d+", str(value or ""))
    return int(match.group(0)) if match else 0


def _v201_school_has_override(
    cur, maktab_id: int, section: str, talim_tili="uz"
) -> bool:
    """Maktab markaziy andozadan aynan shu bo'limda chetlaganmi."""
    _v201_central_school_settings_tables(cur)
    language = _v238_normalize_instruction_language(talim_tili)
    cur.execute("""SELECT 1 FROM maktab_andoza_override_v20_1
                   WHERE maktab_id=%s AND talim_tili=%s
                     AND bolim=%s AND alohida=TRUE""",
                (maktab_id, language, section))
    return bool(cur.fetchone())


def _v194_school_subject_grades(cur, maktab_id: int, talim_tili="uz"):
    """None — eski maktab; dict — fanlar sinflar bo'yicha aniq sozlangan."""
    language = _v238_normalize_instruction_language(talim_tili)
    if (
        language != "uz"
        or not _v201_school_has_override(cur, maktab_id, "fanlar", language)
    ):
        central_result = {}
        for row in _v201_central_rows(cur, language):
            grade = int(row["sinf_darajasi"])
            central_result.setdefault(grade, {})[
                _v1875_subject_key(row["fan_nomi"])
            ] = row["fan_nomi"]
        return central_result or None
    cur.execute("SELECT to_regclass('public.maktab_fan_sinflari_v19_4') AS jadval")
    if not (cur.fetchone() or {}).get("jadval"):
        return None
    cur.execute("""SELECT fan_nomi,sinf_darajasi
                   FROM maktab_fan_sinflari_v19_4
                   WHERE maktab_id=%s ORDER BY sinf_darajasi,fan_nomi""", (maktab_id,))
    rows = cur.fetchall()
    if not rows:
        # Yangi maktab DTSdagi tasodifiy nomlardan emas, admin tasdiqlagan
        # markaziy 1–11-sinf andozasidan boshlaydi.
        central_result = {}
        for row in _v201_central_rows(cur, language):
            grade = int(row["sinf_darajasi"])
            central_result.setdefault(grade, {})[
                _v1875_subject_key(row["fan_nomi"])
            ] = row["fan_nomi"]
        return central_result or None
    result = {}
    for row in rows:
        grade = int(row["sinf_darajasi"])
        result.setdefault(grade, {})[_v1875_subject_key(row["fan_nomi"])] = row["fan_nomi"]
    return result


def _v193_template_rows_for_class(class_row, selected_by_grade=None, curriculum=None):
    grade = _v193_grade_number(class_row.get("sinf"))
    allowed = (selected_by_grade or {}).get(grade) if selected_by_grade is not None else None
    # ``None`` faqat tarixiy UZ fallbackni anglatadi. Bo'sh tuple esa RU/EN
    # andozasi hali tasdiqlanmaganini bildiradi va UZ fanlariga tushmasligi shart.
    curriculum_items = list(
        SAMTM_V19_3_DEFAULT_CURRICULUM
        if curriculum is None else curriculum
    )
    default_hours = {
        _v1875_subject_key(subject): float(hours.get(grade, 0) or 0)
        for subject, hours, *_rest in curriculum_items
    }
    daily_limits = {
        _v1875_subject_key(item[0]): int(
            (item[2] if len(item) > 2 else {}).get(grade, 1) or 1
        )
        for item in curriculum_items
    }
    if allowed is not None:
        # O'quv reja ustunlari faqat maktab 1–11-sinflar uchun saqlagan
        # fanlardan tuziladi. Tayanchda topilmagan yangi fan ham yo'qolmaydi:
        # u 0 soat bilan ko'rinadi va rahbariyat soatini qo'lda kiritadi.
        return [
            {
                "sinf_id": int(class_row["id"]),
                "fan_nomi": subject,
                "haftalik_soat": default_hours.get(key, 0.0),
                "kunlik_max": daily_limits.get(key, 1),
                "manba": "amaldagi_tayanch_reja" if default_hours.get(key, 0) > 0 else "maktab_fan_tanlovi",
            }
            for key, subject in allowed.items()
        ]
    return [
        {
            "sinf_id": int(class_row["id"]),
            "fan_nomi": subject,
            "haftalik_soat": float(hours[grade]),
            "kunlik_max": int((daily or {}).get(grade, 1) or 1),
            "manba": "amaldagi_tayanch_reja",
        }
        for subject, hours, *daily_parts in curriculum_items
        for daily in [daily_parts[0] if daily_parts else {}]
        if float(hours.get(grade, 0)) > 0
    ]


def _v238_class_curriculum_context(cur, maktab_id: int, class_row, cache=None):
    cache = cache if cache is not None else {}
    language = _v238_normalize_instruction_language(
        (class_row or {}).get("talim_tili")
    )
    if language not in cache:
        cache[language] = (
            _v194_school_subject_grades(cur, maktab_id, language),
            _v201_central_curriculum(cur, language),
        )
    selected_by_grade, curriculum = cache[language]
    return language, selected_by_grade, curriculum


def _v238_mark_plan_languages_draft(cur, maktab_id: int, languages):
    """Rejasi o'zgargan tillarning eski tasdiq belgisini xavfsiz bekor qiladi."""
    normalized = sorted({
        _v238_normalize_instruction_language(language)
        for language in languages
    })
    for language in normalized:
        cur.execute("""INSERT INTO aqlli_oquv_reja_til_holati_v238(
                        maktab_id,talim_tili,holat,versiya,yangilangan_at)
                       VALUES(%s,%s,'draft',1,NOW())
                       ON CONFLICT(maktab_id,talim_tili) DO UPDATE SET
                         holat='draft',
                         versiya=aqlli_oquv_reja_til_holati_v238.versiya+1,
                         tasdiqlagan_user_id=NULL,tasdiqlangan_at=NULL,
                         yangilangan_at=NOW()""", (maktab_id, language))


def _v193_ensure_plan(cur, maktab_id: int, classes):
    context_cache = {}
    cur.execute("""INSERT INTO aqlli_oquv_reja_holati_v19_3(maktab_id)
                   VALUES(%s) ON CONFLICT(maktab_id) DO NOTHING""", (maktab_id,))
    cur.execute("""SELECT sinf_id,fan_nomi,haftalik_soat,kunlik_max
                   FROM aqlli_oquv_reja_qatorlari_v19_3 WHERE maktab_id=%s""",
                (maktab_id,))
    saved_rows = cur.fetchall()
    existing = {int(row["sinf_id"]) for row in saved_rows}
    expected_rows = []
    class_by_id = {int(row["id"]): row for row in classes}
    centrally_managed_ids = set()
    for class_row in classes:
        talim_tili, selected_by_grade, curriculum = _v238_class_curriculum_context(
            cur, maktab_id, class_row, context_cache
        )
        if _v201_school_has_override(
            cur, maktab_id, "oquv_reja", talim_tili
        ):
            continue
        # Unapproved RU/EN seed operational reja emas. U mavjud rejani
        # o'chirmaydi va UZ fallbackka ham tushmaydi.
        if not curriculum and not selected_by_grade:
            continue
        centrally_managed_ids.add(int(class_row["id"]))
        expected_rows.extend(
            row for row in _v193_template_rows_for_class(
                class_row, selected_by_grade, curriculum
            ) if float(row["haftalik_soat"] or 0) > 0
        )
    if centrally_managed_ids:
        saved_by_class = {}
        for row in saved_rows:
            saved_by_class.setdefault(int(row["sinf_id"]), set()).add(
                (_v1875_subject_key(row["fan_nomi"]),
                 round(float(row["haftalik_soat"] or 0), 3),
                 int(row.get("kunlik_max") or 1))
            )
        expected_by_class = {}
        for row in expected_rows:
            expected_by_class.setdefault(int(row["sinf_id"]), set()).add(
                (_v1875_subject_key(row["fan_nomi"]),
                 round(float(row["haftalik_soat"] or 0), 3),
                 int(row.get("kunlik_max") or 1))
            )
        changed_class_ids = sorted(
            class_id for class_id in centrally_managed_ids
            if saved_by_class.get(class_id, set())
            != expected_by_class.get(class_id, set())
        )
        if changed_class_ids:
            cur.execute(
                "DELETE FROM aqlli_oquv_reja_qatorlari_v19_3 "
                "WHERE maktab_id=%s AND sinf_id=ANY(%s)",
                (maktab_id, changed_class_ids),
            )
            for item in expected_rows:
                if int(item["sinf_id"]) not in changed_class_ids:
                    continue
                cur.execute("""INSERT INTO aqlli_oquv_reja_qatorlari_v19_3(
                                maktab_id,sinf_id,fan_nomi,haftalik_soat,
                                kunlik_max,manba,yangilangan_at)
                               VALUES(%s,%s,%s,%s,%s,%s,NOW())""",
                            (maktab_id, item["sinf_id"], item["fan_nomi"],
                             item["haftalik_soat"], item["kunlik_max"], item["manba"]))
                cur.execute("""INSERT INTO maktab_fanlari(maktab_id,fan_nomi)
                               VALUES(%s,%s) ON CONFLICT DO NOTHING""",
                            (maktab_id, item["fan_nomi"]))
            cur.execute("""UPDATE aqlli_oquv_reja_holati_v19_3
                           SET holat='draft',versiya=versiya+1,
                               tasdiqlagan_user_id=NULL,tasdiqlangan_at=NULL,
                               yangilangan_at=NOW() WHERE maktab_id=%s""", (maktab_id,))
            _v238_mark_plan_languages_draft(
                cur, maktab_id,
                [class_by_id[class_id].get("talim_tili")
                 for class_id in changed_class_ids],
            )
            return
    missing = [row for row in classes if int(row["id"]) not in existing]
    missing = [
        row for row in missing
        if any(_v238_class_curriculum_context(
            cur, maktab_id, row, context_cache
        )[1:])
    ]
    if not missing:
        return
    for class_row in missing:
        talim_tili, selected_by_grade, curriculum = _v238_class_curriculum_context(
            cur, maktab_id, class_row, context_cache
        )
        template_rows = _v193_template_rows_for_class(class_row, selected_by_grade, curriculum)
        for item in template_rows:
            if float(item["haftalik_soat"] or 0) <= 0:
                continue
            cur.execute("""INSERT INTO aqlli_oquv_reja_qatorlari_v19_3(
                            maktab_id,sinf_id,fan_nomi,haftalik_soat,
                            kunlik_max,manba,yangilangan_at)
                           VALUES(%s,%s,%s,%s,%s,%s,NOW())
                           ON CONFLICT(maktab_id,sinf_id,fan_nomi) DO NOTHING""",
                        (
                            maktab_id, item["sinf_id"], item["fan_nomi"],
                            item["haftalik_soat"], item["kunlik_max"], item["manba"],
                        ))
        for item in template_rows:
            cur.execute("""INSERT INTO maktab_fanlari(maktab_id,fan_nomi)
                           VALUES(%s,%s) ON CONFLICT DO NOTHING""",
                        (maktab_id, item["fan_nomi"]))
    cur.execute("""UPDATE aqlli_oquv_reja_holati_v19_3
                   SET holat='draft',tasdiqlagan_user_id=NULL,
                       tasdiqlangan_at=NULL,yangilangan_at=NOW()
                   WHERE maktab_id=%s""", (maktab_id,))
    _v238_mark_plan_languages_draft(
        cur, maktab_id,
        [row.get("talim_tili") for row in missing],
    )


def _v193_plan_payload(cur, maktab_id: int, classes):
    _v193_ensure_plan(cur, maktab_id, classes)
    context_cache = {}
    class_languages = {
        _v238_normalize_instruction_language(row.get("talim_tili"))
        for row in classes
    }
    custom_plan_by_language = {
        language: _v201_school_has_override(
            cur, maktab_id, "oquv_reja", language
        )
        for language in class_languages
    }
    cur.execute("""SELECT holat,versiya,tasdiqlagan_user_id,
                          tasdiqlangan_at,yangilangan_at
                   FROM aqlli_oquv_reja_holati_v19_3 WHERE maktab_id=%s""",
                (maktab_id,))
    status = dict(cur.fetchone() or {})
    cur.execute("""SELECT talim_tili,holat,versiya,tasdiqlangan_at,yangilangan_at
                     FROM aqlli_oquv_reja_til_holati_v238
                    WHERE maktab_id=%s ORDER BY talim_tili""", (maktab_id,))
    language_statuses = {
        row["talim_tili"]: dict(row) for row in cur.fetchall()
    }
    cur.execute("""SELECT id,sinf_id,fan_nomi,haftalik_soat,kunlik_max,
                          manba,yangilangan_at
                   FROM aqlli_oquv_reja_qatorlari_v19_3
                   WHERE maktab_id=%s
                   ORDER BY sinf_id,fan_nomi""", (maktab_id,))
    rows = [dict(row) for row in cur.fetchall()]
    class_by_id = {int(row["id"]): row for row in classes}
    filtered_rows = []
    for row in rows:
        class_row = class_by_id.get(int(row["sinf_id"]))
        if class_row is None:
            continue
        row["talim_tili"] = _v238_normalize_instruction_language(
            class_row.get("talim_tili")
        )
        _, selected_by_grade, _ = _v238_class_curriculum_context(
            cur, maktab_id, class_row, context_cache
        )
        grade = _v193_grade_number(class_row.get("sinf"))
        allowed = (selected_by_grade or {}).get(grade) if selected_by_grade is not None else None
        if (
            custom_plan_by_language.get(row["talim_tili"], False)
            or allowed is None
            or _v1875_subject_key(row["fan_nomi"]) in allowed
        ):
            filtered_rows.append(row)
    rows = filtered_rows
    class_totals = {}
    for row in rows:
        class_id = int(row["sinf_id"])
        class_totals[class_id] = class_totals.get(class_id, 0.0) + float(row["haftalik_soat"])
    cur.execute("""SELECT sinf_id,fan_nomi,haftalik_soat FROM aqlli_sinf_soati_qoidalari_v2
                   WHERE maktab_id=%s AND faol=TRUE""", (maktab_id,))
    class_hour_rows = {int(row["sinf_id"]): dict(row) for row in cur.fetchall()}
    template_rows = []
    for class_row in classes:
        language, selected_by_grade, curriculum = _v238_class_curriculum_context(
            cur, maktab_id, class_row, context_cache
        )
        class_templates = _v193_template_rows_for_class(
            class_row, selected_by_grade, curriculum
        )
        for item in class_templates:
            item["talim_tili"] = language
        template_rows.extend(class_templates)
    active_year = _v1852_active_year(cur, maktab_id)
    active_year_name = str((active_year or {}).get("nomi") or "").strip()
    return {
        "holat": status.get("holat") or "draft",
        "versiya": int(status.get("versiya") or 1),
        "tasdiqlangan_at": status.get("tasdiqlangan_at"),
        "yangilangan_at": status.get("yangilangan_at"),
        "til_holatlari": language_statuses,
        # UI yilni koddan emas, maktabning faol o'quv yilidan oladi.
        # Shu sabab keyingi o'quv yilida frontend kodi almashtirilmaydi.
        "oquv_yili_nomi": active_year_name,
        "andoza_nomi": "Amaldagi tayanch o‘quv reja",
        "andoza_manbasi": SAMTM_2026_2027_CURRICULUM_SOURCE,
        "andoza_manbalari": SAMTM_2026_2027_CURRICULUM_SOURCES,
        "markaziy_andoza": not any(custom_plan_by_language.values()),
        "markaziy_andoza_til_boyicha": {
            language: not custom
            for language, custom in custom_plan_by_language.items()
        },
        "sinf_soati_avtomatik": True,
        "sinf_soati_haftalik": 1,
        "sinf_soati_nomi": next((row.get("fan_nomi") for row in class_hour_rows.values()), "KELAJAK SOATI"),
        "qatorlar": rows,
        "andoza_qatorlar": template_rows,
        "sinf_jami": [
            {
                "sinf_id": int(class_row["id"]),
                "sinf": f"{class_row['sinf']}-{class_row['harf']}",
                "talim_tili": _v238_normalize_instruction_language(
                    class_row.get("talim_tili")
                ),
                "fan_soati": float(class_totals.get(int(class_row["id"]), 0)),
                "sinf_soati": int(class_hour_rows.get(int(class_row["id"]), {}).get("haftalik_soat") or 0),
                "sinf_soati_nomi": class_hour_rows.get(int(class_row["id"]), {}).get("fan_nomi") or "KELAJAK SOATI",
                "haftalik_soat": float(class_totals.get(int(class_row["id"]), 0))
                + int(class_hour_rows.get(int(class_row["id"]), {}).get("haftalik_soat") or 0),
            }
            for class_row in classes
        ],
        "tanlangan_fan_sinflari": [
            {
                "sinf_id": int(class_row["id"]),
                "sinf_darajasi": _v193_grade_number(class_row.get("sinf")),
                "talim_tili": _v238_normalize_instruction_language(
                    class_row.get("talim_tili")
                ),
                "fanlar": list((
                    _v238_class_curriculum_context(
                        cur, maktab_id, class_row, context_cache
                    )[1] or {}
                ).get(_v193_grade_number(class_row.get("sinf")), {}).values()),
            }
            for class_row in classes
        ],
    }


def _v193_approved_plan_map(cur, maktab_id: int):
    cur.execute("""SELECT holat FROM aqlli_oquv_reja_holati_v19_3
                   WHERE maktab_id=%s""", (maktab_id,))
    status = cur.fetchone()
    if not status or status.get("holat") != "tasdiqlangan":
        return None
    cur.execute("""SELECT sinf_id,fan_nomi,haftalik_soat,kunlik_max
                   FROM aqlli_oquv_reja_qatorlari_v19_3
                   WHERE maktab_id=%s""", (maktab_id,))
    return {
        (int(row["sinf_id"]), _v1875_subject_key(row["fan_nomi"])): dict(row)
        for row in cur.fetchall()
    }


def _v192_group_variants(cur, maktab_id: int):
    systems = _v1876_group_system_catalog(cur, maktab_id)
    cur.execute("""SELECT s.id,s.sinf,s.harf,COALESCE(s.smena,1) AS smena,
                          COALESCE(s.talim_tili,'uz') AS talim_tili,
                          s.bino,s.xona,s.bino_id,s.xona_id,
                          s.rahbar_user_id,COALESCE(u.full_name,'') AS rahbar_ismi
                   FROM maktab_sinflari s
                   LEFT JOIN users u ON u.user_id=s.rahbar_user_id
                   WHERE s.maktab_id=%s
                   ORDER BY CASE WHEN s.sinf::text ~ '^\\d+$' THEN s.sinf::int ELSE 999 END,s.harf""",
                (maktab_id,))
    classes = [dict(row) for row in cur.fetchall()]
    systems_by_class = {}
    for system in systems:
        systems_by_class.setdefault(int(system["sinf_id"]), []).append(system)

    result = []
    for cls in classes:
        class_id = int(cls["id"])
        class_name = f"{cls['sinf']}-{cls['harf']}"
        result.append({
            "sinf_id": class_id,
            "sinf": class_name,
            "guruh_kaliti": "whole",
            "guruh_nomi": "Butun sinf",
            "qisqa": class_name,
            "turi": "whole",
            "tizim_id": None,
            "fanlar": [],
        })
        seen = set()
        for system in systems_by_class.get(class_id, []):
            for group in system.get("guruhlar") or []:
                key = str(group.get("guruh_kaliti") or "").strip()
                if not key or key in seen:
                    continue
                seen.add(key)
                short = {
                    "group_1": "1G",
                    "group_2": "2G",
                    "boys": "O‘",
                    "girls": "Q",
                }.get(key, str(group.get("guruh_nomi") or key)[:8])
                result.append({
                    "sinf_id": class_id,
                    "sinf": class_name,
                    "guruh_kaliti": key,
                    "guruh_nomi": str(group.get("guruh_nomi") or key),
                    "qisqa": short,
                    "turi": str(system.get("turi") or "manual"),
                    "tizim_id": int(system["id"]),
                    "fanlar": list(system.get("fanlar") or []),
                })
    return classes, systems, result


def _v192_assignment_rows(cur, maktab_id: int):
    cur.execute("""SELECT b.id,b.maktab_id,b.user_id,b.sinf_id,b.fan_nomi,
                          COALESCE(NULLIF(b.guruh_kaliti,''),'whole') AS guruh_kaliti,
                          COALESCE(b.haftalik_soat,0) AS haftalik_soat,
                          COALESCE(b.kunlik_max,1) AS kunlik_max,b.xona_id,
                          COALESCE(b.manba,'noma_lum') AS manba,
                          u.full_name,s.sinf,s.harf,COALESCE(s.smena,1) AS smena
                   FROM maktab_dars_birikmalari b
                   JOIN users u ON u.user_id=b.user_id
                   JOIN maktab_sinflari s ON s.id=b.sinf_id
                   WHERE b.maktab_id=%s
                   ORDER BY u.full_name,s.sinf::int,s.harf,b.fan_nomi,b.guruh_kaliti""",
                (maktab_id,))
    return [dict(row) for row in cur.fetchall()]


def _v192_totals(cur, maktab_id: int, rows, classes):
    teacher_base = {}
    class_subject = {}
    for row in rows:
        hours = max(0.0, float(row.get("haftalik_soat") or 0))
        teacher_id = int(row["user_id"])
        teacher_base[teacher_id] = teacher_base.get(teacher_id, 0) + hours
        pair = (int(row["sinf_id"]), _v1875_subject_key(row["fan_nomi"]))
        data = class_subject.setdefault(pair, {"whole": [], "groups": [], "fan_nomi": row["fan_nomi"]})
        if _v1875_group_key(row.get("guruh_kaliti")) == "whole":
            data["whole"].append(hours)
        else:
            data["groups"].append(hours)

    cur.execute("""SELECT q.sinf_id,q.haftalik_soat,s.rahbar_user_id
                   FROM aqlli_sinf_soati_qoidalari_v2 q
                   JOIN maktab_sinflari s ON s.id=q.sinf_id
                   WHERE q.maktab_id=%s AND q.faol=TRUE""", (maktab_id,))
    class_hour_extra = {}
    class_hour_classes = set()
    for row in cur.fetchall():
        class_hour_classes.add(int(row["sinf_id"]))
        teacher_id = row.get("rahbar_user_id")
        if teacher_id is not None:
            class_hour_extra[int(teacher_id)] = class_hour_extra.get(int(teacher_id), 0) + int(row.get("haftalik_soat") or 1)

    class_totals = {int(cls["id"]): 0 for cls in classes}
    subject_totals = []
    warnings = []
    for (class_id, _), data in class_subject.items():
        if data["whole"] and data["groups"]:
            warnings.append(f"{data['fan_nomi']}: butun sinf va guruh qatori birga yozilgan")
        if data["groups"]:
            values = [value for value in data["groups"] if value > 0]
            weekly = max(values, default=0)
            if len(set(values)) > 1:
                warnings.append(f"{data['fan_nomi']}: parallel guruh soatlari teng emas")
        else:
            weekly = sum(data["whole"])
        class_totals[class_id] = class_totals.get(class_id, 0) + weekly
        subject_totals.append({
            "sinf_id": class_id,
            "fan_nomi": data["fan_nomi"],
            "haftalik_soat": weekly,
        })
    for class_id in class_hour_classes:
        class_totals[class_id] = class_totals.get(class_id, 0) + 1
        subject_totals.append({
            "sinf_id": class_id,
            "fan_nomi": "SINF SOATI",
            "haftalik_soat": 1,
        })

    year = _v1852_active_year(cur, maktab_id)
    school_days = 0
    school_weeks = 0.0
    if year:
        current = year["boshlanish"]
        end = year["tugash"]
        weekdays = int(year.get("hafta_kunlari") or 6)
        while current <= end:
            if _v1852_is_school_day(cur, maktab_id, current, weekdays):
                school_days += 1
            current += timedelta(days=1)
        school_weeks = round(school_days / max(1, weekdays), 2)

    teacher_totals = []
    cur.execute("SELECT user_id,full_name FROM users WHERE maktab_id=%s ORDER BY full_name", (maktab_id,))
    for teacher in cur.fetchall():
        teacher_id = int(teacher["user_id"])
        base = round(float(teacher_base.get(teacher_id, 0)), 1)
        extra = int(class_hour_extra.get(teacher_id, 0))
        teacher_totals.append({
            "user_id": teacher_id,
            "full_name": teacher["full_name"],
            "fan_soati": base,
            "sinf_soati": extra,
            "haftalik_jami": base + extra,
        })

    school_weekly = sum(class_totals.values())
    return {
        "oqituvchilar": teacher_totals,
        "sinflar": [
            {
                "sinf_id": int(cls["id"]),
                "sinf": f"{cls['sinf']}-{cls['harf']}",
                "haftalik_soat": round(float(class_totals.get(int(cls["id"]), 0)), 1),
                "yillik_soat": round(float(class_totals.get(int(cls["id"]), 0)) * school_weeks, 1),
            }
            for cls in classes
        ],
        "fanlar": subject_totals,
        "maktab_haftalik_soat": school_weekly,
        "maktab_yillik_soat": round(school_weekly * school_weeks),
        "oquv_kuni": school_days,
        "oquv_haftasi": school_weeks,
        "oqituvchi_soat_jami": sum(item["haftalik_jami"] for item in teacher_totals),
        "ogohlantirishlar": list(dict.fromkeys(warnings)),
    }


def _v2249_ensure_teacher_numbers(cur, maktab_id: int):
    """Har maktab o'qituvchisiga bir marta beriladigan, keyin o'zgarmaydigan jadval raqami."""
    cur.execute(
        """SELECT user_id,jadval_raqami,full_name
           FROM users
           WHERE maktab_id=%s
             AND (LOWER(COALESCE(role,'')) IN ('oqituvchi','teacher')
                  OR LOWER(COALESCE(lavozim,'')) LIKE '%%oqituvchi%%')
           ORDER BY CASE WHEN jadval_raqami IS NULL THEN 1 ELSE 0 END,
                    jadval_raqami NULLS LAST, LOWER(full_name), user_id
           FOR UPDATE""",
        (maktab_id,),
    )
    rows = [dict(row) for row in cur.fetchall()]
    used = {int(row["jadval_raqami"]) for row in rows if row.get("jadval_raqami") is not None}
    next_number = max(used, default=0) + 1
    # Birinchi migratsiyada eski o'qituvchilar alfavit bo'yicha raqamlanadi.
    # Keyin raqam hech qachon F.I.Sh. o'zgargani yoki yangi ustoz qo'shilgani uchun almashmaydi.
    missing = sorted(
        (row for row in rows if row.get("jadval_raqami") is None),
        key=lambda row: (str(row.get("full_name") or "").casefold(), int(row["user_id"])),
    )
    for row in missing:
        while next_number in used:
            next_number += 1
        cur.execute(
            "UPDATE users SET jadval_raqami=%s WHERE user_id=%s AND maktab_id=%s",
            (next_number, int(row["user_id"]), maktab_id),
        )
        row["jadval_raqami"] = next_number
        used.add(next_number)
        next_number += 1
    return {int(row["user_id"]): int(row["jadval_raqami"]) for row in rows if row.get("jadval_raqami") is not None}


def _v204_assignment_revision(rows, classes=None):
    """Exact workload + leadership snapshot hash used to reject stale saves."""
    assignments = sorted(
        (
            int(row.get("user_id") or 0),
            int(row.get("sinf_id") or 0),
            re.sub(r"\s+", " ", str(row.get("fan_nomi") or "")).strip().casefold(),
            str(row.get("guruh_kaliti") or "whole"),
            round(float(row.get("haftalik_soat") or 0), 4),
            int(row.get("kunlik_max") or 1),
            int(row.get("xona_id")) if row.get("xona_id") is not None else None,
        )
        for row in rows or []
    )
    leadership = sorted(
        (
            int(row.get("id") or 0),
            int(row.get("rahbar_user_id"))
            if row.get("rahbar_user_id") is not None else None,
        )
        for row in classes or []
    )
    raw = json.dumps(
        {"assignments": assignments, "leadership": leadership},
        ensure_ascii=False,
        separators=(",", ":"),
    )
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _v204_group_subject_family(value):
    key = _v1875_subject_key(value)
    if re.search(r"chet tili|ingliz tili|english|nemis tili|fransuz tili", key):
        return "chet_tili"
    if re.search(r"rus tili|russki", key):
        return "rus_tili"
    if re.search(r"informatika|axborot texnolog", key):
        return "informatika"
    if re.search(r"jismoniy tarbiya|fizkultura|sport", key):
        return "jismoniy"
    if re.search(r"texnologiya|mehnat", key):
        return "texnologiya"
    return key


def _v204_group_subject_matches(left, right):
    left_key = _v1875_subject_key(left)
    right_key = _v1875_subject_key(right)
    return bool(
        left_key and right_key and (
            left_key == right_key
            or _v204_group_subject_family(left_key)
            == _v204_group_subject_family(right_key)
        )
    )


def _v204_expected_skeleton_rows(cur, maktab_id: int):
    """Return the canonical plan/group keys accepted by both skeleton saves."""
    classes, _systems, variants = _v192_group_variants(cur, maktab_id)
    plan = _v193_plan_payload(cur, maktab_id, classes)
    saved_plan_rows = plan.get("qatorlar") or []
    plan_rows = saved_plan_rows or (plan.get("andoza_qatorlar") or [])
    variants_by_class = {}
    for variant in variants:
        if _v1875_group_key(variant.get("guruh_kaliti")) == "whole":
            continue
        variants_by_class.setdefault(int(variant["sinf_id"]), []).append(variant)

    expected = {}
    for row in plan_rows:
        hours = round(float(row.get("haftalik_soat") or 0), 4)
        if hours <= 0:
            continue
        subject = _v192_clean_subject(row.get("fan_nomi"))
        subject_key = _v1875_subject_key(subject)
        if subject_key in {
            _v1875_subject_key("SINF SOATI"),
            _v1875_subject_key("KELAJAK SOATI"),
        }:
            continue
        class_id = int(row["sinf_id"])
        linked = [
            variant for variant in variants_by_class.get(class_id, [])
            if any(
                _v204_group_subject_matches(item, subject)
                for item in (variant.get("fanlar") or [])
            )
        ]
        group_keys = sorted({
            _v1875_group_key(variant.get("guruh_kaliti"))
            for variant in linked
        }) or ["whole"]
        for group_key in group_keys:
            expected[(class_id, subject_key, group_key)] = {
                "sinf_id": class_id,
                "fan_nomi": subject,
                "fan_kaliti": subject_key,
                "guruh_kaliti": group_key,
                "haftalik_soat": hours,
                "kunlik_max": max(1, min(4, int(row.get("kunlik_max") or 1))),
            }
    return expected


def _v204_resolve_expected_skeleton_key(expected, class_id, subject, group_key):
    """Resolve aliases only when one canonical plan row matches unambiguously."""
    class_id = int(class_id)
    subject_key = _v1875_subject_key(subject)
    group_key = _v1875_group_key(group_key)
    exact = (class_id, subject_key, group_key)
    if exact in expected:
        return exact
    matches = [
        key for key, row in expected.items()
        if key[0] == class_id
        and key[2] == group_key
        and _v204_group_subject_matches(row["fan_nomi"], subject)
    ]
    return matches[0] if len(matches) == 1 else None


def _v204_validate_complete_skeleton_payload(cur, maktab_id: int, qatorlar):
    """Prove full plan/group coverage before the school-wide replacement."""
    expected = _v204_expected_skeleton_rows(cur, maktab_id)
    # _v204_group_subject_matches is applied by the shared resolver so the
    # full and partial endpoints accept exactly the same unambiguous aliases.

    submitted = {}
    for row in qatorlar:
        key = _v204_resolve_expected_skeleton_key(
            expected, row.sinf_id, row.fan_nomi, row.guruh_kaliti,
        )
        if key is None:
            # Noma'lum qator submitted_keys ichida qolib, to'liq snapshot
            # xatosida ortiqcha qator sifatida aniq ko'rinadi.
            key = (
                int(row.sinf_id),
                _v1875_subject_key(row.fan_nomi),
                _v1875_group_key(row.guruh_kaliti),
            )
        submitted.setdefault(key, []).append(row)

    expected_keys = set(expected)
    submitted_keys = set(submitted)
    if expected_keys != submitted_keys:
        missing = sorted(expected_keys - submitted_keys)
        extra = sorted(submitted_keys - expected_keys)
        details = []
        if missing:
            details.append(f"{len(missing)} ta reja qatori yetishmaydi")
        if extra:
            details.append(f"{len(extra)} ta ortiqcha yoki noto'g'ri guruh qatori bor")
        raise HTTPException(
            status_code=409,
            detail="Skelet o'quv reja bilan to'liq mos emas: " + "; ".join(details),
        )

    grouped_teachers = {}
    for key, rows in submitted.items():
        if len(rows) != 1:
            raise HTTPException(
                status_code=409,
                detail="Bitta fan–sinf–guruhga aynan bitta o'qituvchi tanlanishi kerak",
            )
        row = rows[0]
        if abs(
            float(row.haftalik_soat)
            - float(expected[key]["haftalik_soat"])
        ) > 1e-6:
            raise HTTPException(
                status_code=409,
                detail="Skeletdagi haftalik soat tasdiqlangan o'quv reja bilan teng emas",
            )
        if key[2] != "whole":
            grouped_teachers.setdefault(key[:2], []).append(int(row.user_id))
    if any(len(teachers) != len(set(teachers)) for teachers in grouped_teachers.values()):
        raise HTTPException(
            status_code=409,
            detail="Bitta guruhli fanning har bir guruhiga alohida o'qituvchi tanlanishi kerak",
        )
    return {"qator_soni": len(expected_keys)}


def _v192_matrix_payload(cur, maktab_id: int):
    classes, systems, variants = _v192_group_variants(cur, maktab_id)
    plan = _v193_plan_payload(cur, maktab_id, classes)
    plan_subjects_by_class = {}
    for plan_row in plan.get("qatorlar") or plan.get("andoza_qatorlar") or []:
        plan_subjects_by_class.setdefault(int(plan_row["sinf_id"]), []).append(
            plan_row["fan_nomi"]
        )
    rows = _v192_assignment_rows(cur, maktab_id)
    variant_names = {
        (int(row["sinf_id"]), str(row["guruh_kaliti"])): row
        for row in variants
    }
    for row in rows:
        variant = variant_names.get(
            (int(row["sinf_id"]), _v1875_group_key(row.get("guruh_kaliti")))
        )
        row["guruh_nomi"] = (variant or {}).get("guruh_nomi", row["guruh_kaliti"])
        row["guruh_qisqa"] = (variant or {}).get("qisqa", row["guruh_kaliti"])
        row["sinf_nomi"] = f"{row['sinf']}-{row['harf']}"

    _v2249_ensure_teacher_numbers(cur, maktab_id)
    # V22.50: faqat skelet/raqam tanlash oynasi hali fan biriktirilmagan
    # raqamlangan o‘qituvchilarni ham ko‘radi. Jadval generatori esa
    # faqat haqiqiy dars yuklamasi bor o‘qituvchilarni oladi.
    teachers = _v1859_effective_teachers(cur, maktab_id, include_numbered=True)
    cur.execute("SELECT fan_nomi FROM maktab_fanlari WHERE maktab_id=%s ORDER BY fan_nomi", (maktab_id,))
    subjects = [row["fan_nomi"] for row in cur.fetchall()]
    cur.execute("SELECT * FROM aqlli_xonalar_v2 WHERE maktab_id=%s AND faol=TRUE AND darsga_yaroqli=TRUE ORDER BY nomi", (maktab_id,))
    rooms = [dict(row) for row in cur.fetchall()]
    cur.execute("""SELECT avtomatik_tavsiya FROM aqlli_jadval_boshqaruv_v19_2
                   WHERE maktab_id=%s""", (maktab_id,))
    mode = cur.fetchone()
    return {
        "versiya": "teacher-first-approved-curriculum-v19.3",
        "paket": "all-14-sections-updated",
        "sinflar": classes,
        "oqituvchilar": teachers,
        "fanlar": subjects,
        "fan_sinflari": [
            {
                "sinf_id": int(cls["id"]),
                "sinf_darajasi": _v193_grade_number(cls.get("sinf")),
                "talim_tili": _v238_normalize_instruction_language(
                    cls.get("talim_tili")
                ),
                "fanlar": plan_subjects_by_class.get(int(cls["id"]), []),
            }
            for cls in classes
        ],
        "xonalar": rooms,
        "guruh_variantlari": variants,
        "birikmalar": rows,
        "yuklama_revision": _v204_assignment_revision(rows, classes),
        "oquv_reja": plan,
        "hisob": _v192_totals(cur, maktab_id, rows, classes),
        "avtomatik_tavsiya": (
            SAMTM_V19_2_AUTO_SWAP_DEFAULT
            if not mode else bool(mode["avtomatik_tavsiya"])
        ),
    }


@app.get("/api/maktab/aqlli_jadval/v3/yuklama_matritsasi")
def v192_load_matrix(token: str, maktab_id: int):
    actor_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor()
    try:
        _v192_tables(cur)
        if not _v1852_staff(cur, actor_id, maktab_id):
            raise HTTPException(status_code=403, detail="Bu maktab yuklamasini ko'rishga ruxsat yo'q")
        payload = _v192_matrix_payload(cur, maktab_id)
        conn.commit()
        return payload
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


class V193CurriculumItem(BaseModel):
    fan_nomi: str
    haftalik_soat: float
    kunlik_max: int = 1


class V193CurriculumClassSave(BaseModel):
    maktab_id: int
    sinf_id: int
    fanlar: list[V193CurriculumItem]
    talim_tili: Optional[str] = None


class V193CurriculumMatrixItem(BaseModel):
    sinf_id: int
    fan_nomi: str
    haftalik_soat: float
    kunlik_max: int = 1


class V193CurriculumMatrixSave(BaseModel):
    maktab_id: int
    qatorlar: list[V193CurriculumMatrixItem]
    talim_tili: str = "uz"


class V193CurriculumAction(BaseModel):
    maktab_id: int
    talim_tili: str = "uz"


class V20ClassHourPlanItem(BaseModel):
    sinf_id: int
    fan_nomi: str = "KELAJAK SOATI"
    haftalik_soat: int = 1


class V20ClassHourPlanSave(BaseModel):
    maktab_id: int
    qatorlar: list[V20ClassHourPlanItem] = []
    talim_tili: str = "uz"


@app.put("/api/maktab/aqlli_jadval/v3/oquv_reja/sinf_soati")
def v20_class_hour_plan_save(sorov: V20ClassHourPlanSave, token: str):
    """Kelajak soati nomi, soati va qaysi sinflarga tegishliligini saqlaydi."""
    actor_id = _jwt_tekshir(token)
    try:
        language = _v238_normalize_instruction_language(sorov.talim_tili)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    conn = _db(); cur = conn.cursor()
    try:
        _v192_tables(cur)
        if not _v1852_manager(cur, actor_id, sorov.maktab_id):
            raise HTTPException(status_code=403, detail="Kelajak soatini faqat rahbariyat boshqaradi")
        _v201_mark_school_override(
            cur, sorov.maktab_id, "oquv_reja", actor_id, language
        )
        cur.execute(
            "SELECT id FROM maktab_sinflari WHERE maktab_id=%s "
            "AND COALESCE(talim_tili,'uz')=%s",
            (sorov.maktab_id, language),
        )
        valid_ids = {int(row["id"]) for row in cur.fetchall()}
        if not valid_ids:
            raise HTTPException(
                status_code=400,
                detail=f"{language} ta’lim tilidagi sinf topilmadi",
            )
        received = set()
        for item in sorov.qatorlar:
            class_id = int(item.sinf_id)
            if class_id not in valid_ids:
                raise HTTPException(status_code=400, detail="Kelajak soati uchun begona sinf yuborildi")
            name = _v192_clean_subject(item.fan_nomi) or "KELAJAK SOATI"
            hours = int(item.haftalik_soat)
            if hours not in range(1, 6):
                raise HTTPException(status_code=400, detail="Kelajak soati haftasiga 1–5 soat bo‘lishi mumkin")
            received.add(class_id)
            cur.execute("""INSERT INTO aqlli_sinf_soati_qoidalari_v2(
                            maktab_id,sinf_id,hafta_kuni,dars_raqami,faol,
                            yaratgan_user_id,fan_nomi,haftalik_soat,yangilangan_at)
                           VALUES(%s,%s,5,1,TRUE,%s,%s,%s,NOW())
                           ON CONFLICT(maktab_id,sinf_id) DO UPDATE SET
                             faol=TRUE,fan_nomi=EXCLUDED.fan_nomi,
                             haftalik_soat=EXCLUDED.haftalik_soat,
                             yaratgan_user_id=EXCLUDED.yaratgan_user_id,yangilangan_at=NOW()""",
                        (sorov.maktab_id, class_id, actor_id, name, hours))
        cur.execute(
            """UPDATE aqlli_sinf_soati_qoidalari_v2
                  SET faol=FALSE,yangilangan_at=NOW()
                WHERE maktab_id=%s AND sinf_id=ANY(%s)
                  AND NOT(sinf_id=ANY(%s))""",
            (sorov.maktab_id, sorted(valid_ids), sorted(received) or [-1]),
        )
        _v238_mark_plan_languages_draft(cur, sorov.maktab_id, [language])
        cur.execute(
            "UPDATE aqlli_oquv_reja_holati_v19_3 "
            "SET holat='draft',tasdiqlagan_user_id=NULL," 
            "tasdiqlangan_at=NULL,yangilangan_at=NOW() "
            "WHERE maktab_id=%s",
            (sorov.maktab_id,),
        )
        cur.execute("""UPDATE aqlli_jadval_urinishlari_v2 SET holat='bekor'
                         WHERE maktab_id=%s AND holat='draft'""",
                    (sorov.maktab_id,))
        conn.commit()
        return {
            "holat": "saqlandi", "talim_tili": language,
            "faol_sinf_soni": len(received),
        }
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


def _v195_refresh_teacher_summary(cur, maktab_id: int, user_id: int):
    """Aniq birikmalardan o'qituvchining fan, sinf va jami soatini qayta quradi."""
    cur.execute("""SELECT b.sinf_id,b.fan_nomi,COALESCE(b.haftalik_soat,0) AS haftalik_soat,
                          s.sinf,s.harf
                   FROM maktab_dars_birikmalari b
                   JOIN maktab_sinflari s ON s.id=b.sinf_id
                   WHERE b.maktab_id=%s AND b.user_id=%s
                   ORDER BY s.sinf::int,s.harf,b.fan_nomi,b.guruh_kaliti""",
                (maktab_id, user_id))
    rows = [dict(row) for row in cur.fetchall()]
    cur.execute("DELETE FROM maktab_xodim_sinflari WHERE maktab_id=%s AND user_id=%s",
                (maktab_id, user_id))
    by_class = {}
    for row in rows:
        by_class.setdefault(int(row["sinf_id"]), []).append(row["fan_nomi"])
    for class_id, subjects in by_class.items():
        unique_subjects = list(dict.fromkeys(subjects))
        cur.execute("""INSERT INTO maktab_xodim_sinflari(
                        maktab_id,user_id,sinf_id,fanlari)
                       VALUES(%s,%s,%s,%s)""",
                    (maktab_id, user_id, class_id, "; ".join(unique_subjects)))
    subject_list = sorted(
        {row["fan_nomi"] for row in rows}, key=lambda value: value.casefold()
    )
    class_list = sorted(
        {f"{row['sinf']}-{row['harf']}" for row in rows},
        key=_v1859_sinf_sort_key,
    )
    weekly_total = round(sum(float(row.get("haftalik_soat") or 0) for row in rows), 1)
    cur.execute("""UPDATE users
                   SET fanlari=%s,haftalik_dars_soati=%s,oqitadigan_sinflari=%s
                   WHERE user_id=%s AND maktab_id=%s""",
                (
                    "; ".join(subject_list) or None,
                    weekly_total,
                    "; ".join(class_list) or None,
                    user_id,
                    maktab_id,
                ))
    return weekly_total


def _v195_reconcile_teacher_loads_with_plan(cur, maktab_id: int, class_ids=None):
    """Tasdiqlanayotgan reja bilan eski o'qituvchi yuklamalarini moslashtiradi."""
    scoped_ids = sorted({int(value) for value in (class_ids or [])})
    plan_scope = " AND sinf_id=ANY(%s)" if scoped_ids else ""
    plan_args = (maktab_id, scoped_ids) if scoped_ids else (maktab_id,)
    cur.execute("""SELECT sinf_id,fan_nomi,haftalik_soat
                   FROM aqlli_oquv_reja_qatorlari_v19_3
                   WHERE maktab_id=%s""" + plan_scope, plan_args)
    plan = {
        (int(row["sinf_id"]), _v1875_subject_key(row["fan_nomi"])): dict(row)
        for row in cur.fetchall()
    }
    assignment_scope = " AND sinf_id=ANY(%s)" if scoped_ids else ""
    assignment_args = (maktab_id, scoped_ids) if scoped_ids else (maktab_id,)
    cur.execute("""SELECT id,user_id,sinf_id,fan_nomi,
                          COALESCE(NULLIF(guruh_kaliti,''),'whole') AS guruh_kaliti,
                          COALESCE(haftalik_soat,0) AS haftalik_soat
                   FROM maktab_dars_birikmalari
                   WHERE maktab_id=%s""" + assignment_scope + """
                   ORDER BY sinf_id,fan_nomi,guruh_kaliti,id""", assignment_args)
    assignments = [dict(row) for row in cur.fetchall()]
    remaining = {}
    affected_users = set()
    deleted = 0
    reduced = 0
    renamed = 0
    for row in assignments:
        affected_users.add(int(row["user_id"]))
        pair = (int(row["sinf_id"]), _v1875_subject_key(row["fan_nomi"]))
        plan_row = plan.get(pair)
        if not plan_row:
            cur.execute("DELETE FROM maktab_dars_birikmalari WHERE id=%s", (row["id"],))
            deleted += 1
            continue
        exact = (pair[0], pair[1], _v1875_group_key(row.get("guruh_kaliti")))
        if exact not in remaining:
            remaining[exact] = float(plan_row.get("haftalik_soat") or 0)
        old_hours = max(0.0, float(row.get("haftalik_soat") or 0))
        new_hours = min(old_hours, max(0.0, float(remaining[exact])))
        if new_hours <= 0:
            cur.execute("DELETE FROM maktab_dars_birikmalari WHERE id=%s", (row["id"],))
            deleted += 1
            continue
        canonical_subject = str(plan_row["fan_nomi"])
        if new_hours != old_hours or canonical_subject != row["fan_nomi"]:
            cur.execute("""UPDATE maktab_dars_birikmalari
                           SET fan_nomi=%s,haftalik_soat=%s,yangilangan_at=NOW()
                           WHERE id=%s""",
                        (canonical_subject, new_hours, row["id"]))
            if new_hours != old_hours:
                reduced += 1
            if canonical_subject != row["fan_nomi"]:
                renamed += 1
        remaining[exact] -= new_hours
    for user_id in affected_users:
        _v195_refresh_teacher_summary(cur, maktab_id, user_id)
    return {
        "ochirilgan_qator": deleted,
        "qisqartirilgan_qator": reduced,
        "nomi_moslangan_qator": renamed,
        "oqituvchi_soni": len(affected_users),
    }


@app.put("/api/maktab/aqlli_jadval/v3/oquv_reja")
def v193_curriculum_save(sorov: V193CurriculumClassSave, token: str):
    actor_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor()
    try:
        _v192_tables(cur)
        if not _v1852_manager(cur, actor_id, sorov.maktab_id):
            raise HTTPException(status_code=403, detail="O'quv rejasini faqat rahbariyat boshqaradi")
        cur.execute("""SELECT id,sinf,harf,COALESCE(talim_tili,'uz') AS talim_tili
                         FROM maktab_sinflari
                       WHERE id=%s AND maktab_id=%s FOR UPDATE""",
                    (sorov.sinf_id, sorov.maktab_id))
        class_row = cur.fetchone()
        if not class_row:
            raise HTTPException(status_code=404, detail="Sinf topilmadi")
        class_language, selected_by_grade, _curriculum = (
            _v238_class_curriculum_context(
                cur, sorov.maktab_id, class_row, {}
            )
        )
        try:
            requested_language = (
                _v238_normalize_instruction_language(sorov.talim_tili)
                if sorov.talim_tili is not None else class_language
            )
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        if requested_language != class_language:
            raise HTTPException(
                status_code=409,
                detail="Yuborilgan ta’lim tili tanlangan sinf tiliga mos emas",
            )
        _v201_mark_school_override(
            cur, sorov.maktab_id, "oquv_reja", actor_id, class_language
        )
        allowed = (selected_by_grade or {}).get(_v193_grade_number(class_row.get("sinf")), {})
        received = set()
        cleaned = []
        for index, item in enumerate(sorov.fanlar, start=1):
            subject = _v192_clean_subject(item.fan_nomi)
            if not subject:
                raise HTTPException(status_code=400, detail=f"{index}-qator: fan nomi kiritilmagan")
            key = _v1875_subject_key(subject)
            if key == "sinf soati":
                raise HTTPException(
                    status_code=400,
                    detail="SINF SOATI fan qatoriga yozilmaydi; u har sinfga avtomatik 1 soat qo'shiladi",
                )
            if key not in allowed:
                raise HTTPException(
                    status_code=400,
                    detail=(
                        f"{subject}: {class_language} tilidagi "
                        "sinf o‘quv rejasida bu fan yo‘q"
                    ),
                )
            if key in allowed:
                subject = allowed[key]
            if key in received:
                raise HTTPException(status_code=400, detail=f"{subject}: reja qatorida ikki marta yozilgan")
            received.add(key)
            hours = float(item.haftalik_soat)
            daily = int(item.kunlik_max)
            if hours < 0.5 or hours > 20 or hours * 2 != int(hours * 2):
                raise HTTPException(status_code=400, detail=f"{subject}: haftalik soat 0,5–20 oralig'ida, 0,5 qadam bilan bo'lishi kerak")
            if daily < 1 or daily > 4:
                raise HTTPException(status_code=400, detail=f"{subject}: kunlik maksimum 1–4 bo'lishi kerak")
            cleaned.append((subject, hours, daily))
        if not cleaned:
            raise HTTPException(status_code=400, detail="Sinf o'quv rejasida kamida bitta fan bo'lishi kerak")
        total = sum(row[1] for row in cleaned)
        if total > 60:
            raise HTTPException(status_code=400, detail="Sinfning haftalik reja soati 60 dan oshmasligi kerak")
        cur.execute("""INSERT INTO aqlli_oquv_reja_holati_v19_3(maktab_id)
                       VALUES(%s) ON CONFLICT(maktab_id) DO NOTHING""", (sorov.maktab_id,))
        cur.execute("""DELETE FROM aqlli_oquv_reja_qatorlari_v19_3
                       WHERE maktab_id=%s AND sinf_id=%s""",
                    (sorov.maktab_id, sorov.sinf_id))
        for subject, hours, daily in cleaned:
            cur.execute("""INSERT INTO aqlli_oquv_reja_qatorlari_v19_3(
                            maktab_id,sinf_id,fan_nomi,haftalik_soat,
                            kunlik_max,manba,yangilangan_at)
                           VALUES(%s,%s,%s,%s,%s,'qolda',NOW())""",
                        (sorov.maktab_id, sorov.sinf_id, subject, hours, daily))
            cur.execute("""INSERT INTO maktab_fanlari(maktab_id,fan_nomi)
                           VALUES(%s,%s) ON CONFLICT DO NOTHING""",
                        (sorov.maktab_id, subject))
            if class_language == "uz":
                # Bu jadval tarixiy UZ katalogi; RU/EN fanlarini unga yozish
                # o'zbek sinflarining ruxsat etilgan fanlarini ifloslantiradi.
                cur.execute("""INSERT INTO maktab_fan_sinflari_v19_4(maktab_id,fan_nomi,sinf_darajasi)
                               VALUES(%s,%s,%s) ON CONFLICT DO NOTHING""",
                            (sorov.maktab_id, subject, _v193_grade_number(class_row.get("sinf"))))
        cur.execute("""UPDATE aqlli_oquv_reja_holati_v19_3
                       SET holat='draft',versiya=versiya+1,
                           tasdiqlagan_user_id=NULL,tasdiqlangan_at=NULL,
                           yangilangan_at=NOW() WHERE maktab_id=%s""",
                    (sorov.maktab_id,))
        cur.execute("""INSERT INTO aqlli_oquv_reja_til_holati_v238(
                        maktab_id,talim_tili,holat,versiya,yangilangan_at)
                       VALUES(%s,%s,'draft',1,NOW())
                       ON CONFLICT(maktab_id,talim_tili) DO UPDATE SET
                         holat='draft',versiya=aqlli_oquv_reja_til_holati_v238.versiya+1,
                         tasdiqlagan_user_id=NULL,tasdiqlangan_at=NULL,
                         yangilangan_at=NOW()""",
                    (sorov.maktab_id, class_language))
        cur.execute("""UPDATE aqlli_jadval_urinishlari_v2 SET holat='bekor'
                       WHERE maktab_id=%s AND holat='draft'""", (sorov.maktab_id,))
        class_hour_result = _v199_ensure_class_hour_rules(
            cur, sorov.maktab_id, [sorov.sinf_id], actor_id
        )
        result = _v192_matrix_payload(cur, sorov.maktab_id)
        conn.commit()
        return {
            "holat": "draft_saqlandi",
            "sinf": f"{class_row['sinf']}-{class_row['harf']}",
            "talim_tili": class_language,
            "fan_soni": len(cleaned),
            "haftalik_jami": total,
            "sinf_soati": class_hour_result,
            "matritsa": result,
        }
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


@app.put("/api/maktab/aqlli_jadval/v3/oquv_reja/matritsa")
def v193_curriculum_matrix_save(sorov: V193CurriculumMatrixSave, token: str):
    actor_id = _jwt_tekshir(token)
    try:
        language = _v238_normalize_instruction_language(sorov.talim_tili)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    conn = _db(); cur = conn.cursor()
    try:
        _v192_tables(cur)
        if not _v1852_manager(cur, actor_id, sorov.maktab_id):
            raise HTTPException(status_code=403, detail="O'quv rejasini faqat rahbariyat boshqaradi")
        _v201_mark_school_override(
            cur, sorov.maktab_id, "oquv_reja", actor_id, language
        )
        cur.execute("""SELECT id,sinf,harf,COALESCE(talim_tili,'uz') AS talim_tili
                         FROM maktab_sinflari
                       WHERE maktab_id=%s AND COALESCE(talim_tili,'uz')=%s
                       ORDER BY CASE WHEN sinf::text ~ '^\\d+$' THEN sinf::int ELSE 999 END,harf
                       FOR UPDATE""",
                    (sorov.maktab_id, language))
        classes = [dict(row) for row in cur.fetchall()]
        valid_classes = {int(row["id"]): row for row in classes}
        if not valid_classes:
            raise HTTPException(status_code=400, detail="Maktabda birorta ham sinf topilmadi")
        if not sorov.qatorlar:
            raise HTTPException(status_code=400, detail="O'quv reja matritsasi bo'sh")
        context_cache = {}
        allowed_by_class = {}
        for class_id, class_row in valid_classes.items():
            class_language, selected_by_grade, _curriculum = (
                _v238_class_curriculum_context(
                    cur, sorov.maktab_id, class_row, context_cache
                )
            )
            if class_language != language:
                raise HTTPException(
                    status_code=409,
                    detail="Sinf ta'lim tili matritsa tiliga mos emas",
                )
            allowed_by_class[class_id] = (selected_by_grade or {}).get(
                _v193_grade_number(class_row.get("sinf")), {}
            )
        seen = set()
        cleaned = []
        class_totals = {class_id: 0 for class_id in valid_classes}
        for index, item in enumerate(sorov.qatorlar, start=1):
            class_id = int(item.sinf_id)
            if class_id not in valid_classes:
                raise HTTPException(status_code=400, detail=f"{index}-qator: sinf bu maktabga tegishli emas")
            subject = _v192_clean_subject(item.fan_nomi)
            if not subject:
                raise HTTPException(status_code=400, detail=f"{index}-qator: fan nomi kiritilmagan")
            subject_key = _v1875_subject_key(subject)
            if subject_key == "sinf soati":
                raise HTTPException(
                    status_code=400,
                    detail="SINF SOATI fan qatoriga yozilmaydi; u har sinfga avtomatik 1 soat qo'shiladi",
                )
            allowed = allowed_by_class[class_id]
            if subject_key not in allowed:
                raise HTTPException(
                    status_code=400,
                    detail=(
                        f"{subject}: {language} tilidagi sinf o‘quv "
                        "rejasida bu fan yo‘q"
                    ),
                )
            if subject_key in allowed:
                subject = allowed[subject_key]
            hours = float(item.haftalik_soat)
            daily = int(item.kunlik_max)
            if hours < 0.5 or hours > 20 or hours * 2 != int(hours * 2):
                raise HTTPException(status_code=400, detail=f"{subject}: haftalik soat 0,5–20 oralig'ida, 0,5 qadam bilan bo'lishi kerak")
            if daily < 1 or daily > 4:
                raise HTTPException(status_code=400, detail=f"{subject}: kunlik maksimum 1–4 bo'lishi kerak")
            key = (class_id, _v1875_subject_key(subject))
            if key in seen:
                raise HTTPException(
                    status_code=400,
                    detail=f"{valid_classes[class_id]['sinf']}-{valid_classes[class_id]['harf']} / {subject}: ikki marta yuborilgan",
                )
            seen.add(key)
            class_totals[class_id] += hours
            if class_totals[class_id] > 60:
                cls = valid_classes[class_id]
                raise HTTPException(
                    status_code=400,
                    detail=f"{cls['sinf']}-{cls['harf']}: haftalik jami 60 soatdan oshdi",
                )
            cleaned.append((class_id, subject, hours, daily))
        missing_class_ids = [
            class_id for class_id, total in class_totals.items() if total <= 0
        ]
        if missing_class_ids:
            labels = [
                f"{valid_classes[class_id]['sinf']}-{valid_classes[class_id]['harf']}"
                for class_id in missing_class_ids
            ]
            raise HTTPException(
                status_code=409,
                detail=(
                    f"{language} til matritsasi to‘liq emas. Qator yuborilmagan "
                    "sinflar: " + ", ".join(labels[:30])
                ),
            )
        cur.execute("""INSERT INTO aqlli_oquv_reja_holati_v19_3(maktab_id)
                       VALUES(%s) ON CONFLICT(maktab_id) DO NOTHING""", (sorov.maktab_id,))
        target_class_ids = sorted(valid_classes)
        cur.execute(
            "DELETE FROM aqlli_oquv_reja_qatorlari_v19_3 "
            "WHERE maktab_id=%s AND sinf_id=ANY(%s)",
            (sorov.maktab_id, target_class_ids),
        )
        for class_id, subject, hours, daily in cleaned:
            cur.execute("""INSERT INTO aqlli_oquv_reja_qatorlari_v19_3(
                            maktab_id,sinf_id,fan_nomi,haftalik_soat,
                            kunlik_max,manba,yangilangan_at)
                           VALUES(%s,%s,%s,%s,%s,'matritsa',NOW())""",
                        (sorov.maktab_id, class_id, subject, hours, daily))
            cur.execute("""INSERT INTO maktab_fanlari(maktab_id,fan_nomi)
                           VALUES(%s,%s) ON CONFLICT DO NOTHING""",
                        (sorov.maktab_id, subject))
            if language == "uz":
                cur.execute("""INSERT INTO maktab_fan_sinflari_v19_4(maktab_id,fan_nomi,sinf_darajasi)
                               VALUES(%s,%s,%s) ON CONFLICT DO NOTHING""",
                            (sorov.maktab_id, subject,
                             _v193_grade_number(valid_classes[class_id].get("sinf"))))
        cur.execute("""UPDATE aqlli_oquv_reja_holati_v19_3
                       SET holat='draft',versiya=versiya+1,
                           tasdiqlagan_user_id=NULL,tasdiqlangan_at=NULL,
                           yangilangan_at=NOW() WHERE maktab_id=%s""",
                    (sorov.maktab_id,))
        cur.execute("""INSERT INTO aqlli_oquv_reja_til_holati_v238(
                        maktab_id,talim_tili,holat,versiya,yangilangan_at)
                       VALUES(%s,%s,'draft',1,NOW())
                       ON CONFLICT(maktab_id,talim_tili) DO UPDATE SET
                         holat='draft',versiya=aqlli_oquv_reja_til_holati_v238.versiya+1,
                         tasdiqlagan_user_id=NULL,tasdiqlangan_at=NULL,
                         yangilangan_at=NOW()""",
                    (sorov.maktab_id, language))
        cur.execute("""UPDATE aqlli_jadval_urinishlari_v2 SET holat='bekor'
                       WHERE maktab_id=%s AND holat='draft'""", (sorov.maktab_id,))
        class_hour_result = _v199_ensure_class_hour_rules(
            cur, sorov.maktab_id, valid_classes.keys(), actor_id
        )
        matrix = _v192_matrix_payload(cur, sorov.maktab_id)
        conn.commit()
        return {
            "holat": "matritsa_draft_saqlandi",
            "talim_tili": language,
            "sinf_soni": len(valid_classes),
            "fan_qatori": len(cleaned),
            "maktab_haftalik_jami": sum(class_totals.values()),
            "sinf_soati": class_hour_result,
            "matritsa": matrix,
        }
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


@app.post("/api/maktab/aqlli_jadval/v3/oquv_reja/tasdiqlash")
def v193_curriculum_approve(sorov: V193CurriculumAction, token: str):
    actor_id = _jwt_tekshir(token)
    try:
        language = _v238_normalize_instruction_language(sorov.talim_tili)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    conn = _db(); cur = conn.cursor()
    try:
        _v192_tables(cur)
        if not _v1852_manager(cur, actor_id, sorov.maktab_id):
            raise HTTPException(status_code=403, detail="O'quv rejasini faqat rahbariyat tasdiqlaydi")
        all_classes, _systems, _variants = _v192_group_variants(
            cur, sorov.maktab_id
        )
        _v193_ensure_plan(cur, sorov.maktab_id, all_classes)
        classes = [
            row for row in all_classes
            if _v238_normalize_instruction_language(row.get("talim_tili"))
            == language
        ]
        if not classes:
            raise HTTPException(
                status_code=400,
                detail=f"{language} ta’lim tilidagi sinf topilmadi",
            )
        target_class_ids = sorted(int(row["id"]) for row in classes)
        cur.execute("""SELECT sinf_id,COUNT(*) AS fan_soni,SUM(haftalik_soat) AS jami
                       FROM aqlli_oquv_reja_qatorlari_v19_3
                       WHERE maktab_id=%s AND sinf_id=ANY(%s)
                       GROUP BY sinf_id""",
                    (sorov.maktab_id, target_class_ids))
        totals = {int(row["sinf_id"]): row for row in cur.fetchall()}
        missing = [
            f"{row['sinf']}-{row['harf']}"
            for row in classes
            if int(row["id"]) not in totals or int(totals[int(row["id"])].get("jami") or 0) < 1
        ]
        if missing:
            raise HTTPException(
                status_code=400,
                detail="Reja fanlari kiritilmagan sinflar: " + ", ".join(missing[:20]),
            )
        cur.execute("""UPDATE aqlli_sinf_fan_yuklamalari_v2
                       SET haftalik_soat=0,asosiy_oqituvchi_user_id=NULL
                       WHERE maktab_id=%s AND sinf_id=ANY(%s)""",
                    (sorov.maktab_id, target_class_ids))
        cur.execute("""SELECT sinf_id,fan_nomi,haftalik_soat,kunlik_max
                       FROM aqlli_oquv_reja_qatorlari_v19_3
                       WHERE maktab_id=%s AND sinf_id=ANY(%s)""",
                    (sorov.maktab_id, target_class_ids))
        plan_rows = cur.fetchall()
        for row in plan_rows:
            cur.execute("""INSERT INTO aqlli_sinf_fan_yuklamalari_v2(
                            maktab_id,sinf_id,fan_nomi,haftalik_soat,kunlik_max)
                           VALUES(%s,%s,%s,%s,%s)
                           ON CONFLICT(maktab_id,sinf_id,fan_nomi) DO UPDATE SET
                             haftalik_soat=EXCLUDED.haftalik_soat,
                             kunlik_max=EXCLUDED.kunlik_max""",
                        (
                            sorov.maktab_id, row["sinf_id"], row["fan_nomi"],
                            row["haftalik_soat"], row["kunlik_max"],
                        ))
        cur.execute("""INSERT INTO aqlli_oquv_reja_til_holati_v238(
                        maktab_id,talim_tili,holat,tasdiqlagan_user_id,
                        tasdiqlangan_at,yangilangan_at)
                       VALUES(%s,%s,'tasdiqlangan',%s,NOW(),NOW())
                       ON CONFLICT(maktab_id,talim_tili) DO UPDATE SET
                         holat='tasdiqlangan',
                         tasdiqlagan_user_id=EXCLUDED.tasdiqlagan_user_id,
                         tasdiqlangan_at=NOW(),yangilangan_at=NOW()""",
                    (sorov.maktab_id, language, actor_id))
        used_languages = sorted({
            _v238_normalize_instruction_language(row.get("talim_tili"))
            for row in all_classes
        })
        cur.execute("""SELECT talim_tili,holat
                         FROM aqlli_oquv_reja_til_holati_v238
                        WHERE maktab_id=%s AND talim_tili=ANY(%s)""",
                    (sorov.maktab_id, used_languages))
        approved_languages = {
            row["talim_tili"] for row in cur.fetchall()
            if row.get("holat") == "tasdiqlangan"
        }
        all_approved = set(used_languages) <= approved_languages
        cur.execute("""UPDATE aqlli_oquv_reja_holati_v19_3
                       SET holat=%s,
                           tasdiqlagan_user_id=CASE WHEN %s THEN %s ELSE NULL END,
                           tasdiqlangan_at=CASE WHEN %s THEN NOW() ELSE NULL END,
                           yangilangan_at=NOW()
                       WHERE maktab_id=%s""",
                    (
                        "tasdiqlangan" if all_approved else "draft",
                        all_approved, actor_id, all_approved, sorov.maktab_id,
                    ))
        reconcile = _v195_reconcile_teacher_loads_with_plan(cur,
            sorov.maktab_id, target_class_ids)
        class_hour_result = _v199_ensure_class_hour_rules(
            cur, sorov.maktab_id, [row["id"] for row in classes], actor_id
        )
        sync_warnings = []
        if all_approved:
            sync_warnings = _v192_sync_schedule_sources(cur, sorov.maktab_id)
        else:
            sync_warnings.append(
                "Qolgan ta’lim tillari tasdiqlangach umumiy jadval manbasi sinxronlanadi"
            )
        if reconcile["ochirilgan_qator"]:
            sync_warnings.append(
                f"O'quv rejada qolmagan {reconcile['ochirilgan_qator']} ta eski o'qituvchi yuklama qatori olib tashlandi"
            )
        if reconcile["qisqartirilgan_qator"]:
            sync_warnings.append(
                f"{reconcile['qisqartirilgan_qator']} ta o'qituvchi yuklamasi yangi reja soatidan oshmaguncha qisqartirildi"
            )
        cur.execute("""UPDATE aqlli_jadval_urinishlari_v2 SET holat='bekor'
                       WHERE maktab_id=%s AND holat='draft'""", (sorov.maktab_id,))
        result = _v192_matrix_payload(cur, sorov.maktab_id)
        conn.commit()
        return {
            "holat": "tasdiqlangan",
            "talim_tili": language,
            "barcha_tillar_tasdiqlandi": all_approved,
            "sinf_soni": len(classes),
            "fan_qatori": len(plan_rows),
            "yuklama_moslash": reconcile,
            "sinf_soati": class_hour_result,
            "ogohlantirishlar": sync_warnings,
            "matritsa": result,
        }
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


class V192TeacherLoadRow(BaseModel):
    sinf_id: int
    fan_nomi: str
    guruh_kaliti: str = "whole"
    haftalik_soat: float
    kunlik_max: int = 1
    xona_id: Optional[int] = None


class V192TeacherLoadSave(BaseModel):
    maktab_id: int
    user_id: int
    full_name: Optional[str] = None
    mutaxassisligi: Optional[str] = None
    otadigan_fanlari: Optional[list[str]] = None
    haftalik_maqsad_soat: Optional[float] = None
    tugilgan_sana: Optional[date] = None
    tugilgan_yili: Optional[int] = None
    ish_staji: Optional[int] = None
    toifasi: Optional[str] = None
    rahbar_sinf_id: Optional[int] = None
    qatorlar: list[V192TeacherLoadRow]


class V204SkeletonLoadRow(BaseModel):
    user_id: int
    sinf_id: int
    fan_nomi: str
    guruh_kaliti: str = "whole"
    haftalik_soat: float
    kunlik_max: int = 1
    xona_id: Optional[int] = None


class V204SkeletonLeaderRow(BaseModel):
    sinf_id: int
    user_id: Optional[int] = None


class V204SkeletonBulkSave(BaseModel):
    maktab_id: int
    qatorlar: list[V204SkeletonLoadRow] = []
    rahbarlar: list[V204SkeletonLeaderRow] = []
    toliq_snapshot: bool = False
    kutilgan_yuklama_revision: Optional[str] = None


class V204SkeletonAssignmentKey(BaseModel):
    sinf_id: int
    fan_nomi: str
    guruh_kaliti: str = "whole"


class V204SkeletonPatchSave(BaseModel):
    """Only keys explicitly listed here are changed; every other row survives."""
    maktab_id: int
    qatorlar: list[V204SkeletonLoadRow] = []
    ochirilgan_kalitlar: list[V204SkeletonAssignmentKey] = []
    rahbarlar: list[V204SkeletonLeaderRow] = []
    kutilgan_yuklama_revision: Optional[str] = None


class V192ManualTeacherCreate(BaseModel):
    maktab_id: int
    full_name: str
    mutaxassisligi: Optional[str] = None
    otadigan_fanlari: Optional[list[str]] = None
    haftalik_maqsad_soat: Optional[float] = None
    tugilgan_sana: Optional[date] = None
    tugilgan_yili: Optional[int] = None
    ish_staji: Optional[int] = None
    toifasi: Optional[str] = None
    rahbar_sinf_id: Optional[int] = None
    qatorlar: list[V192TeacherLoadRow] = []


def _v194_teacher_profile_values(mutaxassisligi, haftalik_maqsad_soat):
    specialty = re.sub(r"\s+", " ", str(mutaxassisligi or "")).strip() or None
    if specialty and len(specialty) > 120:
        raise HTTPException(status_code=400, detail="Mutaxassislik nomi 120 ta belgidan oshmasligi kerak")
    target = haftalik_maqsad_soat
    if target is not None:
        target = round(float(target), 1)
        if target < 0.5 or target > 60 or abs(target * 2 - round(target * 2)) > 1e-9:
            raise HTTPException(
                status_code=400,
                detail="Haftalik maqsad soati 0,5–60 oralig'ida va 0,5 qadamda bo'lishi kerak",
            )
    return specialty, target


def _v199_save_teacher_leadership(cur, maktab_id: int, user_id: int,
                                  rahbar_sinf_id, actor_id: int):
    """O'qituvchining bitta sinf rahbarligini va uning 1 soat sinf soatini saqlaydi."""
    leader_class = None
    if rahbar_sinf_id is not None:
        cur.execute("""SELECT id,sinf,harf,rahbar_user_id
                       FROM maktab_sinflari
                       WHERE id=%s AND maktab_id=%s FOR UPDATE""",
                    (int(rahbar_sinf_id), int(maktab_id)))
        leader_class = cur.fetchone()
        if not leader_class:
            raise HTTPException(status_code=404, detail="Sinf rahbarligi uchun tanlangan sinf topilmadi")
        current_leader = leader_class.get("rahbar_user_id")
        if current_leader is not None and int(current_leader) != int(user_id):
            raise HTTPException(
                status_code=409,
                detail=f"{leader_class['sinf']}-{leader_class['harf']} sinfiga boshqa rahbar tayinlangan",
            )
    cur.execute("""UPDATE maktab_sinflari SET rahbar_user_id=NULL
                   WHERE maktab_id=%s AND rahbar_user_id=%s
                     AND (%s::INTEGER IS NULL OR id<>%s::INTEGER)""",
                (maktab_id, user_id, rahbar_sinf_id, rahbar_sinf_id))
    if leader_class:
        cur.execute("""UPDATE maktab_sinflari SET rahbar_user_id=%s
                       WHERE id=%s AND maktab_id=%s""",
                    (user_id, int(leader_class["id"]), maktab_id))
        _v199_ensure_class_hour_rules(
            cur, maktab_id, [int(leader_class["id"])], actor_id
        )
    return leader_class


def _v199_teacher_total_with_class_hour(cur, maktab_id: int, user_id: int, result: dict):
    fan_hours = round(float(result.get("haftalik_jami") or 0), 1)
    cur.execute("""SELECT COALESCE(SUM(q.haftalik_soat),0) AS son
                   FROM aqlli_sinf_soati_qoidalari_v2 q
                   JOIN maktab_sinflari s ON s.id=q.sinf_id
                   WHERE q.maktab_id=%s AND q.faol=TRUE AND s.rahbar_user_id=%s""",
                (maktab_id, user_id))
    class_hours = int((cur.fetchone() or {}).get("son") or 0)
    result["fan_soati"] = fan_hours
    result["sinf_soati"] = class_hours
    result["haftalik_jami"] = round(fan_hours + class_hours, 1)
    return result


def _v192_sync_schedule_sources(cur, maktab_id: int):
    rows = _v192_assignment_rows(cur, maktab_id)
    pairs = {}
    for row in rows:
        subject = _v192_clean_subject(row["fan_nomi"])
        pair = pairs.setdefault(
            (int(row["sinf_id"]), _v1875_subject_key(subject)),
            {"sinf_id": int(row["sinf_id"]), "fan_nomi": subject, "rows": []},
        )
        pair["rows"].append(row)

    approved_plan = _v193_approved_plan_map(cur, maktab_id)
    if approved_plan is None:
        # Bir til qayta tahrirlanayotgan paytda global holat draft bo'ladi.
        # Shu vaqtda boshqa tillarning tasdiqlangan yuklama/guruh ma'lumotini
        # nolga tushirish yoki o'chirish mutlaqo mumkin emas.
        return [
            "O'qituvchi yuklamasi qo'lda saqlandi. O'quv reja tasdiqlanmagani uchun avtomatik soat va dars jadvali manbasi hali yoqilmadi"
        ]
    else:
        # Tasdiqlangan rejada hali o'qituvchi biriktirilmagan fanlarning
        # soati yo'qolmaydi. Faqat aniq birikma bor fanlar quyida yangilanadi.
        cur.execute("""UPDATE aqlli_sinf_fan_yuklamalari_v2
                       SET asosiy_oqituvchi_user_id=NULL,xona_id=NULL
                       WHERE maktab_id=%s""", (maktab_id,))
    # Guruh qatorlari UPSERT qilinadi. Maktab bo'yicha ommaviy DELETE eski
    # 1/2 yoki O'g'il/Qiz sxemasini va boshqa til sozlamasini yo'qotardi.
    warnings = []
    for pair in pairs.values():
        whole = [row for row in pair["rows"] if _v1875_group_key(row.get("guruh_kaliti")) == "whole"]
        groups = [row for row in pair["rows"] if _v1875_group_key(row.get("guruh_kaliti")) != "whole"]
        if whole and groups:
            warnings.append(
                f"{pair['fan_nomi']}: butun sinf va guruh qatorlari birga yozilgan"
            )
        active = groups if groups else whole
        hours = [max(0.0, float(row.get("haftalik_soat") or 0)) for row in active]
        weekly = max(hours, default=0) if groups else sum(hours)
        daily = max([int(row.get("kunlik_max") or 1) for row in active], default=1)
        if groups and len(set(value for value in hours if value > 0)) > 1:
            warnings.append(
                f"{pair['fan_nomi']}: parallel guruhlar haftalik soati teng emas"
            )
        primary = int(whole[0]["user_id"]) if len(whole) == 1 and not groups else None
        room = int(whole[0]["xona_id"]) if len(whole) == 1 and whole[0].get("xona_id") else None
        cur.execute("""INSERT INTO aqlli_sinf_fan_yuklamalari_v2(
                        maktab_id,sinf_id,fan_nomi,haftalik_soat,kunlik_max,
                        asosiy_oqituvchi_user_id,xona_id)
                       VALUES(%s,%s,%s,%s,%s,%s,%s)
                       ON CONFLICT(maktab_id,sinf_id,fan_nomi) DO UPDATE SET
                         haftalik_soat=EXCLUDED.haftalik_soat,
                         kunlik_max=EXCLUDED.kunlik_max,
                         asosiy_oqituvchi_user_id=EXCLUDED.asosiy_oqituvchi_user_id,
                         xona_id=EXCLUDED.xona_id""",
                    (
                        maktab_id, pair["sinf_id"], pair["fan_nomi"], weekly,
                        daily, primary, room,
                    ))
        for row in groups:
            cur.execute("""INSERT INTO aqlli_guruh_sozlamalari_v2(
                            maktab_id,sinf_id,fan_nomi,guruh_kaliti,
                            oqituvchi_user_id,xona_id)
                           VALUES(%s,%s,%s,%s,%s,%s)
                           ON CONFLICT(maktab_id,sinf_id,fan_nomi,guruh_kaliti)
                           DO UPDATE SET
                             oqituvchi_user_id=EXCLUDED.oqituvchi_user_id,
                             xona_id=EXCLUDED.xona_id""",
                        (
                            maktab_id, pair["sinf_id"], pair["fan_nomi"],
                            _v1875_group_key(row.get("guruh_kaliti")),
                            row["user_id"], row.get("xona_id"),
                        ))
    cur.execute("""UPDATE aqlli_jadval_urinishlari_v2 SET holat='bekor'
                   WHERE maktab_id=%s AND holat='draft'""", (maktab_id,))
    cur.execute("""UPDATE aqlli_fan_guruh_tasdiqlari_v2
                   SET tasdiqlangan=FALSE,yangilangan_at=NOW()
                   WHERE maktab_id=%s""", (maktab_id,))
    return list(dict.fromkeys(warnings))


def _v192_auto_confirm_exact_pairs(cur, maktab_id: int, actor_id: int):
    """Saytdagi aniq qatorlar to'liq bo'lsa alohida qayta tasdiq talab qilmaydi."""
    systems = _v1876_group_system_catalog(cur, maktab_id)
    systems_by_class = {}
    for system in systems:
        systems_by_class.setdefault(int(system["sinf_id"]), []).append(system)
    pairs = _v1876_pair_assignment_rows(cur, maktab_id)
    confirmed = 0
    pending = 0
    for pair in pairs.values():
        rows = pair["rows"]
        whole = [row for row in rows if row["guruh_kaliti"] == "whole"]
        groups = [row for row in rows if row["guruh_kaliti"] != "whole"]
        mode = None
        system_id = None
        valid = False
        if len(whole) == 1 and not groups and float(whole[0].get("haftalik_soat") or 0) > 0:
            mode = "whole"
            valid = True
        elif groups and not whole:
            group_keys = {str(row["guruh_kaliti"]) for row in groups}
            hours = {float(row.get("haftalik_soat") or 0) for row in groups}
            teachers = {int(row["user_id"]) for row in groups}
            matching_system = next((
                system for system in systems_by_class.get(pair["sinf_id"], [])
                if pair["fan_kaliti"] in (system.get("fan_kalitlari") or [])
                and {
                    str(group["guruh_kaliti"])
                    for group in system.get("guruhlar") or []
                } == group_keys
            ), None)
            if (
                matching_system
                and len(hours) == 1
                and next(iter(hours), 0) > 0
                and len(teachers) == len(groups)
            ):
                mode = "group"
                system_id = int(matching_system["id"])
                valid = True
        pair_systems = systems_by_class.get(pair["sinf_id"], [])
        source_hash = _v1876_pair_hash(rows, pair_systems, pair["fan_kaliti"])
        cur.execute("""INSERT INTO aqlli_fan_guruh_tasdiqlari_v2(
                        maktab_id,sinf_id,fan_nomi,fan_kaliti,turi,tizim_id,
                        manba_hash,tasdiqlangan,tasdiqlagan_user_id,yangilangan_at)
                       VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW())
                       ON CONFLICT(maktab_id,sinf_id,fan_kaliti) DO UPDATE SET
                         fan_nomi=EXCLUDED.fan_nomi,turi=EXCLUDED.turi,
                         tizim_id=EXCLUDED.tizim_id,manba_hash=EXCLUDED.manba_hash,
                         tasdiqlangan=EXCLUDED.tasdiqlangan,
                         tasdiqlagan_user_id=EXCLUDED.tasdiqlagan_user_id,
                         yangilangan_at=NOW()""",
                    (
                        maktab_id, pair["sinf_id"], pair["fan_nomi"],
                        pair["fan_kaliti"], mode or "group", system_id,
                        source_hash, valid, actor_id if valid else None,
                    ))
        if valid:
            confirmed += 1
        else:
            pending += 1
    return {"avto_tasdiqlangan": confirmed, "kutilmoqda": pending}


def _v192_save_teacher_load_rows(
    cur, actor_id: int, maktab_id: int, user_id: int,
    qatorlar: list[V192TeacherLoadRow],
    finalize: bool = True,
):
    cur.execute("SELECT pg_advisory_xact_lock(%s)", (1925000000 + int(maktab_id),))
    cur.execute("""SELECT user_id,full_name FROM users
                   WHERE user_id=%s AND maktab_id=%s FOR UPDATE""",
                (user_id, maktab_id))
    teacher = cur.fetchone()
    if not teacher:
        raise HTTPException(status_code=404, detail="O'qituvchi topilmadi")
    approved_plan = _v193_approved_plan_map(cur, maktab_id)
    classes, _systems, variants = _v192_group_variants(cur, maktab_id)
    valid_classes = {int(row["id"]) for row in classes}
    cur.execute("SELECT fan_nomi FROM maktab_fanlari WHERE maktab_id=%s", (maktab_id,))
    fallback_subjects = {
        _v1875_subject_key(row["fan_nomi"]): row["fan_nomi"]
        for row in cur.fetchall()
    }
    allowed_by_class = {}
    context_cache = {}
    for cls in classes:
        grade = _v193_grade_number(cls.get("sinf"))
        _language, selected_by_grade, _curriculum = (
            _v238_class_curriculum_context(
                cur, maktab_id, cls, context_cache
            )
        )
        allowed_by_class[int(cls["id"])] = (
            dict((selected_by_grade or {}).get(grade, {}))
            if selected_by_grade is not None else dict(fallback_subjects)
        )
    variant_map = {
        (int(row["sinf_id"]), str(row["guruh_kaliti"])): row
        for row in variants
    }
    cleaned_by_key = {}
    duplicate_merged = 0
    plan_overrides = []
    for index, item in enumerate(qatorlar, start=1):
        if int(item.sinf_id) not in valid_classes:
            raise HTTPException(status_code=400, detail=f"{index}-qator: sinf bu maktabga tegishli emas")
        subject = _v192_clean_subject(item.fan_nomi)
        if not subject:
            raise HTTPException(status_code=400, detail=f"{index}-qator: fan tanlanmagan")
        subject_key = _v1875_subject_key(subject)
        plan_item = (
            approved_plan.get((int(item.sinf_id), subject_key))
            if approved_plan is not None else None
        )
        if approved_plan is not None and not plan_item:
            raise HTTPException(
                status_code=400,
                detail=f"{index}-qator: {subject} tanlangan sinfning tasdiqlangan o'quv rejasida yo'q",
            )
        if approved_plan is None:
            allowed_subject = allowed_by_class.get(int(item.sinf_id), {}).get(subject_key)
            if not allowed_subject:
                raise HTTPException(
                    status_code=400,
                    detail=f"{index}-qator: {subject} bu sinf uchun maktab fanlari ro'yxatida yo'q",
                )
            subject = allowed_subject
        group_key = _v1875_group_key(item.guruh_kaliti)
        variant = variant_map.get((int(item.sinf_id), group_key))
        if not variant:
            raise HTTPException(status_code=400, detail=f"{index}-qator: tanlangan guruh bu sinfda yo'q")
        hours = round(float(item.haftalik_soat), 1)
        daily = int(item.kunlik_max)
        if hours < 0.5 or hours > 20 or abs(hours * 2 - round(hours * 2)) > 1e-9:
            raise HTTPException(
                status_code=400,
                detail=f"{index}-qator: haftalik soat 0,5–20 oralig'ida va 0,5 qadamda bo'lishi kerak",
            )
        if daily < 1 or daily > 4:
            raise HTTPException(status_code=400, detail=f"{index}-qator: kunlik maksimum 1–4 bo'lishi kerak")
        key = (int(item.sinf_id), _v1875_subject_key(subject), group_key)
        existing = cleaned_by_key.get(key)
        if existing:
            existing["haftalik_soat"] += hours
            existing["kunlik_max"] = max(existing["kunlik_max"], daily)
            if not existing.get("xona_id") and item.xona_id:
                existing["xona_id"] = int(item.xona_id)
            duplicate_merged += 1
        else:
            cleaned_by_key[key] = {
                "sinf_id": int(item.sinf_id),
                "fan_nomi": subject,
                "fan_kaliti": _v1875_subject_key(subject),
                "guruh_kaliti": group_key,
                "haftalik_soat": hours,
                "kunlik_max": daily,
                "xona_id": int(item.xona_id) if item.xona_id else None,
                "variant": variant,
            }

    cleaned = list(cleaned_by_key.values())
    teacher_group_pairs = {}
    for row in cleaned:
        if row["guruh_kaliti"] == "whole":
            continue
        teacher_group_pairs.setdefault(
            (row["sinf_id"], row["fan_kaliti"]), set()
        ).add(row["guruh_kaliti"])
    if any(len(group_keys) > 1 for group_keys in teacher_group_pairs.values()):
        raise HTTPException(
            status_code=409,
            detail="Bitta o'qituvchiga bir sinfdagi bir guruhli fanning faqat bitta aniq guruhi beriladi",
        )
    for row in cleaned:
        cur.execute("""SELECT COALESCE(SUM(b.haftalik_soat),0) AS band_soat,
                              STRING_AGG(DISTINCT u.full_name, ', ' ORDER BY u.full_name) AS oqituvchilar
                       FROM maktab_dars_birikmalari b
                       JOIN users u ON u.user_id=b.user_id
                       WHERE b.maktab_id=%s AND b.sinf_id=%s
                         AND LOWER(TRIM(b.fan_nomi))=LOWER(TRIM(%s))
                         AND COALESCE(NULLIF(b.guruh_kaliti,''),'whole')=%s
                         AND b.user_id<>%s""",
                    (
                        maktab_id, row["sinf_id"], row["fan_nomi"],
                        row["guruh_kaliti"], user_id,
                    ))
        occupied = cur.fetchone() or {}
        occupied_hours = float(occupied.get("band_soat") or 0)
        if approved_plan is not None:
            plan_item = approved_plan.get((row["sinf_id"], row["fan_kaliti"]))
            plan_hours = float((plan_item or {}).get("haftalik_soat") or 0)
            remaining = round(max(0.0, plan_hours - occupied_hours), 1)
            if row["haftalik_soat"] > remaining:
                owners = occupied.get("oqituvchilar") or "yuqoridagi qatorlar"
                raise HTTPException(
                    status_code=409,
                    detail=(
                        f"{row['variant']['sinf']} / {row['fan_nomi']} / "
                        f"{row['variant']['guruh_nomi']}: reja {plan_hours:g} soat, "
                        f"{owners}da {occupied_hours} soat tanlangan, faqat {remaining:g} soat qoldi"
                    ),
                )
            if row["haftalik_soat"] < remaining:
                plan_overrides.append(
                    f"{row['variant']['sinf']} / {row['fan_nomi']}: "
                    f"yana {remaining - row['haftalik_soat']:g} soat taqsimlanmagan"
                )
    if duplicate_merged:
        plan_overrides.append(
            f"Bir xil fan–sinf–guruhdagi {duplicate_merged} ta takror qator bitta qatorga qo'shildi"
        )

    cur.execute("DELETE FROM maktab_dars_birikmalari WHERE maktab_id=%s AND user_id=%s",
                (maktab_id, user_id))
    for row in cleaned:
        cur.execute("""INSERT INTO maktab_dars_birikmalari(
                        maktab_id,user_id,sinf_id,fan_nomi,guruh_kaliti,
                        haftalik_soat,kunlik_max,xona_id,manba,yangilangan_at)
                       VALUES(%s,%s,%s,%s,%s,%s,%s,%s,'v19.2_qolda',NOW())""",
                    (
                        maktab_id, user_id, row["sinf_id"], row["fan_nomi"],
                        row["guruh_kaliti"], row["haftalik_soat"],
                        row["kunlik_max"], row["xona_id"],
                    ))
        cur.execute("""INSERT INTO maktab_fanlari(maktab_id,fan_nomi)
                       VALUES(%s,%s) ON CONFLICT DO NOTHING""",
                    (maktab_id, row["fan_nomi"]))
        system_id = row["variant"].get("tizim_id")
        if system_id:
            cur.execute("SELECT fanlar FROM maktab_sinf_guruh_tizimlari WHERE id=%s FOR UPDATE", (system_id,))
            system = cur.fetchone()
            current_subjects = list((system or {}).get("fanlar") or [])
            if all(_v1875_subject_key(value) != _v1875_subject_key(row["fan_nomi"]) for value in current_subjects):
                current_subjects.append(row["fan_nomi"])
                cur.execute("""UPDATE maktab_sinf_guruh_tizimlari
                               SET fanlar=%s,yangilangan_at=NOW() WHERE id=%s""",
                            (current_subjects, system_id))

    weekly_total = _v195_refresh_teacher_summary(cur, maktab_id, user_id)
    if not finalize:
        return {
            "holat": "saqlandi",
            "user_id": int(user_id),
            "oqituvchi": teacher["full_name"],
            "qator_soni": len(cleaned),
            "haftalik_jami": weekly_total,
            "ogohlantirishlar": plan_overrides,
        }
    warnings = _v192_sync_schedule_sources(cur, maktab_id)
    warnings.extend(plan_overrides)
    auto_confirmation = _v192_auto_confirm_exact_pairs(cur, maktab_id, actor_id)
    payload = _v192_matrix_payload(cur, maktab_id)
    return {
        "holat": "saqlandi",
        "user_id": int(user_id),
        "oqituvchi": teacher["full_name"],
        "qator_soni": len(cleaned),
        "haftalik_jami": weekly_total,
        "ogohlantirishlar": warnings,
        "guruh_tasdiqlari": auto_confirmation,
        "matritsa": payload,
    }


@app.put("/api/maktab/aqlli_jadval/v3/oqituvchi_yuklamasi")
def v192_teacher_load_save(sorov: V192TeacherLoadSave, token: str):
    from datetime import date as _date

    actor_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor()
    try:
        if not _v1852_manager(cur, actor_id, sorov.maktab_id):
            raise HTTPException(status_code=403, detail="O'qituvchi yuklamasini faqat rahbariyat boshqaradi")
        # Precheck SELECT locklari keyingi ALTER TABLE lockiga upgrade qilinmasin.
        conn.commit()
        _v192_tables(cur)
        # Og'ir schema ensure locklarini yuklama business transactioni
        # davomida ushlab turmaymiz; vakolat keyingi transactionda qayta tekshiriladi.
        conn.commit()
        if not _v1852_manager(cur, actor_id, sorov.maktab_id):
            raise HTTPException(status_code=403, detail="O'qituvchi yuklamasini faqat rahbariyat boshqaradi")
        # Import va manual create bilan bir canonical maktab locki barcha
        # profil/yuklama/rahbarlik mutatsiyalarini ketma-ketlashtiradi.
        cur.execute(
            "SELECT pg_advisory_xact_lock(%s)",
            (1922000000 + int(sorov.maktab_id),),
        )
        cur.execute(
            "SELECT pg_advisory_xact_lock(%s)",
            (1925000000 + int(sorov.maktab_id),),
        )
        supplied_fields = set(
            getattr(sorov, "model_fields_set", getattr(sorov, "__fields_set__", set()))
        )
        if not sorov.qatorlar:
            raise HTTPException(
                status_code=400,
                detail="O'qituvchida kamida bitta aniq fan–sinf–guruh yuklamasi bo'lishi kerak",
            )
        if "full_name" in supplied_fields:
            full_name = re.sub(r"\s+", " ", str(sorov.full_name or "")).strip()
            if len(full_name) < 3 or len(full_name) > 160:
                raise HTTPException(
                    status_code=400,
                    detail="O'qituvchi F.I.Sh. 3–160 ta belgi bo'lishi kerak",
                )
            # Yangi o'qituvchi qo'shish va mavjudini qayta nomlash bir xil
            # maktab lockidan foydalanadi. Shu tariqa bir xil F.I.Sh. parallel
            # so'rovlarda ham yashirin dublikat bo'lib qolmaydi.
            cur.execute(
                """SELECT user_id,full_name FROM users
                   WHERE maktab_id=%s AND user_id<>%s""",
                (sorov.maktab_id, sorov.user_id),
            )
            candidate_name_key = _v237_teacher_name_key(full_name)
            if any(
                _v237_teacher_name_key(row.get("full_name")) == candidate_name_key
                for row in cur.fetchall()
            ):
                raise HTTPException(
                    status_code=409,
                    detail="Bu F.I.Sh. bilan boshqa xodim allaqachon mavjud",
                )
            cur.execute(
                """UPDATE users SET full_name=%s
                   WHERE user_id=%s AND maktab_id=%s""",
                (full_name, sorov.user_id, sorov.maktab_id),
            )
            if int(cur.rowcount or 0) != 1:
                raise HTTPException(status_code=404, detail="O'qituvchi topilmadi")
        # Partial API klienti bitta profil maydonini yuborsa, qolgan eski
        # qiymatlar NULL bilan ustidan yozilmaydi. Frontend barcha maydonlarni
        # (tozalash uchun explicit null bilan) yuboradi; har biri alohida
        # tekshiriladi va faqat yuborilgan ustun yangilanadi.
        profile_updates = []
        profile_values = []
        if "mutaxassisligi" in supplied_fields or "otadigan_fanlari" in supplied_fields:
            specialty_source = sorov.mutaxassisligi
            if not specialty_source and sorov.otadigan_fanlari:
                specialty_source = ";".join(
                    _v1859_fanlarni_ajrat(";".join(sorov.otadigan_fanlari))
                )
            specialty, _ = _v194_teacher_profile_values(
                specialty_source, None
            )
            profile_updates.append("mutaxassisligi=%s")
            profile_values.append(specialty)
        if "haftalik_maqsad_soat" in supplied_fields:
            _, target = _v194_teacher_profile_values(
                None, sorov.haftalik_maqsad_soat
            )
            profile_updates.append("haftalik_maqsad_soat=%s")
            profile_values.append(target)
        if "ish_staji" in supplied_fields:
            work_years = sorov.ish_staji
            if work_years is not None and not 0 <= int(work_years) <= 60:
                raise HTTPException(status_code=400, detail="Ish staji 0–60 yil oralig'ida bo'lishi kerak")
            profile_updates.append("ish_staji=%s")
            profile_values.append(int(work_years) if work_years is not None else None)
        if "tugilgan_sana" in supplied_fields or "tugilgan_yili" in supplied_fields:
            birth_date = sorov.tugilgan_sana if "tugilgan_sana" in supplied_fields else None
            birth_year = (
                birth_date.year if birth_date is not None
                else sorov.tugilgan_yili
            )
            current_year = _date.today().year
            if birth_year is not None and not 1900 <= int(birth_year) <= current_year:
                raise HTTPException(
                    status_code=400,
                    detail=f"Tug'ilgan yil 1900–{current_year} oralig'ida bo'lishi kerak",
                )
            if birth_date is not None and birth_date > _date.today():
                raise HTTPException(status_code=400, detail="Tug'ilgan sana kelajakda bo'lishi mumkin emas")
            profile_updates.append("tugilgan_sana=%s")
            profile_values.append(
                birth_date or (
                    _date(int(birth_year), 1, 1)
                    if birth_year is not None else None
                )
            )
        if "toifasi" in supplied_fields:
            category = re.sub(r"\s+", " ", str(sorov.toifasi or "")).strip() or None
            if category and category not in TOIFALAR:
                raise HTTPException(status_code=400, detail="O'qituvchi toifasi noto'g'ri")
            profile_updates.append("toifasi=%s")
            profile_values.append(category)
        if profile_updates:
            cur.execute(
                f"""UPDATE users SET {', '.join(profile_updates)}
                    WHERE user_id=%s AND maktab_id=%s""",
                (*profile_values, sorov.user_id, sorov.maktab_id),
            )
            if int(cur.rowcount or 0) != 1:
                raise HTTPException(status_code=404, detail="O'qituvchi topilmadi")
        leader_class = None
        if "rahbar_sinf_id" in supplied_fields:
            leader_class = _v199_save_teacher_leadership(
                cur, sorov.maktab_id, sorov.user_id, sorov.rahbar_sinf_id, actor_id
            )
        result = _v192_save_teacher_load_rows(
            cur, actor_id, sorov.maktab_id, sorov.user_id, sorov.qatorlar
        )
        _v199_teacher_total_with_class_hour(
            cur, sorov.maktab_id, sorov.user_id, result
        )
        result.update({
            "rahbar_sinf_id": int(leader_class["id"]) if leader_class else None,
            "rahbar_sinf_nomi": (
                f"{leader_class['sinf']}-{leader_class['harf']}" if leader_class else None
            ),
        })
        conn.commit()
        return result
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


@app.post("/api/maktab/aqlli_jadval/v3/oqituvchi_qoshish")
def v192_manual_teacher_create(sorov: V192ManualTeacherCreate, token: str):
    from datetime import date as _date

    actor_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor()
    try:
        if not _v1852_manager(cur, actor_id, sorov.maktab_id):
            raise HTTPException(status_code=403, detail="O'qituvchini faqat maktab rahbariyati qo'shadi")
        # Precheck SELECT locklari keyingi ALTER TABLE lockiga upgrade qilinmasin.
        conn.commit()
        _v192_tables(cur)
        _xodim_kod_jadvali(cur)
        # Og'ir schema ensure locklari xodim yaratish/raqamlash transactioniga
        # o'tmasin; vakolat va locklar commitdan keyin qayta olinadi.
        conn.commit()
        if not _v1852_manager(cur, actor_id, sorov.maktab_id):
            raise HTTPException(status_code=403, detail="O'qituvchini faqat maktab rahbariyati qo'shadi")
        full_name = re.sub(r"\s+", " ", str(sorov.full_name or "")).strip()
        if len(full_name) < 3 or len(full_name) > 160:
            raise HTTPException(status_code=400, detail="O'qituvchi F.I.Sh. 3–160 ta belgi bo'lishi kerak")
        work_years = sorov.ish_staji
        if work_years is not None and not 0 <= int(work_years) <= 60:
            raise HTTPException(status_code=400, detail="Ish staji 0–60 yil oralig'ida bo'lishi kerak")
        birth_date = sorov.tugilgan_sana
        birth_year = birth_date.year if birth_date is not None else sorov.tugilgan_yili
        current_year = _date.today().year
        if birth_year is not None and not 1900 <= int(birth_year) <= current_year:
            raise HTTPException(
                status_code=400,
                detail=f"Tug'ilgan yil 1900–{current_year} oralig'ida bo'lishi kerak",
            )
        if birth_date is not None and birth_date > _date.today():
            raise HTTPException(status_code=400, detail="Tug'ilgan sana kelajakda bo'lishi mumkin emas")
        category = re.sub(r"\s+", " ", str(sorov.toifasi or "")).strip() or None
        if category and category not in TOIFALAR:
            raise HTTPException(status_code=400, detail="O'qituvchi toifasi noto'g'ri")
        specialty_source = sorov.mutaxassisligi
        if not specialty_source and sorov.otadigan_fanlari:
            specialty_source = ";".join(
                _v1859_fanlarni_ajrat(";".join(sorov.otadigan_fanlari))
            )
        specialty, target_hours = _v194_teacher_profile_values(
            specialty_source, sorov.haftalik_maqsad_soat
        )
        if not sorov.qatorlar:
            raise HTTPException(
                status_code=400,
                detail="Yangi o'qituvchida kamida bitta aniq fan–sinf–guruh yuklamasi bo'lishi kerak",
            )

        # ``users.user_id`` manfiy identifikatori barcha maktablar uchun bitta
        # global ketma-ketlikdan olinadi. Import endpointi bilan parallel
        # qo'shishda ham bir xil ID chiqmasligi uchun lock tartibi doimo
        # GLOBAL -> maktab -> yuklama bo'ladi.
        cur.execute("SELECT pg_advisory_xact_lock(%s)", (2370000001,))
        cur.execute("SELECT pg_advisory_xact_lock(%s)", (1922000000 + int(sorov.maktab_id),))
        cur.execute("SELECT pg_advisory_xact_lock(%s)", (1925000000 + int(sorov.maktab_id),))
        leader_class = None
        if sorov.rahbar_sinf_id is not None:
            cur.execute("""SELECT id,sinf,harf,rahbar_user_id
                           FROM maktab_sinflari
                           WHERE id=%s AND maktab_id=%s
                           FOR UPDATE""",
                        (int(sorov.rahbar_sinf_id), int(sorov.maktab_id)))
            leader_class = cur.fetchone()
            if not leader_class:
                raise HTTPException(status_code=404, detail="Sinf rahbarligi uchun tanlangan sinf topilmadi")
            if leader_class.get("rahbar_user_id") is not None:
                raise HTTPException(
                    status_code=409,
                    detail=f"{leader_class['sinf']}-{leader_class['harf']} sinfiga rahbar allaqachon tayinlangan",
                )
        cur.execute(
            "SELECT user_id,full_name FROM users WHERE maktab_id=%s",
            (sorov.maktab_id,),
        )
        candidate_name_key = _v237_teacher_name_key(full_name)
        if any(
            _v237_teacher_name_key(row.get("full_name")) == candidate_name_key
            for row in cur.fetchall()
        ):
            raise HTTPException(status_code=409, detail="Bu F.I.Sh. bilan xodim allaqachon mavjud")
        cur.execute("SELECT MIN(user_id) AS eng_kichik FROM users WHERE user_id < 0")
        smallest = cur.fetchone()
        new_user_id = (
            int(smallest["eng_kichik"]) - 1
            if smallest and smallest.get("eng_kichik") is not None else -1
        )
        cur.execute("""INSERT INTO users(
                        user_id,full_name,role,maktab_id,lavozim,tugilgan_sana,
                        ish_staji,toifasi,mutaxassisligi,haftalik_maqsad_soat,
                        haftalik_dars_soati)
                       VALUES(%s,%s,'oqituvchi',%s,'fan_oqituvchisi',%s,%s,%s,%s,%s,0)""",
                    (
                        new_user_id, full_name, sorov.maktab_id,
                        birth_date or (_date(int(birth_year), 1, 1) if birth_year is not None else None),
                        int(work_years) if work_years is not None else None,
                        category, specialty, target_hours,
                    ))
        teacher_numbers = _v2249_ensure_teacher_numbers(cur, sorov.maktab_id)
        new_teacher_number = teacher_numbers.get(int(new_user_id))
        plain_code, stored_code = _xodim_kod_yarat()
        cur.execute("INSERT INTO xodim_kod(kod,user_id) VALUES(%s,%s)",
                    (stored_code, new_user_id))
        if leader_class:
            cur.execute("""UPDATE maktab_sinflari
                           SET rahbar_user_id=%s
                           WHERE id=%s AND maktab_id=%s""",
                        (new_user_id, int(leader_class["id"]), int(sorov.maktab_id)))
            _v199_ensure_class_hour_rules(
                cur, sorov.maktab_id, [int(leader_class["id"])], actor_id
            )
        result = _v192_save_teacher_load_rows(
            cur, actor_id, sorov.maktab_id, new_user_id, sorov.qatorlar
        )
        _v199_teacher_total_with_class_hour(
            cur, sorov.maktab_id, new_user_id, result
        )
        result.update({
            "holat": "oqituvchi_va_yuklama_saqlandi",
            "kirish_kodi": plain_code,
            "kirish_kodi_muddati": "2 oy",
            "qolda_kiritildi": True,
            "jadval_raqami": new_teacher_number,
            "mutaxassisligi": specialty,
            "haftalik_maqsad_soat": target_hours,
            "tugilgan_yili": int(birth_year) if birth_year is not None else None,
            "tugilgan_sana": birth_date.isoformat() if birth_date is not None else None,
            "rahbar_sinf_id": int(leader_class["id"]) if leader_class else None,
            "rahbar_sinf_nomi": (
                f"{leader_class['sinf']}-{leader_class['harf']}" if leader_class else None
            ),
        })
        conn.commit()
        return result
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


class V237TeacherImportPayload(BaseModel):
    maktab_id: int
    fayl_nomi: Optional[str] = "Oqituvchilar_import.xlsx"
    xlsx_base64: str


_V237_TEACHER_XLSX_MAX_BYTES = 3 * 1024 * 1024
_V237_TEACHER_XLSX_MAX_UNCOMPRESSED = 15 * 1024 * 1024
_V237_TEACHER_XLSX_MAX_ROWS = 5000
_V237_TEACHER_XLSX_MAX_BASE64_CHARS = (
    4 * ((_V237_TEACHER_XLSX_MAX_BYTES + 2) // 3)
)
_V237_TEACHER_IMPORT_BODY_MAX_BYTES = (
    _V237_TEACHER_XLSX_MAX_BASE64_CHARS + 1024 * 1024
)
_V237_TEACHER_IMPORT_MAX_REQUEST_FRAMES = 8192
_V237_TEACHER_IMPORT_PATHS = frozenset({
    "/api/maktab/aqlli_jadval/v3/oqituvchi_import_preview",
    "/api/maktab/aqlli_jadval/v3/oqituvchi_import_commit",
})


async def _v237_send_teacher_import_body_too_large(send):
    body = (
        '{"detail":"Excel import so‘rovi juda katta; '
        'XLSX fayl 3 MB dan oshmasligi kerak."}'
    ).encode("utf-8")
    await send({
        "type": "http.response.start",
        "status": 413,
        "headers": [
            (b"content-type", b"application/json; charset=utf-8"),
            (b"content-length", str(len(body)).encode("ascii")),
        ],
    })
    await send({"type": "http.response.body", "body": body, "more_body": False})


class _V237TeacherImportBodyLimitMiddleware:
    """Limit only teacher-import JSON before FastAPI buffers or parses it."""

    def __init__(self, app):
        self.app = app

    async def __call__(self, scope, receive, send):
        request_path = str(scope.get("path") or "")
        root_prefix = str(scope.get("root_path") or "").rstrip("/")
        if root_prefix and (
            request_path == root_prefix
            or request_path.startswith(root_prefix + "/")
        ):
            request_path = request_path[len(root_prefix):] or "/"
        if (
            scope.get("type") != "http"
            or str(scope.get("method") or "").upper() != "POST"
            or request_path not in _V237_TEACHER_IMPORT_PATHS
        ):
            return await self.app(scope, receive, send)

        declared_sizes = []
        for name, value in scope.get("headers", ()):
            if name.lower() != b"content-length":
                continue
            try:
                declared_size = int(value)
            except (TypeError, ValueError):
                continue
            if declared_size >= 0:
                declared_sizes.append(declared_size)
        declared_too_large = (
            bool(declared_sizes)
            and max(declared_sizes) > _V237_TEACHER_IMPORT_BODY_MAX_BYTES
        )
        if declared_too_large:
            return await _v237_send_teacher_import_body_too_large(send)

        buffered_body = bytearray()
        request_frames = 0
        terminal_message = None
        while True:
            message = await receive()
            if message.get("type") == "http.request":
                request_frames += 1
                if request_frames > _V237_TEACHER_IMPORT_MAX_REQUEST_FRAMES:
                    buffered_body.clear()
                    message = None
                    return await _v237_send_teacher_import_body_too_large(send)
                chunk = message.get("body") or b""
                if len(chunk) > (
                    _V237_TEACHER_IMPORT_BODY_MAX_BYTES - len(buffered_body)
                ):
                    buffered_body.clear()
                    message = None
                    return await _v237_send_teacher_import_body_too_large(send)
                buffered_body.extend(chunk)
                if message.get("more_body", False):
                    continue
            else:
                terminal_message = message
            break

        combined_request = {
            "type": "http.request",
            "body": bytes(buffered_body),
            "more_body": terminal_message is not None,
        }
        buffered_body.clear()
        message = None
        combined_pending = True
        terminal_pending = terminal_message is not None

        async def replay_receive():
            nonlocal combined_pending, terminal_pending
            if combined_pending:
                combined_pending = False
                return combined_request
            if terminal_pending:
                terminal_pending = False
                return terminal_message
            return await receive()

        return await self.app(scope, replay_receive, send)


if not getattr(app.state, "samtm_v237_teacher_import_body_limit", False):
    app.add_middleware(_V237TeacherImportBodyLimitMiddleware)
    app.state.samtm_v237_teacher_import_body_limit = True


def _v237_teacher_name_clean(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _v237_teacher_name_key(value):
    name = _v237_unicodedata.normalize(
        "NFKC", _v237_teacher_name_clean(value).translate(_V237_APOSTROPHES)
    )
    return re.sub(r"\s+", " ", name).strip().casefold()


def _v237_excel_header_key(value):
    raw = _v237_unicodedata.normalize(
        "NFKC", str(value or "").translate(_V237_APOSTROPHES)
    ).casefold().strip()
    if raw == "№":
        return "number"
    compact = re.sub(r"[^\w]+", "", raw, flags=re.UNICODE)
    if compact in {"n", "no", "nomer", "raqam", "tr", "tartibraqami"}:
        return "number"
    if compact in {
        "fish", "fio", "fullname", "oqituvchi", "oqituvchifish",
        "oqituvchiningfish", "ismfamiliya", "familiyaismsharif",
    }:
        return "full_name"
    if compact in {
        "skeletsoati", "skeletsoat", "haftaliksoat", "haftalikdarssoati",
        "haftalikmaqsadsoat", "darssoati", "soat",
    }:
        return "skeleton_hours"
    return compact


def _v237_decode_teacher_xlsx(payload: V237TeacherImportPayload):
    filename = _v237_teacher_name_clean(payload.fayl_nomi or "Oqituvchilar_import.xlsx")
    if filename and not filename.casefold().endswith(".xlsx"):
        raise ValueError("Faqat .xlsx fayl qabul qilinadi.")
    original_encoded = str(payload.xlsx_base64 or "")
    if len(original_encoded) > _V237_TEACHER_IMPORT_BODY_MAX_BYTES:
        raise ValueError("XLSX fayl 3 MB dan oshmasligi kerak.")
    encoded = original_encoded.strip()
    if encoded.startswith("data:"):
        if "," not in encoded:
            raise ValueError("XLSX base64 data URL noto‘g‘ri.")
        encoded = encoded.split(",", 1)[1]
    encoded = re.sub(r"\s+", "", encoded)
    if not encoded:
        raise ValueError("XLSX fayl yuborilmadi.")
    # Base64 taxminan 4/3 kattalashadi. Decode qilishdan oldin ham qattiq limit.
    if len(encoded) > _V237_TEACHER_XLSX_MAX_BASE64_CHARS + 16:
        raise ValueError("XLSX fayl 3 MB dan oshmasligi kerak.")
    try:
        raw = _v237_base64.b64decode(encoded, validate=True)
    except (_v237_binascii.Error, ValueError) as exc:
        raise ValueError("XLSX base64 ma’lumoti buzilgan.") from exc
    if not raw or len(raw) > _V237_TEACHER_XLSX_MAX_BYTES:
        raise ValueError("XLSX fayl bo‘sh yoki 3 MB dan katta.")
    try:
        with _v237_zipfile.ZipFile(_v237_io.BytesIO(raw)) as archive:
            members = archive.infolist()
            if len(members) > 200:
                raise ValueError("XLSX ichida ortiqcha fayllar mavjud.")
            unpacked = 0
            for member in members:
                normalized_name = str(member.filename or "").replace("\\", "/")
                if normalized_name.startswith("/") or ".." in normalized_name.split("/"):
                    raise ValueError("XLSX ichki fayl yo‘li xavfsiz emas.")
                unpacked += int(member.file_size or 0)
                if unpacked > _V237_TEACHER_XLSX_MAX_UNCOMPRESSED:
                    raise ValueError("XLSX ochilgandagi hajmi xavfsizlik limitidan oshdi.")
    except _v237_zipfile.BadZipFile as exc:
        raise ValueError("Fayl haqiqiy XLSX emas yoki buzilgan.") from exc
    return raw


def _v237_teacher_hour_value(value, row_number):
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, bool):
        raise ValueError(f"{row_number}-qator: Skelet soati raqam bo‘lishi kerak.")
    text = str(value).strip().replace(",", ".")
    try:
        hours = float(text)
    except (TypeError, ValueError) as exc:
        raise ValueError(
            f"{row_number}-qator: Skelet soati 0,5–60 oralig‘ida bo‘lishi kerak."
        ) from exc
    if (
        not _v237_math.isfinite(hours)
        or hours < 0.5
        or hours > 60
        or abs(hours * 2 - round(hours * 2)) > 1e-9
    ):
        raise ValueError(
            f"{row_number}-qator: Skelet soati 0,5–60 oralig‘ida, 0,5 qadamda bo‘lishi kerak."
        )
    return round(hours, 1)


def _v237_parse_teacher_xlsx(raw):
    try:
        import openpyxl
        workbook = openpyxl.load_workbook(
            _v237_io.BytesIO(raw), read_only=True, data_only=False, keep_links=False
        )
    except Exception as exc:
        raise ValueError("XLSX ochilmadi. Fayl buzilmaganini tekshiring.") from exc
    try:
        preferred = []
        fallback = []
        for worksheet in workbook.worksheets:
            sheet_key = _v237_excel_header_key(worksheet.title)
            if sheet_key in {"toldirishnamunasi", "namuna", "korsatma", "qollanma"}:
                continue
            if sheet_key in {"oqituvchilar", "oqituvchi"}:
                preferred.append(worksheet)
            else:
                fallback.append(worksheet)
        candidates = preferred + fallback
        selected = None
        header_row = None
        columns = None
        for worksheet in candidates:
            for row_number, cells in enumerate(
                worksheet.iter_rows(min_row=1, max_row=20, max_col=30), start=1
            ):
                keys = [_v237_excel_header_key(cell.value) for cell in cells]
                mapping = {}
                for index, key in enumerate(keys):
                    if key in {"number", "full_name", "skeleton_hours"} and key not in mapping:
                        mapping[key] = index
                if {"full_name", "skeleton_hours"}.issubset(mapping):
                    selected = worksheet
                    header_row = row_number
                    columns = mapping
                    break
            if selected is not None:
                break
        if selected is None:
            raise ValueError(
                "O‘qituvchilar varag‘ida “F.I.Sh.” va “Skelet soati” sarlavhalari topilmadi."
            )

        parsed = []
        errors = []
        seen = {}
        max_column = min(max(int(selected.max_column or 1), max(columns.values()) + 1), 30)
        scanned = 0
        for row_number, cells in enumerate(
            selected.iter_rows(min_row=header_row + 1, max_col=max_column),
            start=header_row + 1,
        ):
            scanned += 1
            if scanned > _V237_TEACHER_XLSX_MAX_ROWS:
                errors.append({
                    "qator": row_number,
                    "xato": f"Ko‘pi bilan {_V237_TEACHER_XLSX_MAX_ROWS} ta o‘qituvchi yuklash mumkin.",
                })
                break
            formula_cell = next((
                cell for cell in cells
                if cell.data_type == "f" or str(cell.value or "").lstrip().startswith("=")
            ), None)
            if formula_cell is not None:
                errors.append({
                    "qator": row_number,
                    "xato": "Formula qabul qilinmaydi; qiymatni oddiy matn/raqam qilib yozing.",
                })
                continue
            name_value = cells[columns["full_name"]].value
            hour_value = cells[columns["skeleton_hours"]].value
            number_value = (
                cells[columns["number"]].value if "number" in columns else None
            )
            if all(value is None or str(value).strip() == "" for value in (name_value, hour_value, number_value)):
                continue
            name = _v237_teacher_name_clean(name_value)
            if not 3 <= len(name) <= 160:
                errors.append({
                    "qator": row_number,
                    "xato": "O‘qituvchi F.I.Sh. 3–160 ta belgi bo‘lishi kerak.",
                })
                continue
            # F.I.Sh. boshqa ustunda xato bo'lsa ham fayldagi takror sifatida
            # hisoblanadi. Aks holda birinchi nusxaning soati xato bo'lganda
            # ikkinchi nusxa yashirincha yaroqli bo'lib qolardi.
            name_key = _v237_teacher_name_key(name)
            if name_key in seen:
                errors.append({
                    "qator": row_number,
                    "xato": f"{name} faylda takrorlangan (oldingi qator: {seen[name_key]}).",
                })
                continue
            seen[name_key] = row_number
            try:
                target_hours = _v237_teacher_hour_value(hour_value, row_number)
            except ValueError as exc:
                errors.append({"qator": row_number, "xato": str(exc)})
                continue
            parsed.append({
                "qator": row_number,
                "full_name": name,
                "skelet_soati": target_hours,
                "name_key": name_key,
            })
        if not parsed and not errors:
            errors.append({"qator": None, "xato": "Import uchun birorta o‘qituvchi topilmadi."})
        return parsed, errors, selected.title, int(header_row)
    finally:
        workbook.close()


def _v237_teacher_import_db_errors(cur, maktab_id, parsed):
    cur.execute(
        "SELECT user_id,full_name FROM users WHERE maktab_id=%s",
        (maktab_id,),
    )
    existing = {
        _v237_teacher_name_key(row.get("full_name")): dict(row)
        for row in cur.fetchall()
        if _v237_teacher_name_clean(row.get("full_name"))
    }
    errors = []
    for item in parsed:
        old = existing.get(item["name_key"])
        if old:
            errors.append({
                "qator": item["qator"],
                "xato": (
                    f"{item['full_name']} bu maktabda allaqachon mavjud "
                    f"(ID {old['user_id']})."
                ),
            })
    return errors


def _v237_teacher_template_bytes():
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    workbook = openpyxl.Workbook()
    main = workbook.active
    main.title = "O‘qituvchilar"
    main.append(["№", "F.I.Sh.", "Skelet soati"])
    # Asosiy varaq ataylab mutlaqo namunasiz: namuna qatori tasodifan import
    # bo'lib ketmasligi uchun to'rtta misol alohida varaqda turadi.
    example = workbook.create_sheet("To‘ldirish namunasi")
    example.append(["№", "F.I.Sh.", "Skelet soati"])
    examples = [
        (1, "Aliyeva Dilnoza Anvarovna", 18),
        (2, "Karimov Sardor Bahodirovich", 16.5),
        (3, "Rasulova Mohira Otabekovna", 20),
        (4, "Abdullayev Jasur Akmalovich", 12.5),
    ]
    for row in examples:
        example.append(row)
    guide = workbook.create_sheet("Ko‘rsatma")
    guide.append(["O‘QITUVCHILARNI EXCEL ORQALI YUKLASH"])
    guide.append(["1", "Faqat “O‘qituvchilar” varag‘ini to‘ldiring."])
    guide.append(["2", "F.I.Sh. 3–160 ta belgi bo‘lsin; bir xil F.I.Sh. takrorlanmasin."])
    guide.append(["3", "Skelet soati bo‘sh qolishi yoki 0,5–60 oralig‘ida 0,5 qadamda yozilishi mumkin."])
    guide.append(["4", "Formula yozmang. O‘qituvchi raqami sayt tomonidan o‘zgarmas qilib beriladi."])
    guide.append(["5", "“To‘ldirish namunasi” faqat ko‘rish uchun; u import qilinmaydi."])
    guide.merge_cells("A1:C1")
    guide["A1"].font = Font(size=15, bold=True, color="FFFFFF")
    guide["A1"].fill = PatternFill("solid", fgColor="155A7A")

    for worksheet in (main, example):
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="0F7C82")
            cell.alignment = Alignment(horizontal="center")
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = f"A1:C{max(2, worksheet.max_row)}"
        worksheet.column_dimensions["A"].width = 9
        worksheet.column_dimensions["B"].width = 42
        worksheet.column_dimensions["C"].width = 19
        validation = DataValidation(
            type="custom",
            formula1='OR(C2="",AND(ISNUMBER(C2),C2>=0.5,C2<=60,MOD(C2*2,1)=0))',
            allow_blank=True,
        )
        validation.error = "0,5–60 oralig‘ida, 0,5 qadamda kiriting"
        validation.errorTitle = "Skelet soati noto‘g‘ri"
        validation.prompt = "Masalan: 18 yoki 16,5"
        validation.promptTitle = "Skelet soati"
        validation.showErrorMessage = True
        validation.showInputMessage = True
        worksheet.add_data_validation(validation)
        validation.add("C2:C5001")
    guide.column_dimensions["A"].width = 10
    guide.column_dimensions["B"].width = 95
    output = _v237_io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _v237_require_teacher_import_manager(actor_id, maktab_id, action):
    """XLSXni ochishdan oldin arzon va yozuvsiz ruxsat tekshiruvi."""
    conn = _db(); cur = conn.cursor()
    try:
        if not _v1852_manager(cur, actor_id, maktab_id):
            raise HTTPException(
                status_code=403,
                detail=f"O‘qituvchi importini faqat maktab rahbariyati {action}",
            )
    finally:
        cur.close(); conn.close()


@app.get("/api/maktab/aqlli_jadval/v3/oqituvchi_import_shablon")
def v237_teacher_import_template(token: str, maktab_id: int):
    from fastapi.responses import StreamingResponse
    from urllib.parse import quote

    actor_id = _jwt_tekshir(token)
    _v237_require_teacher_import_manager(actor_id, maktab_id, "yuklaydi")
    raw = _v237_teacher_template_bytes()
    filename = "SAMTM_Oqituvchilar_import_shabloni.xlsx"
    return StreamingResponse(
        _v237_io.BytesIO(raw),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}",
            "X-SAMTM-Template-Examples": "4",
        },
    )


@app.post("/api/maktab/aqlli_jadval/v3/oqituvchi_import_preview")
def v237_teacher_import_preview(sorov: V237TeacherImportPayload, token: str):
    actor_id = _jwt_tekshir(token)
    _v237_require_teacher_import_manager(actor_id, sorov.maktab_id, "tekshiradi")
    try:
        raw = _v237_decode_teacher_xlsx(sorov)
        parsed, parse_errors, sheet_name, header_row = _v237_parse_teacher_xlsx(raw)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    conn = _db(); cur = conn.cursor()
    try:
        if not _v1852_manager(cur, actor_id, sorov.maktab_id):
            raise HTTPException(status_code=403, detail="O‘qituvchi importini faqat maktab rahbariyati tekshiradi")
        db_errors = _v237_teacher_import_db_errors(cur, sorov.maktab_id, parsed)
        errors = [dict(item) for item in (parse_errors + db_errors)]
        for error in errors:
            error["excel_qatori"] = error.get("qator")
        error_rows = {
            int(item["qator"])
            for item in errors
            if item.get("qator") is not None
        }
        public_rows = [
            {
                "excel_qatori": int(row["qator"]),
                "qator": int(row["qator"]),
                "full_name": row["full_name"],
                "haftalik_maqsad_soat": row["skelet_soati"],
                "skelet_soati": row["skelet_soati"],
            }
            for row in parsed
            if int(row["qator"]) not in error_rows
        ]
        all_rows = {
            int(row["qator"]) for row in parsed if row.get("qator") is not None
        } | error_rows
        return {
            "holat": "xato" if errors else "tayyor",
            "valid": not errors,
            "varaq": sheet_name,
            "sarlavha_qatori": header_row,
            "qatorlar": public_rows,
            "jami": len(all_rows),
            "yaroqli": len(public_rows),
            "yangi_soni": len(public_rows),
            "xatolar": len(errors),
            "xato_qatorlar": errors,
        }
    finally:
        cur.close(); conn.close()


@app.post("/api/maktab/aqlli_jadval/v3/oqituvchi_import_commit")
def v237_teacher_import_commit(sorov: V237TeacherImportPayload, token: str):
    actor_id = _jwt_tekshir(token)
    _v237_require_teacher_import_manager(actor_id, sorov.maktab_id, "saqlaydi")
    try:
        raw = _v237_decode_teacher_xlsx(sorov)
        parsed, parse_errors, _sheet_name, _header_row = _v237_parse_teacher_xlsx(raw)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    if parse_errors:
        raise HTTPException(
            status_code=400,
            detail={
                "message": "XLSX qatorlarida xato bor; hech narsa yozilmadi",
                "xabar": "XLSX qatorlarida xato bor; hech narsa yozilmadi",
                "xato_qatorlar": parse_errors,
            },
        )
    conn = _db(); cur = conn.cursor()
    try:
        _v192_tables(cur)
        _xodim_kod_jadvali(cur)
        # `_v192_tables` va `_xodim_kod_jadvali` og'ir DDL locklarini import,
        # raqamlash va matritsa davomida ushlab turmaymiz. Keyingi SQL alohida
        # business transaction boshlaydi.
        conn.commit()
        if not _v1852_manager(cur, actor_id, sorov.maktab_id):
            raise HTTPException(status_code=403, detail="O‘qituvchi importini faqat maktab rahbariyati saqlaydi")
        # Negative users.user_id butun users jadvali bo'yicha global. Avval global,
        # keyin barcha teacher create/rename yo'llari ishlatadigan maktab locki
        # olinadi: turli maktab parallel importida ham PK urilmaydi.
        cur.execute("SELECT pg_advisory_xact_lock(%s)", (2370000001,))
        cur.execute(
            "SELECT pg_advisory_xact_lock(%s)",
            (1922000000 + int(sorov.maktab_id),),
        )
        db_errors = _v237_teacher_import_db_errors(cur, sorov.maktab_id, parsed)
        if db_errors:
            raise HTTPException(
                status_code=409,
                detail={
                    "message": "Import qayta yuborilgan yoki F.I.Sh. allaqachon mavjud; hech narsa yozilmadi",
                    "xabar": "Import qayta yuborilgan yoki F.I.Sh. allaqachon mavjud; hech narsa yozilmadi",
                    "xatolar": db_errors,
                },
            )
        cur.execute("SELECT MIN(user_id) AS eng_kichik FROM users WHERE user_id < 0")
        smallest = cur.fetchone() or {}
        next_user_id = (
            int(smallest["eng_kichik"]) - 1
            if smallest.get("eng_kichik") is not None else -1
        )
        created = []
        for item in parsed:
            user_id = next_user_id
            next_user_id -= 1
            cur.execute(
                """INSERT INTO users(
                       user_id,full_name,role,maktab_id,lavozim,
                       haftalik_maqsad_soat,haftalik_dars_soati)
                   VALUES(%s,%s,'oqituvchi',%s,'fan_oqituvchisi',%s,0)""",
                (
                    user_id,
                    item["full_name"],
                    sorov.maktab_id,
                    item["skelet_soati"],
                ),
            )
            plain_code, stored_code = _xodim_kod_yarat()
            cur.execute(
                "INSERT INTO xodim_kod(kod,user_id) VALUES(%s,%s)",
                (stored_code, user_id),
            )
            created.append({
                "user_id": user_id,
                "full_name": item["full_name"],
                "skelet_soati": item["skelet_soati"],
                "haftalik_maqsad_soat": item["skelet_soati"],
                "kirish_kodi": plain_code,
            })
        teacher_numbers = _v2249_ensure_teacher_numbers(cur, sorov.maktab_id)
        for row in created:
            row["jadval_raqami"] = teacher_numbers.get(int(row["user_id"]))
        matrix = _v192_matrix_payload(cur, sorov.maktab_id)
        conn.commit()
        return {
            "holat": "oqituvchilar_import_qilindi",
            "yaratildi": len(created),
            "kodlar": created,
            "kirish_kodlari": created,
            "matritsa": matrix,
        }
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


@app.put("/api/maktab/aqlli_jadval/v3/sinf_skeleti_yuklama_qisman")
def v204_class_skeleton_patch_save(sorov: V204SkeletonPatchSave, token: str):
    """Save only changed Step-3 cells without replacing another teacher's rows."""
    actor_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor()
    try:
        if not _v1852_manager(cur, actor_id, sorov.maktab_id):
            raise HTTPException(
                status_code=403,
                detail="Sinf skeletini faqat maktab rahbariyati saqlaydi",
            )
        # Schema ensure alohida transactionda tugaydi; business transaction
        # faqat bitta maktabning canonical yuklamasini bloklaydi.
        conn.commit()
        _v192_tables(cur)
        conn.commit()
        if not _v1852_manager(cur, actor_id, sorov.maktab_id):
            raise HTTPException(
                status_code=403,
                detail="Sinf skeletini faqat maktab rahbariyati saqlaydi",
            )
        cur.execute(
            "SELECT pg_advisory_xact_lock(%s)",
            (1922000000 + int(sorov.maktab_id),),
        )
        cur.execute(
            "SELECT pg_advisory_xact_lock(%s)",
            (2040000000 + int(sorov.maktab_id),),
        )
        cur.execute(
            "SELECT pg_advisory_xact_lock(%s)",
            (1925000000 + int(sorov.maktab_id),),
        )

        if not (sorov.qatorlar or sorov.ochirilgan_kalitlar or sorov.rahbarlar):
            raise HTTPException(
                status_code=400,
                detail="Saqlash uchun kamida bitta o'zgargan katak yuborilishi kerak",
            )

        # Matrixda ko'rinadigan haqiqiy/raqamlangan o'qituvchilar aynan shu
        # allowlistdan olinadi. Admin, o'quvchi yoki boshqa maktab xodimini
        # API orqali yashirincha darsga/rahbarlikka biriktirib bo'lmaydi.
        school_users = {
            int(row["user_id"]): dict(row)
            for row in _v1859_effective_teachers(
                cur, sorov.maktab_id, include_numbered=True
            )
        }
        cur.execute(
            """SELECT id,sinf,harf,rahbar_user_id FROM maktab_sinflari
               WHERE maktab_id=%s ORDER BY id FOR UPDATE""",
            (sorov.maktab_id,),
        )
        school_classes = {int(row["id"]): dict(row) for row in cur.fetchall()}
        current_rows = _v192_assignment_rows(cur, sorov.maktab_id)
        current_revision = _v204_assignment_revision(
            current_rows, school_classes.values()
        )
        expected_revision = str(
            sorov.kutilgan_yuklama_revision or ""
        ).strip().lower()
        if not expected_revision or expected_revision != current_revision:
            raise HTTPException(
                status_code=409,
                detail=(
                    "O'qituvchi yuklamasi boshqa oynada yangilangan. "
                    "Yangi ma'lumot yo'qolmasligi uchun sahifani qayta yuklang"
                ),
            )

        expected = _v204_expected_skeleton_rows(cur, sorov.maktab_id)
        cur.execute(
            """SELECT id FROM aqlli_xonalar_v2
               WHERE maktab_id=%s AND faol=TRUE AND darsga_yaroqli=TRUE""",
            (sorov.maktab_id,),
        )
        allowed_room_ids = {int(row["id"]) for row in cur.fetchall()}

        # targeted[key] is a cleaned replacement row, or None for an explicit
        # clear. Duplicate/overlapping keys are rejected instead of silently
        # choosing the last browser value.
        targeted = {}
        for index, row in enumerate(sorov.qatorlar, start=1):
            class_id = int(row.sinf_id)
            teacher_id = int(row.user_id)
            if class_id not in school_classes:
                raise HTTPException(
                    status_code=400,
                    detail=f"{index}-qator: sinf bu maktabga tegishli emas",
                )
            if teacher_id not in school_users:
                raise HTTPException(
                    status_code=400,
                    detail=f"{index}-qator: o'qituvchi bu maktabga tegishli emas",
                )
            key = _v204_resolve_expected_skeleton_key(
                expected, class_id, row.fan_nomi, row.guruh_kaliti,
            )
            if key is None:
                raise HTTPException(
                    status_code=400,
                    detail=(
                        f"{index}-qator: fan–sinf–guruh tasdiqlangan "
                        "o'quv rejasida yo'q"
                    ),
                )
            if key in targeted:
                raise HTTPException(
                    status_code=409,
                    detail="Bitta fan–sinf–guruh o'zgarishi ikki marta yuborilgan",
                )
            expected_row = expected[key]
            if abs(
                float(row.haftalik_soat)
                - float(expected_row["haftalik_soat"])
            ) > 1e-6:
                raise HTTPException(
                    status_code=409,
                    detail=(
                        f"{index}-qator: haftalik soat tasdiqlangan "
                        "o'quv reja bilan teng emas"
                    ),
                )
            if int(row.kunlik_max) not in range(1, 5):
                raise HTTPException(
                    status_code=400,
                    detail=f"{index}-qator: kunlik maksimum 1–4 bo'lishi kerak",
                )
            room_id = int(row.xona_id) if row.xona_id is not None else None
            if room_id is not None and room_id not in allowed_room_ids:
                raise HTTPException(
                    status_code=400,
                    detail=f"{index}-qator: xona bu maktabga tegishli emas yoki faol emas",
                )
            targeted[key] = {
                **expected_row,
                "user_id": teacher_id,
                # O'quv reja kunlik maksimumi canonical manba bo'lib qoladi.
                "kunlik_max": int(expected_row["kunlik_max"]),
                "xona_id": room_id,
            }

        for index, item in enumerate(sorov.ochirilgan_kalitlar, start=1):
            class_id = int(item.sinf_id)
            if class_id not in school_classes:
                raise HTTPException(
                    status_code=400,
                    detail=f"O'chirish {index}: sinf bu maktabga tegishli emas",
                )
            key = _v204_resolve_expected_skeleton_key(
                expected, class_id, item.fan_nomi, item.guruh_kaliti,
            )
            if key is None:
                raise HTTPException(
                    status_code=400,
                    detail=(
                        f"O'chirish {index}: fan–sinf–guruh tasdiqlangan "
                        "o'quv rejasida yo'q"
                    ),
                )
            if key in targeted:
                raise HTTPException(
                    status_code=409,
                    detail="Bitta fan–sinf–guruh ham saqlash, ham o'chirishga yuborilgan",
                )
            targeted[key] = None

        current_by_key = {}
        unresolved_current_rows = []
        for row in current_rows:
            key = _v204_resolve_expected_skeleton_key(
                expected, row["sinf_id"], row["fan_nomi"], row["guruh_kaliti"],
            )
            if key is not None:
                current_by_key.setdefault(key, []).append(row)
            else:
                unresolved_current_rows.append(row)

        for key in targeted:
            if len(current_by_key.get(key, [])) > 1:
                raise HTTPException(
                    status_code=409,
                    detail=(
                        "Tanlangan fan–sinf–guruh eski ma'lumotda bir nechta "
                        "o'qituvchiga yozilgan. Ma'lumot yo'qolmasligi uchun "
                        "avval O'qituvchi va yuklama oynasida bitta aniq "
                        "qatorga keltiring"
                    ),
                )
            canonical_subject = expected[key]["fan_nomi"]
            hidden_rows = [
                row for row in unresolved_current_rows
                if int(row.get("sinf_id") or 0) == int(key[0])
                and _v204_group_subject_matches(
                    row.get("fan_nomi"), canonical_subject
                )
            ]
            if hidden_rows:
                raise HTTPException(
                    status_code=409,
                    detail=(
                        "Bu sinf-fanda joriy guruh tizimiga ulanmagan eski "
                        "yuklama bor. Uni O'qituvchi va yuklama oynasida "
                        "tozalamasdan yangi tanlov bilan ustidan yozib bo'lmaydi"
                    ),
                )

        # Patchdan keyingi guruhlar holatini oldindan tekshiramiz. Bitta ustoz
        # ayni sinf-fanning ikki guruhiga tushsa transaction boshlanmaydi.
        touched_pairs = {key[:2] for key in targeted}
        resulting_by_key = dict(current_by_key)
        for key, replacement in targeted.items():
            resulting_by_key[key] = [] if replacement is None else [replacement]
        for pair in touched_pairs:
            group_teacher_ids = []
            for key, rows in resulting_by_key.items():
                if key[:2] != pair or key[2] == "whole":
                    continue
                group_teacher_ids.extend(
                    int(row["user_id"]) for row in rows
                )
            if len(group_teacher_ids) != len(set(group_teacher_ids)):
                raise HTTPException(
                    status_code=409,
                    detail=(
                        "Bitta guruhli fanning har bir guruhiga alohida "
                        "o'qituvchi tanlanishi kerak"
                    ),
                )

        leader_class_ids = [int(item.sinf_id) for item in sorov.rahbarlar]
        if len(leader_class_ids) != len(set(leader_class_ids)):
            raise HTTPException(
                status_code=409,
                detail="Bitta sinf rahbarlar ro'yxatida ikki marta yuborilgan",
            )
        resulting_leaders = {
            class_id: (
                int(row["rahbar_user_id"])
                if row.get("rahbar_user_id") is not None else None
            )
            for class_id, row in school_classes.items()
        }
        for item in sorov.rahbarlar:
            class_id = int(item.sinf_id)
            if class_id not in school_classes:
                raise HTTPException(
                    status_code=400,
                    detail="Sinf rahbarligi uchun noma'lum sinf yuborildi",
                )
            teacher_id = int(item.user_id) if item.user_id is not None else None
            if teacher_id is not None and teacher_id not in school_users:
                raise HTTPException(
                    status_code=400,
                    detail=(
                        "Sinf rahbarligi uchun tanlangan o'qituvchi "
                        "bu maktabga tegishli emas"
                    ),
                )
            resulting_leaders[class_id] = teacher_id
        leader_teacher_ids = [
            teacher_id for teacher_id in resulting_leaders.values()
            if teacher_id is not None
        ]
        if len(leader_teacher_ids) != len(set(leader_teacher_ids)):
            raise HTTPException(
                status_code=409,
                detail=(
                    "Bitta o'qituvchi bir vaqtning o'zida ikki sinf "
                    "rahbari qilib tanlangan"
                ),
            )

        affected_teacher_ids = set()
        delete_ids = []
        for key in targeted:
            for old_row in current_by_key.get(key, []):
                delete_ids.append(int(old_row["id"]))
                affected_teacher_ids.add(int(old_row["user_id"]))
        if delete_ids:
            cur.execute(
                """DELETE FROM maktab_dars_birikmalari
                   WHERE maktab_id=%s AND id=ANY(%s)""",
                (sorov.maktab_id, sorted(set(delete_ids))),
            )

        saved_rows = 0
        cleared_rows = 0
        for key, row in targeted.items():
            if row is None:
                cleared_rows += 1
                continue
            affected_teacher_ids.add(int(row["user_id"]))
            cur.execute(
                """INSERT INTO maktab_dars_birikmalari(
                       maktab_id,user_id,sinf_id,fan_nomi,guruh_kaliti,
                       haftalik_soat,kunlik_max,xona_id,manba,yangilangan_at)
                   VALUES(%s,%s,%s,%s,%s,%s,%s,%s,
                          'v20.4_skelet_qisman',NOW())""",
                (
                    sorov.maktab_id, row["user_id"], row["sinf_id"],
                    row["fan_nomi"], row["guruh_kaliti"],
                    row["haftalik_soat"], row["kunlik_max"], row["xona_id"],
                ),
            )
            cur.execute(
                """INSERT INTO maktab_fanlari(maktab_id,fan_nomi)
                   VALUES(%s,%s) ON CONFLICT DO NOTHING""",
                (sorov.maktab_id, row["fan_nomi"]),
            )
            saved_rows += 1

        leader_rule_class_ids = []
        for item in sorov.rahbarlar:
            teacher_id = int(item.user_id) if item.user_id is not None else None
            old_teacher_id = school_classes[int(item.sinf_id)].get("rahbar_user_id")
            if old_teacher_id is not None:
                affected_teacher_ids.add(int(old_teacher_id))
            if teacher_id is not None:
                affected_teacher_ids.add(teacher_id)
            cur.execute(
                """UPDATE maktab_sinflari SET rahbar_user_id=%s
                   WHERE id=%s AND maktab_id=%s""",
                (teacher_id, int(item.sinf_id), sorov.maktab_id),
            )
            if teacher_id is not None:
                leader_rule_class_ids.append(int(item.sinf_id))
        if leader_rule_class_ids:
            _v199_ensure_class_hour_rules(
                cur, sorov.maktab_id, leader_rule_class_ids, actor_id
            )

        teacher_totals = {}
        for teacher_id in sorted(affected_teacher_ids):
            teacher_totals[str(teacher_id)] = _v195_refresh_teacher_summary(
                cur, sorov.maktab_id, teacher_id
            )
        warnings = _v192_sync_schedule_sources(cur, sorov.maktab_id)
        auto_confirmation = _v192_auto_confirm_exact_pairs(
            cur, sorov.maktab_id, actor_id
        )
        matrix = _v192_matrix_payload(cur, sorov.maktab_id)
        conn.commit()
        return {
            "holat": "joriy_ozgarishlar_saqlandi",
            "qator_soni": saved_rows,
            "ochirilgan_qator_soni": cleared_rows,
            "saqlangan_qator": saved_rows,
            "ochirilgan_kalit": cleared_rows,
            "rahbar_ozgarishi": len(sorov.rahbarlar),
            "oqituvchi_soatlari": teacher_totals,
            "ogohlantirishlar": list(
                dict.fromkeys(str(item) for item in (warnings or []) if item)
            )[:40],
            "guruh_tasdiqlari": auto_confirmation,
            "matritsa": matrix,
        }
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


@app.put("/api/maktab/aqlli_jadval/v3/sinf_skeleti_yuklama")
def v204_class_skeleton_bulk_save(sorov: V204SkeletonBulkSave, token: str):
    """
    V20.4 — sinf skeleti orqali o'qituvchi yuklamasini bir martada saqlash.

    Frontend sinf o'quv rejasini fanlari ko'pdan kamga qarab vaqtinchalik
    kataklarga yoyadi. Bu endpoint kataklarning vaqtini saqlamaydi — faqat
    sinf + fan + guruh + o'qituvchi + haftalik soat bog'lanishini atomik
    saqlaydi. Shuning uchun skelet keyingi exact jadval generatoriga toza
    manba bo'lib xizmat qiladi, lekin yakuniy dars kunlarini oldindan
    qotirib qo'ymaydi.
    """
    actor_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor()
    try:
        if not _v1852_manager(cur, actor_id, sorov.maktab_id):
            raise HTTPException(status_code=403, detail="Sinf skeletini faqat maktab rahbariyati saqlaydi")
        conn.commit()
        _v192_tables(cur)
        conn.commit()
        if not _v1852_manager(cur, actor_id, sorov.maktab_id):
            raise HTTPException(status_code=403, detail="Sinf skeletini faqat maktab rahbariyati saqlaydi")
        cur.execute(
            "SELECT pg_advisory_xact_lock(%s)",
            (1922000000 + int(sorov.maktab_id),),
        )
        cur.execute("SELECT pg_advisory_xact_lock(%s)", (2040000000 + int(sorov.maktab_id),))
        cur.execute("SELECT pg_advisory_xact_lock(%s)", (1925000000 + int(sorov.maktab_id),))

        cur.execute("SELECT user_id,full_name FROM users WHERE maktab_id=%s", (sorov.maktab_id,))
        school_users = {int(row["user_id"]): dict(row) for row in cur.fetchall()}
        cur.execute("SELECT id,sinf,harf,rahbar_user_id FROM maktab_sinflari WHERE maktab_id=%s", (sorov.maktab_id,))
        school_classes = {int(row["id"]): dict(row) for row in cur.fetchall()}

        # Bu endpoint butun maktab yuklamasini almashtiradi. Eski tab yoki eski
        # frontend yangi ma'lumotni ustidan yozmasligi uchun to'liq snapshot va
        # GET matritsasida olingan ayni revision majburiy.
        if not sorov.toliq_snapshot:
            raise HTTPException(
                status_code=400,
                detail="Skelet faqat to'liq snapshot sifatida saqlanadi. Sahifani yangilab qayta urinib ko'ring",
            )
        if not sorov.qatorlar:
            raise HTTPException(
                status_code=400,
                detail="Bo'sh skelet bilan mavjud o'qituvchi yuklamasini o'chirib bo'lmaydi",
            )
        current_assignment_rows = _v192_assignment_rows(cur, sorov.maktab_id)
        current_revision = _v204_assignment_revision(
            current_assignment_rows, school_classes.values()
        )
        expected_revision = str(sorov.kutilgan_yuklama_revision or "").strip().lower()
        if not expected_revision or expected_revision != current_revision:
            raise HTTPException(
                status_code=409,
                detail="O'qituvchi yuklamasi boshqa oynada yangilangan. Yangi ma'lumot yo'qolmasligi uchun sahifani qayta yuklang",
            )

        rows_by_teacher = {}
        for index, row in enumerate(sorov.qatorlar, start=1):
            teacher_id = int(row.user_id)
            class_id = int(row.sinf_id)
            if teacher_id not in school_users:
                raise HTTPException(status_code=400, detail=f"{index}-qator: o'qituvchi bu maktabga tegishli emas")
            if class_id not in school_classes:
                raise HTTPException(status_code=400, detail=f"{index}-qator: sinf bu maktabga tegishli emas")
            rows_by_teacher.setdefault(teacher_id, []).append(V192TeacherLoadRow(
                sinf_id=class_id,
                fan_nomi=row.fan_nomi,
                guruh_kaliti=row.guruh_kaliti or "whole",
                haftalik_soat=row.haftalik_soat,
                kunlik_max=row.kunlik_max,
                xona_id=row.xona_id,
            ))

        _v204_validate_complete_skeleton_payload(
            cur, sorov.maktab_id, sorov.qatorlar
        )

        leader_class_ids = [int(item.sinf_id) for item in sorov.rahbarlar]
        if len(leader_class_ids) != len(set(leader_class_ids)):
            raise HTTPException(status_code=409, detail="Bitta sinf rahbarlar ro'yxatida ikki marta yuborilgan")
        if set(leader_class_ids) != set(school_classes):
            raise HTTPException(
                status_code=400,
                detail="Sinf rahbarlari ham barcha sinflar bo'yicha to'liq snapshot bo'lishi kerak",
            )
        leader_teacher_ids = [int(item.user_id) for item in sorov.rahbarlar if item.user_id is not None]
        if len(leader_teacher_ids) != len(set(leader_teacher_ids)):
            raise HTTPException(status_code=409, detail="Bitta o'qituvchi bir vaqtning o'zida ikki sinf rahbari qilib tanlangan")
        for item in sorov.rahbarlar:
            class_id = int(item.sinf_id)
            if class_id not in school_classes:
                raise HTTPException(status_code=400, detail="Sinf rahbarligi uchun noma'lum sinf yuborildi")
            if item.user_id is not None and int(item.user_id) not in school_users:
                raise HTTPException(status_code=400, detail="Sinf rahbarligi uchun tanlangan o'qituvchi bu maktabga tegishli emas")

        # Barcha yuklama bir tranzaksiyada almashtiriladi. Xato bo'lsa rollback
        # eski holatni to'liq qaytaradi; yarim saqlangan jadval qolmaydi.
        old_assignment_teacher_ids = {
            int(row["user_id"]) for row in current_assignment_rows
        }
        cur.execute(
            """SELECT DISTINCT user_id FROM maktab_xodim_sinflari
               WHERE maktab_id=%s""",
            (sorov.maktab_id,),
        )
        old_derived_teacher_ids = {
            int(row["user_id"]) for row in cur.fetchall()
        }
        affected_teacher_ids = sorted(
            old_assignment_teacher_ids | old_derived_teacher_ids | set(rows_by_teacher)
        )
        cur.execute("DELETE FROM maktab_dars_birikmalari WHERE maktab_id=%s", (sorov.maktab_id,))
        if affected_teacher_ids:
            # Canonical qatorlar to'liq almashtirilganda eski derived fan/sinf
            # summarylari qolib, keyingi matritsaga soxta yuklama bo'lib
            # qaytmasin. Profil, maqsad soati, staj va jadval raqami tegilmaydi.
            cur.execute(
                """DELETE FROM maktab_xodim_sinflari
                   WHERE maktab_id=%s AND user_id=ANY(%s)""",
                (sorov.maktab_id, affected_teacher_ids),
            )
            cur.execute(
                """UPDATE users SET haftalik_dars_soati=0,
                                      fanlari=NULL,oqitadigan_sinflari=NULL
                   WHERE maktab_id=%s AND user_id=ANY(%s)""",
                (sorov.maktab_id, affected_teacher_ids),
            )

        # Sinf rahbarlari ham shu skeletning bir qismi. Frontend barcha sinflarni
        # yuboradi, shuning uchun eski noto'g'ri rahbarlar qolib ketmaydi.
        if sorov.rahbarlar:
            cur.execute("UPDATE maktab_sinflari SET rahbar_user_id=NULL WHERE maktab_id=%s", (sorov.maktab_id,))
            leader_class_ids = []
            for item in sorov.rahbarlar:
                if item.user_id is None:
                    continue
                cur.execute(
                    "UPDATE maktab_sinflari SET rahbar_user_id=%s WHERE id=%s AND maktab_id=%s",
                    (int(item.user_id), int(item.sinf_id), int(sorov.maktab_id)),
                )
                leader_class_ids.append(int(item.sinf_id))
            if leader_class_ids:
                _v199_ensure_class_hour_rules(cur, sorov.maktab_id, leader_class_ids, actor_id)

        saved_teachers = 0
        saved_rows = 0
        saved_hours = 0.0
        warnings = []
        # Deterministik tartib: bir xil payload har safar bir xil validatsiya yo'lidan o'tadi.
        for teacher_id in sorted(rows_by_teacher):
            result = _v192_save_teacher_load_rows(
                cur, actor_id, sorov.maktab_id, teacher_id, rows_by_teacher[teacher_id],
                finalize=False,
            )
            saved_teachers += 1
            saved_rows += int(result.get("qator_soni") or 0)
            saved_hours += float(result.get("haftalik_jami") or 0)
            warnings.extend(result.get("ogohlantirishlar") or [])

        # Yuklamasi bo'lmagan ustozlar uchun summary 0 bo'lib qoladi.
        sync_warnings = _v192_sync_schedule_sources(cur, sorov.maktab_id)
        warnings.extend(sync_warnings or [])
        auto_confirmation = _v192_auto_confirm_exact_pairs(cur, sorov.maktab_id, actor_id)
        matrix = _v192_matrix_payload(cur, sorov.maktab_id)
        conn.commit()
        return {
            "holat": "sinf_skeleti_saqlandi",
            "oqituvchi_soni": saved_teachers,
            "qator_soni": saved_rows,
            "fan_soati": round(saved_hours, 1),
            "ogohlantirishlar": list(dict.fromkeys(str(x) for x in warnings if x))[:40],
            "guruh_tasdiqlari": auto_confirmation,
            "matritsa": matrix,
        }
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


class V195TeacherDelete(BaseModel):
    maktab_id: int
    user_id: int
    tasdiq: bool = False


@app.post("/api/maktab/aqlli_jadval/v3/oqituvchi_ochirish")
def v195_teacher_delete(sorov: V195TeacherDelete, token: str):
    """O'qituvchini maktabdan va barcha faol yuklamalardan to'liq chiqaradi."""
    actor_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor()
    try:
        if not _v1852_manager(cur, actor_id, sorov.maktab_id):
            raise HTTPException(status_code=403, detail="O'qituvchini faqat maktab rahbariyati o'chiradi")
        conn.commit()
        _v192_tables(cur)
        conn.commit()
        if not _v1852_manager(cur, actor_id, sorov.maktab_id):
            raise HTTPException(status_code=403, detail="O'qituvchini faqat maktab rahbariyati o'chiradi")
        if not sorov.tasdiq:
            raise HTTPException(status_code=400, detail="O'chirish uchun Ha tasdig'i kerak")
        if int(sorov.user_id) == int(actor_id):
            raise HTTPException(status_code=400, detail="O'zingizning rahbariyat hisobingizni bu yerdan o'chira olmaysiz")
        cur.execute(
            "SELECT pg_advisory_xact_lock(%s)",
            (1922000000 + int(sorov.maktab_id),),
        )
        cur.execute("SELECT pg_advisory_xact_lock(%s)", (1925000000 + int(sorov.maktab_id),))
        cur.execute("""SELECT user_id,full_name,lavozim FROM users
                       WHERE user_id=%s AND maktab_id=%s FOR UPDATE""",
                    (sorov.user_id, sorov.maktab_id))
        teacher = cur.fetchone()
        if not teacher:
            raise HTTPException(status_code=404, detail="O'qituvchi topilmadi yoki avval o'chirilgan")
        if str(teacher.get("lavozim") or "").lower() in {
            "direktor", "zam_direktor_uquv", "zam_direktor_tarbiya", "owner", "admin"
        }:
            raise HTTPException(status_code=409, detail="Rahbariyat hisobini o'qituvchi oynasidan o'chirib bo'lmaydi")

        cur.execute("""SELECT COUNT(*) AS son,COALESCE(SUM(haftalik_soat),0) AS soat
                       FROM maktab_dars_birikmalari
                       WHERE maktab_id=%s AND user_id=%s""",
                    (sorov.maktab_id, sorov.user_id))
        old_load = cur.fetchone() or {}
        cur.execute("UPDATE maktab_sinflari SET rahbar_user_id=NULL WHERE maktab_id=%s AND rahbar_user_id=%s",
                    (sorov.maktab_id, sorov.user_id))
        if "psixolog_user_id" in _v1857_columns(cur, "maktab_sinflari"):
            cur.execute("UPDATE maktab_sinflari SET psixolog_user_id=NULL WHERE maktab_id=%s AND psixolog_user_id=%s",
                        (sorov.maktab_id, sorov.user_id))
        cur.execute("DELETE FROM maktab_dars_birikmalari WHERE maktab_id=%s AND user_id=%s",
                    (sorov.maktab_id, sorov.user_id))
        cur.execute("DELETE FROM maktab_xodim_sinflari WHERE maktab_id=%s AND user_id=%s",
                    (sorov.maktab_id, sorov.user_id))
        if _v1857_has_columns(cur, "aqlli_oqituvchi_qoidalari_v2", {"maktab_id", "user_id"}):
            cur.execute("DELETE FROM aqlli_oqituvchi_qoidalari_v2 WHERE maktab_id=%s AND user_id=%s",
                        (sorov.maktab_id, sorov.user_id))
        if _v1857_has_columns(cur, "aqlli_oqituvchi_vaqti_v2", {"maktab_id", "user_id"}):
            cur.execute("DELETE FROM aqlli_oqituvchi_vaqti_v2 WHERE maktab_id=%s AND user_id=%s",
                        (sorov.maktab_id, sorov.user_id))
        if _v1857_has_columns(cur, "aqlli_guruh_sozlamalari_v2", {"maktab_id", "oqituvchi_user_id"}):
            cur.execute("DELETE FROM aqlli_guruh_sozlamalari_v2 WHERE maktab_id=%s AND oqituvchi_user_id=%s",
                        (sorov.maktab_id, sorov.user_id))
        if _v1857_has_columns(cur, "aqlli_sinf_fan_yuklamalari_v2", {"maktab_id", "asosiy_oqituvchi_user_id"}):
            cur.execute("""UPDATE aqlli_sinf_fan_yuklamalari_v2
                           SET asosiy_oqituvchi_user_id=NULL
                           WHERE maktab_id=%s AND asosiy_oqituvchi_user_id=%s""",
                        (sorov.maktab_id, sorov.user_id))
        if _v1857_has_columns(cur, "xodim_davomati", {"maktab_id", "user_id"}):
            cur.execute("DELETE FROM xodim_davomati WHERE maktab_id=%s AND user_id=%s",
                        (sorov.maktab_id, sorov.user_id))
        if _v1857_has_columns(cur, "xodim_kod", {"user_id"}):
            cur.execute("DELETE FROM xodim_kod WHERE user_id=%s", (sorov.user_id,))
        if _v1857_has_columns(cur, "dars_jadvali", {"oqituvchi_user_id"}):
            cur.execute("DELETE FROM dars_jadvali WHERE oqituvchi_user_id=%s", (sorov.user_id,))

        cur.execute("""UPDATE users
                       SET maktab_id=NULL,lavozim=NULL,fanlari=NULL,
                           oqitadigan_sinflari=NULL,haftalik_dars_soati=NULL,
                           haftalik_maqsad_soat=NULL,mutaxassisligi=NULL
                       WHERE user_id=%s AND maktab_id=%s""",
                    (sorov.user_id, sorov.maktab_id))
        cur.execute("""UPDATE aqlli_jadval_urinishlari_v2 SET holat='bekor'
                       WHERE maktab_id=%s AND holat IN ('draft','tasdiqlangan')""",
                    (sorov.maktab_id,))
        cur.execute("""UPDATE aqlli_fan_guruh_tasdiqlari_v2
                       SET tasdiqlangan=FALSE,yangilangan_at=NOW()
                       WHERE maktab_id=%s""", (sorov.maktab_id,))
        warnings = _v192_sync_schedule_sources(cur, sorov.maktab_id)
        matrix = _v192_matrix_payload(cur, sorov.maktab_id)
        conn.commit()
        return {
            "holat": "oqituvchi_ochirildi",
            "user_id": int(sorov.user_id),
            "oqituvchi": teacher["full_name"],
            "ochirilgan_qator": int(old_load.get("son") or 0),
            "ochirilgan_soat": int(old_load.get("soat") or 0),
            "ogohlantirishlar": warnings,
            "matritsa": matrix,
        }
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


class V192AutoMode(BaseModel):
    maktab_id: int
    avtomatik_tavsiya: bool


@app.put("/api/maktab/aqlli_jadval/v3/almashtirish_rejimi")
def v192_swap_mode(sorov: V192AutoMode, token: str):
    actor_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor()
    try:
        _v192_tables(cur)
        if not _v1852_manager(cur, actor_id, sorov.maktab_id):
            raise HTTPException(status_code=403, detail="Jadval rejimini faqat rahbariyat o'zgartiradi")
        cur.execute("""INSERT INTO aqlli_jadval_boshqaruv_v19_2(
                        maktab_id,avtomatik_tavsiya,yangilagan_user_id,yangilangan_at)
                       VALUES(%s,%s,%s,NOW())
                       ON CONFLICT(maktab_id) DO UPDATE SET
                         avtomatik_tavsiya=EXCLUDED.avtomatik_tavsiya,
                         yangilagan_user_id=EXCLUDED.yangilagan_user_id,
                         yangilangan_at=NOW()""",
                    (sorov.maktab_id, sorov.avtomatik_tavsiya, actor_id))
        conn.commit()
        return {"holat": "saqlandi", "avtomatik_tavsiya": sorov.avtomatik_tavsiya}
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


def _v192_run_slots(cur, run_id: int):
    cur.execute("""SELECT e.*,s.sinf,s.harf,u.full_name AS oqituvchi_ismi,
                          COALESCE(r.nomi,e.xona_matni) AS xona_nomi
                   FROM aqlli_jadval_slotlari_v2 e
                   JOIN maktab_sinflari s ON s.id=e.sinf_id
                   LEFT JOIN users u ON u.user_id=e.oqituvchi_user_id
                   LEFT JOIN aqlli_xonalar_v2 r ON r.id=e.xona_id
                   WHERE e.urinish_id=%s
                   ORDER BY e.hafta_kuni,e.smena,e.dars_raqami,
                            s.sinf::int,s.harf,e.guruh_kaliti""", (run_id,))
    return [dict(row) for row in cur.fetchall()]


def _v200_xlsx_sheet_name(value, used):
    cleaned = re.sub(r"[\\/*?:\[\]]+", " ", str(value or "Jadval")).strip()
    cleaned = re.sub(r"\s+", " ", cleaned)[:31] or "Jadval"
    candidate = cleaned
    suffix = 2
    while candidate.casefold() in used:
        tail = f" {suffix}"
        candidate = f"{cleaned[:31-len(tail)]}{tail}"
        suffix += 1
    used.add(candidate.casefold())
    return candidate


def _v200_xlsx_slot_text(rows, teacher_view=False):
    lines = []
    seen = set()
    for row in rows:
        group = str(row.get("guruh_kaliti") or "whole")
        group_label = ""
        if group != "whole":
            normalized = group.casefold()
            if normalized in {"boys", "boy", "ogil", "o'g'il", "o‘g‘il"}:
                group_label = "O'g'il bolalar"
            elif normalized in {"girls", "girl", "qiz"}:
                group_label = "Qiz bolalar"
            else:
                match = re.search(r"(\d+)", normalized)
                group_label = f"{match.group(1)}-guruh" if match else group
        week = str(row.get("hafta_turi") or "har_hafta")
        week_label = ""
        if week in {"toq", "juft"}:
            week_label = "TOQ" if week == "toq" else "JUFT"
        subject = str(row.get("fan_nomi") or "Fan")
        class_label = f"{row.get('sinf','')}-{row.get('harf','')}"
        teacher = str(row.get("oqituvchi_ismi") or "O'qituvchi yo'q")
        room = str(row.get("xona_nomi") or row.get("xona_matni") or "Xona yo'q")
        prefix = " · ".join(value for value in [group_label, week_label] if value)
        first = " · ".join(value for value in [class_label if teacher_view else "", prefix, subject] if value)
        second = " · ".join(value for value in ["" if teacher_view else teacher, room] if value)
        text = f"{first}\n{second}" if second else first
        if text not in seen:
            seen.add(text)
            lines.append(text)
    return "\n——\n".join(lines)


def _v200_xlsx_style_workbook(workbook):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    theme = {
        "ink": "18324B", "blue": "155A7A", "teal": "0F7C82",
        "sky": "EDF5FB", "mint": "EAF7F4", "cream": "FAF7F0",
        "line": "DDE6EC", "green": "33755A", "red": "A54242",
    }
    thin = Side(style="thin", color=theme["line"])
    return theme, Font, PatternFill, Alignment, Border, thin


def _v200_xlsx_prepare_schedule_sheet(
    ws, title, subtitle, summary_text, weekdays, rows, teacher_view=False,
    default_shift=1,
):
    import openpyxl
    theme, Font, PatternFill, Alignment, Border, thin = _v200_xlsx_style_workbook(ws.parent)
    last_col = 1 + len(weekdays)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    ws["A1"] = title
    ws["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=theme["blue"])
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 29
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    ws["A2"] = subtitle
    ws["A2"].font = Font(size=9, color="6D7B87")
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=last_col)
    ws["A3"] = summary_text
    ws["A3"].font = Font(size=10, bold=True, color=theme["green"])
    ws["A3"].fill = PatternFill("solid", fgColor=theme["mint"])
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[3].height = 26
    ws.cell(4, 1, "Dars")
    for index, (_, name) in enumerate(weekdays, start=2):
        ws.cell(4, index, name)
    for cell in ws[4]:
        cell.font = Font(size=9, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=theme["teal"])
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    key_map = _v1852_defaultdict(list)
    for row in rows:
        key_map[(int(row["hafta_kuni"]), int(row["smena"]), int(row["dars_raqami"]))].append(row)
    schedule_rows = []
    if teacher_view:
        schedule_rows = [(shift, period, f"{shift}-smena · {period}") for shift in (1, 2) for period in range(1, 7)]
    else:
        class_shift = int(rows[0].get("smena") or default_shift) if rows else int(default_shift or 1)
        schedule_rows = [(class_shift, period, str(period)) for period in range(1, 7)]

    for output_row, (shift, period, label) in enumerate(schedule_rows, start=5):
        ws.cell(output_row, 1, label)
        ws.cell(output_row, 1).font = Font(size=9, bold=True, color=theme["ink"])
        ws.cell(output_row, 1).fill = PatternFill("solid", fgColor=theme["cream"])
        ws.cell(output_row, 1).alignment = Alignment(horizontal="center", vertical="center")
        for column, (day, _) in enumerate(weekdays, start=2):
            cell = ws.cell(output_row, column)
            cell.value = _v200_xlsx_slot_text(
                key_map.get((int(day), shift, period), []),
                teacher_view=teacher_view,
            )
            cell.font = Font(size=8, bold=bool(cell.value), color=theme["ink"])
            cell.fill = PatternFill("solid", fgColor=theme["sky"] if cell.value else "FFFFFF")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        ws.row_dimensions[output_row].height = 42 if teacher_view else 55
    ws.column_dimensions["A"].width = 13
    for column in range(2, last_col + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(column)].width = 24
    ws.freeze_panes = "B5"
    ws.sheet_view.showGridLines = False
    ws.auto_filter.ref = f"A4:{openpyxl.utils.get_column_letter(last_col)}{4+len(schedule_rows)}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = f"A1:{openpyxl.utils.get_column_letter(last_col)}{4+len(schedule_rows)}"
    ws.oddFooter.center.text = "SAMTM Aqlli jadval"
    ws.oddFooter.right.text = "&P / &N"


@app.get("/api/maktab/aqlli_jadval/v3/jadval_xlsx")
def v200_schedule_xlsx(token: str, urinish_id: int, turi: str = "sinflar"):
    """Sinf yoki o'qituvchi kesimida varaqlangan, chopga tayyor XLSX."""
    actor_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor()
    try:
        _v192_tables(cur)
        cur.execute(
            "SELECT * FROM aqlli_jadval_urinishlari_v2 WHERE id=%s",
            (urinish_id,),
        )
        run = cur.fetchone()
        if not run:
            raise HTTPException(status_code=404, detail="Jadval topilmadi")
        maktab_id = int(run["maktab_id"])
        if not _v1852_staff(cur, actor_id, maktab_id):
            raise HTTPException(status_code=403, detail="Bu jadvalni yuklashga ruxsat yo'q")
        export_type = str(turi or "sinflar").strip().lower()
        if export_type not in {"sinflar", "oqituvchilar"}:
            raise HTTPException(status_code=400, detail="turi sinflar yoki oqituvchilar bo'lishi kerak")
        cur.execute("SELECT nomi FROM maktablar WHERE id=%s", (maktab_id,))
        school = cur.fetchone() or {"nomi": "Maktab"}
        slots = _v192_run_slots(cur, urinish_id)
        report = _v1875_schedule_integrity_report(cur, maktab_id, urinish_id)
        class_summary = {int(row["sinf_id"]): row for row in report.get("sinflar", [])}
        teacher_summary = {int(row["user_id"]): row for row in report.get("oqituvchilar", [])}
        year = _v1852_active_year(cur, maktab_id) or {}
        weekdays = [
            (day, _V1852_HAFTA.get(day, str(day)))
            for day in range(1, int(year.get("hafta_kunlari") or 6) + 1)
        ]

        import io
        import openpyxl
        from fastapi.responses import StreamingResponse
        from openpyxl.styles import Alignment, Font, PatternFill
        from urllib.parse import quote

        workbook = openpyxl.Workbook()
        index = workbook.active
        index.title = "Mundarija"
        index.sheet_view.showGridLines = False
        index.append(["SAMTM AQILLI DARS JADVALI"])
        index.append([school.get("nomi") or "Maktab"])
        index.append(["Turi", "Nomi", "Haftalik hisob", "Holat"])
        used = {"mundarija"}
        created = []
        if export_type == "sinflar":
            cur.execute(
                "SELECT id,sinf,harf,COALESCE(smena,1) AS smena FROM maktab_sinflari "
                "WHERE maktab_id=%s ORDER BY sinf::int,harf",
                (maktab_id,),
            )
            entities = [dict(row) for row in cur.fetchall()]
            for entity in entities:
                entity_id = int(entity["id"])
                label = f"{entity['sinf']}-{entity['harf']}"
                rows = [row for row in slots if int(row["sinf_id"]) == entity_id]
                summary = class_summary.get(entity_id, {})
                summary_text = summary.get("tasdiq_matni") or (
                    f"Reja {summary.get('reja', 0)} · jadval {summary.get('jadval', 0)}"
                )
                sheet_name = _v200_xlsx_sheet_name(label, used)
                ws = workbook.create_sheet(sheet_name)
                _v200_xlsx_prepare_schedule_sheet(
                    ws, f"{label} · haftalik dars jadvali",
                    school.get("nomi") or "Maktab", summary_text,
                    weekdays, rows, teacher_view=False,
                    default_shift=int(entity.get("smena") or 1),
                )
                created.append(("Sinf", label, summary_text, "TO'LIQ" if summary.get("mos") else "TEKSHIRISH", sheet_name))
        else:
            cur.execute(
                "SELECT user_id,full_name FROM users WHERE maktab_id=%s ORDER BY full_name,user_id",
                (maktab_id,),
            )
            names = {int(row["user_id"]): row["full_name"] for row in cur.fetchall()}
            entity_ids = sorted(teacher_summary, key=lambda uid: str(names.get(uid, uid)).casefold())
            for entity_id in entity_ids:
                label = names.get(entity_id) or teacher_summary[entity_id].get("full_name") or str(entity_id)
                rows = [row for row in slots if int(row.get("oqituvchi_user_id") or 0) == entity_id]
                summary = teacher_summary.get(entity_id, {})
                summary_text = summary.get("tasdiq_matni") or (
                    f"Reja {summary.get('reja', 0)} · jadval {summary.get('jadval', 0)}"
                )
                sheet_name = _v200_xlsx_sheet_name(label, used)
                ws = workbook.create_sheet(sheet_name)
                _v200_xlsx_prepare_schedule_sheet(
                    ws, f"{label} · haftalik ish jadvali",
                    school.get("nomi") or "Maktab", summary_text,
                    weekdays, rows, teacher_view=True,
                )
                created.append(("O'qituvchi", label, summary_text, "TO'LIQ" if summary.get("mos") else "TEKSHIRISH", sheet_name))

        for row_number, (kind, label, summary_text, status, sheet_name) in enumerate(created, start=4):
            index.cell(row_number, 1, kind)
            index.cell(row_number, 2, label)
            index.cell(row_number, 2).hyperlink = f"#'{sheet_name.replace(chr(39), chr(39)*2)}'!A1"
            index.cell(row_number, 3, summary_text)
            index.cell(row_number, 4, status)
        index.merge_cells("A1:D1")
        index["A1"].font = Font(size=19, bold=True, color="FFFFFF")
        index["A1"].fill = PatternFill("solid", fgColor="155A7A")
        index.merge_cells("A2:D2")
        index["A2"].font = Font(size=11, bold=True, color="0F7C82")
        for cell in index[3]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="0F7C82")
            cell.alignment = Alignment(horizontal="center")
        for width, column in zip((16, 34, 70, 16), ("A", "B", "C", "D")):
            index.column_dimensions[column].width = width
        index.freeze_panes = "A4"
        index.page_setup.orientation = "landscape"
        index.page_setup.fitToWidth = 1
        index.sheet_properties.pageSetUpPr.fitToPage = True

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        filename = f"SAMTM_{export_type}_jadvali_{urinish_id}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
        )
    finally:
        cur.close(); conn.close()


def _v192_bundle_for_slot(slots, slot_id: int):
    source = next((row for row in slots if int(row["id"]) == int(slot_id)), None)
    if not source:
        return None, []
    bundle = [
        row for row in slots
        if int(row["sinf_id"]) == int(source["sinf_id"])
        and int(row["hafta_kuni"]) == int(source["hafta_kuni"])
        and int(row["smena"]) == int(source["smena"])
        and int(row["dars_raqami"]) == int(source["dars_raqami"])
    ]
    return source, bundle


def _v192_bundle_label(bundle):
    if not bundle:
        return "Bo'sh vaqt"
    subjects = list(dict.fromkeys(str(row["fan_nomi"]) for row in bundle))
    teachers = list(dict.fromkeys(
        str(row.get("oqituvchi_ismi") or "O'qituvchi belgilanmagan")
        for row in bundle
    ))
    return f"{' + '.join(subjects)} · {' + '.join(teachers)}"


def _v192_candidate_conflicts(
    moving_bundle, target_day, target_period, all_slots, ignored_ids,
    hard_rows, class_row, class_day_map, room_check=True,
):
    reasons = []
    shift = int(moving_bundle[0]["smena"])
    blocked = _v1856_class_day_block_reason(class_row, target_day, class_day_map)
    if blocked:
        reasons.append(blocked)
    moving_teachers = [
        int(row["oqituvchi_user_id"])
        for row in moving_bundle if row.get("oqituvchi_user_id") is not None
    ]
    if len(moving_teachers) != len(set(moving_teachers)):
        reasons.append("bir o'qituvchi parallel guruhlarda takrorlangan")
    for teacher_id in moving_teachers:
        for rule in hard_rows:
            if int(rule["user_id"]) != teacher_id:
                continue
            if int(rule["hafta_kuni"]) != int(target_day):
                continue
            if rule["turi"] == "metod_kuni":
                reasons.append("o'qituvchining metod kuni")
                break
            if not bool(rule.get("qattiq")):
                continue
            rule_shift = int(rule.get("smena") or 0)
            rule_period = int(rule.get("dars_raqami") or 0)
            if rule_shift in (0, shift) and rule_period in (0, int(target_period)):
                reasons.append("o'qituvchi bu vaqtda band")
                break
        if any(
            int(row["id"]) not in ignored_ids
            and row.get("oqituvchi_user_id") is not None
            and int(row["oqituvchi_user_id"]) == teacher_id
            and int(row["hafta_kuni"]) == int(target_day)
            and int(row["smena"]) == shift
            and int(row["dars_raqami"]) == int(target_period)
            for row in all_slots
        ):
            reasons.append("o'qituvchi boshqa sinfda")
    if room_check:
        for moving in moving_bundle:
            room_id = moving.get("xona_id")
            room_text = str(moving.get("xona_matni") or "").strip().casefold()
            if room_id and any(
                    int(row["id"]) not in ignored_ids
                    and row.get("xona_id") is not None
                    and int(row["xona_id"]) == int(room_id)
                    and int(row["hafta_kuni"]) == int(target_day)
                    and int(row["smena"]) == shift
                    and int(row["dars_raqami"]) == int(target_period)
                    for row in all_slots
            ):
                reasons.append("xona bu vaqtda band")
            elif room_text and any(
                    int(row["id"]) not in ignored_ids
                    and str(row.get("xona_matni") or "").strip().casefold() == room_text
                    and int(row["hafta_kuni"]) == int(target_day)
                    and int(row["smena"]) == shift
                    and int(row["dars_raqami"]) == int(target_period)
                    for row in all_slots
            ):
                reasons.append("sinf xonasi bu vaqtda band")
    return list(dict.fromkeys(reasons))


def _v192_parallel_conflicts(slots):
    grouped = {}
    for row in slots:
        teacher_id = row.get("oqituvchi_user_id")
        if teacher_id is None:
            continue
        week_type = str(row.get("hafta_turi") or "har_hafta")
        phases = ("toq", "juft") if week_type == "har_hafta" else (week_type,)
        for phase in phases:
            key = (
                int(teacher_id), int(row["hafta_kuni"]),
                int(row["smena"]), int(row["dars_raqami"]), phase,
            )
            grouped.setdefault(key, []).append(row)
    result = []
    for key, rows in grouped.items():
        class_ids = {int(row["sinf_id"]) for row in rows}
        if len(class_ids) > 1:
            result.append({
                "oqituvchi_user_id": key[0],
                "oqituvchi_ismi": rows[0].get("oqituvchi_ismi"),
                "hafta_kuni": key[1],
                "smena": key[2],
                "dars_raqami": key[3],
                "hafta_turi": key[4],
                "sinflar": list(dict.fromkeys(f"{row['sinf']}-{row['harf']}" for row in rows)),
                "slot_idlar": [int(row["id"]) for row in rows],
            })
    return result


def _v192_class_periods_contiguous(slots, class_id: int, changes=None):
    """Taklif qilingan ko'chirishdan keyin sinf kunlari oknosiz qolishini tekshiradi."""
    changes = changes or {}
    day_periods = _v1852_defaultdict(set)
    for row in slots:
        if int(row["sinf_id"]) != int(class_id):
            continue
        day = int(row["hafta_kuni"])
        period = int(row["dars_raqami"])
        replacement = changes.get(int(row["id"]))
        if replacement:
            day, period = int(replacement[0]), int(replacement[1])
        day_periods[day].add(period)
    return all(_v1852_gap_count(periods) == 0 for periods in day_periods.values())


def _v213_projected_core_period6_days(slots, class_id: int, changes=None):
    """Qo'lda ko'chirishdan keyingi asosiy-fan 6-dars kunlarini hisoblaydi."""
    changes = changes or {}
    result = set()
    for row in slots:
        if int(row.get("sinf_id") or 0) != int(class_id):
            continue
        day = int(row.get("hafta_kuni") or 0)
        period = int(row.get("dars_raqami") or 0)
        replacement = changes.get(int(row.get("id") or 0))
        if replacement:
            day, period = int(replacement[0]), int(replacement[1])
        if period != 6:
            continue
        grade = _v1874_grade(row)
        profile = _v1874_subject_profile(row.get("fan_nomi"), grade)
        if profile.get("core_priority"):
            result.add(day)
    return result


def _v192_swap_suggestions(cur, run_id: int, slot_id: int):
    cur.execute("SELECT * FROM aqlli_jadval_urinishlari_v2 WHERE id=%s", (run_id,))
    run = cur.fetchone()
    if not run:
        raise HTTPException(status_code=404, detail="Jadval topilmadi")
    slots = _v192_run_slots(cur, run_id)
    source, source_bundle = _v192_bundle_for_slot(slots, slot_id)
    if not source:
        raise HTTPException(status_code=404, detail="Tanlangan dars topilmadi")
    if any(str(row.get("fan_nomi") or "").strip().casefold() == "sinf soati" for row in source_bundle):
        raise HTTPException(status_code=409, detail="Sinf soati faqat 3-bosqichdagi qoida orqali ko'chiriladi")
    cur.execute("SELECT * FROM maktab_sinflari WHERE id=%s", (source["sinf_id"],))
    class_row = cur.fetchone()
    class_day_map = _v1856_class_day_rule_map(
        _v1856_class_day_rule_rows(cur, run["maktab_id"])
    )
    cur.execute("""SELECT * FROM aqlli_oqituvchi_vaqti_v2
                   WHERE maktab_id=%s""", (run["maktab_id"],))
    hard_rows = [dict(row) for row in cur.fetchall()]
    cur.execute("""SELECT dars_soni FROM aqlli_smena_sozlamalari_v2
                   WHERE maktab_id=%s AND smena=%s""",
                (run["maktab_id"], source["smena"]))
    shift = cur.fetchone()
    max_period = int((shift or {}).get("dars_soni") or 7)
    year = _v1852_active_year(cur, run["maktab_id"])
    weekdays = int((year or {}).get("hafta_kunlari") or 6)
    source_ids = {int(row["id"]) for row in source_bundle}
    suggestions = []
    for day in range(1, weekdays + 1):
        for period in range(1, max_period + 1):
            if day == int(source["hafta_kuni"]) and period == int(source["dars_raqami"]):
                continue
            target_bundle = [
                row for row in slots
                if int(row["sinf_id"]) == int(source["sinf_id"])
                and int(row["hafta_kuni"]) == day
                and int(row["smena"]) == int(source["smena"])
                and int(row["dars_raqami"]) == period
            ]
            if any(str(row.get("fan_nomi") or "").strip().casefold() == "sinf soati" for row in target_bundle):
                continue
            target_ids = {int(row["id"]) for row in target_bundle}
            projected_changes = {
                int(row["id"]): (day, period) for row in source_bundle
            }
            projected_changes.update({
                int(row["id"]): (
                    int(source["hafta_kuni"]), int(source["dars_raqami"])
                )
                for row in target_bundle
            })
            if not _v192_class_periods_contiguous(
                slots, int(source["sinf_id"]), projected_changes
            ):
                continue
            if len(_v213_projected_core_period6_days(
                slots, int(source["sinf_id"]), projected_changes
            )) > _V213_CORE_PERIOD6_LIMIT:
                continue
            source_reasons = _v192_candidate_conflicts(
                source_bundle, day, period, slots, source_ids | target_ids,
                hard_rows, class_row, class_day_map,
            )
            target_reasons = []
            if target_bundle:
                target_reasons = _v192_candidate_conflicts(
                    target_bundle, int(source["hafta_kuni"]), int(source["dars_raqami"]),
                    slots, source_ids | target_ids, hard_rows, class_row, class_day_map,
                )
            reasons = list(dict.fromkeys(source_reasons + target_reasons))
            if reasons:
                continue
            kind = "almashtirish" if target_bundle else "kochirish"
            distance = abs(day - int(source["hafta_kuni"])) * 2 + abs(period - int(source["dars_raqami"]))
            teacher_daily = sum(
                1 for row in slots
                if row.get("oqituvchi_user_id") in {
                    item.get("oqituvchi_user_id") for item in source_bundle
                }
                and int(row["hafta_kuni"]) == day
            )
            class_grade = _v1874_grade(class_row)
            source_period_penalty = max(
                _v1874_subject_period_penalty(
                    _v1874_subject_profile(row.get("fan_nomi"), class_grade),
                    class_grade,
                    period,
                    max_period,
                )
                for row in source_bundle
            )
            target_period_penalty = 0
            if target_bundle:
                target_period_penalty = max(
                    _v1874_subject_period_penalty(
                        _v1874_subject_profile(row.get("fan_nomi"), class_grade),
                        class_grade,
                        int(source["dars_raqami"]),
                        max_period,
                    )
                    for row in target_bundle
                )
            score = (
                distance + teacher_daily
                + (3 if kind == "almashtirish" else 0)
                + source_period_penalty + target_period_penalty
            )
            suggestions.append({
                "turi": kind,
                "yangi_hafta_kuni": day,
                "yangi_smena": int(source["smena"]),
                "yangi_dars_raqami": period,
                "baho": score,
                "nishon": _v192_bundle_label(target_bundle),
                "nishon_slot_idlar": sorted(target_ids),
                "izoh": (
                    "Bo'sh katakka okno qoldirmasdan ko'chirish mumkin"
                    if not target_bundle
                    else "Ikki dars joyini oknosiz va xavfsiz almashtirish mumkin"
                ),
            })
    suggestions.sort(key=lambda row: (row["baho"], row["yangi_hafta_kuni"], row["yangi_dars_raqami"]))
    cur.execute("""SELECT avtomatik_tavsiya FROM aqlli_jadval_boshqaruv_v19_2
                   WHERE maktab_id=%s""", (run["maktab_id"],))
    mode = cur.fetchone()
    return {
        "urinish_id": int(run["id"]),
        "maktab_id": int(run["maktab_id"]),
        "manba": {
            "slot_id": int(source["id"]),
            "sinf_id": int(source["sinf_id"]),
            "sinf": f"{source['sinf']}-{source['harf']}",
            "fan": _v192_bundle_label(source_bundle),
            "hafta_kuni": int(source["hafta_kuni"]),
            "smena": int(source["smena"]),
            "dars_raqami": int(source["dars_raqami"]),
            "slot_idlar": sorted(source_ids),
        },
        "tavsiyalar": suggestions[:SAMTM_V19_2_SWAP_SUGGESTION_LIMIT],
        "avtomatik_tavsiya": (
            SAMTM_V19_2_AUTO_SWAP_DEFAULT
            if not mode else bool(mode["avtomatik_tavsiya"])
        ),
        "parallel_ziddiyatlar": _v192_parallel_conflicts(slots),
    }


@app.get("/api/maktab/aqlli_jadval/v3/almashtirish_tavsiyalari")
def v192_swap_suggestions(token: str, urinish_id: int, slot_id: int):
    actor_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor()
    try:
        _v192_tables(cur)
        cur.execute("SELECT maktab_id FROM aqlli_jadval_urinishlari_v2 WHERE id=%s", (urinish_id,))
        run = cur.fetchone()
        if not run:
            raise HTTPException(status_code=404, detail="Jadval topilmadi")
        if not _v1852_staff(cur, actor_id, run["maktab_id"]):
            raise HTTPException(status_code=403, detail="Ruxsat yo'q")
        result = _v192_swap_suggestions(cur, urinish_id, slot_id)
        conn.commit()
        return result
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


class V192SwapApply(BaseModel):
    urinish_id: int
    slot_id: int
    yangi_hafta_kuni: int
    yangi_dars_raqami: int
    turi: str = "almashtirish"


def _v192_clone_run(cur, run, actor_id: int):
    cur.execute("""INSERT INTO aqlli_jadval_urinishlari_v2(
                    maktab_id,holat,yaratgan_user_id,sifat,joylashtirildi,
                    joylashtirilmadi,diagnostika,sozlamalar)
                   VALUES(%s,'draft',%s,%s,%s,%s,%s,%s) RETURNING id""",
                (
                    run["maktab_id"], actor_id, run.get("sifat") or 0,
                    run.get("joylashtirildi") or 0, run.get("joylashtirilmadi") or 0,
                    psycopg2.extras.Json(run.get("diagnostika") or {}),
                    psycopg2.extras.Json({
                        **(run.get("sozlamalar") or {}),
                        "v19_2_manba_urinish_id": int(run["id"]),
                    }),
                ))
    new_run_id = int(cur.fetchone()["id"])
    cur.execute("""INSERT INTO aqlli_jadval_slotlari_v2(
                    urinish_id,maktab_id,sinf_id,hafta_kuni,smena,dars_raqami,
                    fan_nomi,oqituvchi_user_id,guruh_kaliti,xona_id,xona_matni,
                    boshlanish_vaqti,tugash_vaqti,yuklama_id,takror_raqami,hafta_turi)
                   SELECT %s,maktab_id,sinf_id,hafta_kuni,smena,dars_raqami,
                          fan_nomi,oqituvchi_user_id,guruh_kaliti,xona_id,xona_matni,
                          boshlanish_vaqti,tugash_vaqti,yuklama_id,takror_raqami,hafta_turi
                   FROM aqlli_jadval_slotlari_v2 WHERE urinish_id=%s""",
                (new_run_id, run["id"]))
    return new_run_id


class V192SlotRoomUpdate(BaseModel):
    urinish_id: int
    slot_id: int
    xona_id: Optional[int] = None
    xona_matni: Optional[str] = None


@app.put("/api/maktab/aqlli_jadval/v3/slot_xonasi")
def v192_slot_room_update(sorov: V192SlotRoomUpdate, token: str):
    """Jadval katagidagi xonani katalogdan yoki qo'lda xavfsiz tuzatadi."""
    actor_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor()
    try:
        _v192_tables(cur)
        cur.execute(
            "SELECT * FROM aqlli_jadval_urinishlari_v2 WHERE id=%s FOR UPDATE",
            (sorov.urinish_id,),
        )
        original_run = cur.fetchone()
        if not original_run:
            raise HTTPException(status_code=404, detail="Jadval topilmadi")
        if not _v1852_manager(cur, actor_id, original_run["maktab_id"]):
            raise HTTPException(status_code=403, detail="Jadval xonasini faqat rahbariyat o'zgartiradi")

        original_slots = _v192_run_slots(cur, int(original_run["id"]))
        original_slot = next(
            (row for row in original_slots if int(row["id"]) == int(sorov.slot_id)),
            None,
        )
        if not original_slot:
            raise HTTPException(status_code=404, detail="Dars katagi topilmadi")

        run_id = int(original_run["id"])
        slot_id = int(sorov.slot_id)
        if original_run["holat"] == "tasdiqlangan":
            run_id = _v192_clone_run(cur, original_run, actor_id)
            cur.execute(
                """SELECT id FROM aqlli_jadval_slotlari_v2
                   WHERE urinish_id=%s AND sinf_id=%s AND hafta_kuni=%s
                     AND smena=%s AND dars_raqami=%s AND fan_nomi=%s
                     AND guruh_kaliti=%s
                     AND hafta_turi=%s
                     AND COALESCE(oqituvchi_user_id,0)=COALESCE(%s,0)
                     AND COALESCE(takror_raqami,0)=COALESCE(%s,0)
                   ORDER BY id LIMIT 1""",
                (
                    run_id, original_slot["sinf_id"], original_slot["hafta_kuni"],
                    original_slot["smena"], original_slot["dars_raqami"],
                    original_slot["fan_nomi"], original_slot["guruh_kaliti"],
                    original_slot.get("hafta_turi") or "har_hafta",
                    original_slot.get("oqituvchi_user_id"),
                    original_slot.get("takror_raqami"),
                ),
            )
            cloned_slot = cur.fetchone()
            if not cloned_slot:
                raise HTTPException(status_code=409, detail="Yangi draftdagi dars topilmadi")
            slot_id = int(cloned_slot["id"])
        elif original_run["holat"] != "draft":
            raise HTTPException(status_code=409, detail="Faqat draft yoki faol jadval xonasi o'zgartiriladi")

        room_id = int(sorov.xona_id) if sorov.xona_id is not None else None
        room_text = re.sub(r"\s+", " ", str(sorov.xona_matni or "")).strip()[:80] or None
        if room_id is not None:
            cur.execute(
                """SELECT id,nomi FROM aqlli_xonalar_v2
                   WHERE id=%s AND maktab_id=%s AND faol=TRUE AND darsga_yaroqli=TRUE""",
                (room_id, original_run["maktab_id"]),
            )
            selected_room = cur.fetchone()
            if not selected_room:
                raise HTTPException(status_code=400, detail="Tanlangan xona dars o'tishga yaroqli emas")
            room_text = selected_room["nomi"]
        if not room_text:
            cur.execute(
                """SELECT bino,xona
                   FROM maktab_sinflari WHERE id=%s""",
                (original_slot["sinf_id"],),
            )
            class_room = cur.fetchone()
            cur.execute(
                """SELECT id,nomi,faol,darsga_yaroqli FROM aqlli_xonalar_v2
                   WHERE maktab_id=%s""",
                (original_run["maktab_id"],),
            )
            fallback_catalog = [dict(row) for row in cur.fetchall()]
            annotated_home = _v205_annotate_class_home_room(
                dict(class_room or {}), fallback_catalog
            )
            room_id = annotated_home.get("_home_room_id")
            room_text = annotated_home.get("_home_room_text")

        if room_text:
            cur.execute(
                """SELECT id,nomi,faol,darsga_yaroqli FROM aqlli_xonalar_v2
                   WHERE maktab_id=%s""",
                (original_run["maktab_id"],),
            )
            room_catalog = [dict(row) for row in cur.fetchall()]
            room_name_to_id = {
                _v205_room_normalized_name(row.get("nomi")): int(row["id"])
                for row in room_catalog
            }
            matching_catalog_room = next(
                (
                    row for row in room_catalog
                    if _v205_room_normalized_name(row.get("nomi"))
                    == _v205_room_normalized_name(room_text)
                ),
                None,
            )
            if matching_catalog_room:
                if not bool(matching_catalog_room.get("faol")) or not bool(
                    matching_catalog_room.get("darsga_yaroqli")
                ):
                    raise HTTPException(
                        status_code=400,
                        detail="Tanlangan xona dars o'tishga yaroqli emas",
                    )
                room_id = int(matching_catalog_room["id"])
                room_text = str(matching_catalog_room.get("nomi") or room_text)
            selected_room_key = _v205_persisted_room_key(
                room_id, room_text, room_name_to_id
            )
            cur.execute(
                """SELECT e.id,e.xona_id,e.xona_matni,e.hafta_turi,
                          e.smena,e.dars_raqami,e.boshlanish_vaqti,e.tugash_vaqti,
                          COALESCE(r.nomi,e.xona_matni) AS xona_nomi
                   FROM aqlli_jadval_slotlari_v2 e
                   LEFT JOIN aqlli_xonalar_v2 r ON r.id=e.xona_id
                   WHERE e.urinish_id=%s AND e.id<>%s AND e.hafta_kuni=%s""",
                (
                    run_id, slot_id, original_slot["hafta_kuni"],
                ),
            )
            other_slots = [dict(row) for row in cur.fetchall()]
            cur.execute(
                "SELECT * FROM aqlli_smena_sozlamalari_v2 WHERE maktab_id=%s",
                (original_run["maktab_id"],),
            )
            room_update_intervals = {}
            for shift_row in cur.fetchall():
                shift_number = int(shift_row["smena"])
                for time_row in _v1852_shift_slots(shift_row):
                    room_update_intervals[(
                        shift_number, int(time_row["dars_raqami"])
                    )] = time_row
            canonical_time = room_update_intervals.get((
                int(original_slot["smena"]), int(original_slot["dars_raqami"])
            ))
            target_start = _v196_clock_minutes(
                canonical_time.get("boshlanish") if canonical_time
                else original_slot.get("boshlanish_vaqti")
            )
            target_end = _v196_clock_minutes(
                canonical_time.get("tugash") if canonical_time
                else original_slot.get("tugash_vaqti")
            )
            target_phase = str(original_slot.get("hafta_turi") or "har_hafta")
            occupied = False
            for row in other_slots:
                other_key = _v205_persisted_room_key(
                    row.get("xona_id"), row.get("xona_matni"), room_name_to_id
                )
                if not selected_room_key or other_key != selected_room_key:
                    continue
                other_phase = str(row.get("hafta_turi") or "har_hafta")
                if not (
                    target_phase == "har_hafta"
                    or other_phase == "har_hafta"
                    or target_phase == other_phase
                ):
                    continue
                other_time = room_update_intervals.get((
                    int(row.get("smena") or 0), int(row.get("dars_raqami") or 0)
                ))
                other_start = _v196_clock_minutes(
                    other_time.get("boshlanish") if other_time
                    else row.get("boshlanish_vaqti")
                )
                other_end = _v196_clock_minutes(
                    other_time.get("tugash") if other_time
                    else row.get("tugash_vaqti")
                )
                if None not in (target_start, target_end, other_start, other_end):
                    overlaps = int(target_start) < int(other_end) and int(other_start) < int(target_end)
                else:
                    overlaps = (
                        int(row.get("smena") or 0) == int(original_slot["smena"])
                        and int(row.get("dars_raqami") or 0) == int(original_slot["dars_raqami"])
                    )
                if overlaps:
                    occupied = True
                    break
            if occupied:
                raise HTTPException(status_code=409, detail=f"{room_text} bu vaqtda boshqa darsga band")

        cur.execute(
            """UPDATE aqlli_jadval_slotlari_v2
               SET xona_id=%s,xona_matni=%s WHERE id=%s""",
            (room_id, room_text, slot_id),
        )
        conn.commit()
        return {
            "holat": "xona_yangilandi",
            "urinish_id": run_id,
            "slot_id": slot_id,
            "xona_id": room_id,
            "xona": room_text,
            "yangi_draft": run_id != int(original_run["id"]),
        }
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


@app.post("/api/maktab/aqlli_jadval/v3/almashtirish")
def v192_swap_apply(sorov: V192SwapApply, token: str):
    actor_id = _jwt_tekshir(token)
    conn = _db(); cur = conn.cursor()
    try:
        _v192_tables(cur)
        cur.execute("SELECT * FROM aqlli_jadval_urinishlari_v2 WHERE id=%s FOR UPDATE", (sorov.urinish_id,))
        original_run = cur.fetchone()
        if not original_run:
            raise HTTPException(status_code=404, detail="Jadval topilmadi")
        if not _v1852_manager(cur, actor_id, original_run["maktab_id"]):
            raise HTTPException(status_code=403, detail="Jadvalni faqat rahbariyat o'zgartiradi")
        source_slots = _v192_run_slots(cur, original_run["id"])
        original_source = next((row for row in source_slots if int(row["id"]) == int(sorov.slot_id)), None)
        if not original_source:
            raise HTTPException(status_code=404, detail="Dars topilmadi")

        run_id = int(original_run["id"])
        slot_id = int(sorov.slot_id)
        if original_run["holat"] == "tasdiqlangan":
            run_id = _v192_clone_run(cur, original_run, actor_id)
            cur.execute("""SELECT id FROM aqlli_jadval_slotlari_v2
                           WHERE urinish_id=%s AND sinf_id=%s AND hafta_kuni=%s
                             AND smena=%s AND dars_raqami=%s AND fan_nomi=%s
                             AND guruh_kaliti=%s LIMIT 1""",
                        (
                            run_id, original_source["sinf_id"], original_source["hafta_kuni"],
                            original_source["smena"], original_source["dars_raqami"],
                            original_source["fan_nomi"], original_source["guruh_kaliti"],
                        ))
            slot_id = int(cur.fetchone()["id"])
        elif original_run["holat"] != "draft":
            raise HTTPException(status_code=409, detail="Faqat draft yoki faol jadval o'zgartiriladi")

        report = _v192_swap_suggestions(cur, run_id, slot_id)
        candidate = next((
            row for row in report["tavsiyalar"]
            if int(row["yangi_hafta_kuni"]) == int(sorov.yangi_hafta_kuni)
            and int(row["yangi_dars_raqami"]) == int(sorov.yangi_dars_raqami)
            and str(row["turi"]) == str(sorov.turi)
        ), None)
        if not candidate:
            raise HTTPException(
                status_code=409,
                detail="Tanlangan joy endi xavfsiz emas. Tavsiyalarni qayta yuklang",
            )
        slots = _v192_run_slots(cur, run_id)
        source, source_bundle = _v192_bundle_for_slot(slots, slot_id)
        source_ids = [int(row["id"]) for row in source_bundle]
        target_bundle = [
            row for row in slots
            if int(row["sinf_id"]) == int(source["sinf_id"])
            and int(row["hafta_kuni"]) == int(candidate["yangi_hafta_kuni"])
            and int(row["smena"]) == int(source["smena"])
            and int(row["dars_raqami"]) == int(candidate["yangi_dars_raqami"])
        ]
        target_ids = [int(row["id"]) for row in target_bundle]
        old_day = int(source["hafta_kuni"])
        old_period = int(source["dars_raqami"])
        projected_changes = {
            int(row["id"]): (
                int(candidate["yangi_hafta_kuni"]),
                int(candidate["yangi_dars_raqami"]),
            )
            for row in source_bundle
        }
        projected_changes.update({
            int(row["id"]): (old_day, old_period)
            for row in target_bundle
        })
        projected_core_period6_days = _v213_projected_core_period6_days(
            slots, int(source["sinf_id"]), projected_changes
        )
        if len(projected_core_period6_days) > _V213_CORE_PERIOD6_LIMIT:
            raise HTTPException(
                status_code=409,
                detail=(
                    f"{_V213_CORE_PERIOD6_REASON}. O'zgarish bekor qilindi"
                ),
            )
        cur.execute("""UPDATE aqlli_jadval_slotlari_v2
                       SET hafta_kuni=7 WHERE id=ANY(%s)""", (source_ids,))
        if target_ids:
            cur.execute("""UPDATE aqlli_jadval_slotlari_v2
                           SET hafta_kuni=%s,dars_raqami=%s WHERE id=ANY(%s)""",
                        (old_day, old_period, target_ids))
        cur.execute("""UPDATE aqlli_jadval_slotlari_v2
                       SET hafta_kuni=%s,dars_raqami=%s WHERE id=ANY(%s)""",
                    (
                        int(candidate["yangi_hafta_kuni"]),
                        int(candidate["yangi_dars_raqami"]),
                        source_ids,
                    ))
        after_slots = _v192_run_slots(cur, run_id)
        parallel_conflicts = _v192_parallel_conflicts(after_slots)
        if parallel_conflicts:
            raise HTTPException(
                status_code=409,
                detail="O'qituvchi parallel darsga tushib qoldi; o'zgarish bekor qilindi",
            )
        core_period6_days = _v213_projected_core_period6_days(
            after_slots, int(source["sinf_id"])
        )
        if len(core_period6_days) > _V213_CORE_PERIOD6_LIMIT:
            raise HTTPException(
                status_code=409,
                detail=(
                    f"{_V213_CORE_PERIOD6_REASON}. O'zgarish bekor qilindi"
                ),
            )
        cur.execute("""INSERT INTO aqlli_jadval_ozgarish_log_v19_2(
                        maktab_id,urinish_id,manba_slot_id,eski_hafta_kuni,
                        eski_dars_raqami,yangi_hafta_kuni,yangi_dars_raqami,
                        turi,bajargan_user_id)
                       VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                    (
                        original_run["maktab_id"], run_id, slot_id, old_day, old_period,
                        candidate["yangi_hafta_kuni"], candidate["yangi_dars_raqami"],
                        candidate["turi"], actor_id,
                    ))
        exact = _v1875_schedule_integrity_report(cur, original_run["maktab_id"], run_id)
        hygiene = _v1874_schedule_hygiene_violations(cur, original_run["maktab_id"], run_id)
        class_blocks = _v1856_schedule_block_violations(cur, original_run["maktab_id"], run_id)
        diagnostic = {
            "tasdiqlash_mumkin": bool(exact.get("tayyor") and not hygiene and not class_blocks),
            "jadval_mosligi": exact,
            "gigiyena_xatolari": hygiene,
            "sinf_kun_xatolari": class_blocks,
            "v19_2_ozgartirish": {
                "turi": candidate["turi"],
                "eski": [old_day, old_period],
                "yangi": [candidate["yangi_hafta_kuni"], candidate["yangi_dars_raqami"]],
            },
        }
        cur.execute("""UPDATE aqlli_jadval_urinishlari_v2
                       SET diagnostika=%s WHERE id=%s""",
                    (psycopg2.extras.Json(diagnostic), run_id))
        conn.commit()
        return {
            "holat": "almashtirildi" if target_ids else "ko'chirildi",
            "urinish_id": run_id,
            "yangi_draft": run_id != int(original_run["id"]),
            "tasdiqlash_mumkin": diagnostic["tasdiqlash_mumkin"],
            "diagnostika": diagnostic,
        }
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


# ========================= V19.2 END =========================

# ═══════════════════════════════════════════════════════════
# V19.6 — PEDAGOGIK VA O'QITUVCHIGA QULAY JOYLASHTIRISH
# 0,5 fanlar A/B haftada aniq ko'rinadi. Generator sinf yoshiga mos
# dars vaqtini, og'ir/yengil fan almashuvini, jismoniy tarbiyadan
# keyingi tiklanishni va o'qituvchining oknosiz ixcham ish kunini
# birgalikda ballaydi.
# ═══════════════════════════════════════════════════════════

_v196_base_build_jobs = _v1874_build_jobs
_v196_base_candidate_reasons = _v1874_candidate_reasons
_v196_base_place_job = _v1874_place_job


def _v196_rotation_profile(job, context=None):
    """A/B slotning ikkala faniga ham mos yagona vaqt profilini qaytaradi."""
    profiles = []
    members = job.get("rotation_members") or []
    class_row = (context or {}).get("classes", {}).get(job.get("sinf_id"), {})
    grade = int(job.get("v1874_grade") or _v1874_grade(class_row))
    for member in members:
        profiles.append(
            member.get("v1874_profile")
            or _v1874_subject_profile(member.get("fan"), grade)
        )
    if not profiles:
        return _v1874_profile_for_job(job, context or {})

    # Eng talabchan a'zo slot vaqtini belgilaydi. "light" faqat ikkala fan
    # ham yengil bo'lganda rost: og'ir + yengil juftlik ertaroq joylashadi.
    most_difficult = max(profiles, key=lambda item: int(item.get("difficulty") or 0))
    return {
        **most_difficult,
        "key": " / ".join(str(profile.get("key") or "") for profile in profiles),
        "academic": any(bool(profile.get("academic")) for profile in profiles),
        "physical": any(bool(profile.get("physical")) for profile in profiles),
        "technology": any(bool(profile.get("technology")) for profile in profiles),
        "light": all(bool(profile.get("light")) for profile in profiles),
        "primary_light": all(bool(profile.get("primary_light")) for profile in profiles),
        "primary_core": any(bool(profile.get("primary_core")) for profile in profiles),
        "heavy": any(bool(profile.get("heavy")) for profile in profiles),
        "written_heavy": any(bool(profile.get("written_heavy")) for profile in profiles),
        "math": any(bool(profile.get("math")) for profile in profiles),
        "native_language": any(bool(profile.get("native_language")) for profile in profiles),
        "core_science": any(bool(profile.get("core_science")) for profile in profiles),
        "core_priority": any(bool(profile.get("core_priority")) for profile in profiles),
        "difficulty": max(int(profile.get("difficulty") or 0) for profile in profiles),
    }


def _v220_fractional_member_domain(member, context):
    """Use the exact adapter itself as the canonical hard-domain source."""

    if not _V220DefaultTimetableAdapter or not _v220_candidate_hard_violations:
        return frozenset()
    rows = _V220DefaultTimetableAdapter().build_candidates(member, context)
    return frozenset(
        (int(row["day"]), int(row["period"]))
        for row in rows
        if not _v220_candidate_hard_violations(row, member, context)
    )


def _v220_fractional_pair_plan(domains):
    """Maximize compatible A/B pairs, then their shared-domain size."""

    domains = tuple(frozenset(domain or ()) for domain in domains)
    size = len(domains)
    if size > 16:
        # Defensive linear fallback for an abnormally large single class.
        remaining = list(range(size))
        plan = []
        while remaining:
            first = remaining.pop(0)
            candidates = [
                (len(domains[first] & domains[partner]), -partner, partner)
                for partner in remaining
            ]
            overlap, _tie_breaker, partner = max(
                candidates, default=(0, 0, -1)
            )
            if overlap > 0:
                remaining.remove(partner)
                plan.append((first, partner))
            else:
                plan.append((first,))
        return tuple(plan)

    memo = {}

    def solve(remaining):
        if not remaining:
            return 0, 0, ()
        if remaining in memo:
            return memo[remaining]
        first = remaining[0]
        child = solve(remaining[1:])
        best = (child[0], child[1], ((first,),) + child[2])
        for position, partner in enumerate(remaining[1:], start=1):
            overlap = len(domains[first] & domains[partner])
            if overlap <= 0:
                continue
            rest = remaining[1:position] + remaining[position + 1:]
            child = solve(rest)
            candidate = (
                child[0] + 1,
                child[1] + overlap,
                ((first, partner),) + child[2],
            )
            if candidate[:2] > best[:2] or (
                candidate[:2] == best[:2] and candidate[2] < best[2]
            ):
                best = candidate
        memo[remaining] = best
        return best

    return solve(tuple(range(size)))[2]


def _v220_repair_fractional_pairs(jobs, context):
    """Re-pair 0.5-hour members only when their hard domains intersect."""

    regular_jobs = []
    half_by_class = _v1852_defaultdict(list)
    for job in jobs:
        members = list(job.get("rotation_members") or [])
        if not members:
            regular_jobs.append(job)
            continue
        for member in members:
            signature = (
                int(member.get("sinf_id") or job.get("sinf_id") or 0),
                int(member.get("smena") or job.get("smena") or 1),
                tuple(
                    str(group.get("guruh_kaliti") or "whole")
                    for group in member.get("groups") or []
                ),
            )
            half_by_class[signature].append(member)

    paired_count = 0
    standalone_count = 0
    for signature, members in sorted(half_by_class.items(), key=lambda item: item[0]):
        ordered = sorted(
            members,
            key=lambda item: (
                str(item.get("fan") or "").casefold(),
                int(item.get("load_id") or 0),
                str(item.get("job_id") or ""),
            ),
        )
        domains = [
            _v220_fractional_member_domain(member, context)
            for member in ordered
        ]
        plan = _v220_fractional_pair_plan(domains)
        single_phase = 0
        for indices in plan:
            selected = [ordered[index] for index in indices]
            if len(selected) == 2:
                phases = ("toq", "juft")
                paired_count += 1
            else:
                phases = (("toq", "juft")[single_phase % 2],)
                single_phase += 1
                standalone_count += 1
            rotation_members = [
                {**member, "hafta_turi": phases[index]}
                for index, member in enumerate(selected)
            ]
            first = selected[0]
            combined = {
                **first,
                "job_id": "rotation:" + ":".join(
                    str(member.get("job_id")) for member in selected
                ),
                "fan": " / ".join(str(member.get("fan") or "") for member in selected),
                "groups": [],
                "teacher_options": [],
                "room_id": None,
                "rotation_weight": 1.0,
                "rotation_members": rotation_members,
                "difficulty": max(
                    float(member.get("difficulty") or 0) for member in selected
                ) + 25,
                "weight": max(int(member.get("weight") or 1) for member in selected),
                "preferred_last": min(
                    int(member.get("preferred_last") or 5) for member in selected
                ),
            }
            combined["v1874_profile"] = _v196_rotation_profile(combined, context)
            regular_jobs.append(combined)

    warnings = []
    if paired_count or standalone_count:
        warnings.append(
            f"A/B hafta: {paired_count} ta 0,5+0,5 juft umumiy legal katak "
            f"bo'yicha tanlandi; {standalone_count} ta 0,5 dars qattiq "
            "qoidani buzmaslik uchun alohida hafta fazasida qoldi"
        )
    return regular_jobs, warnings


def _v196_candidate_reasons(
    job, day, period, selected_teachers, room_keys, state, context
):
    reasons = list(_v196_base_candidate_reasons(
        job, day, period, selected_teachers, room_keys, state, context
    ))
    # Oddiy urinishda fan kuniga bir marta qo‘yiladi. Jadvalning hamma darsi
    # sig‘may qolsa, generatorning zaxira urinishigina bitta sinfda haftasiga
    # avval 1 kun, mutlaqo iloj bo‘lmasa 2 kungacha shu fanni ikki marta
    # qo‘yishi mumkin. O‘qituvchining ayni vaqtdagi bandligi filtrlanmaydi.
    emergency_repeat_days = int(context.get("v203_emergency_repeat_days") or 0)
    if emergency_repeat_days and "fan kunlik maksimumga yetgan" in reasons:
        subject_counts = state.get("subject_daily", {})
        subject_key = str(job.get("fan") or "").casefold()
        current_count = int(subject_counts.get((job["sinf_id"], subject_key, day), 0))
        already_repeat_days = {
            int(subject_day)
            for (class_id, _subject, subject_day), count in subject_counts.items()
            if int(class_id) == int(job["sinf_id"])
            and str(_subject) == subject_key
            and int(count or 0) >= 2
        }
        may_use_day = int(day) in already_repeat_days or len(already_repeat_days) < emergency_repeat_days
        if current_count < 2 and may_use_day:
            reasons = [reason for reason in reasons if reason not in {
                "fan kunlik maksimumga yetgan",
                "boshlang‘ich sinfda bir fan shu kuni takror qo‘yilmaydi",
            }]
    profile = _v196_rotation_profile(job, context)
    if int(period) == 1 and (
        profile.get("physical") or profile.get("technology")
    ):
        reasons.append(
            "jismoniy tarbiya va texnologiya 1-darsga qo'yilmaydi"
        )

    # V22.48: post-processing 4–5 soatli fanni 2+2(+1) ko'rinishida
    # yig'ishi mumkin, ammo bir fan uchun juft darsli kunlar soni exact
    # validator limitidan oshmaydi. Bu o'qituvchi oynosini tuzatib bo'lgach
    # butun natijaning rollback bo'lishini oldini oladi.
    subject_key = str(job.get("fan") or "").casefold()
    class_id = int(job.get("sinf_id") or 0)
    current_count = int((state.get("subject_daily", {}) or {}).get(
        (class_id, subject_key, int(day)), 0
    ))
    profile = _v196_rotation_profile(job, context)
    repeat_limit = int(
        (context or {}).get("practical_repeat_day_limit", 1)
        if (profile.get("physical") or profile.get("technology"))
        else (context or {}).get("max_subject_repeat_days", 2)
    )
    paired_days = {
        int(subject_day)
        for (subject_class, subject, subject_day), count
        in (state.get("subject_daily", {}) or {}).items()
        if int(subject_class) == class_id
        and str(subject) == subject_key
        and int(count or 0) >= 2
    }
    if (
        current_count == 1
        and int(day) not in paired_days
        and len(paired_days) >= max(0, repeat_limit)
        and "fan kunlik maksimumga yetgan" not in reasons
    ):
        reasons.append("fan takror kuni limiti to'lgan")

    # Qurish vaqtida vaqtinchalik sinf oknosini qattiq taqiqlamaymiz. Aks
    # holda navbatda oldin kelgan ikki-o'qituvchili Chet tili 3-darsga bo'sh
    # bo'lsa ham, 1–2-dars hali joylashmagani uchungina nomzodsiz qolardi.
    # Nomzod ballida okno 4350 ball bilan keskin jazolanadi; qurish tugagach
    # _v196_compact_class_gaps uni yopadi va saqlashdan oldingi qattiq
    # tekshiruv bitta ham sinf oknosi qolgan draftni qabul qilmaydi.
    return list(dict.fromkeys(reasons))


def _v1852_build_jobs(classes, loads, assignments, group_settings, teachers):
    jobs, warnings = _v196_base_build_jobs(
        classes, loads, assignments, group_settings, teachers
    )
    rotation_count = 0
    for job in jobs:
        grade = _v1874_grade(classes.get(int(job.get("sinf_id") or 0), {}))
        # V22.48: barcha fanlarda bir sinf-kun uchun limit 2; 4–5 soatli fan 2+2(+1) bo'lishi mumkin.
        # Bu majburiy juft dars emas; faqat 3 legal kun ichiga 5 soat kabi
        # yuklamani 2+2+1 ko'rinishida sig'dirish imkonidir.
        job["daily_max"] = 2
        for member in job.get("rotation_members") or []:
            member["daily_max"] = 2
        if job.get("rotation_members"):
            rotation_count += 1
            job["v1874_profile"] = _v196_rotation_profile(
                job, {"classes": classes}
            )
    warnings.append(
        "V19.8 pedagogik strategiya faol: ona tili, adabiyot, matematika, "
        "algebra, geometriya, fizika, kimyo va biologiya 1–5-darsga; faqat "
        "boshqa legal katak qolmasa 6-darsga (bir sinfda haftasiga ko'pi bilan "
        "2 kun); jismoniy tarbiya 3–6-darsga ustuvor. O'qituvchining ichki "
        "oknosi va ikki smena orasidagi uzoq kutish birgalikda kamaytiriladi"
    )
    if rotation_count:
        warnings.append(
            f"A/B ko'rinishi: {rotation_count} ta jadval katagida 0,5 fanlar "
            "TOQ/JUFT hafta yorlig'i bilan almashadi"
        )
    return jobs, warnings




def _v196_teacher_shift_demand(jobs):
    """O'qituvchining qaysi smenalarda darsi borligini oldindan hisoblaydi."""
    result = _v1852_defaultdict(lambda: _v1852_defaultdict(float))
    for job in jobs:
        shift = int(job.get("smena") or 1)
        members = job.get("rotation_members") or []
        source_jobs = members or [job]
        contribution = 0.5 if members else 1.0
        for source in source_jobs:
            teachers = _v1852_job_teacher_ids(source)
            for teacher in teachers:
                result[int(teacher)][shift] += contribution
    return {teacher: dict(shifts) for teacher, shifts in result.items()}


def _v196_shift_max_period(context, shift):
    slots = (context.get("shifts", {}).get(int(shift), {}) or {}).get("slotlar") or []
    return max([int(slot.get("dars_raqami") or 0) for slot in slots] or [6])


def _v196_clock_minutes(value):
    text = _v1852_time_str(value)
    try:
        hour, minute = text.split(":", 1)
        return int(hour) * 60 + int(minute[:2])
    except (TypeError, ValueError, AttributeError):
        return None


def _v196_slot_interval(context, shift, period):
    slots = (context.get("shifts", {}).get(int(shift), {}) or {}).get("slotlar") or []
    row = next(
        (slot for slot in slots if int(slot.get("dars_raqami") or 0) == int(period)),
        None,
    )
    if not row:
        return None
    start = _v196_clock_minutes(row.get("boshlanish"))
    end = _v196_clock_minutes(row.get("tugash"))
    return (start, end) if start is not None and end is not None else None


def _v196_cross_shift_gap_minutes(state, teacher, day, context, extra=None):
    """1-smena oxiri bilan 2-smena boshi orasidagi haqiqiy daqiqalar."""
    by_shift = {}
    for shift in (1, 2):
        by_shift[shift] = set(
            int(value) for value in state.get("teacher_periods", {}).get(
                (int(teacher), int(day), shift), set()
            )
        )
    if extra:
        shift, period = extra
        by_shift.setdefault(int(shift), set()).add(int(period))
    if not by_shift.get(1) or not by_shift.get(2):
        return None
    first_intervals = [
        _v196_slot_interval(context, 1, period) for period in by_shift[1]
    ]
    second_intervals = [
        _v196_slot_interval(context, 2, period) for period in by_shift[2]
    ]
    first_intervals = [row for row in first_intervals if row]
    second_intervals = [row for row in second_intervals if row]
    if not first_intervals or not second_intervals:
        return None
    latest_first_end = max(row[1] for row in first_intervals)
    earliest_second_start = min(row[0] for row in second_intervals)
    return int(max(0, earliest_second_start - latest_first_end))


def _v200_teacher_day_timeline(state, teacher, day, context, extra=None):
    """Ikki smenani bitta haqiqiy vaqt chizig'iga birlashtiradi."""
    intervals = set()
    for shift in (1, 2):
        periods = state.get("teacher_periods", {}).get(
            (int(teacher), int(day), shift), set()
        )
        for period in periods:
            interval = _v196_slot_interval(context, shift, period)
            if interval:
                intervals.add((int(interval[0]), int(interval[1]), shift, int(period)))
    if extra:
        shift, period = int(extra[0]), int(extra[1])
        interval = _v196_slot_interval(context, shift, period)
        if interval:
            intervals.add((int(interval[0]), int(interval[1]), shift, period))
    return sorted(intervals, key=lambda row: (row[0], row[1], row[2], row[3]))


def _v200_teacher_day_idle(state, teacher, day, context, extra=None):
    """Oddiy tanaffusdan katta bo'sh kutishni haqiqiy daqiqada qaytaradi.

    25 daqiqagacha bo'lgan odatiy tanaffus okno emas. Bir yoki ikki smena
    alohida sanalmaydi: ustozning ertalabki birinchi darsidan kechki oxirgi
    darsigacha bo'lgan bitta real ish kuni tekshiriladi.
    """
    timeline = _v200_teacher_day_timeline(
        state, teacher, day, context, extra=extra
    )
    gaps = []
    for previous, current in zip(timeline, timeline[1:]):
        minutes = max(0, int(current[0]) - int(previous[1]))
        if minutes > 25:
            gaps.append(minutes)
    return {
        "jami_daqiqa": int(sum(gaps)),
        "eng_katta_daqiqa": int(max(gaps or [0])),
        "okno_soni": int(len(gaps)),
        "ikki_soatdan_uzoq": int(sum(1 for value in gaps if value > 120)),
        "uch_soatdan_uzoq": int(sum(1 for value in gaps if value > 180)),
    }


def _v200_all_teacher_idle_signature(state, context):
    days = {
        (int(teacher), int(day))
        for (teacher, day), count in state.get("teacher_daily", {}).items()
        if float(count or 0) > 0
    }
    profiles = [
        _v200_teacher_day_idle(state, teacher, day, context)
        for teacher, day in days
    ]
    return (
        sum(row["uch_soatdan_uzoq"] for row in profiles),
        sum(row["ikki_soatdan_uzoq"] for row in profiles),
        max([row["eng_katta_daqiqa"] for row in profiles] or [0]),
        sum(row["jami_daqiqa"] for row in profiles),
        sum(row["okno_soni"] for row in profiles),
    )




def _v196_cross_shift_edge_blocks(state, teacher, day, context, extra=None):
    """Ikki smena orasidagi oldini olish mumkin bo'lgan bo'sh bloklar.

    1-smena oxiridagi bo'sh darslar + 2-smena boshidagi bo'sh darslar olinadi.
    Masalan, 1-smena 1–3 va 2-smena 4–5 bo'lsa 3+3=6 blok; 1-smena
    4–6 va 2-smena 1–2 bo'lsa 0 blok. Faqat ikkala smena shu kunda faol
    bo'lgandagina qiymat qaytariladi.
    """
    by_shift = {}
    for shift in (1, 2):
        by_shift[shift] = set(
            int(value) for value in state.get("teacher_periods", {}).get(
                (int(teacher), int(day), shift), set()
            )
        )
    if extra:
        shift, period = extra
        by_shift.setdefault(int(shift), set()).add(int(period))
    if not by_shift.get(1) or not by_shift.get(2):
        return None
    return int(
        max(0, _v196_shift_max_period(context, 1) - max(by_shift[1]))
        + max(0, min(by_shift[2]) - 1)
    )


def _v196_teacher_demand(jobs):
    demand = _v1852_defaultdict(float)
    for job in jobs:
        members = job.get("rotation_members") or []
        if members:
            for member in members:
                teachers = [
                    teacher for teacher in _v199_rotation_member_teachers(member)
                    if teacher is not None
                ]
                for teacher in set(teachers):
                    demand[int(teacher)] += 0.5
            continue
        if job.get("groups"):
            for teacher in {
                group.get("teacher") for group in job.get("groups") or []
                if group.get("teacher") is not None
            }:
                demand[int(teacher)] += 1.0
            continue
        options = [teacher for teacher in job.get("teacher_options") or [] if teacher is not None]
        share = 1.0 / max(1, len(options))
        for teacher in options:
            demand[int(teacher)] += share
    return dict(demand)




def _v196_teacher_target_days(demand, rules):
    """Haftalik yuklamani iloji boricha 1–4 ixcham ish kuniga yig'adi.

    V22.48: 1 soat tabiiy ravishda 1 kunda qoladi. 2–6 soatli yuklama 2
    kunga, 7–10 soat 3 kunga, 11–18 soat esa 3–4 kunga yig'iladi. Kunlik
    maksimum qattiq qoida bo'lib qoladi; maqsad sig'masa minimum kun soni
    avtomatik oshadi.
    """
    demand = float(demand or 0)
    daily_limit = max(1, min(6, int((rules or {}).get("kunlik_max") or 6)))
    minimum = max(1, int(math.ceil(demand / daily_limit)))
    if demand <= 1:
        return 1
    if demand <= 6:
        return max(2, minimum)
    if demand <= 10:
        return max(3, minimum)
    if demand <= 15:
        return max(3, minimum)
    if demand <= 18:
        return max(4, minimum)
    compact_capacity = max(1, min(5, daily_limit))
    return max(minimum, int(math.ceil(demand / compact_capacity)))

def _v201_teacher_fallback_days(demand, rules):
    """Maqsad sig'masa ruxsat etiladigan birinchi zaxira kun soni."""
    demand = float(demand or 0)
    target = _v196_teacher_target_days(demand, rules)
    if demand <= 2:
        return 1
    if demand <= 6:
        return max(target, 3)
    if demand <= 10:
        return max(target, 4)
    if demand <= 15:
        return max(target, 4)
    if demand < 20:
        return max(target, 5)
    return target


def _v196_class_distribution(jobs, context):
    """Har bir sinf uchun 4/5/6 darsli barqaror kun maqsadini tayyorlaydi."""
    totals = _v1852_Counter(int(job["sinf_id"]) for job in jobs)
    result = {}
    weekdays = int(context.get("weekdays") or 6)
    for class_id, total in totals.items():
        class_row = context.get("classes", {}).get(class_id, {})
        grade = _v1874_grade(class_row)
        allowed = []
        for day in range(1, weekdays + 1):
            if 1 <= grade <= 4 and day == 6:
                continue
            if _v1856_class_day_block_reason(
                class_row, day, context.get("class_day_blocks", {})
            ):
                continue
            allowed.append(day)
        if not allowed:
            continue
        low, remainder = divmod(int(total), len(allowed))
        result[class_id] = {
            "total": int(total),
            "days": tuple(allowed),
            "low": int(low),
            "high": int(low + (1 if remainder else 0)),
            "remainder": int(remainder),
        }
    return result


def _v196_class_distribution_metrics(state, context):
    imbalance = 0
    short_days = 0
    for class_id, target in (context.get("v196_class_distribution") or {}).items():
        days = list(target["days"])
        actual = sorted(
            int(state.get("class_daily_total", {}).get((class_id, day), 0))
            for day in days
        )
        ideal = sorted(
            [target["low"]] * (len(days) - target["remainder"])
            + [target["high"]] * target["remainder"]
        )
        imbalance += sum(abs(a - b) for a, b in zip(actual, ideal))
        short_days += sum(1 for count in actual if 0 < count < target["low"])
    return int(imbalance), int(short_days)




def _v1852_place_job(job, day, period, teachers, room_keys, state, context):
    _v196_base_place_job(job, day, period, teachers, room_keys, state, context)
    teacher_jobs = state.setdefault(
        "v196_teacher_period_jobs", _v1852_defaultdict(dict)
    )
    for teacher in teachers:
        if teacher is not None:
            teacher_jobs[(int(teacher), int(day), int(job.get("smena") or 1))][int(period)] = job


def _v196_class_gap_count(state):
    """Sinf kunidagi bosh va ichki bo'sh darslarning aniq soni."""
    return int(sum(
        max(set(period_jobs.keys())) - len(set(period_jobs.keys()))
        if period_jobs else 0
        for period_jobs in state.get("class_period_jobs", {}).values()
    ))


def _v196_movable_placement(placement):
    job = placement.get("job") or {}
    return not (
        job.get("is_class_hour")
        or int(job.get("fixed_day") or 0)
        or int(job.get("fixed_period") or 0)
    )


def _v196_place_exact(
    job, day, period, state, context, rng, expected_teachers=None
):
    candidates, _ = _v1852_open_candidates(
        job, state, context, exact=(int(day), int(period))
    )
    if expected_teachers is not None:
        expected = {
            int(teacher) for teacher in expected_teachers
            if teacher is not None
        }
        candidates = [
            candidate for candidate in candidates
            if {
                int(teacher) for teacher in candidate[3]
                if teacher is not None
            } == expected
        ]
    if not candidates:
        return False
    _, target_day, target_period, teachers, rooms = candidates[0]
    _v1852_place_job(
        job, target_day, target_period, teachers, rooms, state, context
    )
    return True


def _v206_deadline_reached(context):
    context = context or {}
    deadline = float(context.get("v206_deadline") or 0)
    now = _samtm_time.monotonic()
    if deadline and now >= deadline:
        return True
    if context.get("v2244_cancelled"):
        return True
    # Ichki optimizerlar bu funksiyani minglab marta chaqiradi. DB signalini
    # har iteratsiyada emas, 0.25 soniyada bir tekshirib, To'xtatishni tez va
    # bazaga ortiqcha yuk bermasdan ishlatamiz.
    cancel_requested = context.get("v2244_cancel_requested")
    if callable(cancel_requested):
        next_check = float(context.get("v2244_cancel_next_check") or 0)
        if now >= next_check:
            context["v2244_cancel_next_check"] = now + 0.25
            try:
                if cancel_requested():
                    context["v2244_cancelled"] = True
                    return True
            except Exception:
                pass
    return False


def _v196_compact_class_gaps(state, context, rng, max_moves=32):
    """1–2–3–bo'sh–5 ni 1–2–3–4 ko'rinishiga xavfsiz siqadi."""
    moves = 0
    while moves < int(max_moves):
        if _v206_deadline_reached(context):
            break
        before_gap = _v196_class_gap_count(state)
        if before_gap <= 0:
            break
        changed = False
        class_days = sorted(
            state.get("class_period_jobs", {}).items(),
            key=lambda item: (
                -(max(item[1]) - len(item[1]) if item[1] else 0),
                int(item[0][0]), int(item[0][1]),
            ),
        )
        for (class_id, day), period_jobs in class_days:
            if _v206_deadline_reached(context):
                break
            periods = sorted(int(value) for value in period_jobs)
            if not periods:
                continue
            missing = next(
                (value for value in range(1, max(periods)) if value not in period_jobs),
                None,
            )
            if missing is None:
                continue
            donors = sorted(
                [
                    placement for placement in state.get("placements", [])
                    if int(placement["job"].get("sinf_id") or 0) == int(class_id)
                    and int(placement.get("day") or 0) == int(day)
                    and int(placement.get("period") or 0) > int(missing)
                    and _v196_movable_placement(placement)
                ],
                key=lambda placement: int(placement.get("period") or 0),
                reverse=True,
            )
            for donor in donors:
                donor_id = id(donor)
                trial = _v1852_rebuild_schedule_state(
                    [
                        placement for placement in state.get("placements", [])
                        if id(placement) != donor_id
                    ],
                    context,
                )
                if not _v196_place_exact(
                    donor["job"], day, missing, trial, context, rng
                ):
                    continue
                if _v196_class_gap_count(trial) >= before_gap:
                    continue
                state = trial
                moves += 1
                changed = True
                break
            if changed:
                break
        if not changed:
            break
    state["class_gap_count"] = _v196_class_gap_count(state)
    return state


def _v226_frozen_class_day_signature(state):
    return {
        (int(class_id), int(day)): int(count or 0)
        for (class_id, day), count
        in (state.get("class_daily_total", {}) or {}).items()
        if int(count or 0) > 0
    }


def _v226_class_day_counts_match(state, context):
    frozen = (context or {}).get("v226_frozen_class_day_counts")
    if not frozen:
        return True
    current = _v226_frozen_class_day_signature(state)
    normalized = {
        (int(key[0]), int(key[1])): int(value or 0)
        for key, value in frozen.items()
        if int(value or 0) > 0
    }
    return current == normalized


def _v196_balance_class_days(state, context, rng, max_moves=24):
    """Boshlang'ich exact jadvaldan keyin sinfning kunlik soati QOTIRILADI.

    O'qituvchi oynosini tuzatish birinchi exact bosqich yaratgan teng
    sinf-kun soatlarini boshqa kunga ko'chira olmaydi. Keyingi qulaylashtirish
    faqat shu kataklar ichida
    fanlarni almashtiradi yoki ikki kun o'rtasida 1:1 swap qiladi.
    """
    if (context or {}).get("v226_frozen_class_day_counts"):
        state["v226_class_day_counts_frozen"] = True
        return state

    moves = 0
    while moves < int(max_moves):
        if _v206_deadline_reached(context):
            break
        before_imbalance, _ = _v196_class_distribution_metrics(state, context)
        if before_imbalance <= 0:
            break
        changed = False
        for class_id, target in sorted(
            (context.get("v196_class_distribution") or {}).items()
        ):
            counts = {
                int(day): int(
                    state.get("class_daily_total", {}).get((class_id, day), 0)
                )
                for day in target["days"]
            }
            recipients = sorted(
                [day for day, count in counts.items() if count < int(target["low"])],
                key=lambda day: (counts[day], day),
            )
            donors = sorted(
                [day for day, count in counts.items() if count > int(target["high"])],
                key=lambda day: (-counts[day], day),
            )
            if not recipients:
                recipients = sorted(
                    [day for day in counts if counts[day] == min(counts.values())]
                )
            if not donors:
                donors = sorted(
                    [day for day in counts if counts[day] == max(counts.values())]
                )
            for recipient in recipients:
                for donor_day in donors:
                    if counts[donor_day] - counts[recipient] < 2:
                        continue
                    recipient_periods = state.get("class_period_jobs", {}).get(
                        (class_id, recipient), {}
                    )
                    target_period = len(recipient_periods) + 1
                    if target_period in recipient_periods:
                        continue
                    donor_placements = sorted(
                        [
                            placement for placement in state.get("placements", [])
                            if int(placement["job"].get("sinf_id") or 0) == int(class_id)
                            and int(placement.get("day") or 0) == int(donor_day)
                            and _v196_movable_placement(placement)
                        ],
                        key=lambda placement: int(placement.get("period") or 0),
                        reverse=True,
                    )
                    for donor in donor_placements:
                        donor_id = id(donor)
                        trial = _v1852_rebuild_schedule_state(
                            [
                                placement for placement in state.get("placements", [])
                                if id(placement) != donor_id
                            ],
                            context,
                        )
                        if not _v196_place_exact(
                            donor["job"], recipient, target_period,
                            trial, context, rng,
                        ):
                            continue
                        trial = _v196_compact_class_gaps(
                            trial, context, rng, max_moves=4
                        )
                        after_imbalance, _ = _v196_class_distribution_metrics(
                            trial, context
                        )
                        if after_imbalance >= before_imbalance:
                            continue
                        if _v196_class_gap_count(trial) > _v196_class_gap_count(state):
                            continue
                        state = trial
                        moves += 1
                        changed = True
                        break
                    if changed:
                        break
                if changed:
                    break
            if changed:
                break
        if not changed:
            break
    return state

def _v196_teacher_gap_metrics(state, context):
    """O'qituvchining smena ichidagi oynalarini kunlar kesimida sanaydi."""
    total = 0
    gap_shift_days = 0
    multi_gap_shift_days = 0
    max_gap = 0
    gap_rows = []
    gap_days_by_teacher = _v1852_defaultdict(set)
    for key, periods in state.get("teacher_periods", {}).items():
        teacher_id, day, shift = key
        ordered_periods = sorted({int(value) for value in periods})
        gap = int(_v1852_gap_count(set(periods)))
        total += gap
        max_gap = max(max_gap, gap)
        if gap > 0:
            gap_shift_days += 1
            missing_periods = [
                period for period in range(
                    min(ordered_periods), max(ordered_periods) + 1
                )
                if period not in set(ordered_periods)
            ]
            gap_days_by_teacher[int(teacher_id)].add(int(day))
            gap_rows.append({
                "oqituvchi_id": int(teacher_id),
                "kun": int(day),
                "smena": int(shift),
                "darslar": ordered_periods,
                "bosh_darslar": missing_periods,
                "okno_soni": len(missing_periods),
            })
        if gap > 1:
            multi_gap_shift_days += 1
    unified = _v200_all_teacher_idle_signature(state, context)
    overflow_rows = [
        {
            "oqituvchi_id": int(teacher_id),
            "oknoli_kunlar": sorted(int(day) for day in days),
            "oknoli_kun_soni": len(days),
            "maksimum_kun": 2,
        }
        for teacher_id, days in sorted(gap_days_by_teacher.items())
        if len(days) > 2
    ]
    return {
        "oqituvchi_ichki_okno": int(total),
        "oqituvchi_oknoli_smena_kun": int(gap_shift_days),
        "oqituvchi_kop_oknoli_smena_kun": int(multi_gap_shift_days),
        "eng_katta_ichki_okno": int(max_gap),
        "oqituvchi_ichki_okno_kunlari": gap_rows,
        "oqituvchi_ichki_okno_ortiqcha_kunlar": overflow_rows,
        "oqituvchi_ichki_okno_ortiqcha_ustoz": len(overflow_rows),
        "oqituvchi_birlashgan_3soat_okno": int(unified[0]),
        "oqituvchi_birlashgan_2soat_okno": int(unified[1]),
        "oqituvchi_birlashgan_eng_katta_daqiqa": int(unified[2]),
        "oqituvchi_birlashgan_okno_daqiqa": int(unified[3]),
        "oqituvchi_birlashgan_okno_soni": int(unified[4]),
    }


def _v196_teacher_comfort_signature(state, context):
    """Kichik qiymat — pedagogik va o'qituvchi uchun yaxshiroq jadval.

    Sinf uzluksizligi hamda asosiy/amaliy fan vaqti birinchi o'rinda qoladi.
    Shundan keyin 3 soatdan uzun smena oralig'i, 2 soatdan uzun istisnolar va
    nihoyat smena ichidagi ko'p/har kungi oynalar kamaytiriladi.
    """
    metrics = _v196_attempt_metrics(state, context)
    excess_cross = max(
        0,
        int(metrics.get("oqituvchi_smenalar_orasi_daqiqa", 0))
        - 120 * int(metrics.get("ikki_smenali_2soatdan_uzoq", 0)),
    )
    return (
        int(_v196_class_gap_count(state)),
        # O'qituvchi oynasini qisqartirish sinfning 5+5+5+5+5+5 kabi
        # teng taqsimotini 1+5+6+6+6+6 ga buzishi mumkin emas.
        int(metrics.get("sinf_kun_taqsimoti_farqi", 0)),
        int(metrics.get("sinf_qisqa_kunlari", 0)),
        # 6-darsdagi asosiy fan faqat boshqa legal kombinatsiya qolmaganda
        # ishlatiladi. Ta'mirlash o'qituvchi oynasini qisqartirish bahonasida
        # bunday kunni ko'paytirmasligi uchun u comfort mezonidan oldin turadi.
        int(metrics.get("asosiy_fan_6_limitdan_ortiq", 0)),
        int(metrics.get("asosiy_fan_6_kunlari", 0)),
        # Ikki smenali ustoz uchun 4 soatdan uzun kutish hech qachon tayyor
        # variant hisoblanmaydi; 2 soatdan uzun istisno esa haftada ko'pi
        # bilan bir kun bo'lishi mumkin. Lokal swap avval shu ikki buzilishni
        # yo'qotishga urinadi, keyin 1 soatlik maqsadga yaqinlashtiradi.
        int(metrics.get("ikki_smenali_4soatdan_uzoq", 0)),
        int(metrics.get("ikki_smenali_istisno_ortiqcha_kun", 0)),
        int(metrics.get("ikki_smenali_1soatdan_uzoq", 0)),
        int(metrics.get("oqituvchi_birlashgan_3soat_okno", 0)),
        int(metrics.get("oqituvchi_birlashgan_2soat_okno", 0)),
        int(metrics.get("oqituvchi_birlashgan_eng_katta_daqiqa", 0)),
        int(metrics.get("oqituvchi_birlashgan_okno_daqiqa", 0)),
        int(metrics.get("oqituvchi_birlashgan_okno_soni", 0)),
        int(metrics.get("ikki_smenali_uzoq_tanaffus", 0)),
        int(metrics.get("ikki_smenali_2soatdan_uzoq", 0)),
        int(metrics.get("eng_uzoq_smena_oraligi_daqiqa", 0)),
        int(excess_cross),
        int(metrics.get("oqituvchi_ichki_okno_ortiqcha_ustoz", 0)),
        int(metrics.get("oqituvchi_kop_oknoli_smena_kun", 0)),
        int(metrics.get("eng_katta_ichki_okno", 0)),
        int(metrics.get("oqituvchi_oknoli_smena_kun", 0)),
        int(metrics.get("oqituvchi_ichki_okno", 0)),
        int(metrics.get("amaliy_fan_1_2", 0)),
        int(metrics.get("jismoniydan_keyin_ogir_fan", 0)),
        int(metrics.get("asosiy_fan_5_6", 0)),
        int(metrics.get("ketma_ket_ogir_fan", 0)),
    )


def _v196_swap_same_class_day(state, first, second, context, rng):
    """Bir sinf-kundagi ikki dars o'rnini, ustozlarini saqlab almashtiradi."""
    removed = {id(first), id(second)}
    base = [
        placement for placement in state.get("placements", [])
        if id(placement) not in removed
    ]
    orders = ((second, first), (first, second))
    targets = {
        id(first): int(second.get("period") or 0),
        id(second): int(first.get("period") or 0),
    }
    for leading, trailing in orders:
        trial = _v1852_rebuild_schedule_state(base, context)
        if not _v196_place_exact(
            leading["job"], leading["day"], targets[id(leading)],
            trial, context, rng, expected_teachers=leading.get("teachers"),
        ):
            continue
        if not _v196_place_exact(
            trailing["job"], trailing["day"], targets[id(trailing)],
            trial, context, rng, expected_teachers=trailing.get("teachers"),
        ):
            continue
        if _v196_class_gap_count(trial) <= _v196_class_gap_count(state):
            return trial
    return None


def _v196_teacher_window_candidates(state, context, limit=96):
    """Eng yomon o'qituvchi-kunlardan xavfsiz swap nomzodlarini oladi.

    Avvalgi versiya faqat smena ichidagi okno yoki 120 daqiqadan katta
    smenalararo tanaffusni ko'rardi. Endi 1- va 2-smena bitta real vaqt
    chizig'i: odatiy tanaffusdan katta har qanday kutish optimizatsiyaga
    kiradi. Qaysi swap foydali ekani keyingi bosqichda to'liq jadvalni qayta
    hisoblash orqali aniqlanadi; shuning uchun bu yerda faqat yo'nalish
    bo'yicha taxmin qilib yaxshi variantni tasodifan chiqarib tashlamaymiz.
    """
    issues = []
    teacher_days = {
        (int(teacher), int(day))
        for (teacher, day), count in state.get("teacher_daily", {}).items()
        if float(count or 0) > 0
    }
    for teacher, day in teacher_days:
        first = set(state.get("teacher_periods", {}).get((teacher, day, 1), set()))
        second = set(state.get("teacher_periods", {}).get((teacher, day, 2), set()))
        internal = int(_v1852_gap_count(first) + _v1852_gap_count(second))
        unified = _v200_teacher_day_idle(state, teacher, day, context)
        if internal <= 0 and int(unified["okno_soni"]) <= 0:
            continue
        issues.append((
            int(unified["uch_soatdan_uzoq"]),
            int(unified["ikki_soatdan_uzoq"]),
            int(unified["eng_katta_daqiqa"]),
            int(unified["jami_daqiqa"]),
            int(unified["okno_soni"]),
            internal,
            teacher,
            day,
        ))

    placements = list(state.get("placements", []))
    pairs = []
    seen = set()
    for _, _, unified_max, _, _, internal, teacher, day in sorted(issues, reverse=True)[:12]:
        owned = [
            placement for placement in placements
            if int(placement.get("day") or 0) == int(day)
            and int(teacher) in {
                int(value) for value in placement.get("teachers") or []
                if value is not None
            }
            and _v196_movable_placement(placement)
        ]
        for first in owned:
            shift = int(first["job"].get("smena") or 1)
            current_periods = set(
                state.get("teacher_periods", {}).get(
                    (int(teacher), int(day), shift), set()
                )
            )
            before_internal = int(_v1852_gap_count(current_periods))
            class_id = int(first["job"].get("sinf_id") or 0)
            alternatives = [
                other for other in placements
                if int(other.get("day") or 0) == int(day)
                and int(other["job"].get("sinf_id") or 0) == class_id
                and int(other.get("period") or 0) != int(first.get("period") or 0)
                and _v196_movable_placement(other)
            ]
            for second in alternatives:
                moved_periods = (
                    current_periods - {int(first.get("period") or 0)}
                ) | {int(second.get("period") or 0)}
                improves_internal = (
                    int(_v1852_gap_count(moved_periods)) < before_internal
                )
                moves_toward_shift_edge = bool(
                    unified_max > 25 and (
                        (shift == 1 and int(second["period"]) > int(first["period"]))
                        or (shift == 2 and int(second["period"]) < int(first["period"]))
                    )
                )
                # Bir smenadagi darsni ichkariga siljitish ba'zan boshqa
                # smenadagi darsga yoki shu kunning boshqa darsiga aynan
                # tutashtiradi. Buni oddiy yo'nalish testi oldindan bilmaydi;
                # real unified signature sinab ko'rib, faqat yaxshilangan
                # holatni qabul qiladi.
                if not improves_internal and not moves_toward_shift_edge and internal <= 0:
                    continue
                key = tuple(sorted((id(first), id(second))))
                if key in seen:
                    continue
                seen.add(key)
                pairs.append((first, second))
                if len(pairs) >= int(limit):
                    return pairs
    return pairs


def _v225_teacher_ids(state):
    return sorted({
        int(teacher_id)
        for placement in state.get("placements", []) or []
        for teacher_id in placement.get("teachers", []) or []
        if teacher_id is not None
    })


def _v225_teacher_days(state, teacher_id):
    return len({
        int(day) for (owner, day), count
        in (state.get("teacher_daily", {}) or {}).items()
        if int(owner) == int(teacher_id) and float(count or 0) > 0
    })


def _v225_teacher_load(state, teacher_id):
    return sum(
        float(count or 0) for (owner, _day), count
        in (state.get("teacher_daily", {}) or {}).items()
        if int(owner) == int(teacher_id)
    )


def _v226_teacher_day_target(state, context, teacher_id):
    rules_map = (context or {}).get("rules") or {}
    defaults = (context or {}).get("default_rules") or {"kunlik_max": 6}
    rules = rules_map.get(int(teacher_id), defaults) or defaults
    return int(_v196_teacher_target_days(
        _v225_teacher_load(state, teacher_id), rules
    ))


def _v2258_presence_limit_minutes(lesson_count):
    """Kunlik dars soni uchun foydalanuvchi belgilagan real vaqt chegarasi.

    4–5 dars birinchi darsdan oxirgisigacha ko'pi bilan 8 soatga, 6–7 dars
    esa ko'pi bilan 10 soatga sig'ishi kerak. Qisqaroq variant doim yaxshiroq;
    bu qiymatlar minimum emas, ortiqcha maktabda ushlab qolishning yuqori
    chegarasidir.
    """
    lessons = max(0, int(lesson_count or 0))
    if lessons <= 1:
        return 120
    if lessons == 2:
        return 240
    if lessons == 3:
        return 360
    if lessons <= 5:
        return 480
    return 600


def _v2258_teacher_presence_policy(state, context, teacher_id):
    """TOQ/JUFT haftalarni aralashtirmay real maktabda qolishni baholaydi."""
    teacher_id = int(teacher_id)
    phase_days = {
        "toq": _v1852_defaultdict(set),
        "juft": _v1852_defaultdict(set),
    }
    for placement in state.get("placements", []) or []:
        job = placement.get("job") or {}
        teacher_phases = _v212_job_teacher_phases(
            job, placement.get("teachers") or []
        ).get(teacher_id, set())
        if not teacher_phases:
            continue
        shift = int(job.get("smena") or 1)
        period = int(placement.get("period") or 0)
        day = int(placement.get("day") or 0)
        interval = _v196_slot_interval(context, shift, period)
        if not day or not interval:
            continue
        token = (int(interval[0]), int(interval[1]), shift, period)
        for actual_phase in ("toq", "juft"):
            if any(
                _v209_week_phases_overlap(actual_phase, configured_phase)
                for configured_phase in teacher_phases
            ):
                phase_days[actual_phase][day].add(token)

    # Oddiy har-hafta jadval ikkala fazada bir xil ko'rinadi va bir marta
    # sanaladi. A/B jadval farq qilsa, ikki haqiqiy hafta alohida baholanadi.
    canonical = [phase_days["toq"]]
    if phase_days["juft"] != phase_days["toq"]:
        canonical.append(phase_days["juft"])

    one_lesson_days = 0
    overstay_total = 0
    overstay_max = 0
    span_total = 0
    work_days = 0
    target_days = 0
    excess_days = 0
    dual_shift_days = 0
    all_active_shifts = set()
    dual_shift_phases = 0
    active_shifts_by_phase = []
    low_single_shift_days = 0
    rules_map = (context or {}).get("rules") or {}
    defaults = (context or {}).get("default_rules") or {"kunlik_max": 6}
    rules = rules_map.get(teacher_id, defaults) or defaults
    for days in canonical:
        phase_load = 0
        phase_work_days = 0
        phase_active_shifts = set()
        for day, raw_timeline in sorted(days.items()):
            timeline = sorted(set(raw_timeline))
            if not timeline:
                continue
            lesson_count = len(timeline)
            active_shifts = {int(value[2]) for value in timeline}
            phase_active_shifts.update(active_shifts)
            all_active_shifts.update(active_shifts)
            span = max(0, int(timeline[-1][1]) - int(timeline[0][0]))
            allowed = int(_v2258_presence_limit_minutes(lesson_count))
            excess = max(0, span - allowed)
            one_lesson_days += int(lesson_count == 1)
            span_total += span
            overstay_total += excess
            overstay_max = max(overstay_max, excess)
            phase_load += lesson_count
            phase_work_days += 1
            dual_shift_days += int(len(active_shifts) >= 2)
            low_single_shift_days += int(
                len(active_shifts) == 1 and lesson_count <= 3
            )
        dual_shift_phases += int(len(phase_active_shifts) >= 2)
        active_shifts_by_phase.append(
            sorted(int(value) for value in phase_active_shifts)
        )
        phase_target = int(_v196_teacher_target_days(phase_load, rules))
        work_days += phase_work_days
        target_days += phase_target
        excess_days += max(0, phase_work_days - phase_target)
    return {
        "bitta_darsli_kun": int(one_lesson_days),
        "ortiqcha_qolish_jami_daqiqa": int(overstay_total),
        "eng_katta_ortiqcha_qolish_daqiqa": int(overstay_max),
        "maktabda_jami_daqiqa": int(span_total),
        "ish_kunlari": int(work_days),
        "maqsad_ish_kunlari": int(target_days),
        "ortiqcha_ish_kunlari": int(excess_days),
        "hafta_fazalari": int(len(canonical)),
        "ikki_smenali_kunlar": int(dual_shift_days),
        "ikki_smenali_haftalar": int(dual_shift_phases),
        # TOQ va JUFT haqiqiy haftalar aralashmaydi. Ustoz faqat TOQda
        # 1-smena, faqat JUFTda 2-smena ishlasa, bir real haftada ikki smenada
        # qolmaydi va dual-shift deb noto'g'ri ustuvorlashtirilmasin.
        "ikkala_smenada_darsi_bor": bool(dual_shift_phases > 0),
        "faol_smenalar": sorted(int(value) for value in all_active_shifts),
        "faol_smenalar_fazalar": active_shifts_by_phase,
        "bitta_smenali_kam_darsli_kunlar": int(low_single_shift_days),
    }


def _v225_teacher_score(state, context, teacher_id):
    """Bitta o'qituvchi uchun worst-first lexicographic baho.

    Foydalanuvchi uchun asosiy muammo — real okno. Shu sabab avval eng katta
    va jami kutish, keyin okno soni kamayadi. Faqat shundan keyin 4–5 darsni
    8 soatga, 6–7 darsni 10 soatga sig'dirish va ortiqcha ish kunini yig'ish
    baholanadi. Bitta ustozning ish kunini qisqartirish qolgan ustozlarning
    oknosini navbatsiz qoldira olmaydi.
    """
    snapshot = _v214_teacher_phase_window_snapshot(state, context, teacher_id)
    presence = _v2258_teacher_presence_policy(
        state, context, teacher_id
    )
    return (
        int(snapshot.get("eng_katta_daqiqa") or 0),
        int(snapshot.get("jami_daqiqa") or 0),
        int(snapshot.get("okno_soni") or 0),
        int(snapshot.get("ichki_okno") or 0),
        int(presence.get("eng_katta_ortiqcha_qolish_daqiqa") or 0),
        int(presence.get("ortiqcha_qolish_jami_daqiqa") or 0),
        int(presence.get("bitta_darsli_kun") or 0),
        int(presence.get("ortiqcha_ish_kunlari") or 0),
        int(presence.get("maktabda_jami_daqiqa") or 0),
        int(presence.get("ish_kunlari") or 0),
    )


def _v236_teacher_phase_window_vector(state, context, teacher_id):
    """TOQ va JUFT oynolarini alohida qotiradigan solishtirish vektori."""
    snapshot = _v214_teacher_phase_window_snapshot(
        state, context, teacher_id
    )
    return {
        str(row.get("hafta_turi") or "har_hafta"): (
            int(row.get("eng_katta_daqiqa") or 0),
            int(row.get("jami_daqiqa") or 0),
            int(row.get("okno_soni") or 0),
            int(row.get("ichki_okno") or 0),
        )
        for row in snapshot.get("fazalar", []) or []
    }


def _v236_teacher_phase_day_lesson_counts(state, teacher_id):
    """Ustozning haqiqiy TOQ/JUFT kunlik dars sonini qaytaradi.

    ``teacher_daily`` A/B aylanishidagi har bir 0.5-soatli darsni 0.5 deb
    saqlaydi. Oyna optimizatori uchun esa ustoz o'sha haqiqiy hafta fazasida
    nechta darsga kirishi muhim: masalan, to'rtta TOQ darsi 2.0 emas, 4 ta
    darsdir. Har-haftalik dars ikkala haqiqiy fazada bittadan sanaladi.
    """
    teacher_id = int(teacher_id)
    counts = {}
    for placement in (state or {}).get("placements", []) or []:
        job = placement.get("job") or {}
        teacher_phases = _v212_job_teacher_phases(
            job, placement.get("teachers") or []
        ).get(teacher_id, set())
        if not teacher_phases:
            continue
        day = int(placement.get("day") or 0)
        if day <= 0:
            continue
        for actual_phase in ("toq", "juft"):
            if any(
                _v209_week_phases_overlap(actual_phase, configured_phase)
                for configured_phase in teacher_phases
            ):
                key = (actual_phase, day)
                counts[key] = int(counts.get(key, 0)) + 1
    return counts

def _v225_changed_teachers(first, second):
    return {
        int(value)
        for placement in (first, second)
        for value in placement.get("teachers", []) or []
        if value is not None
    }


def _v226_teacher_batch_limit(teacher_count):
    teacher_count = max(0, int(teacher_count or 0))
    if teacher_count <= 30:
        return min(10, teacher_count)
    if teacher_count <= 60:
        return 15
    if teacher_count <= 100:
        return 25
    if teacher_count <= 150:
        return 35
    return 50


def _v225_target_order(state, context):
    """Avval ikki smenali, keyin faqat kam darsli bir smenali ustozlar."""
    teacher_ids = _v225_teacher_ids(state)
    rows = []
    for teacher_id in teacher_ids:
        score = _v225_teacher_score(state, context, teacher_id)
        presence = _v2258_teacher_presence_policy(
            state, context, teacher_id
        )
        dual_shift_days = int(presence.get("ikki_smenali_kunlar") or 0)
        both_shifts = bool(
            presence.get("ikkala_smenada_darsi_bor")
            or dual_shift_days > 0
        )
        low_single_shift_days = int(
            presence.get("bitta_smenali_kam_darsli_kunlar") or 0
        )
        # Birinchi 8 mezon haqiqiy muammo; jami maktab vaqti va ish kunlari
        # faqat teng variantlar orasidagi tie-breaker. Oknosiz ustozni bekorga
        # qayta-qayta o'ynamaymiz.
        if not any(int(value) > 0 for value in score[:8]):
            continue
        # Faqat bitta smenada 4+ ketma-ket darsi bor ustozga tegilmaydi.
        # Bitta smenalidan faqat 1–3 darsli kuni, 1 darsli kun/ortiqcha ish
        # kuni yoki haqiqiy oknosi bo'lganlar yengil siqiladi.
        if not both_shifts:
            has_single_shift_problem = bool(
                low_single_shift_days > 0
                and (
                    any(int(value) > 0 for value in score[:4])
                    or int(score[6]) > 0
                    or int(score[7]) > 0
                )
            )
            if not has_single_shift_problem:
                continue
        load = float(_v225_teacher_load(state, teacher_id))
        rows.append((
            -int(both_shifts),
            -int(dual_shift_days),
            *(-int(value) for value in score),
            load,
            int(teacher_id),
        ))
    return [row[-1] for row in sorted(rows)]

def _v225_teacher_candidates(state, context, teacher_id, limit=None):
    """Target ustoz uchun aynan uning o'z same-day/across-day swaplarini beradi.

    Avval singleton/eng kam darsli kundagi darsni ustoz allaqachon ishlaydigan
    kunga yig'uvchi 1:1 swaplar, keyin shu kun ichidagi tepa/past siqishlar
    sinab ko'riladi. Sinfning kunlik dars SONI o'zgarmaydi.
    """
    teacher_id = int(teacher_id)
    placements = list(state.get("placements", []) or [])
    daily = {
        int(day): float(count or 0)
        for (owner, day), count in (state.get("teacher_daily", {}) or {}).items()
        if int(owner) == teacher_id
    }
    phase_daily = _v236_teacher_phase_day_lesson_counts(state, teacher_id)
    presence = _v2258_teacher_presence_policy(state, context, teacher_id)
    both_shifts = bool(
        presence.get("ikkala_smenada_darsi_bor")
        or int(presence.get("ikki_smenali_kunlar") or 0) > 0
    )

    def applicable_phase_counts(row, day=None):
        """Row ko'chsa ta'sir qiladigan haqiqiy fazalardagi kunlik sanoq."""
        row_day = int(day if day is not None else row.get("day") or 0)
        teacher_phases = _v212_job_teacher_phases(
            row.get("job") or {}, row.get("teachers") or []
        ).get(teacher_id, set())
        return tuple(
            int(phase_daily.get((actual_phase, row_day), 0))
            for actual_phase in ("toq", "juft")
            if any(
                _v209_week_phases_overlap(actual_phase, configured_phase)
                for configured_phase in teacher_phases
            )
        )

    def row_day_load(row, day=None):
        row_day = int(day if day is not None else row.get("day") or 0)
        if both_shifts:
            # Ikki smenali ustozlar uchun avvalgi tartiblash saqlanadi.
            return float(daily.get(row_day, 0) or 0)
        counts = applicable_phase_counts(row, row_day)
        return max(counts, default=0)

    def single_shift_donor_allowed(row):
        counts = applicable_phase_counts(row)
        # Noma'lum fazani xavfsizlik uchun ko'chirmaymiz. Row tegishli bo'lgan
        # har bir haqiqiy fazada aynan 1–3 dars bo'lishi shart; 4+ bo'lsa yo'q.
        return bool(counts) and all(1 <= int(count) <= 3 for count in counts)

    owned = [
        row for row in placements
        if teacher_id in {
            int(value) for value in row.get("teachers", []) or []
            if value is not None
        }
        and _v196_movable_placement(row)
        # Faqat bitta smenada ishlaydigan ustozning 4+ darsli kuniga
        # tegilmaydi. Ustoz ikki smenada ishlasa barcha muammoli kunlari
        # tekshiriladi; bitta smenalidan faqat 1–3 darsli kun donor bo'ladi.
        and (
            both_shifts
            or single_shift_donor_allowed(row)
        )
    ]
    if not owned:
        return
    seen = set()
    yielded = 0

    # 1) Kam soatli ustozni 2–4 faol kunga yig'ish: avval eng bo'sh kun donor.
    owned_days = {
        int(day) for (_phase, day), count in phase_daily.items()
        if int(count or 0) > 0
    }
    for first in sorted(
        owned,
        key=lambda row: (
            row_day_load(row),
            int(row.get("day") or 0), int(row.get("period") or 0),
        ),
    ):
        class_id = int((first.get("job") or {}).get("sinf_id") or 0)
        first_day = int(first.get("day") or 0)
        alternatives = [
            second for second in placements
            if _v196_movable_placement(second)
            and int((second.get("job") or {}).get("sinf_id") or 0) == class_id
            and int(second.get("day") or 0) != first_day
        ]
        alternatives.sort(key=lambda second: (
            0 if int(second.get("day") or 0) in owned_days else 1,
            -row_day_load(first, int(second.get("day") or 0)),
            int(second.get("day") or 0), int(second.get("period") or 0),
        ))
        for second in alternatives:
            key = ("across_day", min(id(first), id(second)), max(id(first), id(second)))
            if key in seen:
                continue
            seen.add(key)
            yield "across_day", first, second
            yielded += 1
            if limit is not None and yielded >= int(limit):
                return

    # 2) Shu kun ichida 1/3/5 kabi oynoni tepaga/pastga siqish.
    for first in owned:
        day = int(first.get("day") or 0)
        class_id = int((first.get("job") or {}).get("sinf_id") or 0)
        alternatives = [
            second for second in placements
            if _v196_movable_placement(second)
            and int(second.get("day") or 0) == day
            and int((second.get("job") or {}).get("sinf_id") or 0) == class_id
            and int(second.get("period") or 0) != int(first.get("period") or 0)
        ]
        for second in alternatives:
            key = ("same_day", min(id(first), id(second)), max(id(first), id(second)))
            if key in seen:
                continue
            seen.add(key)
            yield "same_day", first, second
            yielded += 1
            if limit is not None and yielded >= int(limit):
                return

def _v196_optimize_teacher_windows(state, context, rng, max_swaps=None):
    """Barcha ustozni round-robin yaxshilaydi, bir ustozda qolib ketmaydi.

    Har aylanishda har bir muammoli o'qituvchi uchun ko'pi bilan bitta eng
    foydali swap qabul qilinadi. Keyin navbat keyingi ustozga o'tadi. Barcha
    ustozlar tekshirilgach yangi aylanish boshlanadi; birorta foydali hard-safe
    variant qolmaganda yoki foydalanuvchi To'xtatishni bosganda tugaydi.
    """
    swap_limit = (
        None if max_swaps is None or int(max_swaps or 0) <= 0
        else int(max_swaps)
    )

    def limit_reached():
        return swap_limit is not None and swaps >= swap_limit

    def non_worsening(after, before):
        """Har bir o'lchov alohida yomonlashmasligini talab qiladi."""
        return len(after) == len(before) and all(
            int(after[index]) <= int(before[index])
            for index in range(len(before))
        )

    def strictly_improves(after, before):
        return non_worsening(after, before) and any(
            int(after[index]) < int(before[index])
            for index in range(len(before))
        )

    def placement_signature(row):
        job = row.get("job") or {}
        return (
            job.get("id"), job.get("yuklama_id"),
            int(job.get("sinf_id") or 0),
            str(job.get("fan") or "").casefold(),
            str(job.get("guruh_kaliti") or "whole"),
            int(job.get("takror_raqami") or 0),
            str(job.get("hafta_turi") or "har_hafta"),
            int(row.get("day") or 0), int(row.get("period") or 0),
            tuple(sorted(
                int(value) for value in row.get("teachers", []) or []
                if value is not None
            )),
        )

    swaps = 0
    trials = 0
    passes = 0
    frozen = {}
    improved = []
    target_ids_seen = set()
    rejected_targets = set()
    rejected_candidate_signatures = {}
    callback_rejections = 0
    callback_rejection_reasons = {}
    cancelled_during_callback = False
    phase_guard = globals().get("_v236_teacher_phase_window_vector")

    while not limit_reached() and not _v206_deadline_reached(context):
        targets = _v225_target_order(state, context)
        target_ids_seen.update(int(value) for value in targets)
        if not targets:
            break
        passes += 1
        pass_swaps = 0

        for target_index, teacher_id in enumerate(targets, start=1):
            if limit_reached() or _v206_deadline_reached(context):
                break
            scan_callback = (context or {}).get(
                "v236_teacher_scan_callback"
            )
            if callable(scan_callback):
                try:
                    scan_callback(
                        passes, target_index, len(targets), teacher_id,
                        trials, swaps,
                    )
                except Exception:
                    pass
            before = _v225_teacher_score(state, context, teacher_id)
            before_global = _v196_teacher_comfort_signature(state, context)
            before_class_safety = (
                int(_v196_class_gap_count(state)),
                *_v196_class_distribution_metrics(state, context),
            )
            current_score_cache = {int(teacher_id): before}
            current_phase_cache = {}

            def current_score(owner):
                owner = int(owner)
                if owner not in current_score_cache:
                    current_score_cache[owner] = _v225_teacher_score(
                        state, context, owner
                    )
                return current_score_cache[owner]

            def current_phases(owner):
                owner = int(owner)
                if owner not in current_phase_cache:
                    current_phase_cache[owner] = phase_guard(
                        state, context, owner
                    )
                return current_phase_cache[owner]

            blocked = rejected_candidate_signatures.setdefault(
                int(teacher_id), set()
            )

            # Runtime bir ranked variantni (masalan A/B signature yoki DB
            # persist sabab) rad etsa, shu ustozning keyingi xavfsiz varianti
            # ham sinab ko'riladi. Bitta rad javobi ustozni yoki qolgan 55
            # ustozni butunlay tashlab yubormaydi.
            while not limit_reached() and not _v206_deadline_reached(context):
                best = None
                for move_kind, first, second in _v225_teacher_candidates(
                    state, context, teacher_id, limit=None
                ):
                    if _v206_deadline_reached(context):
                        break
                    candidate_signature = (
                        str(move_kind), placement_signature(first),
                        placement_signature(second),
                    )
                    if candidate_signature in blocked:
                        continue
                    changed = _v225_changed_teachers(first, second)
                    if int(teacher_id) not in changed:
                        continue
                    before_changed_scores = {
                        int(owner): current_score(owner)
                        for owner in changed
                    }
                    before_changed_phases = {
                        int(owner): current_phases(owner)
                        for owner in changed
                    } if callable(phase_guard) else {}
                    trials += 1
                    if move_kind == "across_day":
                        trial = _v219_swap_across_class_days(
                            state, first, second, context, rng
                        )
                    else:
                        trial = _v196_swap_same_class_day(
                            state, first, second, context, rng
                        )
                    if trial is None or not _v226_class_day_counts_match(trial, context):
                        continue
                    trial_class_safety = (
                        int(_v196_class_gap_count(trial)),
                        *_v196_class_distribution_metrics(trial, context),
                    )
                    if not non_worsening(
                        trial_class_safety, before_class_safety
                    ):
                        continue
                    after_changed_scores = {
                        int(owner): _v225_teacher_score(trial, context, owner)
                        for owner in changed
                    }
                    # Lexicographic solishtirish keyingi ko'rsatkichning katta
                    # yomonlashishini yashirishi mumkin edi. Endi darsi ko'chgan
                    # har bir ustozning HAR BIR okno/qolish/ish-kuni o'lchovi
                    # alohida ravishda yomonlashmasligi shart.
                    if any(
                        not non_worsening(
                            after_changed_scores[int(owner)], before_score
                        )
                        for owner, before_score in before_changed_scores.items()
                    ):
                        continue
                    if any(
                        not non_worsening(
                            after_changed_scores[int(owner)], frozen[int(owner)]
                        )
                        for owner in changed if int(owner) in frozen
                    ):
                        continue
                    if callable(phase_guard):
                        after_changed_phases = {
                            int(owner): phase_guard(trial, context, owner)
                            for owner in changed
                        }
                        phase_worsened = False
                        for owner, before_phases in before_changed_phases.items():
                            after_phases = after_changed_phases.get(owner, {})
                            if set(after_phases) != set(before_phases) or any(
                                not non_worsening(
                                    after_phases[phase], before_vector
                                )
                                for phase, before_vector in before_phases.items()
                            ):
                                phase_worsened = True
                                break
                        if phase_worsened:
                            continue
                    target_score = after_changed_scores[int(teacher_id)]
                    if not strictly_improves(target_score, before):
                        continue
                    global_score = _v196_teacher_comfort_signature(trial, context)
                    if not non_worsening(global_score, before_global):
                        continue
                    rank = (target_score, global_score)
                    if best is None or rank < best[0]:
                        best = (
                            rank, trial, target_score, candidate_signature,
                        )

                if best is None:
                    break

                candidate_state = best[1]
                after_score = best[2]
                candidate_signature = best[3]
                progress_callback = (context or {}).get(
                    "v2253_improvement_callback"
                )
                accepted_by_runtime = True
                rejection_reason = "runtime_rejected"
                if isinstance(context, dict):
                    context.pop("v2253_last_rejection_reason", None)
                if callable(progress_callback):
                    try:
                        accepted_by_runtime = bool(progress_callback(
                            int(swaps + 1), int(teacher_id), before,
                            after_score, candidate_state,
                        ))
                        if isinstance(context, dict):
                            rejection_reason = str(
                                context.get("v2253_last_rejection_reason")
                                or rejection_reason
                            )
                    except Exception as callback_error:
                        accepted_by_runtime = False
                        rejection_reason = (
                            "callback_exception:"
                            + type(callback_error).__name__
                        )
                if not accepted_by_runtime:
                    callback_rejections += 1
                    rejected_targets.add(int(teacher_id))
                    blocked.add(candidate_signature)
                    callback_rejection_reasons[rejection_reason] = int(
                        callback_rejection_reasons.get(rejection_reason, 0)
                    ) + 1
                    if _v206_deadline_reached(context):
                        cancelled_during_callback = True
                        break
                    continue

                state = candidate_state
                swaps += 1
                pass_swaps += 1
                if int(teacher_id) not in improved:
                    improved.append(int(teacher_id))
                frozen[int(teacher_id)] = after_score
                break

            if cancelled_during_callback:
                break

        if cancelled_during_callback or pass_swaps == 0:
            break

    state["v196_teacher_window_swaps"] = int(swaps)
    state["v225_teacher_window_trials"] = int(trials)
    state["v225_teacher_targets"] = int(len(target_ids_seen))
    state["v225_teacher_improved"] = list(improved)
    state["v225_teacher_frozen"] = int(len(frozen))
    state["v225_teacher_callback_rejections"] = int(callback_rejections)
    state["v225_teacher_rejected_targets"] = sorted(int(value) for value in rejected_targets)
    state["v225_teacher_rejection_reasons"] = dict(
        sorted(callback_rejection_reasons.items())
    )
    state["v225_teacher_passes"] = int(passes)
    state["v225_optimizer_mode"] = "v236_okno_first_all_teachers_continue_on_reject"
    state["v226_class_day_counts_preserved"] = bool(
        _v226_class_day_counts_match(state, context)
    )
    return state

def _v219_subject_repeat_signature(state, context=None):
    """Kunlik ruxsat 2 bo'lsa ham ishlatilgan ikkinchi fanlarni sanaydi.

    Kichik qiymat yaxshi. ``daily_max=2`` majburiyat emas, faqat yechimni
    saqlab qoluvchi zaxira; shu sabab qulaylik bosqichi imkon topganda bunday
    takrorni boshqa kundagi fan bilan xavfsiz almashtiradi.
    """
    context = context or {}
    practical_pairs = set()
    periods_by_key = _v1852_defaultdict(set)
    for placement in state.get("placements", []) or []:
        job = placement.get("job") or {}
        pair = (
            int(job.get("sinf_id") or 0),
            str(job.get("fan") or "").casefold(),
        )
        profile = _v196_rotation_profile(job, context)
        if profile.get("physical") or profile.get("technology"):
            practical_pairs.add(pair)
        periods_by_key[(
            pair[0], pair[1], int(placement.get("day") or 0)
        )].add(int(placement.get("period") or 0))
    excess = []
    adjacent = 0
    for key, raw_count in (state.get("subject_daily", {}) or {}).items():
        try:
            count = int(raw_count or 0)
        except (TypeError, ValueError):
            count = 0
        if count > 1 and tuple(key[:2]) not in practical_pairs:
            excess.append((tuple(key), count - 1))
            ordered = sorted(periods_by_key.get(tuple(key), set()))
            adjacent += sum(
                1 for left, right in zip(ordered, ordered[1:])
                if int(right) - int(left) == 1
            )
    return (
        sum(value for _key, value in excess),
        int(adjacent),
        len(excess),
        tuple(sorted((str(key), value) for key, value in excess)),
    )


def _v219_swap_across_class_days(state, first, second, context, rng):
    """Bir sinfning ikki kunidagi darslarini hard-safe tarzda almashtiradi."""
    first_job = first.get("job") or {}
    second_job = second.get("job") or {}
    if (
        int(first_job.get("sinf_id") or 0)
        != int(second_job.get("sinf_id") or 0)
        or int(first.get("day") or 0) == int(second.get("day") or 0)
    ):
        return None
    removed = {id(first), id(second)}
    base = [
        placement for placement in state.get("placements", [])
        if id(placement) not in removed
    ]
    targets = {
        id(first): (int(second.get("day") or 0), int(second.get("period") or 0)),
        id(second): (int(first.get("day") or 0), int(first.get("period") or 0)),
    }
    for leading, trailing in ((first, second), (second, first)):
        trial = _v1852_rebuild_schedule_state(base, context)
        lead_day, lead_period = targets[id(leading)]
        trail_day, trail_period = targets[id(trailing)]
        if not _v196_place_exact(
            leading["job"], lead_day, lead_period, trial, context, rng,
            expected_teachers=leading.get("teachers"),
        ):
            continue
        if not _v196_place_exact(
            trailing["job"], trail_day, trail_period, trial, context, rng,
            expected_teachers=trailing.get("teachers"),
        ):
            continue
        if _v196_class_gap_count(trial) <= _v196_class_gap_count(state):
            return trial
    return None


def _v219_reduce_avoidable_subject_repeats(
    state, context, rng, max_swaps=16, max_trials=120
):
    """Algebra+Algebra kabi takrorni boshqa kun faniga xavfsiz almashtiradi.

    Avval ayni o'qituvchining shu sinfdagi boshqa fani (masalan Geometriya),
    keyin boshqa legal fan sinab ko'riladi. Har sinov butun state'ni qayta
    quradi; qizil/BAND, metod kuni, sinf/o'qituvchi/xona kolliziyasi, kunlik
    limit va real smena vaqti yana tekshiriladi. O'qituvchi qulayligi yoki
    sinf oknosi yomonlashadigan swap qabul qilinmaydi.
    """
    before = _v219_subject_repeat_signature(state, context)
    state["v219_subject_repeat_before"] = int(before[0])
    swaps = 0
    trials = 0
    while swaps < int(max_swaps) and trials < int(max_trials):
        if _v206_deadline_reached(context):
            break
        current_repeat = _v219_subject_repeat_signature(state, context)
        if current_repeat[0] <= 0:
            break
        current_comfort = _v196_teacher_comfort_signature(state, context)
        placements = list(state.get("placements", []))
        repeated = {
            (int(key[0]), str(key[1]), int(key[2]))
            for key, count in (state.get("subject_daily", {}) or {}).items()
            if int(count or 0) > 1
        }
        targets = [
            placement for placement in placements
            if _v196_movable_placement(placement)
            and not (
                _v196_rotation_profile(
                    placement.get("job") or {}, context
                ).get("physical")
                or _v196_rotation_profile(
                    placement.get("job") or {}, context
                ).get("technology")
            )
            and (
                int((placement.get("job") or {}).get("sinf_id") or 0),
                str((placement.get("job") or {}).get("fan") or "").casefold(),
                int(placement.get("day") or 0),
            ) in repeated
        ]
        best = None
        for first in sorted(
            targets,
            key=lambda row: (
                int((row.get("job") or {}).get("sinf_id") or 0),
                int(row.get("day") or 0), int(row.get("period") or 0),
            ),
        ):
            if _v206_deadline_reached(context) or trials >= int(max_trials):
                break
            first_job = first.get("job") or {}
            class_id = int(first_job.get("sinf_id") or 0)
            subject = str(first_job.get("fan") or "").casefold()
            first_teachers = {
                int(value) for value in first.get("teachers") or []
                if value is not None
            }
            alternatives = [
                other for other in placements
                if _v196_movable_placement(other)
                and int((other.get("job") or {}).get("sinf_id") or 0) == class_id
                and int(other.get("day") or 0) != int(first.get("day") or 0)
                and str((other.get("job") or {}).get("fan") or "").casefold()
                != subject
                and int((state.get("subject_daily", {}) or {}).get((
                    class_id, subject, int(other.get("day") or 0)
                ), 0) or 0) == 0
            ]
            alternatives.sort(key=lambda other: (
                0 if first_teachers.intersection({
                    int(value) for value in other.get("teachers") or []
                    if value is not None
                }) else 1,
                int(other.get("day") or 0), int(other.get("period") or 0),
                str((other.get("job") or {}).get("fan") or "").casefold(),
            ))
            for second in alternatives:
                if _v206_deadline_reached(context) or trials >= int(max_trials):
                    break
                trials += 1
                trial = _v219_swap_across_class_days(
                    state, first, second, context, rng
                )
                if trial is None:
                    continue
                repeat_signature = _v219_subject_repeat_signature(trial, context)
                if repeat_signature >= current_repeat:
                    continue
                comfort = _v196_teacher_comfort_signature(trial, context)
                if comfort > current_comfort:
                    continue
                shared_teacher = bool(first_teachers.intersection({
                    int(value) for value in second.get("teachers") or []
                    if value is not None
                }))
                rank = (
                    repeat_signature,
                    0 if shared_teacher else 1,
                    comfort,
                )
                if best is None or rank < best[0]:
                    best = (rank, trial)
            # Boshqa kunga tarqatish topilmasa, A-A-G ni A-G-A qilish ham
            # foydali: fan o'sha kuni qoladi, ammo ikki bir xil dars orasiga
            # boshqa fan kiradi. Bu ham to'liq hard-safe swap orqali o'tadi.
            if best is None or best[0][0][0] >= current_repeat[0]:
                same_day = [
                    other for other in placements
                    if _v196_movable_placement(other)
                    and int((other.get("job") or {}).get("sinf_id") or 0)
                    == class_id
                    and int(other.get("day") or 0)
                    == int(first.get("day") or 0)
                    and str((other.get("job") or {}).get("fan") or "").casefold()
                    != subject
                ]
                for second in same_day:
                    if _v206_deadline_reached(context) or trials >= int(max_trials):
                        break
                    trials += 1
                    trial = _v196_swap_same_class_day(
                        state, first, second, context, rng
                    )
                    if trial is None:
                        continue
                    repeat_signature = _v219_subject_repeat_signature(
                        trial, context
                    )
                    if repeat_signature >= current_repeat:
                        continue
                    comfort = _v196_teacher_comfort_signature(trial, context)
                    if comfort > current_comfort:
                        continue
                    rank = (repeat_signature, 0, comfort)
                    if best is None or rank < best[0]:
                        best = (rank, trial)
            if best is not None and best[0][0][0] == 0:
                break
        if best is None:
            break
        state = best[1]
        swaps += 1
    after = _v219_subject_repeat_signature(state, context)
    state["v219_subject_repeat_before"] = int(before[0])
    state["v219_subject_repeat_after"] = int(after[0])
    state["v219_subject_adjacent_after"] = int(after[1])
    state["v219_subject_repeat_swaps"] = int(swaps)
    state["v219_subject_repeat_trials"] = int(trials)
    return state


def _v214_teacher_phase_window_snapshot(state, context, teacher_id):
    """TOQ/JUFT haftani aralashtirmasdan bitta ustoz oynalarini hisoblaydi.

    ``teacher_periods`` indeksi fizik kataklar uchun qulay, ammo u A/B
    aylanishidagi TOQ va JUFT darslarni bitta hafta deb ko'rsatadi. Hisobot
    foydalanuvchiga yolg'on oyna aytmasligi uchun placementlar fazasi bo'yicha
    qayta yig'iladi. TOQ/JUFT farq qilsa umumiy son va daqiqa ikki haftalik
    sikl yig'indisi, maksimum esa ikkala haftaning kattasidir; oddiy
    ``har_hafta`` jadval ikki marta sanalmaydi.
    """
    teacher_id = int(teacher_id)
    phase_periods = {
        "toq": _v1852_defaultdict(set),
        "juft": _v1852_defaultdict(set),
    }
    for placement in state.get("placements", []):
        phases = _v212_job_teacher_phases(
            placement.get("job") or {}, placement.get("teachers") or []
        ).get(teacher_id, set())
        if not phases:
            continue
        day = int(placement.get("day") or 0)
        shift = int((placement.get("job") or {}).get("smena") or 1)
        period = int(placement.get("period") or 0)
        for actual_phase in ("toq", "juft"):
            if any(
                _v209_week_phases_overlap(actual_phase, configured_phase)
                for configured_phase in phases
            ):
                phase_periods[actual_phase][(day, shift)].add(period)

    def phase_result(phase):
        by_day = _v1852_defaultdict(list)
        internal = 0
        rows = []
        for (day, shift), periods in sorted(phase_periods[phase].items()):
            ordered = sorted(int(value) for value in periods if int(value) > 0)
            if not ordered:
                continue
            missing = [
                period for period in range(min(ordered), max(ordered) + 1)
                if period not in periods
            ]
            internal += len(missing)
            for period in ordered:
                interval = _v196_slot_interval(context, shift, period)
                if interval:
                    by_day[int(day)].append((
                        int(interval[0]), int(interval[1]), int(shift), int(period)
                    ))
            if missing:
                rows.append({
                    "hafta_turi": phase,
                    "kun": int(day),
                    "kun_nomi": _V1852_HAFTA.get(int(day), str(day)),
                    "smena": int(shift),
                    "darslar": ordered,
                    "bosh_darslar": missing,
                    "ichki_okno": len(missing),
                })
        gap_count = 0
        gap_minutes = 0
        max_minutes = 0
        for day, timeline in sorted(by_day.items()):
            timeline = sorted(set(timeline))
            day_gaps = []
            for previous, current in zip(timeline, timeline[1:]):
                minutes = max(0, int(current[0]) - int(previous[1]))
                if minutes > 25:
                    day_gaps.append(minutes)
            if day_gaps:
                gap_count += len(day_gaps)
                gap_minutes += sum(day_gaps)
                max_minutes = max(max_minutes, max(day_gaps))
                rows.append({
                    "hafta_turi": phase,
                    "kun": int(day),
                    "kun_nomi": _V1852_HAFTA.get(int(day), str(day)),
                    "smena": 0,
                    "darslar": [],
                    "bosh_darslar": [],
                    "real_okno_soni": len(day_gaps),
                    "real_okno_daqiqalari": day_gaps,
                    "jami_daqiqa": sum(day_gaps),
                })
        return {
            "hafta_turi": phase,
            "okno_soni": int(gap_count),
            "jami_daqiqa": int(gap_minutes),
            "eng_katta_daqiqa": int(max_minutes),
            "ichki_okno": int(internal),
            "kunlar": rows,
        }

    toq = phase_result("toq")
    juft = phase_result("juft")
    # Har bir faza alohida saqlanadi. Avvalgi lexicographic ``selected``
    # bitta haftadagi oyna sonini, boshqa haftadagi 300 daqiqalik kutishni esa
    # yashirib qo'yishi mumkin edi. Endi TOQ/JUFT farq qilsa son va daqiqalar
    # aniq ikki haftalik sikl bo'yicha yig'iladi, maksimum esa ikkala haftadan
    # olinadi. Oddiy ``har_hafta`` jadval bir marta sanaladi.
    same = phase_periods["toq"] == phase_periods["juft"]
    canonical_phases = [toq, juft]
    if same:
        display = {
            **toq,
            "hafta_turi": "har_hafta",
            "kunlar": [
                {**row, "hafta_turi": "har_hafta"}
                for row in toq["kunlar"]
            ],
        }
        phase_rows = [display]
        aggregate_rows = [toq]
    else:
        phase_rows = canonical_phases
        aggregate_rows = canonical_phases
    selected = max(
        canonical_phases,
        key=lambda row: (
            int(row["okno_soni"]), int(row["jami_daqiqa"]),
            int(row["eng_katta_daqiqa"]), int(row["ichki_okno"]),
            1 if row["hafta_turi"] == "toq" else 0,
        ),
    )
    return {
        "okno_soni": int(sum(row["okno_soni"] for row in aggregate_rows)),
        "jami_daqiqa": int(sum(row["jami_daqiqa"] for row in aggregate_rows)),
        "eng_katta_daqiqa": int(max(
            [row["eng_katta_daqiqa"] for row in aggregate_rows] or [0]
        )),
        "ichki_okno": int(sum(row["ichki_okno"] for row in aggregate_rows)),
        "eng_yomon_hafta": selected["hafta_turi"],
        "hisoblash_davri": "har_hafta" if same else "ikki_haftalik_sikl",
        "kunlar": [row for phase in phase_rows for row in phase["kunlar"]],
        "haftalar": phase_rows,
        # Solishtirish doimo ikkala real faza bilan bajariladi. ``haftalar``
        # esa UI uchun bir xil TOQ/JUFT jadvalni ``har_hafta`` deb ixchamlaydi.
        "fazalar": canonical_phases,
    }


def _v214_effective_availability_rows(rows, teacher_id, day, shift, period):
    """Bitta fizik katakni yopayotgan raw qizil/metod qatorlarini qaytaradi."""
    result = []
    for source in rows or []:
        if int(source.get("user_id") or 0) != int(teacher_id):
            continue
        if int(source.get("hafta_kuni") or 0) != int(day):
            continue
        kind = str(source.get("turi") or "")
        if kind == "metod_kuni":
            result.append(dict(source))
            continue
        if kind != "band" or not bool(source.get("qattiq")):
            continue
        source_shift = int(source.get("smena") or 0)
        source_period = int(source.get("dars_raqami") or 0)
        if source_shift not in (0, int(shift)):
            continue
        if source_period not in (0, int(period)):
            continue
        result.append(dict(source))
    return result


def _v214_single_slot_write_plan(
    rows, context, teacher_id, day, shift, period, additional_slots=None
):
    """What-if istisnosini DB/UI orqali qayta yaratish uchun to'liq reja.

    Kun/smena umbrella qatorini shunchaki yumshatish uning barcha katagini
    ochib yuboradi. Metod kuni esa ``qattiq=False`` bo'lsa ham generator uchun
    qattiq. Shu sabab manba qatorlari o'chiriladi va targetdan boshqa qamrov
    exact ``band/qattiq`` qatorlarga materializatsiya qilinadi. Rejada manba
    id yo'q bo'lsa tavsiya actionable emas va umuman qaytarilmaydi.
    """
    teacher_id, day = int(teacher_id), int(day)
    shift, period = int(shift), int(period)
    sources = [dict(row) for row in (rows or [])]
    if not sources or any(row.get("id") is None for row in sources):
        return None

    all_slots = sorted({
        (int(shift_key), int(slot.get("dars_raqami") or 0))
        for shift_key, shift_row in (context.get("shifts") or {}).items()
        for slot in (shift_row or {}).get("slotlar") or []
        if int(slot.get("dars_raqami") or 0) > 0
    })
    target_slots = {(shift, period)} | {
        (int(value[0]), int(value[1]))
        for value in (additional_slots or [])
    }
    if not target_slots or not target_slots.issubset(set(all_slots)):
        return None

    def source_covers(source, candidate_shift, candidate_period):
        if str(source.get("turi") or "") == "metod_kuni":
            return True
        source_shift = int(source.get("smena") or 0)
        source_period = int(source.get("dars_raqami") or 0)
        return (
            source_shift in (0, int(candidate_shift))
            and source_period in (0, int(candidate_period))
        )

    preserved_slots = {
        (candidate_shift, candidate_period)
        for candidate_shift, candidate_period in all_slots
        if (candidate_shift, candidate_period) not in target_slots
        and any(
            source_covers(source, candidate_shift, candidate_period)
            for source in sources
        )
    }
    source_ids = sorted({int(row["id"]) for row in sources})
    hard_rows = [
        {
            "user_id": teacher_id,
            "hafta_kuni": day,
            "smena": candidate_shift,
            "dars_raqami": candidate_period,
            "turi": "band",
            "qattiq": True,
            "izoh": "Metod/qizil qoida: exact katak saqlandi",
        }
        for candidate_shift, candidate_period in sorted(preserved_slots)
    ]
    target_soft_rows = [
        {
            "user_id": teacher_id,
            "hafta_kuni": day,
            "smena": target_shift,
            "dars_raqami": target_period,
            "turi": "band",
            "qattiq": False,
            "izoh": "Oyna hisobotidagi exact-katak istisnosi",
        }
        for target_shift, target_period in sorted(target_slots)
    ]
    source_types = {str(row.get("turi") or "") for row in sources}
    return {
        "to_liq": True,
        "manba_idlar": source_ids,
        "manba_qatorlari": [
            {
                key: row.get(key)
                for key in (
                    "id", "user_id", "hafta_kuni", "smena",
                    "dars_raqami", "turi", "qattiq", "izoh",
                )
            }
            for row in sources
        ],
        "amallar": [
            {
                "amal": "ochirish",
                "availability_id": source_id,
            }
            for source_id in source_ids
        ] + [
            {"amal": "yaratish", "qator": row}
            for row in hard_rows
        ] + [
            {"amal": "yaratish", "qator": row}
            for row in target_soft_rows
        ],
        "qolgan_qattiq_kataklar": hard_rows,
        "ochiladigan_katak": target_soft_rows[0],
        "ochiladigan_kataklar": target_soft_rows,
        "istisno_soati": len(target_soft_rows),
        "metod_qatori_almashtiriladi": "metod_kuni" in source_types,
        "izoh": (
            "Manba qatorlarining barchasini shunchaki yumshatmang: "
            "ko'rsatilgan idlarni o'chirib, qamrovdagi boshqa kataklarni "
            "exact qizil/qattiq qilib qayta yarating; faqat ko'rsatilgan "
            f"{len(target_soft_rows)} ta target katakni yashil/yumshoq qiling."
        ),
    }


def _v214_context_with_single_slot_exception(
    context, teacher_id, day, shift, period
):
    """Faqat target katakni ochib, umbrella qoidaning qolganini yopiq qoldiradi."""
    teacher_id, day = int(teacher_id), int(day)
    shift, period = int(shift), int(period)
    hard = set(context.get("hard", set()))
    soft = set(context.get("soft", set()))
    method_hard = set(context.get("method_hard", set()))
    method_soft = set(context.get("method_soft", set()))
    was_method = (teacher_id, day) in method_hard
    was_red = _v1852_blocked(hard, teacher_id, day, shift, period)
    method_origin_pairs = set(
        context.get("_v214_method_origin_pairs", set())
    )
    synthetic_method_hard = set(
        context.get("_v214_synthetic_method_hard", set())
    )
    target_token = (teacher_id, day, shift, period)
    continued_method = (
        (teacher_id, day) in method_origin_pairs
        and target_token in synthetic_method_hard
    )
    # What-if faqat metod kunining aniq 1–2 katagini ko'radi. Qizil/BAND
    # katakni hatto hisobot uchun ham ochib sinamaymiz va tavsiya qilmaymiz.
    if (not was_method and not continued_method) or (
        was_red and not continued_method
    ):
        return None, None

    all_slots = []
    for shift_key, shift_row in sorted(
        (context.get("shifts") or {}).items(), key=lambda item: int(item[0])
    ):
        for slot in (shift_row or {}).get("slotlar") or []:
            all_slots.append((int(shift_key), int(slot.get("dars_raqami") or 0)))

    def key_matches(key, candidate_shift, candidate_period):
        return (
            int(key[0]) == teacher_id
            and int(key[1]) == day
            and int(key[2]) in (0, int(candidate_shift))
            and int(key[3]) in (0, int(candidate_period))
        )

    removed_keys = [
        key for key in hard if key_matches(key, shift, period)
    ]
    for key in removed_keys:
        hard.discard(key)
        for other_shift, other_period in all_slots:
            if (other_shift, other_period) == (shift, period):
                continue
            if key_matches(key, other_shift, other_period):
                hard.add((teacher_id, day, other_shift, other_period))
    if was_method:
        method_hard.discard((teacher_id, day))
        # Amaldagi availability modeli yumshoq metod kunini ifodalamaydi:
        # ``metod_kuni`` qatori qattiq=False bo'lsa ham qattiq sanaladi.
        # Sinov yuqoridagi yozish rejasiga aynan mos bo'lishi uchun metod qatori
        # olib tashlanadi, qolgan kataklar exact hard, target esa soft bo'ladi.
        method_soft.discard((teacher_id, day))
        for other_shift, other_period in all_slots:
            if (other_shift, other_period) != (shift, period):
                hard.add((teacher_id, day, other_shift, other_period))
                synthetic_method_hard.add(
                    (teacher_id, day, other_shift, other_period)
                )
        method_origin_pairs.add((teacher_id, day))
    elif continued_method:
        synthetic_method_hard.discard(target_token)
    soft.add((teacher_id, day, shift, period))
    trial = dict(context)
    trial["hard"] = hard
    trial["soft"] = soft
    trial["method_hard"] = method_hard
    trial["method_soft"] = method_soft
    trial["_v214_method_origin_pairs"] = method_origin_pairs
    trial["_v214_synthetic_method_hard"] = synthetic_method_hard
    return trial, "metod_kuni"


def _v214_teacher_window_relaxation_report(
    state, context, teachers, *, max_suggestions=8, max_probes=48
):
    """Joriy valid draftni o'zgartirmay, bitta-katak what-if hisobotini beradi.

    Raqamli kamayish faqat ikki dars exact qayta joylashtirilib, barcha qattiq
    qoidalar (e'lon qilingan bitta availability istisnosidan tashqari) va
    pedagogik himoyalar qayta tekshirilganda chiqadi. Deadline tugashi yoki
    diagnostika xatosi jadvalni saqlashga xalaqit bermaydi.
    """
    started = _samtm_time.monotonic()
    try:
        deadline = float(context.get("v214_analysis_deadline") or 0)
    except (TypeError, ValueError):
        deadline = 0.0
    if not deadline:
        deadline = started + 0.8
    max_suggestions = max(0, min(12, int(max_suggestions or 0)))
    max_probes = max(0, min(180, int(max_probes or 0)))
    availability_rows = list(context.get("availability_rows") or [])

    teacher_ids = sorted({
        int(value) for value in teachers
    } | {
        int(teacher)
        for placement in state.get("placements", [])
        for teacher in (placement.get("teachers") or [])
        if teacher is not None
    })
    snapshots = {
        teacher_id: _v214_teacher_phase_window_snapshot(
            state, context, teacher_id
        )
        for teacher_id in teacher_ids
    }

    def phase_metrics(snapshot):
        """Snapshotni TOQ/JUFT bo'yicha taqqoslanadigan tuple'ga aylantiradi."""
        result = {}
        rows = list(snapshot.get("fazalar") or snapshot.get("haftalar") or [])
        for row in rows:
            phase = str(row.get("hafta_turi") or "har_hafta")
            metric = (
                int(row.get("okno_soni") or 0),
                int(row.get("jami_daqiqa") or 0),
                int(row.get("eng_katta_daqiqa") or 0),
                int(row.get("ichki_okno") or 0),
            )
            if phase == "har_hafta":
                result["toq"] = metric
                result["juft"] = metric
            else:
                result[phase] = metric
        zero = (0, 0, 0, 0)
        result.setdefault("toq", zero)
        result.setdefault("juft", zero)
        return result

    def snapshot_not_worse(after, before, *, require_improvement=False):
        before_phases = phase_metrics(before)
        after_phases = phase_metrics(after)
        strictly_better = False
        for phase in ("toq", "juft"):
            before_metric = before_phases[phase]
            after_metric = after_phases[phase]
            if any(
                int(after_value) > int(before_value)
                for after_value, before_value in zip(
                    after_metric, before_metric
                )
            ):
                return False
            if any(
                int(after_value) < int(before_value)
                for after_value, before_value in zip(
                    after_metric, before_metric
                )
            ):
                strictly_better = True
        return strictly_better if require_improvement else True

    affected = []
    for teacher_id, snapshot in snapshots.items():
        if not snapshot["okno_soni"] and not snapshot["ichki_okno"]:
            continue
        teacher_name = str(
            (teachers.get(teacher_id) or {}).get("full_name") or teacher_id
        )
        affected.append({
            "oqituvchi_id": int(teacher_id),
            "full_name": teacher_name,
            **snapshot,
        })
    affected.sort(key=lambda row: (
        -int(row["jami_daqiqa"]), -int(row["okno_soni"]),
        -int(row["ichki_okno"]), str(row["full_name"]), int(row["oqituvchi_id"]),
    ))

    result = {
        "jami_okno_oldin": int(sum(row["okno_soni"] for row in affected)),
        "jami_ichki_okno_oldin": int(sum(row["ichki_okno"] for row in affected)),
        "jami_okno_daqiqa_oldin": int(sum(row["jami_daqiqa"] for row in affected)),
        "eng_katta_okno_daqiqa": int(max(
            [row["eng_katta_daqiqa"] for row in affected] or [0]
        )),
        "oknoli_oqituvchi_soni": len(affected),
        "oknoli_oqituvchilar": affected,
        "tavsiyalar": [],
        "tekshirilgan_variantlar": 0,
        # Bu butun constraint space isboti emas: eng yomon ustozlar atrofidagi
        # deterministik, chegaralangan ikki-dars swap qidiruvidir. Oynasi bor
        # holatda uni hech qachon ``to'liq`` deb ko'rsatmaymiz.
        "cheklangan": bool(affected),
        "hisobot_toliq": not bool(affected),
        "qidiruv_turi": (
            "chegaralangan_lokal" if affected else "tekshiruv_shart_emas"
        ),
        "cheklash_sabablari": (
            [
                "faqat eng ustuvor 10 o'qituvchi",
                "har o'qituvchida ko'pi bilan 12 manba dars",
                "faqat bir sinf ichidagi ikki-dars exact swap",
            ]
            if affected else []
        ),
        "avtomatik_qollanmadi": True,
        "avval_qayta_yaratish": bool(
            context.get("v214_avval_qayta_yaratish")
        ),
        "izoh": (
            "Haqiqiy ko'rsatkich TOQ/JUFT haftani alohida hisoblaydi. "
            + (
                "Qidiruv to'liq tugamadi; tekshirilmagan variantlar bor. "
                if affected else ""
            )
            +
            "Tavsiya jadvalga qo'llanmadi; vaqtni tahrirlab, yagona "
            "generatorni qayta ishga tushirish kerak."
        ),
    }
    if not affected or max_suggestions <= 0 or max_probes <= 0:
        return result

    placements = list(state.get("placements", []))
    by_class = _v1852_defaultdict(list)
    for placement in placements:
        by_class[int((placement.get("job") or {}).get("sinf_id") or 0)].append(
            placement
        )
    before_metrics = dict(
        state.get("v196_metrics") or _v196_attempt_metrics(state, context)
    )
    # `_v196_attempt_metrics` qaytaradigan barcha sonli pedagogik/qulaylik
    # mezonlari himoyalanadi. Faqat ayrim "muhim" kalitlarni sanash boshqa
    # ustozga yangi uzun oyna, qo'shimcha ish kuni yoki smena oralig'i berib
    # qo'yishi mumkin edi. Ro'yxatlarni (diagnostika qatorlarini) emas, ularning
    # quyidagi aniq sonli agregatlarini oldin→keyin solishtiramiz.
    protected_metric_keys = (
        "oqituvchi_bitta_darsli_kun", "oqituvchi_faol_kun",
        "ketma_ket_ogir_fan", "jismoniydan_keyin_ogir_fan",
        "9_11_birinchi_dars_ogir", "asosiy_fan_5_6",
        "asosiy_fan_6_kunlari", "asosiy_fan_6_limitdan_ortiq",
        "amaliy_fan_1_2",
        "oqituvchi_smenalar_orasi_blok",
        "oqituvchi_smenalar_orasi_daqiqa",
        "eng_uzoq_smena_oraligi_daqiqa",
        "ikki_smenali_1soatdan_uzoq", "ikki_smenali_2soatdan_uzoq",
        "ikki_smenali_uzoq_tanaffus", "ikki_smenali_4soatdan_uzoq",
        "ikki_smenali_istisno_ortiqcha_kun",
        "10_19_ortiqcha_kun", "10_19_limitdan_ortiq_kun",
        "10_19_yonma_yon_kun", "10_19_notekis_kun",
        "sinf_kun_taqsimoti_farqi", "sinf_qisqa_kunlari",
        "oqituvchi_ichki_okno", "oqituvchi_oknoli_smena_kun",
        "oqituvchi_kop_oknoli_smena_kun", "eng_katta_ichki_okno",
        "oqituvchi_ichki_okno_ortiqcha_ustoz",
        "oqituvchi_birlashgan_3soat_okno",
        "oqituvchi_birlashgan_2soat_okno",
        "oqituvchi_birlashgan_eng_katta_daqiqa",
        "oqituvchi_birlashgan_okno_daqiqa",
        "oqituvchi_birlashgan_okno_soni",
    )
    before_repeat_days = int(sum(
        1 for count in state.get("subject_daily", {}).values()
        if int(count or 0) >= 2
    ))
    best_by_teacher = {}
    probes = 0
    source_rows_truncated = False
    # Boshlang'ich metod kuni uchun tekshirilgan 2-soatlik variantga kamida
    # sakkizta lokal swap sinovi qolsin. Kichik test limitida esa barcha
    # probe bitta-katak tahliliga beriladi.
    single_probe_limit = (
        max_probes - 8 if max_probes >= 16 else max_probes
    )

    for teacher_row in affected[:10]:
        if probes >= single_probe_limit or _samtm_time.monotonic() >= deadline:
            break
        teacher_id = int(teacher_row["oqituvchi_id"])
        gap_days = {
            int(row.get("kun") or 0)
            for row in teacher_row.get("kunlar") or []
            if row.get("real_okno_soni") or row.get("ichki_okno")
        }
        sources = [
            placement for placement in placements
            if teacher_id in {
                int(value) for value in placement.get("teachers") or []
                if value is not None
            }
            and int(placement.get("day") or 0) in gap_days
            and _v196_movable_placement(placement)
        ]
        sources.sort(key=lambda placement: (
            int(placement.get("day") or 0),
            int((placement.get("job") or {}).get("smena") or 1),
            int(placement.get("period") or 0),
            int((placement.get("job") or {}).get("load_id") or 0),
        ))
        if len(sources) > 12:
            source_rows_truncated = True
        pair_candidates = []
        for first in sources[:12]:
            first_job = first.get("job") or {}
            class_id = int(first_job.get("sinf_id") or 0)
            first_day = int(first.get("day") or 0)
            first_shift = int(first_job.get("smena") or 1)
            first_period = int(first.get("period") or 0)
            for second in by_class.get(class_id, []):
                if second is first or not _v196_movable_placement(second):
                    continue
                second_job = second.get("job") or {}
                target_day = int(second.get("day") or 0)
                target_shift = int(second_job.get("smena") or 1)
                target_period = int(second.get("period") or 0)
                if (target_day, target_shift, target_period) == (
                    first_day, first_shift, first_period
                ):
                    continue
                method_target = (
                    teacher_id, target_day
                ) in context.get("method_hard", set())
                red_target = _v1852_blocked(
                    context.get("hard", set()), teacher_id,
                    target_day, target_shift, target_period,
                )
                if not method_target and not red_target:
                    continue
                class_row = context.get("classes", {}).get(class_id, {})
                grade = int(
                    first_job.get("v1874_grade") or _v1874_grade(class_row)
                )
                # Yuqori sinflarda metod kuni faqat katta (120+ daqiqa)
                # kutishni tuzatish uchun oxirgi variant. Boshlang'ich 1–4 da
                # esa bir yoki ikki exact soatlik metod istisnosi kichikroq
                # oynani ham yopishi mumkin; faqat Shanba qat'iy qoladi.
                if (
                    method_target
                    and int(teacher_row["jami_daqiqa"]) <= 120
                    and not 1 <= grade <= 4
                ):
                    continue
                if _v1856_class_day_block_reason(
                    class_row, target_day, context.get("class_day_blocks", {})
                ):
                    continue
                if 1 <= grade <= 4 and target_day == 6:
                    continue
                if grade <= 4 and target_period >= 6:
                    continue
                rows = _v214_effective_availability_rows(
                    availability_rows, teacher_id, target_day,
                    target_shift, target_period,
                )
                write_plan = _v214_single_slot_write_plan(
                    rows, context, teacher_id, target_day,
                    target_shift, target_period,
                )
                # Manba idlari va qolgan kataklarni saqlaydigan to'liq write
                # plan bo'lmasa foydalanuvchi what-if holatini qaytara olmaydi.
                # Bunday variant raqamli jihatdan yaxshi bo'lsa ham aytilmaydi.
                if write_plan is None:
                    continue
                pair_candidates.append((
                    1 if method_target else 0,
                    0 if target_day == first_day else 1,
                    target_day, target_shift, target_period,
                    first, second, rows, write_plan,
                ))

        for _, _, target_day, target_shift, target_period, first, second, rows, write_plan in sorted(
            pair_candidates,
            key=lambda item: (
                item[0], item[1], item[2], item[3], item[4],
                int((item[5].get("job") or {}).get("load_id") or 0),
                int((item[6].get("job") or {}).get("load_id") or 0),
            ),
        ):
            if probes >= single_probe_limit or _samtm_time.monotonic() >= deadline:
                break
            trial_context, kind = _v214_context_with_single_slot_exception(
                context, teacher_id, target_day, target_shift, target_period
            )
            if trial_context is None:
                continue
            trial_context["v206_deadline"] = deadline
            removed_ids = {id(first), id(second)}
            base = [
                placement for placement in placements
                if id(placement) not in removed_ids
            ]
            trial = None
            for leading, trailing in ((first, second), (second, first)):
                if probes >= single_probe_limit or _samtm_time.monotonic() >= deadline:
                    break
                probes += 1
                candidate_state = _v1852_rebuild_schedule_state(
                    base, trial_context
                )
                targets = {
                    id(first): (int(second.get("day") or 0), int(second.get("period") or 0)),
                    id(second): (int(first.get("day") or 0), int(first.get("period") or 0)),
                }
                lead_day, lead_period = targets[id(leading)]
                tail_day, tail_period = targets[id(trailing)]
                if not _v196_place_exact(
                    leading.get("job") or {}, lead_day, lead_period,
                    candidate_state, trial_context,
                    _v1852_random.Random(teacher_id * 1009 + probes),
                    expected_teachers=leading.get("teachers"),
                ):
                    continue
                if not _v196_place_exact(
                    trailing.get("job") or {}, tail_day, tail_period,
                    candidate_state, trial_context,
                    _v1852_random.Random(teacher_id * 1013 + probes),
                    expected_teachers=trailing.get("teachers"),
                ):
                    continue
                trial = candidate_state
                break
            if trial is None or len(trial.get("placements", [])) != len(placements):
                continue
            if _v196_class_gap_count(trial) != _v196_class_gap_count(state):
                continue

            changed_teachers = {
                int(value)
                for placement in (first, second)
                for value in (placement.get("teachers") or [])
                if value is not None
            }
            after_snapshots = {
                value: _v214_teacher_phase_window_snapshot(
                    trial, trial_context, value
                )
                for value in changed_teachers
            }
            before_target = snapshots[teacher_id]
            after_target = after_snapshots.get(teacher_id, before_target)
            target_improved = snapshot_not_worse(
                after_target, before_target, require_improvement=True
            )
            if not target_improved:
                continue
            other_worsened = any(
                not snapshot_not_worse(
                    after_snapshots[value], snapshots[value]
                )
                for value in changed_teachers if value != teacher_id
            )
            if other_worsened:
                continue
            after_metrics = _v196_attempt_metrics(trial, trial_context)
            if any(
                int(after_metrics.get(key, 0)) > int(before_metrics.get(key, 0))
                for key in protected_metric_keys
            ):
                continue
            after_repeat_days = int(sum(
                1 for count in trial.get("subject_daily", {}).values()
                if int(count or 0) >= 2
            ))
            if after_repeat_days > before_repeat_days:
                continue

            first_job = first.get("job") or {}
            second_job = second.get("job") or {}
            class_row = context.get("classes", {}).get(first_job.get("sinf_id"), {})
            class_name = f"{class_row.get('sinf','')}-{class_row.get('harf','')}"
            availability_ids = list(write_plan["manba_idlar"])
            action = (
                f"{_V1852_HAFTA.get(target_day, target_day)} kuni "
                f"{target_shift}-smena {target_period}-dars uchun "
                + (
                    "metod kuni va ustma-ust qizil manba qatorlarining "
                    "barchasini yozish rejasiga ko'ra o'chirib/ajratib, "
                    "qamrovdagi boshqa kataklarni exact qizil qiling; faqat "
                    "shu target katakni yashil/yumshoq qiling."
                    if kind == "metod_va_qizil"
                    else
                    "metod kuni qatorini exact qizil kataklarga ajratib, faqat "
                    "shu katakni yashil/yumshoq qiling; qolgan kataklar yopiq qolsin."
                    if kind == "metod_kuni"
                    else "ustma-ust qizil manba qatorlarining barchasini "
                    "yozish rejasiga ko'ra o'chirib/ajrating: faqat shu "
                    "target katakni yashil/yumshoq, qolgan qamrovni exact "
                    "qizil/qattiq qiling."
                )
            )
            suggestion = {
                "oqituvchi_id": teacher_id,
                "full_name": teacher_row["full_name"],
                "turi": kind,
                "availability_idlar": availability_ids,
                "yozish_rejasi": write_plan,
                "tahrir_rejasi": write_plan,
                "kun": int(target_day),
                "kun_nomi": _V1852_HAFTA.get(target_day, str(target_day)),
                "smena": int(target_shift),
                "dars_raqami": int(target_period),
                "dars_raqamlari": [int(target_period)],
                "istisno_kataklari": [{
                    "smena": int(target_shift),
                    "dars_raqami": int(target_period),
                }],
                "istisno_soati": 1,
                "boshlangich_sinf": bool(1 <= grade <= 4),
                "oldin_okno": int(before_target["okno_soni"]),
                "keyin_okno": int(after_target["okno_soni"]),
                "kamayish": int(before_target["okno_soni"] - after_target["okno_soni"]),
                "oldin_ichki_okno": int(before_target["ichki_okno"]),
                "keyin_ichki_okno": int(after_target["ichki_okno"]),
                "oldin_daqiqa": int(before_target["jami_daqiqa"]),
                "keyin_daqiqa": int(after_target["jami_daqiqa"]),
                "daqiqa_kamayishi": int(before_target["jami_daqiqa"] - after_target["jami_daqiqa"]),
                "isbotlangan": True,
                "ishonch": "local_proven",
                "qayta_yaratish_shart": True,
                "fan": str(first_job.get("fan") or ""),
                "sinf": class_name,
                "amal_matni": action,
                "sabab": (
                    f"{class_name} {first_job.get('fan','')} darsi "
                    f"{int(first.get('day') or 0)}-kun {int(first.get('period') or 0)}-darsdan "
                    f"{target_day}-kun {target_period}-darsga o'tkazilib, "
                    f"{second_job.get('fan','')} bilan xavfsiz almashtirildi. "
                    "Sinf, xona, smena, parallel ustoz va qattiq qoidalar "
                    "tekshirildi; asosiy/og'ir/amaliy fan, kun taqsimoti, "
                    "takror va o'qituvchi qulayligi ko'rsatkichlari yomonlashmadi."
                ),
            }
            current_best = best_by_teacher.get(teacher_id)
            suggestion_rank = (
                -int(suggestion["daqiqa_kamayishi"]),
                -int(suggestion["kamayish"]),
                1 if kind in {"metod_kuni", "metod_va_qizil"} else 0,
                int(target_day), int(target_shift), int(target_period),
            )
            if current_best is None or suggestion_rank < current_best[0]:
                best_by_teacher[teacher_id] = (
                    suggestion_rank, suggestion, trial, trial_context,
                    set(changed_teachers),
                )

    # Boshlang'ich sinf uchun Shanbadan boshqa metod kunida foyda bergan
    # 1-soatlik variant topilsa, o'sha kunning yana bitta aniq katagini ochish
    # ham alohida SINOV qilinadi. Ikki katak birdan taxmin qilinmaydi: birinchi
    # valid trial ustiga ikkinchi exact swap qo'yilib, oldin→keyin qayta
    # o'lchanadi. Qolgan metod kuni yopiq qoladi.
    double_suggestions = []
    for teacher_id, best_value in sorted(best_by_teacher.items()):
        if probes >= max_probes or _samtm_time.monotonic() >= deadline:
            break
        _, single, first_trial, first_trial_context, first_changed = best_value
        if (
            single.get("turi") not in {"metod_kuni", "metod_va_qizil"}
            or not single.get("boshlangich_sinf")
            or int(single.get("kun") or 0) == 6
        ):
            continue
        first_snapshot = _v214_teacher_phase_window_snapshot(
            first_trial, first_trial_context, teacher_id
        )
        if not first_snapshot["okno_soni"] and not first_snapshot["ichki_okno"]:
            continue
        gap_days = {
            int(row.get("kun") or 0)
            for row in first_snapshot.get("kunlar") or []
            if row.get("real_okno_soni") or row.get("ichki_okno")
        }
        current_placements = list(first_trial.get("placements", []))
        current_by_class = _v1852_defaultdict(list)
        for placement in current_placements:
            current_by_class[int(
                (placement.get("job") or {}).get("sinf_id") or 0
            )].append(placement)
        second_pairs = []
        used_slots = {
            (int(row.get("smena") or 0), int(row.get("dars_raqami") or 0))
            for row in single.get("istisno_kataklari") or []
        }
        for first in current_placements:
            first_job = first.get("job") or {}
            if (
                teacher_id not in {
                    int(value) for value in first.get("teachers") or []
                    if value is not None
                }
                or int(first.get("day") or 0) not in gap_days
                or not _v196_movable_placement(first)
            ):
                continue
            class_id = int(first_job.get("sinf_id") or 0)
            class_row = context.get("classes", {}).get(class_id, {})
            grade = int(
                first_job.get("v1874_grade") or _v1874_grade(class_row)
            )
            if not 1 <= grade <= 4:
                continue
            for second in current_by_class.get(class_id, []):
                if second is first or not _v196_movable_placement(second):
                    continue
                second_job = second.get("job") or {}
                target_day = int(second.get("day") or 0)
                target_shift = int(second_job.get("smena") or 1)
                target_period = int(second.get("period") or 0)
                if target_day != int(single.get("kun") or 0):
                    continue
                if target_day == 6 or target_period >= 6:
                    continue
                if (target_shift, target_period) in used_slots:
                    continue
                if not _v1852_blocked(
                    first_trial_context.get("hard", set()), teacher_id,
                    target_day, target_shift, target_period,
                ):
                    continue
                if _v1856_class_day_block_reason(
                    class_row, target_day,
                    first_trial_context.get("class_day_blocks", {}),
                ):
                    continue
                second_pairs.append((
                    target_day, target_shift, target_period, first, second,
                ))

        best_double = None
        for target_day, target_shift, target_period, first, second in sorted(
            second_pairs,
            key=lambda item: (
                item[0], item[1], item[2],
                int((item[3].get("job") or {}).get("load_id") or 0),
                int((item[4].get("job") or {}).get("load_id") or 0),
            ),
        ):
            if probes >= max_probes or _samtm_time.monotonic() >= deadline:
                break
            first_exception = list(single.get("istisno_kataklari") or [])
            if not first_exception:
                continue
            first_exception_shift = int(first_exception[0].get("smena") or 0)
            first_exception_period = int(
                first_exception[0].get("dars_raqami") or 0
            )
            combined_source_rows = {}
            for source in (
                _v214_effective_availability_rows(
                    availability_rows, teacher_id, target_day,
                    first_exception_shift, first_exception_period,
                )
                + _v214_effective_availability_rows(
                    availability_rows, teacher_id, target_day,
                    target_shift, target_period,
                )
            ):
                if source.get("id") is not None:
                    combined_source_rows[int(source["id"])] = dict(source)
            combined_write_plan = _v214_single_slot_write_plan(
                list(combined_source_rows.values()), context,
                teacher_id, target_day,
                first_exception_shift, first_exception_period,
                additional_slots=[(target_shift, target_period)],
            )
            if combined_write_plan is None:
                continue
            second_context, _ = _v214_context_with_single_slot_exception(
                first_trial_context, teacher_id,
                target_day, target_shift, target_period,
            )
            if second_context is None:
                continue
            second_context["v206_deadline"] = deadline
            removed_ids = {id(first), id(second)}
            base = [
                placement for placement in current_placements
                if id(placement) not in removed_ids
            ]
            second_trial = None
            for leading, trailing in ((first, second), (second, first)):
                if probes >= max_probes or _samtm_time.monotonic() >= deadline:
                    break
                probes += 1
                candidate_state = _v1852_rebuild_schedule_state(
                    base, second_context
                )
                targets = {
                    id(first): (
                        int(second.get("day") or 0),
                        int(second.get("period") or 0),
                    ),
                    id(second): (
                        int(first.get("day") or 0),
                        int(first.get("period") or 0),
                    ),
                }
                lead_day, lead_period = targets[id(leading)]
                tail_day, tail_period = targets[id(trailing)]
                if not _v196_place_exact(
                    leading.get("job") or {}, lead_day, lead_period,
                    candidate_state, second_context,
                    _v1852_random.Random(teacher_id * 2017 + probes),
                    expected_teachers=leading.get("teachers"),
                ):
                    continue
                if not _v196_place_exact(
                    trailing.get("job") or {}, tail_day, tail_period,
                    candidate_state, second_context,
                    _v1852_random.Random(teacher_id * 2027 + probes),
                    expected_teachers=trailing.get("teachers"),
                ):
                    continue
                second_trial = candidate_state
                break
            if (
                second_trial is None
                or len(second_trial.get("placements", [])) != len(placements)
                or _v196_class_gap_count(second_trial)
                != _v196_class_gap_count(state)
            ):
                continue
            changed_teachers = set(first_changed) | {
                int(value)
                for placement in (first, second)
                for value in (placement.get("teachers") or [])
                if value is not None
            }
            after_snapshots = {
                value: _v214_teacher_phase_window_snapshot(
                    second_trial, second_context, value
                )
                for value in changed_teachers
            }
            after_target = after_snapshots.get(teacher_id, first_snapshot)
            further_improved = snapshot_not_worse(
                after_target, first_snapshot, require_improvement=True
            )
            if not further_improved:
                continue
            if any(
                not snapshot_not_worse(
                    after_snapshots[value], snapshots[value]
                )
                for value in changed_teachers if value != teacher_id
            ):
                continue
            after_metrics = _v196_attempt_metrics(second_trial, second_context)
            if any(
                int(after_metrics.get(key, 0)) > int(before_metrics.get(key, 0))
                for key in protected_metric_keys
            ):
                continue
            after_repeat_days = int(sum(
                1 for count in second_trial.get("subject_daily", {}).values()
                if int(count or 0) >= 2
            ))
            if after_repeat_days > before_repeat_days:
                continue

            slots = sorted(
                list(single.get("istisno_kataklari") or []) + [{
                    "smena": int(target_shift),
                    "dars_raqami": int(target_period),
                }],
                key=lambda row: (
                    int(row.get("smena") or 0),
                    int(row.get("dars_raqami") or 0),
                ),
            )
            slot_text = " va ".join(
                f"{int(row['smena'])}-smena {int(row['dars_raqami'])}-dars"
                for row in slots
            )
            combined = dict(single)
            combined_types = {
                str(row.get("turi") or "")
                for row in combined_source_rows.values()
            }
            combined_kind = (
                "metod_va_qizil"
                if "metod_kuni" in combined_types and "band" in combined_types
                else "metod_kuni"
            )
            combined.update({
                "turi": combined_kind,
                "availability_idlar": sorted(combined_source_rows),
                "yozish_rejasi": combined_write_plan,
                "tahrir_rejasi": combined_write_plan,
                "dars_raqami": int(slots[0]["dars_raqami"]),
                "dars_raqamlari": [
                    int(row["dars_raqami"]) for row in slots
                ],
                "istisno_kataklari": slots,
                "istisno_soati": 2,
                "keyin_okno": int(after_target["okno_soni"]),
                "kamayish": int(
                    snapshots[teacher_id]["okno_soni"]
                    - after_target["okno_soni"]
                ),
                "keyin_ichki_okno": int(after_target["ichki_okno"]),
                "keyin_daqiqa": int(after_target["jami_daqiqa"]),
                "daqiqa_kamayishi": int(
                    snapshots[teacher_id]["jami_daqiqa"]
                    - after_target["jami_daqiqa"]
                ),
                "amal_matni": (
                    f"{_V1852_HAFTA.get(target_day, target_day)} "
                    + (
                        "metod kuni va ustma-ust qizil manba qatorlarining "
                        "barchasini yozish rejasiga ko'ra o'chirib/ajratib, "
                        if combined_kind == "metod_va_qizil"
                        else "metod kuni manba qatorini yozish rejasiga "
                        "ko'ra o'chirib, "
                    )
                    + f"faqat {slot_text} target kataklarini yashil/yumshoq "
                    "qiling; qamrovdagi boshqa kataklarni exact qizil/qattiq "
                    "qilib qayta yarating."
                ),
                "sabab": (
                    str(single.get("sabab") or "")
                    + " Ikkinchi boshlang'ich dars ham alohida exact swapda "
                    "tekshirildi; ikkala istisno birga qo'llanganda boshqa "
                    "ustoz yoki pedagogik ko'rsatkich yomonlashmadi."
                ),
            })
            rank = (
                -int(combined["daqiqa_kamayishi"]),
                -int(combined["kamayish"]),
                int(target_shift), int(target_period),
            )
            if best_double is None or rank < best_double[0]:
                best_double = (rank, combined)
        if best_double is not None:
            double_suggestions.append(best_double[1])

    suggestions = [value[1] for value in best_by_teacher.values()]
    suggestions.extend(double_suggestions)
    suggestions.sort(key=lambda row: (
        -int(row["daqiqa_kamayishi"]), -int(row["kamayish"]),
        str(row["full_name"]), int(row["kun"]),
        int(row["smena"]), int(row["dars_raqami"]),
    ))
    omitted_suggestions = max(0, len(suggestions) - max_suggestions)
    result["tavsiyalar"] = suggestions[:max_suggestions]
    result["tekshirilgan_variantlar"] = int(probes)
    if len(affected) > 10:
        result["cheklash_sabablari"].append(
            f"{len(affected) - 10} o'qituvchi ustuvor limitdan tashqarida qoldi"
        )
    if source_rows_truncated:
        result["cheklash_sabablari"].append(
            "ayrim o'qituvchida 12 tadan ortiq manba dars bor"
        )
    if probes >= max_probes:
        result["cheklash_sabablari"].append(
            f"{max_probes} ta sinov limiti tugadi"
        )
    if _samtm_time.monotonic() >= deadline:
        result["cheklash_sabablari"].append("hisobot vaqt limiti tugadi")
    if omitted_suggestions:
        result["cheklash_sabablari"].append(
            f"{omitted_suggestions} ta pastroq ustuvor tavsiya ko'rsatilmagan"
        )
    # Oynasi mavjud bo'lsa bu har doim chegaralangan lokal dalil; butun yechim
    # fazosi to'liq ko'rildi degan yolg'on signal berilmaydi.
    result["cheklangan"] = bool(affected)
    result["hisobot_toliq"] = not bool(affected)
    if not result["tavsiyalar"]:
        result["izoh"] += (
            " Chegaralangan lokal sinovlarda aniq foydali tahrir "
            "isbotlanmadi; tekshirilmagan variantlar mavjud."
        )
    return result


def _v196_attempt_metrics(state, context):
    single_teacher_days = sum(
        1 for periods in state.get("teacher_periods", {}).values()
        if len(set(periods)) == 1
    )
    teacher_active_days = len({
        (int(teacher), int(day))
        for (teacher, day), count in state.get("teacher_daily", {}).items()
        if float(count or 0) > 0
    })
    heavy_pairs = 0
    pe_before_heavy = 0
    upper_first_heavy = 0
    late_core = 0
    early_practical = 0
    core_period6_by_class = _v1852_defaultdict(set)
    by_class_day = state.get("class_period_jobs", {})
    for (class_id, class_day), period_jobs in by_class_day.items():
        for period, job in period_jobs.items():
            profile = _v196_rotation_profile(job, context)
            grade = int(
                job.get("v1874_grade")
                or _v1874_grade(context.get("classes", {}).get(job.get("sinf_id"), {}))
            )
            if grade >= 9 and int(period) == 1 and profile.get("heavy"):
                upper_first_heavy += 1
            if int(period) >= 5 and profile.get("core_priority"):
                late_core += 1
            if int(period) == 6 and profile.get("core_priority"):
                core_period6_by_class[int(class_id)].add(int(class_day))
            if int(period) <= 2 and (
                profile.get("physical") or profile.get("technology")
            ):
                early_practical += 1
            next_job = period_jobs.get(int(period) + 1)
            if not next_job:
                continue
            next_profile = _v196_rotation_profile(next_job, context)
            if profile.get("heavy") and next_profile.get("heavy"):
                heavy_pairs += 1
            if profile.get("physical") and next_profile.get("written_heavy"):
                pe_before_heavy += 1
    cross_shift_blocks = 0
    cross_shift_over_one_hour = 0
    cross_shift_long_days = 0
    cross_shift_over_two_hours = 0
    cross_shift_over_four_hours = 0
    cross_shift_total_minutes = 0
    cross_shift_max_minutes = 0
    cross_shift_exception_days = _v1852_defaultdict(list)
    teacher_days = {
        (int(teacher), int(day))
        for (teacher, day), count in state.get("teacher_daily", {}).items()
        if float(count or 0) > 0
    }
    compact_extra_days = 0
    compact_overflow_days = 0
    compact_adjacent_days = 0
    compact_unbalanced_days = 0
    demands = context.get("v196_teacher_demand") or {}
    for teacher, demand in demands.items():
        demand = float(demand or 0)
        rules = context.get("rules", {}).get(
            int(teacher), context.get("default_rules", {})
        )
        target_days = _v196_teacher_target_days(demand, rules)
        active = sorted(
            day for uid, day in teacher_days if int(uid) == int(teacher)
        )
        fallback_days = _v201_teacher_fallback_days(demand, rules)
        compact_extra_days += max(0, len(active) - target_days)
        compact_overflow_days += max(0, len(active) - fallback_days)
        adjacent_pairs = sum(
            1 for left, right in zip(active, active[1:])
            if int(right) - int(left) == 1
        )
        allowed_adjacent = 0 if target_days <= 3 else 1
        compact_adjacent_days += max(0, adjacent_pairs - allowed_adjacent)
        daily_loads = [
            float(state.get("teacher_daily", {}).get((int(teacher), day), 0) or 0)
            for day in active
        ]
        if daily_loads:
            compact_unbalanced_days += max(
                0, int(math.ceil(max(daily_loads) - min(daily_loads) - 2))
            )
    for teacher, day in sorted(teacher_days):
        blocks = _v196_cross_shift_edge_blocks(state, teacher, day, context)
        minutes = _v196_cross_shift_gap_minutes(state, teacher, day, context)
        if blocks is None or minutes is None:
            continue
        cross_shift_blocks += int(blocks)
        cross_shift_total_minutes += int(minutes)
        cross_shift_max_minutes = max(cross_shift_max_minutes, int(minutes))
        if int(minutes) > 60:
            cross_shift_over_one_hour += 1
        if int(minutes) > 120:
            cross_shift_over_two_hours += 1
            cross_shift_exception_days[int(teacher)].append(
                (int(day), int(minutes))
            )
        if int(minutes) > 180:
            cross_shift_long_days += 1
        if int(minutes) > 240:
            cross_shift_over_four_hours += 1
    cross_shift_exception_overflow = sum(
        max(0, len(rows) - 1)
        for rows in cross_shift_exception_days.values()
    )
    cross_shift_policy_problems = []
    for teacher, rows in sorted(cross_shift_exception_days.items()):
        ordered_rows = sorted(rows, key=lambda item: (item[0], item[1]))
        over_four = [
            {"kun": int(day), "daqiqalar": int(minutes)}
            for day, minutes in ordered_rows if int(minutes) > 240
        ]
        if len(ordered_rows) > 1 or over_four:
            cross_shift_policy_problems.append({
                "oqituvchi_id": int(teacher),
                "120_daqiqadan_uzoq_kunlar": [
                    {"kun": int(day), "daqiqalar": int(minutes)}
                    for day, minutes in ordered_rows
                ],
                "240_daqiqadan_uzoq_kunlar": over_four,
            })
    class_imbalance, class_short_days = _v196_class_distribution_metrics(state, context)
    teacher_gap_metrics = _v196_teacher_gap_metrics(state, context)
    core_period6_days = sum(
        len(days) for days in core_period6_by_class.values()
    )
    core_period6_overflow = sum(
        max(0, len(days) - _V213_CORE_PERIOD6_LIMIT)
        for days in core_period6_by_class.values()
    )
    return {
        "oqituvchi_bitta_darsli_kun": int(single_teacher_days),
        "oqituvchi_faol_kun": int(teacher_active_days),
        "ketma_ket_ogir_fan": int(heavy_pairs),
        "jismoniydan_keyin_ogir_fan": int(pe_before_heavy),
        "9_11_birinchi_dars_ogir": int(upper_first_heavy),
        "asosiy_fan_5_6": int(late_core),
        "asosiy_fan_6_kunlari": int(core_period6_days),
        "asosiy_fan_6_limitdan_ortiq": int(core_period6_overflow),
        "amaliy_fan_1_2": int(early_practical),
        "oqituvchi_smenalar_orasi_blok": int(cross_shift_blocks),
        "oqituvchi_smenalar_orasi_daqiqa": int(cross_shift_total_minutes),
        "eng_uzoq_smena_oraligi_daqiqa": int(cross_shift_max_minutes),
        "ikki_smenali_1soatdan_uzoq": int(cross_shift_over_one_hour),
        "ikki_smenali_2soatdan_uzoq": int(cross_shift_over_two_hours),
        "ikki_smenali_uzoq_tanaffus": int(cross_shift_long_days),
        "ikki_smenali_4soatdan_uzoq": int(cross_shift_over_four_hours),
        "ikki_smenali_istisno_ortiqcha_kun": int(cross_shift_exception_overflow),
        "ikki_smenali_oraliq_muammolari": cross_shift_policy_problems,
        "10_19_ortiqcha_kun": int(compact_extra_days),
        "10_19_limitdan_ortiq_kun": int(compact_overflow_days),
        "10_19_yonma_yon_kun": int(compact_adjacent_days),
        "10_19_notekis_kun": int(compact_unbalanced_days),
        "sinf_kun_taqsimoti_farqi": class_imbalance,
        "sinf_qisqa_kunlari": class_short_days,
        **teacher_gap_metrics,
    }


# V20.4 — qattiq real-vaqt to'qnashuvi va amaliy fanlarning juft darsi.
# Bu wrapper V19.6/V19.8 dagi barcha sinf, guruh, xona va qizil vaqt
# tekshiruvlarini saqlaydi; faqat yetishmagan ikki holatni qo'shadi.
_v204_base_candidate_reasons = _v196_candidate_reasons


def _v209_week_phases_overlap(first, second):
    first = str(first or "har_hafta")
    second = str(second or "har_hafta")
    return first == "har_hafta" or second == "har_hafta" or first == second


def _v209_job_room_keys_by_phase(job, fallback_room_keys, classes):
    """A/B aylanishidagi xona resurslarini toq/juft fazaga ajratadi."""
    members = (job or {}).get("rotation_members") or []
    if not members:
        phase = str((job or {}).get("hafta_turi") or "har_hafta")
        return {phase: {key for key in (fallback_room_keys or []) if key}}
    result = _v1852_defaultdict(set)
    for member in members:
        phase = str(member.get("hafta_turi") or "har_hafta")
        member_teachers = _v199_rotation_member_teachers(member)
        for key in _v1852_room_keys(member, member_teachers, classes):
            if key:
                result[phase].add(key)
    return dict(result)


def _v209_phase_room_collision(
    incoming_job, incoming_room_keys, existing_placement, classes
):
    incoming = _v209_job_room_keys_by_phase(
        incoming_job, incoming_room_keys, classes
    )
    existing_job = (existing_placement or {}).get("job") or {}
    existing = _v209_job_room_keys_by_phase(
        existing_job,
        (existing_placement or {}).get("room_keys") or [],
        classes,
    )
    return any(
        _v209_week_phases_overlap(incoming_phase, existing_phase)
        and bool(incoming_keys.intersection(existing_keys))
        for incoming_phase, incoming_keys in incoming.items()
        for existing_phase, existing_keys in existing.items()
    )


def _v212_job_teacher_phases(job, fallback_teachers=None):
    """Job ustozlarini TOQ/JUFT/har_hafta fazasi bilan qaytaradi.

    A/B aylanishidagi ikki fan bir fizik katakda turadi, ammo ularning
    ustozlari turli haftada ishlaydi. Eski ``teacher_busy`` indeksida faza
    yo'q edi va TOQ ustoz sabab JUFT ustozning legal darsi ham yopilardi.
    """
    result = _v1852_defaultdict(set)
    members = (job or {}).get("rotation_members") or []
    if members:
        for member in members:
            phase = str(member.get("hafta_turi") or "har_hafta")
            for teacher in _v199_rotation_member_teachers(member):
                if teacher is not None:
                    result[int(teacher)].add(phase)
    else:
        phase = str((job or {}).get("hafta_turi") or "har_hafta")
        teachers = (
            list(fallback_teachers)
            if fallback_teachers is not None
            else list(_v1852_job_teacher_ids(job or {}))
        )
        for teacher in teachers:
            if teacher is not None:
                result[int(teacher)].add(phase)
    return {teacher: set(phases) for teacher, phases in result.items()}


def _v212_teacher_phase_collision(incoming_job, incoming_teachers, placement):
    """Ikki paketda ayni ustozning bir-birini kesadigan hafta fazasi bormi."""
    incoming = _v212_job_teacher_phases(incoming_job, incoming_teachers)
    existing_job = (placement or {}).get("job") or {}
    existing = _v212_job_teacher_phases(
        existing_job, (placement or {}).get("teachers") or []
    )
    return any(
        _v209_week_phases_overlap(incoming_phase, existing_phase)
        for teacher in set(incoming).intersection(existing)
        for incoming_phase in incoming[teacher]
        for existing_phase in existing[teacher]
    )


def _v205_class_hour_red_day_exception(job, teacher, day, period, context):
    """Faqat qo'lda tanlangan qat'iy Kelajak soati metod kunidan o'tadi.

    Qizil/BAND bu yordamchida ham hech qachon ochilmaydi.
    """
    if not job.get("is_class_hour") or teacher is None:
        return False
    teacher = int(teacher)
    day = int(day)
    period = int(period)
    if (
        int(job.get("fixed_day") or 0) != day
        or int(job.get("fixed_period") or 0) != period
    ):
        return False
    return (teacher, day) in context.get("method_hard", set())


def _v213_core_period6_days(state, class_id, context):
    """Sinfning asosiy fan 6-darsga tushgan aniq hafta kunlarini qaytaradi.

    A/B aylanish katagida ikki fan bitta fizik slotni egallaydi; ulardan biri
    asosiy fan bo'lsa ham kun bir marta sanaladi. Hisob ``class_period_jobs``
    dan olinadi, shuning uchun rebuild va bounded-repair holatlarida ham bir xil.
    """
    result = set()
    for (placed_class_id, placed_day), period_jobs in (
        state.get("class_period_jobs", {}) or {}
    ).items():
        if int(placed_class_id) != int(class_id):
            continue
        placed_job = (period_jobs or {}).get(6)
        if placed_job and _v196_rotation_profile(
            placed_job, context
        ).get("core_priority"):
            result.add(int(placed_day))
    return result


def _v1852_candidate_reasons(
    job, day, period, selected_teachers, room_keys, state, context
):
    reasons = list(_v204_base_candidate_reasons(
        job, day, period, selected_teachers, room_keys, state, context
    ))
    profile = _v196_rotation_profile(job, context)

    # Asosiy fanlar avval 1–5-darslardan joy oladi. Boshqa legal kombinatsiya
    # qolmasa 6-dars zaxira sifatida ishlaydi, ammo bir sinfning haftasida
    # bunday kunlar ikkitadan oshmaydi. Bu pedagogik afzallik emas, foydalanuvchi
    # belgilagan yakuniy chegara; completion bosqichi ham uni olib tashlamaydi.
    if int(period) == 6 and profile.get("core_priority"):
        core_period6_days = _v213_core_period6_days(
            state, job.get("sinf_id"), context
        )
        if (
            int(day) not in core_period6_days
            and len(core_period6_days) >= _V213_CORE_PERIOD6_LIMIT
        ):
            reasons.append(_V213_CORE_PERIOD6_REASON)

    # Eski room_busy to'plami hafta fazasini saqlamaydi. Shuning uchun u
    # toq haftadagi xona sabab juft haftadagi boshqa darsni ham yopardi.
    # Placementlar fazani saqlaydi: ayni slotda faqat kesishadigan fazadagi
    # bir xil jismoniy xona bo'lsa `xona band` qattiq sababi qoladi.
    if "xona band" in reasons:
        exact_phase_collision = any(
            int(placement.get("day") or 0) == int(day)
            and int((placement.get("job") or {}).get("smena") or 1)
            == int(job.get("smena") or 1)
            and int(placement.get("period") or 0) == int(period)
            and _v209_phase_room_collision(
                job, room_keys, placement, context.get("classes", {})
            )
            for placement in state.get("placements", [])
        )
        if not exact_phase_collision:
            reasons = [reason for reason in reasons if reason != "xona band"]

    # ``teacher_busy`` ham eski fazasiz indeks. Ayni TOQ ustozning katagi
    # JUFT haftadagi o'sha ustoz yoki boshqa aylanish paketini asossiz
    # bloklamasin; faqat kesishadigan hafta fazasi real kolliziya hisoblanadi.
    if "o'qituvchi boshqa darsda" in reasons:
        exact_teacher_collision = any(
            int(placement.get("day") or 0) == int(day)
            and int((placement.get("job") or {}).get("smena") or 1)
            == int(job.get("smena") or 1)
            and int(placement.get("period") or 0) == int(period)
            and _v212_teacher_phase_collision(
                job, selected_teachers, placement
            )
            for placement in state.get("placements", [])
        )
        if not exact_teacher_collision:
            reasons = [
                reason for reason in reasons
                if reason != "o'qituvchi boshqa darsda"
            ]

    # Sinf rahbari keyin tayinlanishi mumkin. Shunday sinfning qat'iy
    # KELAJAK/SINF SOATI katagi o'qituvchisiz bo'lsa ham yo'qolmaydi.
    if job.get("is_class_hour") and not any(
        teacher is not None for teacher in selected_teachers
    ):
        reasons = [
            reason for reason in reasons
            if reason != "o'qituvchi biriktirilmagan"
        ]

    # Faqat sinf rahbarining administrator tanlagan qat'iy Kelajak soati
    # katagi uchun metod-kuni istisno. Qizil/BAND bu maxsus darsda ham
    # ochilmaydi; parallel dars to'qnashuvi avvalgidek qattiq qoladi.
    if job.get("is_class_hour") and any(
        _v205_class_hour_red_day_exception(job, teacher, day, period, context)
        for teacher in selected_teachers
        if teacher is not None
    ):
        reasons = [
            reason for reason in reasons
            if reason != "o'qituvchining metod kuni"
        ]

    # Jismoniy tarbiya yoki texnologiyada administrator daily_max>=2 ni aniq
    # belgilagan bo'lsa, faqat bir marta yonma-yon juft dars mumkin. Avvalgi kod
    # current_count < 2 bo'lsa shunchaki olib tashlar, natijada bir haftada bir
    # nechta va hatto yonma-yon bo'lmagan takror paydo bo'lishi mumkin edi.
    if (
        profile.get("physical") or profile.get("technology")
    ) and int(job.get("daily_max") or 1) > 1:
        subject_key = str(job.get("fan") or "").casefold()
        class_id = job.get("sinf_id")
        subject_daily = state.get("subject_daily", {})
        current_count = int(subject_daily.get(
            (class_id, subject_key, int(day)), 0
        ))
        existing_periods = set(
            state.get("class_subject_periods", {}).get(
                (class_id, subject_key, int(day)), set()
            )
        )
        paired_days = {
            int(subject_day)
            for (subject_class, subject, subject_day), count
            in subject_daily.items()
            if int(subject_class) == int(class_id)
            and str(subject) == subject_key
            and int(count or 0) >= 2
        }
        adjacent_pair = bool(
            current_count == 1
            and not paired_days
            and any(abs(int(period) - int(old_period)) == 1 for old_period in existing_periods)
        )
        if adjacent_pair:
            reasons = [
                reason for reason in reasons
                if reason != "fan kunlik maksimumga yetgan"
            ]
        elif current_count >= 1 and "fan kunlik maksimumga yetgan" not in reasons:
            reasons.append("fan kunlik maksimumga yetgan")

    # Parallel dars faqat "bir xil smena+dars raqami" emas. Smenalar
    # vaqti bir-biriga tegib yoki ustma-ust kelsa ham o'qituvchi ikki joyda
    # bo'la olmaydi. Haqiqiy boshlanish/tugash daqiqalari qattiq tekshiriladi.
    new_interval = _v196_slot_interval(
        context, int(job.get("smena") or 1), int(period)
    )
    if new_interval:
        new_start, new_end = new_interval
        for placement in state.get("placements", []):
            if int(placement.get("day") or 0) != int(day):
                continue
            if not _v212_teacher_phase_collision(
                job, selected_teachers, placement
            ):
                continue
            old_job = placement.get("job") or {}
            old_interval = _v196_slot_interval(
                context,
                int(old_job.get("smena") or 1),
                int(placement.get("period") or 0),
            )
            if not old_interval:
                continue
            old_start, old_end = old_interval
            if int(new_start) < int(old_end) and int(old_start) < int(new_end):
                reasons.append(
                    "o'qituvchi boshqa smenadagi dars bilan real vaqtda ustma-ust"
                )
                break
        # Xona ham dars raqami bilan emas, haqiqiy vaqt oralig'i bilan band.
        # Masalan, 1-smena 6-darsi 2-smena 1-darsi bilan ustma-ust tushishi
        # mumkin. Kanonik room:<id>/text:<nom> kalitlari bir xil jismoniy
        # xonaning katalog va sinf matni ko'rinishlarini birlashtiradi.
        requested_rooms = {key for key in room_keys if key}
        if requested_rooms:
            for placement in state.get("placements", []):
                if int(placement.get("day") or 0) != int(day):
                    continue
                if not _v209_phase_room_collision(
                    job, room_keys, placement, context.get("classes", {})
                ):
                    continue
                old_job = placement.get("job") or {}
                old_interval = _v196_slot_interval(
                    context,
                    int(old_job.get("smena") or 1),
                    int(placement.get("period") or 0),
                )
                if not old_interval:
                    continue
                old_start, old_end = old_interval
                if int(new_start) < int(old_end) and int(old_start) < int(new_end):
                    reasons.append(
                        "xona band: boshqa smenadagi dars bilan real vaqtda ustma-ust"
                    )
                    break
    # V20.7: faqat pedagogik tavsiyalar bosqichma-bosqich yumshaydi.
    # Qizil vaqt, parallel, sinf bandligi, smena va metod kuni modulda
    # qat'iy deb tasniflangan va hech qachon filtrdan olib tashlanmaydi.
    return _timetable_filter_reasons(
        list(dict.fromkeys(reasons)),
        context.get("v207_policy_stage", "strict"),
    )


# ═══════════════════════════════════════════════════════════════════════════
# V20.9 — ADMIN MAKTAB USTASI: BITTA ATOMAR, KANONIK BACKEND KONTRAKTI
#
# Frontend parallel sonlarini 1/2-smena uchun alohida hisoblab, tayyor sinf
# qatorlari sifatida yuboradi. Backend frontendga ishonib ketmaydi: sinf
# harflarini, bino+xona identifikatorini, xizmat xonalarini va bitta uy
# xonasining ikki sinfga berilmaganini transactiondan oldin qayta tekshiradi.
# ═══════════════════════════════════════════════════════════════════════════

SAMTM_SCHOOL_CREATION_PATCH = "v20.9-contract-b-f-safe"

_V209_MEMBER_ROLE_ORDER = {
    "sinf_rahbari": 1,
    "psixolog": 2,
    "direktor": 3,
}


class V209SchoolCreationRoom(BaseModel):
    number: str
    floor: int = 1
    turi: str = "classroom"
    darsga_yaroqli: bool = True


class V209SchoolCreationBuilding(BaseModel):
    key: str
    name: str
    floors: int = 1
    rooms: list[V209SchoolCreationRoom] = []


class V209SchoolCreationClass(BaseModel):
    name: str
    shift: int = 1
    talim_tili: str = "uz"
    leader_user_id: Optional[int] = None
    psychologist_user_id: Optional[int] = None
    building_key: Optional[str] = None
    room_number: Optional[str] = None


class V209SchoolCreationRequest(BaseModel):
    token: str
    name: str
    school_number: Optional[str] = None
    region: str
    district: str
    shift_count: int = 1
    alifbo_turi: str = "latin_xalqaro"
    director_user_id: Optional[int] = None
    buildings: list[V209SchoolCreationBuilding] = []
    classes: list[V209SchoolCreationClass] = []


def _v209_model_dict(value):
    if isinstance(value, dict):
        return dict(value)
    dump = getattr(value, "model_dump", None)
    if callable(dump):
        return dump()
    legacy_dump = getattr(value, "dict", None)
    if callable(legacy_dump):
        return legacy_dump()
    return dict(value)


def _v209_clean_text(value, maximum=160):
    return re.sub(r"\s+", " ", str(value or "")).strip()[:maximum]


def _v209_creation_signature(actor_id, school_name, school_number, region, district):
    """Double-click qulfi uchun ajralishi aniq, maktab raqamli imzo."""
    parts = [
        str(int(actor_id)),
        _v209_clean_text(school_name, 200).casefold(),
        _v209_clean_text(school_number, 40).casefold(),
        _v209_clean_text(region, 120).casefold(),
        _v209_clean_text(district, 120).casefold(),
    ]
    return "|".join(f"{len(part)}:{part}" for part in parts)


def _v209_keep_highest_member_role(member_roles, user_id, role):
    """Bir foydalanuvchining maktabdagi eng yuqori vakolatini saqlaydi."""
    if user_id is None:
        return
    if role not in _V209_MEMBER_ROLE_ORDER:
        raise ValueError(f"Noma'lum maktab roli: {role}")
    user_id = int(user_id)
    current = member_roles.get(user_id)
    if (
        current is None
        or _V209_MEMBER_ROLE_ORDER[role] > _V209_MEMBER_ROLE_ORDER[current]
    ):
        member_roles[user_id] = role


def _v209_normalize_class_name(value, alphabet_type="latin_xalqaro"):
    """`1-A`, `1-Б`, `1-Rus` yoki `1-001` ni raqam + yorliqqa ajratadi."""
    raw = _v209_clean_text(value, 64)
    # Raqam bilan boshlanuvchi erkin yorliq faqat aniq ajratgichdan keyin
    # qabul qilinadi. Shu sabab oddiy ``11`` xato qilib 1-sinf + ``1``
    # paralleliga aylanmaydi, lekin ``1-001`` qonuniy bo'ladi.
    match = re.fullmatch(r"(1[01]|[1-9])\s*[-–—_]\s*(.+)", raw)
    if not match:
        match = re.fullmatch(r"(1[01]|[1-9])\s+(.+)", raw)
    if not match:
        match = re.fullmatch(r"(1[01]|[1-9])([^\d\s].*)", raw)
    if not match:
        return None
    try:
        label = _v237_clean_parallel_label(match.group(2))
    except ValueError:
        return None
    alphabet = _V237_CLASS_ALPHABETS.get(
        str(alphabet_type or "latin_xalqaro").strip().lower(),
        _V237_CLASS_ALPHABETS["latin_xalqaro"],
    )
    canonical = {
        _v237_parallel_label_key(item): item for item in alphabet
    }.get(_v237_parallel_label_key(label), label)
    return int(match.group(1)), canonical


def _v209_normalize_materialized_classes(raw_classes, shift_count,
                                         alphabet_type="latin_xalqaro"):
    """Frontend sinflarini Unicode alifbo va erkin yorliqlar bilan tekshiradi."""
    try:
        shift_count = int(shift_count)
    except (TypeError, ValueError) as exc:
        raise ValueError("Smena soni 1 yoki 2 bo‘lishi kerak.") from exc
    if shift_count not in (1, 2):
        raise ValueError("Smena soni 1 yoki 2 bo‘lishi kerak.")
    alphabet_type = str(alphabet_type or "latin_xalqaro").strip().lower()
    if alphabet_type not in _V237_CLASS_ALPHABETS:
        raise ValueError("Sinf alifbosi noto‘g‘ri.")
    alphabet = _V237_CLASS_ALPHABETS[alphabet_type]
    alphabet_indexes = {
        _v237_parallel_label_key(label): index
        for index, label in enumerate(alphabet)
    }
    rows = [_v209_model_dict(item) for item in (raw_classes or [])]
    if not rows:
        raise ValueError("Kamida bitta sinf yarating.")

    normalized = []
    names = set()
    by_grade = _v1852_defaultdict(list)
    for item in rows:
        parsed = _v209_normalize_class_name(item.get("name"), alphabet_type)
        if parsed is None:
            raise ValueError(
                f"“{_v209_clean_text(item.get('name'), 64) or '?'}” sinf nomi noto‘g‘ri. "
                "Masalan 1-A, 1-Б yoki 1-Rus ko‘rinishida kiriting."
            )
        grade, letter = parsed
        name = f"{grade}-{letter}"
        name_key = (grade, _v237_parallel_label_key(letter))
        if name_key in names:
            raise ValueError(f"{name} sinfi ikki marta yuborilgan.")
        names.add(name_key)
        try:
            shift = int(item.get("shift") or 1)
        except (TypeError, ValueError) as exc:
            raise ValueError(f"{name} uchun smena noto‘g‘ri.") from exc
        if shift not in (1, 2) or (shift_count == 1 and shift != 1):
            raise ValueError(
                f"{name} uchun {shift}-smena maktabning {shift_count} smena sozlamasiga mos emas."
            )
        try:
            instruction_language = _v238_normalize_instruction_language(
                item.get("talim_tili")
            )
        except ValueError as exc:
            raise ValueError(f"{name}: {exc}") from exc
        row = {
            **item,
            "name": name,
            "grade": grade,
            "letter": letter,
            "shift": shift,
            "talim_tili": instruction_language,
        }
        normalized.append(row)
        by_grade[grade].append(row)

    for grade, grade_rows in sorted(by_grade.items()):
        ordered = sorted(
            grade_rows,
            key=lambda row: _v237_class_sort_key({
                "id": 0, "sinf": grade, "harf": row["letter"],
            }, alphabet_type),
        )
        actual_keys = [_v237_parallel_label_key(row["letter"]) for row in ordered]
        # Tanlangan alifbodagi standart harflar ishlatilsa A/B/... yoki
        # А/Б/... uzluksizligi saqlanadi. Erkin nomlar (Rus, STEM, A-1)
        # esa takrorlanmasa qabul qilinadi va tabiiy tartiblanadi.
        if all(key in alphabet_indexes for key in actual_keys):
            expected = list(alphabet[:len(ordered)])
            expected_keys = [_v237_parallel_label_key(value) for value in expected]
        else:
            expected = []
            expected_keys = actual_keys
        if actual_keys != expected_keys:
            raise ValueError(
                f"{grade}-sinf parallellari uzluksiz emas: "
                f"kutilgan {', '.join(expected)}, yuborilgan "
                f"{', '.join(row['letter'] for row in ordered)}."
            )
        # 1-smena harflari tugagach 2-smena davom etadi; A/C ni birinchi,
        # B/D ni ikkinchi smenaga aralashtirib yuborish qabul qilinmaydi.
        second_started = False
        for row in ordered:
            if row["shift"] == 2:
                second_started = True
            elif second_started:
                raise ValueError(
                    f"{grade}-sinf smenalari harflar bo‘yicha uzluksiz emas: "
                    "avval 1-smena, keyin 2-smena bo‘lishi kerak."
                )
    return sorted(
        normalized,
        key=lambda row: _v237_class_sort_key({
            "id": 0, "sinf": row["grade"], "harf": row["letter"],
        }, alphabet_type),
    )


def _v209_normalize_buildings(raw_buildings):
    buildings = []
    key_seen = set()
    name_seen = set()
    catalog_seen = {}
    rooms = {}
    for source in raw_buildings or []:
        item = _v209_model_dict(source)
        client_key = _v209_clean_text(item.get("key"), 160)
        name = _v209_clean_text(item.get("name"), 160)
        if not client_key or not name:
            raise ValueError("Har bir binoning kaliti va nomi bo‘lishi kerak.")
        key_norm = client_key.casefold()
        name_norm = name.casefold()
        if key_norm in key_seen:
            raise ValueError(f"{name} bino kaliti ikki marta yuborilgan.")
        if name_norm in name_seen:
            raise ValueError(f"{name} bino nomi ikki marta yuborilgan.")
        key_seen.add(key_norm)
        name_seen.add(name_norm)
        try:
            floors = int(item.get("floors") or 1)
        except (TypeError, ValueError) as exc:
            raise ValueError(f"{name} qavat soni noto‘g‘ri.") from exc
        if floors not in range(1, 21):
            raise ValueError(f"{name} qavat soni 1–20 oralig‘ida bo‘lishi kerak.")

        room_sources = item.get("rooms") or []
        if not room_sources:
            raise ValueError(
                f"{name} binosida kamida bitta xona bo‘lishi kerak. "
                "Aks holda bino va xonalarni keyin kiritishni tanlang."
            )

        building_rooms = []
        local_seen = set()
        for room_source in room_sources:
            room = _v209_model_dict(room_source)
            number = _v209_clean_text(room.get("number"), 100)
            if not number:
                raise ValueError(f"{name} binosida xona raqami bo‘sh.")
            if re.search(r"[;,\n]", str(room.get("number") or "")):
                raise ValueError(
                    f"{name} / {number}: har bir xona alohida qator bo‘lishi kerak."
                )
            number_norm = number.casefold()
            if number_norm in local_seen:
                raise ValueError(f"{name} / {number} xonasi ikki marta yuborilgan.")
            local_seen.add(number_norm)
            try:
                floor = int(room.get("floor") or 1)
            except (TypeError, ValueError) as exc:
                raise ValueError(f"{name} / {number} qavati noto‘g‘ri.") from exc
            if floor not in range(1, floors + 1):
                raise ValueError(
                    f"{name} / {number}: qavat 1–{floors} oralig‘ida bo‘lishi kerak."
                )
            room_type = _v209_clean_text(room.get("turi") or "classroom", 32).lower()
            teaching = bool(room.get("darsga_yaroqli", True))
            if room_type == "non_teaching" or not teaching:
                room_type = "non_teaching"
                teaching = False
            elif room_type in {"oddiy", "classroom"}:
                room_type = "classroom"
            elif room_type in {"maxsus", "reserve", "sport"}:
                room_type = {"maxsus": "reserve"}.get(room_type, room_type)
            else:
                raise ValueError(f"{name} / {number}: xona turi noto‘g‘ri.")
            catalog_name = f"{name} {number}"
            catalog_norm = catalog_name.casefold()
            previous_room = catalog_seen.get(catalog_norm)
            if previous_room is not None:
                raise ValueError(
                    "Xona katalog nomi to‘qnashdi: "
                    f"{previous_room['building']} / {previous_room['number']} va "
                    f"{name} / {number}. Bino yoki xona nomini aniqlashtiring."
                )
            catalog_seen[catalog_norm] = {
                "building": name,
                "number": number,
            }
            normalized_room = {
                "number": number,
                "number_norm": number_norm,
                "floor": floor,
                "turi": room_type,
                "darsga_yaroqli": teaching,
                # Eski generator aynan `bino + xona` matnini katalogdan
                # qidiradi. Shu nom ikki binodagi bir xil 101 ni ajratadi.
                "catalog_name": catalog_name,
                "building_key_norm": key_norm,
            }
            building_rooms.append(normalized_room)
            rooms[(key_norm, number_norm)] = normalized_room
        buildings.append({
            "key": client_key,
            "key_norm": key_norm,
            "name": name,
            "floors": floors,
            "rooms": building_rooms,
        })
    return buildings, rooms


def _v209_validate_home_rooms(classes, buildings, rooms):
    building_map = {item["key_norm"]: item for item in buildings}
    occupied = {}
    result = []
    for source in classes:
        row = dict(source)
        building_key = _v209_clean_text(row.get("building_key"), 160)
        room_number = _v209_clean_text(row.get("room_number"), 100)
        if bool(building_key) != bool(room_number):
            raise ValueError(
                f"{row['name']} uchun bino va xona birga tanlanishi kerak."
            )
        row["home_room"] = None
        if building_key:
            building_norm = building_key.casefold()
            building = building_map.get(building_norm)
            if building is None:
                raise ValueError(f"{row['name']} uchun tanlangan bino topilmadi.")
            room_key = (building_norm, room_number.casefold())
            room = rooms.get(room_key)
            if room is None:
                raise ValueError(
                    f"{row['name']} uchun {building['name']} / {room_number} xona topilmadi."
                )
            if not room["darsga_yaroqli"] or room["turi"] == "non_teaching":
                raise ValueError(
                    f"{row['name']} uchun {building['name']} / {room_number} "
                    "dars o‘tilmaydigan xona tanlangan."
                )
            shift = 2 if int(row.get("shift") or 1) == 2 else 1
            occupied_key = (shift, room_key)
            owner = occupied.get(occupied_key)
            if owner:
                raise ValueError(
                    "Xona bir smenada takror biriktirilgan: "
                    f"{building['name']} / {room_number}. Bitta xona "
                    f"{shift}-smenada faqat bitta sinfga beriladi "
                    f"({owner} va {row['name']})."
                )
            occupied[occupied_key] = row["name"]
            row["home_room"] = {"building": building, "room": room}
        result.append(row)
    return result


def _v209_school_creation_tables(cur):
    _maktab_jadvali(cur)
    _maktab_sinflari_jadvali(cur)
    _muassasa_jadvali(cur)
    _v1852_tables(cur)
    cur.execute("ALTER TABLE maktablar ADD COLUMN IF NOT EXISTS maktab_raqami TEXT")
    cur.execute("ALTER TABLE maktab_sinflari ADD COLUMN IF NOT EXISTS smena INTEGER DEFAULT 1")
    cur.execute(
        "ALTER TABLE maktab_sinflari ADD COLUMN IF NOT EXISTS "
        "talim_tili TEXT NOT NULL DEFAULT 'uz'"
    )
    cur.execute("ALTER TABLE maktab_sinflari ADD COLUMN IF NOT EXISTS bino TEXT")
    cur.execute("ALTER TABLE maktab_sinflari ADD COLUMN IF NOT EXISTS xona TEXT")
    cur.execute("ALTER TABLE maktab_sinflari ADD COLUMN IF NOT EXISTS bino_id BIGINT")
    cur.execute("ALTER TABLE maktab_sinflari ADD COLUMN IF NOT EXISTS xona_id BIGINT")
    cur.execute("ALTER TABLE maktab_sinflari ADD COLUMN IF NOT EXISTS psixolog_user_id BIGINT")
    cur.execute("""CREATE TABLE IF NOT EXISTS aqlli_binolar_v2(
        id BIGSERIAL PRIMARY KEY,
        maktab_id INTEGER NOT NULL REFERENCES maktablar(id) ON DELETE CASCADE,
        client_key TEXT NOT NULL,
        nomi TEXT NOT NULL,
        qavat_soni INTEGER NOT NULL DEFAULT 1 CHECK(qavat_soni BETWEEN 1 AND 20),
        faol BOOLEAN NOT NULL DEFAULT TRUE,
        yaratilgan_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
        UNIQUE(maktab_id,client_key)
    )""")
    cur.execute("ALTER TABLE aqlli_xonalar_v2 ADD COLUMN IF NOT EXISTS bino_id BIGINT")
    cur.execute("ALTER TABLE aqlli_xonalar_v2 ADD COLUMN IF NOT EXISTS xona_raqami TEXT")
    cur.execute("ALTER TABLE aqlli_xonalar_v2 ADD COLUMN IF NOT EXISTS qavat INTEGER")
    cur.execute("""CREATE UNIQUE INDEX IF NOT EXISTS uq_aqlli_xona_bino_raqam_v209
                   ON aqlli_xonalar_v2(maktab_id,bino_id,lower(trim(xona_raqami)))
                   WHERE bino_id IS NOT NULL AND xona_raqami IS NOT NULL""")


def _v209_require_known_users(cur, user_ids):
    requested = {int(value) for value in user_ids if value is not None}
    if not requested:
        return
    cur.execute(
        "SELECT user_id FROM users WHERE user_id = ANY(%s)",
        (sorted(requested),),
    )
    found = {int(row["user_id"]) for row in cur.fetchall()}
    missing = sorted(requested - found)
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Tanlangan foydalanuvchi topilmadi: ID {', '.join(map(str, missing))}.",
        )


def _v209_attach_school_member(cur, user_id, school_id, role):
    if user_id is None:
        return
    user_id = int(user_id)
    cur.execute(
        """INSERT INTO foydalanuvchi_muassasalari(
               user_id,muassasa_turi,muassasa_id,lavozim
           ) VALUES(%s,'maktab',%s,%s)
           ON CONFLICT(user_id,muassasa_turi,muassasa_id)
           DO UPDATE SET lavozim=EXCLUDED.lavozim""",
        (user_id, school_id, role),
    )
    # Ko'p maktabli foydalanuvchining faol workspace'ini zo'rlab
    # almashtirmaymiz; faqat hali maktabsiz bo'lsa yangi maktabni faollashtiramiz.
    cur.execute(
        """UPDATE users
              SET maktab_id=COALESCE(maktab_id,%s),
                  lavozim=COALESCE(NULLIF(lavozim,''),%s)
            WHERE user_id=%s""",
        (school_id, role, user_id),
    )


def _v209_remove_earlier_school_creation_route():
    """Starlette birinchi mos route'ni oladi; eski POST ni aniq almashtiramiz."""
    router = getattr(app, "router", None)
    routes = list(getattr(router, "routes", []) or [])
    if router is None:
        return
    router.routes[:] = [
        route for route in routes
        if not (
            getattr(route, "path", None) == "/api/admin/maktab-yaratish-v2"
            and "POST" in (getattr(route, "methods", set()) or set())
        )
    ]


_v209_remove_earlier_school_creation_route()


@app.post("/api/admin/maktab-yaratish-v2")
def v209_admin_create_school(sorov: V209SchoolCreationRequest):
    """Maktab, bino, xona va sinflarni bitta transactionda yaratadi."""
    actor_id = _admin_tekshir(sorov.token)
    school_name = _v209_clean_text(sorov.name, 200)
    region = _v209_clean_text(sorov.region, 120)
    district = _v209_clean_text(sorov.district, 120)
    school_number = _v209_clean_text(sorov.school_number, 40) or None
    if len(school_name) < 2:
        raise HTTPException(status_code=400, detail="Maktab nomini to‘liq kiriting.")
    if not region or not district:
        raise HTTPException(status_code=400, detail="Viloyat va tumanni tanlang.")
    try:
        alphabet_type = str(sorov.alifbo_turi or "latin_xalqaro").strip().lower()
        classes = _v209_normalize_materialized_classes(
            sorov.classes, sorov.shift_count, alphabet_type
        )
        buildings, rooms = _v209_normalize_buildings(sorov.buildings)
        classes = _v209_validate_home_rooms(classes, buildings, rooms)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    selected_users = [sorov.director_user_id]
    for row in classes:
        selected_users.extend((
            row.get("leader_user_id"), row.get("psychologist_user_id")
        ))

    conn = _db()
    cur = conn.cursor()
    try:
        _v209_school_creation_tables(cur)
        _v209_require_known_users(cur, selected_users)
        # Bitta adminning tasodifiy double-clicki ikki aynan teng maktab
        # yaratmasin. Advisory lock transaction oxirigacha shu imzoni yopadi.
        creation_signature = _v209_creation_signature(
            actor_id, school_name, school_number, region, district
        )
        cur.execute("SELECT pg_advisory_xact_lock(hashtext(%s))", (creation_signature,))
        cur.execute(
            """SELECT id FROM maktablar
                WHERE lower(trim(nomi))=lower(trim(%s))
                  AND lower(trim(COALESCE(maktab_raqami,'')))=lower(trim(%s))
                  AND lower(trim(COALESCE(viloyat,'')))=lower(trim(%s))
                  AND lower(trim(COALESCE(tuman,'')))=lower(trim(%s))
                  AND yaratilgan_at >= NOW() - INTERVAL '2 minutes'
                ORDER BY id DESC LIMIT 1""",
            (school_name, school_number or "", region, district),
        )
        if cur.fetchone():
            raise HTTPException(
                status_code=409,
                detail="Aynan shu maktab hozirgina yaratildi. Sahifani yangilab ro‘yxatdan oching.",
            )

        cur.execute(
            """INSERT INTO maktablar(
                   nomi,maktab_raqami,viloyat,tuman,smena_soni,
                   direktor_user_id,pulli,oylik_tolov,alifbo_turi
               ) VALUES(%s,%s,%s,%s,%s,%s,FALSE,NULL,%s)
               RETURNING id""",
            (
                school_name, school_number, region, district,
                int(sorov.shift_count), sorov.director_user_id, alphabet_type,
            ),
        )
        school_id = int(cur.fetchone()["id"])

        building_ids = {}
        room_ids = {}
        room_payload = []
        for building in buildings:
            cur.execute(
                """INSERT INTO aqlli_binolar_v2(
                       maktab_id,client_key,nomi,qavat_soni,faol
                   ) VALUES(%s,%s,%s,%s,TRUE) RETURNING id""",
                (school_id, building["key"], building["name"], building["floors"]),
            )
            building_id = int(cur.fetchone()["id"])
            building_ids[building["key_norm"]] = building_id
            for room in building["rooms"]:
                cur.execute(
                    """INSERT INTO aqlli_xonalar_v2(
                           maktab_id,nomi,turi,faol,bino_id,xona_raqami,qavat,
                           darsga_yaroqli
                       ) VALUES(%s,%s,%s,TRUE,%s,%s,%s,%s)
                       RETURNING id""",
                    (
                        school_id, room["catalog_name"], room["turi"],
                        building_id, room["number"], room["floor"],
                        room["darsga_yaroqli"],
                    ),
                )
                room_id = int(cur.fetchone()["id"])
                room_key = (building["key_norm"], room["number_norm"])
                room_ids[room_key] = room_id
                room_payload.append({
                    "id": room_id,
                    "bino_id": building_id,
                    "bino": building["name"],
                    "raqam": room["number"],
                    "nomi": room["catalog_name"],
                    "turi": room["turi"],
                    "darsga_yaroqli": room["darsga_yaroqli"],
                })

        class_payload = []
        # Bir xil foydalanuvchi bir nechta rolda tanlansa eng yuqori rol
        # oxirida yoziladi: sinf rahbari < psixolog < direktor.
        member_roles = {}
        for row in classes:
            home = row.get("home_room")
            building_id = room_id = None
            building_name = room_number = None
            if home:
                building = home["building"]
                room = home["room"]
                room_key = (building["key_norm"], room["number_norm"])
                building_id = building_ids[building["key_norm"]]
                room_id = room_ids[room_key]
                building_name = building["name"]
                room_number = room["number"]
            cur.execute(
                """INSERT INTO maktab_sinflari(
                       maktab_id,sinf,harf,smena,bino,xona,bino_id,xona_id,
                       rahbar_user_id,psixolog_user_id,talim_tili
                   ) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                   RETURNING id""",
                (
                    school_id, str(row["grade"]), row["letter"], row["shift"],
                    building_name, room_number, building_id, room_id,
                    row.get("leader_user_id"), row.get("psychologist_user_id"),
                    row["talim_tili"],
                ),
            )
            class_id = int(cur.fetchone()["id"])
            _v209_keep_highest_member_role(
                member_roles, row.get("leader_user_id"), "sinf_rahbari"
            )
            _v209_keep_highest_member_role(
                member_roles, row.get("psychologist_user_id"), "psixolog"
            )
            class_payload.append({
                "id": class_id,
                "name": row["name"],
                "smena": row["shift"],
                "bino_id": building_id,
                "xona_id": room_id,
                "bino": building_name,
                "xona": room_number,
                "talim_tili": row["talim_tili"],
            })

        _v209_keep_highest_member_role(
            member_roles, sorov.director_user_id, "direktor"
        )
        for user_id, role in sorted(
            member_roles.items(),
            key=lambda pair: _V209_MEMBER_ROLE_ORDER[pair[1]],
        ):
            _v209_attach_school_member(cur, user_id, school_id, role)

        _v1852_default_shifts(cur, school_id, int(sorov.shift_count))
        conn.commit()
        school = {
            "id": school_id,
            "maktab_id": school_id,
            "name": school_name,
            "nomi": school_name,
            "school_number": school_number,
            "maktab_raqami": school_number,
            "region": region,
            "viloyat": region,
            "district": district,
            "tuman": district,
            "shift_count": int(sorov.shift_count),
            "smena_soni": int(sorov.shift_count),
            "alifbo_turi": alphabet_type,
        }
        return {
            "holat": "yaratildi",
            "school": school,
            "maktab_id": school_id,
            "binolar": len(buildings),
            "xonalar": room_payload,
            "sinflar": class_payload,
            "darsga_yaroqsiz_xona_soni": sum(
                1 for room in room_payload if not room["darsga_yaroqli"]
            ),
        }
    except Exception:
        conn.rollback()
        raise
    finally:
        cur.close()
        conn.close()


# ========================= V19.6 END =========================

# Preserve Python monolith semantics: late definitions must be visible to
# earlier platform routes such as the employee import endpoint.
for _v19_name, _v19_value in list(globals().items()):
    if _v19_name not in _V19_IMPORTED_NAMES and not _v19_name.startswith("__"):
        setattr(_platform, _v19_name, _v19_value)

__all__ = [name for name in globals() if not name.startswith("__")]
