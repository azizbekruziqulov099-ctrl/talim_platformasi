import unittest
import sqlite3
import tempfile
from pathlib import Path
import zipfile

import openpyxl

try:  # fullstack repository: backend/tests va backend/modules
    from backend.modules.test_template_import import (
        authoritative_topic_scope,
        canonical_subject_name,
        discover_test_worksheets,
        embedded_images_by_row,
        embedded_images_by_sheet_from_xlsx,
        is_named_test_sheet,
        normalize_difficulty,
        row_values_by_header,
        subject_matches,
        worksheet_subject_hint,
    )
except ModuleNotFoundError:  # backend-only repository: tests va modules
    from modules.test_template_import import (
        authoritative_topic_scope,
        canonical_subject_name,
        discover_test_worksheets,
        embedded_images_by_row,
        embedded_images_by_sheet_from_xlsx,
        is_named_test_sheet,
        normalize_difficulty,
        row_values_by_header,
        subject_matches,
        worksheet_subject_hint,
    )


class _AnchorFrom:
    row = 6


class _Anchor:
    _from = _AnchorFrom()


class _Image:
    anchor = _Anchor()
    format = "PNG"

    @staticmethod
    def _data():
        return b"image-bytes"


class _SecondImage(_Image):
    @staticmethod
    def _data():
        return b"second-image"


class TestTemplateImportHelpers(unittest.TestCase):
    def test_discovers_all_twenty_two_test_sheets_not_only_the_first(self):
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)
        for number in range(1, 23):
            worksheet = workbook.create_sheet(f"TESTLAR_Fan_{number}")
            worksheet.append(["TOPIC_CODE", " Question ", "correct_answer"])
            worksheet.append([f"1-FAN-{number}", f"Savol {number}", "A"])

        workbook.create_sheet("MALUMOT").append(["Topic code", "Fan"])
        # Hatto asosiy ustunlarga o'xshasa ham, TESTLAR varaqlari mavjud
        # paytda yordamchi varaq importga aralashmasligi kerak.
        workbook.create_sheet("IZOH").append(["topic_code", "question"])
        broken = workbook.create_sheet("TESTLAR_Buzilgan")
        broken.append(["topic_code", "correct_answer"])

        selected, skipped = discover_test_worksheets(workbook)

        self.assertEqual(len(selected), 22)
        self.assertEqual([item.name for item in selected[:2]], ["TESTLAR_Fan_1", "TESTLAR_Fan_2"])
        self.assertEqual(sum(item.worksheet.max_row - 1 for item in selected), 22)
        self.assertEqual(skipped[0]["varaq"], "TESTLAR_Buzilgan")
        self.assertEqual(skipped[0]["yetishmagan_ustunlar"], ["question"])

    def test_accepts_renamed_sheet_by_signature_when_no_testlar_name_exists(self):
        workbook = openpyxl.Workbook()
        workbook.active.title = "Mening savollarim"
        workbook.active.append(["topic_code", "question", "correct_answer"])
        workbook.active.append(["1-MAT-1", "2 + 2?"])

        selected, skipped = discover_test_worksheets(workbook)

        self.assertEqual([item.name for item in selected], ["Mening savollarim"])
        self.assertEqual(skipped, [])
        self.assertTrue(is_named_test_sheet("testlar_Fizika"))
        self.assertTrue(is_named_test_sheet("TESTLAR Fizika"))

    def test_accepts_renamed_valid_sheet_alongside_named_test_sheet(self):
        workbook = openpyxl.Workbook()
        workbook.active.title = "TESTLAR_Matematika"
        workbook.active.append(["topic_code", "question", "correct_answer"])
        workbook.active.append(["1-MAT-1", "2 + 2?", "A"])
        renamed = workbook.create_sheet("Qo'shimcha savollar")
        renamed.append(["topic_code", "question", "correct_answer"])
        renamed.append(["1-MAT-2", "3 + 3?", "B"])

        selected, skipped = discover_test_worksheets(workbook)

        self.assertEqual(
            [item.name for item in selected],
            ["TESTLAR_Matematika", "Qo'shimcha savollar"],
        )
        self.assertEqual(skipped, [])

    def test_embedded_image_is_keyed_by_one_based_excel_row(self):
        worksheet = type("Worksheet", (), {"_images": [_Image()]})()

        result = embedded_images_by_row(worksheet)

        self.assertEqual(result.source_count, 1)
        self.assertEqual(result.by_row, {7: (b"image-bytes", "png")})
        self.assertEqual(result.errors, [])
        self.assertEqual(result.warnings, [])

    def test_first_image_wins_when_two_images_share_one_row(self):
        worksheet = type("Worksheet", (), {"_images": [_Image(), _SecondImage()]})()

        result = embedded_images_by_row(worksheet)

        self.assertEqual(result.by_row, {7: (b"image-bytes", "png")})
        self.assertEqual(len(result.warnings), 1)

    def test_read_only_xlsx_image_extractor_uses_drawing_relationships(self):
        with tempfile.NamedTemporaryFile(suffix=".xlsx") as xlsx_file:
            with zipfile.ZipFile(xlsx_file.name, "w") as archive:
                archive.writestr(
                    "xl/workbook.xml",
                    '''<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"
                       xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
                       <sheets><sheet name="TESTLAR_MATEMATIKA" sheetId="1" r:id="rId1"/></sheets>
                       </workbook>''',
                )
                archive.writestr(
                    "xl/_rels/workbook.xml.rels",
                    '''<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
                       <Relationship Id="rId1" Target="worksheets/sheet1.xml"/>
                       </Relationships>''',
                )
                archive.writestr(
                    "xl/worksheets/sheet1.xml",
                    '''<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"
                       xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
                       <drawing r:id="rId1"/></worksheet>''',
                )
                archive.writestr(
                    "xl/worksheets/_rels/sheet1.xml.rels",
                    '''<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
                       <Relationship Id="rId1" Target="../drawings/drawing1.xml"/>
                       </Relationships>''',
                )
                archive.writestr(
                    "xl/drawings/drawing1.xml",
                    '''<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
                       xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"
                       xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
                       <xdr:oneCellAnchor><xdr:from><xdr:row>6</xdr:row></xdr:from>
                       <xdr:pic><xdr:blipFill><a:blip r:embed="rId5"/></xdr:blipFill></xdr:pic>
                       </xdr:oneCellAnchor></xdr:wsDr>''',
                )
                archive.writestr(
                    "xl/drawings/_rels/drawing1.xml.rels",
                    '''<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
                       <Relationship Id="rId5" Target="../media/image1.png"/>
                       </Relationships>''',
                )
                archive.writestr("xl/media/image1.png", b"zip-image-bytes")

            result = embedded_images_by_sheet_from_xlsx(
                xlsx_file.name, {"TESTLAR_MATEMATIKA"}
            )["TESTLAR_MATEMATIKA"]

        self.assertEqual(result.source_count, 1)
        self.assertEqual(result.by_row, {7: (b"zip-image-bytes", "png")})
        self.assertEqual(result.errors, [])

    def test_blank_cells_do_not_shift_school_fields_into_answer_columns(self):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.append([
            "index", "topic_code", "difficulty", "situation", "question",
            "option_a", "option_b", "option_c", "option_d", "correct_answer",
        ])
        worksheet.append([
            None, "6-MAT-001", "SCHOOL", "🏫 Oddiy", "2 + 2 nechchi?",
            "2", "4", "5", "6", "4",
        ])

        headers = [str(cell.value or "").strip().lower() for cell in worksheet[1]]
        mapped = row_values_by_header(headers, next(worksheet.iter_rows(min_row=2, max_row=2)))

        self.assertIsNone(mapped["index"])
        self.assertEqual(mapped["difficulty"], "SCHOOL")
        self.assertEqual(mapped["situation"], "🏫 Oddiy")
        self.assertEqual(mapped["option_a"], "2")
        self.assertEqual(mapped["option_b"], "4")

    def test_difficulty_apostrophe_variants_are_normalized(self):
        self.assertEqual(normalize_difficulty("o‘rta"), "o'rta")
        self.assertEqual(normalize_difficulty("o’rta"), "o'rta")
        self.assertEqual(normalize_difficulty("o'rta"), "o'rta")
        self.assertEqual(normalize_difficulty("OSON"), "oson")
        self.assertIsNone(normalize_difficulty(None))

    def test_subject_sheet_hint_and_canonical_matching_are_safe(self):
        self.assertEqual(worksheet_subject_hint("TESTLAR_Matematika"), "Matematika")
        self.assertEqual(worksheet_subject_hint("testlar O'zbek tili"), "O'zbek tili")
        self.assertIsNone(worksheet_subject_hint("TESTLAR"))
        self.assertEqual(canonical_subject_name("O'zbek tili"), canonical_subject_name("Ozbek-tili"))
        self.assertTrue(subject_matches("Matematika", "Matematika"))
        self.assertFalse(subject_matches("Matematika", "Fizika"))

    def test_malumot_is_authoritative_for_grade_subject_and_topic_prefix(self):
        metadata = {
            "7-01-01-01-01-01-001": {
                "grade": "7", "subject_name": "ALGEBRA", "mavzu_name": "sonli ifodalar",
            },
            "7-02-01-01-01-01-001": {
                "grade": "7", "subject_name": "GEOMETRIYA", "mavzu_name": "eng sodda shakllar",
            },
            "7-03-01-01-01-01-001": {
                "grade": "7", "subject_name": "INGLIZ TILI", "mavzu_name": "summer holidays",
            },
            "6-02-01-01-01-01-001": {
                "grade": "6", "subject_name": "MATEMATIKA", "mavzu_name": "kasrlar",
            },
        }

        scope, errors = authoritative_topic_scope(metadata, "7")

        self.assertEqual(errors, [])
        self.assertEqual(set(scope), {
            "7-01-01-01-01-01-001",
            "7-02-01-01-01-01-001",
            "7-03-01-01-01-01-001",
        })
        self.assertEqual(scope["7-02-01-01-01-01-001"]["subject_code"], "02")
        self.assertEqual(scope["7-02-01-01-01-01-001"]["subject_name"], "GEOMETRIYA")

    def test_one_subject_code_cannot_point_to_two_subjects(self):
        metadata = {
            "7-02-01-01-01-01-001": {"grade": "7", "subject_name": "GEOMETRIYA"},
            "7-02-01-01-01-02-001": {"grade": "7", "subject_name": "INGLIZ TILI"},
        }

        scope, errors = authoritative_topic_scope(metadata, "7")

        self.assertEqual(len(scope), 1)
        self.assertTrue(any("bitta fan kodi ikki fanga" in error for error in errors))


class TestTemplateImportEndpointContract(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        main_path = Path(__file__).resolve().parents[1] / "main.py"
        main_source = main_path.read_text(encoding="utf-8")
        start = main_source.index("async def shablon_import")
        end = main_source.index("\n@app.get(\"/api/test_rasmi/", start)
        cls.endpoint = main_source[start:end]

    def test_multi_sheet_loop_and_atomic_single_commit(self):
        self.assertIn("for test_varaq in test_varaqlar", self.endpoint)
        self.assertNotIn("SAVEPOINT shablon_import_qatori", self.endpoint)
        self.assertIn("def import_batch", self.endpoint)
        self.assertIn("len(import_qatorlari) >= 500", self.endpoint)
        self.assertIn("psycopg2.extras.execute_values", self.endpoint)
        self.assertEqual(self.endpoint.count("conn.commit()"), 1)
        self.assertIn("conn.rollback()", self.endpoint)

    def test_large_workbook_is_streamed_from_disk_and_read_only(self):
        self.assertIn('tempfile.NamedTemporaryFile(suffix=".xlsx")', self.endpoint)
        self.assertIn("await fayl.read(1024 * 1024)", self.endpoint)
        self.assertIn("read_only=True", self.endpoint)
        self.assertIn("embedded_images_by_sheet_from_xlsx", self.endpoint)
        self.assertNotIn("content = await fayl.read()", self.endpoint)
        self.assertNotIn("zf.testzip()", self.endpoint)

    def test_import_atomically_replaces_only_selected_grade_or_topic_scope(self):
        full_grade_delete = self.endpoint.index(
            "DELETE FROM generated_tests AS gt"
        )
        single_subject_delete = self.endpoint.index(
            "DELETE FROM generated_tests WHERE topic_code=ANY(%s)"
        )
        first_insert = self.endpoint.index("INSERT INTO generated_tests")
        self.assertLess(full_grade_delete, first_insert)
        self.assertLess(single_subject_delete, first_insert)
        self.assertIn(
            "split_part(COALESCE(gt.topic_code, ''), '-', 1)=%s",
            self.endpoint,
        )
        self.assertIn("if barcha_fanlar", self.endpoint)
        self.assertIn("import_scope_topic_codes", self.endpoint)
        self.assertIn(
            '"almashtirishda_ochirilgan_eski_test_soni"', self.endpoint
        )

    def test_duplicate_lookup_cannot_move_or_delete_another_subject_or_grade(self):
        self.assertIn("AND gt.topic_code=ANY(%s)", self.endpoint)
        self.assertNotIn(
            "FROM (VALUES %s) AS v(id, topic_code)", self.endpoint
        )
        self.assertNotIn("begona_nusxalar =", self.endpoint)

    def test_malformed_named_sheet_is_rejected_before_database_write(self):
        validation = self.endpoint.index("if buzuq_test_varaqlar")
        database_open = self.endpoint.index("conn = _db()")
        self.assertLess(validation, database_open)
        self.assertIn("status_code=400", self.endpoint[validation:database_open])

    def test_duplicate_fingerprint_uses_all_answers_and_question_type(self):
        for column in (
            "option_a", "option_b", "option_c", "option_d",
            "correct_answer", "question_type",
        ):
            self.assertIn(f"COALESCE(gt.{column}", self.endpoint)

    def test_question_type_schema_guard_runs_before_duplicate_query(self):
        schema_guard = self.endpoint.index(
            "ALTER TABLE generated_tests ADD COLUMN IF NOT EXISTS \"\n"
            "            \"question_type"
        )
        duplicate_query = self.endpoint.index("COALESCE(gt.question_type")
        self.assertLess(schema_guard, duplicate_query)
        self.assertIn("SET question_type='single_choice'", self.endpoint)

    def test_response_contains_multi_sheet_diagnostics(self):
        for key in (
            '"import_qilingan_varaq_soni"',
            '"import_qilingan_varaqlar"',
            '"korilgan_savollar_soni"',
            '"fayldagi_topic_code_soni"',
            '"varaq_diagnostika"',
        ):
            self.assertIn(key, self.endpoint)

    def test_subject_and_grade_scope_are_checked_before_any_question_insert(self):
        self.assertIn("kutilgan_sinf: str = None", self.endpoint)
        self.assertIn("kutilgan_fan: str = None", self.endpoint)
        self.assertIn("SELECT topic_code, grade, subject_name", self.endpoint)
        self.assertIn("authoritative_topic_scope", self.endpoint)
        self.assertIn("xaritalash_xatolari", self.endpoint)
        self.assertIn("worksheet_subject_hint", self.endpoint)
        self.assertIn("Hech bir savol saqlanmadi", self.endpoint)
        validation = self.endpoint.index("if xaritalash_xatolari")
        first_insert = self.endpoint.index("INSERT INTO generated_tests")
        self.assertLess(validation, first_insert)

    def test_all_malumot_topic_codes_repair_dts_subject_before_questions(self):
        mapping = self.endpoint.index("for topic_code, info in authoritative_scope.items()")
        exact_code = self.endpoint.index(
            "if raw_code in authoritative_scope and raw_code in sinf_mavzulari_by_code"
        )
        fallback = self.endpoint.index("resolved = resolve_topic_code_for_scope", exact_code)
        repair = self.endpoint.index("UPDATE dts_tree AS d", mapping)
        first_insert = self.endpoint.index("INSERT INTO generated_tests")
        self.assertLess(mapping, repair)
        self.assertLess(repair, first_insert)
        self.assertLess(exact_code, fallback)
        self.assertIn("MALUMOTdagi tanlangan sinf/fanga tegishli emas", self.endpoint)
        self.assertIn('str(info["subject_code"]).strip()', self.endpoint)
        self.assertIn('str(info["subject_name"]).strip()', self.endpoint)
        self.assertIn('"malumotdan_tekshirilgan_dts_kodlari"', self.endpoint)


class TestRailwaySqliteCopyBlock(unittest.TestCase):
    def test_copy_block_adds_missing_columns_and_is_idempotent(self):
        guide_path = Path(__file__).resolve().parents[2] / "V18_9_COPY_QILISH.md"
        guide = guide_path.read_text(encoding="utf-8")
        start = guide.index("def ensure_questions_import_schema")
        end = guide.index("\n```", start)
        namespace = {}
        exec(guide[start:end], namespace)

        db = sqlite3.connect(":memory:")
        db.execute('CREATE TABLE "questions" (id INTEGER PRIMARY KEY, question TEXT)')
        db.execute('INSERT INTO "questions" (question) VALUES (?)', ("Eski savol",))

        ensure_schema = namespace["ensure_questions_import_schema"]
        ensure_schema(db)
        ensure_schema(db)

        columns = {
            row[1] for row in db.execute('PRAGMA table_info("questions")').fetchall()
        }
        self.assertIn("question_type", columns)
        self.assertIn("maqsad", columns)
        self.assertIn("rasm_malumot", columns)
        value = db.execute(
            'SELECT "question_type" FROM "questions" WHERE id=1'
        ).fetchone()[0]
        self.assertEqual(value, "single_choice")
        db.close()


if __name__ == "__main__":
    unittest.main()
