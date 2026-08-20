"""SamTM V18.23 — maktabni puxta sozlash va xavfsiz Excel importi.

Bu router eski maktab endpointlarini buzmasdan yangi, ikki bosqichli import
oqimini beradi: fayl avval qoralamaga tekshiriladi, keyin foydalanuvchi
tasdig'idan so'ng bitta tranzaksiyada saqlanadi.
"""

from __future__ import annotations

import hashlib
import io
import json
import os
import re
import secrets
import string
import unicodedata
from datetime import datetime, timezone
from difflib import SequenceMatcher
from typing import Any, Callable, Optional

from fastapi import APIRouter, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field


ROLE_LABELS = {
    "direktor": "Direktor",
    "zam_direktor_uquv": "O'quv ishlari bo'yicha direktor o'rinbosari",
    "zam_direktor_tarbiya": "Ma'naviy-ma'rifiy ishlar bo'yicha direktor o'rinbosari",
    "psixolog": "Psixolog",
    "kotib": "Kotib",
    "fan_oqituvchisi": "Fan o'qituvchisi",
}
ROLE_ALIASES = {
    "direktor": "direktor",
    "oquv ishlari boyicha direktor orinbosari": "zam_direktor_uquv",
    "manaviy marifiy ishlar boyicha direktor orinbosari": "zam_direktor_tarbiya",
    "psixolog": "psixolog",
    "kotib": "kotib",
    "fan oqituvchisi": "fan_oqituvchisi",
    "oqituvchi": "fan_oqituvchisi",
}
DEFAULT_SUBJECTS = [
    "Algebra", "Geometriya", "Matematika", "Ona tili", "Adabiyot",
    "O'qish savodxonligi", "Ingliz tili", "Rus tili", "Tarix",
    "O'zbekiston tarixi", "Jahon tarixi", "Biologiya", "Geografiya",
    "Fizika", "Kimyo", "Informatika", "Texnologiya", "Tabiiy fan",
    "Tarbiya", "Musiqa", "Tasviriy san'at", "Huquq", "Iqtisod",
    "Jismoniy tarbiya",
]


def normalize_text(value: Any) -> str:
    """Ism/fan/lavozimni solishtirish uchun barqaror shakl."""
    text = unicodedata.normalize("NFKD", str(value or "").strip().lower())
    text = text.replace("ʻ", "'").replace("’", "'").replace("`", "'").replace("‘", "'")
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.replace("o'", "o").replace("g'", "g")
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def normalize_class_code(value: Any) -> str:
    """5a, 5-a, 5 A va 5-A ni yagona 5-A ko'rinishiga keltiradi."""
    raw = str(value or "").strip().upper()
    raw = raw.replace("–", "-").replace("—", "-")
    raw = re.sub(r"\bSINF(I)?\b", "", raw, flags=re.IGNORECASE)
    compact = re.sub(r"[\s_-]+", "", raw)
    match = re.fullmatch(r"(1[01]|[1-9])([A-Z])", compact)
    if not match:
        raise ValueError("Sinf 1-A, 5B yoki 11-D ko'rinishida bo'lishi kerak")
    return f"{int(match.group(1))}-{match.group(2)}"


def split_class_code(value: Any) -> tuple[str, str, str]:
    code = normalize_class_code(value)
    grade, letter = code.split("-", 1)
    return code, grade, letter


def _best_subject(value: Any, catalog: list[str]) -> dict[str, Any]:
    raw = str(value or "").strip()
    normalized = normalize_text(raw)
    if not normalized:
        return {"input": raw, "subject": None, "score": 0, "needs_confirmation": False}
    exact = {normalize_text(item): item for item in catalog}
    if normalized in exact:
        return {"input": raw, "subject": exact[normalized], "score": 100, "needs_confirmation": False, "alternatives": [{"subject": exact[normalized], "score": 100}]}
    ranked = sorted(
        ((round(SequenceMatcher(None, normalized, key).ratio() * 100), label) for key, label in exact.items()),
        reverse=True,
    )
    score, label = ranked[0] if ranked else (0, None)
    alternatives = [
        {"subject": candidate_label, "score": candidate_score}
        for candidate_score, candidate_label in ranked[:5]
        if candidate_score >= 72 and candidate_score >= score - 5
    ]
    return {
        "input": raw,
        "subject": label if score >= 72 else None,
        "score": score,
        "needs_confirmation": score >= 72 and (score < 90 or len(alternatives) > 1),
        "alternatives": alternatives,
    }


def parse_teaching_assignments(value: Any, catalog: list[str]) -> list[dict[str, Any]]:
    """`5A: Tarix; 6-B: Tarix` va faqat `Tarix` yozuvlarini ajratadi."""
    text = str(value or "").strip()
    if not text:
        return []
    parts = [part.strip() for part in re.split(r"[;\n]+", text) if part.strip()]
    result = []
    for part in parts:
        class_code = None
        subject_text = part
        match = re.match(r"^\s*(1[01]|[1-9])\s*[-_ ]?\s*([A-Za-z])\s*(?::|-)?\s*(.+?)\s*$", part)
        if match:
            class_code = normalize_class_code(f"{match.group(1)}-{match.group(2)}")
            subject_text = match.group(3)
        subject = _best_subject(subject_text, catalog)
        result.append({"class_code": class_code, **subject})
    return result


def _header_key(value: Any) -> str:
    return normalize_text(value).replace(" ", "_")


def _year(value: Any, field: str, required: bool = False) -> Optional[int]:
    if value in (None, ""):
        if required:
            raise ValueError(f"{field} majburiy")
        return None
    try:
        result = int(float(value))
    except (TypeError, ValueError):
        raise ValueError(f"{field} to'g'ri yil bo'lishi kerak")
    current = datetime.now(timezone.utc).year
    if result < 1900 or result > current:
        raise ValueError(f"{field} 1900–{current} oralig'ida bo'lishi kerak")
    return result


def _integer(value: Any, field: str, minimum: int = 0) -> Optional[int]:
    if value in (None, ""):
        return None
    try:
        result = int(float(value))
    except (TypeError, ValueError):
        raise ValueError(f"{field} butun son bo'lishi kerak")
    if result < minimum:
        raise ValueError(f"{field} {minimum} dan kichik bo'lmasligi kerak")
    return result


def _pin_hash(pin: str) -> str:
    secret = secrets.token_hex(16)
    digest = hashlib.pbkdf2_hmac("sha256", pin.encode("utf-8"), secret.encode("ascii"), 180_000).hex()
    return f"pbkdf2_sha256$180000${secret}${digest}"


def _pin_matches(pin: str, stored: Optional[str]) -> bool:
    if not re.fullmatch(r"\d{4}", str(pin or "")):
        return False
    if not stored:
        fallback = os.getenv("MUASSASA_OCHIRISH_PAROLI", "").strip()
        return bool(re.fullmatch(r"\d{4}", fallback)) and secrets.compare_digest(pin, fallback)
    try:
        algorithm, rounds, salt, expected = stored.split("$", 3)
        if algorithm != "pbkdf2_sha256":
            return False
        actual = hashlib.pbkdf2_hmac(
            "sha256", pin.encode("utf-8"), salt.encode("ascii"), int(rounds)
        ).hex()
        return secrets.compare_digest(actual, expected)
    except (TypeError, ValueError):
        return False


# Ishlaydigan router ham, mustaqil testlar ham aynan bitta qoidalar
# modulidan foydalanadi. Yuqoridagi nomlar eski patchlar bilan import
# mosligini saqlash uchun qoldirilgan, bu import ularni yagona manbaga
# bog'laydi.
from .school_import_rules import (
    best_subject as _best_subject,
    normalize_class_code,
    normalize_text,
    parse_teaching_assignments,
    pin_hash as _pin_hash,
    pin_matches as _pin_matches,
    split_class_code,
)


def _next_negative_user_id(cur) -> int:
    cur.execute("SELECT pg_advisory_xact_lock(74125, 20260816)")
    cur.execute("SELECT MIN(user_id) AS min_id FROM users WHERE user_id < 0")
    row = cur.fetchone()
    return int(row["min_id"] - 1) if row and row["min_id"] is not None else -1


def _catalog(cur) -> list[str]:
    try:
        cur.execute("SELECT to_regclass('public.dts_tree') AS table_name")
        if cur.fetchone()["table_name"]:
            cur.execute("SELECT DISTINCT subject_name FROM dts_tree WHERE subject_name IS NOT NULL ORDER BY subject_name")
            found = [str(row["subject_name"]).strip() for row in cur.fetchall() if str(row["subject_name"] or "").strip()]
            if found:
                return list(dict.fromkeys(found + DEFAULT_SUBJECTS))
    except Exception:
        pass
    return DEFAULT_SUBJECTS.copy()


def _is_admin(cur, user_id: int) -> bool:
    cur.execute("SELECT 1 FROM admin_akkaunt WHERE uid=%s", (user_id,))
    return cur.fetchone() is not None


def _resolve_school_id(
    cur, scope_id: int, user_id: int, create_link: bool = True,
    prefer_context: bool = False,
) -> int:
    """Legacy maktab ID va V17 learning_context ID ikkalasini qabul qiladi."""
    if prefer_context:
        cur.execute("SELECT maktab_id FROM school_context_links WHERE context_id=%s", (scope_id,))
        row = cur.fetchone()
        if row:
            return int(row["maktab_id"])
    else:
        cur.execute("SELECT id FROM maktablar WHERE id=%s AND deleted_at IS NULL", (scope_id,))
        row = cur.fetchone()
        if row:
            return int(row["id"])
        cur.execute("SELECT maktab_id FROM school_context_links WHERE context_id=%s", (scope_id,))
        row = cur.fetchone()
        if row:
            return int(row["maktab_id"])
    if not create_link:
        raise HTTPException(status_code=404, detail="Maktab topilmadi")
    cur.execute(
        """SELECT id,name,owner_user_id,region,district
             FROM learning_contexts
            WHERE id=%s AND context_type='school' AND active=TRUE""",
        (scope_id,),
    )
    context = cur.fetchone()
    if not context:
        raise HTTPException(status_code=404, detail="Maktab topilmadi")
    cur.execute(
        """INSERT INTO maktablar(
               nomi,viloyat,tuman,smena_soni,creator_user_id,creation_source,
               payment_exempt,lifecycle_status
           ) VALUES(%s,%s,%s,1,%s,'v17_context',FALSE,'active') RETURNING id""",
        (context["name"], context["region"], context["district"], context["owner_user_id"] or user_id),
    )
    school_id = int(cur.fetchone()["id"])
    cur.execute(
        "INSERT INTO school_context_links(context_id,maktab_id) VALUES(%s,%s) ON CONFLICT DO NOTHING",
        (scope_id, school_id),
    )
    return school_id


def _can_manage(
    cur, user_id: int, scope_id: int, *, mutation: bool = False,
    prefer_context: bool = False,
) -> tuple[int, bool]:
    admin = _is_admin(cur, user_id)
    school_id = _resolve_school_id(cur, scope_id, user_id, prefer_context=prefer_context)
    if mutation:
        cur.execute("SELECT to_regclass('public.organization_trials') AS table_name")
        if cur.fetchone()["table_name"]:
            cur.execute(
                """SELECT (
                          o.lifecycle_status='read_only'
                          OR (o.lifecycle_status='trial' AND o.trial_ends_at<=NOW())
                       ) AS effective_read_only
                     FROM organization_trials o
                     LEFT JOIN school_context_links link ON link.context_id=o.context_id
                    WHERE o.context_id=%s OR link.maktab_id=%s
                    ORDER BY CASE WHEN o.context_id=%s THEN 0 ELSE 1 END LIMIT 1""",
                (scope_id, school_id, scope_id),
            )
            organization = cur.fetchone()
            if organization and organization["effective_read_only"]:
                raise HTTPException(
                    status_code=423,
                    detail={
                        "code": "ORGANIZATION_READ_ONLY",
                        "message": "30 kunlik sinov tugagan. Ma'lumotlar saqlangan; faollashtirgach tahrirlash ochiladi.",
                        "activation_price_uzs": 200_000,
                    },
                )
    if admin:
        return school_id, True
    cur.execute("SELECT direktor_user_id,creator_user_id FROM maktablar WHERE id=%s", (school_id,))
    school = cur.fetchone()
    if school and user_id in (school["direktor_user_id"], school["creator_user_id"]):
        return school_id, False
    cur.execute(
        "SELECT 1 FROM users WHERE user_id=%s AND maktab_id=%s AND lavozim IN ('direktor','zam_direktor_uquv','zam_direktor_tarbiya')",
        (user_id, school_id),
    )
    if cur.fetchone():
        return school_id, False
    cur.execute(
        """SELECT 1 FROM foydalanuvchi_muassasalari
            WHERE user_id=%s AND muassasa_turi='maktab' AND muassasa_id=%s
              AND lavozim IN ('owner','direktor','zam_direktor_uquv','zam_direktor_tarbiya')""",
        (user_id, school_id),
    )
    if cur.fetchone():
        return school_id, False
    cur.execute(
        """SELECT 1
             FROM school_context_links link
             JOIN learning_contexts c ON c.id=link.context_id
             LEFT JOIN context_memberships cm
               ON cm.context_id=c.id AND cm.user_id=%s AND cm.status='active'
            WHERE link.maktab_id=%s
              AND (c.owner_user_id=%s OR cm.member_role IN ('manager','director','admin'))""",
        (user_id, school_id, user_id),
    )
    if cur.fetchone():
        return school_id, False
    raise HTTPException(status_code=403, detail="Bu maktabni boshqarishga ruxsat yo'q")


class ClassInput(BaseModel):
    code: str
    shift_no: int = Field(default=1, ge=1, le=2)
    homeroom_teacher_user_id: Optional[int] = None
    psychologist_user_id: Optional[int] = None
    building_name: Optional[str] = Field(default=None, max_length=120)
    room_number: Optional[str] = Field(default=None, max_length=40)


class AdminSchoolCreate(BaseModel):
    token: str
    name: str = Field(min_length=2, max_length=160)
    region: Optional[str] = Field(default=None, max_length=120)
    district: Optional[str] = Field(default=None, max_length=120)
    shift_count: int = Field(default=1, ge=1, le=2)
    director_user_id: Optional[int] = None
    classes: list[ClassInput] = Field(default_factory=list)


class ClassBatch(BaseModel):
    token: str
    classes: list[ClassInput]
    scope_kind: str = "school"


class DeleteSchool(BaseModel):
    token: str
    confirmation_name: str
    pin: Optional[str] = None
    reason: Optional[str] = Field(default=None, max_length=240)


class ImportCommit(BaseModel):
    token: str
    decisions: dict[str, Any] = Field(default_factory=dict)


def create_school_institution_v23_router(
    verify_token: Callable[[str], int], db_factory: Callable[[], Any]
) -> APIRouter:
    router = APIRouter(prefix="/api/maktab-v23", tags=["school-v23"])

    def connection():
        conn = db_factory()
        return conn, conn.cursor()

    def close(conn, cur):
        cur.close()
        conn.close()

    def class_rows(cur, school_id: int) -> list[dict[str, Any]]:
        cur.execute(
            """SELECT s.id,s.normalized_code,s.sinf,s.harf,s.shift_no,
                      s.rahbar_user_id,leader.full_name AS rahbar_ismi,
                      s.psychologist_user_id,psych.full_name AS psixolog_ismi,
                      s.building_name,s.room_number,s.qoshilish_paroli
                 FROM maktab_sinflari s
                 LEFT JOIN users leader ON leader.user_id=s.rahbar_user_id
                 LEFT JOIN users psych ON psych.user_id=s.psychologist_user_id
                WHERE s.maktab_id=%s
                ORDER BY NULLIF(regexp_replace(s.sinf,'[^0-9]','','g'),'')::INTEGER,s.harf""",
            (school_id,),
        )
        result = []
        seen = set()
        for raw in cur.fetchall():
            row = dict(raw)
            try:
                canonical = normalize_class_code(row.get("normalized_code") or f'{row["sinf"]}-{row["harf"]}')
            except ValueError:
                canonical = row.get("normalized_code") or f'{row["sinf"]}-{row["harf"]}'
            if canonical in seen:
                continue
            seen.add(canonical)
            row["normalized_code"] = canonical
            result.append(row)
        return result

    def upsert_classes(cur, school_id: int, shift_count: int, items: list[ClassInput]) -> list[dict[str, Any]]:
        seen: set[str] = set()
        for item in items:
            code, grade, letter = split_class_code(item.code)
            if code in seen:
                raise HTTPException(status_code=400, detail=f"{code} ikki marta kiritilgan")
            seen.add(code)
            if item.shift_no > shift_count:
                raise HTTPException(status_code=400, detail=f"{code}: maktab {shift_count} smenali")
            for person_id, label in (
                (item.homeroom_teacher_user_id, "Sinf rahbari"),
                (item.psychologist_user_id, "Psixolog"),
            ):
                if person_id is not None:
                    cur.execute("SELECT 1 FROM users WHERE user_id=%s", (person_id,))
                    if not cur.fetchone():
                        raise HTTPException(status_code=400, detail=f"{code}: {label.lower()} topilmadi")
            password = "".join(secrets.choice(string.digits) for _ in range(4))
            cur.execute(
                """SELECT id FROM maktab_sinflari
                    WHERE maktab_id=%s
                      AND COALESCE(normalized_code,CONCAT(sinf,'-',UPPER(harf)))=%s
                    ORDER BY id LIMIT 1""",
                (school_id, code),
            )
            existing_class = cur.fetchone()
            if existing_class:
                cur.execute(
                    """UPDATE maktab_sinflari SET normalized_code=%s,
                              shift_no=%s,rahbar_user_id=%s,psychologist_user_id=%s,
                              building_name=%s,room_number=%s,updated_at=NOW()
                        WHERE id=%s""",
                    (
                        code, item.shift_no,
                        item.homeroom_teacher_user_id, item.psychologist_user_id,
                        (item.building_name or "").strip() or None,
                        (item.room_number or "").strip() or None,
                        existing_class["id"],
                    ),
                )
                continue
            cur.execute(
                """INSERT INTO maktab_sinflari(
                       maktab_id,sinf,harf,normalized_code,shift_no,rahbar_user_id,
                       psychologist_user_id,building_name,room_number,qoshilish_paroli
                   ) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                   ON CONFLICT (maktab_id,sinf,harf) DO UPDATE SET
                       normalized_code=EXCLUDED.normalized_code,
                       shift_no=EXCLUDED.shift_no,
                       rahbar_user_id=EXCLUDED.rahbar_user_id,
                       psychologist_user_id=EXCLUDED.psychologist_user_id,
                       building_name=EXCLUDED.building_name,
                       room_number=EXCLUDED.room_number,
                       updated_at=NOW()""",
                (
                    school_id, grade, letter, code, item.shift_no,
                    item.homeroom_teacher_user_id, item.psychologist_user_id,
                    (item.building_name or "").strip() or None,
                    (item.room_number or "").strip() or None, password,
                ),
            )
        return class_rows(cur, school_id)

    @router.post("/admin/create")
    def admin_create(payload: AdminSchoolCreate):
        user_id = verify_token(payload.token)
        conn, cur = connection()
        try:
            if not _is_admin(cur, user_id):
                raise HTTPException(status_code=403, detail="Faqat admin uchun")
            deletion_pin = "".join(secrets.choice(string.digits) for _ in range(4))
            cur.execute(
                """INSERT INTO maktablar(
                       nomi,viloyat,tuman,smena_soni,direktor_user_id,creator_user_id,
                       creation_source,payment_exempt,lifecycle_status,pulli,oylik_tolov,
                       deletion_pin_hash
                   ) VALUES(%s,%s,%s,%s,%s,%s,'admin_grant',TRUE,'active',FALSE,NULL,%s)
                   RETURNING id""",
                (
                    payload.name.strip(), payload.region, payload.district,
                    payload.shift_count, payload.director_user_id, user_id,
                    _pin_hash(deletion_pin),
                ),
            )
            school_id = int(cur.fetchone()["id"])
            classes = upsert_classes(cur, school_id, payload.shift_count, payload.classes)
            cur.execute(
                """INSERT INTO learning_contexts(
                       context_type,name,owner_user_id,region,district,external_type,
                       external_id,active,metadata
                   ) VALUES('school',%s,%s,%s,%s,'maktab',%s,TRUE,%s::jsonb)
                   ON CONFLICT (external_type,external_id) WHERE external_type IS NOT NULL
                   DO UPDATE SET name=EXCLUDED.name,owner_user_id=EXCLUDED.owner_user_id,
                                 active=TRUE,updated_at=NOW()
                   RETURNING id""",
                (
                    payload.name.strip(), user_id, payload.region, payload.district,
                    school_id,
                    json.dumps({"creation_source": "admin_grant", "charged_uzs": 0}),
                ),
            )
            context_id = int(cur.fetchone()["id"])
            cur.execute(
                "INSERT INTO school_context_links(context_id,maktab_id) VALUES(%s,%s) ON CONFLICT DO NOTHING",
                (context_id, school_id),
            )
            cur.execute(
                """INSERT INTO context_memberships(
                       context_id,user_id,member_role,status,source,approved_by_user_id,metadata
                   ) VALUES(%s,%s,'admin','active','admin_grant',%s,'{}'::jsonb)
                   ON CONFLICT DO NOTHING""",
                (context_id, user_id, user_id),
            )
            conn.commit()
            return {
                "status": "created", "school_id": school_id, "context_id": context_id,
                "lifecycle_status": "active", "activation_source": "admin_grant",
                "charged_uzs": 0, "payment_exempt": True, "classes": classes,
                # Faqat yaratish javobida qaytadi; bazada faqat xeshi saqlanadi.
                "deletion_pin": deletion_pin,
            }
        except Exception:
            conn.rollback()
            raise
        finally:
            close(conn, cur)

    @router.get("/admin/schools")
    def admin_schools(token: str):
        user_id = verify_token(token)
        conn, cur = connection()
        try:
            if not _is_admin(cur, user_id):
                raise HTTPException(status_code=403, detail="Faqat admin uchun")
            cur.execute(
                """SELECT m.id,m.nomi,m.viloyat,m.tuman,m.smena_soni,
                          m.direktor_user_id,u.full_name AS direktor_ismi,
                          m.creator_user_id,m.creation_source,m.payment_exempt,
                          m.lifecycle_status,m.yaratilgan_at,
                          (m.creator_user_id=%s) AS own_creation,
                          (SELECT COUNT(DISTINCT COALESCE(s.normalized_code,CONCAT(s.sinf,'-',UPPER(s.harf))))
                             FROM maktab_sinflari s WHERE s.maktab_id=m.id) AS class_count
                     FROM maktablar m
                     LEFT JOIN users u ON u.user_id=m.direktor_user_id
                    WHERE m.deleted_at IS NULL ORDER BY m.yaratilgan_at DESC,m.nomi""",
                (user_id,),
            )
            return {"schools": [dict(row) for row in cur.fetchall()]}
        finally:
            close(conn, cur)

    @router.delete("/{scope_id}")
    def delete_school(scope_id: int, payload: DeleteSchool):
        user_id = verify_token(payload.token)
        conn, cur = connection()
        try:
            if not _is_admin(cur, user_id):
                raise HTTPException(status_code=403, detail="Faqat admin uchun")
            school_id = _resolve_school_id(cur, scope_id, user_id, create_link=False)
            cur.execute(
                "SELECT nomi,creator_user_id,deletion_pin_hash,deleted_at FROM maktablar WHERE id=%s FOR UPDATE",
                (school_id,),
            )
            school = cur.fetchone()
            if not school or school["deleted_at"] is not None:
                raise HTTPException(status_code=404, detail="Maktab topilmadi")
            if normalize_text(payload.confirmation_name) != normalize_text(school["nomi"]):
                raise HTTPException(status_code=400, detail="Tasdiqlash uchun maktab nomini aynan kiriting")
            own = school["creator_user_id"] == user_id
            if not own:
                subject = f"school-delete:{school_id}:{user_id}"
                cur.execute("SELECT attempts,locked_until FROM school_delete_attempts WHERE subject=%s", (subject,))
                attempt = cur.fetchone()
                now = datetime.now(timezone.utc)
                if attempt and attempt["locked_until"] and attempt["locked_until"] > now:
                    raise HTTPException(status_code=429, detail="Ko'p xato urinish. 30 daqiqadan keyin qayta urinib ko'ring")
                if not _pin_matches(payload.pin or "", school["deletion_pin_hash"]):
                    cur.execute(
                        """INSERT INTO school_delete_attempts(subject,attempts,window_started,locked_until,updated_at)
                           VALUES(%s,1,NOW(),NULL,NOW())
                           ON CONFLICT(subject) DO UPDATE SET
                             attempts=CASE WHEN school_delete_attempts.window_started<NOW()-INTERVAL '15 minutes' THEN 1 ELSE school_delete_attempts.attempts+1 END,
                             window_started=CASE WHEN school_delete_attempts.window_started<NOW()-INTERVAL '15 minutes' THEN NOW() ELSE school_delete_attempts.window_started END,
                             locked_until=CASE WHEN school_delete_attempts.attempts+1>=10 THEN NOW()+INTERVAL '30 minutes' ELSE school_delete_attempts.locked_until END,
                             updated_at=NOW()""",
                        (subject,),
                    )
                    conn.commit()
                    raise HTTPException(status_code=400, detail="4 xonali o'chirish paroli noto'g'ri")
                cur.execute("DELETE FROM school_delete_attempts WHERE subject=%s", (subject,))
            cur.execute(
                """UPDATE maktablar SET deleted_at=NOW(),deleted_by_user_id=%s,
                          delete_reason=%s,lifecycle_status='archived',updated_at=NOW()
                    WHERE id=%s""",
                (user_id, (payload.reason or "").strip() or None, school_id),
            )
            cur.execute("UPDATE learning_contexts SET active=FALSE,updated_at=NOW() WHERE external_type='maktab' AND external_id=%s", (school_id,))
            conn.commit()
            return {"status": "archived", "school_id": school_id, "password_required": not own}
        except HTTPException:
            if not conn.closed:
                conn.rollback()
            raise
        except Exception:
            conn.rollback()
            raise
        finally:
            close(conn, cur)

    @router.get("/{scope_id}/classes")
    def classes(scope_id: int, token: str, scope_kind: str = "school"):
        user_id = verify_token(token)
        conn, cur = connection()
        try:
            school_id, _ = _can_manage(cur, user_id, scope_id, prefer_context=scope_kind == "context")
            cur.execute("SELECT smena_soni FROM maktablar WHERE id=%s", (school_id,))
            school = cur.fetchone()
            conn.commit()
            return {"school_id": school_id, "shift_count": school["smena_soni"], "classes": class_rows(cur, school_id)}
        finally:
            close(conn, cur)

    @router.get("/{scope_id}/people")
    def school_people(scope_id: int, token: str, query: str = "", role: str = "", scope_kind: str = "school"):
        user_id = verify_token(token)
        conn, cur = connection()
        try:
            school_id, _ = _can_manage(cur, user_id, scope_id, prefer_context=scope_kind == "context")
            filters = ["u.maktab_id=%s", "u.full_name ILIKE %s"]
            params: list[Any] = [school_id, f"%{query.strip()}%"]
            if role == "psychologist":
                filters.append("u.lavozim='psixolog'")
            elif role == "teacher":
                filters.append("u.role='oqituvchi'")
            cur.execute(
                f"""SELECT u.user_id,u.full_name,u.role,u.lavozim
                       FROM users u WHERE {' AND '.join(filters)}
                      ORDER BY u.full_name LIMIT 20""",
                tuple(params),
            )
            result = {"people": [dict(row) for row in cur.fetchall()]}
            conn.commit()
            return result
        finally:
            close(conn, cur)

    @router.post("/{scope_id}/classes")
    def save_classes(scope_id: int, payload: ClassBatch):
        user_id = verify_token(payload.token)
        conn, cur = connection()
        try:
            school_id, _ = _can_manage(
                cur, user_id, scope_id, mutation=True,
                prefer_context=payload.scope_kind == "context",
            )
            cur.execute("SELECT smena_soni FROM maktablar WHERE id=%s FOR UPDATE", (school_id,))
            school = cur.fetchone()
            result = upsert_classes(cur, school_id, int(school["smena_soni"]), payload.classes)
            conn.commit()
            return {"status": "saved", "classes": result}
        except Exception:
            conn.rollback()
            raise
        finally:
            close(conn, cur)

    def workbook_response(workbook, filename: str):
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    def style_template(workbook, sheet, required_columns: set[int]):
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.worksheet.datavalidation import DataValidation

        for cell in sheet[1]:
            required = cell.column in required_columns
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="B0553A" if required else "6F7F8F")
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        instruction = workbook.create_sheet("YORIQNOMA")
        instruction["A1"] = "* Qizil ustun — majburiy"
        instruction["A2"] = "Kulrang ustun — ixtiyoriy; bo'sh qoldirish mumkin"
        instruction["A3"] = "Import avval tekshiriladi. Tasdiqlamaguningizcha bazaga hech narsa yozilmaydi."
        instruction["A1"].font = Font(bold=True, color="B0553A")
        instruction.column_dimensions["A"].width = 100
        return DataValidation

    @router.get("/{scope_id}/students/template")
    def student_template(scope_id: int, token: str, class_ids: str = "", scope_kind: str = "school"):
        user_id = verify_token(token)
        conn, cur = connection()
        try:
            school_id, _ = _can_manage(cur, user_id, scope_id, prefer_context=scope_kind == "context")
            all_classes = class_rows(cur, school_id)
            selected_ids = {int(value) for value in class_ids.split(",") if value.strip().isdigit()}
            selected = [row for row in all_classes if not selected_ids or int(row["id"]) in selected_ids]
            if not selected:
                raise HTTPException(status_code=400, detail="Kamida bitta sinfni tanlang")
            conn.commit()
        finally:
            close(conn, cur)
        import openpyxl
        from openpyxl.styles import Font
        from openpyxl.worksheet.datavalidation import DataValidation

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "OQUVCHILAR"
        headers = [
            "Sinf *", "O'quvchi F.I.Sh. *", "O'quvchi tug'ilgan yili *",
            "Ota/ona turi *", "Ota/ona F.I.Sh. *", "Ota/ona tug'ilgan yili *",
            "Mahalla (ixtiyoriy)", "Yashash manzili (ixtiyoriy)",
        ]
        ws.append(headers)
        style_template(wb, ws, set(range(1, 7)))
        classes_sheet = wb.create_sheet("SINFLAR")
        classes_sheet.append(["Maktabdagi tanlangan sinflar"])
        classes_sheet["A1"].font = Font(bold=True)
        for row in selected:
            classes_sheet.append([row["normalized_code"] or f'{row["sinf"]}-{row["harf"]}'])
        classes_sheet.sheet_state = "hidden"
        dv_class = DataValidation(type="list", formula1=f"=SINFLAR!$A$2:$A${len(selected)+1}", allow_blank=False)
        dv_parent = DataValidation(type="list", formula1='"Ota,Ona,Vasiy"', allow_blank=False)
        ws.add_data_validation(dv_class); dv_class.add("A2:A2000")
        ws.add_data_validation(dv_parent); dv_parent.add("D2:D2000")
        sample_class = selected[0]["normalized_code"] or f'{selected[0]["sinf"]}-{selected[0]["harf"]}'
        ws.append([sample_class, "Aliyev Akmal Vali o'g'li", 2015, "Ota", "Aliyev Vali Karimovich", 1988, "Navbahor", "Samarqand shahri"])
        for column, width in zip("ABCDEFGH", [14, 34, 24, 18, 34, 25, 25, 42]):
            ws.column_dimensions[column].width = width
        return workbook_response(wb, "oquvchilar_shabloni.xlsx")

    @router.get("/{scope_id}/staff/template")
    def staff_template(scope_id: int, token: str, scope_kind: str = "school"):
        user_id = verify_token(token)
        conn, cur = connection()
        try:
            school_id, _ = _can_manage(cur, user_id, scope_id, prefer_context=scope_kind == "context")
            school_classes = class_rows(cur, school_id)
            subjects = _catalog(cur)
            conn.commit()
        finally:
            close(conn, cur)
        import openpyxl
        from openpyxl.styles import Font
        from openpyxl.worksheet.datavalidation import DataValidation

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "XODIMLAR"
        ws.append([
            "F.I.Sh. *", "Lavozim *", "Mutaxassisligi * (o'qituvchi uchun)",
            "Toifasi (ixtiyoriy)", "Tug'ilgan yili (ixtiyoriy)",
            "Ish staji (ixtiyoriy)", "Yashash manzili (ixtiyoriy)",
            "Sinf rahbarligi (ixtiyoriy)", "Dars birikmalari (ixtiyoriy)",
        ])
        style_template(wb, ws, {1, 2, 3})
        helper = wb.create_sheet("SINFLAR_VA_FANLAR")
        helper.append(["Sinflar", "Fanlar", "Lavozimlar"])
        helper["A1"].font = helper["B1"].font = helper["C1"].font = Font(bold=True)
        for index, row in enumerate(school_classes, 2):
            helper.cell(index, 1, row["normalized_code"] or f'{row["sinf"]}-{row["harf"]}')
        for index, subject in enumerate(subjects, 2):
            helper.cell(index, 2, subject)
        for index, label in enumerate(ROLE_LABELS.values(), 2):
            helper.cell(index, 3, label)
        helper.sheet_state = "hidden"
        dv_role = DataValidation(type="list", formula1=f"=SINFLAR_VA_FANLAR!$C$2:$C${len(ROLE_LABELS)+1}")
        ws.add_data_validation(dv_role); dv_role.add("B2:B2000")
        if school_classes:
            dv_class = DataValidation(type="list", formula1=f"=SINFLAR_VA_FANLAR!$A$2:$A${len(school_classes)+1}", allow_blank=True)
            ws.add_data_validation(dv_class); dv_class.add("H2:H2000")
        ws.append([
            "Yusupov Sardor Baxtiyorovich", "Fan o'qituvchisi", "Tarix",
            "1-toifali", 1988, 8, "Samarqand shahri", "5A",
            "5A: Tarix; 5-B: Tarix; 6a: Jahon tarixi",
        ])
        guide = wb["YORIQNOMA"]
        guide["A5"] = "Dars birikmalari: 5A: Tarix; 5-B: Tarix. Birikmalarni nuqtali vergul yoki yangi qator bilan ajrating."
        guide["A6"] = "Faqat Tarix yozilsa, xodimning fani saqlanadi, ammo sinfga biriktirilmaydi."
        guide["A7"] = "5a, 5-a, 5 A va 5-A yozuvlari bitta 5-A sinf sifatida olinadi."
        for column, width in zip("ABCDEFGHI", [34, 44, 28, 24, 23, 22, 40, 28, 50]):
            ws.column_dimensions[column].width = width
        return workbook_response(wb, "xodimlar_shabloni.xlsx")

    def load_sheet(content: bytes, expected: str):
        import openpyxl
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Excel faylini o'qib bo'lmadi: {exc}")
        return workbook[expected] if expected in workbook.sheetnames else workbook.active

    def rows_as_dicts(sheet) -> list[tuple[int, dict[str, Any]]]:
        headers = [_header_key(cell.value) for cell in sheet[1]]
        result = []
        for number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
            if not any(value not in (None, "") for value in values):
                continue
            result.append((number, {headers[index]: value for index, value in enumerate(values) if index < len(headers)}))
        return result

    def field(row: dict[str, Any], *starts: str):
        normalized_starts = [normalize_text(start).replace(" ", "_") for start in starts]
        for key, value in row.items():
            cleaned = key.replace("_ixtiyoriy", "").replace("_majburiy", "").strip("_")
            if any(cleaned.startswith(start) for start in normalized_starts):
                return value
        return None

    def create_job(cur, school_id: int, user_id: int, kind: str, payload: list[dict[str, Any]]) -> int:
        cur.execute(
            """INSERT INTO school_import_jobs(school_id,import_type,status,payload,created_by_user_id)
               VALUES(%s,%s,'previewed',%s::jsonb,%s) RETURNING id""",
            (school_id, kind, json.dumps(payload, ensure_ascii=False), user_id),
        )
        return int(cur.fetchone()["id"])

    @router.post("/{scope_id}/students/preview")
    async def preview_students(scope_id: int, token: str, scope_kind: str = "school", file: UploadFile = File(...)):
        user_id = verify_token(token)
        content = await file.read()
        if len(content) > 10 * 1024 * 1024:
            raise HTTPException(status_code=413, detail="Excel fayli 10 MB dan katta bo'lmasligi kerak")
        sheet = load_sheet(content, "OQUVCHILAR")
        conn, cur = connection()
        try:
            school_id, _ = _can_manage(
                cur, user_id, scope_id, mutation=True,
                prefer_context=scope_kind == "context",
            )
            classes_by_code = {
                (row["normalized_code"] or f'{row["sinf"]}-{row["harf"]}'): row
                for row in class_rows(cur, school_id)
            }
            preview = []
            batch_parents = []
            for row_number, row in rows_as_dicts(sheet):
                errors, warnings = [], []
                try:
                    class_code = normalize_class_code(field(row, "sinf"))
                except ValueError as exc:
                    class_code = str(field(row, "sinf") or "")
                    errors.append(str(exc))
                student_name = str(field(row, "oquvchi_f_i_sh", "oquvchi_fish") or "").strip()
                parent_name = str(field(row, "ota_ona_f_i_sh", "ota_ona_fish") or "").strip()
                parent_type = str(field(row, "ota_ona_turi") or "").strip().capitalize()
                try:
                    student_year = _year(field(row, "oquvchi_tugilgan_yili"), "O'quvchi tug'ilgan yili", True)
                except ValueError as exc:
                    student_year = None; errors.append(str(exc))
                try:
                    parent_year = _year(field(row, "ota_ona_tugilgan_yili"), "Ota/ona tug'ilgan yili", True)
                except ValueError as exc:
                    parent_year = None; errors.append(str(exc))
                if class_code not in classes_by_code:
                    errors.append(f"{class_code or 'Sinf'} maktabda topilmadi")
                if len(student_name) < 3:
                    errors.append("O'quvchi F.I.Sh. majburiy")
                if len(parent_name) < 3:
                    errors.append("Ota/ona F.I.Sh. majburiy")
                if normalize_text(parent_type) not in {"ota", "ona", "vasiy"}:
                    errors.append("Ota/ona turi: Ota, Ona yoki Vasiy bo'lishi kerak")
                candidates = []
                exact_parent_id = None
                if parent_name and parent_year:
                    cur.execute(
                        """SELECT p.user_id,u.full_name,p.birth_year,p.normalized_name,
                                  ARRAY_REMOVE(ARRAY_AGG(a.normalized_alias),NULL) AS aliases
                             FROM school_person_profiles p
                             JOIN users u ON u.user_id=p.user_id
                             LEFT JOIN school_person_aliases a ON a.user_id=p.user_id
                            WHERE p.person_type='parent' AND p.birth_year=%s
                            GROUP BY p.user_id,u.full_name,p.birth_year,p.normalized_name""",
                        (parent_year,),
                    )
                    wanted = normalize_text(parent_name)
                    exact_candidates = []
                    for candidate in cur.fetchall():
                        names = [candidate["normalized_name"], *(candidate.get("aliases") or [])]
                        score = max(round(SequenceMatcher(None, wanted, name).ratio() * 100) for name in names if name)
                        if score == 100:
                            exact_candidates.append(candidate)
                        elif score >= 72:
                            candidates.append({
                                "user_id": int(candidate["user_id"]), "full_name": candidate["full_name"],
                                "birth_year": candidate["birth_year"], "score": score,
                            })
                    if len(exact_candidates) == 1:
                        exact_parent_id = int(exact_candidates[0]["user_id"])
                    elif len(exact_candidates) > 1:
                        candidates.extend({
                            "user_id": int(candidate["user_id"]), "full_name": candidate["full_name"],
                            "birth_year": candidate["birth_year"], "score": 100,
                        } for candidate in exact_candidates)
                    if exact_parent_id is None:
                        for previous in batch_parents:
                            if previous["birth_year"] != parent_year:
                                continue
                            score = round(SequenceMatcher(None, wanted, previous["normalized_name"]).ratio() * 100)
                            if score >= 72:
                                candidates.append({
                                    "user_id": f'batch:{previous["row_number"]}',
                                    "full_name": previous["full_name"],
                                    "birth_year": previous["birth_year"],
                                    "score": score,
                                    "source": "uploaded_file",
                                })
                    candidates.sort(key=lambda item: item["score"], reverse=True)
                    if candidates and exact_parent_id is None:
                        warnings.append("O'xshash ota/ona topildi — bir xil yoki har xil ekanini tanlang")
                    batch_parents.append({
                        "row_number": row_number, "full_name": parent_name,
                        "normalized_name": wanted, "birth_year": parent_year,
                    })
                item = {
                    "row_number": row_number, "class_code": class_code,
                    "class_id": classes_by_code.get(class_code, {}).get("id"),
                    "student_name": student_name, "student_birth_year": student_year,
                    "parent_type": parent_type, "parent_name": parent_name,
                    "parent_birth_year": parent_year,
                    "mahalla": str(field(row, "mahalla") or "").strip() or None,
                    "address": str(field(row, "yashash_manzili") or "").strip() or None,
                    "exact_parent_id": exact_parent_id, "parent_candidates": candidates[:5],
                    "errors": errors, "warnings": warnings,
                    "status": "error" if errors else ("decision_required" if candidates and exact_parent_id is None else "ready"),
                }
                preview.append(item)
            job_id = create_job(cur, school_id, user_id, "students", preview)
            conn.commit()
            return {
                "job_id": job_id, "rows": preview,
                "summary": {
                    "total": len(preview),
                    "ready": sum(row["status"] == "ready" for row in preview),
                    "decision_required": sum(row["status"] == "decision_required" for row in preview),
                    "errors": sum(row["status"] == "error" for row in preview),
                },
            }
        except Exception:
            conn.rollback()
            raise
        finally:
            close(conn, cur)

    @router.post("/imports/{job_id}/students/commit")
    def commit_students(job_id: int, payload: ImportCommit):
        user_id = verify_token(payload.token)
        conn, cur = connection()
        try:
            cur.execute("SELECT * FROM school_import_jobs WHERE id=%s FOR UPDATE", (job_id,))
            job = cur.fetchone()
            if not job or job["import_type"] != "students":
                raise HTTPException(status_code=404, detail="O'quvchi import qoralamasi topilmadi")
            school_id, _ = _can_manage(cur, user_id, int(job["school_id"]), mutation=True)
            if job["status"] == "committed":
                return {"status": "committed", "reused": True, **(job["result"] or {})}
            rows = job["payload"] if isinstance(job["payload"], list) else json.loads(job["payload"])
            if any(row["status"] == "error" for row in rows):
                raise HTTPException(status_code=409, detail="Avval Excel xatolarini tuzating")
            created_students = created_parents = linked = 0
            access_codes = []
            next_user_id = _next_negative_user_id(cur)
            parent_by_row: dict[int, int] = {}
            for row in rows:
                decision = payload.decisions.get(str(row["row_number"]), {})
                if row["status"] == "decision_required" and "same_parent_id" not in decision:
                    raise HTTPException(status_code=409, detail=f'{row["row_number"]}-qator: ota/ona o\'xshashligini tasdiqlang')
                cur.execute(
                    """SELECT p.user_id FROM school_person_profiles p
                       WHERE p.school_id=%s AND p.person_type='student'
                         AND p.normalized_name=%s AND p.birth_year=%s LIMIT 1""",
                    (school_id, normalize_text(row["student_name"]), row["student_birth_year"]),
                )
                existing_student = cur.fetchone()
                if existing_student:
                    student_id = int(existing_student["user_id"])
                    cur.execute(
                        "UPDATE users SET full_name=%s,maktab_id=%s,class=%s,class_letter=%s WHERE user_id=%s",
                        (row["student_name"], school_id, row["class_code"].split("-")[0], row["class_code"].split("-")[1], student_id),
                    )
                    cur.execute(
                        """UPDATE school_person_profiles SET address=COALESCE(%s,address),
                                  mahalla=COALESCE(%s,mahalla),updated_at=NOW() WHERE user_id=%s""",
                        (row["address"], row["mahalla"], student_id),
                    )
                else:
                    student_id = next_user_id; next_user_id -= 1
                    cur.execute("INSERT INTO users(user_id,full_name,role,maktab_id,class,class_letter) VALUES(%s,%s,'oquvchi',%s,%s,%s)", (
                        student_id, row["student_name"], school_id,
                        row["class_code"].split("-")[0], row["class_code"].split("-")[1],
                    ))
                    cur.execute(
                        """INSERT INTO school_person_profiles(user_id,school_id,person_type,birth_year,normalized_name,address,mahalla)
                           VALUES(%s,%s,'student',%s,%s,%s,%s)""",
                        (student_id, school_id, row["student_birth_year"], normalize_text(row["student_name"]), row["address"], row["mahalla"]),
                    )
                    created_students += 1
                selected_parent = decision.get("same_parent_id", row.get("exact_parent_id"))
                if isinstance(selected_parent, str) and selected_parent.startswith("batch:"):
                    try:
                        source_row = int(selected_parent.split(":", 1)[1])
                        parent_id = parent_by_row[source_row]
                    except (KeyError, TypeError, ValueError):
                        raise HTTPException(status_code=400, detail=f'{row["row_number"]}-qator: fayldagi ota/ona tanlovi yaroqsiz')
                    canonical_name = str(decision.get("canonical_name") or "").strip()
                    if canonical_name:
                        cur.execute("SELECT full_name FROM users WHERE user_id=%s", (parent_id,))
                        previous = cur.fetchone()
                        if previous and normalize_text(previous["full_name"]) != normalize_text(canonical_name):
                            cur.execute(
                                "INSERT INTO school_person_aliases(user_id,alias_name,normalized_alias) VALUES(%s,%s,%s) ON CONFLICT DO NOTHING",
                                (parent_id, previous["full_name"], normalize_text(previous["full_name"])),
                            )
                        cur.execute("UPDATE users SET full_name=%s WHERE user_id=%s", (canonical_name, parent_id))
                        cur.execute("UPDATE school_person_profiles SET normalized_name=%s,updated_at=NOW() WHERE user_id=%s", (normalize_text(canonical_name), parent_id))
                elif selected_parent in ("new", "different", 0, "0", None):
                    parent_id = next_user_id; next_user_id -= 1
                    cur.execute("INSERT INTO users(user_id,full_name,role) VALUES(%s,%s,'ota-ona')", (parent_id, row["parent_name"]))
                    cur.execute(
                        """INSERT INTO school_person_profiles(user_id,school_id,person_type,birth_year,normalized_name,address,mahalla,guardian_type)
                           VALUES(%s,%s,'parent',%s,%s,%s,%s,%s)""",
                        (parent_id, school_id, row["parent_birth_year"], normalize_text(row["parent_name"]), row["address"], row["mahalla"], normalize_text(row["parent_type"])),
                    )
                    created_parents += 1
                else:
                    parent_id = int(selected_parent)
                    allowed_ids = {int(item["user_id"]) for item in row.get("parent_candidates", [])}
                    if row.get("exact_parent_id"):
                        allowed_ids.add(int(row["exact_parent_id"]))
                    if parent_id not in allowed_ids:
                        raise HTTPException(status_code=400, detail=f'{row["row_number"]}-qator: ota/ona tanlovi yaroqsiz')
                    canonical_name = str(decision.get("canonical_name") or "").strip()
                    if canonical_name:
                        cur.execute("SELECT full_name FROM users WHERE user_id=%s", (parent_id,))
                        previous = cur.fetchone()
                        if previous and normalize_text(previous["full_name"]) != normalize_text(canonical_name):
                            cur.execute(
                                "INSERT INTO school_person_aliases(user_id,alias_name,normalized_alias) VALUES(%s,%s,%s) ON CONFLICT DO NOTHING",
                                (parent_id, previous["full_name"], normalize_text(previous["full_name"])),
                            )
                        cur.execute("UPDATE users SET full_name=%s WHERE user_id=%s", (canonical_name, parent_id))
                        cur.execute("UPDATE school_person_profiles SET normalized_name=%s,updated_at=NOW() WHERE user_id=%s", (normalize_text(canonical_name), parent_id))
                    if normalize_text(row["parent_name"]) not in {normalize_text(canonical_name), ""}:
                        cur.execute(
                            "INSERT INTO school_person_aliases(user_id,alias_name,normalized_alias) VALUES(%s,%s,%s) ON CONFLICT DO NOTHING",
                            (parent_id, row["parent_name"], normalize_text(row["parent_name"])),
                        )
                parent_by_row[int(row["row_number"])] = parent_id
                cur.execute("SELECT 1 FROM parent_child WHERE parent_id=%s AND child_id=%s", (parent_id, student_id))
                if not cur.fetchone():
                    cur.execute("INSERT INTO parent_child(parent_id,child_id) VALUES(%s,%s)", (parent_id, student_id))
                    linked += 1
                cur.execute("SELECT 1 FROM maktab_sinf_azolari WHERE sinf_id=%s AND user_id=%s", (row["class_id"], student_id))
                already_in_class = cur.fetchone() is not None
                cur.execute(
                    """DELETE FROM maktab_sinf_azolari roster
                        USING maktab_sinflari class_row
                        WHERE roster.sinf_id=class_row.id AND roster.user_id=%s
                          AND class_row.maktab_id=%s AND roster.sinf_id<>%s""",
                    (student_id, school_id, row["class_id"]),
                )
                if not already_in_class:
                    cur.execute("INSERT INTO maktab_sinf_azolari(sinf_id,user_id) VALUES(%s,%s)", (row["class_id"], student_id))
            result = {"created_students": created_students, "created_parents": created_parents, "parent_child_links": linked, "access_codes": access_codes}
            cur.execute("UPDATE school_import_jobs SET status='committed',decisions=%s::jsonb,result=%s::jsonb,committed_at=NOW() WHERE id=%s", (
                json.dumps(payload.decisions, ensure_ascii=False), json.dumps(result, ensure_ascii=False), job_id,
            ))
            conn.commit()
            return {"status": "committed", **result}
        except Exception:
            conn.rollback()
            raise
        finally:
            close(conn, cur)

    @router.post("/{scope_id}/staff/preview")
    async def preview_staff(scope_id: int, token: str, scope_kind: str = "school", file: UploadFile = File(...)):
        user_id = verify_token(token)
        content = await file.read()
        if len(content) > 10 * 1024 * 1024:
            raise HTTPException(status_code=413, detail="Excel fayli 10 MB dan katta bo'lmasligi kerak")
        sheet = load_sheet(content, "XODIMLAR")
        conn, cur = connection()
        try:
            school_id, _ = _can_manage(
                cur, user_id, scope_id, mutation=True,
                prefer_context=scope_kind == "context",
            )
            valid_classes = {
                (row["normalized_code"] or f'{row["sinf"]}-{row["harf"]}'): int(row["id"])
                for row in class_rows(cur, school_id)
            }
            catalog = _catalog(cur)
            preview = []
            leadership_claims: dict[str, int] = {}
            for row_number, row in rows_as_dicts(sheet):
                errors, warnings = [], []
                name = str(field(row, "f_i_sh", "fish") or "").strip()
                role_text = str(field(row, "lavozim") or "").strip()
                role = ROLE_ALIASES.get(normalize_text(role_text))
                specialty = str(field(row, "mutaxassisligi") or "").strip()
                if len(name) < 3: errors.append("F.I.Sh. majburiy")
                if not role: errors.append("Lavozim ro'yxatdagi qiymatlardan biri bo'lishi kerak")
                if role == "fan_oqituvchisi" and not specialty: errors.append("O'qituvchi uchun mutaxassislik majburiy")
                try:
                    birth_year = _year(field(row, "tugilgan_yili"), "Tug'ilgan yili")
                except ValueError as exc:
                    birth_year = None; errors.append(str(exc))
                try:
                    experience = _integer(field(row, "ish_staji"), "Ish staji")
                except ValueError as exc:
                    experience = None; errors.append(str(exc))
                leadership_raw = str(field(row, "sinf_rahbarligi") or "").strip()
                leadership = None
                if leadership_raw:
                    try:
                        leadership = normalize_class_code(leadership_raw)
                        if leadership not in valid_classes:
                            errors.append(f"{leadership} maktabda yaratilmagan")
                        elif leadership in leadership_claims:
                            errors.append(f"{leadership} rahbarligi {leadership_claims[leadership]}-qatorda ham kiritilgan")
                        else:
                            leadership_claims[leadership] = row_number
                    except ValueError as exc:
                        errors.append(str(exc))
                assignments = parse_teaching_assignments(field(row, "dars_birikmalari", "oqitadigan_fanlari"), catalog)
                for assignment in assignments:
                    if assignment["class_code"] and assignment["class_code"] not in valid_classes:
                        errors.append(f'{assignment["class_code"]} maktabda yaratilmagan')
                    if not assignment["subject"]:
                        errors.append(f'"{assignment["input"]}" fani topilmadi')
                    elif assignment["needs_confirmation"]:
                        warnings.append(f'{assignment["input"]} → {assignment["subject"]} ({assignment["score"]}%): tasdiqlang')
                cur.execute(
                    """SELECT u.user_id,u.full_name,p.normalized_name
                         FROM users u
                         LEFT JOIN school_staff_profiles p ON p.user_id=u.user_id
                        WHERE u.maktab_id=%s""",
                    (school_id,),
                )
                wanted_name = normalize_text(name)
                existing = next(
                    (
                        candidate for candidate in cur.fetchall()
                        if (candidate["normalized_name"] or normalize_text(candidate["full_name"])) == wanted_name
                    ),
                    None,
                )
                preview.append({
                    "row_number": row_number, "name": name, "role": role,
                    "role_label": ROLE_LABELS.get(role, role_text), "specialty": specialty,
                    "category": str(field(row, "toifasi") or "").strip() or None,
                    "birth_year": birth_year, "experience_years": experience,
                    "address": str(field(row, "yashash_manzili") or "").strip() or None,
                    "leadership_class": leadership,
                    "leadership_class_id": valid_classes.get(leadership),
                    "assignments": assignments, "existing_user_id": int(existing["user_id"]) if existing else None,
                    "errors": list(dict.fromkeys(errors)), "warnings": warnings,
                    "status": "error" if errors else ("decision_required" if warnings else "ready"),
                })
            job_id = create_job(cur, school_id, user_id, "staff", preview)
            conn.commit()
            return {
                "job_id": job_id, "rows": preview,
                "summary": {
                    "total": len(preview), "ready": sum(r["status"] == "ready" for r in preview),
                    "decision_required": sum(r["status"] == "decision_required" for r in preview),
                    "errors": sum(r["status"] == "error" for r in preview),
                },
            }
        except Exception:
            conn.rollback()
            raise
        finally:
            close(conn, cur)

    @router.post("/imports/{job_id}/staff/commit")
    def commit_staff(job_id: int, payload: ImportCommit):
        user_id = verify_token(payload.token)
        conn, cur = connection()
        try:
            cur.execute("SELECT * FROM school_import_jobs WHERE id=%s FOR UPDATE", (job_id,))
            job = cur.fetchone()
            if not job or job["import_type"] != "staff":
                raise HTTPException(status_code=404, detail="Xodim import qoralamasi topilmadi")
            school_id, _ = _can_manage(cur, user_id, int(job["school_id"]), mutation=True)
            if job["status"] == "committed":
                return {"status": "committed", "reused": True, **(job["result"] or {})}
            rows = job["payload"] if isinstance(job["payload"], list) else json.loads(job["payload"])
            if any(row["status"] == "error" for row in rows):
                raise HTTPException(status_code=409, detail="Avval Excel xatolarini tuzating")
            next_user_id = _next_negative_user_id(cur)
            created = updated = 0
            access_codes = []
            for row in rows:
                if row["status"] == "decision_required" and not payload.decisions.get(str(row["row_number"]), {}).get("accept_subjects"):
                    raise HTTPException(status_code=409, detail=f'{row["row_number"]}-qator: fan tavsiyalarini tasdiqlang')
                employee_id = row.get("existing_user_id")
                if employee_id:
                    employee_id = int(employee_id); updated += 1
                    cur.execute("UPDATE users SET full_name=%s,lavozim=%s,fanlari=%s,ish_staji=%s,toifasi=%s,maktab_id=%s WHERE user_id=%s", (
                        row["name"], row["role"], row["specialty"] or None, row["experience_years"], row["category"], school_id, employee_id,
                    ))
                else:
                    employee_id = next_user_id; next_user_id -= 1; created += 1
                    cur.execute("INSERT INTO users(user_id,full_name,role,maktab_id,lavozim,fanlari,ish_staji,toifasi) VALUES(%s,%s,'oqituvchi',%s,%s,%s,%s,%s)", (
                        employee_id, row["name"], school_id, row["role"], row["specialty"] or None, row["experience_years"], row["category"],
                    ))
                    plain = "".join(secrets.choice(string.ascii_uppercase + string.digits) for _ in range(12))
                    stored = "sha256:" + hashlib.sha256(plain.encode("utf-8")).hexdigest()
                    cur.execute("INSERT INTO xodim_kod(kod,user_id) VALUES(%s,%s)", (stored, employee_id))
                    access_codes.append({"name": row["name"], "code": plain})
                cur.execute(
                    """INSERT INTO school_staff_profiles(user_id,school_id,normalized_name,specialty,category,birth_year,experience_years,address)
                       VALUES(%s,%s,%s,%s,%s,%s,%s,%s)
                       ON CONFLICT(user_id) DO UPDATE SET school_id=EXCLUDED.school_id,
                         normalized_name=EXCLUDED.normalized_name,specialty=EXCLUDED.specialty,
                         category=EXCLUDED.category,birth_year=EXCLUDED.birth_year,
                         experience_years=EXCLUDED.experience_years,address=EXCLUDED.address,updated_at=NOW()""",
                    (employee_id, school_id, normalize_text(row["name"]), row["specialty"] or None, row["category"], row["birth_year"], row["experience_years"], row["address"]),
                )
                if row["role"] == "direktor":
                    cur.execute("UPDATE maktablar SET direktor_user_id=%s,updated_at=NOW() WHERE id=%s", (employee_id, school_id))
                if row.get("leadership_class_id"):
                    cur.execute("UPDATE maktab_sinflari SET rahbar_user_id=%s,updated_at=NOW() WHERE id=%s AND maktab_id=%s", (employee_id, row["leadership_class_id"], school_id))
                cur.execute("DELETE FROM school_teacher_assignments WHERE school_id=%s AND teacher_user_id=%s", (school_id, employee_id))
                row_decision = payload.decisions.get(str(row["row_number"]), {})
                subject_choices = row_decision.get("subject_choices", {})
                for assignment_index, assignment in enumerate(row.get("assignments", [])):
                    subject_name = assignment["subject"]
                    if assignment.get("needs_confirmation"):
                        subject_name = subject_choices.get(str(assignment_index))
                        allowed_subjects = {option["subject"] for option in assignment.get("alternatives", [])}
                        if subject_name not in allowed_subjects:
                            raise HTTPException(status_code=409, detail=f'{row["row_number"]}-qator: fan variantini tanlang')
                    class_id = None
                    if assignment.get("class_code"):
                        cur.execute("SELECT id FROM maktab_sinflari WHERE maktab_id=%s AND normalized_code=%s", (school_id, assignment["class_code"]))
                        found = cur.fetchone(); class_id = found["id"] if found else None
                    cur.execute(
                        """INSERT INTO school_teacher_assignments(school_id,teacher_user_id,class_id,subject_name,subject_normalized)
                           VALUES(%s,%s,%s,%s,%s) ON CONFLICT DO NOTHING""",
                        (school_id, employee_id, class_id, subject_name, normalize_text(subject_name)),
                    )
            result = {"created_staff": created, "updated_staff": updated, "access_codes": access_codes}
            cur.execute("UPDATE school_import_jobs SET status='committed',decisions=%s::jsonb,result=%s::jsonb,committed_at=NOW() WHERE id=%s", (
                json.dumps(payload.decisions, ensure_ascii=False), json.dumps(result, ensure_ascii=False), job_id,
            ))
            conn.commit()
            return {"status": "committed", **result}
        except Exception:
            conn.rollback()
            raise
        finally:
            close(conn, cur)

    return router
