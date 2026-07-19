"""main.py — SamTM Ta'lim veb-sayt backend'i (v3).

Haqiqiy jadvallarga ulangan + Google orqali kirish (OAuth) qo'shildi.
"""
import os
import re
import io
import math
import secrets
import string
import httpx
import psycopg2
import psycopg2.extras
from typing import Optional
from datetime import datetime, timedelta, timezone
from jose import jwt, JWTError
from fastapi import FastAPI, HTTPException, Query, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import RedirectResponse, Response
from pydantic import BaseModel

DATABASE_URL = os.getenv("DATABASE_URL", "")
BOT_TOKEN = os.getenv("BOT_TOKEN", "")
GOOGLE_CLIENT_ID = os.getenv("GOOGLE_CLIENT_ID", "")
GOOGLE_CLIENT_SECRET = os.getenv("GOOGLE_CLIENT_SECRET", "")
JWT_MAXFIY_KALIT = os.getenv("JWT_MAXFIY_KALIT", "")
BAZA_URL = os.getenv("BAZA_URL", "https://talimplatformasi-production.up.railway.app")
FRONTEND_URL = os.getenv("FRONTEND_URL", "https://talimplatformasi-production.up.railway.app")
REDIRECT_URI = f"{BAZA_URL}/auth/google/callback"

app = FastAPI(title="SamTM Ta'lim API")
app.add_middleware(
    CORSMiddleware, allow_origins=["*"], allow_methods=["GET", "POST", "PUT", "DELETE"], allow_headers=["*"],
)


def _db():
    return psycopg2.connect(DATABASE_URL, cursor_factory=psycopg2.extras.RealDictCursor)


# Fan kodiga qarab dashboard rangi — yangi fan qo'shilsa shu ro'yxatga qo'shiladi
FAN_RANG = {
    "MAT": "#C89B3C", "TIL": "#2D8B8B", "ADB": "#8B5FBF",
    "TAB": "#B0553A", "RUS": "#4A7C9E", "ENG": "#7C9E4A",
}


@app.get("/")
def salomat():
    return {"holat": "ishlayapti"}


@app.get("/api/bola/{bola_id}/bilim")
def bola_bilimi(bola_id: int, sinf: str = None):
    """Bolaning fan-mavzu bo'yicha bilim darajasi — FAQAT bolaning O'ZI
    sinfiga tegishli mavzular bo'yicha. sinf berilmasa, avtomatik bola
    profilidagi class ustunidan olinadi. MUHIM: agar bola profilida
    sinf umuman ko'rsatilmagan bo'lsa — BARCHA sinflarni ARALASH
    ko'rsatish O'RNIGA bo'sh natija qaytariladi (aks holda 1-sinf
    bolasiga Algebra kabi butunlay boshqa sinflarning fanlari chiqib
    ketardi, chunki sinfsiz cheklov qo'yib bo'lmaydi)."""
    try:
        conn = _db()
        cur = conn.cursor()

        cur.execute("SELECT full_name, class FROM users WHERE user_id=%s", (bola_id,))
        bola = cur.fetchone()
        if not bola:
            raise HTTPException(status_code=404, detail="Bola topilmadi")

        if not sinf:
            if not bola["class"]:
                cur.close()
                conn.close()
                return {
                    "bola": {"ism": bola["full_name"]}, "umumiy_foiz": 0, "fanlar": [],
                    "jami_mavzu": 0, "otilgan_mavzu": 0, "sinf_sozlanmagan": True,
                }
            sinf = str(bola["class"]).replace("-sinf", "").strip()

        sinf_shart = "AND d.grade = %s" if sinf else ""
        params = (bola_id, sinf) if sinf else (bola_id,)

        cur.execute(f"""
            SELECT d.subject_code, d.subject_name, d.topic_code,
                   COALESCE(d.mavzu_name, d.bolim_name, d.bob_name) AS mavzu_nomi,
                   lt.score
            FROM dts_tree d
            LEFT JOIN learned_topics lt
                ON lt.topic_code = d.topic_code AND lt.user_id = %s
            WHERE 1=1 {sinf_shart}
            ORDER BY d.subject_code, d.topic_code
        """, params)
        qatorlar = cur.fetchall()
        cur.close()
        conn.close()

        fanlar = {}
        for q in qatorlar:
            kod = q["subject_code"] or "BOSHQA"
            if kod not in fanlar:
                fanlar[kod] = {
                    "nom": q["subject_name"] or kod, "qisqa": kod,
                    "rang": FAN_RANG.get(kod, "#8A8578"), "mavzular": [],
                }
            if q["score"] is not None:   # faqat o'rganilgan mavzular ko'rsatiladi
                fanlar[kod]["mavzular"].append({
                    "nom": q["mavzu_nomi"], "foiz": q["score"],
                })

        # Hali birorta ham mavzu o'rganilmagan fanlarni chiqarmaymiz
        natija_royxat = [f for f in fanlar.values() if f["mavzular"]]
        for f in natija_royxat:
            f["foiz"] = round(sum(m["foiz"] for m in f["mavzular"]) / len(f["mavzular"]))

        umumiy = round(sum(f["foiz"] for f in natija_royxat) / len(natija_royxat)) if natija_royxat else 0
        jami_mavzu_soni = len({q["topic_code"] for q in qatorlar})
        otilgan_mavzu_soni = len({q["topic_code"] for q in qatorlar if q["score"] is not None})

        return {
            "bola": {"ism": bola["full_name"]}, "umumiy_foiz": umumiy, "fanlar": natija_royxat,
            "jami_mavzu": jami_mavzu_soni, "otilgan_mavzu": otilgan_mavzu_soni,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/ota/{ota_id}/farzandlar")
def ota_farzandlari(ota_id: int):
    """Ota-onaning barcha ulangan farzandlari ro'yxati."""
    try:
        conn = _db()
        cur = conn.cursor()
        cur.execute("""
            SELECT u.user_id, u.full_name FROM parent_child pc
            JOIN users u ON u.user_id = pc.child_id
            WHERE pc.parent_id = %s
        """, (ota_id,))
        r = cur.fetchall()
        cur.close(); conn.close()
        return {"farzandlar": r}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ═══════════════════════════════════════════════════════════
# GOOGLE ORQALI KIRISH (OAuth)
# ═══════════════════════════════════════════════════════════

def _jwt_yarat(user_id: int) -> str:
    """30 kun amal qiladigan sessiya tokeni yaratadi."""
    payload = {
        "user_id": user_id,
        "exp": datetime.now(timezone.utc) + timedelta(days=30),
    }
    return jwt.encode(payload, JWT_MAXFIY_KALIT, algorithm="HS256")


def _jwt_tekshir(token: str) -> int:
    """Tokenni tekshiradi, user_id qaytaradi. Noto'g'ri bo'lsa xato beradi."""
    try:
        payload = jwt.decode(token, JWT_MAXFIY_KALIT, algorithms=["HS256"])
        return payload["user_id"]
    except JWTError:
        raise HTTPException(status_code=401, detail="Sessiya eskirgan, qaytadan kiring")


@app.get("/auth/google/login")
def google_login():
    """Foydalanuvchini Google'ning kirish sahifasiga yo'naltiradi."""
    url = (
        "https://accounts.google.com/o/oauth2/v2/auth"
        f"?client_id={GOOGLE_CLIENT_ID}"
        f"&redirect_uri={REDIRECT_URI}"
        "&response_type=code"
        "&scope=openid%20email%20profile"
        "&access_type=online"
    )
    return RedirectResponse(url)


@app.get("/auth/google/callback")
async def google_callback(code: str = None, error: str = None):
    """Google qaytargandan keyin ishlaydi — email oladi, bog'langan-bog'lanmaganini
    tekshiradi, mos ekranga yo'naltiradi."""
    if error or not code:
        return RedirectResponse(f"{FRONTEND_URL}/?xato=kirish_bekor")

    async with httpx.AsyncClient() as client:
        token_resp = await client.post(
            "https://oauth2.googleapis.com/token",
            data={
                "client_id": GOOGLE_CLIENT_ID,
                "client_secret": GOOGLE_CLIENT_SECRET,
                "code": code,
                "grant_type": "authorization_code",
                "redirect_uri": REDIRECT_URI,
            },
        )
        token_data = token_resp.json()
        if "access_token" not in token_data:
            return RedirectResponse(f"{FRONTEND_URL}/?xato=google_token")

        userinfo_resp = await client.get(
            "https://www.googleapis.com/oauth2/v2/userinfo",
            headers={"Authorization": f"Bearer {token_data['access_token']}"},
        )
        userinfo = userinfo_resp.json()

    email = userinfo.get("email")
    ism = userinfo.get("name", "")
    if not email:
        return RedirectResponse(f"{FRONTEND_URL}/?xato=email_topilmadi")

    conn = _db()
    cur = conn.cursor()
    cur.execute("SELECT user_id FROM google_hisob WHERE google_email=%s", (email,))
    r = cur.fetchone()
    cur.close()
    conn.close()

    if r:
        token = _jwt_yarat(r["user_id"])
        return RedirectResponse(f"{FRONTEND_URL}/kabinet?token={token}")
    else:
        return RedirectResponse(f"{FRONTEND_URL}/ulash?email={email}&ism={ism}")


class UlashSorov(BaseModel):
    email: str
    kod: str


class RoyxatSorov(BaseModel):
    email: str
    ism: str
    rol: str          # 'oquvchi' | 'ota-ona' | 'oqituvchi'
    sinf: Optional[str] = None  # faqat rol='oquvchi' bo'lsa
    region: Optional[str] = None
    district: Optional[str] = None
    tugilgan_sana: Optional[str] = None
    maktab_raqami: Optional[str] = None

RUXSAT_ETILGAN_ROLLAR = {"oquvchi", "ota-ona", "oqituvchi"}


@app.get("/auth/ism_tekshir")
def ism_tekshir(ism: str):
    """Botda shu ismga o'xshash foydalanuvchi bor-yo'qligini tekshiradi —
    saytdan yangi ro'yxatdan o'tishda, odam bilmasdan ikkinchi
    (dublikat) hisob ochib qo'ymasligi uchun ogohlantirish beriladi.
    Faqat BOTDAN kelgan (musbat user_id) foydalanuvchilar orasidan
    qidiradi — saytdan ro'yxatdan o'tganlar (manfiy ID) hisobga olinmaydi."""
    birinchi_soz = ism.strip().split()[0] if ism.strip() else ""
    if len(birinchi_soz) < 3:
        return {"oxshash": []}

    conn = _db()
    cur = conn.cursor()
    cur.execute("""
        SELECT full_name, role FROM users
        WHERE full_name ILIKE %s AND user_id > 0
        LIMIT 3
    """, (f"%{birinchi_soz}%",))
    natija = cur.fetchall()
    cur.close()
    conn.close()
    return {"oxshash": natija}


@app.post("/auth/royxat")
def yangi_royxat(sorov: RoyxatSorov):
    """Botsiz, to'g'ridan saytdan YANGI foydalanuvchi yaratadi.
    Telegram ID bilan TO'QNASHMASLIGI uchun MANFIY user_id beriladi
    (haqiqiy Telegram ID doim musbat bo'ladi)."""
    if sorov.rol not in RUXSAT_ETILGAN_ROLLAR:
        raise HTTPException(status_code=400, detail=f"Noto'g'ri rol: {sorov.rol}")
    if not sorov.ism.strip():
        raise HTTPException(status_code=400, detail="Ism kiritilmagan")

    conn = _db()
    cur = conn.cursor()
    cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS tugilgan_sana DATE")
    cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS maktab_raqami TEXT")

    cur.execute("SELECT user_id FROM google_hisob WHERE google_email=%s", (sorov.email,))
    if cur.fetchone():
        cur.close(); conn.close()
        raise HTTPException(status_code=400, detail="Bu email allaqachon ulangan — kirish orqali davom eting")

    cur.execute("SELECT MIN(user_id) AS eng_kichik FROM users WHERE user_id < 0")
    r = cur.fetchone()
    yangi_id = (r["eng_kichik"] - 1) if r and r["eng_kichik"] is not None else -1

    cur.execute(
        """INSERT INTO users(user_id, full_name, role, class, region, district, tugilgan_sana, maktab_raqami)
           VALUES(%s,%s,%s,%s,%s,%s,%s,%s)""",
        (yangi_id, sorov.ism.strip(), sorov.rol, sorov.sinf if sorov.rol == "oquvchi" else None,
         sorov.region, sorov.district, sorov.tugilgan_sana, sorov.maktab_raqami),
    )
    cur.execute(
        "INSERT INTO google_hisob(google_email, user_id) VALUES(%s,%s)",
        (sorov.email, yangi_id),
    )
    conn.commit()
    cur.close()
    conn.close()

    token = _jwt_yarat(yangi_id)
    return {"token": token, "user_id": yangi_id, "holat": "royxatdan otdi"}


@app.post("/auth/ulash")
def hisob_ulash(sorov: UlashSorov):
    """Google hisobini bot user_id'siga kod orqali bog'laydi. Ikki xil
    kod manbasini tekshiradi: botdagi veb_ulash_kod (15 daqiqa amal
    qiladi) VA maktab xodimlari uchun xodim_kod (30 kun amal qiladi,
    admin Excel orqali xodim import qilganda yaratiladi) — shu sabab
    bitta "kod kiritish" ekrani ikkalasi uchun ham ishlaydi."""
    email, kod = sorov.email, sorov.kod
    conn = _db()
    cur = conn.cursor()
    cur.execute("""
        SELECT user_id, ishlatildi,
               (yaratildi > NOW() - INTERVAL '15 minutes') AS hali_yangi
        FROM veb_ulash_kod WHERE kod=%s
    """, (kod,))
    r = cur.fetchone()
    muddat_matni = "15 daqiqa"
    jadval_nomi = "veb_ulash_kod"

    if not r:
        cur.execute("SELECT 1 FROM information_schema.tables WHERE table_name='xodim_kod'")
        if cur.fetchone():
            cur.execute("""
                SELECT user_id, ishlatildi,
                       (yaratildi > NOW() - INTERVAL '30 days') AS hali_yangi
                FROM xodim_kod WHERE kod=%s
            """, (kod,))
            r = cur.fetchone()
            muddat_matni = "30 kun"
            jadval_nomi = "xodim_kod"

    if not r:
        cur.close(); conn.close()
        raise HTTPException(status_code=400, detail="Kod noto'g'ri")
    if r["ishlatildi"]:
        cur.close(); conn.close()
        raise HTTPException(status_code=400, detail="Kod allaqachon ishlatilgan")
    if not r["hali_yangi"]:
        cur.close(); conn.close()
        raise HTTPException(status_code=400, detail=f"Kod muddati tugagan ({muddat_matni}) — qaytadan so'rang")

    cur.execute("""
        INSERT INTO google_hisob (google_email, user_id) VALUES (%s,%s)
        ON CONFLICT (google_email) DO UPDATE SET user_id=EXCLUDED.user_id
    """, (email, r["user_id"]))
    cur.execute(f"UPDATE {jadval_nomi} SET ishlatildi=TRUE WHERE kod=%s", (kod,))
    conn.commit()
    cur.close()
    conn.close()

    token = _jwt_yarat(r["user_id"])
    return {"token": token, "holat": "ulandi"}


@app.get("/auth/men")
def joriy_foydalanuvchi(token: str):
    """Token orqali 'bu kim' ekanini tasdiqlaydi — frontend sahifa yuklanganda
    ishlatadi. Admin bo'lsa, is_admin=true qaytadi — frontend shunga qarab
    sinf-cheklovini olib tashlaydi (admin barcha sinflarni ko'rishi kerak)."""
    user_id = _jwt_tekshir(token)
    conn = _db()
    cur = conn.cursor()
    cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS tugilgan_sana DATE")
    cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS maktab_raqami TEXT")
    cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS jins TEXT")
    cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS oqituvchi_fani TEXT")
    cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS maktab_id INTEGER")
    cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS markaz_id INTEGER")
    cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS bogcha_id INTEGER")
    cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS universitet_id INTEGER")
    cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS lavozim TEXT")
    cur.execute(
        "SELECT user_id, full_name, role, class, class_letter, school_type, "
        "region, district, tugilgan_sana, maktab_raqami, jins, oqituvchi_fani, "
        "maktab_id, markaz_id, bogcha_id, universitet_id, lavozim FROM users WHERE user_id=%s",
        (user_id,),
    )
    r = cur.fetchone()
    if not r:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Foydalanuvchi topilmadi")

    if r["maktab_id"]:
        _maktab_jadvali(cur)
        cur.execute("SELECT nomi FROM maktablar WHERE id=%s", (r["maktab_id"],))
        m = cur.fetchone()
        r["maktab_nomi"] = m["nomi"] if m else None
    else:
        r["maktab_nomi"] = None

    cur.execute("SELECT 1 FROM admin_akkaunt WHERE uid=%s", (user_id,))
    r["is_admin"] = cur.fetchone() is not None
    cur.close()
    conn.close()
    return r


# ═══════════════════════════════════════════════════════════
# TEST YECHISH (saytdan, botsiz)
# ═══════════════════════════════════════════════════════════

@app.get("/api/mavzular")
def mavzular_royxati(sinf: str = None, turi: str = "oddiy", faqat_testli: bool = True):
    """Fan/mavzularni qaytaradi — Fan → Sinf → Mavzu tartibida.

    MUHIM: bitta "mavzu" ostida bir nechta "kichik mavzu" bo'lishi mumkin
    (har biri o'z topic_code'iga ega) — lekin o'quvchiga BITTA mavzu
    IKKI MARTA (har kichik mavzu uchun alohida) ko'rinishi noto'g'ri va
    chalkashtiruvchi edi. Shu sabab bu yerda MAVZU darajasida guruhlaymiz:
    har mavzu — bitta yozuv, ichida esa BARCHA kichik mavzularning
    topic_code'lari "topic_codes" ro'yxatida jamlanadi. Test yechilganda
    shu ro'yxatdagi barcha kodlardan ARALASH (random) savol olinadi
    (/api/test_aralash orqali) — shunday qilib bitta "mavzu" tanlansa,
    uning barcha kichik mavzularidan birgalikda savol chiqadi.

    faqat_testli=True (standart, test yechish uchun) — faqat
    generated_tests'da HAQIQATAN savoli bor kichik mavzularni hisobga
    oladi (agar bir mavzuning faqat qismi testli bo'lsa, faqat o'sha
    testli qismidan savol olinadi). faqat_testli=False (admin
    kontent-yaratish oqimlari uchun) — testi hali yo'q mavzularni ham
    ko'rsatadi va BARCHA kichik mavzu kodlarini beradi.

    grade ustuni ba'zan "3-4", "5-6" kabi ORALIQ ko'rinishida bo'ladi —
    bular ODDIY maktab sinfi EMAS, balki TO'GARAKNING O'Z maxsus
    guruhlari. turi="oddiy" (standart) — faqat sof raqamli sinflar
    (1,2,...11). turi="togarak" — faqat ORALIQ (to'garak) guruhlari."""
    if sinf:
        sinf = sinf.replace("-sinf", "").strip()

    togarak_mi = turi == "togarak"
    grade_shart = "d.grade !~ '^[0-9]+$'" if togarak_mi else "d.grade ~ '^[0-9]+$'"

    conn = _db()
    cur = conn.cursor()
    shart = grade_shart
    params = []
    if sinf:
        shart += " AND d.grade = %s"
        params.append(sinf)
    cur.execute(f"""
        SELECT d.subject_code, d.subject_name, d.grade,
               COALESCE(d.mavzu_name, d.bolim_name, d.bob_name) AS nomi,
               array_agg(DISTINCT d.topic_code ORDER BY d.topic_code) AS barcha_kodlar,
               array_agg(DISTINCT d.topic_code ORDER BY d.topic_code)
                   FILTER (WHERE d.topic_code IN (SELECT DISTINCT topic_code FROM generated_tests)) AS testli_kodlar,
               COUNT(gt.id) AS savol_soni
        FROM dts_tree d
        LEFT JOIN generated_tests gt ON gt.topic_code = d.topic_code
        WHERE {shart} AND d.is_deleted = FALSE
        GROUP BY d.subject_code, d.subject_name, d.grade, COALESCE(d.mavzu_name, d.bolim_name, d.bob_name)
        ORDER BY d.subject_code, d.grade, MIN(d.topic_code)
    """, params)
    qatorlar = cur.fetchall()
    cur.close()
    conn.close()

    fanlar = {}
    for q in qatorlar:
        kodlar = q["testli_kodlar"] if faqat_testli else q["barcha_kodlar"]
        if faqat_testli and not kodlar:
            continue  # bu mavzuning hech bir kichik qismida test yo'q — test yechish ro'yxatida ko'rsatmaymiz

        fkod = q["subject_code"] or "BOSHQA"
        if fkod not in fanlar:
            fanlar[fkod] = {"nom": q["subject_name"] or fkod, "qisqa": fkod, "sinflar": {}}

        skod = q["grade"]
        if skod not in fanlar[fkod]["sinflar"]:
            fanlar[fkod]["sinflar"][skod] = {"sinf": skod, "mavzular": []}
        fanlar[fkod]["sinflar"][skod]["mavzular"].append({
            "topic_codes": kodlar, "nomi": q["nomi"], "savol_soni": q["savol_soni"],
        })

    natija = []
    for f in fanlar.values():
        if togarak_mi:
            # "3-4", "5-6" kabi — matn bo'yicha saralaymiz (raqamga aylantirib bo'lmaydi)
            f["sinflar"] = sorted(f["sinflar"].values(), key=lambda s: s["sinf"])
        else:
            # sinflarni SONLI tartibda saralaymiz (1,2,...,11 — "11" harflar bo'yicha "2"dan oldin kelib qolmasin)
            f["sinflar"] = sorted(f["sinflar"].values(), key=lambda s: int(s["sinf"]))
        natija.append(f)
    return {"fanlar": natija}


def _qoshimcha_test_shartlari(rasimli: bool, vaqtli: bool, yozuvli: bool):
    """rasimli/vaqtli/yozuvli — None bo'lsa cheklanmaydi (aralash), True/False
    bo'lsa mos savollar filtrlanadi. SQL parcha va parametrlarni qaytaradi."""
    shartlar = []
    params = []
    if rasimli is True:
        shartlar.append("COALESCE(NULLIF(image_file_id, ''), image_url, '') != ''")
    elif rasimli is False:
        shartlar.append("COALESCE(NULLIF(image_file_id, ''), image_url, '') = ''")
    if vaqtli is True:
        shartlar.append("COALESCE(time_limit, 0) > 0")
    elif vaqtli is False:
        shartlar.append("COALESCE(time_limit, 0) = 0")
    if yozuvli is True:
        shartlar.append("question_type = 'write_answer'")
    elif yozuvli is False:
        shartlar.append("question_type != 'write_answer'")
    return ("".join(f" AND {s}" for s in shartlar), params)


@app.get("/api/test/{topic_code}/soni")
def test_savollari_soni(topic_code: str, qiyinlik: str = None, rasimli: bool = None, vaqtli: bool = None, yozuvli: bool = None):
    """Tanlangan sozlamalar (qiyinlik/rasm/vaqt/javob turi) bo'yicha nechta
    savol MAVJUDLIGINI qaytaradi — test boshlanishidan OLDIN frontend shu
    yordamida haqiqiy sonni ko'rsatadi."""
    conn = _db()
    cur = conn.cursor()
    shart = "topic_code = %s"
    params = [topic_code]
    if qiyinlik:
        shart += " AND difficulty = %s"
        params.append(qiyinlik)
    qoshimcha, qoshimcha_params = _qoshimcha_test_shartlari(rasimli, vaqtli, yozuvli)
    shart += qoshimcha
    params += qoshimcha_params
    cur.execute(f"SELECT COUNT(*) AS soni FROM generated_tests WHERE {shart}", params)
    soni = cur.fetchone()["soni"]
    cur.close()
    conn.close()
    return {"soni": soni}


class AralashSoniSorovi(BaseModel):
    topic_codes: list = []
    qiyinlik: Optional[str] = None
    rasimli: Optional[bool] = None
    vaqtli: Optional[bool] = None
    yozuvli: Optional[bool] = None


@app.post("/api/test_aralash/soni")
def aralash_savollari_soni(sorov: AralashSoniSorovi):
    """Aralash (bir nechta mavzu) tanlanganda — sozlamalarga mos nechta
    savol mavjudligini qaytaradi. topic_codes ichida bo'sh/noto'g'ri
    qiymat bo'lsa ham (masalan null) 422 bermasdan, shunchaki e'tiborsiz
    qoldiradi — frontendga har doim aniq javob (soni: N) qaytadi."""
    kodlar = [str(k).strip() for k in sorov.topic_codes if k and str(k).strip()]
    if not kodlar:
        return {"soni": 0}
    conn = _db()
    cur = conn.cursor()
    shart = "topic_code = ANY(%s)"
    params = [kodlar]
    if sorov.qiyinlik:
        shart += " AND difficulty = %s"
        params.append(sorov.qiyinlik)
    qoshimcha, qoshimcha_params = _qoshimcha_test_shartlari(sorov.rasimli, sorov.vaqtli, sorov.yozuvli)
    shart += qoshimcha
    params += qoshimcha_params
    cur.execute(f"SELECT COUNT(*) AS soni FROM generated_tests WHERE {shart}", params)
    soni = cur.fetchone()["soni"]
    cur.close()
    conn.close()
    return {"soni": soni}


@app.get("/api/test/{topic_code}")
def test_savollari(
    topic_code: str, soni: int = 10, qiyinlik: str = None,
    rasimli: bool = None, vaqtli: bool = None, yozuvli: bool = None,
):
    """Berilgan mavzu bo'yicha tasodifiy savollarni qaytaradi.
    qiyinlik berilsa (oson/o'rta/qiyin/murakkab), faqat o'sha darajadagi
    savollar tanlanadi — bo'lmasa (aralash) barcha darajalardan aralash.
    rasimli/vaqtli/yozuvli — True/False bo'lsa mos savollargina tanlanadi,
    berilmasa (None) hammasidan aralash."""
    conn = _db()
    cur = conn.cursor()
    shart = "topic_code = %s"
    params = [topic_code]
    if qiyinlik:
        shart += " AND difficulty = %s"
        params.append(qiyinlik)
    qoshimcha, qoshimcha_params = _qoshimcha_test_shartlari(rasimli, vaqtli, yozuvli)
    shart += qoshimcha
    params += qoshimcha_params
    params.append(soni)
    cur.execute(f"""
        SELECT id, question, option_a, option_b, option_c, option_d,
               question_type, is_latex, time_limit, difficulty,
               COALESCE(NULLIF(image_file_id, ''), image_url) AS rasm_id
        FROM generated_tests
        WHERE {shart}
        ORDER BY RANDOM()
        LIMIT %s
    """, params)
    savollar = cur.fetchall()
    cur.close()
    conn.close()

    if not savollar:
        raise HTTPException(status_code=404, detail="Bu mavzuda (tanlangan sozlamalar bo'yicha) savol topilmadi")

    # DIQQAT: bu yerda [ru]/[en] teglarini ATAYLAB OLIB TASHLAMAYMIZ —
    # frontend ularni ko'rsatishda yashiradi, lekin ovoz o'qishda AYNAN shu
    # teglar orqali qaysi so'z qaysi tilda o'qilishini aniqlaydi. Faqat
    # "10.0" -> "10" kabi raqam artefaktini tozalaymiz.
    for s in savollar:
        s["question"] = _raqam_artefaktini_tozala(s["question"])
        for maydon in ("option_a", "option_b", "option_c", "option_d"):
            s[maydon] = _raqam_artefaktini_tozala(s[maydon])

    # correct_answer va explanation FRONTENDGA yubormaymiz — bular javob
    # berilgandan KEYIN, /api/test/javob_tekshir orqali ochiladi
    return {"topic_code": topic_code, "savollar": savollar}


class AralashTestSorovi(BaseModel):
    topic_codes: list = []
    soni: int = 10
    qiyinlik: Optional[str] = None
    rasimli: Optional[bool] = None
    vaqtli: Optional[bool] = None
    yozuvli: Optional[bool] = None


@app.post("/api/test_aralash")
def aralash_test_savollari(sorov: AralashTestSorovi):
    """Bir nechta TANLANGAN mavzudan aralashtirib savollar oladi —
    o'quvchi bir nechta mavzuni bir vaqtda takrorlashi uchun."""
    kodlar = [str(k).strip() for k in sorov.topic_codes if k and str(k).strip()]
    if not kodlar:
        raise HTTPException(status_code=400, detail="Kamida bitta mavzu tanlang")

    conn = _db()
    cur = conn.cursor()
    shart = "topic_code = ANY(%s)"
    params = [kodlar]
    if sorov.qiyinlik:
        shart += " AND difficulty = %s"
        params.append(sorov.qiyinlik)
    qoshimcha, qoshimcha_params = _qoshimcha_test_shartlari(sorov.rasimli, sorov.vaqtli, sorov.yozuvli)
    shart += qoshimcha
    params += qoshimcha_params
    params.append(sorov.soni)
    cur.execute(f"""
        SELECT id, topic_code, question, option_a, option_b, option_c, option_d,
               question_type, is_latex, time_limit, difficulty,
               COALESCE(NULLIF(image_file_id, ''), image_url) AS rasm_id
        FROM generated_tests
        WHERE {shart}
        ORDER BY RANDOM()
        LIMIT %s
    """, params)
    savollar = cur.fetchall()
    cur.close()
    conn.close()

    if not savollar:
        raise HTTPException(status_code=404, detail="Tanlangan mavzu/sozlamalarda savol topilmadi")

    for s in savollar:
        s["question"] = _raqam_artefaktini_tozala(s["question"])
        for maydon in ("option_a", "option_b", "option_c", "option_d"):
            s[maydon] = _raqam_artefaktini_tozala(s[maydon])

    return {"topic_codes": kodlar, "savollar": savollar}


class BittaJavob(BaseModel):
    savol_id: int
    tanlangan: str


def _raqam_artefaktini_tozala(matn):
    """"10.0" kabi butun sonlarni "10" ga soddalashtiradi — teglarga tegmaydi."""
    if not matn:
        return matn
    tozalangan = matn.strip()
    if re.fullmatch(r"-?\d+\.0+", tozalangan):
        tozalangan = tozalangan.split(".")[0]
    return tozalangan


def _matnni_tozala(matn):
    """[ru]...[/ru] kabi teglarni olib tashlaydi, va "10.0" kabi butun
    sonlarni "10" ga soddalashtiradi — ham ko'rsatish, ham solishtirish
    uchun ishlatiladi."""
    if not matn:
        return matn
    tozalangan = re.sub(r"\[/?[a-zA-Z]+\]", "", matn).strip()
    if re.fullmatch(r"-?\d+\.0+", tozalangan):
        tozalangan = tozalangan.split(".")[0]
    return tozalangan


def _togri_harfni_top(option_a, option_b, option_c, option_d, correct_answer):
    """correct_answer ustuni ba'zan harf (A/B/C/D), ba'zan variantning
    TO'LIQ MATNI (masalan "20.0" yoki "[ru]родной язык[/ru]") ko'rinishida
    saqlangan — ikkalasini ham qamrab olib, HAQIQIY to'g'ri harfni
    aniqlaydi. Teglar va sonlar formatidagi farqlar e'tiborga olinmaydi."""
    ca = _matnni_tozala((correct_answer or "").strip())
    if ca.upper() in ("A", "B", "C", "D"):
        return ca.upper()
    variantlar = {"A": option_a, "B": option_b, "C": option_c, "D": option_d}
    ca_kichik = ca.lower()
    for harf, matn in variantlar.items():
        if (_matnni_tozala(matn) or "").lower() == ca_kichik:
            return harf
    return None


def _yozma_javob_togrimi(given: str, correct: str) -> bool:
    """Yozuvli (write_answer) javoblarni tekshiradi — botdagi
    check_text_answer/is_match bilan bir xil qoidalar."""
    given = _matnni_tozala(given or "").strip().lower()
    correct = _matnni_tozala(correct or "").strip().lower()
    if given == correct:
        return True
    try:
        return float(given) == float(correct)
    except (ValueError, TypeError):
        pass
    if len(correct) <= 5:
        return given == correct
    if len(correct) > 10 and correct in given:
        return True
    return False


@app.post("/api/test/javob_tekshir")
def javob_tekshir(j: BittaJavob):
    """Bitta savolga berilgan javobni DARHOL tekshiradi — to'g'ri javob
    va tushuntirishni shu yerda ochadi (foydalanuvchi javob bergandan
    keyin, savol ko'rsatilganda EMAS — aks holda oldindan ko'rinib qolardi).
    Yozuvli (write_answer) savollarda harf emas, yozilgan matn solishtiriladi."""
    conn = _db()
    cur = conn.cursor()
    cur.execute("""SELECT option_a, option_b, option_c, option_d, correct_answer,
                          explanation, question_type
                   FROM generated_tests WHERE id=%s""", (j.savol_id,))
    r = cur.fetchone()
    cur.close()
    conn.close()
    if not r:
        raise HTTPException(status_code=404, detail="Savol topilmadi")

    if r["question_type"] == "write_answer":
        togri = _yozma_javob_togrimi(j.tanlangan, r["correct_answer"])
        togri_javob = _matnni_tozala(r["correct_answer"])
    else:
        togri_javob = _togri_harfni_top(r["option_a"], r["option_b"], r["option_c"], r["option_d"], r["correct_answer"])
        togri = (j.tanlangan or "").strip().upper() == togri_javob

    return {"togrimi": togri, "togri_javob": togri_javob, "tushuntirish": _matnni_tozala(r["explanation"])}


@app.get("/api/rasm/{file_id}")
async def rasm_proxy(file_id: str):
    """Telegram'da saqlangan rasmni saytda ko'rsatish uchun oraliq xizmat.

    MUHIM: generated_tests.image_url ko'pincha haqiqiy Telegram file_id
    EMAS — "1-02-1-01-01-01-001-1" kabi KOLLAJ KODI bo'ladi. Botning o'zi
    ham bu kodni to'g'ridan-to'g'ri ishlatmaydi — avval "images" jadvalidan
    (name→file_id) haqiqiy Telegram file_id'ni qidiradi (Talim.py'dagi
    bilan AYNAN bir xil mantiq). Shu sabab bu yerda ham AVVAL images
    jadvalidan qidiramiz, faqat topilmasa file_id'ning O'ZINI ishlatamiz."""
    if not BOT_TOKEN:
        raise HTTPException(status_code=500, detail="Bot tokeni sozlanmagan")
    if file_id.startswith("http"):
        # Ba'zi eski yozuvlarda image_url to'g'ridan URL bo'lishi mumkin
        return RedirectResponse(file_id)

    haqiqiy_file_id = file_id
    try:
        conn = _db()
        cur = conn.cursor()
        cur.execute("SELECT file_id FROM images WHERE name=%s LIMIT 1", (file_id,))
        r = cur.fetchone()
        cur.close()
        conn.close()
        if r and r["file_id"]:
            haqiqiy_file_id = r["file_id"]
    except Exception:
        pass  # images jadvali bo'lmasa ham, file_id'ning o'zi bilan urinib ko'ramiz

    async with httpx.AsyncClient() as client:
        meta = await client.get(f"https://api.telegram.org/bot{BOT_TOKEN}/getFile",
                                 params={"file_id": haqiqiy_file_id})
        meta_data = meta.json()
        if not meta_data.get("ok"):
            raise HTTPException(status_code=404, detail="Rasm topilmadi")
        file_path = meta_data["result"]["file_path"]
        img = await client.get(f"https://api.telegram.org/file/bot{BOT_TOKEN}/{file_path}")
        return Response(content=img.content, media_type="image/jpeg")


EDGE_OVOZ = {
    "qiz": "uz-UZ-MadinaNeural",
    "ogil": "uz-UZ-SardorNeural",
}
_TIL_OVOZLARI = {
    "en": {"qiz": "en-US-JennyNeural", "ogil": "en-US-GuyNeural"},
    "ru": {"qiz": "ru-RU-SvetlanaNeural", "ogil": "ru-RU-DmitryNeural"},
}

# ── Ovoz uchun matnni tayyorlash — botdagi ovoz.py bilan bir xil qoidalar ──
_BIRLIK = ["", "bir", "ikki", "uch", "to'rt", "besh", "olti", "yetti", "sakkiz", "to'qqiz"]
_ONLIK = ["", "o'n", "yigirma", "o'ttiz", "qirq", "ellik", "oltmish", "yetmish", "sakson", "to'qson"]
_TARTIB = {
    "bir": "birinchi", "ikki": "ikkinchi", "uch": "uchinchi", "to'rt": "to'rtinchi",
    "besh": "beshinchi", "olti": "oltinchi", "yetti": "yettinchi", "sakkiz": "sakkizinchi",
    "to'qqiz": "to'qqizinchi", "o'n": "o'ninchi", "yigirma": "yigirmanchi", "o'ttiz": "o'ttizinchi",
    "qirq": "qirqinchi", "ellik": "ellikinchi", "oltmish": "oltmishinchi", "yetmish": "yetmishinchi",
    "sakson": "saksoninchi", "to'qson": "to'qsoninchi", "yuz": "yuzinchi", "ming": "minginchi",
}


def _son_soz(n: int) -> str:
    if n == 0:
        return "nol"
    if n < 0:
        return "minus " + _son_soz(-n)
    q = []
    if n >= 1000:
        m = n // 1000
        q.append("ming" if m == 1 else _son_soz(m) + " ming")
        n %= 1000
    if n >= 100:
        y = n // 100
        q.append("yuz" if y == 1 else _BIRLIK[y] + " yuz")
        n %= 100
    if n >= 10:
        q.append(_ONLIK[n // 10])
        n %= 10
    if n > 0:
        q.append(_BIRLIK[n])
    return " ".join(x for x in q if x)


_MATH_MAP = [
    (r"\s*\+\s*", " qo'shuv "),
    (r"(?<=\d)\s*-\s*(?=\d)", " ayirish "),
    (r"\s*×\s*|\s*\*\s*", " ko'paytiruv "),
    (r"\s*÷\s*", " bo'linadi "),
    (r"\s*=\s*", " teng "),
    (r"\s*>\s*", " katta "),
    (r"\s*<\s*", " kichik "),
    (r"\s*%\s*", " foiz "),
    (r"\s*≈\s*", " taxminan "),
]


_APOSTROF_VARIANTLARI = "\u2018\u2019\u02BB\u02BC\u0060\u00B4\u2032"


def _apostrofni_tuzat(matn: str) -> str:
    """o'/g' dan keyingi turli tirnoq-apostrof belgilarini ('  '  ʻ  ʼ  `  ´)
    bitta standart apostrofga keltiradi — aks holda ovoz ularni "o'"/"g'"
    deb emas, oddiy "o"/"g" deb yoki umuman boshqacha o'qib yuboradi."""
    return re.sub(rf"([oOgG])[{_APOSTROF_VARIANTLARI}']", r"\1'", matn)


def _c_va_w_tuzat(matn: str) -> str:
    """"c" harfini (agar "ch" qismi bo'lmasa) inglizcha qoidaga ko'ra
    s/k tovushiga, "w" ni esa "v" ga almashtiradi — o'zbekcha ovoz "c"ni
    "ch" deb, "w"ni esa noto'g'ri o'qib yuborishining oldini oladi."""
    natija = []
    n = len(matn)
    i = 0
    while i < n:
        ch = matn[i]
        if ch.lower() == "c" and (i + 1 >= n or matn[i + 1].lower() != "h"):
            keyingi = matn[i + 1] if i + 1 < n else ""
            alm = "s" if keyingi.lower() in ("e", "i", "y") else "k"
            natija.append(alm.upper() if ch.isupper() else alm)
        elif ch.lower() == "w":
            natija.append("V" if ch.isupper() else "v")
        else:
            natija.append(ch)
        i += 1
    return "".join(natija)


def _ovoz_uchun_tayyorla(matn: str) -> str:
    """Xom matn -> ovoz aniq o'qiydigan matn — botdagi ovoz.py:tayyorla
    bilan bir xil (matematik belgilar so'zga, sonlar so'zga, teglar tozalanadi)."""
    m = _matnni_tozala(matn) or ""
    m = _apostrofni_tuzat(m)
    m = _c_va_w_tuzat(m)
    m = re.sub(r"<[^>]+>", " ", m)
    m = re.sub(r"[_`#]+", "", m)  # * ni bu yerda OLIB TASHLAMAYMIZ — pastda MATH_MAP "ko'paytiruv"ga o'giradi
    m = re.sub(r"[\U0001F300-\U0001FAFF\u2600-\u27BF]", " ", m)
    m = re.sub(r"https?://\S+", " havola ", m)

    # Kasrlar: 1/2 -> ikkidan bir (matematikadan oldin)
    def _kasr(x):
        a, b = int(x.group(1)), int(x.group(2))
        return f" {_son_soz(b)}dan {_son_soz(a)} "
    m = re.sub(r"\b(\d{1,3})\s*/\s*(\d{1,3})\b", _kasr, m)

    for naqsh, alm in _MATH_MAP:
        m = re.sub(naqsh, alm, m)

    # 5-sinf -> beshinchi sinf
    def _t(x):
        n = int(x.group(1))
        soz = _son_soz(n).split()
        soz[-1] = _TARTIB.get(soz[-1], soz[-1] + "inchi")
        return f"{' '.join(soz)} {x.group(2)}"
    m = re.sub(r"\b(\d{1,4})-(sinf|mashq|dars|savol|misol|bob|bet|mavzu|qism)\b", _t, m, flags=re.I)

    # 3,5 -> uch butun besh
    def _b(x):
        return f"{_son_soz(int(x.group(1)))} butun {_son_soz(int(x.group(2)))}"
    m = re.sub(r"\b(\d+)[,.](\d+)\b", _b, m)

    # Qolgan sonlar so'zga
    def _o(x):
        n = int(x.group(0))
        return _son_soz(n) if n < 1000000 else x.group(0)
    m = re.sub(r"\b\d{1,6}\b", _o, m)

    # Tinish belgilarini pauzaga aylantirish
    m = m.replace(":", ",").replace(";", ",")
    m = re.sub(r"\s*[\(\[\{]\s*", ", ", m)
    m = re.sub(r"\s*[\)\]\}]\s*", ", ", m)
    m = re.sub(r'["«»„“”]', " ", m)
    m = re.sub(r"\s*[–—/|]\s*", ", ", m)
    m = re.sub(r"\s*[•▪●○*]\s*", ", ", m)
    m = re.sub(r"[…]+", ".", m)
    m = re.sub(r"\.{2,}", ".", m)
    m = re.sub(r"(?<=\w)-(?=\w)", " ", m)
    m = re.sub(r"(,\s*){2,}", ", ", m)
    m = re.sub(r"\s+([.,!?])", r"\1", m)
    m = re.sub(r",\s*([.!?])", r"\1", m)
    m = re.sub(r"([.!?])\s*[.,]+", r"\1", m)
    m = re.sub(r"([.!?])\s*([.!?])", r"\1", m)
    m = re.sub(r"\s{2,}", " ", m).strip()
    return m.strip(" ,.")


_TIL_TEG_NAQSHI = re.compile(r"\[(en|ru)\](.*?)\[/\1\]", re.S | re.I)


def _ovoz_qismlarga_bol(matn: str):
    """Matnni [en]...[/en] / [ru]...[/ru] teglariga qarab bo'laklarga
    ajratadi — har bo'lak (til, matn). til=None bo'lsa standart
    o'zbekcha ovoz va matematik-son qoidalari bilan o'qiladi."""
    qismlar = []
    oxiri = 0
    for m in _TIL_TEG_NAQSHI.finditer(matn):
        oldingi = matn[oxiri:m.start()]
        if oldingi.strip():
            qismlar.append((None, oldingi))
        til, ichi = m.group(1).lower(), m.group(2)
        if ichi.strip():
            qismlar.append((til, ichi))
        oxiri = m.end()
    qolgan = matn[oxiri:]
    if qolgan.strip():
        qismlar.append((None, qolgan))
    return qismlar or [(None, matn)]


@app.get("/api/ovoz")
async def ovoz_oqish(matn: str, jins: str = "qiz"):
    """Berilgan matnni ovozga aylantirib beradi (mp3). [en]/[ru] teglari
    ichidagi qismlar o'sha tilning ovozida, qolgani o'zbekcha (matematik
    belgilar/sonlar so'zga o'girilib) o'qiladi — botdagi ovoz_ikki_tilli
    bilan bir xil mantiq, faqat ikkita til uchun kengaytirilgan."""
    if not matn or not matn.strip():
        raise HTTPException(status_code=400, detail="Matn berilmagan")
    try:
        import edge_tts
    except ImportError:
        raise HTTPException(status_code=500, detail="edge-tts o'rnatilmagan")

    matn = matn[:1500]
    buf = io.BytesIO()
    ovoz_bormi = False
    for til, bolak in _ovoz_qismlarga_bol(matn):
        if til in _TIL_OVOZLARI:
            voice = _TIL_OVOZLARI[til].get(jins, _TIL_OVOZLARI[til]["qiz"])
            tayyor = re.sub(r"<[^>]+>", " ", bolak).strip()
        else:
            voice = EDGE_OVOZ.get(jins, EDGE_OVOZ["qiz"])
            tayyor = _ovoz_uchun_tayyorla(bolak)
        if not tayyor.strip():
            continue
        com = edge_tts.Communicate(tayyor, voice)
        async for chunk in com.stream():
            if chunk["type"] == "audio":
                buf.write(chunk["data"])
                ovoz_bormi = True

    if not ovoz_bormi:
        raise HTTPException(status_code=500, detail="Ovoz yaratilmadi")
    buf.seek(0)
    return Response(content=buf.getvalue(), media_type="audio/mpeg")


class JavobItem(BaseModel):
    savol_id: int
    tanlangan: str


class TestNatijaSorov(BaseModel):
    token: str
    topic_code: Optional[str] = None       # bitta mavzu bo'lsa
    topic_codes: Optional[list] = None  # aralash (bir nechta mavzu) bo'lsa
    javoblar: list[JavobItem]
    # UMUMIY natija foizini TANLANGAN (masalan 10 ta) savol soniga nisbatan
    # hisoblash uchun — javob berilmagan savollar ham hisobga olinishi kerak
    # (aks holda 10 tadan 5 tasiga javob berib, hammasi to'g'ri bo'lsa, "100%"
    # ko'rsatib qo'yardi, holbuki haqiqatda 50%). Berilmasa — eski xulq-atvorga
    # (faqat javob berilganlar soniga nisbatan) qaytiladi.
    jami_savol_soni: Optional[int] = None


@app.post("/api/test/natija")
def test_natijasini_saqla(sorov: TestNatijaSorov):
    """Test yakunlanganda — har javobni backendda tekshiradi, foizni
    hisoblaydi, learned_topics'ga yozadi (bot ishlatgan JADVALNING O'ZIGA —
    shuning uchun dashboard darhol yangilanadi). Yozuvli (write_answer)
    savollar ham to'g'ri tekshiriladi, va xato qilingan savollar ro'yxati
    (sharh bilan) qaytariladi. Aralash (bir nechta mavzu) test bo'lsa, HAR
    BIR mavzu o'ziga tegishli savollar asosida alohida baholanadi."""
    user_id = _jwt_tekshir(sorov.token)

    conn = _db()
    cur = conn.cursor()

    savol_idlar = [j.savol_id for j in sorov.javoblar]
    cur.execute(
        """SELECT id, topic_code, question, option_a, option_b, option_c, option_d,
                  correct_answer, question_type, explanation
           FROM generated_tests WHERE id = ANY(%s)""",
        (savol_idlar,),
    )
    savollar_map = {r["id"]: r for r in cur.fetchall()}

    togri_soni = 0
    xatolar = []
    natija_har_mavzu = {}  # topic_code -> {"togri": n, "jami": n}
    for j in sorov.javoblar:
        r = savollar_map.get(j.savol_id)
        if not r:
            continue
        if r["question_type"] == "write_answer":
            togri = _yozma_javob_togrimi(j.tanlangan, r["correct_answer"])
            togri_javob = _matnni_tozala(r["correct_answer"])
        else:
            togri_harf = _togri_harfni_top(r["option_a"], r["option_b"], r["option_c"], r["option_d"], r["correct_answer"])
            togri = (j.tanlangan or "").strip().upper() == togri_harf
            togri_javob = togri_harf

        tk = r["topic_code"]
        natija_har_mavzu.setdefault(tk, {"togri": 0, "jami": 0})
        natija_har_mavzu[tk]["jami"] += 1
        if togri:
            togri_soni += 1
            natija_har_mavzu[tk]["togri"] += 1
        else:
            xatolar.append({
                "savol_id": j.savol_id,
                "savol": _matnni_tozala(r["question"]),
                "sizning_javob": j.tanlangan or "(javob berilmadi)",
                "togri_javob": togri_javob,
                "tushuntirish": _matnni_tozala(r["explanation"]),
            })

    # UMUMIY foiz — agar frontend "jami_savol_soni" yuborsa (tanlangan
    # savollar soni), o'shanga nisbatan hisoblanadi — javob berilmagan
    # savollar ham "noto'g'ri" sifatida hisobga kiradi. FAQAT shu
    # ko'rsatkichga (natija ekranidagi statistika) tegishli — pastdagi
    # mavzu bo'yicha learned_topics hisobiga ASLO ta'sir qilmaydi.
    jami = sorov.jami_savol_soni if sorov.jami_savol_soni else len(sorov.javoblar)
    foiz = round((togri_soni / jami) * 100) if jami else 0

    # Har bir mavzu (aralash bo'lsa — bir nechtasi) o'ziga tegishli
    # savollar asosida alohida learned_topics'ga yoziladi.
    # MUHIM: bu FAQAT haqiqatan JAVOB BERILGAN savollar asosida hisoblanadi
    # (yuqoridagi tuzatish bunga tegmaydi) — o'quvchi o'zi urinib ko'rgan
    # mavzular bo'yicha bilim darajasi shu tarzda avvalgidek qoladi.
    for tk, hisob in natija_har_mavzu.items():
        if not tk:
            continue
        mavzu_foizi = round((hisob["togri"] / hisob["jami"]) * 100) if hisob["jami"] else 0
        cur.execute("""
            INSERT INTO learned_topics(user_id, topic_code, score, repeat_count, learned_at, next_repeat)
            VALUES(%s,%s,%s,1,NOW(),CURRENT_DATE + INTERVAL '7 days')
            ON CONFLICT (user_id, topic_code) DO UPDATE SET
                score = EXCLUDED.score,
                repeat_count = learned_topics.repeat_count + 1,
                learned_at = NOW(),
                next_repeat = CURRENT_DATE + INTERVAL '7 days'
        """, (user_id, tk, mavzu_foizi))
    conn.commit()
    cur.close()
    conn.close()

    return {"togri": togri_soni, "jami": jami, "foiz": foiz, "xatolar": xatolar}


# ═══════════════════════════════════════════════════════════
# SAYTDAN BOTGA ULASH — teskari yo'nalish
# (Saytda ro'yxatdan o'tgan, botni ham ishlatmoqchi bo'lganlar uchun)
# ═══════════════════════════════════════════════════════════

@app.post("/auth/sayt_kod_yarat")
def sayt_kod_yarat(token: str):
    """Saytda kirgan foydalanuvchi uchun BOTGA ulash kodi yaratadi.
    Bot bu kodni ko'rib, shu web_user_id'dagi ma'lumotni haqiqiy
    Telegram user_id'ga ko'chiradi."""
    user_id = _jwt_tekshir(token)

    kod = "".join(secrets.choice(string.ascii_uppercase + string.digits) for _ in range(6))
    conn = _db()
    cur = conn.cursor()
    cur.execute("""CREATE TABLE IF NOT EXISTS sayt_ulash_kod(
        kod TEXT PRIMARY KEY, web_user_id BIGINT REFERENCES users(user_id),
        yaratildi TIMESTAMP DEFAULT NOW(), ishlatildi BOOLEAN DEFAULT FALSE)""")
    cur.execute("INSERT INTO sayt_ulash_kod(kod, web_user_id) VALUES(%s,%s)", (kod, user_id))
    conn.commit()
    cur.close()
    conn.close()

    return {"kod": kod}


# ═══════════════════════════════════════════════════════════
# O'QITUVCHI — baholash
# ═══════════════════════════════════════════════════════════

@app.get("/api/oqituvchi/togaraklar")
def oqituvchi_togaraklari(token: str):
    """O'qituvchining o'ziga tegishli barcha to'garaklarini qaytaradi."""
    user_id = _jwt_tekshir(token)
    conn = _db()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, nomi, fan, max_talaba,
               (SELECT COUNT(*) FROM togarak_azolar WHERE togarak_id=togaraklar.id AND aktiv=TRUE) AS azo_soni
        FROM togaraklar
        WHERE teacher_id=%s AND aktiv=TRUE
        ORDER BY nomi
    """, (user_id,))
    natija = cur.fetchall()
    cur.close()
    conn.close()
    return {"togaraklar": natija}


@app.get("/api/oqituvchi/togarak/{togarak_id}/azolar")
def togarak_azolari(togarak_id: int, token: str):
    """Berilgan to'garakdagi o'quvchilarni, ularning OXIRGI bahosi bilan
    qaytaradi. Faqat shu to'garakning o'z o'qituvchisi ko'ra oladi."""
    user_id = _jwt_tekshir(token)
    conn = _db()
    cur = conn.cursor()

    cur.execute("SELECT teacher_id FROM togaraklar WHERE id=%s", (togarak_id,))
    r = cur.fetchone()
    if not r or r["teacher_id"] != user_id:
        cur.close()
        conn.close()
        raise HTTPException(status_code=403, detail="Bu to'garak sizga tegishli emas")

    cur.execute("""
        SELECT u.user_id, u.full_name,
               (SELECT baho FROM togarak_baholar tb
                WHERE tb.togarak_id=%s AND tb.user_id=u.user_id
                ORDER BY tb.created_at DESC LIMIT 1) AS oxirgi_baho
        FROM togarak_azolar ta
        JOIN users u ON u.user_id = ta.user_id
        WHERE ta.togarak_id=%s AND ta.aktiv=TRUE
        ORDER BY u.full_name
    """, (togarak_id, togarak_id))
    azolar = cur.fetchall()
    cur.close()
    conn.close()
    return {"azolar": azolar}


class BahoSorov(BaseModel):
    token: str
    togarak_id: int
    user_id: int
    baho: int
    izoh: Optional[str] = None


@app.post("/api/oqituvchi/baho_qoy")
def baho_qoy(sorov: BahoSorov):
    """Bitta o'quvchiga baho qo'yadi. Faqat to'garakning o'z o'qituvchisi,
    va faqat o'sha to'garak a'zosiga baho qo'ya oladi."""
    teacher_id = _jwt_tekshir(sorov.token)
    conn = _db()
    cur = conn.cursor()

    cur.execute("SELECT teacher_id FROM togaraklar WHERE id=%s", (sorov.togarak_id,))
    r = cur.fetchone()
    if not r or r["teacher_id"] != teacher_id:
        cur.close()
        conn.close()
        raise HTTPException(status_code=403, detail="Bu to'garak sizga tegishli emas")

    cur.execute(
        "SELECT 1 FROM togarak_azolar WHERE togarak_id=%s AND user_id=%s AND aktiv=TRUE",
        (sorov.togarak_id, sorov.user_id),
    )
    if not cur.fetchone():
        cur.close()
        conn.close()
        raise HTTPException(status_code=400, detail="Bu o'quvchi shu to'garak a'zosi emas")

    if not (0 <= sorov.baho <= 100):
        cur.close()
        conn.close()
        raise HTTPException(status_code=400, detail="Baho 0-100 oralig'ida bo'lishi kerak")

    cur.execute(
        """INSERT INTO togarak_baholar(togarak_id, user_id, baho, izoh, teacher_id)
           VALUES(%s,%s,%s,%s,%s)""",
        (sorov.togarak_id, sorov.user_id, sorov.baho, sorov.izoh, teacher_id),
    )
    conn.commit()
    cur.close()
    conn.close()
    return {"holat": "saqlandi"}


# ═══════════════════════════════════════════════════════════
# OTA-ONA ↔ FARZAND — botdagi ota_ona.py bilan AYNAN BIR XIL jadval
# (farzand_kod, parent_child) — shu sabab botda yaratilgan kodni
# saytda kiritish ham, aksincha ham ishlaydi.
# ═══════════════════════════════════════════════════════════

FARZAND_KOD_MUDDATI = 15  # daqiqa


def _ota_ona_jadvallari(cur):
    cur.execute("""CREATE TABLE IF NOT EXISTS farzand_kod(
        kod TEXT PRIMARY KEY, child_id BIGINT NOT NULL, muddat TIMESTAMP NOT NULL
    )""")
    cur.execute("""CREATE TABLE IF NOT EXISTS parent_child(
        id SERIAL PRIMARY KEY, parent_id BIGINT NOT NULL, child_id BIGINT NOT NULL
    )""")


@app.post("/api/farzand/kod_yarat")
def farzand_kod_yarat(token: str):
    """O'quvchi (farzand) ota-onasini ulash uchun 6 xonali kod oladi —
    botdagi bilan bir xil jadvalga yoziladi, 15 daqiqa amal qiladi."""
    child_id = _jwt_tekshir(token)
    conn = _db()
    cur = conn.cursor()
    _ota_ona_jadvallari(cur)
    cur.execute("DELETE FROM farzand_kod WHERE child_id=%s OR muddat < NOW()", (child_id,))
    kod = None
    for _ in range(10):
        taklif = "".join(secrets.choice(string.digits) for _ in range(6))
        cur.execute("SELECT 1 FROM farzand_kod WHERE kod=%s", (taklif,))
        if not cur.fetchone():
            kod = taklif
            break
    if not kod:
        cur.close(); conn.close()
        raise HTTPException(status_code=500, detail="Kod yaratib bo'lmadi, qayta urinib ko'ring")
    cur.execute(
        "INSERT INTO farzand_kod(kod, child_id, muddat) VALUES(%s,%s,%s)",
        (kod, child_id, datetime.now() + timedelta(minutes=FARZAND_KOD_MUDDATI)),
    )
    conn.commit()
    cur.close()
    conn.close()
    return {"kod": kod, "amal_qilish_daqiqasi": FARZAND_KOD_MUDDATI}


@app.post("/api/ota/farzand_boglash")
def ota_farzand_boglash(token: str, kod: str):
    """Ota-ona farzanddan olgan 6 xonali kodni kiritib, hisobni bog'laydi."""
    parent_id = _jwt_tekshir(token)
    conn = _db()
    cur = conn.cursor()
    _ota_ona_jadvallari(cur)
    cur.execute("DELETE FROM farzand_kod WHERE muddat < NOW()")
    cur.execute("SELECT child_id FROM farzand_kod WHERE kod=%s", (kod.strip(),))
    r = cur.fetchone()
    if not r:
        cur.close(); conn.close()
        raise HTTPException(status_code=400, detail="Kod noto'g'ri yoki muddati o'tgan")
    child_id = r["child_id"]
    if child_id == parent_id:
        cur.close(); conn.close()
        raise HTTPException(status_code=400, detail="O'zingizni ulay olmaysiz")

    cur.execute(
        "INSERT INTO parent_child(parent_id, child_id) VALUES(%s,%s) ON CONFLICT DO NOTHING RETURNING id",
        (parent_id, child_id),
    )
    yangi_boglanish = cur.fetchone() is not None
    cur.execute("DELETE FROM farzand_kod WHERE kod=%s", (kod.strip(),))
    cur.execute("SELECT full_name FROM users WHERE user_id=%s", (child_id,))
    ism_row = cur.fetchone()
    conn.commit()
    cur.close()
    conn.close()
    return {
        "holat": "ulandi" if yangi_boglanish else "allaqachon_ulangan",
        "farzand_ismi": ism_row["full_name"] if ism_row else "",
    }


@app.delete("/api/ota/farzand_uzish")
def ota_farzand_uzish(token: str, farzand_id: int):
    """Ota-ona farzand bilan bog'lanishni uzadi."""
    parent_id = _jwt_tekshir(token)
    conn = _db()
    cur = conn.cursor()
    cur.execute("DELETE FROM parent_child WHERE parent_id=%s AND child_id=%s", (parent_id, farzand_id))
    ochirildi = cur.rowcount > 0
    conn.commit()
    cur.close()
    conn.close()
    if not ochirildi:
        raise HTTPException(status_code=404, detail="Bunday bog'lanish topilmadi")
    return {"holat": "uzildi"}


# ═══════════════════════════════════════════════════════════
# PROFIL — tahrirlash va rol almashtirish
# ═══════════════════════════════════════════════════════════

class ProfilYangilash(BaseModel):
    token: str
    full_name: Optional[str] = None
    region: Optional[str] = None
    district: Optional[str] = None
    tugilgan_sana: Optional[str] = None
    maktab_raqami: Optional[str] = None
    maktab_turi: Optional[str] = None   # oddiy | xususiy | ixtisoslashgan | prezident
    sinf: Optional[str] = None          # 1..11
    sinf_harfi: Optional[str] = None    # A, B, V ...
    jins: Optional[str] = None          # ogil | qiz — dizayn uchun (o'quvchi va o'qituvchi)
    oqituvchi_fani: Optional[str] = None  # o'qituvchining o'zi o'qitadigan fan — dizayn uchun
    maktab_id: Optional[int] = None     # ro'yxatdagi (tizimga qo'shilgan) maktabga ANIQ bog'lanish


MAKTAB_TURLARI = {
    "oddiy": "🏫 Oddiy davlat maktabi",
    "xususiy": "🏢 Xususiy",
    "ixtisoslashgan": "⭐ Ixtisoslashgan (IDUM)",
    "prezident": "🏆 Prezident maktabi",
}


@app.put("/api/profil")
def profil_yangila(sorov: ProfilYangilash):
    """Foydalanuvchi o'z profilini yangilaydi."""
    user_id = _jwt_tekshir(sorov.token)
    if sorov.full_name is not None and not sorov.full_name.strip():
        raise HTTPException(status_code=400, detail="Ism bo'sh bo'lishi mumkin emas")
    if sorov.maktab_turi is not None and sorov.maktab_turi not in MAKTAB_TURLARI:
        raise HTTPException(status_code=400, detail="Noto'g'ri maktab turi")
    if sorov.sinf is not None and sorov.sinf not in [str(i) for i in range(1, 12)]:
        raise HTTPException(status_code=400, detail="Sinf 1 dan 11 gacha bo'lishi kerak")
    if sorov.jins is not None and sorov.jins not in ("ogil", "qiz"):
        raise HTTPException(status_code=400, detail="Noto'g'ri jins qiymati")

    conn = _db()
    cur = conn.cursor()
    cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS tugilgan_sana DATE")
    cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS maktab_raqami TEXT")
    cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS jins TEXT")
    cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS oqituvchi_fani TEXT")

    maydonlar = []
    qiymatlar = []
    if sorov.full_name is not None:
        maydonlar.append("full_name=%s")
        qiymatlar.append(sorov.full_name.strip())
    if sorov.region is not None:
        maydonlar.append("region=%s")
        qiymatlar.append(sorov.region.strip())
    if sorov.district is not None:
        maydonlar.append("district=%s")
        qiymatlar.append(sorov.district.strip())
    if sorov.tugilgan_sana is not None:
        maydonlar.append("tugilgan_sana=%s")
        qiymatlar.append(sorov.tugilgan_sana)
    if sorov.maktab_raqami is not None:
        maydonlar.append("maktab_raqami=%s")
        qiymatlar.append(sorov.maktab_raqami.strip())
    if sorov.maktab_turi is not None:
        maydonlar.append("school_type=%s")
        qiymatlar.append(MAKTAB_TURLARI[sorov.maktab_turi])
    if sorov.sinf is not None:
        maydonlar.append("class=%s")
        qiymatlar.append(sorov.sinf)
    if sorov.sinf_harfi is not None:
        maydonlar.append("class_letter=%s")
        qiymatlar.append(sorov.sinf_harfi.strip().upper())
    if sorov.jins is not None:
        maydonlar.append("jins=%s")
        qiymatlar.append(sorov.jins)
    if sorov.oqituvchi_fani is not None:
        maydonlar.append("oqituvchi_fani=%s")
        qiymatlar.append(sorov.oqituvchi_fani.strip())
    if sorov.maktab_id is not None:
        _maktab_jadvali(cur)
        cur.execute("SELECT 1 FROM maktablar WHERE id=%s", (sorov.maktab_id,))
        if not cur.fetchone():
            cur.close(); conn.close()
            raise HTTPException(status_code=400, detail="Ko'rsatilgan maktab topilmadi")
        cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS maktab_id INTEGER")
        maydonlar.append("maktab_id=%s")
        qiymatlar.append(sorov.maktab_id)

    if not maydonlar:
        cur.close()
        conn.close()
        return {"holat": "ozgarish_yoq"}

    qiymatlar.append(user_id)
    cur.execute(f"UPDATE users SET {', '.join(maydonlar)} WHERE user_id=%s", qiymatlar)
    conn.commit()
    cur.close()
    conn.close()
    return {"holat": "saqlandi"}


class RolOzgartirish(BaseModel):
    token: str
    yangi_rol: str
    tasdiqlayman: bool = False


RUXSAT_ETILGAN_ROLLAR2 = {"oquvchi", "ota-ona", "oqituvchi"}
ROL_BEPUL_LIMIT = 2          # necha marta ERKIN (kod so'ramasdan) rol almashtirish mumkin
ROL_KOD_AMAL_MUDDATI = 10    # daqiqa
ROL_OYLIK_LIMIT_KUN = 30     # kod bilan almashtirilgach, keyingisi uchun necha kun kutish kerak


def _rol_ustunlarini_tayyorla(cur):
    cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS rol_ozgarish_soni INTEGER DEFAULT 0")
    cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS oxirgi_rol_ozgarish TIMESTAMP")
    cur.execute("""CREATE TABLE IF NOT EXISTS rol_tasdiq_kod(
        user_id BIGINT PRIMARY KEY REFERENCES users(user_id),
        kod TEXT NOT NULL, yangi_rol TEXT NOT NULL, yaratilgan_at TIMESTAMP DEFAULT NOW()
    )""")


def _email_yubor(qabul_qiluvchi: str, mavzu: str, matn: str) -> bool:
    """SMTP orqali email yuboradi. SMTP_HOST/SMTP_USER/SMTP_PASSWORD Railway'da
    o'rnatilgan bo'lishi kerak (masalan Gmail App Password) — aks holda False
    qaytaradi va konsolga log yozadi."""
    host = os.getenv("SMTP_HOST", "smtp.gmail.com")
    port = int(os.getenv("SMTP_PORT", "587"))
    user = os.getenv("SMTP_USER")
    parol = os.getenv("SMTP_PASSWORD")
    if not user or not parol:
        print(f"[EMAIL YUBORILMADI — SMTP sozlanmagan] {qabul_qiluvchi}: {matn}")
        return False
    try:
        import smtplib
        from email.mime.text import MIMEText
        msg = MIMEText(matn, "plain", "utf-8")
        msg["Subject"] = mavzu
        msg["From"] = user
        msg["To"] = qabul_qiluvchi
        with smtplib.SMTP(host, port, timeout=10) as s:
            s.starttls()
            s.login(user, parol)
            s.sendmail(user, [qabul_qiluvchi], msg.as_string())
        return True
    except Exception as e:
        print(f"[EMAIL XATO] {e}")
        return False


@app.put("/api/rol_ozgartir")
def rol_ozgartir(sorov: RolOzgartirish):
    """Foydalanuvchi rolini o'zgartiradi.
    - Admin uchun — CHEKLOVSIZ (sinab ko'rish uchun).
    - Oddiy foydalanuvchi uchun — hayotda 2 marta ERKIN (faqat tasdiq bilan),
      3-martadan boshlab Gmail'ga yuborilgan kod bilan, va kod bilan
      almashtirilgach keyingisi uchun 30 kun kutish kerak."""
    user_id = _jwt_tekshir(sorov.token)
    if sorov.yangi_rol not in RUXSAT_ETILGAN_ROLLAR2:
        raise HTTPException(status_code=400, detail=f"Noto'g'ri rol: {sorov.yangi_rol}")

    conn = _db()
    cur = conn.cursor()
    _rol_ustunlarini_tayyorla(cur)

    cur.execute("SELECT role, rol_ozgarish_soni, oxirgi_rol_ozgarish FROM users WHERE user_id=%s", (user_id,))
    r = cur.fetchone()
    if not r:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Foydalanuvchi topilmadi")

    cur.execute("SELECT 1 FROM admin_akkaunt WHERE uid=%s", (user_id,))
    admin_mi = cur.fetchone() is not None

    hozirgi_rol = r["role"]
    if hozirgi_rol == sorov.yangi_rol:
        cur.close(); conn.close()
        return {"holat": "ozgarish_yoq"}

    soni = r["rol_ozgarish_soni"] or 0

    # ADMIN — cheklovsiz, sinab ko'rish uchun
    if admin_mi:
        if not sorov.tasdiqlayman:
            cur.close(); conn.close()
            return {"holat": "tasdiq_kerak", "hozirgi_rol": hozirgi_rol, "yangi_rol": sorov.yangi_rol, "admin_test": True}
        cur.execute("UPDATE users SET role=%s WHERE user_id=%s", (sorov.yangi_rol, user_id))
        conn.commit(); cur.close(); conn.close()
        return {"holat": "saqlandi", "yangi_rol": sorov.yangi_rol}

    # ODDIY FOYDALANUVCHI — hali bepul limitdan foydalanmagan
    if soni < ROL_BEPUL_LIMIT:
        if not sorov.tasdiqlayman:
            cur.close(); conn.close()
            return {
                "holat": "tasdiq_kerak", "hozirgi_rol": hozirgi_rol, "yangi_rol": sorov.yangi_rol,
                "qolgan_bepul": ROL_BEPUL_LIMIT - soni,
            }
        cur.execute(
            "UPDATE users SET role=%s, rol_ozgarish_soni=rol_ozgarish_soni+1, oxirgi_rol_ozgarish=NOW() WHERE user_id=%s",
            (sorov.yangi_rol, user_id),
        )
        conn.commit(); cur.close(); conn.close()
        return {"holat": "saqlandi", "yangi_rol": sorov.yangi_rol, "qolgan_bepul": ROL_BEPUL_LIMIT - soni - 1}

    # BEPUL LIMIT TUGAGAN — 30 kunlik muddat tekshiriladi
    if r["oxirgi_rol_ozgarish"]:
        keyingi = r["oxirgi_rol_ozgarish"] + timedelta(days=ROL_OYLIK_LIMIT_KUN)
        if datetime.now() < keyingi:
            cur.close(); conn.close()
            raise HTTPException(
                status_code=429,
                detail=f"Rol almashtirish limiti tugagan. Keyingi imkoniyat: {keyingi.strftime('%d.%m.%Y')}",
            )

    cur.close(); conn.close()
    return {"holat": "kod_kerak", "hozirgi_rol": hozirgi_rol, "yangi_rol": sorov.yangi_rol}


class RolKodSorash(BaseModel):
    token: str
    yangi_rol: str


@app.post("/api/rol_kod_yubor")
def rol_kod_yubor(sorov: RolKodSorash):
    """Bepul limit tugagan foydalanuvchi uchun — Gmail'ga tasdiqlash kodi yuboradi."""
    user_id = _jwt_tekshir(sorov.token)
    if sorov.yangi_rol not in RUXSAT_ETILGAN_ROLLAR2:
        raise HTTPException(status_code=400, detail="Noto'g'ri rol")

    conn = _db()
    cur = conn.cursor()
    _rol_ustunlarini_tayyorla(cur)
    cur.execute("SELECT google_email FROM google_hisob WHERE user_id=%s LIMIT 1", (user_id,))
    r = cur.fetchone()
    if not r or not r["google_email"]:
        cur.close(); conn.close()
        raise HTTPException(status_code=400, detail="Gmail hisobingiz ulanmagan — avval botdagi kabinet orqali ulang")

    email = r["google_email"]
    kod = "".join(secrets.choice(string.digits) for _ in range(6))
    cur.execute("""
        INSERT INTO rol_tasdiq_kod(user_id, kod, yangi_rol, yaratilgan_at)
        VALUES(%s,%s,%s,NOW())
        ON CONFLICT (user_id) DO UPDATE SET kod=EXCLUDED.kod, yangi_rol=EXCLUDED.yangi_rol, yaratilgan_at=NOW()
    """, (user_id, kod, sorov.yangi_rol))
    conn.commit()
    cur.close(); conn.close()

    yuborildi = _email_yubor(
        email, "SamTM Ta'lim — rol o'zgartirish kodi",
        f"Rolni \"{sorov.yangi_rol}\"ga o'zgartirish uchun tasdiqlash kodi: {kod}\n"
        f"Kod {ROL_KOD_AMAL_MUDDATI} daqiqa amal qiladi. Agar bu so'rovni siz yubormagan bo'lsangiz, e'tiborsiz qoldiring.",
    )
    yashirilgan = re.sub(r"(?<=.{2}).(?=[^@]*@)", "*", email)
    return {"holat": "yuborildi" if yuborildi else "smtp_sozlanmagan", "email": yashirilgan}


class RolKodTasdiqlash(BaseModel):
    token: str
    kod: str


@app.post("/api/rol_kod_tasdiqla")
def rol_kod_tasdiqla(sorov: RolKodTasdiqlash):
    """Yuborilgan kodni tekshiradi va to'g'ri bo'lsa rolni o'zgartiradi."""
    user_id = _jwt_tekshir(sorov.token)
    conn = _db()
    cur = conn.cursor()
    _rol_ustunlarini_tayyorla(cur)
    cur.execute("SELECT kod, yangi_rol, yaratilgan_at FROM rol_tasdiq_kod WHERE user_id=%s", (user_id,))
    r = cur.fetchone()
    if not r:
        cur.close(); conn.close()
        raise HTTPException(status_code=400, detail="Avval kod so'rang")
    if datetime.now() - r["yaratilgan_at"] > timedelta(minutes=ROL_KOD_AMAL_MUDDATI):
        cur.close(); conn.close()
        raise HTTPException(status_code=400, detail="Kod muddati tugagan — qaytadan so'rang")
    if sorov.kod.strip() != r["kod"]:
        cur.close(); conn.close()
        raise HTTPException(status_code=400, detail="Kod noto'g'ri")

    cur.execute(
        "UPDATE users SET role=%s, rol_ozgarish_soni=rol_ozgarish_soni+1, oxirgi_rol_ozgarish=NOW() WHERE user_id=%s",
        (r["yangi_rol"], user_id),
    )
    cur.execute("DELETE FROM rol_tasdiq_kod WHERE user_id=%s", (user_id,))
    conn.commit()
    cur.close(); conn.close()
    return {"holat": "saqlandi", "yangi_rol": r["yangi_rol"]}


# ═══════════════════════════════════════════════════════════
# O'QITUVCHI — yangi to'garak yaratish
# ═══════════════════════════════════════════════════════════

class TogarakYaratish(BaseModel):
    token: str
    nomi: str
    fan: str
    sinf: Optional[str] = None   # "1".."11" (oddiy) yoki "3-4" kabi (to'garak guruhi)
    parol: Optional[str] = None
    max_talaba: Optional[int] = None
    oylik_summa: Optional[int] = None
    universitet_guruh_id: Optional[int] = None  # professor shu fanini ANIQ universitet guruhi uchun o'qitsa


@app.post("/api/oqituvchi/togarak_yarat")
def togarak_yarat(sorov: TogarakYaratish):
    """O'qituvchi yangi to'garak yaratadi — bot ishlatadigan AYNAN SHU
    jadvalga (togaraklar) yoziladi, shuning uchun bot va sayt bir xil
    ma'lumotni ko'radi. Fan+sinf tanlanganda — o'sha fan/sinfga tegishli
    BARCHA mavzular avtomatik ravishda to'garakning "ta'lim yo'li"ga
    bog'lanadi (togarak_mavzulari)."""
    teacher_id = _jwt_tekshir(sorov.token)
    if not sorov.nomi.strip():
        raise HTTPException(status_code=400, detail="To'garak nomi kiritilmagan")
    if not sorov.fan.strip():
        raise HTTPException(status_code=400, detail="Fan kiritilmagan")
    if sorov.max_talaba is not None and sorov.max_talaba < 1:
        raise HTTPException(status_code=400, detail="Maksimal talaba soni kamida 1 bo'lishi kerak")

    conn = _db()
    cur = conn.cursor()
    cur.execute("ALTER TABLE togaraklar ADD COLUMN IF NOT EXISTS sinf TEXT")
    cur.execute("ALTER TABLE togaraklar ADD COLUMN IF NOT EXISTS markaz_id INTEGER")
    cur.execute("ALTER TABLE togaraklar ADD COLUMN IF NOT EXISTS universitet_guruh_id INTEGER")
    cur.execute("""CREATE TABLE IF NOT EXISTS togarak_mavzulari(
        togarak_id INTEGER REFERENCES togaraklar(id),
        topic_code TEXT,
        PRIMARY KEY (togarak_id, topic_code)
    )""")
    # Agar o'qituvchi biror o'quv markaziga tegishli bo'lsa (xodim
    # importi orqali "Fan o'qituvchisi" sifatida qo'shilgan bo'lsa) —
    # yaratayotgan guruhi AVTOMATIK shu markazga bog'lanadi, markaz
    # direktori/administratori uni darhol "Markaz" boshqaruv panelida
    # ko'radi — qo'lda bog'lash shart emas.
    cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS markaz_id INTEGER")
    cur.execute("SELECT markaz_id FROM users WHERE user_id=%s", (teacher_id,))
    ur = cur.fetchone()
    teacher_markaz_id = ur["markaz_id"] if ur else None

    # Professor bu fanni ANIQ bitta universitet guruhi uchun o'qitayotgan
    # bo'lsa — shu guruhga bog'laydi, guruh kuratori/dekani keyin BUTUN
    # guruhning shu fandagi bilim darajasini ko'ra oladi.
    universitet_guruh_id = None
    if sorov.universitet_guruh_id is not None:
        cur.execute("SELECT id FROM universitet_guruhlari WHERE id=%s", (sorov.universitet_guruh_id,))
        if cur.fetchone():
            universitet_guruh_id = sorov.universitet_guruh_id

    sinf_qiymati = sorov.sinf.strip() if sorov.sinf else None
    cur.execute("""
        INSERT INTO togaraklar(nomi, fan, teacher_id, sinf, parol, max_talaba, oylik_summa, aktiv, markaz_id, universitet_guruh_id)
        VALUES(%s,%s,%s,%s,%s,%s,%s,TRUE,%s,%s) RETURNING id
    """, (sorov.nomi.strip(), sorov.fan.strip(), teacher_id, sinf_qiymati,
          sorov.parol.strip() if sorov.parol else None,
          sorov.max_talaba, sorov.oylik_summa, teacher_markaz_id, universitet_guruh_id))
    yangi_id = cur.fetchone()["id"]

    bogliq_mavzu_soni = 0
    if sinf_qiymati:
        cur.execute("""
            SELECT topic_code FROM dts_tree
            WHERE grade=%s AND subject_name=%s AND is_deleted=FALSE
              AND topic_code IN (SELECT DISTINCT topic_code FROM generated_tests)
        """, (sinf_qiymati, sorov.fan.strip()))
        mavzu_kodlari = [r["topic_code"] for r in cur.fetchall()]
        for kod in mavzu_kodlari:
            cur.execute(
                "INSERT INTO togarak_mavzulari(togarak_id, topic_code) VALUES(%s,%s) ON CONFLICT DO NOTHING",
                (yangi_id, kod),
            )
        bogliq_mavzu_soni = len(mavzu_kodlari)

    conn.commit()
    cur.close()
    conn.close()
    return {"holat": "yaratildi", "togarak_id": yangi_id, "boglangan_mavzu_soni": bogliq_mavzu_soni}


# ═══════════════════════════════════════════════════════════
# TA'LIM YO'LI — o'quvchining fan bo'yicha ketma-ket mavzular ustidan
# qanday bosib o'tayotgani (ota-ona/o'quvchi/o'qituvchi ko'radi)
# ═══════════════════════════════════════════════════════════

def _chorak_taqsimoti(mavzular: list) -> list:
    """Mavzular ro'yxatini (har birida "chorak" maydoni bor) 1/2/3/4-chorak
    bo'yicha guruhlab, har chorakning necha foizi bosib o'tilganini
    hisoblaydi — chorak ma'lumoti yo'q mavzular hisobga olinmaydi."""
    guruhlar = {}
    for m in mavzular:
        ch = (m.get("chorak") or "").strip()
        if not ch:
            continue
        guruhlar.setdefault(ch, {"jami": 0, "otilgan": 0})
        guruhlar[ch]["jami"] += 1
        if m["score"] is not None:
            guruhlar[ch]["otilgan"] += 1
    natija = []
    for ch in sorted(guruhlar.keys(), key=lambda x: (len(x), x)):
        g = guruhlar[ch]
        natija.append({
            "chorak": ch, "jami_mavzu": g["jami"], "otilgan_mavzu": g["otilgan"],
            "foiz": round((g["otilgan"] / g["jami"]) * 100) if g["jami"] else 0,
        })
    return natija


# ═══════════════════════════════════════════════════════════
# ESDAN CHIQISH XAVFI + BUGUNGI TAVSIYA — sof matematik formula,
# HECH QANDAY AI ishlatilmaydi. Faqat mavjud learned_topics
# ma'lumotidan (ball, oxirgi o'rganilgan sana, necha marta
# takrorlangan) hisoblanadi — shu sabab BEPUL va har doim ANIQ
# (bir xil kirish — doim bir xil natija).
#
# Mantiq: xotira "barqarorligi" har muvaffaqiyatli takrorda oshadi
# (spaced-repetition tamoyili), past ball bilan o'rganilgan mavzu esa
# tezroq "unutiladi" deb hisoblanadi.
# ═══════════════════════════════════════════════════════════

_ESDAN_CHIQISH_ASOSIY_INTERVAL = 10   # kun — birinchi marta o'rgangandan keyin "e'tibor zonasi"
_ESDAN_CHIQISH_OSISH_KOEF = 2.3       # har takrorda xotira necha barobar "mustahkamlanadi"


def _esdan_chiqish_foizi(ortacha_ball: float, kunlar_otgan: int, takror_soni: float) -> int:
    """0-100 oralig'ida "unutish ehtimoli" — AI emas, sof formula."""
    if kunlar_otgan <= 0:
        return 0
    barqarorlik = (
        _ESDAN_CHIQISH_ASOSIY_INTERVAL
        * (_ESDAN_CHIQISH_OSISH_KOEF ** max(0, (takror_soni or 1) - 1))
        * max(0.5, (ortacha_ball or 0) / 100)
    )
    foiz = 100 * (1 - math.exp(-kunlar_otgan / barqarorlik))
    return round(foiz)


def _xavf_darajasi(foiz: int) -> str:
    if foiz >= 60:
        return "yuqori"
    if foiz >= 30:
        return "orta"
    return "past"


@app.get("/api/bola/{bola_id}/bugungi_tavsiya")
def bugungi_tavsiya(bola_id: int, limit: int = 8):
    """O'quvchining O'Z SINFI bo'yicha, avval o'rgangan mavzularini
    "unutish xavfi"ga qarab SARALAB, bugun eng birinchi takrorlash
    kerak bo'lganlarini qaytaradi. To'liq AI'siz, sof hisob-kitob."""
    conn = _db()
    cur = conn.cursor()
    cur.execute("SELECT class FROM users WHERE user_id=%s", (bola_id,))
    u = cur.fetchone()
    if not u or not u["class"]:
        cur.close(); conn.close()
        return {"tavsiyalar": [], "sinf_sozlanmagan": True}
    sinf = str(u["class"]).replace("-sinf", "").strip()

    cur.execute("""
        SELECT COALESCE(d.mavzu_name, d.bolim_name, d.bob_name) AS nomi, d.subject_name AS fan,
               MAX(lt.learned_at) AS oxirgi_sana,
               AVG(lt.score) AS ortacha_ball,
               AVG(lt.repeat_count) AS ortacha_takror
        FROM dts_tree d
        JOIN learned_topics lt ON lt.topic_code = d.topic_code AND lt.user_id = %s
        WHERE d.grade = %s AND d.is_deleted = FALSE
        GROUP BY COALESCE(d.mavzu_name, d.bolim_name, d.bob_name), d.subject_name
    """, (bola_id, sinf))
    qatorlar = cur.fetchall()
    cur.close()
    conn.close()

    bugun = datetime.now()
    tavsiyalar = []
    for r in qatorlar:
        kunlar_otgan = (bugun - r["oxirgi_sana"]).days if r["oxirgi_sana"] else 0
        foiz = _esdan_chiqish_foizi(r["ortacha_ball"], kunlar_otgan, r["ortacha_takror"])
        tavsiyalar.append({
            "nomi": r["nomi"], "fan": r["fan"], "kunlar_otgan": kunlar_otgan,
            "oxirgi_ball": round(r["ortacha_ball"]) if r["ortacha_ball"] is not None else None,
            "esdan_chiqish_foizi": foiz, "daraja": _xavf_darajasi(foiz),
        })

    tavsiyalar.sort(key=lambda t: t["esdan_chiqish_foizi"], reverse=True)
    # Faqat haqiqatan e'tiborga loyiq (past emas) darajadagilarni ko'rsatamiz —
    # "past" xavfli mavzularni bugun takrorlashga majburlash shart emas.
    ehtiyoj_borlari = [t for t in tavsiyalar if t["daraja"] != "past"]
    return {"tavsiyalar": ehtiyoj_borlari[:limit], "sinf_sozlanmagan": False}


@app.get("/api/bola/{bola_id}/haftalik_xulosa")
def haftalik_xulosa(bola_id: int):
    """O'quvchi uchun oxirgi 7 kunlik xulosa — QAYSI mavzular ishlangan,
    o'rtacha ball, nechta YANGI mavzu o'rgangan, qaysilari qiyinlik
    qilgan, va nechta kun KETMA-KET mashq qilingan (streak). To'liq
    mavjud learned_topics ma'lumotidan hisoblanadi — AI shart emas."""
    conn = _db()
    cur = conn.cursor()

    cur.execute("""
        SELECT COALESCE(d.mavzu_name, d.bolim_name, d.bob_name) AS nomi, d.subject_name AS fan,
               MAX(lt.score) AS ball, MAX(lt.repeat_count) AS takror_soni
        FROM learned_topics lt
        JOIN dts_tree d ON d.topic_code = lt.topic_code
        WHERE lt.user_id = %s AND lt.learned_at >= NOW() - INTERVAL '7 days'
        GROUP BY COALESCE(d.mavzu_name, d.bolim_name, d.bob_name), d.subject_name
    """, (bola_id,))
    hafta_qatorlari = cur.fetchall()

    # Streak (ketma-ket kunlar) — BUTUN tarixdan, faqat shu haftadan emas,
    # chunki "necha kundan beri uzluksiz mashq qilyapsiz" savoli haftadan
    # oshib ketishi mumkin.
    cur.execute("""
        SELECT DISTINCT learned_at::date AS kun FROM learned_topics
        WHERE user_id=%s ORDER BY kun DESC
    """, (bola_id,))
    kunlar = [r["kun"] for r in cur.fetchall()]
    cur.close()
    conn.close()

    ketma_ket = 0
    if kunlar:
        bugun = datetime.now().date()
        # Bugun hali mashq qilinmagan bo'lsa ham, kechadan boshlab hisoblaymiz —
        # aks holda kun tugamasdan streak "0" ko'rinib, foydalanuvchini
        # asossiz xafa qilmasin.
        joriy_kun = bugun if bugun in kunlar else bugun - timedelta(days=1)
        while joriy_kun in kunlar:
            ketma_ket += 1
            joriy_kun -= timedelta(days=1)

    jami_mavzu = len(hafta_qatorlari)
    ortacha_ball = round(sum(r["ball"] for r in hafta_qatorlari) / jami_mavzu) if jami_mavzu else 0
    yangi_mavzular = [r["nomi"] for r in hafta_qatorlari if (r["takror_soni"] or 1) == 1]
    zaif_mavzular = sorted(
        [{"nomi": r["nomi"], "ball": r["ball"]} for r in hafta_qatorlari if r["ball"] is not None and r["ball"] < 60],
        key=lambda x: x["ball"],
    )[:5]

    # Fanlar bo'yicha o'rtacha — eng yaxshi natijali fanni topish uchun
    fanlar_hisobi = {}
    for r in hafta_qatorlari:
        fanlar_hisobi.setdefault(r["fan"], []).append(r["ball"] or 0)
    eng_yaxshi_fan = None
    if fanlar_hisobi:
        eng_yaxshi_fan = max(fanlar_hisobi, key=lambda f: sum(fanlar_hisobi[f]) / len(fanlar_hisobi[f]))

    return {
        "jami_mavzu": jami_mavzu, "ortacha_ball": ortacha_ball,
        "yangi_mavzular_soni": len(yangi_mavzular), "yangi_mavzular": yangi_mavzular[:5],
        "zaif_mavzular": zaif_mavzular, "eng_yaxshi_fan": eng_yaxshi_fan,
        "ketma_ket_kun": ketma_ket,
    }


@app.get("/api/bola/{bola_id}/yol")
def talim_yoli_oddiy(bola_id: int, fan: str):
    """Oddiy (majburiy) o'quv dasturi bo'yicha — o'quvchining O'Z SINFI
    (avtomatik aniqlanadi) va berilgan fan uchun BARCHA mavzularni
    ketma-ket, har biriga o'quvchining natijasi (score, agar hali
    yechmagan bo'lsa — yo'q) bilan qaytaradi."""
    conn = _db()
    cur = conn.cursor()
    cur.execute("SELECT class FROM users WHERE user_id=%s", (bola_id,))
    u = cur.fetchone()
    if not u or not u["class"]:
        cur.close(); conn.close()
        raise HTTPException(status_code=400, detail="O'quvchining sinfi aniqlanmagan")
    sinf = str(u["class"]).replace("-sinf", "").strip()

    # MUHIM: bitta "mavzu" ostida bir nechta "kichik mavzu" bo'lishi mumkin
    # (har biri o'z topic_code'iga ega) — lekin o'quvchiga YO'L sifatida
    # kichik mavzular emas, faqat MAVZU darajasi ko'rsatilishi kerak.
    # Shu sabab MAVZU nomi bo'yicha guruhlaymiz: bir nechta kichik mavzu —
    # bitta yo'l bandi, ballari o'rtacha olinadi. Chorak (quarter) ham shu
    # yerda olinadi — pastda chorak bo'yicha taqsimot hisoblanadi.
    cur.execute("""
        SELECT COALESCE(d.mavzu_name, d.bolim_name, d.bob_name) AS nomi,
               MIN(d.topic_code) AS topic_code,
               MIN(d.quarter) AS chorak,
               COUNT(*) AS jami_kichik,
               COUNT(lt.score) AS otilgan_kichik,
               AVG(lt.score) AS ortacha_ball
        FROM dts_tree d
        LEFT JOIN learned_topics lt ON lt.topic_code = d.topic_code AND lt.user_id = %s
        WHERE d.grade = %s AND d.subject_name = %s AND d.is_deleted = FALSE
          AND d.topic_code IN (SELECT DISTINCT topic_code FROM generated_tests)
        GROUP BY COALESCE(d.mavzu_name, d.bolim_name, d.bob_name)
        ORDER BY MIN(d.topic_code)
    """, (bola_id, sinf, fan))
    xom_qatorlar = cur.fetchall()
    cur.close()
    conn.close()

    mavzular = [{
        "topic_code": r["topic_code"], "nomi": r["nomi"], "chorak": r["chorak"],
        "score": round(r["ortacha_ball"]) if r["otilgan_kichik"] > 0 else None,
        "otilgan_kichik": r["otilgan_kichik"], "jami_kichik": r["jami_kichik"],
    } for r in xom_qatorlar]
    choraklar = _chorak_taqsimoti(mavzular)

    otilgan = sum(1 for m in mavzular if m["score"] is not None)
    jami = len(mavzular)
    ortacha = round(sum(m["score"] for m in mavzular if m["score"] is not None) / otilgan) if otilgan else 0
    return {
        "sinf": sinf, "jami_mavzu": jami, "otilgan_mavzu": otilgan,
        "yol_foizi": round((otilgan / jami) * 100) if jami else 0,
        "samaradorlik_foizi": ortacha,
        "mavzular": mavzular, "choraklar": choraklar,
    }


@app.get("/api/bola/{bola_id}/togarak_yoli/{togarak_id}")
def talim_yoli_togarak(bola_id: int, togarak_id: int):
    """To'garakning O'ZIGA XOS ta'lim yo'li — faqat shu to'garakka
    biriktirilgan mavzular (togarak_mavzulari) bo'yicha. Bu — o'quvchi
    to'garakka QO'SHILGANDAGINA ko'rinadigan qo'shimcha statistika,
    oddiy sinf statistikasiga ARALASHMAYDI."""
    conn = _db()
    cur = conn.cursor()
    cur.execute("SELECT nomi, fan, sinf FROM togaraklar WHERE id=%s", (togarak_id,))
    tg = cur.fetchone()
    if not tg:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="To'garak topilmadi")

    cur.execute("""
        SELECT COALESCE(d.mavzu_name, d.bolim_name, d.bob_name) AS nomi,
               MIN(d.topic_code) AS topic_code,
               MIN(d.quarter) AS chorak,
               COUNT(*) AS jami_kichik,
               COUNT(lt.score) AS otilgan_kichik,
               AVG(lt.score) AS ortacha_ball
        FROM togarak_mavzulari tm
        JOIN dts_tree d ON d.topic_code = tm.topic_code
        LEFT JOIN learned_topics lt ON lt.topic_code = d.topic_code AND lt.user_id = %s
        WHERE tm.togarak_id = %s
        GROUP BY COALESCE(d.mavzu_name, d.bolim_name, d.bob_name)
        ORDER BY MIN(d.topic_code)
    """, (bola_id, togarak_id))
    xom_qatorlar = cur.fetchall()
    cur.close()
    conn.close()

    mavzular = [{
        "topic_code": r["topic_code"], "nomi": r["nomi"], "chorak": r["chorak"],
        "score": round(r["ortacha_ball"]) if r["otilgan_kichik"] > 0 else None,
        "otilgan_kichik": r["otilgan_kichik"], "jami_kichik": r["jami_kichik"],
    } for r in xom_qatorlar]
    choraklar = _chorak_taqsimoti(mavzular)

    otilgan = sum(1 for m in mavzular if m["score"] is not None)
    jami = len(mavzular)
    ortacha = round(sum(m["score"] for m in mavzular if m["score"] is not None) / otilgan) if otilgan else 0
    return {
        "togarak_nomi": tg["nomi"], "fan": tg["fan"], "sinf": tg["sinf"],
        "jami_mavzu": jami, "otilgan_mavzu": otilgan,
        "yol_foizi": round((otilgan / jami) * 100) if jami else 0,
        "samaradorlik_foizi": ortacha,
        "mavzular": mavzular, "choraklar": choraklar,
    }


@app.get("/api/bola/{bola_id}/togaraklarim")
def bolaning_togaraklari(bola_id: int):
    """O'quvchi a'zo bo'lgan barcha faol to'garaklar ro'yxati — 'ta'lim
    yo'li' ekranida to'garak yo'lini alohida ko'rsatish uchun."""
    conn = _db()
    cur = conn.cursor()
    cur.execute("""
        SELECT t.id, t.nomi, t.fan, t.sinf
        FROM togarak_azolar ta
        JOIN togaraklar t ON t.id = ta.togarak_id
        WHERE ta.user_id = %s AND ta.aktiv = TRUE AND t.aktiv = TRUE
        ORDER BY t.nomi
    """, (bola_id,))
    natija = cur.fetchall()
    cur.close()
    conn.close()
    return {"togaraklar": natija}


# ═══════════════════════════════════════════════════════════
# TO'GARAKKA QO'SHILISH (parol orqali — barcha rollar uchun)
# ═══════════════════════════════════════════════════════════

class TogarakqaQoshilish(BaseModel):
    token: str
    parol: str


@app.post("/api/togarakka_qoshil")
def togarakka_qoshil(sorov: TogarakqaQoshilish):
    """Foydalanuvchi (o'quvchi, ota-ona va h.k.) parol orqali to'garakka
    qo'shiladi — bot orqali qo'shilgan bilan BIR XIL jadvalga yoziladi."""
    user_id = _jwt_tekshir(sorov.token)
    if not sorov.parol.strip():
        raise HTTPException(status_code=400, detail="Parol kiritilmagan")

    conn = _db()
    cur = conn.cursor()
    cur.execute("SELECT id, nomi, max_talaba FROM togaraklar WHERE parol=%s AND aktiv=TRUE", (sorov.parol.strip(),))
    t = cur.fetchone()
    if not t:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Bunday parolli to'garak topilmadi")

    cur.execute(
        "SELECT 1 FROM togarak_azolar WHERE togarak_id=%s AND user_id=%s AND aktiv=TRUE",
        (t["id"], user_id),
    )
    if cur.fetchone():
        cur.close(); conn.close()
        raise HTTPException(status_code=400, detail="Siz allaqachon shu to'garak a'zosisiz")

    if t["max_talaba"]:
        cur.execute("SELECT COUNT(*) AS soni FROM togarak_azolar WHERE togarak_id=%s AND aktiv=TRUE", (t["id"],))
        joriy = cur.fetchone()["soni"]
        if joriy >= t["max_talaba"]:
            cur.close(); conn.close()
            raise HTTPException(status_code=400, detail="To'garak to'lgan")

    cur.execute("INSERT INTO togarak_azolar(togarak_id, user_id, aktiv) VALUES(%s,%s,TRUE)", (t["id"], user_id))
    conn.commit()
    cur.close()
    conn.close()
    return {"holat": "qoshildi", "togarak_nomi": t["nomi"]}


@app.get("/api/mening_togaraklarim")
def mening_togaraklarim(token: str):
    """Foydalanuvchi a'zo bo'lgan barcha to'garaklarni qaytaradi."""
    user_id = _jwt_tekshir(token)
    conn = _db()
    cur = conn.cursor()
    cur.execute("""
        SELECT tg.id, tg.nomi, tg.fan
        FROM togarak_azolar ta
        JOIN togaraklar tg ON tg.id = ta.togarak_id
        WHERE ta.user_id=%s AND ta.aktiv=TRUE
    """, (user_id,))
    natija = cur.fetchall()
    cur.close()
    conn.close()
    return {"togaraklar": natija}


# ═══════════════════════════════════════════════════════════
# ADMIN — Test shablon (Excel) yuklab olish va import qilish
# Botdagi _generate_template / import_tests_excel mantig'iga mos
# ═══════════════════════════════════════════════════════════

def _admin_tekshir(token: str):
    user_id = _jwt_tekshir(token)
    conn = _db()
    cur = conn.cursor()
    cur.execute("SELECT 1 FROM admin_akkaunt WHERE uid=%s", (user_id,))
    natija = cur.fetchone()
    cur.close()
    conn.close()
    if not natija:
        raise HTTPException(status_code=403, detail="Faqat admin uchun")
    return user_id


# ═══════════════════════════════════════════════════════════
# MAKTAB TIZIMI — 1-BOSQICH: maktab yaratish
# (2-bosqich: xodimlarni Excel orqali kiritish, 3-bosqich: sinflar,
#  4-bosqich: o'quvchi qo'shilishi, 5-bosqich: sinf tahlili — keyinroq)
# ═══════════════════════════════════════════════════════════

def _maktab_jadvali(cur):
    cur.execute("""CREATE TABLE IF NOT EXISTS maktablar(
        id SERIAL PRIMARY KEY,
        nomi TEXT NOT NULL,
        viloyat TEXT, tuman TEXT,
        smena_soni INTEGER NOT NULL DEFAULT 1,
        direktor_user_id BIGINT REFERENCES users(user_id),
        yaratilgan_at TIMESTAMP DEFAULT NOW()
    )""")


class MaktabYaratish(BaseModel):
    token: str
    nomi: str
    viloyat: Optional[str] = None
    tuman: Optional[str] = None
    smena_soni: int = 1
    direktor_user_id: Optional[int] = None


@app.post("/api/admin/maktab_yarat")
def maktab_yarat(sorov: MaktabYaratish):
    """1-bosqich: yangi maktabni tizimga qo'shadi. Direktor keyinroq ham
    (xodimlar Excel orqali import qilinganda) belgilanishi mumkin —
    shu sabab bu yerda ixtiyoriy."""
    _admin_tekshir(sorov.token)
    if not sorov.nomi.strip():
        raise HTTPException(status_code=400, detail="Maktab nomi kiritilmagan")
    if sorov.smena_soni not in (1, 2):
        raise HTTPException(status_code=400, detail="Smena soni 1 yoki 2 bo'lishi kerak")

    conn = _db()
    cur = conn.cursor()
    _maktab_jadvali(cur)
    if sorov.direktor_user_id is not None:
        cur.execute("SELECT 1 FROM users WHERE user_id=%s", (sorov.direktor_user_id,))
        if not cur.fetchone():
            cur.close(); conn.close()
            raise HTTPException(status_code=400, detail="Ko'rsatilgan direktor foydalanuvchisi topilmadi")
    cur.execute("""
        INSERT INTO maktablar(nomi, viloyat, tuman, smena_soni, direktor_user_id)
        VALUES(%s,%s,%s,%s,%s) RETURNING id
    """, (sorov.nomi.strip(), sorov.viloyat, sorov.tuman, sorov.smena_soni, sorov.direktor_user_id))
    yangi_id = cur.fetchone()["id"]
    conn.commit()
    cur.close()
    conn.close()
    return {"holat": "yaratildi", "maktab_id": yangi_id}


@app.get("/api/admin/maktablar")
def maktablar_royxati(token: str):
    """Barcha maktablar ro'yxati — direktor ismi bilan birga."""
    _admin_tekshir(token)
    conn = _db()
    cur = conn.cursor()
    _maktab_jadvali(cur)
    cur.execute("""
        SELECT m.id, m.nomi, m.viloyat, m.tuman, m.smena_soni, m.direktor_user_id,
               u.full_name AS direktor_ismi
        FROM maktablar m
        LEFT JOIN users u ON u.user_id = m.direktor_user_id
        ORDER BY m.nomi
    """)
    natija = cur.fetchall()
    cur.close()
    conn.close()
    return {"maktablar": natija}


@app.put("/api/admin/maktab_direktor")
def maktab_direktor_belgila(token: str, maktab_id: int, direktor_user_id: int):
    """Mavjud maktabga direktorni keyinroq belgilash/almashtirish uchun."""
    _admin_tekshir(token)
    conn = _db()
    cur = conn.cursor()
    _maktab_jadvali(cur)
    cur.execute("SELECT 1 FROM users WHERE user_id=%s", (direktor_user_id,))
    if not cur.fetchone():
        cur.close(); conn.close()
        raise HTTPException(status_code=400, detail="Foydalanuvchi topilmadi")
    cur.execute("UPDATE maktablar SET direktor_user_id=%s WHERE id=%s", (direktor_user_id, maktab_id))
    ozgardi = cur.rowcount > 0
    conn.commit()
    cur.close()
    conn.close()
    if not ozgardi:
        raise HTTPException(status_code=404, detail="Maktab topilmadi")
    return {"holat": "saqlandi"}


@app.get("/api/admin/foydalanuvchi_qidir")
def admin_foydalanuvchi_qidir(token: str, ism: str):
    """Admin uchun — ism bo'yicha foydalanuvchi qidiradi (masalan
    direktor sifatida tayinlash uchun kerakli odamni topish)."""
    _admin_tekshir(token)
    if len(ism.strip()) < 2:
        return {"natijalar": []}
    conn = _db()
    cur = conn.cursor()
    cur.execute("""
        SELECT user_id, full_name, role FROM users
        WHERE full_name ILIKE %s
        ORDER BY full_name LIMIT 10
    """, (f"%{ism.strip()}%",))
    natija = cur.fetchall()
    cur.close()
    conn.close()
    return {"natijalar": natija}


# ═══════════════════════════════════════════════════════════
# MAKTAB TIZIMI — 2-BOSQICH: xodimlarni Excel orqali kiritish
# Har bir xodim uchun avtomatik KIRISH KODI (mavjud veb_ulash_kod
# mexanizmiga o'xshash, lekin uzoqroq — 30 kun — amal qiladigan)
# yaratiladi. Agar "Sinf rahbarligi" to'ldirilgan bo'lsa — o'sha
# sinf (maktab_sinflari) ham shu bilan birga yaratiladi/yangilanadi,
# 4 xonali qo'shilish paroli bilan.
# ═══════════════════════════════════════════════════════════

LAVOZIMLAR = {
    "direktor": "Direktor",
    "zam_direktor_uquv": "O'quv ishlari bo'yicha direktor o'rinbosari",
    "zam_direktor_tarbiya": "Ma'naviy-ma'rifiy ishlar bo'yicha direktor o'rinbosari",
    "psixolog": "Psixolog",
    "kotib": "Kotib",
    "fan_oqituvchisi": "Fan o'qituvchisi",
}
_LAVOZIM_MATNDAN = {v.lower(): k for k, v in LAVOZIMLAR.items()}


def _xodim_kod_jadvali(cur):
    cur.execute("""CREATE TABLE IF NOT EXISTS xodim_kod(
        kod TEXT PRIMARY KEY, user_id BIGINT NOT NULL REFERENCES users(user_id),
        yaratildi TIMESTAMP DEFAULT NOW(), ishlatildi BOOLEAN DEFAULT FALSE
    )""")


def _maktab_sinflari_jadvali(cur):
    cur.execute("""CREATE TABLE IF NOT EXISTS maktab_sinflari(
        id SERIAL PRIMARY KEY,
        maktab_id INTEGER NOT NULL REFERENCES maktablar(id),
        sinf TEXT NOT NULL, harf TEXT NOT NULL,
        rahbar_user_id BIGINT REFERENCES users(user_id),
        qoshilish_paroli TEXT,
        yaratilgan_at TIMESTAMP DEFAULT NOW(),
        UNIQUE(maktab_id, sinf, harf)
    )""")


@app.get("/api/admin/xodim_shablon")
def xodim_shablon(token: str):
    """Xodimlarni import qilish uchun Excel shablonini beradi —
    F.I.Sh, Lavozim, Sinf rahbarligi ustunlari bilan."""
    _admin_tekshir(token)
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    import io
    from fastapi.responses import StreamingResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "XODIMLAR"
    ustunlar = ["F.I.Sh", "Lavozim", "Sinf rahbarligi (ixtiyoriy)"]
    for col, h in enumerate(ustunlar, 1):
        c = ws.cell(1, col, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1B4B7A")

    namunalar = [
        ("Aliyev Vali Aliyevich", "Direktor", ""),
        ("Karimova Nilufar Rustamovna", "O'quv ishlari bo'yicha direktor o'rinbosari", ""),
        ("Yusupov Sardor Bahtiyorovich", "Fan o'qituvchisi", "5-A"),
        ("Nazarova Feruza Odilovna", "Fan o'qituvchisi", ""),
    ]
    for r in namunalar:
        ws.append(r)
    for col, w in zip("ABC", [30, 45, 25]):
        ws.column_dimensions[col].width = w

    ws2 = wb.create_sheet("IZOH")
    ws2.cell(1, 1, "Lavozim ustuniga faqat shu variantlardan birini yozing:").font = Font(bold=True)
    for i, nom in enumerate(LAVOZIMLAR.values(), 2):
        ws2.cell(i, 1, f"• {nom}")
    ws2.cell(len(LAVOZIMLAR) + 3, 1,
             "Sinf rahbarligi — faqat shu odam biror sinfga rahbar bo'lsa to'ldiring (masalan: 5-A). Bo'sh qoldirsa ham bo'ladi.")
    ws2.column_dimensions["A"].width = 70

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=xodimlar_shablon.xlsx"},
    )


@app.post("/api/admin/xodim_import")
async def xodim_import(token: str, maktab_id: int, fayl: UploadFile = File(...)):
    """To'ldirilgan xodimlar shablonini import qiladi — har biriga
    hisob va 30 kun amal qiladigan KIRISH KODI yaratadi. "Sinf
    rahbarligi" to'ldirilgan bo'lsa, o'sha sinfni ham (yangi 4 xonali
    qo'shilish paroli bilan) yaratadi/yangilaydi."""
    _admin_tekshir(token)
    import openpyxl
    import io

    conn = _db()
    cur = conn.cursor()
    _maktab_jadvali(cur)
    cur.execute("SELECT id FROM maktablar WHERE id=%s", (maktab_id,))
    if not cur.fetchone():
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Maktab topilmadi")

    content = await fayl.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        cur.close(); conn.close()
        raise HTTPException(status_code=400, detail=f"Excel o'qib bo'lmadi: {e}")
    ws = wb["XODIMLAR"] if "XODIMLAR" in wb.sheetnames else wb.active

    _xodim_kod_jadvali(cur)
    _maktab_sinflari_jadvali(cur)
    cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS maktab_id INTEGER")
    cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS lavozim TEXT")

    natijalar = []
    xato_soni = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0] or not str(row[0]).strip():
            continue
        fish = str(row[0]).strip()
        lavozim_matni = str(row[1]).strip() if len(row) > 1 and row[1] else "Fan o'qituvchisi"
        sinf_rahbarligi = str(row[2]).strip() if len(row) > 2 and row[2] else ""
        lavozim_kaliti = _LAVOZIM_MATNDAN.get(lavozim_matni.lower(), "fan_oqituvchisi")

        try:
            cur.execute("SELECT MIN(user_id) AS eng_kichik FROM users WHERE user_id < 0")
            r = cur.fetchone()
            yangi_id = (r["eng_kichik"] - 1) if r and r["eng_kichik"] is not None else -1

            cur.execute("""
                INSERT INTO users(user_id, full_name, role, maktab_id, lavozim)
                VALUES(%s,%s,'oqituvchi',%s,%s)
            """, (yangi_id, fish, maktab_id, lavozim_kaliti))

            if lavozim_kaliti == "direktor":
                cur.execute("UPDATE maktablar SET direktor_user_id=%s WHERE id=%s", (yangi_id, maktab_id))

            kirish_kodi = "".join(secrets.choice(string.ascii_uppercase + string.digits) for _ in range(6))
            cur.execute(
                "INSERT INTO xodim_kod(kod, user_id) VALUES(%s,%s)",
                (kirish_kodi, yangi_id),
            )

            sinf_paroli = None
            if sinf_rahbarligi and "-" in sinf_rahbarligi:
                sinf_qismi, harf_qismi = sinf_rahbarligi.split("-", 1)
                sinf_paroli = "".join(secrets.choice(string.digits) for _ in range(4))
                cur.execute("""
                    INSERT INTO maktab_sinflari(maktab_id, sinf, harf, rahbar_user_id, qoshilish_paroli)
                    VALUES(%s,%s,%s,%s,%s)
                    ON CONFLICT (maktab_id, sinf, harf) DO UPDATE SET
                        rahbar_user_id = EXCLUDED.rahbar_user_id,
                        qoshilish_paroli = EXCLUDED.qoshilish_paroli
                """, (maktab_id, sinf_qismi.strip(), harf_qismi.strip().upper(), yangi_id, sinf_paroli))

            conn.commit()
            natijalar.append({
                "fish": fish, "lavozim": LAVOZIMLAR.get(lavozim_kaliti, lavozim_matni),
                "kirish_kodi": kirish_kodi, "sinf_rahbarligi": sinf_rahbarligi or None,
                "sinf_paroli": sinf_paroli,
            })
        except Exception:
            conn.rollback()
            xato_soni += 1

    cur.close()
    conn.close()
    return {"natijalar": natijalar, "xato_soni": xato_soni}


# ═══════════════════════════════════════════════════════════
# MAKTAB TIZIMI — 3-BOSQICH: sinflarni ko'rish/boshqarish
# (Excel import paytida "sinf rahbarligi" orqali avtomatik yaratilgan
#  sinflarni ko'rish, qo'lda yangi sinf qo'shish, parolni qayta tashlash)
# ═══════════════════════════════════════════════════════════

@app.get("/api/admin/maktab_sinflari")
def maktab_sinflari_royxati(token: str, maktab_id: int):
    _admin_tekshir(token)
    conn = _db()
    cur = conn.cursor()
    _maktab_sinflari_jadvali(cur)
    cur.execute("""
        SELECT s.id, s.sinf, s.harf, s.qoshilish_paroli, u.full_name AS rahbar_ismi
        FROM maktab_sinflari s
        LEFT JOIN users u ON u.user_id = s.rahbar_user_id
        WHERE s.maktab_id=%s
        ORDER BY s.sinf::int, s.harf
    """, (maktab_id,))
    natija = cur.fetchall()
    cur.close()
    conn.close()
    return {"sinflar": natija}


class SinfYaratish(BaseModel):
    token: str
    maktab_id: int
    sinf: str
    harf: str
    rahbar_user_id: Optional[int] = None


@app.post("/api/admin/maktab_sinf_yarat")
def maktab_sinf_yarat(sorov: SinfYaratish):
    """Qo'lda, Excel'siz ham bitta sinf qo'shish imkoni — masalan
    keyinroq yangi sinf ochilganda qayta Excel import shart emas."""
    _admin_tekshir(sorov.token)
    conn = _db()
    cur = conn.cursor()
    _maktab_sinflari_jadvali(cur)
    paroli = "".join(secrets.choice(string.digits) for _ in range(4))
    cur.execute("""
        INSERT INTO maktab_sinflari(maktab_id, sinf, harf, rahbar_user_id, qoshilish_paroli)
        VALUES(%s,%s,%s,%s,%s)
        ON CONFLICT (maktab_id, sinf, harf) DO UPDATE SET
            rahbar_user_id = COALESCE(EXCLUDED.rahbar_user_id, maktab_sinflari.rahbar_user_id)
        RETURNING id, qoshilish_paroli
    """, (sorov.maktab_id, sorov.sinf.strip(), sorov.harf.strip().upper(), sorov.rahbar_user_id, paroli))
    natija = cur.fetchone()
    conn.commit()
    cur.close()
    conn.close()
    return {"holat": "yaratildi", "sinf_id": natija["id"], "qoshilish_paroli": natija["qoshilish_paroli"]}


@app.put("/api/admin/sinf_parolini_tashla")
def sinf_parolini_tashla(token: str, sinf_id: int):
    """Sinf rahbari yoki admin — qo'shilish parolini qayta yaratadi
    (masalan parol tarqalib ketgan bo'lsa)."""
    _admin_tekshir(token)
    conn = _db()
    cur = conn.cursor()
    yangi_parol = "".join(secrets.choice(string.digits) for _ in range(4))
    cur.execute("UPDATE maktab_sinflari SET qoshilish_paroli=%s WHERE id=%s", (yangi_parol, sinf_id))
    ozgardi = cur.rowcount > 0
    conn.commit()
    cur.close()
    conn.close()
    if not ozgardi:
        raise HTTPException(status_code=404, detail="Sinf topilmadi")
    return {"holat": "yangilandi", "qoshilish_paroli": yangi_parol}


# ═══════════════════════════════════════════════════════════
# PULLI MAKTAB — TO'LOV TIZIMI
# Bu — HISOB-KITOB/YOZUV tizimi (naqd yoki tashqi o'tkazma orqali
# to'langan to'lovni QAYD ETISH), TO'LOV SHLYUZI (Payme/Click orqali
# ONLAYN to'lov qabul qilish) EMAS — bu ALOHIDA, ancha katta ish, agar
# kerak bo'lsa keyinroq alohida gaplashamiz.
#
# Bildirishnoma — botga emas, saytdagi "Xabarlar" bo'limiga (hozircha
# "Tez orada" bo'sh turgan joy) yoziladi — ota-ona kirganda ko'radi.
# ═══════════════════════════════════════════════════════════

def _tolov_jadvallari(cur):
    cur.execute("ALTER TABLE maktablar ADD COLUMN IF NOT EXISTS pulli BOOLEAN DEFAULT FALSE")
    cur.execute("ALTER TABLE maktablar ADD COLUMN IF NOT EXISTS oylik_tolov INTEGER")
    cur.execute("""CREATE TABLE IF NOT EXISTS tolovlar(
        id SERIAL PRIMARY KEY,
        user_id BIGINT NOT NULL REFERENCES users(user_id),
        maktab_id INTEGER REFERENCES maktablar(id),
        oy TEXT NOT NULL,
        summa_kerak INTEGER NOT NULL,
        tolangan_summa INTEGER NOT NULL DEFAULT 0,
        tolov_sanasi DATE,
        qayd_etildi_at TIMESTAMP DEFAULT NOW(),
        UNIQUE(user_id, maktab_id, oy)
    )""")
    # MUHIM: markaz/to'garak to'lovlari uchun ham ishlatiladi — maktab_id
    # bo'sh (NULL) qoldirilib, togarak_id to'ldiriladi. NULL != NULL
    # bo'lgani uchun standart SQL semantikasida bu ikkala UNIQUE cheklov
    # bir-biriga XALAQIT bermaydi (mustaqil ishlaydi).
    cur.execute("ALTER TABLE tolovlar ADD COLUMN IF NOT EXISTS togarak_id INTEGER REFERENCES togaraklar(id)")
    cur.execute("ALTER TABLE tolovlar ALTER COLUMN maktab_id DROP NOT NULL")
    cur.execute("CREATE UNIQUE INDEX IF NOT EXISTS tolovlar_togarak_unique ON tolovlar(user_id, togarak_id, oy)")
    cur.execute("""CREATE TABLE IF NOT EXISTS bildirishnomalar(
        id SERIAL PRIMARY KEY,
        user_id BIGINT NOT NULL REFERENCES users(user_id),
        matn TEXT NOT NULL,
        turi TEXT DEFAULT 'umumiy',
        oqildimi BOOLEAN DEFAULT FALSE,
        yaratildi TIMESTAMP DEFAULT NOW()
    )""")


class MaktabTolovSozlash(BaseModel):
    token: str
    maktab_id: int
    pulli: bool
    oylik_tolov: Optional[int] = None


@app.put("/api/admin/maktab_tolov_sozlash")
def maktab_tolov_sozlash(sorov: MaktabTolovSozlash):
    _admin_tekshir(sorov.token)
    if sorov.pulli and not sorov.oylik_tolov:
        raise HTTPException(status_code=400, detail="Pulli maktab uchun oylik to'lov summasini kiriting")
    conn = _db()
    cur = conn.cursor()
    _tolov_jadvallari(cur)
    cur.execute(
        "UPDATE maktablar SET pulli=%s, oylik_tolov=%s WHERE id=%s",
        (sorov.pulli, sorov.oylik_tolov if sorov.pulli else None, sorov.maktab_id),
    )
    conn.commit()
    cur.close()
    conn.close()
    return {"holat": "saqlandi"}


@app.get("/api/oqituvchi/mening_sinflarim")
def mening_rasmiy_sinflarim(token: str):
    """O'qituvchi RAHBAR bo'lgan rasmiy maktab sinflari — to'garakdan
    farqli, bu yerga o'quvchi PAROL bilan TASDIQLAB qo'shiladi (4-bosqich),
    shu sabab o'quvchi soni ham faqat TASDIQLANGAN a'zolarni hisoblaydi."""
    user_id = _jwt_tekshir(token)
    conn = _db()
    cur = conn.cursor()
    _maktab_sinflari_jadvali(cur)
    _tolov_jadvallari(cur)
    _sinf_azolari_jadvali(cur)
    cur.execute("""
        SELECT s.id, s.sinf, s.harf, s.qoshilish_paroli, m.id AS maktab_id, m.nomi AS maktab_nomi, m.pulli, m.oylik_tolov,
               (SELECT COUNT(*) FROM maktab_sinf_azolari a WHERE a.sinf_id=s.id) AS oquvchi_soni
        FROM maktab_sinflari s
        JOIN maktablar m ON m.id = s.maktab_id
        WHERE s.rahbar_user_id=%s
        ORDER BY s.sinf::int, s.harf
    """, (user_id,))
    natija = cur.fetchall()
    cur.close()
    conn.close()
    return {"sinflar": natija}



def sinf_tolovlari(token: str, sinf_id: int, oy: str):
    """Sinf rahbari (yoki admin) uchun — shu oy uchun sinfga TASDIQLAB
    qo'shilgan (4-bosqich) har bir o'quvchining to'lov holatini
    ko'rsatadi. `oy` format: "2026-07"."""
    user_id = _jwt_tekshir(token)
    conn = _db()
    cur = conn.cursor()
    _tolov_jadvallari(cur)
    _sinf_azolari_jadvali(cur)
    cur.execute("SELECT maktab_id, sinf, harf, rahbar_user_id FROM maktab_sinflari WHERE id=%s", (sinf_id,))
    s = cur.fetchone()
    if not s:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Sinf topilmadi")
    ruxsat = s["rahbar_user_id"] == user_id or _maktab_boshqaruvchi_mi(cur, user_id, s["maktab_id"])
    if not ruxsat:
        cur.close(); conn.close()
        raise HTTPException(status_code=403, detail="Faqat shu sinf rahbari, maktab rahbariyati yoki admin ko'ra oladi")

    cur.execute("SELECT oylik_tolov FROM maktablar WHERE id=%s", (s["maktab_id"],))
    maktab = cur.fetchone()
    kerakli_summa = (maktab["oylik_tolov"] if maktab else None) or 0

    cur.execute("""
        SELECT u.user_id, u.full_name FROM maktab_sinf_azolari a
        JOIN users u ON u.user_id = a.user_id
        WHERE a.sinf_id=%s
        ORDER BY u.full_name
    """, (sinf_id,))
    oquvchilar = cur.fetchall()

    cur.execute("SELECT user_id, tolangan_summa, tolov_sanasi FROM tolovlar WHERE maktab_id=%s AND oy=%s", (s["maktab_id"], oy))
    tolovlar_map = {r["user_id"]: r for r in cur.fetchall()}
    cur.close()
    conn.close()

    natija = []
    for o in oquvchilar:
        t = tolovlar_map.get(o["user_id"])
        tolangan = t["tolangan_summa"] if t else 0
        natija.append({
            "user_id": o["user_id"], "full_name": o["full_name"],
            "kerakli_summa": kerakli_summa, "tolangan_summa": tolangan,
            "qarzdor": tolangan < kerakli_summa,
            "tolov_sanasi": t["tolov_sanasi"].isoformat() if t and t["tolov_sanasi"] else None,
        })
    return {"oquvchilar": natija, "kerakli_summa": kerakli_summa}


class TolovBelgilash(BaseModel):
    token: str
    user_id: int
    maktab_id: int
    oy: str
    tolangan_summa: int


@app.post("/api/oqituvchi/tolov_belgila")
def tolov_belgila(sorov: TolovBelgilash):
    """Sinf rahbari — o'quvchi naqd/o'tkazma orqali to'laganda, shu
    yerda QAYD ETADI. To'landi deb belgilangach, ota-onaga saytdagi
    Xabarlar bo'limiga bildirishnoma yoziladi."""
    murojaatchi_id = _jwt_tekshir(sorov.token)
    conn = _db()
    cur = conn.cursor()
    _tolov_jadvallari(cur)
    _sinf_azolari_jadvali(cur)

    # XAVFSIZLIK: faqat admin, shu maktab rahbariyati (direktor/
    # o'rinbosarlar), yoki aynan shu o'quvchi a'zo bo'lgan sinfning
    # rahbari to'lov belgilashi mumkin — boshqa hech kim emas.
    cur.execute("SELECT 1 FROM admin_akkaunt WHERE uid=%s", (murojaatchi_id,))
    ruxsat = cur.fetchone() is not None
    if not ruxsat:
        cur.execute(
            "SELECT 1 FROM users WHERE user_id=%s AND maktab_id=%s AND lavozim IN ('direktor','zam_direktor_uquv','zam_direktor_tarbiya')",
            (murojaatchi_id, sorov.maktab_id),
        )
        ruxsat = cur.fetchone() is not None
    if not ruxsat:
        cur.execute("""
            SELECT 1 FROM maktab_sinf_azolari a
            JOIN maktab_sinflari s ON s.id = a.sinf_id
            WHERE a.user_id=%s AND s.maktab_id=%s AND s.rahbar_user_id=%s
        """, (sorov.user_id, sorov.maktab_id, murojaatchi_id))
        ruxsat = cur.fetchone() is not None
    if not ruxsat:
        cur.close(); conn.close()
        raise HTTPException(status_code=403, detail="Faqat shu o'quvchining sinf rahbari, maktab rahbariyati yoki admin to'lov belgilay oladi")

    cur.execute("""
        INSERT INTO tolovlar(user_id, maktab_id, oy, summa_kerak, tolangan_summa, tolov_sanasi)
        VALUES(%s,%s,%s,%s,%s, CURRENT_DATE)
        ON CONFLICT (user_id, maktab_id, oy) DO UPDATE SET
            tolangan_summa = EXCLUDED.tolangan_summa, tolov_sanasi = CURRENT_DATE
        RETURNING summa_kerak
    """, (sorov.user_id, sorov.maktab_id, sorov.oy, sorov.tolangan_summa, sorov.tolangan_summa))
    conn.commit()

    # Ota-onaga bildirishnoma — parent_child jadvali orqali (mavjud,
    # ota-ona↔farzand bog'lash uchun ishlatiladigan) shu o'quvchining
    # ota-onasini topamiz.
    cur.execute("SELECT full_name FROM users WHERE user_id=%s", (sorov.user_id,))
    oquvchi = cur.fetchone()
    cur.execute("SELECT parent_id FROM parent_child WHERE child_id=%s", (sorov.user_id,))
    ota_onalar = cur.fetchall()
    for oo in ota_onalar:
        cur.execute(
            "INSERT INTO bildirishnomalar(user_id, matn, turi) VALUES(%s,%s,'tolov')",
            (oo["parent_id"], f"{oquvchi['full_name']} uchun {sorov.oy} oyi to'lovi qabul qilindi: {sorov.tolangan_summa:,} so'm".replace(",", " ")),
        )
    conn.commit()
    cur.close()
    conn.close()
    return {"holat": "saqlandi"}


@app.get("/api/bildirishnomalar")
def bildirishnomalarim(token: str):
    user_id = _jwt_tekshir(token)
    conn = _db()
    cur = conn.cursor()
    _tolov_jadvallari(cur)
    cur.execute("""
        SELECT id, matn, turi, oqildimi, yaratildi FROM bildirishnomalar
        WHERE user_id=%s ORDER BY yaratildi DESC LIMIT 50
    """, (user_id,))
    natija = cur.fetchall()
    cur.execute("UPDATE bildirishnomalar SET oqildimi=TRUE WHERE user_id=%s AND oqildimi=FALSE", (user_id,))
    conn.commit()
    cur.close()
    conn.close()
    return {"bildirishnomalar": natija}


# ═══════════════════════════════════════════════════════════
# MAKTAB TIZIMI — 4-BOSQICH: o'quvchi qo'shilishi
#
# MUHIM: o'quvchining ESKI "maktab_raqami" (erkin matn) MOSLASH uchun
# ISHONCHSIZ — shu sabab bu bosqich YANGI "maktab_id" (ro'yxatdagi
# maktabga aniq bog'lanish) ustiga quriladi. Eski matn maydoni ham
# qoladi (ro'yxatda yo'q maktablar uchun), lekin AVTOMATIK SINF
# TOPISH faqat maktab_id orqali ishlaydi — bu ANIQ, matn solishtirish
# EMAS.
#
# Qo'shilish — profil mosligi + 4 xonali parol (ikkalasi ham kerak),
# rahbar esa xato qo'shilgan a'zoni istalgan vaqt CHIQARIB yuborishi
# mumkin.
# ═══════════════════════════════════════════════════════════

def _sinf_azolari_jadvali(cur):
    cur.execute("""CREATE TABLE IF NOT EXISTS maktab_sinf_azolari(
        id SERIAL PRIMARY KEY,
        sinf_id INTEGER NOT NULL REFERENCES maktab_sinflari(id),
        user_id BIGINT NOT NULL REFERENCES users(user_id),
        qoshilgan_at TIMESTAMP DEFAULT NOW(),
        UNIQUE(sinf_id, user_id)
    )""")


@app.get("/api/maktab_qidir")
def maktab_qidir(nomi: str):
    """HAMMA uchun ochiq (admin shart emas) — o'quvchi profilida
    ro'yxatdagi maktabini nomi bo'yicha qidirib tanlashi uchun."""
    if len(nomi.strip()) < 2:
        return {"natijalar": []}
    conn = _db()
    cur = conn.cursor()
    _maktab_jadvali(cur)
    cur.execute("""
        SELECT id, nomi, viloyat, tuman FROM maktablar
        WHERE nomi ILIKE %s ORDER BY nomi LIMIT 10
    """, (f"%{nomi.strip()}%",))
    natija = cur.fetchall()
    cur.close()
    conn.close()
    return {"natijalar": natija}


@app.get("/api/oquvchi/mos_sinf")
def oquvchi_mos_sinf(token: str):
    """O'quvchining PROFILIDAGI (maktab_id + class + class_letter)
    ma'lumoti biror rasmiy sinfga MOS kelsa — va u hali A'ZO
    bo'lmagan bo'lsa — o'sha sinf haqida ma'lumot qaytaradi (parol
    BERMAYDI, faqat "shunday sinf bor" deb bildiradi)."""
    user_id = _jwt_tekshir(token)
    conn = _db()
    cur = conn.cursor()
    _maktab_sinflari_jadvali(cur)
    _sinf_azolari_jadvali(cur)
    cur.execute("SELECT maktab_id, class, class_letter FROM users WHERE user_id=%s", (user_id,))
    u = cur.fetchone()
    if not u or not u["maktab_id"] or not u["class"] or not u["class_letter"]:
        cur.close(); conn.close()
        return {"topildi": False}

    cur.execute("""
        SELECT s.id, s.sinf, s.harf, m.nomi AS maktab_nomi, ur.full_name AS rahbar_ismi
        FROM maktab_sinflari s
        JOIN maktablar m ON m.id = s.maktab_id
        LEFT JOIN users ur ON ur.user_id = s.rahbar_user_id
        WHERE s.maktab_id=%s AND s.sinf=%s AND s.harf=%s
    """, (u["maktab_id"], u["class"], u["class_letter"]))
    sinf = cur.fetchone()
    if not sinf:
        cur.close(); conn.close()
        return {"topildi": False}

    cur.execute("SELECT 1 FROM maktab_sinf_azolari WHERE sinf_id=%s AND user_id=%s", (sinf["id"], user_id))
    azo_mi = cur.fetchone() is not None
    cur.close()
    conn.close()
    if azo_mi:
        return {"topildi": False}  # allaqachon qo'shilgan — qayta so'ramaymiz
    return {
        "topildi": True, "sinf_id": sinf["id"], "sinf_nomi": f"{sinf['sinf']}-{sinf['harf']}",
        "maktab_nomi": sinf["maktab_nomi"], "rahbar_ismi": sinf["rahbar_ismi"],
    }


@app.post("/api/oquvchi/sinfga_qoshil")
def oquvchi_sinfga_qoshil(token: str, sinf_id: int, parol: str):
    """O'quvchi 4 xonali parolni kiritib, rasmiy sinfga QO'SHILADI
    (tasdiqlaydi). Profil mosligi YETARLI EMAS — parol ham to'g'ri
    bo'lishi kerak, shu orqali tasodifiy/xato qo'shilish oldi olinadi."""
    user_id = _jwt_tekshir(token)
    conn = _db()
    cur = conn.cursor()
    _sinf_azolari_jadvali(cur)
    cur.execute("SELECT qoshilish_paroli FROM maktab_sinflari WHERE id=%s", (sinf_id,))
    s = cur.fetchone()
    if not s:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Sinf topilmadi")
    if s["qoshilish_paroli"] != parol.strip():
        cur.close(); conn.close()
        raise HTTPException(status_code=400, detail="Parol noto'g'ri")
    cur.execute(
        "INSERT INTO maktab_sinf_azolari(sinf_id, user_id) VALUES(%s,%s) ON CONFLICT DO NOTHING",
        (sinf_id, user_id),
    )
    conn.commit()
    cur.close()
    conn.close()
    return {"holat": "qoshildi"}


@app.get("/api/oqituvchi/sinf_azolari")
def sinf_azolari_royxati(token: str, sinf_id: int):
    """Rahbar (yoki admin) — sinfga TASDIQLAB qo'shilgan o'quvchilar
    ro'yxati. Xato qo'shilganlarni shu yerdan chiqarish mumkin."""
    user_id = _jwt_tekshir(token)
    conn = _db()
    cur = conn.cursor()
    _sinf_azolari_jadvali(cur)
    cur.execute("SELECT maktab_id, rahbar_user_id FROM maktab_sinflari WHERE id=%s", (sinf_id,))
    s = cur.fetchone()
    if not s:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Sinf topilmadi")
    ruxsat = s["rahbar_user_id"] == user_id or _maktab_boshqaruvchi_mi(cur, user_id, s["maktab_id"])
    if not ruxsat:
        cur.close(); conn.close()
        raise HTTPException(status_code=403, detail="Faqat shu sinf rahbari, maktab rahbariyati yoki admin ko'ra oladi")
    cur.execute("""
        SELECT a.id AS azolik_id, u.user_id, u.full_name, a.qoshilgan_at
        FROM maktab_sinf_azolari a JOIN users u ON u.user_id = a.user_id
        WHERE a.sinf_id=%s ORDER BY u.full_name
    """, (sinf_id,))
    natija = cur.fetchall()
    cur.close()
    conn.close()
    return {"azolar": natija}


@app.delete("/api/oqituvchi/sinf_azosini_chiqar")
def sinf_azosini_chiqar(token: str, azolik_id: int):
    """Rahbar — xato qo'shilgan (masalan boshqa sinf o'quvchisi
    tasodifan mos kelib qolgan) a'zoni sinfdan chiqaradi."""
    user_id = _jwt_tekshir(token)
    conn = _db()
    cur = conn.cursor()
    _sinf_azolari_jadvali(cur)
    cur.execute("""
        SELECT s.maktab_id, s.rahbar_user_id FROM maktab_sinf_azolari a
        JOIN maktab_sinflari s ON s.id = a.sinf_id WHERE a.id=%s
    """, (azolik_id,))
    r = cur.fetchone()
    if not r:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="A'zolik topilmadi")
    ruxsat = r["rahbar_user_id"] == user_id or _maktab_boshqaruvchi_mi(cur, user_id, r["maktab_id"])
    if not ruxsat:
        cur.close(); conn.close()
        raise HTTPException(status_code=403, detail="Faqat shu sinf rahbari, maktab rahbariyati yoki admin chiqara oladi")
    cur.execute("DELETE FROM maktab_sinf_azolari WHERE id=%s", (azolik_id,))
    conn.commit()
    cur.close()
    conn.close()
    return {"holat": "chiqarildi"}


def _maktab_boshqaruvchi_mi(cur, user_id, maktab_id):
    """True — agar user shu maktabning direktori/o'rinbosari (yoki
    umumiy admin) bo'lsa."""
    cur.execute("SELECT 1 FROM admin_akkaunt WHERE uid=%s", (user_id,))
    if cur.fetchone():
        return True
    cur.execute("SELECT lavozim FROM users WHERE user_id=%s AND maktab_id=%s", (user_id, maktab_id))
    r = cur.fetchone()
    return bool(r and r["lavozim"] in ("direktor", "zam_direktor_uquv", "zam_direktor_tarbiya"))


@app.get("/api/maktab/dashboard")
def maktab_dashboard(token: str, maktab_id: int):
    """Direktor/o'rinbosarlar uchun — BUTUN maktab bitta ekranda:
    barcha sinflar, har birining o'quvchi soni va rahbari, va (agar
    maktab pulli bo'lsa) har bir sinf hamda umumiy maktab bo'yicha
    shu oy to'lov holati (nechtasi to'lagan, nechtasi qarzdor)."""
    user_id = _jwt_tekshir(token)
    conn = _db()
    cur = conn.cursor()
    _maktab_jadvali(cur); _maktab_sinflari_jadvali(cur); _sinf_azolari_jadvali(cur); _tolov_jadvallari(cur)
    if not _maktab_boshqaruvchi_mi(cur, user_id, maktab_id):
        cur.close(); conn.close()
        raise HTTPException(status_code=403, detail="Faqat maktab rahbariyati (direktor/o'rinbosar) yoki admin ko'ra oladi")

    cur.execute("SELECT nomi, pulli, oylik_tolov FROM maktablar WHERE id=%s", (maktab_id,))
    maktab = cur.fetchone()
    if not maktab:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Maktab topilmadi")

    joriy_oy = datetime.now().strftime("%Y-%m")
    oylik_tolov = maktab["oylik_tolov"] or 0
    cur.execute("""
        SELECT s.id, s.sinf, s.harf, u.full_name AS rahbar_ismi,
               COUNT(DISTINCT a.user_id) AS oquvchi_soni,
               COUNT(DISTINCT t.user_id) FILTER (WHERE t.tolangan_summa >= %s) AS tolagan_soni
        FROM maktab_sinflari s
        LEFT JOIN users u ON u.user_id = s.rahbar_user_id
        LEFT JOIN maktab_sinf_azolari a ON a.sinf_id = s.id
        LEFT JOIN tolovlar t ON t.user_id = a.user_id AND t.maktab_id = s.maktab_id AND t.oy = %s
        WHERE s.maktab_id=%s
        GROUP BY s.id, s.sinf, s.harf, u.full_name
        ORDER BY s.sinf::int, s.harf
    """, (oylik_tolov, joriy_oy, maktab_id))
    sinflar = cur.fetchall()
    cur.close()
    conn.close()

    jami_oquvchi = sum(s["oquvchi_soni"] for s in sinflar)
    jami_tolagan = sum(s["tolagan_soni"] for s in sinflar)
    return {
        "maktab_nomi": maktab["nomi"], "pulli": maktab["pulli"], "oylik_tolov": maktab["oylik_tolov"],
        "sinflar": sinflar,
        "tolov_xulosasi": (
            {"jami_oquvchi": jami_oquvchi, "tolagan": jami_tolagan, "qarzdor": jami_oquvchi - jami_tolagan}
            if maktab["pulli"] else None
        ),
    }


# ═══════════════════════════════════════════════════════════
# O'QUV MARKAZI TIZIMI — maktabdan farqli, MAVJUD to'garak (guruh)
# infratuzilmasi USTIGA quriladi (togaraklar, togarak_azolar) —
# takrorlanish emas, ANIQ QAYTA ISHLATISH: markazning har bir
# "guruhi" — oddiy to'garak, faqat endi markaz_id bilan bog'langan.
#
# Rollar: markaz_direktor, administrator (ikkalasi ham markazning
# BARCHA guruhlarini ko'radi/boshqaradi), fan_oqituvchisi (faqat
# o'z guruhlarini).
# ═══════════════════════════════════════════════════════════

MARKAZ_LAVOZIMLARI = {
    "markaz_direktor": "Markaz direktori",
    "administrator": "Administrator",
    "fan_oqituvchisi": "Fan o'qituvchisi",
}
_MARKAZ_LAVOZIM_MATNDAN = {v.lower(): k for k, v in MARKAZ_LAVOZIMLARI.items()}


def _markaz_jadvali(cur):
    cur.execute("""CREATE TABLE IF NOT EXISTS oquv_markazlari(
        id SERIAL PRIMARY KEY,
        nomi TEXT NOT NULL,
        viloyat TEXT, tuman TEXT,
        direktor_user_id BIGINT REFERENCES users(user_id),
        yaratilgan_at TIMESTAMP DEFAULT NOW()
    )""")
    cur.execute("ALTER TABLE togaraklar ADD COLUMN IF NOT EXISTS markaz_id INTEGER")
    cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS markaz_id INTEGER")


def _markaz_boshqaruvchi_mi(cur, user_id, markaz_id):
    """True — agar user shu markazning direktori/administratori
    (yoki umumiy admin) bo'lsa."""
    cur.execute("SELECT 1 FROM admin_akkaunt WHERE uid=%s", (user_id,))
    if cur.fetchone():
        return True
    cur.execute("SELECT lavozim FROM users WHERE user_id=%s AND markaz_id=%s", (user_id, markaz_id))
    r = cur.fetchone()
    return bool(r and r["lavozim"] in ("markaz_direktor", "administrator"))


class MarkazYaratish(BaseModel):
    token: str
    nomi: str
    viloyat: Optional[str] = None
    tuman: Optional[str] = None
    direktor_user_id: Optional[int] = None


@app.post("/api/admin/markaz_yarat")
def markaz_yarat(sorov: MarkazYaratish):
    _admin_tekshir(sorov.token)
    if not sorov.nomi.strip():
        raise HTTPException(status_code=400, detail="Markaz nomi kiritilmagan")
    conn = _db()
    cur = conn.cursor()
    _markaz_jadvali(cur)
    if sorov.direktor_user_id is not None:
        cur.execute("SELECT 1 FROM users WHERE user_id=%s", (sorov.direktor_user_id,))
        if not cur.fetchone():
            cur.close(); conn.close()
            raise HTTPException(status_code=400, detail="Ko'rsatilgan direktor foydalanuvchisi topilmadi")
    cur.execute("""
        INSERT INTO oquv_markazlari(nomi, viloyat, tuman, direktor_user_id)
        VALUES(%s,%s,%s,%s) RETURNING id
    """, (sorov.nomi.strip(), sorov.viloyat, sorov.tuman, sorov.direktor_user_id))
    yangi_id = cur.fetchone()["id"]
    conn.commit()
    cur.close()
    conn.close()
    return {"holat": "yaratildi", "markaz_id": yangi_id}


@app.get("/api/admin/markazlar")
def markazlar_royxati(token: str):
    _admin_tekshir(token)
    conn = _db()
    cur = conn.cursor()
    _markaz_jadvali(cur)
    cur.execute("""
        SELECT m.id, m.nomi, m.viloyat, m.tuman, m.direktor_user_id, u.full_name AS direktor_ismi
        FROM oquv_markazlari m
        LEFT JOIN users u ON u.user_id = m.direktor_user_id
        ORDER BY m.nomi
    """)
    natija = cur.fetchall()
    cur.close()
    conn.close()
    return {"markazlar": natija}


@app.get("/api/admin/markaz_xodim_shablon")
def markaz_xodim_shablon(token: str):
    _admin_tekshir(token)
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    import io
    from fastapi.responses import StreamingResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "XODIMLAR"
    for col, h in enumerate(["F.I.Sh", "Lavozim"], 1):
        c = ws.cell(1, col, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1B4B7A")
    for r in [("Rasulov Jasur Anvarovich", "Markaz direktori"),
              ("Toshmatova Malika Sobirovna", "Administrator"),
              ("Ergashev Ulug'bek Ilhomovich", "Fan o'qituvchisi")]:
        ws.append(r)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 25

    ws2 = wb.create_sheet("IZOH")
    ws2.cell(1, 1, "Lavozim ustuniga faqat shu variantlardan birini yozing:").font = Font(bold=True)
    for i, nom in enumerate(MARKAZ_LAVOZIMLARI.values(), 2):
        ws2.cell(i, 1, f"• {nom}")
    ws2.column_dimensions["A"].width = 60

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=markaz_xodimlar_shablon.xlsx"},
    )


@app.post("/api/admin/markaz_xodim_import")
async def markaz_xodim_import(token: str, markaz_id: int, fayl: UploadFile = File(...)):
    """Xuddi maktab xodim importi kabi — har biriga hisob va 30 kunlik
    kirish kodi yaratadi. "Fan o'qituvchisi" bo'lganlar keyinchalik
    to'garak (guruh) yaratganda, u AVTOMATIK shu markazga bog'lanadi."""
    _admin_tekshir(token)
    import openpyxl
    import io

    conn = _db()
    cur = conn.cursor()
    _markaz_jadvali(cur)
    cur.execute("SELECT id FROM oquv_markazlari WHERE id=%s", (markaz_id,))
    if not cur.fetchone():
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Markaz topilmadi")

    content = await fayl.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        cur.close(); conn.close()
        raise HTTPException(status_code=400, detail=f"Excel o'qib bo'lmadi: {e}")
    ws = wb["XODIMLAR"] if "XODIMLAR" in wb.sheetnames else wb.active

    _xodim_kod_jadvali(cur)
    natijalar = []
    xato_soni = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0] or not str(row[0]).strip():
            continue
        fish = str(row[0]).strip()
        lavozim_matni = str(row[1]).strip() if len(row) > 1 and row[1] else "Fan o'qituvchisi"
        lavozim_kaliti = _MARKAZ_LAVOZIM_MATNDAN.get(lavozim_matni.lower(), "fan_oqituvchisi")

        try:
            cur.execute("SELECT MIN(user_id) AS eng_kichik FROM users WHERE user_id < 0")
            r = cur.fetchone()
            yangi_id = (r["eng_kichik"] - 1) if r and r["eng_kichik"] is not None else -1

            cur.execute("""
                INSERT INTO users(user_id, full_name, role, markaz_id, lavozim)
                VALUES(%s,%s,'oqituvchi',%s,%s)
            """, (yangi_id, fish, markaz_id, lavozim_kaliti))

            if lavozim_kaliti == "markaz_direktor":
                cur.execute("UPDATE oquv_markazlari SET direktor_user_id=%s WHERE id=%s", (yangi_id, markaz_id))

            kirish_kodi = "".join(secrets.choice(string.ascii_uppercase + string.digits) for _ in range(6))
            cur.execute("INSERT INTO xodim_kod(kod, user_id) VALUES(%s,%s)", (kirish_kodi, yangi_id))

            conn.commit()
            natijalar.append({
                "fish": fish, "lavozim": MARKAZ_LAVOZIMLARI.get(lavozim_kaliti, lavozim_matni),
                "kirish_kodi": kirish_kodi,
            })
        except Exception:
            conn.rollback()
            xato_soni += 1

    cur.close()
    conn.close()
    return {"natijalar": natijalar, "xato_soni": xato_soni}


@app.get("/api/markaz/dashboard")
def markaz_dashboard(token: str, markaz_id: int):
    """Markaz direktori/administratori uchun — MARKAZGA BOG'LANGAN
    BARCHA guruhlarni (to'garaklarni) bitta ekranda ko'rsatadi —
    a'zo soni, oylik summa, o'qituvchi. Aniq boshqarish uchun."""
    user_id = _jwt_tekshir(token)
    conn = _db()
    cur = conn.cursor()
    _markaz_jadvali(cur)
    if not _markaz_boshqaruvchi_mi(cur, user_id, markaz_id):
        cur.close(); conn.close()
        raise HTTPException(status_code=403, detail="Faqat markaz direktori/administratori ko'ra oladi")

    cur.execute("""
        SELECT t.id, t.nomi, t.fan, t.sinf, t.oylik_summa, t.parol, u.full_name AS oqituvchi_ismi,
               (SELECT COUNT(*) FROM togarak_azolar WHERE togarak_id=t.id AND aktiv=TRUE) AS azo_soni
        FROM togaraklar t
        LEFT JOIN users u ON u.user_id = t.teacher_id
        WHERE t.markaz_id=%s AND t.aktiv=TRUE
        ORDER BY t.nomi
    """, (markaz_id,))
    guruhlar = cur.fetchall()
    cur.close()
    conn.close()
    return {"guruhlar": guruhlar}


@app.get("/api/markaz/guruh_tolovlari")
def markaz_guruh_tolovlari(token: str, togarak_id: int, oy: str):
    """Bitta guruh (to'garak) uchun — shu oy to'lov holati, faqat
    TASDIQLANGAN (togarak_azolar, aktiv) a'zolar bo'yicha."""
    user_id = _jwt_tekshir(token)
    conn = _db()
    cur = conn.cursor()
    _tolov_jadvallari(cur)
    cur.execute("SELECT teacher_id, markaz_id, oylik_summa FROM togaraklar WHERE id=%s", (togarak_id,))
    t = cur.fetchone()
    if not t:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Guruh topilmadi")
    ruxsat = t["teacher_id"] == user_id or (t["markaz_id"] and _markaz_boshqaruvchi_mi(cur, user_id, t["markaz_id"]))
    if not ruxsat:
        cur.close(); conn.close()
        raise HTTPException(status_code=403, detail="Faqat shu guruh o'qituvchisi yoki markaz rahbariyati ko'ra oladi")

    kerakli_summa = t["oylik_summa"] or 0
    cur.execute("""
        SELECT u.user_id, u.full_name FROM togarak_azolar ta
        JOIN users u ON u.user_id = ta.user_id
        WHERE ta.togarak_id=%s AND ta.aktiv=TRUE
        ORDER BY u.full_name
    """, (togarak_id,))
    azolar = cur.fetchall()
    cur.execute("SELECT user_id, tolangan_summa, tolov_sanasi FROM tolovlar WHERE togarak_id=%s AND oy=%s", (togarak_id, oy))
    tolovlar_map = {r["user_id"]: r for r in cur.fetchall()}
    cur.close()
    conn.close()

    natija = []
    for a in azolar:
        tt = tolovlar_map.get(a["user_id"])
        tolangan = tt["tolangan_summa"] if tt else 0
        natija.append({
            "user_id": a["user_id"], "full_name": a["full_name"],
            "kerakli_summa": kerakli_summa, "tolangan_summa": tolangan,
            "qarzdor": tolangan < kerakli_summa,
        })
    return {"oquvchilar": natija, "kerakli_summa": kerakli_summa}


class MarkazTolovBelgilash(BaseModel):
    token: str
    user_id: int
    togarak_id: int
    oy: str
    tolangan_summa: int


@app.post("/api/markaz/tolov_belgila")
def markaz_tolov_belgila(sorov: MarkazTolovBelgilash):
    user_id_qiluvchi = _jwt_tekshir(sorov.token)
    conn = _db()
    cur = conn.cursor()
    _tolov_jadvallari(cur)

    # XAVFSIZLIK: faqat shu guruh (to'garak) o'qituvchisi, markaz
    # rahbariyati, yoki admin to'lov belgilay oladi.
    cur.execute("SELECT teacher_id, markaz_id FROM togaraklar WHERE id=%s", (sorov.togarak_id,))
    t = cur.fetchone()
    if not t:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Guruh topilmadi")
    ruxsat = t["teacher_id"] == user_id_qiluvchi or (t["markaz_id"] and _markaz_boshqaruvchi_mi(cur, user_id_qiluvchi, t["markaz_id"]))
    if not ruxsat:
        cur.close(); conn.close()
        raise HTTPException(status_code=403, detail="Faqat shu guruh o'qituvchisi yoki markaz rahbariyati to'lov belgilay oladi")

    cur.execute("""
        INSERT INTO tolovlar(user_id, togarak_id, oy, summa_kerak, tolangan_summa, tolov_sanasi)
        VALUES(%s,%s,%s,%s,%s, CURRENT_DATE)
        ON CONFLICT (user_id, togarak_id, oy) DO UPDATE SET
            tolangan_summa = EXCLUDED.tolangan_summa, tolov_sanasi = CURRENT_DATE
    """, (sorov.user_id, sorov.togarak_id, sorov.oy, sorov.tolangan_summa, sorov.tolangan_summa))
    conn.commit()

    cur.execute("SELECT full_name FROM users WHERE user_id=%s", (sorov.user_id,))
    oquvchi = cur.fetchone()
    cur.execute("SELECT parent_id FROM parent_child WHERE child_id=%s", (sorov.user_id,))
    for oo in cur.fetchall():
        cur.execute(
            "INSERT INTO bildirishnomalar(user_id, matn, turi) VALUES(%s,%s,'tolov')",
            (oo["parent_id"], f"{oquvchi['full_name']} uchun {sorov.oy} oyi to'lovi qabul qilindi: {sorov.tolangan_summa:,} so'm".replace(",", " ")),
        )
    conn.commit()
    cur.close()
    conn.close()
    return {"holat": "saqlandi"}


# ═══════════════════════════════════════════════════════════
# ALLAQACHON HISOBI BOR o'qituvchi — Excel import'da yaratilgan
# SHAXSIY kirish_kodi orqali, YANGI hisob ochmasdan, o'ZINING mavjud
# hisobiga maktab/markaz+lavozimni OLIB OLADI. Bu kirish_kodi bilan
# BOG'LIQ PLACEHOLDER hisobning maktab_id/markaz_id/lavozim'i qanday
# BO'LSA — ANIQ O'SHANI oladi (o'zi tanlab olmaydi, shu sabab
# xavfsiz — faqat Excel'da unga MO'LJALLANGAN lavozim beriladi).
# ═══════════════════════════════════════════════════════════

@app.post("/api/oqituvchi/kirish_kodi_orqali_qoshil")
def kirish_kodi_orqali_qoshil(token: str, kirish_kodi: str):
    """token — chaqiruvchining O'Z (allaqachon mavjud) hisobi.
    kirish_kodi — Excel import paytida SHU KISHI uchun mo'ljallab
    yaratilgan kod (xodim_kod jadvali). Kod to'g'ri bo'lsa —
    chaqiruvchining O'Z hisobiga o'sha maktab_id/markaz_id/lavozim
    yoziladi, kod esa "ishlatildi" deb belgilanadi (qayta ishlatib
    bo'lmaydi)."""
    user_id = _jwt_tekshir(token)
    conn = _db()
    cur = conn.cursor()
    _xodim_kod_jadvali(cur)
    cur.execute("""
        SELECT xk.user_id AS placeholder_id, xk.ishlatildi,
               (xk.yaratildi > NOW() - INTERVAL '30 days') AS hali_yangi
        FROM xodim_kod xk WHERE xk.kod=%s
    """, (kirish_kodi.strip(),))
    kod = cur.fetchone()
    if not kod:
        cur.close(); conn.close()
        raise HTTPException(status_code=400, detail="Kod noto'g'ri")
    if kod["ishlatildi"]:
        cur.close(); conn.close()
        raise HTTPException(status_code=400, detail="Kod allaqachon ishlatilgan")
    if not kod["hali_yangi"]:
        cur.close(); conn.close()
        raise HTTPException(status_code=400, detail="Kod muddati tugagan (30 kun) — admindan yangisini so'rang")

    cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS maktab_id INTEGER")
    cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS markaz_id INTEGER")
    cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS bogcha_id INTEGER")
    cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS lavozim TEXT")
    cur.execute(
        "SELECT maktab_id, markaz_id, bogcha_id, lavozim FROM users WHERE user_id=%s",
        (kod["placeholder_id"],),
    )
    p = cur.fetchone()
    if not p or (not p["maktab_id"] and not p["markaz_id"] and not p["bogcha_id"]):
        cur.close(); conn.close()
        raise HTTPException(status_code=400, detail="Bu kodga tegishli muassasa topilmadi")

    cur.execute(
        "UPDATE users SET maktab_id=%s, markaz_id=%s, bogcha_id=%s, lavozim=%s WHERE user_id=%s",
        (p["maktab_id"], p["markaz_id"], p["bogcha_id"], p["lavozim"], user_id),
    )
    if p["maktab_id"] and p["lavozim"] == "direktor":
        cur.execute("UPDATE maktablar SET direktor_user_id=%s WHERE id=%s", (user_id, p["maktab_id"]))
    if p["markaz_id"] and p["lavozim"] == "markaz_direktor":
        cur.execute("UPDATE oquv_markazlari SET direktor_user_id=%s WHERE id=%s", (user_id, p["markaz_id"]))
    if p["bogcha_id"] and p["lavozim"] == "bogcha_direktor":
        cur.execute("UPDATE bogchalar SET direktor_user_id=%s WHERE id=%s", (user_id, p["bogcha_id"]))
    cur.execute("UPDATE xodim_kod SET ishlatildi=TRUE WHERE kod=%s", (kirish_kodi.strip(),))
    conn.commit()

    natija = {"holat": "qoshildi", "lavozim": p["lavozim"]}
    if p["maktab_id"]:
        cur.execute("SELECT nomi FROM maktablar WHERE id=%s", (p["maktab_id"],))
        m = cur.fetchone()
        natija["joy_nomi"] = m["nomi"] if m else None
    elif p["markaz_id"]:
        cur.execute("SELECT nomi FROM oquv_markazlari WHERE id=%s", (p["markaz_id"],))
        m = cur.fetchone()
        natija["joy_nomi"] = m["nomi"] if m else None
    elif p["bogcha_id"]:
        cur.execute("SELECT nomi FROM bogchalar WHERE id=%s", (p["bogcha_id"],))
        m = cur.fetchone()
        natija["joy_nomi"] = m["nomi"] if m else None
    cur.close()
    conn.close()
    return natija


# ═══════════════════════════════════════════════════════════
# BOG'CHA TIZIMI — maktab/markazga o'xshash, lekin FARQI: bolalar
# (bog'cha yoshidagilar) hisobga EGA BO'LMAYDI — shu sabab GURUHGA
# QO'SHILISH parol bilan EMAS, OPA tomonidan TO'G'RIDAN-TO'G'RI ism
# kiritib qo'shiladi. Bola uchun baribir "placeholder" hisob
# yaratamiz (manfiy user_id, hech qachon login qilmaydi) — shu orqali
# to'lov va ota-onaga bildirishnoma tizimlarini QAYTA ISHLATAMIZ,
# noldan qurmaymiz.
# ═══════════════════════════════════════════════════════════

BOGCHA_LAVOZIMLARI = {
    "bogcha_direktor": "Bog'cha direktori",
    "bogcha_zam": "Bog'cha zam direktori",
    "bogcha_opa": "Bog'cha opasi (tarbiyachi)",
}
_BOGCHA_LAVOZIM_MATNDAN = {v.lower(): k for k, v in BOGCHA_LAVOZIMLARI.items()}
BOGCHA_TURLARI = {"xususiy": "Xususiy", "davlat": "Davlat"}


def _bogcha_jadvali(cur):
    cur.execute("""CREATE TABLE IF NOT EXISTS bogchalar(
        id SERIAL PRIMARY KEY,
        nomi TEXT NOT NULL,
        turi TEXT NOT NULL DEFAULT 'xususiy',
        viloyat TEXT, tuman TEXT,
        direktor_user_id BIGINT REFERENCES users(user_id),
        oylik_tolov INTEGER,
        yaratilgan_at TIMESTAMP DEFAULT NOW()
    )""")
    cur.execute("""CREATE TABLE IF NOT EXISTS bogcha_guruhlari(
        id SERIAL PRIMARY KEY,
        bogcha_id INTEGER NOT NULL REFERENCES bogchalar(id),
        nomi TEXT NOT NULL,
        opa_user_id BIGINT REFERENCES users(user_id),
        qoshilish_paroli TEXT,
        yaratilgan_at TIMESTAMP DEFAULT NOW()
    )""")
    cur.execute("""CREATE TABLE IF NOT EXISTS bogcha_guruh_bolalari(
        id SERIAL PRIMARY KEY,
        guruh_id INTEGER NOT NULL REFERENCES bogcha_guruhlari(id),
        bola_user_id BIGINT NOT NULL REFERENCES users(user_id),
        qoshilgan_at TIMESTAMP DEFAULT NOW()
    )""")
    cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS bogcha_id INTEGER")


def _bogcha_boshqaruvchi_mi(cur, user_id, bogcha_id):
    cur.execute("SELECT 1 FROM admin_akkaunt WHERE uid=%s", (user_id,))
    if cur.fetchone():
        return True
    cur.execute("SELECT lavozim FROM users WHERE user_id=%s AND bogcha_id=%s", (user_id, bogcha_id))
    r = cur.fetchone()
    return bool(r and r["lavozim"] in ("bogcha_direktor", "bogcha_zam"))


class BogchaYaratish(BaseModel):
    token: str
    nomi: str
    turi: str = "xususiy"
    viloyat: Optional[str] = None
    tuman: Optional[str] = None
    direktor_user_id: Optional[int] = None
    oylik_tolov: Optional[int] = None


@app.post("/api/admin/bogcha_yarat")
def bogcha_yarat(sorov: BogchaYaratish):
    _admin_tekshir(sorov.token)
    if not sorov.nomi.strip():
        raise HTTPException(status_code=400, detail="Bog'cha nomi kiritilmagan")
    if sorov.turi not in BOGCHA_TURLARI:
        raise HTTPException(status_code=400, detail="Noto'g'ri bog'cha turi")
    conn = _db()
    cur = conn.cursor()
    _bogcha_jadvali(cur)
    if sorov.direktor_user_id is not None:
        cur.execute("SELECT 1 FROM users WHERE user_id=%s", (sorov.direktor_user_id,))
        if not cur.fetchone():
            cur.close(); conn.close()
            raise HTTPException(status_code=400, detail="Ko'rsatilgan direktor foydalanuvchisi topilmadi")
    cur.execute("""
        INSERT INTO bogchalar(nomi, turi, viloyat, tuman, direktor_user_id, oylik_tolov)
        VALUES(%s,%s,%s,%s,%s,%s) RETURNING id
    """, (sorov.nomi.strip(), sorov.turi, sorov.viloyat, sorov.tuman, sorov.direktor_user_id, sorov.oylik_tolov))
    yangi_id = cur.fetchone()["id"]
    conn.commit()
    cur.close()
    conn.close()
    return {"holat": "yaratildi", "bogcha_id": yangi_id}


@app.get("/api/admin/bogchalar")
def bogchalar_royxati(token: str):
    _admin_tekshir(token)
    conn = _db()
    cur = conn.cursor()
    _bogcha_jadvali(cur)
    cur.execute("""
        SELECT b.id, b.nomi, b.turi, b.viloyat, b.tuman, b.direktor_user_id, b.oylik_tolov,
               u.full_name AS direktor_ismi
        FROM bogchalar b
        LEFT JOIN users u ON u.user_id = b.direktor_user_id
        ORDER BY b.nomi
    """)
    natija = cur.fetchall()
    cur.close()
    conn.close()
    return {"bogchalar": natija}


@app.get("/api/admin/bogcha_xodim_shablon")
def bogcha_xodim_shablon(token: str):
    _admin_tekshir(token)
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    import io
    from fastapi.responses import StreamingResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "XODIMLAR"
    for col, h in enumerate(["F.I.Sh", "Lavozim", "Guruh rahbarligi (ixtiyoriy)"], 1):
        c = ws.cell(1, col, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1B4B7A")
    for r in [("Xolmatova Gulnora Rahimovna", "Bog'cha direktori", ""),
              ("Sodiqova Dilfuza Nematovna", "Bog'cha zam direktori", ""),
              ("Yusupova Shahnoza Karimovna", "Bog'cha opasi (tarbiyachi)", "Quyoshcha guruhi"),
              ("Rahimova Zulfiya To'raevna", "Bog'cha opasi (tarbiyachi)", "Kichkintoylar guruhi")]:
        ws.append(r)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 28

    ws2 = wb.create_sheet("IZOH")
    ws2.cell(1, 1, "Lavozim ustuniga faqat shu variantlardan birini yozing:").font = Font(bold=True)
    for i, nom in enumerate(BOGCHA_LAVOZIMLARI.values(), 2):
        ws2.cell(i, 1, f"• {nom}")
    ws2.cell(len(BOGCHA_LAVOZIMLARI) + 3, 1,
             "Guruh rahbarligi — faqat 'Bog'cha opasi' bo'lgan xodim uchun to'ldiring (masalan: Quyoshcha guruhi). Har xil nom — har xil guruh.")
    ws2.column_dimensions["A"].width = 70

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=bogcha_xodimlar_shablon.xlsx"},
    )


@app.post("/api/admin/bogcha_xodim_import")
async def bogcha_xodim_import(token: str, bogcha_id: int, fayl: UploadFile = File(...)):
    """Xuddi maktab/markaz xodim importi kabi. "Guruh rahbarligi"
    to'ldirilgan bo'lsa (faqat bog'cha opalari uchun mazmunli) — o'sha
    nomdagi guruh yaratiladi/yangilanadi, 4 xonali (odatda ota-onaga
    emas, guruh ICHKI hisoboti uchun) parol biriktiriladi."""
    _admin_tekshir(token)
    import openpyxl
    import io

    conn = _db()
    cur = conn.cursor()
    _bogcha_jadvali(cur)
    cur.execute("SELECT id FROM bogchalar WHERE id=%s", (bogcha_id,))
    if not cur.fetchone():
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Bog'cha topilmadi")

    content = await fayl.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        cur.close(); conn.close()
        raise HTTPException(status_code=400, detail=f"Excel o'qib bo'lmadi: {e}")
    ws = wb["XODIMLAR"] if "XODIMLAR" in wb.sheetnames else wb.active

    _xodim_kod_jadvali(cur)
    natijalar = []
    xato_soni = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0] or not str(row[0]).strip():
            continue
        fish = str(row[0]).strip()
        lavozim_matni = str(row[1]).strip() if len(row) > 1 and row[1] else "Bog'cha opasi (tarbiyachi)"
        guruh_nomi = str(row[2]).strip() if len(row) > 2 and row[2] else ""
        lavozim_kaliti = _BOGCHA_LAVOZIM_MATNDAN.get(lavozim_matni.lower(), "bogcha_opa")

        try:
            cur.execute("SELECT MIN(user_id) AS eng_kichik FROM users WHERE user_id < 0")
            r = cur.fetchone()
            yangi_id = (r["eng_kichik"] - 1) if r and r["eng_kichik"] is not None else -1

            cur.execute("""
                INSERT INTO users(user_id, full_name, role, bogcha_id, lavozim)
                VALUES(%s,%s,'oqituvchi',%s,%s)
            """, (yangi_id, fish, bogcha_id, lavozim_kaliti))

            if lavozim_kaliti == "bogcha_direktor":
                cur.execute("UPDATE bogchalar SET direktor_user_id=%s WHERE id=%s", (yangi_id, bogcha_id))

            kirish_kodi = "".join(secrets.choice(string.ascii_uppercase + string.digits) for _ in range(6))
            cur.execute("INSERT INTO xodim_kod(kod, user_id) VALUES(%s,%s)", (kirish_kodi, yangi_id))

            guruh_paroli = None
            if guruh_nomi and lavozim_kaliti == "bogcha_opa":
                guruh_paroli = "".join(secrets.choice(string.digits) for _ in range(4))
                cur.execute("""
                    INSERT INTO bogcha_guruhlari(bogcha_id, nomi, opa_user_id, qoshilish_paroli)
                    VALUES(%s,%s,%s,%s) RETURNING id
                """, (bogcha_id, guruh_nomi, yangi_id, guruh_paroli))

            conn.commit()
            natijalar.append({
                "fish": fish, "lavozim": BOGCHA_LAVOZIMLARI.get(lavozim_kaliti, lavozim_matni),
                "kirish_kodi": kirish_kodi, "guruh_nomi": guruh_nomi or None,
            })
        except Exception:
            conn.rollback()
            xato_soni += 1

    cur.close()
    conn.close()
    return {"natijalar": natijalar, "xato_soni": xato_soni}


@app.get("/api/opa/mening_guruhlarim")
def opa_mening_guruhlarim(token: str):
    """Bog'cha opasi RAHBAR bo'lgan guruhlari — har birida nechta
    bola borligi bilan."""
    user_id = _jwt_tekshir(token)
    conn = _db()
    cur = conn.cursor()
    _bogcha_jadvali(cur)
    cur.execute("""
        SELECT g.id, g.nomi, g.qoshilish_paroli, b.id AS bogcha_id, b.nomi AS bogcha_nomi, b.oylik_tolov,
               (SELECT COUNT(*) FROM bogcha_guruh_bolalari WHERE guruh_id=g.id) AS bola_soni
        FROM bogcha_guruhlari g
        JOIN bogchalar b ON b.id = g.bogcha_id
        WHERE g.opa_user_id=%s
        ORDER BY g.nomi
    """, (user_id,))
    natija = cur.fetchall()
    cur.close()
    conn.close()
    return {"guruhlar": natija}


class BolaQoshish(BaseModel):
    token: str
    guruh_id: int
    bola_ismi: str
    ota_ona_user_id: Optional[int] = None  # mavjud ota-ona hisobiga bog'lash uchun (ixtiyoriy)


@app.get("/api/opa/ota_ona_qidir")
def opa_ota_ona_qidir(token: str, ism: str):
    """Opa (yoki har qanday tizimga kirgan xodim) uchun — bola
    qo'shayotganda uning ota-onasini ISM bo'yicha qidirib topish va
    bog'lash uchun. Faqat role='ota-ona' hisoblar orasidan qidiradi."""
    _jwt_tekshir(token)
    if len(ism.strip()) < 2:
        return {"natijalar": []}
    conn = _db()
    cur = conn.cursor()
    cur.execute("""
        SELECT user_id, full_name FROM users
        WHERE role='ota-ona' AND full_name ILIKE %s
        ORDER BY full_name LIMIT 10
    """, (f"%{ism.strip()}%",))
    natija = cur.fetchall()
    cur.close()
    conn.close()
    return {"natijalar": natija}


@app.get("/api/oqituvchi/universitet_guruh_qidir")
def oqituvchi_universitet_guruh_qidir(token: str, nomi: str):
    """Professor to'garak (kurs) yaratayotganda — bu kursni qaysi
    universitet guruhi uchun o'qitayotganini nomi bo'yicha qidirib
    topishi uchun."""
    _jwt_tekshir(token)
    if len(nomi.strip()) < 1:
        return {"natijalar": []}
    conn = _db()
    cur = conn.cursor()
    _universitet_jadvali(cur)
    cur.execute("""
        SELECT g.id, g.nomi, g.kurs, g.yonalish, k.nomi AS kafedra_nomi
        FROM universitet_guruhlari g LEFT JOIN kafedralar k ON k.id = g.kafedra_id
        WHERE g.nomi ILIKE %s ORDER BY g.nomi LIMIT 10
    """, (f"%{nomi.strip()}%",))
    natija = cur.fetchall()
    cur.close()
    conn.close()
    return {"natijalar": natija}



@app.post("/api/opa/bola_qoshish")
def opa_bola_qoshish(sorov: BolaQoshish):
    """Opa bolani ISMI bilan TO'G'RIDAN-TO'G'RI guruhga qo'shadi —
    bola hisobga ega bo'lmagani uchun, ichki "placeholder" hisob
    yaratiladi (bu hisob hech qachon login qilmaydi, faqat to'lov/
    bildirishnoma tizimlarini qayta ishlatish uchun kerak)."""
    user_id = _jwt_tekshir(sorov.token)
    if not sorov.bola_ismi.strip():
        raise HTTPException(status_code=400, detail="Bola ismini kiriting")
    conn = _db()
    cur = conn.cursor()
    _bogcha_jadvali(cur)
    cur.execute("SELECT opa_user_id FROM bogcha_guruhlari WHERE id=%s", (sorov.guruh_id,))
    g = cur.fetchone()
    if not g:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Guruh topilmadi")
    cur.execute("SELECT 1 FROM admin_akkaunt WHERE uid=%s", (user_id,))
    admin_mi = cur.fetchone() is not None
    if not admin_mi and g["opa_user_id"] != user_id:
        cur.close(); conn.close()
        raise HTTPException(status_code=403, detail="Faqat shu guruh opasi yoki admin qo'sha oladi")

    cur.execute("SELECT MIN(user_id) AS eng_kichik FROM users WHERE user_id < 0")
    r = cur.fetchone()
    yangi_id = (r["eng_kichik"] - 1) if r and r["eng_kichik"] is not None else -1
    cur.execute("INSERT INTO users(user_id, full_name, role) VALUES(%s,%s,'oquvchi')", (yangi_id, sorov.bola_ismi.strip()))
    cur.execute("INSERT INTO bogcha_guruh_bolalari(guruh_id, bola_user_id) VALUES(%s,%s)", (sorov.guruh_id, yangi_id))
    if sorov.ota_ona_user_id is not None:
        cur.execute("SELECT 1 FROM users WHERE user_id=%s", (sorov.ota_ona_user_id,))
        if cur.fetchone():
            cur.execute(
                "INSERT INTO parent_child(parent_id, child_id) VALUES(%s,%s) ON CONFLICT DO NOTHING",
                (sorov.ota_ona_user_id, yangi_id),
            )
    conn.commit()
    cur.close()
    conn.close()
    return {"holat": "qoshildi", "bola_id": yangi_id}


@app.get("/api/opa/guruh_bolalari")
def opa_guruh_bolalari(token: str, guruh_id: int):
    user_id = _jwt_tekshir(token)
    conn = _db()
    cur = conn.cursor()
    _bogcha_jadvali(cur)
    cur.execute("SELECT opa_user_id, bogcha_id FROM bogcha_guruhlari WHERE id=%s", (guruh_id,))
    g = cur.fetchone()
    if not g:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Guruh topilmadi")
    cur.execute("SELECT 1 FROM admin_akkaunt WHERE uid=%s", (user_id,))
    admin_mi = cur.fetchone() is not None
    if not admin_mi and g["opa_user_id"] != user_id:
        cur.close(); conn.close()
        raise HTTPException(status_code=403, detail="Faqat shu guruh opasi yoki admin ko'ra oladi")

    cur.execute("SELECT oylik_tolov FROM bogchalar WHERE id=%s", (g["bogcha_id"],))
    b = cur.fetchone()
    kerakli_summa = (b["oylik_tolov"] if b else None) or 0

    cur.execute("""
        SELECT gb.id AS roster_id, u.user_id, u.full_name
        FROM bogcha_guruh_bolalari gb JOIN users u ON u.user_id = gb.bola_user_id
        WHERE gb.guruh_id=%s ORDER BY u.full_name
    """, (guruh_id,))
    bolalar = cur.fetchall()

    joriy_oy = datetime.now().strftime("%Y-%m")
    cur.execute("ALTER TABLE tolovlar ADD COLUMN IF NOT EXISTS bogcha_guruh_id INTEGER REFERENCES bogcha_guruhlari(id)")
    cur.execute("CREATE UNIQUE INDEX IF NOT EXISTS tolovlar_bogcha_unique ON tolovlar(user_id, bogcha_guruh_id, oy)")
    cur.execute("SELECT user_id, tolangan_summa FROM tolovlar WHERE bogcha_guruh_id=%s AND oy=%s", (guruh_id, joriy_oy))
    tolovlar_map = {r["user_id"]: r["tolangan_summa"] for r in cur.fetchall()}
    cur.close()
    conn.close()

    natija = []
    for bl in bolalar:
        tolangan = tolovlar_map.get(bl["user_id"], 0)
        natija.append({
            "roster_id": bl["roster_id"], "user_id": bl["user_id"], "full_name": bl["full_name"],
            "kerakli_summa": kerakli_summa, "tolangan_summa": tolangan, "qarzdor": tolangan < kerakli_summa,
        })
    return {"bolalar": natija, "kerakli_summa": kerakli_summa, "oy": joriy_oy}


@app.delete("/api/opa/bolani_chiqar")
def opa_bolani_chiqar(token: str, roster_id: int):
    user_id = _jwt_tekshir(token)
    conn = _db()
    cur = conn.cursor()
    cur.execute("""
        SELECT g.opa_user_id FROM bogcha_guruh_bolalari gb
        JOIN bogcha_guruhlari g ON g.id = gb.guruh_id WHERE gb.id=%s
    """, (roster_id,))
    r = cur.fetchone()
    if not r:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Topilmadi")
    cur.execute("SELECT 1 FROM admin_akkaunt WHERE uid=%s", (user_id,))
    admin_mi = cur.fetchone() is not None
    if not admin_mi and r["opa_user_id"] != user_id:
        cur.close(); conn.close()
        raise HTTPException(status_code=403, detail="Faqat shu guruh opasi yoki admin chiqara oladi")
    cur.execute("DELETE FROM bogcha_guruh_bolalari WHERE id=%s", (roster_id,))
    conn.commit()
    cur.close()
    conn.close()
    return {"holat": "chiqarildi"}


class BogchaTolovBelgilash(BaseModel):
    token: str
    bola_user_id: int
    guruh_id: int
    oy: str
    tolangan_summa: int


@app.post("/api/opa/tolov_belgila")
def opa_tolov_belgila(sorov: BogchaTolovBelgilash):
    user_id = _jwt_tekshir(sorov.token)
    conn = _db()
    cur = conn.cursor()
    _tolov_jadvallari(cur)
    cur.execute("ALTER TABLE tolovlar ADD COLUMN IF NOT EXISTS bogcha_guruh_id INTEGER REFERENCES bogcha_guruhlari(id)")
    cur.execute("CREATE UNIQUE INDEX IF NOT EXISTS tolovlar_bogcha_unique ON tolovlar(user_id, bogcha_guruh_id, oy)")

    # XAVFSIZLIK: faqat shu guruh opasi yoki admin to'lov belgilay oladi.
    cur.execute("SELECT opa_user_id FROM bogcha_guruhlari WHERE id=%s", (sorov.guruh_id,))
    g = cur.fetchone()
    if not g:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Guruh topilmadi")
    cur.execute("SELECT 1 FROM admin_akkaunt WHERE uid=%s", (user_id,))
    admin_mi = cur.fetchone() is not None
    if not admin_mi and g["opa_user_id"] != user_id:
        cur.close(); conn.close()
        raise HTTPException(status_code=403, detail="Faqat shu guruh opasi yoki admin to'lov belgilay oladi")

    cur.execute("""
        INSERT INTO tolovlar(user_id, bogcha_guruh_id, oy, summa_kerak, tolangan_summa, tolov_sanasi)
        VALUES(%s,%s,%s,%s,%s, CURRENT_DATE)
        ON CONFLICT (user_id, bogcha_guruh_id, oy) DO UPDATE SET
            tolangan_summa = EXCLUDED.tolangan_summa, tolov_sanasi = CURRENT_DATE
    """, (sorov.bola_user_id, sorov.guruh_id, sorov.oy, sorov.tolangan_summa, sorov.tolangan_summa))
    conn.commit()

    cur.execute("SELECT full_name FROM users WHERE user_id=%s", (sorov.bola_user_id,))
    bola = cur.fetchone()
    cur.execute("SELECT parent_id FROM parent_child WHERE child_id=%s", (sorov.bola_user_id,))
    for oo in cur.fetchall():
        cur.execute(
            "INSERT INTO bildirishnomalar(user_id, matn, turi) VALUES(%s,%s,'tolov')",
            (oo["parent_id"], f"{bola['full_name']} uchun {sorov.oy} oyi bog'cha to'lovi qabul qilindi: {sorov.tolangan_summa:,} so'm".replace(",", " ")),
        )
    conn.commit()
    cur.close()
    conn.close()
    return {"holat": "saqlandi"}


# ═══════════════════════════════════════════════════════════
# OLIY TA'LIM TIZIMI — 4 QAVATLI TUZILMA:
#   Universitet → Fakultet → Kafedra → Guruh (talabalar)
#
# Maktabdan farqi: chuqurroq ierarxiya. Har qavat o'zining rahbarini
# (rektor/dekan/kafedra mudiri/guruh kuratori) belgilashi mumkin.
# Talaba guruhga sinf kabi PAROL bilan qo'shiladi (avtomatik profil
# mosligisiz — chunki talaba profilida fakultet/kafedra maydonlari
# hozircha yo'q, sodda parol-orqali-qo'shilish yetarli).
# ═══════════════════════════════════════════════════════════

UNIVERSITET_LAVOZIMLARI = {
    "rektor": "Rektor",
    "prorektor": "Prorektor",
    "dekan": "Dekan",
    "kafedra_mudiri": "Kafedra mudiri",
    "professor_oqituvchi": "Professor-o'qituvchi",
}
_UNIVERSITET_LAVOZIM_MATNDAN = {v.lower(): k for k, v in UNIVERSITET_LAVOZIMLARI.items()}


def _universitet_jadvali(cur):
    cur.execute("""CREATE TABLE IF NOT EXISTS universitetlar(
        id SERIAL PRIMARY KEY, nomi TEXT NOT NULL, viloyat TEXT, tuman TEXT,
        rektor_user_id BIGINT REFERENCES users(user_id), yaratilgan_at TIMESTAMP DEFAULT NOW()
    )""")
    cur.execute("""CREATE TABLE IF NOT EXISTS fakultetlar(
        id SERIAL PRIMARY KEY, universitet_id INTEGER NOT NULL REFERENCES universitetlar(id),
        nomi TEXT NOT NULL, dekan_user_id BIGINT REFERENCES users(user_id), yaratilgan_at TIMESTAMP DEFAULT NOW()
    )""")
    cur.execute("""CREATE TABLE IF NOT EXISTS kafedralar(
        id SERIAL PRIMARY KEY, fakultet_id INTEGER NOT NULL REFERENCES fakultetlar(id),
        nomi TEXT NOT NULL, mudir_user_id BIGINT REFERENCES users(user_id), yaratilgan_at TIMESTAMP DEFAULT NOW()
    )""")
    cur.execute("""CREATE TABLE IF NOT EXISTS universitet_guruhlari(
        id SERIAL PRIMARY KEY, kafedra_id INTEGER NOT NULL REFERENCES kafedralar(id),
        nomi TEXT NOT NULL, kurs INTEGER, yonalish TEXT,
        rahbar_user_id BIGINT REFERENCES users(user_id), qoshilish_paroli TEXT,
        yaratilgan_at TIMESTAMP DEFAULT NOW()
    )""")
    cur.execute("""CREATE TABLE IF NOT EXISTS universitet_guruh_azolari(
        id SERIAL PRIMARY KEY, guruh_id INTEGER NOT NULL REFERENCES universitet_guruhlari(id),
        user_id BIGINT NOT NULL REFERENCES users(user_id), qoshilgan_at TIMESTAMP DEFAULT NOW(),
        UNIQUE(guruh_id, user_id)
    )""")
    cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS universitet_id INTEGER")


def _universitet_boshqaruvchi_mi(cur, user_id, universitet_id):
    cur.execute("SELECT 1 FROM admin_akkaunt WHERE uid=%s", (user_id,))
    if cur.fetchone():
        return True
    cur.execute("SELECT lavozim FROM users WHERE user_id=%s AND universitet_id=%s", (user_id, universitet_id))
    r = cur.fetchone()
    return bool(r and r["lavozim"] in ("rektor", "prorektor"))


class UniversitetYaratish(BaseModel):
    token: str
    nomi: str
    viloyat: Optional[str] = None
    tuman: Optional[str] = None
    rektor_user_id: Optional[int] = None


@app.post("/api/admin/universitet_yarat")
def universitet_yarat(sorov: UniversitetYaratish):
    _admin_tekshir(sorov.token)
    if not sorov.nomi.strip():
        raise HTTPException(status_code=400, detail="Universitet nomi kiritilmagan")
    conn = _db()
    cur = conn.cursor()
    _universitet_jadvali(cur)
    if sorov.rektor_user_id is not None:
        cur.execute("SELECT 1 FROM users WHERE user_id=%s", (sorov.rektor_user_id,))
        if not cur.fetchone():
            cur.close(); conn.close()
            raise HTTPException(status_code=400, detail="Ko'rsatilgan rektor foydalanuvchisi topilmadi")
    cur.execute("""
        INSERT INTO universitetlar(nomi, viloyat, tuman, rektor_user_id)
        VALUES(%s,%s,%s,%s) RETURNING id
    """, (sorov.nomi.strip(), sorov.viloyat, sorov.tuman, sorov.rektor_user_id))
    yangi_id = cur.fetchone()["id"]
    conn.commit()
    cur.close()
    conn.close()
    return {"holat": "yaratildi", "universitet_id": yangi_id}


@app.get("/api/admin/universitetlar")
def universitetlar_royxati(token: str):
    _admin_tekshir(token)
    conn = _db()
    cur = conn.cursor()
    _universitet_jadvali(cur)
    cur.execute("""
        SELECT u.id, u.nomi, u.viloyat, u.tuman, u.rektor_user_id, us.full_name AS rektor_ismi,
               (SELECT COUNT(*) FROM fakultetlar WHERE universitet_id=u.id) AS fakultet_soni
        FROM universitetlar u LEFT JOIN users us ON us.user_id = u.rektor_user_id
        ORDER BY u.nomi
    """)
    natija = cur.fetchall()
    cur.close()
    conn.close()
    return {"universitetlar": natija}


class FakultetYaratish(BaseModel):
    token: str
    universitet_id: int
    nomi: str
    dekan_user_id: Optional[int] = None


@app.post("/api/admin/fakultet_yarat")
def fakultet_yarat(sorov: FakultetYaratish):
    _admin_tekshir(sorov.token)
    if not sorov.nomi.strip():
        raise HTTPException(status_code=400, detail="Fakultet nomi kiritilmagan")
    conn = _db()
    cur = conn.cursor()
    _universitet_jadvali(cur)
    cur.execute("""
        INSERT INTO fakultetlar(universitet_id, nomi, dekan_user_id)
        VALUES(%s,%s,%s) RETURNING id
    """, (sorov.universitet_id, sorov.nomi.strip(), sorov.dekan_user_id))
    yangi_id = cur.fetchone()["id"]
    conn.commit()
    cur.close()
    conn.close()
    return {"holat": "yaratildi", "fakultet_id": yangi_id}


@app.get("/api/admin/fakultetlar")
def fakultetlar_royxati(token: str, universitet_id: int):
    _admin_tekshir(token)
    conn = _db()
    cur = conn.cursor()
    _universitet_jadvali(cur)
    cur.execute("""
        SELECT f.id, f.nomi, f.dekan_user_id, u.full_name AS dekan_ismi,
               (SELECT COUNT(*) FROM kafedralar WHERE fakultet_id=f.id) AS kafedra_soni
        FROM fakultetlar f LEFT JOIN users u ON u.user_id = f.dekan_user_id
        WHERE f.universitet_id=%s ORDER BY f.nomi
    """, (universitet_id,))
    natija = cur.fetchall()
    cur.close()
    conn.close()
    return {"fakultetlar": natija}


class KafedraYaratish(BaseModel):
    token: str
    fakultet_id: int
    nomi: str
    mudir_user_id: Optional[int] = None


@app.post("/api/admin/kafedra_yarat")
def kafedra_yarat(sorov: KafedraYaratish):
    _admin_tekshir(sorov.token)
    if not sorov.nomi.strip():
        raise HTTPException(status_code=400, detail="Kafedra nomi kiritilmagan")
    conn = _db()
    cur = conn.cursor()
    _universitet_jadvali(cur)
    cur.execute("""
        INSERT INTO kafedralar(fakultet_id, nomi, mudir_user_id)
        VALUES(%s,%s,%s) RETURNING id
    """, (sorov.fakultet_id, sorov.nomi.strip(), sorov.mudir_user_id))
    yangi_id = cur.fetchone()["id"]
    conn.commit()
    cur.close()
    conn.close()
    return {"holat": "yaratildi", "kafedra_id": yangi_id}


@app.get("/api/admin/kafedralar")
def kafedralar_royxati(token: str, fakultet_id: int):
    _admin_tekshir(token)
    conn = _db()
    cur = conn.cursor()
    _universitet_jadvali(cur)
    cur.execute("""
        SELECT k.id, k.nomi, k.mudir_user_id, u.full_name AS mudir_ismi,
               (SELECT COUNT(*) FROM universitet_guruhlari WHERE kafedra_id=k.id) AS guruh_soni
        FROM kafedralar k LEFT JOIN users u ON u.user_id = k.mudir_user_id
        WHERE k.fakultet_id=%s ORDER BY k.nomi
    """, (fakultet_id,))
    natija = cur.fetchall()
    cur.close()
    conn.close()
    return {"kafedralar": natija}


class UniversitetGuruhYaratish(BaseModel):
    token: str
    kafedra_id: int
    nomi: str
    kurs: Optional[int] = None
    yonalish: Optional[str] = None
    rahbar_user_id: Optional[int] = None


@app.post("/api/admin/universitet_guruh_yarat")
def universitet_guruh_yarat(sorov: UniversitetGuruhYaratish):
    _admin_tekshir(sorov.token)
    if not sorov.nomi.strip():
        raise HTTPException(status_code=400, detail="Guruh nomi kiritilmagan")
    if sorov.kurs is not None and sorov.kurs not in range(1, 7):
        raise HTTPException(status_code=400, detail="Kurs 1 dan 6 gacha bo'lishi kerak")
    conn = _db()
    cur = conn.cursor()
    _universitet_jadvali(cur)
    paroli = "".join(secrets.choice(string.digits) for _ in range(4))
    cur.execute("""
        INSERT INTO universitet_guruhlari(kafedra_id, nomi, kurs, yonalish, rahbar_user_id, qoshilish_paroli)
        VALUES(%s,%s,%s,%s,%s,%s) RETURNING id
    """, (sorov.kafedra_id, sorov.nomi.strip(), sorov.kurs, sorov.yonalish, sorov.rahbar_user_id, paroli))
    yangi_id = cur.fetchone()["id"]
    conn.commit()
    cur.close()
    conn.close()
    return {"holat": "yaratildi", "guruh_id": yangi_id, "qoshilish_paroli": paroli}


@app.get("/api/admin/universitet_guruhlari")
def universitet_guruhlari_royxati(token: str, kafedra_id: int):
    _admin_tekshir(token)
    conn = _db()
    cur = conn.cursor()
    _universitet_jadvali(cur)
    cur.execute("""
        SELECT g.id, g.nomi, g.kurs, g.yonalish, g.qoshilish_paroli, g.rahbar_user_id, u.full_name AS rahbar_ismi,
               (SELECT COUNT(*) FROM universitet_guruh_azolari WHERE guruh_id=g.id) AS talaba_soni
        FROM universitet_guruhlari g LEFT JOIN users u ON u.user_id = g.rahbar_user_id
        WHERE g.kafedra_id=%s ORDER BY g.kurs NULLS LAST, g.nomi
    """, (kafedra_id,))
    natija = cur.fetchall()
    cur.close()
    conn.close()
    return {"guruhlar": natija}


@app.post("/api/talaba/guruhga_qoshil")
def talaba_guruhga_qoshil(token: str, parol: str):
    """Talaba universitet guruhiga FAQAT 4 xonali parol bilan
    qo'shiladi — guruh ID'sini bilishi shart emas (to'garakka
    qo'shilish bilan bir xil uslub)."""
    user_id = _jwt_tekshir(token)
    conn = _db()
    cur = conn.cursor()
    _universitet_jadvali(cur)
    cur.execute("SELECT id, nomi FROM universitet_guruhlari WHERE qoshilish_paroli=%s", (parol.strip(),))
    g = cur.fetchone()
    if not g:
        cur.close(); conn.close()
        raise HTTPException(status_code=400, detail="Parol noto'g'ri")
    cur.execute(
        "INSERT INTO universitet_guruh_azolari(guruh_id, user_id) VALUES(%s,%s) ON CONFLICT DO NOTHING",
        (g["id"], user_id),
    )
    conn.commit()
    cur.close()
    conn.close()
    return {"holat": "qoshildi", "guruh_nomi": g["nomi"]}


@app.get("/api/universitet/mening_guruhlarim")
def universitet_mening_guruhlarim(token: str):
    """Kurator RAHBAR bo'lgan universitet guruhlari ro'yxati."""
    user_id = _jwt_tekshir(token)
    conn = _db()
    cur = conn.cursor()
    _universitet_jadvali(cur)
    cur.execute("""
        SELECT g.id, g.nomi, g.kurs, g.yonalish,
               (SELECT COUNT(*) FROM universitet_guruh_azolari WHERE guruh_id=g.id) AS talaba_soni
        FROM universitet_guruhlari g WHERE g.rahbar_user_id=%s ORDER BY g.nomi
    """, (user_id,))
    natija = cur.fetchall()
    cur.close()
    conn.close()
    return {"guruhlar": natija}


@app.get("/api/universitet/guruh_bilimi")
def universitet_guruh_bilimi(token: str, guruh_id: int):
    """Guruh kuratori/dekan/rektor uchun — ENG MUHIM ko'rsatkich:
    guruhga bog'langan HAR BIR fan (professor kursi) bo'yicha, HAR BIR
    talabaning silabusdagi mavzularni qanchalik bilishi — mavjud
    "Bilim" mexanizmi (learned_topics) orqali, GPA emas, aniq va
    tushunarli % ko'rinishida."""
    user_id = _jwt_tekshir(token)
    conn = _db()
    cur = conn.cursor()
    _universitet_jadvali(cur)
    cur.execute("SELECT rahbar_user_id, nomi FROM universitet_guruhlari WHERE id=%s", (guruh_id,))
    g = cur.fetchone()
    if not g:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Guruh topilmadi")
    cur.execute("SELECT 1 FROM admin_akkaunt WHERE uid=%s", (user_id,))
    admin_mi = cur.fetchone() is not None
    if not admin_mi and g["rahbar_user_id"] != user_id:
        cur.close(); conn.close()
        raise HTTPException(status_code=403, detail="Faqat shu guruh kuratori yoki admin ko'ra oladi")

    # Guruhga bog'langan barcha fanlar (professor kurslari)
    cur.execute("""
        SELECT t.id AS togarak_id, t.nomi AS kurs_nomi, t.fan, u.full_name AS professor_ismi
        FROM togaraklar t LEFT JOIN users u ON u.user_id = t.teacher_id
        WHERE t.universitet_guruh_id=%s AND t.aktiv=TRUE
        ORDER BY t.fan
    """, (guruh_id,))
    kurslar = cur.fetchall()

    # Guruhdagi barcha talabalar
    cur.execute("""
        SELECT u.user_id, u.full_name FROM universitet_guruh_azolari ga
        JOIN users u ON u.user_id = ga.user_id WHERE ga.guruh_id=%s ORDER BY u.full_name
    """, (guruh_id,))
    talabalar = cur.fetchall()

    natija_kurslar = []
    for k in kurslar:
        cur.execute("""
            SELECT ta.user_id,
                   COUNT(DISTINCT tm.topic_code) AS jami_mavzu,
                   COUNT(DISTINCT lt.topic_code) AS ishlangan_mavzu,
                   AVG(lt.score) AS ortacha_ball
            FROM togarak_azolar ta
            JOIN togarak_mavzulari tm ON tm.togarak_id = ta.togarak_id
            LEFT JOIN learned_topics lt ON lt.topic_code = tm.topic_code AND lt.user_id = ta.user_id
            WHERE ta.togarak_id=%s AND ta.aktiv=TRUE
            GROUP BY ta.user_id
        """, (k["togarak_id"],))
        talaba_natijalari = {r["user_id"]: r for r in cur.fetchall()}

        talabalar_royxati = []
        for t in talabalar:
            r = talaba_natijalari.get(t["user_id"])
            if r and r["jami_mavzu"]:
                foiz = round((r["ishlangan_mavzu"] or 0) / r["jami_mavzu"] * 100)
                ball = round(r["ortacha_ball"]) if r["ortacha_ball"] is not None else None
            else:
                foiz, ball = 0, None
            talabalar_royxati.append({
                "user_id": t["user_id"], "full_name": t["full_name"],
                "otilgan_foiz": foiz, "ortacha_ball": ball,
            })
        natija_kurslar.append({
            "togarak_id": k["togarak_id"], "kurs_nomi": k["kurs_nomi"], "fan": k["fan"],
            "professor_ismi": k["professor_ismi"], "talabalar": talabalar_royxati,
        })

    cur.close()
    conn.close()
    return {"guruh_nomi": g["nomi"], "talaba_soni": len(talabalar), "kurslar": natija_kurslar}


# ═══════════════════════════════════════════════════════════
# SINOV MUHITI — bir bosishda 4 tizimning HAMMASINI (maktab, bog'cha,
# markaz, universitet) soxta odamlar bilan to'liq yaratadi, VA admin
# ularning HAR BIRI sifatida (Google login'siz) darhol kira oladi —
# faqat SINOV/TEST maqsadida, faqat admin ishlatishi mumkin.
#
# XAVFSIZLIK: "sifatida kirish" tokeni ODDIY tokendan farqli —
# atigi 2 SOAT amal qiladi (30 kun emas), shu orqali xavf chegaralanadi.
# ═══════════════════════════════════════════════════════════

def _sinov_jwt_yarat(user_id: int) -> str:
    """Admin uchun — 'sifatida kirish' tokeni. Uzoq muddatli, chuqur
    sinov (darslar qo'yish, baholash, kontent yuklash) uchun oddiy
    foydalanuvchi seansi bilan BIR XIL — 30 kun amal qiladi."""
    payload = {"user_id": user_id, "exp": datetime.now(timezone.utc) + timedelta(days=30)}
    return jwt.encode(payload, JWT_MAXFIY_KALIT, algorithm="HS256")


def _keyingi_manfiy_id(cur):
    """Joriy eng kichik (manfiy) user_id'dan BOSHLAB pastga hisoblash
    uchun boshlang'ich nuqta."""
    cur.execute("SELECT MIN(user_id) AS eng_kichik FROM users WHERE user_id < 0")
    r = cur.fetchone()
    return (r["eng_kichik"] - 1) if r and r["eng_kichik"] is not None else -1


@app.post("/api/admin/sinov_muhit_yarat")
def sinov_muhit_yarat(token: str):
    """HAQIQIY HAJMDAGI sinov muhiti — universitet (~1000 talaba, 6
    fakultet, 12 kafedra, 36 guruh), maktab (~550 o'quvchi, 1-11 sinf
    x A/B, pulli), markaz (~150 talaba, 5 guruh), bog'cha (~75 bola +
    75 ota-ona, 3 guruh). Talaba/o'quvchi/bola darajasida HAMMASI
    OMMAVIY (bulk) kiritiladi — minglab alohida so'rov EMAS, shu
    sabab tez va ishonchli ishlaydi."""
    _admin_tekshir(token)
    conn = _db()
    cur = conn.cursor()
    try:
        _maktab_jadvali(cur); _maktab_sinflari_jadvali(cur); _sinf_azolari_jadvali(cur)
        _markaz_jadvali(cur); _bogcha_jadvali(cur); _universitet_jadvali(cur)
        _tolov_jadvallari(cur)
        cur.execute("ALTER TABLE togaraklar ADD COLUMN IF NOT EXISTS sinf TEXT")
        cur.execute("ALTER TABLE togaraklar ADD COLUMN IF NOT EXISTS markaz_id INTEGER")
        cur.execute("ALTER TABLE togaraklar ADD COLUMN IF NOT EXISTS universitet_guruh_id INTEGER")
        cur.execute("""CREATE TABLE IF NOT EXISTS togarak_azolar(
            id SERIAL PRIMARY KEY, togarak_id INTEGER REFERENCES togaraklar(id),
            user_id BIGINT REFERENCES users(user_id), aktiv BOOLEAN DEFAULT TRUE, qoshilgan_at TIMESTAMP DEFAULT NOW()
        )""")
        cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS maktab_id INTEGER")
        cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS markaz_id INTEGER")
        cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS bogcha_id INTEGER")
        cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS universitet_id INTEGER")
        cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS lavozim TEXT")

        belgi = datetime.now().strftime("%H%M%S")
        keyingi_id = [_keyingi_manfiy_id(cur)]

        def yid():
            v = keyingi_id[0]
            keyingi_id[0] -= 1
            return v

        def parol4():
            return "".join(secrets.choice(string.digits) for _ in range(4))

        hisoblar = []   # "sifatida kirish" ro'yxati — rahbariyat + har birdan namuna
        sonlar = {}     # ommaviy yaratilgan talaba/o'quvchi/bola sonlari

        def rahbar_qosh(ism, lavozim_kaliti, izoh):
            uid = yid()
            cur.execute("INSERT INTO users(user_id, full_name, role, lavozim) VALUES(%s,%s,'oqituvchi',%s)", (uid, ism, lavozim_kaliti))
            hisoblar.append({"user_id": uid, "full_name": ism, "izoh": izoh})
            return uid

        # ═══════════════ 1) MAKTAB — ~550 o'quvchi, 1-11 sinf x A/B (22 sinf), pulli ═══════════════
        maktab_direktor = rahbar_qosh(f"Sinov Direktor {belgi}", "direktor", "Maktab direktori")
        cur.execute("""
            INSERT INTO maktablar(nomi, viloyat, tuman, smena_soni, direktor_user_id, pulli, oylik_tolov)
            VALUES(%s,%s,%s,1,%s,TRUE,500000) RETURNING id
        """, (f"Sinov Maktabi {belgi}", "Samarqand", "Samarqand shahri", maktab_direktor))
        maktab_id = cur.fetchone()["id"]
        cur.execute("UPDATE users SET maktab_id=%s WHERE user_id=%s", (maktab_id, maktab_direktor))

        JAMI_OQUVCHI = 550
        SINFLAR = [(str(s), h) for s in range(1, 12) for h in ("A", "B")]
        har_sinfga = JAMI_OQUVCHI // len(SINFLAR)
        joriy_oy = datetime.now().strftime("%Y-%m")
        bugun = datetime.now().date()

        maktab_talaba_q, sinf_azo_q, tolov_q = [], [], []
        for sinf, harf in SINFLAR:
            rahbar = rahbar_qosh(f"Sinov Ustoz {sinf}-{harf} {belgi}", "fan_oqituvchisi", f"{sinf}-{harf} sinf rahbari")
            cur.execute("UPDATE users SET maktab_id=%s WHERE user_id=%s", (maktab_id, rahbar))
            cur.execute("""
                INSERT INTO maktab_sinflari(maktab_id, sinf, harf, rahbar_user_id, qoshilish_paroli)
                VALUES(%s,%s,%s,%s,%s) RETURNING id
            """, (maktab_id, sinf, harf, rahbar, parol4()))
            sinf_id = cur.fetchone()["id"]
            for i in range(har_sinfga):
                uid = yid()
                maktab_talaba_q.append((uid, f"O'quvchi {sinf}-{harf}-{i+1} {belgi}", "oquvchi", maktab_id, sinf, harf))
                sinf_azo_q.append((sinf_id, uid))
                if i % 2 == 0:  # yarmi to'lagan, yarmi qarzdor — ikkalasini sinash uchun
                    tolov_q.append((uid, maktab_id, joriy_oy, 500000, 500000, bugun))

        psycopg2.extras.execute_values(cur, "INSERT INTO users(user_id, full_name, role, maktab_id, class, class_letter) VALUES %s", maktab_talaba_q)
        psycopg2.extras.execute_values(cur, "INSERT INTO maktab_sinf_azolari(sinf_id, user_id) VALUES %s", sinf_azo_q)
        if tolov_q:
            psycopg2.extras.execute_values(cur, "INSERT INTO tolovlar(user_id, maktab_id, oy, summa_kerak, tolangan_summa, tolov_sanasi) VALUES %s", tolov_q)
        sonlar["maktab_oquvchi"] = len(maktab_talaba_q)
        if maktab_talaba_q:
            n = maktab_talaba_q[0]
            hisoblar.append({"user_id": n[0], "full_name": n[1], "izoh": f"Namuna o'quvchi ({n[4]}-{n[5]})"})

        # ═══════════════ 2) BOG'CHA — ~75 bola + 75 ota-ona, 3 guruh ═══════════════
        bogcha_direktor = rahbar_qosh(f"Sinov BDirektor {belgi}", "bogcha_direktor", "Bog'cha direktori")
        cur.execute("""
            INSERT INTO bogchalar(nomi, turi, viloyat, tuman, direktor_user_id, oylik_tolov)
            VALUES(%s,'xususiy',%s,%s,%s,800000) RETURNING id
        """, (f"Sinov Bog'chasi {belgi}", "Samarqand", "Samarqand shahri", bogcha_direktor))
        bogcha_id = cur.fetchone()["id"]
        cur.execute("UPDATE users SET bogcha_id=%s WHERE user_id=%s", (bogcha_id, bogcha_direktor))

        BOGCHA_GURUHLAR = ["Kichkintoylar", "Quyoshcha", "Yulduzcha"]
        JAMI_BOLA = 75
        har_bguruhga = JAMI_BOLA // len(BOGCHA_GURUHLAR)

        bola_q, otaona_q, roster_q, parent_child_q = [], [], [], []
        for gi, guruh_nomi in enumerate(BOGCHA_GURUHLAR):
            opa = rahbar_qosh(f"Sinov Opa{gi+1} {belgi}", "bogcha_opa", f"{guruh_nomi} guruhi tarbiyachisi")
            cur.execute("UPDATE users SET bogcha_id=%s WHERE user_id=%s", (bogcha_id, opa))
            cur.execute("""
                INSERT INTO bogcha_guruhlari(bogcha_id, nomi, opa_user_id, qoshilish_paroli)
                VALUES(%s,%s,%s,%s) RETURNING id
            """, (bogcha_id, f"{guruh_nomi} guruhi", opa, parol4()))
            guruh_id = cur.fetchone()["id"]
            for i in range(har_bguruhga):
                bola_id, ota_id = yid(), yid()
                bola_q.append((bola_id, f"Bola {guruh_nomi}-{i+1} {belgi}", "oquvchi"))
                otaona_q.append((ota_id, f"OtaOna {guruh_nomi}-{i+1} {belgi}", "ota-ona"))
                roster_q.append((guruh_id, bola_id))
                parent_child_q.append((ota_id, bola_id))

        psycopg2.extras.execute_values(cur, "INSERT INTO users(user_id, full_name, role) VALUES %s", bola_q + otaona_q)
        psycopg2.extras.execute_values(cur, "INSERT INTO bogcha_guruh_bolalari(guruh_id, bola_user_id) VALUES %s", roster_q)
        psycopg2.extras.execute_values(cur, "INSERT INTO parent_child(parent_id, child_id) VALUES %s", parent_child_q)
        sonlar["bogcha_bola"] = len(bola_q)
        sonlar["bogcha_otaona"] = len(otaona_q)
        if bola_q:
            hisoblar.append({"user_id": otaona_q[0][0], "full_name": otaona_q[0][1], "izoh": f"Namuna ota-ona (farzandi: {bola_q[0][1]})"})

        # ═══════════════ 3) MARKAZ — ~150 talaba, 5 guruh ═══════════════
        markaz_direktor = rahbar_qosh(f"Sinov MDirektor {belgi}", "markaz_direktor", "Markaz direktori")
        cur.execute("""
            INSERT INTO oquv_markazlari(nomi, viloyat, tuman, direktor_user_id)
            VALUES(%s,%s,%s,%s) RETURNING id
        """, (f"Sinov Markazi {belgi}", "Samarqand", "Samarqand shahri", markaz_direktor))
        markaz_id = cur.fetchone()["id"]
        cur.execute("UPDATE users SET markaz_id=%s WHERE user_id=%s", (markaz_id, markaz_direktor))

        JAMI_MTALABA, MARKAZ_GURUH_SONI = 150, 5
        har_mguruhga = JAMI_MTALABA // MARKAZ_GURUH_SONI

        markaz_talaba_q, markaz_azo_q = [], []
        for gi in range(MARKAZ_GURUH_SONI):
            oq = rahbar_qosh(f"Sinov MOqituvchi{gi+1} {belgi}", "fan_oqituvchisi", f"Markaz {gi+1}-guruh o'qituvchisi")
            cur.execute("UPDATE users SET markaz_id=%s WHERE user_id=%s", (markaz_id, oq))
            cur.execute("""
                INSERT INTO togaraklar(nomi, fan, teacher_id, sinf, parol, max_talaba, oylik_summa, aktiv, markaz_id)
                VALUES(%s,'Matematika',%s,'5',%s,60,300000,TRUE,%s) RETURNING id
            """, (f"Sinov Guruh-{gi+1} {belgi}", oq, parol4(), markaz_id))
            togarak_id = cur.fetchone()["id"]
            for i in range(har_mguruhga):
                uid = yid()
                markaz_talaba_q.append((uid, f"MTalaba {gi+1}-{i+1} {belgi}", "oquvchi"))
                markaz_azo_q.append((togarak_id, uid, True))

        psycopg2.extras.execute_values(cur, "INSERT INTO users(user_id, full_name, role) VALUES %s", markaz_talaba_q)
        psycopg2.extras.execute_values(cur, "INSERT INTO togarak_azolar(togarak_id, user_id, aktiv) VALUES %s", markaz_azo_q)
        sonlar["markaz_talaba"] = len(markaz_talaba_q)
        if markaz_talaba_q:
            hisoblar.append({"user_id": markaz_talaba_q[0][0], "full_name": markaz_talaba_q[0][1], "izoh": "Namuna markaz talabasi"})

        # ═══════════════ 4) UNIVERSITET — ~1000 talaba, 6 fakultet x 2 kafedra x 3 guruh (36 guruh) ═══════════════
        rektor = rahbar_qosh(f"Sinov Rektor {belgi}", "rektor", "Rektor")
        cur.execute("""
            INSERT INTO universitetlar(nomi, viloyat, tuman, rektor_user_id)
            VALUES(%s,%s,%s,%s) RETURNING id
        """, (f"Sinov Universiteti {belgi}", "Samarqand", "Samarqand shahri", rektor))
        universitet_id = cur.fetchone()["id"]
        cur.execute("UPDATE users SET universitet_id=%s WHERE user_id=%s", (universitet_id, rektor))
        prorektor = rahbar_qosh(f"Sinov Prorektor {belgi}", "prorektor", "Prorektor")
        cur.execute("UPDATE users SET universitet_id=%s WHERE user_id=%s", (universitet_id, prorektor))

        JAMI_TALABA, FAKULTET_SONI, KAFEDRA_PER_FAKULTET, GURUH_PER_KAFEDRA = 1000, 6, 2, 3
        jami_guruh = FAKULTET_SONI * KAFEDRA_PER_FAKULTET * GURUH_PER_KAFEDRA
        har_guruhga = JAMI_TALABA // jami_guruh

        uni_talaba_q, uni_azo_q, uni_togarak_azo_q = [], [], []
        for fi in range(1, FAKULTET_SONI + 1):
            dekan = rahbar_qosh(f"Sinov Dekan-{fi} {belgi}", "dekan", f"{fi}-fakultet dekani")
            cur.execute("INSERT INTO fakultetlar(universitet_id, nomi, dekan_user_id) VALUES(%s,%s,%s) RETURNING id",
                        (universitet_id, f"Sinov Fakultet-{fi} {belgi}", dekan))
            fakultet_id = cur.fetchone()["id"]
            for ki in range(1, KAFEDRA_PER_FAKULTET + 1):
                mudir = rahbar_qosh(f"Sinov Mudir-{fi}.{ki} {belgi}", "kafedra_mudiri", f"{fi}.{ki}-kafedra mudiri")
                cur.execute("INSERT INTO kafedralar(fakultet_id, nomi, mudir_user_id) VALUES(%s,%s,%s) RETURNING id",
                            (fakultet_id, f"Sinov Kafedra-{fi}.{ki} {belgi}", mudir))
                kafedra_id = cur.fetchone()["id"]
                for gi in range(1, GURUH_PER_KAFEDRA + 1):
                    professor = rahbar_qosh(f"Sinov Professor-{fi}.{ki}.{gi} {belgi}", "professor_oqituvchi", f"{fi}.{ki}.{gi}-guruh kuratori")
                    cur.execute("UPDATE users SET universitet_id=%s WHERE user_id=%s", (universitet_id, professor))
                    cur.execute("""
                        INSERT INTO universitet_guruhlari(kafedra_id, nomi, kurs, yonalish, rahbar_user_id, qoshilish_paroli)
                        VALUES(%s,%s,%s,'Matematika',%s,%s) RETURNING id
                    """, (kafedra_id, f"{fi}{ki}{gi}-guruh", ((gi - 1) % 4) + 1, professor, parol4()))
                    uni_guruh_id = cur.fetchone()["id"]
                    cur.execute("""
                        INSERT INTO togaraklar(nomi, fan, teacher_id, sinf, parol, aktiv, universitet_guruh_id)
                        VALUES(%s,'Matematik tahlil',%s,%s,%s,TRUE,%s) RETURNING id
                    """, (f"Sinov Kurs-{fi}.{ki}.{gi} {belgi}", professor, f"{fi}{ki}{gi}-guruh", parol4(), uni_guruh_id))
                    uni_togarak_id = cur.fetchone()["id"]
                    for i in range(har_guruhga):
                        uid = yid()
                        uni_talaba_q.append((uid, f"Talaba {fi}.{ki}.{gi}-{i+1} {belgi}", "oquvchi"))
                        uni_azo_q.append((uni_guruh_id, uid))
                        uni_togarak_azo_q.append((uni_togarak_id, uid, True))

        psycopg2.extras.execute_values(cur, "INSERT INTO users(user_id, full_name, role) VALUES %s", uni_talaba_q)
        psycopg2.extras.execute_values(cur, "INSERT INTO universitet_guruh_azolari(guruh_id, user_id) VALUES %s", uni_azo_q)
        psycopg2.extras.execute_values(cur, "INSERT INTO togarak_azolar(togarak_id, user_id, aktiv) VALUES %s", uni_togarak_azo_q)
        sonlar["universitet_talaba"] = len(uni_talaba_q)
        if uni_talaba_q:
            hisoblar.append({"user_id": uni_talaba_q[0][0], "full_name": uni_talaba_q[0][1], "izoh": "Namuna talaba"})

        conn.commit()
    except Exception as e:
        conn.rollback()
        cur.close()
        conn.close()
        raise HTTPException(status_code=500, detail=f"Sinov muhitini yaratishda xato: {e}")

    cur.close()
    conn.close()
    return {
        "hisoblar": hisoblar,
        "sonlar": sonlar,
        "izoh": (
            f"Jami: {len(hisoblar)} ta rahbariyat/namuna hisob (pastda), "
            f"+ {sonlar.get('maktab_oquvchi',0)} maktab o'quvchisi, "
            f"{sonlar.get('bogcha_bola',0)} bog'cha bolasi (+{sonlar.get('bogcha_otaona',0)} ota-ona), "
            f"{sonlar.get('markaz_talaba',0)} markaz talabasi, "
            f"{sonlar.get('universitet_talaba',0)} universitet talabasi ommaviy yaratildi. "
            "Talaba/o'quvchi hisoblariga alohida kirish kerak bo'lsa — ID raqamini so'rang, alohida token beraman."
        ),
    }


@app.post("/api/admin/sifatida_kirish")
def admin_sifatida_kirish(token: str, user_id: int):
    """Admin — istalgan (odatda sinov) hisob sifatida DARHOL kirish
    uchun token oladi, Google login shart emas. 30 kun amal qiladi
    (uzoq muddatli sinov uchun). Faqat admin chaqira oladi."""
    _admin_tekshir(token)
    conn = _db()
    cur = conn.cursor()
    cur.execute("SELECT 1 FROM users WHERE user_id=%s", (user_id,))
    if not cur.fetchone():
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Foydalanuvchi topilmadi")
    cur.close()
    conn.close()
    return {"token": _sinov_jwt_yarat(user_id)}



class TestShablonGuruh(BaseModel):
    diff: str    # oson | o'rta | qiyin | murakkab
    turi: str    # single_choice | write_answer
    soni: int    # 0, 5, 10, 15, 20 ...


class TestShablonSorov(BaseModel):
    topic_codes: list[str]
    guruhlar: list[TestShablonGuruh]


_YOSH_GURUHI = {"1": "6-7", "2": "7-8", "3": "8-9", "4": "9-10", "5": "10-11",
                "6": "11-12", "7": "12-13", "8": "13-14", "9": "14-15", "10": "15-16", "11": "16-17"}


# ═══════════════════════════════════════════════════════════
# ADMIN — Topik mavzular (kontent auditi): qaysi mavzuda test
# bor, qaysisida yo'q — Sinf → Fan → Mavzu albomi
# ═══════════════════════════════════════════════════════════

@app.get("/api/admin/topik_sinflar")
def topik_sinflar(token: str):
    """dts_tree'da mavzusi yaratilgan barcha sinflar ro'yxati (oddiy va
    to'garak sinflari alohida-alohida qaytariladi)."""
    _admin_tekshir(token)
    conn = _db()
    cur = conn.cursor()
    cur.execute("SELECT DISTINCT grade FROM dts_tree WHERE is_deleted=FALSE")
    hammasi = [r["grade"] for r in cur.fetchall() if r["grade"]]
    cur.close()
    conn.close()
    oddiy = sorted([g for g in hammasi if g.isdigit()], key=int)
    togarak = sorted([g for g in hammasi if not g.isdigit()])
    return {"oddiy": oddiy, "togarak": togarak}


@app.get("/api/admin/topik_fanlar")
def topik_fanlar(sinf: str, token: str):
    """Berilgan sinfda mavzusi yaratilgan fanlar ro'yxati (test bor-yo'qligidan
    qat'i nazar — bu TEST bilan cheklanmagan, TO'LIQ kontent auditi)."""
    _admin_tekshir(token)
    conn = _db()
    cur = conn.cursor()
    cur.execute("""
        SELECT DISTINCT subject_name, COUNT(*) OVER (PARTITION BY subject_name) AS mavzu_soni
        FROM dts_tree WHERE grade=%s AND is_deleted=FALSE
    """, (sinf,))
    fanlar = [{"nom": r["subject_name"], "mavzu_soni": r["mavzu_soni"]} for r in cur.fetchall()]
    cur.close()
    conn.close()
    return {"fanlar": fanlar}


@app.get("/api/admin/topik_umumiy_korinish")
def topik_umumiy_korinish(token: str):
    """BARCHA sinf va fanlar bo'yicha bir zumda umumiy ko'rinish — har
    sinfga alohida kirmasdan, qaysi fanda nechta mavzu va shundan
    nechtasida test borligini BITTA so'rov bilan qaytaradi (admin
    "Umumiy ko'rinish" tugmasi uchun)."""
    _admin_tekshir(token)
    conn = _db()
    cur = conn.cursor()
    cur.execute("""
        WITH mavzu_guruhlari AS (
            SELECT d.grade, d.subject_name,
                   COALESCE(d.mavzu_name, d.bolim_name, d.bob_name) AS mavzu_nomi,
                   COUNT(DISTINCT gt.topic_code) > 0 AS test_bormi
            FROM dts_tree d
            LEFT JOIN generated_tests gt ON gt.topic_code = d.topic_code
            WHERE d.is_deleted = FALSE
            GROUP BY d.grade, d.subject_name, COALESCE(d.mavzu_name, d.bolim_name, d.bob_name)
        )
        SELECT grade, subject_name,
               COUNT(*) AS jami_mavzu,
               COUNT(*) FILTER (WHERE test_bormi) AS testli_mavzu
        FROM mavzu_guruhlari
        GROUP BY grade, subject_name
        ORDER BY grade, subject_name
    """)
    qatorlar = cur.fetchall()
    cur.close()
    conn.close()

    sinflar = {}
    for r in qatorlar:
        g = r["grade"]
        sinflar.setdefault(g, {"sinf": g, "fanlar": []})
        sinflar[g]["fanlar"].append({
            "nom": r["subject_name"], "jami_mavzu": r["jami_mavzu"], "testli_mavzu": r["testli_mavzu"],
        })
    natija = list(sinflar.values())
    natija.sort(key=lambda s: (0, int(s["sinf"])) if s["sinf"].isdigit() else (1, s["sinf"]))
    return {"sinflar": natija}


# ═══════════════════════════════════════════════════════════
# MAVZU TUSHUNTIRISHLARI — offlayn (Colab'da, tekin) tayyorlangan
# AI tushuntirishlarini saqlash va o'quvchiga ko'rsatish. Jonli AI
# SERVERI YO'Q — bu yerda faqat OLDINDAN yozilgan tushuntirish
# bazadan o'qiladi, shu sabab hech qanday qo'shimcha xarajat yo'q.
# ═══════════════════════════════════════════════════════════

def _tushuntirish_jadvali(cur):
    cur.execute("""CREATE TABLE IF NOT EXISTS mavzu_tushuntirishlari(
        sinf TEXT NOT NULL, fan TEXT NOT NULL, mavzu_nomi TEXT NOT NULL,
        tushuntirish TEXT NOT NULL, yaratilgan_at TIMESTAMP DEFAULT NOW(),
        PRIMARY KEY (sinf, fan, mavzu_nomi)
    )""")


@app.get("/api/mavzu_tushuntirish")
def mavzu_tushuntirish_ol(sinf: str, fan: str, mavzu: str):
    """O'quvchi (yoki har kim) uchun — berilgan mavzuning oldindan
    tayyorlangan AI tushuntirishini qaytaradi. Agar hali yozilmagan
    bo'lsa — topilmadi=true bilan bo'sh qaytadi (xato emas)."""
    conn = _db()
    cur = conn.cursor()
    _tushuntirish_jadvali(cur)
    cur.execute(
        "SELECT tushuntirish FROM mavzu_tushuntirishlari WHERE sinf=%s AND fan=%s AND mavzu_nomi=%s",
        (sinf, fan, mavzu),
    )
    r = cur.fetchone()
    cur.close()
    conn.close()
    if not r:
        return {"topildi": False, "tushuntirish": None}
    return {"topildi": True, "tushuntirish": r["tushuntirish"]}


@app.post("/api/admin/tushuntirish_import")
async def tushuntirish_import(token: str, fayl: UploadFile = File(...)):
    """Offlayn (Colab'da) tayyorlangan Excel faylni import qiladi —
    ustunlar: Sinf, Fan, Mavzu, Tushuntirish. Mavjud (sinf+fan+mavzu)
    yozuv bo'lsa — YANGILANADI (qayta generatsiya qilib yuklash mumkin)."""
    _admin_tekshir(token)
    import openpyxl
    import io

    content = await fayl.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel o'qib bo'lmadi: {e}")
    ws = wb.active

    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    kerakli = {"Sinf", "Fan", "Mavzu", "Tushuntirish"}
    if not kerakli.issubset(set(headers)):
        raise HTTPException(status_code=400, detail=f"Ustunlar mos emas — kerak: {', '.join(kerakli)}")
    idx = {h: i for i, h in enumerate(headers)}

    conn = _db()
    cur = conn.cursor()
    _tushuntirish_jadvali(cur)
    saqlandi, xato_soni = 0, 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) <= max(idx.values()):
            continue
        sinf, fan, mavzu, tushuntirish = row[idx["Sinf"]], row[idx["Fan"]], row[idx["Mavzu"]], row[idx["Tushuntirish"]]
        if not (sinf and fan and mavzu and tushuntirish):
            continue
        try:
            cur.execute("""
                INSERT INTO mavzu_tushuntirishlari(sinf, fan, mavzu_nomi, tushuntirish, yaratilgan_at)
                VALUES(%s,%s,%s,%s,NOW())
                ON CONFLICT (sinf, fan, mavzu_nomi) DO UPDATE SET
                    tushuntirish = EXCLUDED.tushuntirish, yaratilgan_at = NOW()
            """, (str(sinf).strip(), str(fan).strip(), str(mavzu).strip(), str(tushuntirish).strip()))
            conn.commit()
            saqlandi += 1
        except Exception:
            conn.rollback()
            xato_soni += 1
    cur.close()
    conn.close()
    return {"saqlandi": saqlandi, "xato": xato_soni}


@app.get("/api/admin/topik_royxat")
def topik_royxat(sinf: str, fan: str, token: str):
    """Berilgan sinf+fan uchun MAVZU darajasidagi (kichik mavzular
    birlashtirilgan) to'liq ro'yxat — har biriga chorak/bob/bo'lim,
    nechta kichik mavzu borligi, va ENG MUHIMI — shu mavzuga TEST
    borligi yoki YO'QLIGI (test_bormi) qo'shib qaytariladi."""
    _admin_tekshir(token)
    conn = _db()
    cur = conn.cursor()
    cur.execute("""
        SELECT COALESCE(d.mavzu_name, d.bolim_name, d.bob_name) AS nomi,
               MIN(d.topic_code) AS topic_code,
               array_agg(DISTINCT d.topic_code ORDER BY d.topic_code) AS barcha_kodlar,
               MIN(d.quarter) AS chorak, MIN(d.bob_name) AS bob, MIN(d.bolim_name) AS bolim,
               COUNT(*) AS kichik_soni,
               COUNT(DISTINCT gt.topic_code) AS test_bor_soni
        FROM dts_tree d
        LEFT JOIN generated_tests gt ON gt.topic_code = d.topic_code
        WHERE d.grade=%s AND d.subject_name=%s AND d.is_deleted=FALSE
        GROUP BY COALESCE(d.mavzu_name, d.bolim_name, d.bob_name)
        ORDER BY MIN(d.topic_code)
    """, (sinf, fan))
    qatorlar = cur.fetchall()
    cur.close()
    conn.close()
    mavzular = [{
        "nomi": r["nomi"], "topic_code": r["topic_code"], "topic_codes": r["barcha_kodlar"], "chorak": r["chorak"],
        "bob": r["bob"], "bolim": r["bolim"], "kichik_soni": r["kichik_soni"],
        "test_bormi": r["test_bor_soni"] > 0,
    } for r in qatorlar]
    return {"sinf": sinf, "fan": fan, "mavzular": mavzular}


@app.delete("/api/admin/mavzu_testlarini_ochir")
def mavzu_testlarini_ochir(token: str, topic_codes: str):
    """Berilgan mavzuga tegishli BARCHA kichik mavzularning testlarini
    o'chiradi. topic_codes — vergul bilan ajratilgan kodlar ro'yxati
    (bitta mavzuning barcha kichik mavzu kodlari)."""
    _admin_tekshir(token)
    kodlar = [k.strip() for k in topic_codes.split(",") if k.strip()]
    if not kodlar:
        raise HTTPException(status_code=400, detail="Mavzu kodi berilmagan")
    conn = _db()
    cur = conn.cursor()
    cur.execute("DELETE FROM generated_tests WHERE topic_code = ANY(%s)", (kodlar,))
    ochirilgan = cur.rowcount
    conn.commit()
    cur.close()
    conn.close()
    return {"holat": "ochirildi", "ochirilgan_soni": ochirilgan}


@app.delete("/api/admin/fan_testlarini_ochir")
def fan_testlarini_ochir(token: str, sinf: str, fan: str):
    """Berilgan sinf+fanga tegishli BARCHA mavzularning BARCHA testlarini
    o'chiradi — butun fan bo'yicha umumiy tozalash uchun."""
    _admin_tekshir(token)
    conn = _db()
    cur = conn.cursor()
    cur.execute("""
        DELETE FROM generated_tests WHERE topic_code IN (
            SELECT topic_code FROM dts_tree WHERE grade=%s AND subject_name=%s AND is_deleted=FALSE
        )
    """, (sinf, fan))
    ochirilgan = cur.rowcount
    conn.commit()
    cur.close()
    conn.close()
    return {"holat": "ochirildi", "ochirilgan_soni": ochirilgan}


@app.get("/api/admin/mavzu_rasmlari")
def mavzu_rasmlari(token: str, topic_codes: str):
    """Berilgan mavzu(lar)ning testlaridagi BARCHA rasm havolalarini
    (takrorlarsiz) qaytaradi — admin ularni ko'rib, to'g'ri yuklanganini
    tekshirishi uchun. LaTeX ifodalar ham shu ro'yxatga tushishi mumkin —
    frontend ularni /api/rasm orqali so'raganda tabiiy ravishda
    "topilmadi" chiqadi (bu — kutilgan holat, xato emas)."""
    _admin_tekshir(token)
    kodlar = [k.strip() for k in topic_codes.split(",") if k.strip()]
    if not kodlar:
        return {"rasmlar": []}
    conn = _db()
    cur = conn.cursor()
    cur.execute("""
        SELECT DISTINCT COALESCE(NULLIF(image_file_id, ''), image_url) AS rasm_id
        FROM generated_tests
        WHERE topic_code = ANY(%s) AND COALESCE(NULLIF(image_file_id, ''), image_url, '') != ''
    """, (kodlar,))
    rasmlar = [r["rasm_id"] for r in cur.fetchall()]
    cur.close()
    conn.close()
    return {"rasmlar": rasmlar}


@app.post("/api/admin/shablon_yukla")
def shablon_yukla(sorov: TestShablonSorov, token: str):
    """Tanlangan mavzular + har bir qiyinlik darajasi uchun tanlangan
    son/tur (tugmali yoki yozuvli) bo'yicha bo'sh Excel shablon yaratadi —
    UCH varaqli, haqiqiy namunaga (TESTLAR/MALUMOT/RASM_MALUMOTI) mos:
    - TESTLAR: to'ldiriladigan savollar
    - MALUMOT: tanlangan mavzular haqida ma'lumot (nazorat uchun)
    - RASM_MALUMOTI: har savolga tegishli rasm o'rni — description
      yozilsa, botdagi AI rasm generatori shu tavsif bo'yicha rasm
      yaratadi (yoki admin qo'lda kollaj orqali yuklaydi)."""
    _admin_tekshir(token)
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    import io
    from fastapi.responses import StreamingResponse

    kodlar = [k.strip() for k in sorov.topic_codes if k.strip()]
    if not kodlar:
        raise HTTPException(status_code=400, detail="Kamida bitta mavzu tanlang")
    guruhlar = [g for g in sorov.guruhlar if g.soni > 0]
    if not guruhlar:
        raise HTTPException(status_code=400, detail="Kamida bitta qiyinlik darajasidan son tanlang")

    conn = _db()
    cur = conn.cursor()
    cur.execute("""
        SELECT topic_code, grade, subject_name, quarter, bob_name, bolim_name,
               mavzu_name, kichik_name
        FROM dts_tree WHERE topic_code = ANY(%s) AND is_deleted=FALSE
    """, (kodlar,))
    tc_map = {r["topic_code"]: r for r in cur.fetchall()}
    cur.close()
    conn.close()

    wb = openpyxl.Workbook()

    # ═══ 1) TESTLAR — to'ldiriladigan savollar ═══
    ws = wb.active
    ws.title = "TESTLAR"
    testlar_ustunlari = [
        "topic_code", "difficulty", "situation", "question",
        "option_a", "option_b", "option_c", "option_d",
        "correct_answer", "explanation", "question_type", "is_latex",
        "image_url", "audio_text", "language", "life_level", "age_group", "time_limit",
    ]
    diff_colors = {"oson": "E2EFDA", "o'rta": "FFF2CC", "qiyin": "FCE4D6", "murakkab": "F2CEEF"}

    for col, h in enumerate(testlar_ustunlari, 1):
        cell = ws.cell(1, col, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center")

    rasm_qatorlari = []  # (image_id, topic_code) — RASM_MALUMOTI uchun
    row_num = 2
    for kod in kodlar:
        info = tc_map.get(kod)
        grade = str(info["grade"]) if info else ""
        age_group = _YOSH_GURUHI.get(grade, "")
        for g in guruhlar:
            color = diff_colors.get(g.diff, "F2F2F2")
            for i in range(1, g.soni + 1):
                image_id = f"{kod}-{i}"
                ws.cell(row_num, 1, kod)
                ws.cell(row_num, 2, g.diff)
                ws.cell(row_num, 3, "oddiy")
                ws.cell(row_num, 11, g.turi)
                ws.cell(row_num, 12, False)
                ws.cell(row_num, 13, image_id)
                ws.cell(row_num, 15, "uz")
                ws.cell(row_num, 16, 1)
                ws.cell(row_num, 17, age_group)
                ws.cell(row_num, 18, 60 if g.turi == "write_answer" else 55)
                for col in range(1, len(testlar_ustunlari) + 1):
                    ws.cell(row_num, col).fill = PatternFill("solid", fgColor=color)
                    ws.cell(row_num, col).alignment = Alignment(wrap_text=True)
                rasm_qatorlari.append((image_id, kod))
                row_num += 1

    widths = [22, 10, 10, 45, 18, 18, 18, 18, 15, 35, 15, 8, 22, 20, 8, 8, 8, 10]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = w

    # ═══ 2) MALUMOT — tanlangan mavzular haqida (faqat nazorat uchun, o'zgartirmang) ═══
    ws2 = wb.create_sheet("MALUMOT")
    malumot_ustunlari = ["#", "Topic code", "Sinf", "Fan", "Chorak", "Bob", "Bolim", "Mavzu", "Kichik mavzu", "Test soni"]
    for col, h in enumerate(malumot_ustunlari, 1):
        cell = ws2.cell(1, col, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="70AD47")
    for idx, kod in enumerate(kodlar, 1):
        info = tc_map.get(kod)
        jami_soni = sum(g.soni for g in guruhlar)
        ws2.append([
            idx, kod,
            str(info["grade"]) if info else "", info["subject_name"] if info else "",
            info["quarter"] if info else "", info["bob_name"] if info else "",
            info["bolim_name"] if info else "", info["mavzu_name"] if info else "",
            info["kichik_name"] if info else "", jami_soni,
        ])
    for col, w in zip(range(1, 11), [5, 22, 6, 16, 8, 30, 30, 22, 30, 10]):
        ws2.column_dimensions[ws2.cell(1, col).column_letter].width = w

    # ═══ 3) RASM_MALUMOTI — har savol-rasm juftligi uchun tavsif ═══
    ws3 = wb.create_sheet("RASM_MALUMOTI")
    for col, h in enumerate(["image_id", "topic_code", "image_description"], 1):
        cell = ws3.cell(1, col, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="ED7D31")
    for image_id, kod in rasm_qatorlari:
        ws3.append([image_id, kod, ""])
    for col, w in zip(range(1, 4), [26, 22, 55]):
        ws3.column_dimensions[ws3.cell(1, col).column_letter].width = w
    ws3.cell(1, 4, "☝️ Har qatorga rasmda NIMA bo'lishi kerakligini yozing — botdagi AI rasm generatori shu tavsif bo'yicha rasm yaratadi. Rasm kerak bo'lmagan savollar uchun qatorni o'chiring.").font = Font(italic=True, color="8A8578")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=test_shablon.xlsx"},
    )


@app.post("/api/admin/shablon_import")
async def shablon_import(token: str, fayl: UploadFile = File(...)):
    """To'ldirilgan Excel shablonni import qiladi — botning
    import_tests_excel funksiyasidagi duplikat-tekshiruvi bilan bir xil."""
    _admin_tekshir(token)
    import openpyxl
    import io

    content = await fayl.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel o'qib bo'lmadi: {e}")

    ws = wb["TESTLAR"] if "TESTLAR" in wb.sheetnames else wb.active
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    if "topic_code" not in headers:
        raise HTTPException(status_code=400, detail="Excel ustunlari mos emas — 'topic_code' topilmadi")

    conn = _db()
    cur = conn.cursor()
    saved = 0
    duplicates = 0
    errors = 0

    for row in ws.iter_rows(min_row=2):
        d = {headers[i]: cell.value for i, cell in enumerate(row) if i < len(headers) and headers[i]}
        tc = d.get("topic_code")
        q = d.get("question")
        if not tc or not q or str(tc).strip() == "" or str(q).strip() == "":
            continue
        try:
            tc_s = str(tc).strip()
            q_s = str(q).strip()
            opt_a = str(d.get("option_a") or "").strip()
            correct = str(d.get("correct_answer") or "").strip()

            cur.execute("""
                SELECT 1 FROM generated_tests
                WHERE topic_code=%s AND question=%s AND option_a=%s AND correct_answer=%s
                LIMIT 1
            """, (tc_s, q_s, opt_a, correct))
            if cur.fetchone():
                duplicates += 1
                continue

            cur.execute("""
                INSERT INTO generated_tests
                (topic_code, difficulty, situation, question, option_a, option_b, option_c, option_d,
                 correct_answer, explanation, question_type, is_latex, image_url, audio_text,
                 language, life_level, age_group, time_limit)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (
                tc_s, d.get("difficulty"), d.get("situation") or "oddiy", q_s,
                d.get("option_a"), d.get("option_b"), d.get("option_c"), d.get("option_d"),
                d.get("correct_answer"), d.get("explanation"),
                d.get("question_type") or "single_choice",
                bool(d.get("is_latex")) if d.get("is_latex") not in (None, "") else False,
                d.get("image_url"), d.get("audio_text"), d.get("language") or "uz",
                d.get("life_level") or 1, d.get("age_group"), d.get("time_limit") or 60,
            ))
            conn.commit()
            saved += 1
        except Exception as e:
            conn.rollback()
            errors += 1

    cur.close()
    conn.close()
    return {"saved": saved, "duplicates": duplicates, "errors": errors}


# ═══════════════════════════════════════════════════════════
# ADMIN — Topik shablon (dts_tree uchun) yuklab olish va import qilish
# Botdagi shablon_yaratish.py (_create_shablon / handle_shablon_document)
# mantig'iga mos
# ═══════════════════════════════════════════════════════════

class TopikShablonSorov(BaseModel):
    sinf: str
    fan: str
    mavzular: str  # ko'p qatorli matn: "1 / Colours\n1 / Numbers\n2 / Animals"


def _mavzularni_parse(text: str):
    """Botdagi bilan bir xil parser: 'chorak / mavzu' yoki 'chorak mavzu' formatini o'qiydi."""
    natija = []
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        if "/" in line:
            parts = line.split("/", 1)
            chorak_raqam = "".join(ch for ch in parts[0].strip() if ch.isdigit())
            mavzu = parts[1].strip()
        else:
            parts = line.split(None, 1)
            chorak_raqam = parts[0].strip() if parts else "1"
            mavzu = parts[1].strip() if len(parts) > 1 else line
        if mavzu and chorak_raqam:
            natija.append((chorak_raqam, mavzu))
    return natija


@app.post("/api/admin/topik_shablon")
def topik_shablon(sorov: TopikShablonSorov, token: str):
    """Sinf + fan + mavzular ro'yxati bo'yicha DTS (topik kod) shablonini
    Excel qilib yaratadi — MALUMOT varag'i ko'rinishida (haqiqiy namunaga mos)."""
    _admin_tekshir(token)
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    import io
    from fastapi.responses import StreamingResponse

    mavzular = _mavzularni_parse(sorov.mavzular)
    if not mavzular:
        raise HTTPException(status_code=400, detail="Mavzular topilmadi — 'chorak / mavzu' formatida yozing")

    sinf, fan = sorov.sinf.strip(), sorov.fan.strip()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MALUMOT"

    headers = ["#", "Topic code", "Sinf", "Fan", "Chorak", "Bob", "Bolim", "Mavzu", "Kichik mavzu", "Test soni"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="70AD47")
        cell.alignment = Alignment(horizontal="center")

    chorak_colors = {"1": "DEEAF1", "2": "E2EFDA", "3": "FFF2CC", "4": "FCE4D6"}
    row_num = 2
    idx = 1
    for chorak, mavzu in mavzular:
        color = chorak_colors.get(str(chorak), "F2F2F2")
        for _ in range(2):  # botdagi kabi mavzu boshiga 2 qator (Bob/Bolim/Kichik mavzu uchun 2 xil variant)
            ws.cell(row_num, 1, value=idx)
            # "Topic code" ATAYLAB BO'SH qoldiriladi — import qilinganda avtomatik yaratiladi
            ws.cell(row_num, 3, value=sinf)
            ws.cell(row_num, 4, value=fan)
            ws.cell(row_num, 5, value=chorak)
            ws.cell(row_num, 8, value=mavzu)
            ws.cell(row_num, 10, value=0)
            for col in range(1, 11):
                ws.cell(row_num, col).fill = PatternFill("solid", fgColor=color)
                ws.cell(row_num, col).alignment = Alignment(horizontal="left", wrap_text=True)
            row_num += 1
            idx += 1

    for col, width in zip(range(1, 11), [5, 22, 6, 16, 8, 30, 30, 22, 30, 10]):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = width

    ws2 = wb.create_sheet("IZOH")
    ws2.cell(1, 1, value="📋 TO'LDIRISH QO'LLANMASI").font = Font(bold=True, size=14)
    izohlar = [
        (3, "#", "O'zgartirmang"),
        (4, "Topic code", "BO'SH QOLDIRING — import qilinganda avtomatik yaratiladi"),
        (5, "Sinf / Fan / Chorak", "O'zgartirmang — avtomatik to'ldirilgan"),
        (6, "Bob", "To'ldiring: masalan '1-bob. Sonlar'"),
        (7, "Bolim", "To'ldiring: masalan \"1-bo'lim. Narsalarning to'plamlari\""),
        (8, "Mavzu", "O'zgartirmang — mavzu nomi avtomatik"),
        (9, "Kichik mavzu", "To'ldiring: mavzuning kichik qismi"),
        (10, "Test soni", "O'zgartirmang — 0, keyin avtomatik yangilanadi"),
    ]
    for r, ustun, izoh in izohlar:
        ws2.cell(r, 1, value=ustun).font = Font(bold=True)
        ws2.cell(r, 2, value=izoh)
    ws2.column_dimensions['A'].width = 18
    ws2.column_dimensions['B'].width = 55

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"shablon_{sinf}sinf_{fan.replace(' ', '_')[:20]}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


@app.post("/api/admin/topik_import")
async def topik_import(token: str, fayl: UploadFile = File(...)):
    """To'ldirilgan Topik (MALUMOT) shablonini dts_tree jadvaliga import
    qiladi. "Topic code" ustuni bo'sh bo'lsa avtomatik yaratiladi, to'ldirilgan
    bo'lsa — AYNAN o'sha kod bilan saqlanadi (mavjud mavzuni yangilash uchun)."""
    _admin_tekshir(token)
    import openpyxl
    import io

    content = await fayl.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel o'qib bo'lmadi: {e}")
    ws = wb["MALUMOT"] if "MALUMOT" in wb.sheetnames else wb.active

    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    # Eski (DTS_SHABLON) va yangi (MALUMOT) formatlarini ikkalasini ham qo'llab-quvvatlaymiz
    eski_format = "Sinf" in headers and headers[0] == "Sinf"

    conn = _db()
    cur = conn.cursor()
    added, skipped = 0, 0

    for r in range(2, ws.max_row + 1):
        if eski_format:
            berilgan_kod = None
            sinf, fan, chorak, bob, bolim, mavzu, kichik = (ws.cell(r, c).value for c in range(1, 8))
        else:
            berilgan_kod = ws.cell(r, 2).value
            sinf, fan, chorak, bob, bolim, mavzu, kichik = (ws.cell(r, c).value for c in range(3, 10))

        if not sinf or not mavzu:
            continue

        if berilgan_kod and str(berilgan_kod).strip():
            topic_code = str(berilgan_kod).strip()
        else:
            cur.execute("""
                SELECT topic_code FROM dts_tree
                WHERE grade=%s AND subject_name=%s
                ORDER BY topic_code DESC LIMIT 1
            """, (str(sinf), str(fan) if fan else ""))
            row = cur.fetchone()
            if row:
                last = row["topic_code"]
                parts = last.rsplit('-', 1)
                new_num = str(int(parts[1]) + 1).zfill(3)
                topic_code = f"{parts[0]}-{new_num}"
            else:
                # MUHIM: bu fan uchun BIRINCHI marta mavzu qo'shilyapti — "01"ni
                # QATTIQ KODLAMAYMIZ, chunki boshqa fan allaqachon "01"ni band
                # qilgan bo'lishi mumkin (aks holda ikkala fan bitta topic_code'ga
                # to'qnashib, natijalar noto'g'ri fanga yozilib qoladi).
                # Shu grade uchun band qilingan barcha fan-segmentlarini
                # tekshirib, BO'SH birinchi raqamni tanlaymiz.
                cur.execute(
                    "SELECT DISTINCT SPLIT_PART(topic_code, '-', 2) AS seg FROM dts_tree WHERE grade=%s",
                    (str(sinf),),
                )
                band_segmentlar = {r2["seg"] for r2 in cur.fetchall()}
                fan_segmenti = "01"
                for n in range(1, 100):
                    nomzod = str(n).zfill(2)
                    if nomzod not in band_segmentlar:
                        fan_segmenti = nomzod
                        break
                topic_code = f"{sinf}-{fan_segmenti}-{chorak or 1}-01-01-01-001"

        try:
            cur.execute("""
                INSERT INTO dts_tree
                (topic_code, grade, subject_name, quarter,
                 bob_name, bolim_name, mavzu_name, kichik_name, is_deleted)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,FALSE)
                ON CONFLICT (topic_code) DO UPDATE SET
                    bob_name = EXCLUDED.bob_name, bolim_name = EXCLUDED.bolim_name,
                    kichik_name = EXCLUDED.kichik_name
            """, (
                topic_code, str(sinf), str(fan) if fan else "",
                str(chorak) if chorak else "1", str(bob) if bob else "",
                str(bolim) if bolim else "", str(mavzu) if mavzu else "",
                str(kichik) if kichik else "",
            ))
            conn.commit()
            added += 1
        except Exception:
            conn.rollback()
            skipped += 1

    cur.close()
    conn.close()
    return {"added": added, "skipped": skipped}
