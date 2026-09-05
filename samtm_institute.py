"""SamTM Institute V20.

Institut -> fakultet -> kafedra -> ta'lim yo'nalishi ierarxiyasi,
lavozimga asoslangan ruxsatlar va HEMISgacha bo'lgan qabul kontingenti.

Muhim tamoyillar:
- import avval preview qilinadi, keyin bitta tranzaksiyada commit bo'ladi;
- talabaning maxfiy ma'lumoti faqat vakolatli rolga beriladi;
- xodim/talaba uchun doim bir martalik, 2 oylik kirish kodi beriladi;
- qabulda dublikat universitet + JSHSHIR bo'yicha bloklanadi/upsert qilinadi.
"""
from __future__ import annotations

import base64
import hashlib
import io
import json
import math
import re
import secrets
import string
import unicodedata
from datetime import date, datetime, timedelta
from difflib import SequenceMatcher
from typing import Any, Optional

from cryptography.fernet import Fernet, InvalidToken
from fastapi import APIRouter, File, Header, HTTPException, Query, UploadFile
from fastapi.concurrency import run_in_threadpool
from fastapi.responses import Response
from pydantic import BaseModel, Field


SAMTM_INSTITUTE_RELEASE = "institute-workspace-isolation-hierarchy-reports-v24-rev87"
router = APIRouter(prefix="/api/institut/v20", tags=["Institut V20"])
PLATFORM = None
_SCHEMA_READY = False
_V17_SOURCE_TABLES_READY: Optional[bool] = None

TA_LIM_SHAKLLARI = ["Kunduzgi", "Kechki", "Sirtqi", "Masofaviy", "Dual ta'lim"]
TA_LIM_TILLARI = ["O‘zbekcha", "Ruscha", "Tojikcha", "Qoraqalpoqcha", "Inglizcha"]
DARAJALAR = ["Bakalavriat", "Magistratura", "Doktorantura"]

ROLE_LABELS = {
    "owner": "Institut egasi",
    "rektor": "Rektor",
    "prorektor": "Prorektor",
    "institut_admin": "Institut administratori",
    "qabul_operator": "Qabul operatori",
    "dekan": "Dekan",
    "zam_dekan": "Dekan o‘rinbosari",
    "manaviyatchi": "Ma’naviy-ma’rifiy ishlar mas’uli",
    "fakultet_admin": "Fakultet administratori",
    "kafedra_mudiri": "Kafedra mudiri",
    "professor_oqituvchi": "Professor-o‘qituvchi",
    "tyutor": "Tyutor",
    "talaba": "Talaba",
}

INSTITUTE_WIDE = {"owner", "rektor", "prorektor", "institut_admin"}
FACULTY_WIDE = {"dekan", "zam_dekan", "manaviyatchi", "fakultet_admin"}
DEPARTMENT_WIDE = {"kafedra_mudiri"}
# Qabul vakolati tuzilma/xodim vakolatidan ataylab ajratilgan. Masalan,
# qabul operatorini INSTITUTE_WIDE ga qo'shish unga tuzilmani tahrirlash
# huquqini ham berib yuboradi.
ADMISSION_VIEW_ROLES = (
    INSTITUTE_WIDE
    | {"fakultet_admin", "qabul_operator", "dekan", "zam_dekan"}
    | DEPARTMENT_WIDE
    | {"tyutor"}
)
PRIVATE_ROLES = ADMISSION_VIEW_ROLES
MANAGE_STRUCTURE_ROLES = INSTITUTE_WIDE
MANAGE_STAFF_ROLES = INSTITUTE_WIDE | {"dekan", "fakultet_admin"}
ADMISSION_IMPORT_ROLES = {"institut_admin", "fakultet_admin", "qabul_operator"}
MARK_DOCUMENT_ROLES = ADMISSION_IMPORT_ROLES
MARK_HEMIS_ROLES = {"kafedra_mudiri", "tyutor"}
# Eski nomni import qiladigan boshqa modul bo'lsa, huquq kengayib ketmasdan
# yangi HEMIS qoidalariga yo'naltiriladi.
MARK_DATABASE_ROLES = MARK_HEMIS_ROLES
REPORT_VIEW_ROLES = ADMISSION_VIEW_ROLES
PASSWORD_VIEW_ROLES = {"institut_admin"}
ADMIN_ROLES = {"institut_admin", "fakultet_admin"}


def register_institute(app, platform):
    global PLATFORM
    PLATFORM = platform
    app.include_router(router)

    @app.on_event("startup")
    def migrate_institute_v20():
        global _SCHEMA_READY
        conn = platform._db(); cur = conn.cursor()
        try:
            cur.execute("SELECT pg_advisory_lock(%s)", (20005400,))
            _institut_v20_jadvallari(cur)
            conn.commit(); _SCHEMA_READY = True
        finally:
            try:
                cur.execute("SELECT pg_advisory_unlock(%s)", (20005400,)); conn.commit()
            except Exception:
                conn.rollback()
            cur.close(); conn.close()


def _p():
    if PLATFORM is None:
        raise RuntimeError("samtm_institute.register_institute chaqirilmagan")
    return PLATFORM


def _token(token: Optional[str], authorization: Optional[str]) -> str:
    return _p()._jwt_header_yoki_query(token, authorization)


def _uid(token: Optional[str], authorization: Optional[str]) -> int:
    return _p()._jwt_tekshir(_token(token, authorization))


def _norm(value: Any) -> str:
    text = unicodedata.normalize("NFKC", str(value or "")).strip()
    text = text.replace("ʻ", "'").replace("’", "'").replace("‘", "'").replace("`", "'")
    return re.sub(r"\s+", " ", text).strip()


def _key(value: Any) -> str:
    text = _norm(value).casefold()
    text = "".join(ch for ch in unicodedata.normalize("NFKD", text) if not unicodedata.combining(ch))
    return re.sub(r"[^a-z0-9а-яё]+", "", text)


_PROGRAM_STOPWORDS = {
    "talim", "yonalishi", "yonalish", "kafedrasi", "kafedra", "fakulteti", "fakultet",
    "bakalavriat", "kunduzgi", "kechki", "sirtqi", "dual", "tumani", "tuman",
}


def _match_words(value: Any) -> list[str]:
    text = _norm(value).casefold()
    text = "".join(ch for ch in unicodedata.normalize("NFKD", text) if not unicodedata.combining(ch))
    words = re.findall(r"[a-z0-9а-яё]+", text)
    return [word for word in words if word not in _PROGRAM_STOPWORDS]


def _base_program_name(value: Any) -> str:
    """Qavsdagi hudud/shakl izohini olib, asosiy yo'nalish nomini qaytaradi."""
    return _norm(re.sub(r"\([^)]*\)", " ", str(value or "")))


def _admission_program_name(value: Any) -> str:
    """Qabul faylidagi hududli yozuvni bitta asosiy yo'nalishga yig'adi.

    Masalan, ``Boshlang'ich ta'lim (Qo'shrabot tumani)`` alohida
    yo'nalish emas: ``Boshlang'ich ta'lim`` ichida, o'z shakli va tili
    bilan saqlanadi. Boshqa mazmunli qavslar avtomatik olib tashlanmaydi.
    """
    text = _norm(value)
    return _norm(re.sub(
        r"\((?=[^)]*\b(?:tumani|shahri|viloyati|hududi)\b)[^)]*\)",
        " ",
        text,
        flags=re.IGNORECASE,
    ))


def _name_similarity(source: Any, target: Any) -> int:
    """Qisqartma, apostrof va 1–2 harf xatosiga chidamli 0..100 ball."""
    source_key, target_key = _key(source), _key(target)
    if not source_key or not target_key:
        return 0
    if source_key == target_key:
        return 100
    source_base, target_base = _key(_base_program_name(source)), _key(_base_program_name(target))
    if source_base and source_base == target_base:
        return 96
    ratio = SequenceMatcher(None, source_key, target_key).ratio()
    source_words, target_words = set(_match_words(source)), set(_match_words(target))
    union = source_words | target_words
    token_score = len(source_words & target_words) / len(union) if union else 0.0
    contains = 0.92 if min(len(source_key), len(target_key)) >= 6 and (source_key in target_key or target_key in source_key) else 0.0
    acronym = "".join(word[0] for word in _match_words(target) if word)
    acronym_score = 0.90 if len(source_key) >= 2 and source_key == acronym else 0.0
    return int(round(100 * max(ratio, token_score, contains, acronym_score)))


def _rank_programs(name: str, rows: list[dict[str, Any]], limit: int = 5) -> list[dict[str, Any]]:
    ranked = []
    for row in rows:
        name_score = _name_similarity(name, row.get("nomi"))
        code_score = 100 if row.get("kodi") and _key(name) == _key(row.get("kodi")) else 0
        score = max(name_score, code_score)
        ranked.append({
            "id": int(row["id"]), "nomi": row["nomi"], "kodi": row.get("kodi"),
            "fakultet_nomi": row.get("fakultet_nomi"), "kafedra_nomi": row.get("kafedra_nomi"),
            "kafedra_id": int(row["kafedra_id"]), "moslik_foizi": score,
        })
    return sorted(ranked, key=lambda item: (-item["moslik_foizi"], item["nomi"]))[:limit]


def _rank_departments(name: str, rows: list[dict[str, Any]], limit: int = 5) -> list[dict[str, Any]]:
    ranked = [{
        "id": int(row["id"]), "nomi": row["nomi"], "fakultet_nomi": row.get("fakultet_nomi"),
        "moslik_foizi": _name_similarity(name, row["nomi"]),
    } for row in rows]
    return sorted(ranked, key=lambda item: (-item["moslik_foizi"], item["nomi"]))[:limit]


def _digits(value: Any) -> str:
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return re.sub(r"\D", "", str(value or ""))


def _text_number(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return _norm(value)


def _float(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    try:
        return float(str(value).replace(" ", "").replace(",", "."))
    except (TypeError, ValueError):
        return None


def _iso_date(value: Any) -> Optional[str]:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = _norm(value)
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text[:10], fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _telefon(value: Any) -> Optional[str]:
    d = _digits(value)
    if len(d) == 9:
        return "+998" + d
    if len(d) == 12 and d.startswith("998"):
        return "+" + d
    return None


def _mask_phone(value: Optional[str]) -> Optional[str]:
    d = _digits(value)
    if len(d) < 7:
        return None
    return f"+{d[:3]} ** *** {d[-2:]}"


def _mask_pin(value: Optional[str]) -> Optional[str]:
    d = _digits(value)
    if len(d) < 6:
        return None
    return d[:2] + "**********" + d[-2:]


def _canonical_choice(value: Any, choices: list[str]) -> Optional[str]:
    k = _key(value)
    aliases = {
        "kunduzgi": "Kunduzgi", "kechki": "Kechki", "sirtqi": "Sirtqi",
        "masofaviy": "Masofaviy", "dualtalim": "Dual ta'lim", "dual": "Dual ta'lim",
        "ozbekcha": "O‘zbekcha", "uzbekcha": "O‘zbekcha", "uzbek": "O‘zbekcha",
        "ruscha": "Ruscha", "russian": "Ruscha", "tojikcha": "Tojikcha",
        "qoraqalpoqcha": "Qoraqalpoqcha", "inglizcha": "Inglizcha",
        "bakalavriat": "Bakalavriat", "magistratura": "Magistratura", "doktorantura": "Doktorantura",
    }
    result = aliases.get(k)
    return result if result in choices else None


def _header_map(row: list[Any]) -> dict[str, int]:
    return {_key(value): i for i, value in enumerate(row) if _norm(value)}


def _find_header_row(rows: list[list[Any]], required: list[str], sheet_name: str) -> tuple[int, dict[str, int]]:
    """Bezak/sarlavha qatorlari bo'lsa ham haqiqiy ustun qatorini topadi."""
    wanted = {_key(value) for value in required}
    for index, row in enumerate(rows[:25]):
        headers = _header_map(row)
        if wanted.issubset(headers):
            return index, headers
    raise HTTPException(
        status_code=400,
        detail=f"{sheet_name} varag'ida ustunlar topilmadi: " + ", ".join(required),
    )


def _cell(row: list[Any], headers: dict[str, int], *names: str) -> Any:
    for name in names:
        idx = headers.get(_key(name))
        if idx is not None and idx < len(row):
            return row[idx]
    return None


def _workbook_rows(content: bytes, filename: str) -> dict[str, list[list[Any]]]:
    """XLS va XLSX ni formulalarsiz, ixcham 2D qatorlarga o'qiydi."""
    lower = (filename or "").lower()
    if lower.endswith(".xls") and not lower.endswith(".xlsx"):
        try:
            import xlrd
        except ImportError as exc:
            raise HTTPException(status_code=500, detail=".xls o'qish uchun xlrd o'rnatilmagan") from exc
        try:
            book = xlrd.open_workbook(file_contents=content)
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"XLS fayl ochilmadi: {exc}") from exc
        result = {}
        for sheet in book.sheets():
            rows = []
            for r in range(sheet.nrows):
                values = []
                for c in range(sheet.ncols):
                    cell = sheet.cell(r, c)
                    value = cell.value
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        try:
                            value = datetime(*xlrd.xldate_as_tuple(value, book.datemode))
                        except Exception:
                            pass
                    values.append(value)
                rows.append(values)
            result[sheet.name] = rows
        return result
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"XLSX fayl ochilmadi: {exc}") from exc
    result = {ws.title: [list(r) for r in ws.iter_rows(values_only=True)] for ws in wb.worksheets}
    wb.close()
    return result


def _active_rows(sheets: dict[str, list[list[Any]]], preferred: tuple[str, ...] = ()) -> list[list[Any]]:
    for wanted in preferred:
        for name, rows in sheets.items():
            if _key(name) == _key(wanted):
                return rows
    return next(iter(sheets.values()), [])


def _named_rows(sheets: dict[str, list[list[Any]]], wanted: str) -> list[list[Any]]:
    for name, rows in sheets.items():
        if _key(name) == _key(wanted):
            return rows
    return []


def _institut_v20_jadvallari(cur):
    """V20 sxemasi. Runtime startupda bir marta bajaradi."""
    if PLATFORM is not None:
        PLATFORM._universitet_jadvali(cur)
    cur.execute("""CREATE TABLE IF NOT EXISTS universitet_yonalishlari(
        id BIGSERIAL PRIMARY KEY,
        universitet_id INTEGER NOT NULL REFERENCES universitetlar(id) ON DELETE CASCADE,
        fakultet_id INTEGER NOT NULL REFERENCES fakultetlar(id) ON DELETE CASCADE,
        kafedra_id INTEGER NOT NULL REFERENCES kafedralar(id) ON DELETE CASCADE,
        kodi TEXT, nomi TEXT NOT NULL, daraja TEXT NOT NULL DEFAULT 'Bakalavriat',
        faol BOOLEAN NOT NULL DEFAULT TRUE,
        yaratilgan_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
        UNIQUE(universitet_id,kafedra_id,nomi,daraja)
    )""")
    cur.execute("""CREATE TABLE IF NOT EXISTS universitet_yonalish_variantlari(
        id BIGSERIAL PRIMARY KEY,
        yonalish_id BIGINT NOT NULL REFERENCES universitet_yonalishlari(id) ON DELETE CASCADE,
        talim_shakli TEXT NOT NULL, talim_tili TEXT NOT NULL,
        faol BOOLEAN NOT NULL DEFAULT TRUE,
        UNIQUE(yonalish_id,talim_shakli,talim_tili)
    )""")
    cur.execute("""CREATE TABLE IF NOT EXISTS universitet_xodim_rollari(
        id BIGSERIAL PRIMARY KEY,
        universitet_id INTEGER NOT NULL REFERENCES universitetlar(id) ON DELETE CASCADE,
        fakultet_id INTEGER REFERENCES fakultetlar(id) ON DELETE CASCADE,
        kafedra_id INTEGER REFERENCES kafedralar(id) ON DELETE CASCADE,
        yonalish_id BIGINT REFERENCES universitet_yonalishlari(id) ON DELETE CASCADE,
        user_id BIGINT NOT NULL REFERENCES users(user_id) ON DELETE CASCADE,
        rol TEXT NOT NULL,
        faol BOOLEAN NOT NULL DEFAULT TRUE,
        yaratilgan_by BIGINT REFERENCES users(user_id) ON DELETE SET NULL,
        yaratilgan_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
        UNIQUE(universitet_id,user_id,rol,fakultet_id,kafedra_id,yonalish_id)
    )""")
    cur.execute("""CREATE UNIQUE INDEX IF NOT EXISTS uq_uni_fakultet_dekan
        ON universitet_xodim_rollari(fakultet_id)
        WHERE faol=TRUE AND rol='dekan'""")
    cur.execute("""CREATE UNIQUE INDEX IF NOT EXISTS uq_uni_fakultet_manaviy
        ON universitet_xodim_rollari(fakultet_id)
        WHERE faol=TRUE AND rol='manaviyatchi'""")
    cur.execute("""CREATE UNIQUE INDEX IF NOT EXISTS uq_uni_kafedra_mudir
        ON universitet_xodim_rollari(kafedra_id)
        WHERE faol=TRUE AND rol='kafedra_mudiri'""")
    cur.execute("""CREATE TABLE IF NOT EXISTS universitet_xodim_profili(
        xodim_rol_id BIGINT PRIMARY KEY REFERENCES universitet_xodim_rollari(id) ON DELETE CASCADE,
        ilmiy_daraja TEXT, ilmiy_unvon TEXT, staj_yil NUMERIC(5,1),
        mutaxassislik TEXT, qisqa_izoh TEXT,
        yangilangan_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
    )""")
    # REV57: eski Admin → Muassasalar ekranidagi rahbar ustunlari va yangi
    # rol jadvali ikki tomonga sinxron bo'lsin. Shu ko'prik bo'lmasa xodim
    # import qilingan bo'lsa ham eski ekranda "belgilanmagan" qizil chiqadi.
    cur.execute("""INSERT INTO universitet_xodim_rollari(universitet_id,user_id,rol)
        SELECT u.id,u.rektor_user_id,'rektor' FROM universitetlar u
        WHERE u.rektor_user_id IS NOT NULL AND NOT EXISTS(
          SELECT 1 FROM universitet_xodim_rollari xr
          WHERE xr.universitet_id=u.id AND xr.rol='rektor' AND xr.faol=TRUE)""")
    cur.execute("""INSERT INTO universitet_xodim_rollari(universitet_id,fakultet_id,user_id,rol)
        SELECT f.universitet_id,f.id,f.dekan_user_id,'dekan' FROM fakultetlar f
        WHERE f.dekan_user_id IS NOT NULL AND NOT EXISTS(
          SELECT 1 FROM universitet_xodim_rollari xr
          WHERE xr.fakultet_id=f.id AND xr.rol='dekan' AND xr.faol=TRUE)""")
    cur.execute("""INSERT INTO universitet_xodim_rollari(universitet_id,fakultet_id,kafedra_id,user_id,rol)
        SELECT f.universitet_id,f.id,k.id,k.mudir_user_id,'kafedra_mudiri'
        FROM kafedralar k JOIN fakultetlar f ON f.id=k.fakultet_id
        WHERE k.mudir_user_id IS NOT NULL AND NOT EXISTS(
          SELECT 1 FROM universitet_xodim_rollari xr
          WHERE xr.kafedra_id=k.id AND xr.rol='kafedra_mudiri' AND xr.faol=TRUE)""")
    cur.execute("""UPDATE universitetlar u SET rektor_user_id=xr.user_id
        FROM universitet_xodim_rollari xr
        WHERE xr.universitet_id=u.id AND xr.rol='rektor' AND xr.faol=TRUE
          AND u.rektor_user_id IS DISTINCT FROM xr.user_id""")
    cur.execute("""UPDATE fakultetlar f SET dekan_user_id=xr.user_id
        FROM universitet_xodim_rollari xr
        WHERE xr.fakultet_id=f.id AND xr.rol='dekan' AND xr.faol=TRUE
          AND f.dekan_user_id IS DISTINCT FROM xr.user_id""")
    cur.execute("""UPDATE kafedralar k SET mudir_user_id=xr.user_id
        FROM universitet_xodim_rollari xr
        WHERE xr.kafedra_id=k.id AND xr.rol='kafedra_mudiri' AND xr.faol=TRUE
          AND k.mudir_user_id IS DISTINCT FROM xr.user_id""")
    cur.execute("""CREATE TABLE IF NOT EXISTS universitet_tyutor_yonalishlari(
        id BIGSERIAL PRIMARY KEY,
        universitet_id INTEGER NOT NULL REFERENCES universitetlar(id) ON DELETE CASCADE,
        tyutor_user_id BIGINT NOT NULL REFERENCES users(user_id) ON DELETE CASCADE,
        fakultet_id INTEGER REFERENCES fakultetlar(id) ON DELETE CASCADE,
        yonalish_id BIGINT REFERENCES universitet_yonalishlari(id) ON DELETE CASCADE,
        talim_shakli TEXT, talim_tili TEXT, faol BOOLEAN NOT NULL DEFAULT TRUE,
        yaratilgan_by BIGINT REFERENCES users(user_id) ON DELETE SET NULL,
        yaratilgan_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
    )""")
    cur.execute("ALTER TABLE universitet_tyutor_yonalishlari ADD COLUMN IF NOT EXISTS fakultet_id INTEGER REFERENCES fakultetlar(id) ON DELETE CASCADE")
    cur.execute("ALTER TABLE universitet_tyutor_yonalishlari ADD COLUMN IF NOT EXISTS qabul_turi TEXT")
    cur.execute("ALTER TABLE universitet_tyutor_yonalishlari ALTER COLUMN yonalish_id DROP NOT NULL")
    cur.execute("""DELETE FROM universitet_tyutor_yonalishlari old USING universitet_tyutor_yonalishlari newer
        WHERE old.id<newer.id AND old.universitet_id=newer.universitet_id
          AND old.tyutor_user_id=newer.tyutor_user_id
          AND old.fakultet_id IS NOT DISTINCT FROM newer.fakultet_id
          AND old.yonalish_id IS NOT DISTINCT FROM newer.yonalish_id
          AND old.talim_shakli IS NOT DISTINCT FROM newer.talim_shakli
          AND old.talim_tili IS NOT DISTINCT FROM newer.talim_tili""")
    cur.execute("DROP INDEX IF EXISTS uq_uni_tyutor_qamrov")
    cur.execute("""CREATE UNIQUE INDEX uq_uni_tyutor_qamrov
        ON universitet_tyutor_yonalishlari(
          universitet_id,tyutor_user_id,COALESCE(fakultet_id,-1),COALESCE(yonalish_id,-1),
          COALESCE(talim_shakli,''),COALESCE(talim_tili,''),COALESCE(qabul_turi,''))""")
    cur.execute("""CREATE TABLE IF NOT EXISTS universitet_qabul_talabalari(
        id BIGSERIAL PRIMARY KEY,
        universitet_id INTEGER NOT NULL REFERENCES universitetlar(id) ON DELETE CASCADE,
        yonalish_id BIGINT NOT NULL REFERENCES universitet_yonalishlari(id) ON DELETE RESTRICT,
        guruh_id INTEGER REFERENCES universitet_guruhlari(id) ON DELETE SET NULL,
        user_id BIGINT REFERENCES users(user_id) ON DELETE SET NULL,
        abitur_id TEXT NOT NULL, jshshir TEXT NOT NULL, jshshir_hash CHAR(64) NOT NULL,
        familiya TEXT NOT NULL, ism TEXT NOT NULL, ota_ism TEXT,
        tugilgan_sana DATE, pasport_seriya TEXT, pasport_raqam TEXT,
        tavsiya_turi TEXT, talim_shakli TEXT NOT NULL, talim_tili TEXT NOT NULL,
        telefon TEXT, telegram_username TEXT, max_username TEXT,
        ball NUMERIC(7,2), doimiy_region TEXT, doimiy_tuman TEXT,
        maktab_region TEXT, maktab_tuman TEXT, maktab_turi TEXT, maktab_nomi TEXT,
        tugatgan_yili INTEGER, attestat TEXT, otm_nomi TEXT,
        qabul_bosqichi SMALLINT NOT NULL DEFAULT 1 CHECK(qabul_bosqichi BETWEEN 1 AND 4),
        hujjat_topshirgan_at TIMESTAMPTZ, bazaga_kiritilgan_at TIMESTAMPTZ,
        saytga_kiritilgan_at TIMESTAMPTZ, birinchi_kirish_at TIMESTAMPTZ,
        import_batch_id BIGINT, yaratilgan_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
        yangilangan_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
        UNIQUE(universitet_id,jshshir_hash), UNIQUE(universitet_id,abitur_id)
    )""")
    cur.execute("ALTER TABLE universitet_qabul_talabalari ADD COLUMN IF NOT EXISTS bazaga_kiritilgan_at TIMESTAMPTZ")
    cur.execute("ALTER TABLE universitet_qabul_talabalari ADD COLUMN IF NOT EXISTS birinchi_kirish_at TIMESTAMPTZ")
    # Qabul xodimi bilan telefon suhbati izohi (nega topshirmadi, qachon keladi...)
    cur.execute("ALTER TABLE universitet_qabul_talabalari ADD COLUMN IF NOT EXISTS aloqa_izohi TEXT")
    cur.execute("ALTER TABLE universitet_qabul_talabalari ADD COLUMN IF NOT EXISTS aloqa_izohi_at TIMESTAMPTZ")
    cur.execute("ALTER TABLE universitet_qabul_talabalari ADD COLUMN IF NOT EXISTS aloqa_izohi_by BIGINT")
    # Tuzilma elementlari o‘chirilmaydi: 1 yil yumshoq arxivda saqlanadi.
    cur.execute("ALTER TABLE fakultetlar ADD COLUMN IF NOT EXISTS faol BOOLEAN NOT NULL DEFAULT TRUE")
    cur.execute("ALTER TABLE fakultetlar ADD COLUMN IF NOT EXISTS arxiv_at TIMESTAMPTZ")
    cur.execute("ALTER TABLE fakultetlar ADD COLUMN IF NOT EXISTS arxiv_until TIMESTAMPTZ")
    cur.execute("ALTER TABLE kafedralar ADD COLUMN IF NOT EXISTS faol BOOLEAN NOT NULL DEFAULT TRUE")
    cur.execute("ALTER TABLE kafedralar ADD COLUMN IF NOT EXISTS arxiv_at TIMESTAMPTZ")
    cur.execute("ALTER TABLE kafedralar ADD COLUMN IF NOT EXISTS arxiv_until TIMESTAMPTZ")
    cur.execute("ALTER TABLE universitet_yonalishlari ADD COLUMN IF NOT EXISTS arxiv_at TIMESTAMPTZ")
    cur.execute("ALTER TABLE universitet_yonalishlari ADD COLUMN IF NOT EXISTS arxiv_until TIMESTAMPTZ")
    cur.execute("""CREATE TABLE IF NOT EXISTS universitet_tuzilma_arxivi(
        id BIGSERIAL PRIMARY KEY, universitet_id INTEGER NOT NULL,
        obyekt_turi TEXT NOT NULL CHECK(obyekt_turi IN ('fakultet','kafedra','yonalish')),
        obyekt_id INTEGER NOT NULL, nomi TEXT NOT NULL, hisoblar JSONB NOT NULL DEFAULT '{}'::jsonb,
        arxivlagan_by BIGINT, arxiv_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
        arxiv_until TIMESTAMPTZ NOT NULL DEFAULT (NOW()+INTERVAL '1 year'),
        tiklangan_at TIMESTAMPTZ, faol BOOLEAN NOT NULL DEFAULT TRUE
    )""")
    cur.execute("""CREATE UNIQUE INDEX IF NOT EXISTS uq_uni_tuzilma_faol_arxiv
        ON universitet_tuzilma_arxivi(universitet_id,obyekt_turi,obyekt_id) WHERE faol=TRUE""")
    cur.execute("ALTER TABLE universitet_qabul_talabalari DROP CONSTRAINT IF EXISTS universitet_qabul_talabalari_qabul_bosqichi_check")
    cur.execute("""UPDATE universitet_qabul_talabalari SET
        qabul_bosqichi=CASE WHEN qabul_bosqichi=3 AND user_id>=0 THEN 4 ELSE qabul_bosqichi END,
        bazaga_kiritilgan_at=CASE WHEN qabul_bosqichi>=3 THEN COALESCE(bazaga_kiritilgan_at,saytga_kiritilgan_at,yangilangan_at) ELSE bazaga_kiritilgan_at END,
        birinchi_kirish_at=CASE WHEN qabul_bosqichi=3 AND user_id>=0 THEN COALESCE(birinchi_kirish_at,saytga_kiritilgan_at,yangilangan_at) ELSE birinchi_kirish_at END
        WHERE qabul_bosqichi>=3""")
    cur.execute("ALTER TABLE universitet_qabul_talabalari ADD CONSTRAINT universitet_qabul_talabalari_qabul_bosqichi_check CHECK(qabul_bosqichi BETWEEN 1 AND 4)")
    cur.execute("""CREATE INDEX IF NOT EXISTS ix_uni_qabul_filter
        ON universitet_qabul_talabalari(universitet_id,yonalish_id,qabul_bosqichi,talim_shakli,talim_tili,ball DESC)""")
    cur.execute("""CREATE INDEX IF NOT EXISTS ix_uni_qabul_hujjat_sana
        ON universitet_qabul_talabalari(universitet_id,hujjat_topshirgan_at)
        WHERE hujjat_topshirgan_at IS NOT NULL""")
    cur.execute("""CREATE INDEX IF NOT EXISTS ix_uni_qabul_hemis_sana
        ON universitet_qabul_talabalari(universitet_id,bazaga_kiritilgan_at)
        WHERE bazaga_kiritilgan_at IS NOT NULL""")
    cur.execute("""CREATE INDEX IF NOT EXISTS ix_uni_yonalish_scope
        ON universitet_yonalishlari(universitet_id,fakultet_id,kafedra_id,id)
        WHERE faol=TRUE""")
    cur.execute("""CREATE TABLE IF NOT EXISTS universitet_import_batchlar(
        id BIGSERIAL PRIMARY KEY,
        universitet_id INTEGER REFERENCES universitetlar(id) ON DELETE CASCADE,
        import_turi TEXT NOT NULL, fayl_nomi TEXT NOT NULL, fayl_sha256 CHAR(64) NOT NULL,
        payload JSONB NOT NULL, xulosa JSONB NOT NULL,
        yaratilgan_by BIGINT NOT NULL REFERENCES users(user_id) ON DELETE CASCADE,
        holat TEXT NOT NULL DEFAULT 'preview',
        yaratilgan_at TIMESTAMPTZ NOT NULL DEFAULT NOW(), commit_at TIMESTAMPTZ
    )""")
    cur.execute("""CREATE TABLE IF NOT EXISTS universitet_taklif_kodlari(
        id BIGSERIAL PRIMARY KEY,
        universitet_id INTEGER NOT NULL REFERENCES universitetlar(id) ON DELETE CASCADE,
        xodim_rol_id BIGINT REFERENCES universitet_xodim_rollari(id) ON DELETE CASCADE,
        qabul_talaba_id BIGINT REFERENCES universitet_qabul_talabalari(id) ON DELETE CASCADE,
        placeholder_user_id BIGINT NOT NULL REFERENCES users(user_id) ON DELETE CASCADE,
        kod_hash TEXT NOT NULL UNIQUE, kod_shifr TEXT, turi TEXT NOT NULL,
        yaratilgan_by BIGINT REFERENCES users(user_id) ON DELETE SET NULL,
        yaratilgan_at TIMESTAMPTZ NOT NULL DEFAULT NOW(), ishlatildi_at TIMESTAMPTZ
    )""")
    cur.execute("ALTER TABLE universitet_taklif_kodlari ADD COLUMN IF NOT EXISTS kod_shifr TEXT")
    cur.execute("""CREATE TABLE IF NOT EXISTS universitet_workspace_map(
        context_id BIGINT PRIMARY KEY,
        universitet_id INTEGER NOT NULL UNIQUE REFERENCES universitetlar(id) ON DELETE CASCADE,
        yaratilgan_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
    )""")
    cur.execute("""CREATE TABLE IF NOT EXISTS universitet_audit_log(
        id BIGSERIAL PRIMARY KEY, universitet_id INTEGER NOT NULL,
        actor_user_id BIGINT NOT NULL, amal TEXT NOT NULL,
        obyekt_turi TEXT, obyekt_id BIGINT, tafsilot JSONB,
        yaratilgan_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
    )""")


def _ensure_schema(cur):
    if not _SCHEMA_READY:
        _institut_v20_jadvallari(cur)


def _is_global_admin(cur, user_id: int) -> bool:
    cur.execute("SELECT 1 FROM admin_akkaunt WHERE uid=%s", (user_id,))
    return cur.fetchone() is not None


def _require_active_university_source(cur, university_id: int) -> None:
    """V17 workspace'dan yaratilgan institut hali faol ekanini tekshiradi.

    Eski, qo'lda yaratilgan ``universitetlar`` yozuvida xarita bo'lmaydi va u
    odatdagidek ishlaydi. Xaritalangan yozuv esa faqat faol institut contexti
    bilan ochiladi; arxivlangan yoki adashib maktabga bog'langan yozuv UI'ga
    qayta sizib chiqmaydi.
    """
    global _V17_SOURCE_TABLES_READY
    if _V17_SOURCE_TABLES_READY is None:
        cur.execute("""SELECT
            to_regclass('public.universitet_workspace_map') IS NOT NULL
            AND to_regclass('public.organization_trials') IS NOT NULL
            AND to_regclass('public.learning_contexts') IS NOT NULL AS tayyor""")
        readiness = cur.fetchone()
        _V17_SOURCE_TABLES_READY = bool(readiness and readiness["tayyor"])
    if not _V17_SOURCE_TABLES_READY:
        return
    cur.execute("""SELECT uwm.context_id,o.organization_type,
            c.context_type,
            c.active AS context_active,
            LOWER(COALESCE(o.lifecycle_status,'')) AS lifecycle_status
        FROM universitet_workspace_map uwm
        LEFT JOIN organization_trials o ON o.context_id=uwm.context_id
        LEFT JOIN learning_contexts c ON c.id=uwm.context_id
        WHERE uwm.universitet_id=%s""", (university_id,))
    source = cur.fetchone()
    if not source:
        return
    if not (
        source.get("organization_type") == "institute"
        and source.get("context_type") == "university"
        and bool(source.get("context_active"))
        and source.get("lifecycle_status") in {"trial", "read_only", "active"}
    ):
        raise HTTPException(
            status_code=410,
            detail="Institut ish maydoni arxivlangan yoki faol emas",
        )


def _resolve_university(cur, user_id: int, workspace_id: Optional[int], create: bool = True) -> int:
    """Faol V17 institut contextini legacy universitet IDga xaritalaydi.

    Muhim: bu funksiya boshqa turdagi yoki arxivlangan workspace'ni institut
    sifatida hech qachon ochmaydi. Mavjud xarita ham qayta tekshiriladi;
    shuning uchun eski noto'g'ri ``school -> universitet`` bog'lanishi institut
    oynasiga sizib kira olmaydi.
    """
    if not workspace_id:
        cur.execute("""SELECT u.universitet_id
            FROM users u JOIN universitetlar universitet
              ON universitet.id=u.universitet_id
            WHERE u.user_id=%s""", (user_id,))
        row = cur.fetchone()
        if row and row["universitet_id"]:
            university_id = int(row["universitet_id"])
            _require_active_university_source(cur, university_id)
            return university_id
        # V2255: users.universitet_id bo'sh bo'lsa ham — xodim roli yoki talaba yozuvi
        # bo'yicha institut topiladi (dekan, tyutor, mudir kirganda 404 bo'lmasin).
        cur.execute("""SELECT r.universitet_id FROM universitet_xodim_rollari r
                       JOIN universitetlar u ON u.id=r.universitet_id
                       WHERE r.user_id=%s AND r.faol=TRUE ORDER BY r.id LIMIT 1""", (user_id,))
        row = cur.fetchone()
        if not row:
            cur.execute("""SELECT t.universitet_id FROM universitet_qabul_talabalari t
                           JOIN universitetlar u ON u.id=t.universitet_id
                           WHERE t.user_id=%s ORDER BY t.id LIMIT 1""", (user_id,))
            row = cur.fetchone()
        if row and row["universitet_id"]:
            university_id = int(row["universitet_id"])
            _require_active_university_source(cur, university_id)
            cur.execute("UPDATE users SET universitet_id=%s WHERE user_id=%s AND universitet_id IS NULL", (university_id, user_id))
            return university_id
        raise HTTPException(status_code=404, detail="Institut ish maydoni topilmadi")

    # Avval contextning o'zi tekshiriladi. Bu tekshiruv xarita topilgan holatda
    # ham INSERT/UPDATE yoki return'dan oldin bajarilishi shart.
    cur.execute("""SELECT o.display_name,o.creator_user_id,
            LOWER(COALESCE(o.lifecycle_status,'')) lifecycle_status,
            EXISTS(
              SELECT 1 FROM context_memberships cm
              WHERE cm.context_id=o.context_id AND cm.user_id=%s
                AND cm.status='active'
            ) faol_azo
        FROM organization_trials o
        JOIN learning_contexts c ON c.id=o.context_id
        WHERE o.context_id=%s
          AND o.organization_type='institute'
          AND c.context_type='university'
          AND c.active=TRUE
          AND LOWER(COALESCE(o.lifecycle_status,''))
              IN ('trial','read_only','active')""", (user_id, workspace_id))
    org = cur.fetchone()
    if not org:
        raise HTTPException(
            status_code=410,
            detail="Institut ish maydoni arxivlangan, faol emas yoki turi institut emas",
        )

    cur.execute("""SELECT uwm.universitet_id
        FROM universitet_workspace_map uwm
        JOIN universitetlar u ON u.id=uwm.universitet_id
        WHERE uwm.context_id=%s""", (workspace_id,))
    mapped = cur.fetchone()
    if mapped:
        return int(mapped["universitet_id"])
    if not create:
        raise HTTPException(status_code=404, detail="Institut xaritasi topilmadi")
    if int(org.get("creator_user_id") or 0) != int(user_id) and not _is_global_admin(cur, user_id):
        raise HTTPException(status_code=403, detail="Institut xaritasini faqat yaratuvchi yoki super administrator ochadi")
    cur.execute("INSERT INTO universitetlar(nomi) VALUES(%s) RETURNING id", (org["display_name"],))
    university_id = int(cur.fetchone()["id"])
    cur.execute("INSERT INTO universitet_workspace_map(context_id,universitet_id) VALUES(%s,%s)", (workspace_id, university_id))
    _assign_role(cur, university_id, user_id, "owner", created_by=user_id)
    cur.execute("UPDATE users SET universitet_id=COALESCE(universitet_id,%s),lavozim=COALESCE(lavozim,'owner') WHERE user_id=%s", (university_id, user_id))
    return university_id


def _roles(cur, user_id: int, university_id: int) -> list[dict[str, Any]]:
    roles = []
    if _is_global_admin(cur, user_id):
        roles.append({"rol": "institut_admin", "fakultet_id": None, "kafedra_id": None, "yonalish_id": None, "global_admin": True})
    cur.execute("""SELECT rol,fakultet_id,kafedra_id,yonalish_id
        FROM universitet_xodim_rollari
        WHERE universitet_id=%s AND user_id=%s AND faol=TRUE""", (university_id, user_id))
    roles.extend(dict(r) for r in cur.fetchall())
    cur.execute("SELECT id,yonalish_id FROM universitet_qabul_talabalari WHERE universitet_id=%s AND user_id=%s", (university_id, user_id))
    student = cur.fetchone()
    if student:
        roles.append({"rol": "talaba", "fakultet_id": None, "kafedra_id": None, "yonalish_id": student["yonalish_id"], "qabul_id": student["id"]})
    return roles


def _require_member(cur, user_id: int, university_id: int) -> list[dict[str, Any]]:
    _require_active_university_source(cur, university_id)
    roles = _roles(cur, user_id, university_id)
    if not roles:
        raise HTTPException(status_code=403, detail="Siz bu institutga biriktirilmagansiz")
    return roles


def _role_names(roles: list[dict[str, Any]]) -> set[str]:
    return {r["rol"] for r in roles}


def _operator_has_institute_scope(roles: list[dict[str, Any]]) -> bool:
    """Fakultetsiz qabul operatori faqat qabul modulida institut qamroviga ega."""
    return any(
        role.get("rol") == "qabul_operator" and not role.get("fakultet_id")
        for role in roles
    )


def _scope_program_ids(cur, university_id: int, roles: list[dict[str, Any]]) -> Optional[set[int]]:
    names = _role_names(roles)
    if names & INSTITUTE_WIDE or _operator_has_institute_scope(roles):
        return None
    ids: set[int] = set()
    faculties = {
        int(r["fakultet_id"])
        for r in roles
        if r.get("fakultet_id")
        and (r["rol"] in FACULTY_WIDE or r["rol"] == "qabul_operator")
    }
    departments = {int(r["kafedra_id"]) for r in roles if r.get("kafedra_id") and r["rol"] in DEPARTMENT_WIDE}
    direct = {int(r["yonalish_id"]) for r in roles if r.get("yonalish_id") and r["rol"] not in {"tyutor", "talaba"}}
    ids |= direct
    if faculties:
        cur.execute("SELECT id FROM universitet_yonalishlari WHERE universitet_id=%s AND fakultet_id=ANY(%s) AND faol=TRUE", (university_id, list(faculties)))
        ids |= {int(r["id"]) for r in cur.fetchall()}
    if departments:
        cur.execute("SELECT id FROM universitet_yonalishlari WHERE universitet_id=%s AND kafedra_id=ANY(%s) AND faol=TRUE", (university_id, list(departments)))
        ids |= {int(r["id"]) for r in cur.fetchall()}
    tutors = [r for r in roles if r["rol"] == "tyutor"]
    if tutors:
        user_ids = []
        # roles current userga tegishli, caller user_id tashqarida; assignmentdan topamiz.
        # Bu yerda rol yozuvlari user_id bermaydi, shuning uchun keyingi helper caller bilan ishlaydi.
    return ids


def _scope_program_ids_for_user(cur, university_id: int, user_id: int, roles: list[dict[str, Any]]) -> Optional[set[int]]:
    ids = _scope_program_ids(cur, university_id, roles)
    if ids is None:
        return None
    if "tyutor" in _role_names(roles):
        cur.execute("""SELECT DISTINCT y.id yonalish_id FROM universitet_yonalishlari y
            JOIN universitet_tyutor_yonalishlari ty ON ty.universitet_id=y.universitet_id
              AND (ty.fakultet_id IS NULL OR ty.fakultet_id=y.fakultet_id)
              AND (ty.yonalish_id IS NULL OR ty.yonalish_id=y.id)
            WHERE y.universitet_id=%s AND y.faol=TRUE AND ty.tyutor_user_id=%s AND ty.faol=TRUE""", (university_id, user_id))
        ids |= {int(r["yonalish_id"]) for r in cur.fetchall()}
    return ids


def _student_scope_clause(cur, university_id: int, user_id: int, roles: list[dict[str, Any]],
                          student_alias: str = "qt", program_alias: str = "y") -> tuple[str, list[Any]]:
    """Talaba qatoriga rol + fakultet/yo'nalish/shakl/til qamrovini qo'llaydi."""
    names = _role_names(roles)
    if names & INSTITUTE_WIDE or _operator_has_institute_scope(roles):
        return "TRUE", []
    clauses: list[str] = []
    params: list[Any] = []
    program_ids = _scope_program_ids(cur, university_id, [r for r in roles if r["rol"] != "tyutor"])
    if program_ids:
        clauses.append(f"{student_alias}.yonalish_id=ANY(%s)")
        params.append(sorted(program_ids))
    if "tyutor" in names:
        clauses.append(f"""EXISTS(SELECT 1 FROM universitet_tyutor_yonalishlari ty
            WHERE ty.universitet_id=%s AND ty.tyutor_user_id=%s AND ty.faol=TRUE
              AND (ty.fakultet_id IS NULL OR ty.fakultet_id={program_alias}.fakultet_id)
              AND (ty.yonalish_id IS NULL OR ty.yonalish_id={student_alias}.yonalish_id)
              AND (ty.talim_shakli IS NULL OR ty.talim_shakli={student_alias}.talim_shakli)
              AND (ty.talim_tili IS NULL OR ty.talim_tili={student_alias}.talim_tili)
              AND (ty.qabul_turi IS NULL OR ty.qabul_turi=CASE
                    WHEN LOWER(COALESCE({student_alias}.tavsiya_turi,'')) LIKE '%%grant%%'
                      OR LOWER(COALESCE({student_alias}.tavsiya_turi,'')) LIKE '%%budjet%%' THEN 'grant'
                    WHEN LOWER(COALESCE({student_alias}.tavsiya_turi,'')) LIKE '%%kontrakt%%'
                      OR LOWER(COALESCE({student_alias}.tavsiya_turi,'')) LIKE '%%shartnoma%%' THEN 'kontrakt'
                    ELSE NULL END))""")
        params.extend([university_id, user_id])
    return ("(" + " OR ".join(clauses) + ")", params) if clauses else ("FALSE", [])


def _student_access_allowed(cur, university_id: int, user_id: int, roles: list[dict[str, Any]], row: dict[str, Any]) -> bool:
    """Bitta talaba uchun qamrov tekshiruvi; tyutorning shakl va tilini ham saqlaydi."""
    names = _role_names(roles)
    if names & INSTITUTE_WIDE or _operator_has_institute_scope(roles):
        return True
    program_id = int(row["yonalish_id"])
    program_ids = _scope_program_ids(cur, university_id, [r for r in roles if r["rol"] != "tyutor"])
    if program_id in (program_ids or set()):
        return True
    if "tyutor" not in names:
        return False
    admission_value = row.get("tavsiya_turi")
    cur.execute("""SELECT 1 FROM universitet_tyutor_yonalishlari ty
        JOIN universitet_yonalishlari y ON y.id=%s AND y.universitet_id=ty.universitet_id
        WHERE ty.universitet_id=%s AND ty.tyutor_user_id=%s AND ty.faol=TRUE
          AND (ty.fakultet_id IS NULL OR ty.fakultet_id=y.fakultet_id)
          AND (ty.yonalish_id IS NULL OR ty.yonalish_id=%s)
          AND (ty.talim_shakli IS NULL OR ty.talim_shakli=%s)
          AND (ty.talim_tili IS NULL OR ty.talim_tili=%s)
          AND (ty.qabul_turi IS NULL OR ty.qabul_turi=CASE
                WHEN LOWER(COALESCE(%s,'')) LIKE '%%grant%%'
                  OR LOWER(COALESCE(%s,'')) LIKE '%%budjet%%' THEN 'grant'
                WHEN LOWER(COALESCE(%s,'')) LIKE '%%kontrakt%%'
                  OR LOWER(COALESCE(%s,'')) LIKE '%%shartnoma%%' THEN 'kontrakt'
                ELSE NULL END)
        LIMIT 1""", (
            program_id, university_id, user_id, program_id,
            row["talim_shakli"], row["talim_tili"],
            admission_value, admission_value, admission_value, admission_value,
        ))
    return cur.fetchone() is not None


def _has_any(roles: list[dict[str, Any]], allowed: set[str]) -> bool:
    return bool(_role_names(roles) & allowed)


def _validate_assignment_scope(cur, university_id: int, roles: list[dict[str, Any]],
                               faculty_id: Optional[int], department_id: Optional[int],
                               program_id: Optional[int]) -> Optional[int]:
    """Tanlangan fakultet/kafedra/yo'nalish bir institut va bir zanjirda ekanini tekshiradi."""
    candidates: list[int] = []
    selected_department_id: Optional[int] = None
    if faculty_id:
        cur.execute("SELECT id FROM fakultetlar WHERE id=%s AND universitet_id=%s AND faol=TRUE", (faculty_id, university_id))
        row = cur.fetchone()
        if not row: raise HTTPException(status_code=400, detail="Fakultet bu institutga tegishli emas")
        candidates.append(int(row["id"]))
    if department_id:
        cur.execute("""SELECT f.id FROM kafedralar k JOIN fakultetlar f ON f.id=k.fakultet_id
            WHERE k.id=%s AND f.universitet_id=%s
              AND k.faol=TRUE AND f.faol=TRUE""", (department_id, university_id))
        row = cur.fetchone()
        if not row: raise HTTPException(status_code=400, detail="Kafedra bu institutga tegishli emas")
        candidates.append(int(row["id"]))
        selected_department_id = int(department_id)
    if program_id:
        cur.execute("""SELECT y.fakultet_id,y.kafedra_id
            FROM universitet_yonalishlari y
            JOIN fakultetlar f ON f.id=y.fakultet_id
            JOIN kafedralar k ON k.id=y.kafedra_id AND k.fakultet_id=f.id
            WHERE y.id=%s AND y.universitet_id=%s
              AND y.faol=TRUE AND k.faol=TRUE AND f.faol=TRUE""", (program_id, university_id))
        row = cur.fetchone()
        if not row: raise HTTPException(status_code=400, detail="Yo'nalish bu institutga tegishli emas")
        candidates.append(int(row["fakultet_id"]))
        if selected_department_id is not None and int(row["kafedra_id"]) != selected_department_id:
            raise HTTPException(
                status_code=400,
                detail="Tanlangan yo'nalish aynan tanlangan kafedraga tegishli emas",
            )
    if len(set(candidates)) > 1:
        raise HTTPException(status_code=400, detail="Fakultet, kafedra va yo'nalish bir-biriga mos emas")
    target_faculty = candidates[0] if candidates else None
    if not (_role_names(roles) & INSTITUTE_WIDE) and not _operator_has_institute_scope(roles):
        allowed = {
            int(r["fakultet_id"])
            for r in roles
            if r.get("fakultet_id")
            and (r["rol"] in FACULTY_WIDE or r["rol"] == "qabul_operator")
        }
        if target_faculty is None or target_faculty not in allowed:
            raise HTTPException(status_code=403, detail="Faqat o'zingizga biriktirilgan fakultet doirasida ishlashingiz mumkin")
    return target_faculty


def _audit(cur, university_id: int, user_id: int, action: str, object_type: str = "", object_id: Optional[int] = None, detail: Optional[dict] = None):
    cur.execute("""INSERT INTO universitet_audit_log(universitet_id,actor_user_id,amal,obyekt_turi,obyekt_id,tafsilot)
        VALUES(%s,%s,%s,%s,%s,%s::jsonb)""", (university_id, user_id, action, object_type or None, object_id, json.dumps(detail or {}, ensure_ascii=False)))


def _sync_legacy_leader(cur, university_id: int, user_id: int, role: str,
                        faculty_id: Optional[int], department_id: Optional[int]) -> None:
    """Yangi rolni eski Muassasalar kartasidagi rahbar maydoniga ham yozadi."""
    if role == "rektor":
        cur.execute("UPDATE universitetlar SET rektor_user_id=%s WHERE id=%s", (user_id, university_id))
    elif role == "dekan" and faculty_id:
        cur.execute("UPDATE fakultetlar SET dekan_user_id=%s WHERE id=%s AND universitet_id=%s", (user_id, faculty_id, university_id))
    elif role == "kafedra_mudiri" and department_id:
        cur.execute("""UPDATE kafedralar k SET mudir_user_id=%s FROM fakultetlar f
            WHERE k.id=%s AND k.fakultet_id=f.id AND f.universitet_id=%s""", (user_id, department_id, university_id))


def _normalized_role_scope(role: str, faculty_id: Optional[int], department_id: Optional[int],
                           program_id: Optional[int]) -> tuple[Optional[int], Optional[int], Optional[int]]:
    """Rahbariyat rolini eng keng haqiqiy qamrovga kanoniklashtiradi."""
    if role in INSTITUTE_WIDE:
        return None, None, None
    if role in FACULTY_WIDE:
        return faculty_id, None, None
    if role in DEPARTMENT_WIDE:
        return faculty_id, department_id, None
    if role == "qabul_operator":
        return faculty_id, None, None
    return faculty_id, department_id, program_id


def _assign_role(cur, university_id: int, user_id: int, role: str, faculty_id: Optional[int] = None, department_id: Optional[int] = None, program_id: Optional[int] = None, created_by: Optional[int] = None) -> int:
    if role not in ROLE_LABELS or role == "talaba":
        raise HTTPException(status_code=400, detail=f"Noto'g'ri lavozim: {role}")
    if role in FACULTY_WIDE and not faculty_id:
        raise HTTPException(status_code=400, detail="Bu lavozim uchun fakultet tanlanishi shart")
    if role in DEPARTMENT_WIDE and not department_id:
        raise HTTPException(status_code=400, detail="Kafedra mudiri uchun kafedra tanlanishi shart")
    faculty_id, department_id, program_id = _normalized_role_scope(
        role, faculty_id, department_id, program_id
    )
    cur.execute("""SELECT id FROM universitet_xodim_rollari
        WHERE universitet_id=%s AND user_id=%s AND rol=%s
          AND fakultet_id IS NOT DISTINCT FROM %s
          AND kafedra_id IS NOT DISTINCT FROM %s
          AND yonalish_id IS NOT DISTINCT FROM %s
        LIMIT 1""", (university_id, user_id, role, faculty_id, department_id, program_id))
    existing = cur.fetchone()
    if existing:
        cur.execute("UPDATE universitet_xodim_rollari SET faol=TRUE WHERE id=%s", (existing["id"],))
        _sync_legacy_leader(cur, university_id, user_id, role, faculty_id, department_id)
        return int(existing["id"])
    singleton_where, singleton_params = None, []
    if role == "rektor":
        singleton_where, singleton_params = "universitet_id=%s AND rol='rektor'", [university_id]
    elif role in {"dekan", "manaviyatchi"} and faculty_id:
        singleton_where, singleton_params = "fakultet_id=%s AND rol=%s", [faculty_id, role]
    elif role == "kafedra_mudiri" and department_id:
        singleton_where, singleton_params = "kafedra_id=%s AND rol='kafedra_mudiri'", [department_id]
    if singleton_where:
        cur.execute(f"SELECT id,user_id FROM universitet_xodim_rollari WHERE {singleton_where} AND faol=TRUE ORDER BY id LIMIT 1", singleton_params)
        occupied = cur.fetchone()
        if occupied:
            if int(occupied["user_id"]) != int(user_id):
                cur.execute("DELETE FROM universitet_xodim_profili WHERE xodim_rol_id=%s", (occupied["id"],))
            cur.execute("""UPDATE universitet_xodim_rollari
                SET user_id=%s,universitet_id=%s,fakultet_id=%s,kafedra_id=%s,yonalish_id=%s,yaratilgan_by=%s
                WHERE id=%s""", (user_id, university_id, faculty_id, department_id, program_id, created_by, occupied["id"]))
            _sync_legacy_leader(cur, university_id, user_id, role, faculty_id, department_id)
            return int(occupied["id"])
    if role == "zam_dekan":
        cur.execute("SELECT COUNT(*) AS n FROM universitet_xodim_rollari WHERE fakultet_id=%s AND rol='zam_dekan' AND faol=TRUE", (faculty_id,))
        if int(cur.fetchone()["n"]) >= 2:
            raise HTTPException(status_code=409, detail="Bu fakultetda 2 ta dekan o'rinbosari allaqachon bor")
    cur.execute("""INSERT INTO universitet_xodim_rollari(
        universitet_id,fakultet_id,kafedra_id,yonalish_id,user_id,rol,yaratilgan_by)
        VALUES(%s,%s,%s,%s,%s,%s,%s)
        RETURNING id""",
        (university_id, faculty_id, department_id, program_id, user_id, role, created_by))
    role_id = int(cur.fetchone()["id"])
    _sync_legacy_leader(cur, university_id, user_id, role, faculty_id, department_id)
    # V2255: xodim kirganda to'g'ridan-to'g'ri shu institutga tushishi uchun
    cur.execute("UPDATE users SET universitet_id=%s WHERE user_id=%s AND universitet_id IS NULL", (university_id, user_id))
    return role_id


def _save_staff_profile(cur, role_id: int, *, ilmiy_daraja: Any = None,
                        ilmiy_unvon: Any = None, staj_yil: Any = None,
                        mutaxassislik: Any = None, qisqa_izoh: Any = None,
                        force: bool = False) -> None:
    """Kasbiy ma'lumotni rol yozuvidan alohida va ixtiyoriy saqlaydi.

    Eski xodimlar uchun barcha maydonlar ``NULL`` bo'lib qolishi mumkin;
    shu sababli yangi yo'nalish jamoasi ekrani eski bazani ham buzmaydi.
    """
    experience = _float(staj_yil)
    if experience is not None and not 0 <= experience <= 80:
        raise HTTPException(status_code=400, detail="Mehnat staji 0–80 yil oralig'ida bo'lishi kerak")
    values = (
        _norm(ilmiy_daraja) or None,
        _norm(ilmiy_unvon) or None,
        experience,
        _norm(mutaxassislik) or None,
        _norm(qisqa_izoh)[:500] or None,
    )
    if not force and not any(value is not None for value in values):
        return
    cur.execute("""INSERT INTO universitet_xodim_profili(
        xodim_rol_id,ilmiy_daraja,ilmiy_unvon,staj_yil,mutaxassislik,qisqa_izoh)
        VALUES(%s,%s,%s,%s,%s,%s)
        ON CONFLICT(xodim_rol_id) DO UPDATE SET
          ilmiy_daraja=EXCLUDED.ilmiy_daraja,
          ilmiy_unvon=EXCLUDED.ilmiy_unvon,
          staj_yil=EXCLUDED.staj_yil,
          mutaxassislik=EXCLUDED.mutaxassislik,
          qisqa_izoh=EXCLUDED.qisqa_izoh,
          yangilangan_at=NOW()""", (role_id, *values))


def _new_placeholder(cur, full_name: str, university_id: int, role: str, phone: Optional[str], created_by: int, faculty_id: Optional[int] = None, department_id: Optional[int] = None, program_id: Optional[int] = None) -> tuple[int, int, str]:
    p = _p()
    p._xodim_kod_jadvali(cur)
    cur.execute("""SELECT u.user_id,xr.id role_id FROM universitet_xodim_rollari xr
        JOIN users u ON u.user_id=xr.user_id
        WHERE xr.universitet_id=%s AND xr.rol=%s AND xr.faol=TRUE
          AND xr.fakultet_id IS NOT DISTINCT FROM %s
          AND xr.kafedra_id IS NOT DISTINCT FROM %s
          AND xr.yonalish_id IS NOT DISTINCT FROM %s
          AND LOWER(REGEXP_REPLACE(TRIM(u.full_name),'\\s+',' ','g'))
              =LOWER(REGEXP_REPLACE(TRIM(%s),'\\s+',' ','g'))
        ORDER BY xr.id LIMIT 1""", (university_id, role, faculty_id, department_id, program_id, full_name))
    existing = cur.fetchone()
    if existing:
        user_id, role_id = int(existing["user_id"]), int(existing["role_id"])
        cur.execute("UPDATE users SET full_name=%s,universitet_id=%s,lavozim=%s WHERE user_id=%s", (full_name, university_id, role, user_id))
        _sync_legacy_leader(cur, university_id, user_id, role, faculty_id, department_id)
        cur.execute("""UPDATE xodim_kod SET ishlatildi=TRUE WHERE ishlatildi=FALSE AND kod IN(
            SELECT kod_hash FROM universitet_taklif_kodlari WHERE xodim_rol_id=%s)""", (role_id,))
    else:
        cur.execute("SELECT pg_advisory_xact_lock(%s)", (20005401,))
        cur.execute("SELECT MIN(user_id) AS min_id FROM users WHERE user_id<0")
        row = cur.fetchone()
        user_id = int(row["min_id"] - 1) if row and row["min_id"] is not None else -1
        cur.execute("INSERT INTO users(user_id,full_name,role,universitet_id,lavozim) VALUES(%s,%s,'oqituvchi',%s,%s)", (user_id, full_name, university_id, role))
        role_id = _assign_role(cur, university_id, user_id, role, faculty_id, department_id, program_id, created_by)
    if phone:
        p._telefon_jadvallari(cur)
        cur.execute("INSERT INTO telefon_hisob(telefon,user_id) VALUES(%s,%s) ON CONFLICT(telefon) DO UPDATE SET user_id=EXCLUDED.user_id", (phone, user_id))
    while True:
        plain, stored = p._xodim_kod_yarat()
        cur.execute("SELECT 1 FROM xodim_kod WHERE kod=%s", (stored,))
        if not cur.fetchone():
            break
    cur.execute("INSERT INTO xodim_kod(kod,user_id) VALUES(%s,%s)", (stored, user_id))
    cur.execute("""INSERT INTO universitet_taklif_kodlari(
        universitet_id,xodim_rol_id,placeholder_user_id,kod_hash,kod_shifr,turi,yaratilgan_by)
        VALUES(%s,%s,%s,%s,%s,'xodim',%s)""",
        (university_id, role_id, user_id, stored, _seal_invite_code(plain), created_by))
    return user_id, role_id, plain


def _invite_fernet() -> Fernet:
    """Bir martalik kodni bazada ochiq saqlamasdan qayta ko'rish kaliti."""
    secret = str(getattr(_p(), "JWT_MAXFIY_KALIT", "") or "")
    if len(secret.encode("utf-8")) < 32:
        raise HTTPException(status_code=500, detail="Kirish kodlarini himoyalash kaliti sozlanmagan")
    key = base64.urlsafe_b64encode(hashlib.sha256(("samtm-invite-v1:" + secret).encode("utf-8")).digest())
    return Fernet(key)


def _seal_invite_code(code: str) -> str:
    return _invite_fernet().encrypt(code.encode("utf-8")).decode("ascii")


def _open_invite_code(ciphertext: Optional[str]) -> Optional[str]:
    if not ciphertext:
        return None
    try:
        return _invite_fernet().decrypt(ciphertext.encode("ascii")).decode("utf-8")
    except (InvalidToken, ValueError, UnicodeError):
        return None


def _create_student_invite(cur, row: dict[str, Any], actor_id: int) -> str:
    p = _p()
    current_user_id = int(row["user_id"]) if row.get("user_id") is not None else None
    if current_user_id is not None and current_user_id >= 0:
        raise HTTPException(status_code=409, detail="Talaba sayt akkauntiga allaqachon ulangan")
    # REV59: birinchi kirishgacha aynan o'sha kodni istalgancha ko'rish va
    # qayta yuborish mumkin. Hash tekshiruv uchun, shifr esa vakolatli admin
    # uchun qayta ko'rsatish uchun saqlanadi.
    cur.execute("""SELECT tk.kod_shifr FROM universitet_taklif_kodlari tk
        JOIN xodim_kod xk ON xk.kod=tk.kod_hash
        WHERE tk.qabul_talaba_id=%s AND xk.ishlatildi=FALSE
          AND xk.yaratildi>NOW()-INTERVAL '2 months'
        ORDER BY tk.id DESC LIMIT 1""", (row["id"],))
    reusable = cur.fetchone()
    reusable_code = _open_invite_code(reusable["kod_shifr"]) if reusable else None
    if reusable_code:
        return reusable_code
    if current_user_id is not None:
        user_id = current_user_id
        cur.execute("""UPDATE xodim_kod SET ishlatildi=TRUE WHERE kod IN (
            SELECT kod_hash FROM universitet_taklif_kodlari WHERE qabul_talaba_id=%s
        ) AND ishlatildi=FALSE""", (row["id"],))
    else:
        cur.execute("SELECT pg_advisory_xact_lock(%s)", (20005401,))
        cur.execute("SELECT MIN(user_id) AS min_id FROM users WHERE user_id<0")
        r = cur.fetchone()
        user_id = int(r["min_id"] - 1) if r and r["min_id"] is not None else -1
        full_name = " ".join(x for x in [row["familiya"], row["ism"], row.get("ota_ism")] if x)
        cur.execute("INSERT INTO users(user_id,full_name,role,universitet_id,lavozim) VALUES(%s,%s,'oquvchi',%s,'talaba')", (user_id, full_name, row["universitet_id"]))
        if row.get("telefon"):
            p._telefon_jadvallari(cur)
            cur.execute("INSERT INTO telefon_hisob(telefon,user_id) VALUES(%s,%s) ON CONFLICT(telefon) DO NOTHING", (row["telefon"], user_id))
    p._xodim_kod_jadvali(cur)
    while True:
        plain, stored = p._xodim_kod_yarat()
        cur.execute("SELECT 1 FROM xodim_kod WHERE kod=%s", (stored,))
        if not cur.fetchone():
            break
    cur.execute("INSERT INTO xodim_kod(kod,user_id) VALUES(%s,%s)", (stored, user_id))
    cur.execute("""INSERT INTO universitet_taklif_kodlari(
        universitet_id,qabul_talaba_id,placeholder_user_id,kod_hash,kod_shifr,turi,yaratilgan_by)
        VALUES(%s,%s,%s,%s,%s,'talaba',%s)""",
        (row["universitet_id"], row["id"], user_id, stored, _seal_invite_code(plain), actor_id))
    # Taklif yuborish hali "saytga kirgan" degani emas. 4-bosqich faqat kod
    # haqiqiy akkaunt tomonidan qabul qilinganda redeem_code ichida belgilanadi.
    cur.execute("UPDATE universitet_qabul_talabalari SET user_id=%s,yangilangan_at=NOW() WHERE id=%s", (user_id, row["id"]))
    return plain


def _credentials_xlsx(rows: list[dict[str, Any]], filename: str, title: str) -> dict[str, str]:
    """Bir martalik kodlarni javob bilan birga, bazada ochiq saqlamasdan XLSX qiladi."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="Kirish kodlari XLSX uchun openpyxl o'rnatilmagan") from exc
    wb = Workbook(); ws = wb.active; ws.title = "KIRISH KODLARI"
    headers = ["№", "F.I.Sh.", "Rol", "Fakultet", "Kafedra", "Yo'nalish", "Ta'lim", "Telefon", "Kirish kodi", "Amal qilish muddati"]
    ws.append([title] + [None] * (len(headers) - 1)); ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=15); ws["A1"].fill = PatternFill("solid", fgColor="0D7A77"); ws["A1"].alignment = Alignment(horizontal="center")
    ws.append(headers)
    for cell in ws[2]:
        cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor="173E5B"); cell.alignment = Alignment(horizontal="center", vertical="center")
    for index, item in enumerate(rows, 1):
        education = " · ".join(x for x in [item.get("talim_shakli"), item.get("talim_tili")] if x)
        ws.append([index, item.get("fish"), item.get("lavozim") or item.get("rol"), item.get("fakultet"),
                   item.get("kafedra"), item.get("yonalish"), education, item.get("telefon"),
                   item.get("kirish_kodi"), item.get("kod_muddati") or "2 oy"])
    widths = [7, 34, 27, 28, 30, 37, 24, 19, 20, 20]
    for index, width in enumerate(widths, 1): ws.column_dimensions[chr(64 + index)].width = width
    ws.freeze_panes = "A3"; ws.auto_filter.ref = f"A2:J{max(2, ws.max_row)}"; ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 27; ws.row_dimensions[2].height = 24
    for row in ws.iter_rows(min_row=3):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    stream = io.BytesIO(); wb.save(stream); wb.close()
    return {"fayl_nomi": filename, "base64": base64.b64encode(stream.getvalue()).decode("ascii")}


def _structure_template_xlsx(university: dict[str, Any], structure_rows: list[dict[str, Any]],
                             faculties: list[dict[str, Any]], departments: list[dict[str, Any]]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation
    wb = Workbook(); institute = wb.active; institute.title = "INSTITUT"
    structure = wb.create_sheet("TUZILMA"); staff = wb.create_sheet("XODIMLAR"); notes = wb.create_sheet("IZOH"); lists = wb.create_sheet("ROYXATLAR")
    teal, navy, cream, green = "0D7A77", "173E5B", "FFF8E8", "EAF7EF"
    institute.append(["INSTITUT ASOSIY MA'LUMOTLARI", None]); institute.merge_cells("A1:B1")
    institute.append([]); institute.append(["Maydon", "Qiymat"])
    for item in [("Institut nomi", university["nomi"]), ("Viloyat", university.get("viloyat")), ("Tuman", university.get("tuman")), ("Izoh", "Institut nomini o'zgartirmang")]: institute.append(item)
    structure.append(["FAKULTET → KAFEDRA → TA'LIM YO'NALISHI"] + [None] * 6); structure.merge_cells("A1:G1")
    structure.append(["Saytda mavjud fakultet va kafedralar oldindan yozildi. Bo'sh yo'nalish kataklarini to'ldiring."] + [None] * 6); structure.merge_cells("A2:G2")
    structure.append([]); structure.append(["Fakultet", "Kafedra", "Yo'nalish kodi", "Yo'nalish nomi", "Daraja", "Ta'lim shakli", "Ta'lim tili"])
    if structure_rows:
        for row in structure_rows: structure.append([row.get("fakultet"), row.get("kafedra"), row.get("kodi"), row.get("yonalish"), row.get("daraja"), row.get("talim_shakli"), row.get("talim_tili")])
    else:
        for row in departments: structure.append([row.get("fakultet"), row.get("kafedra"), None, None, "Bakalavriat", "Kunduzgi", "O‘zbekcha"])
    staff.append(["INSTITUT XODIMLARI VA ROLLARI"] + [None] * 11); staff.merge_cells("A1:L1")
    staff.append(["F.I.Sh. kataklarini to'ldiring. Importdan keyin har biriga 2 oylik kirish kodi Excelda tushadi."] + [None] * 11); staff.merge_cells("A2:L2")
    staff.append([]); staff.append(["F.I.Sh.", "Telefon", "Lavozim", "Fakultet", "Kafedra", "Yo'nalish",
                                    "Ilmiy daraja", "Ilmiy unvon", "Staj (yil)", "Mutaxassislik",
                                    "Qisqa izoh", "Shablon izohi"])
    staff.append([None, None, "Rektor", None, None, None, None, None, None, None, None, "Majburiy"])
    for faculty in faculties:
        for role, note in [("Dekan", "1 ta"), ("Dekan o'rinbosari", "1/2"), ("Dekan o'rinbosari", "2/2"), ("Ma'naviy-ma'rifiy ishlar mas'uli", "1 ta")]:
            staff.append([None, None, role, faculty["nomi"], None, None, None, None, None, None, None, note])
    for department in departments:
        staff.append([None, None, "Kafedra mudiri", department["fakultet"], department["kafedra"], None,
                      None, None, None, None, None, "1 ta"])
    tutor_seen = set()
    for row in structure_rows:
        tutor_key = (_key(row.get("fakultet")), _key(row.get("kafedra")), _key(row.get("yonalish")))
        if not row.get("yonalish") or tutor_key in tutor_seen: continue
        tutor_seen.add(tutor_key)
        staff.append([None, None, "Tyutor", row.get("fakultet"), row.get("kafedra"), row.get("yonalish"),
                      None, None, None, None, None, "Kerakli miqdorda qatorni nusxalang"])
    notes.append(["QADAM", "NIMA QILASIZ"]); notes.append([1, "INSTITUT varag'idagi nomni o'zgartirmang."]); notes.append([2, "TUZILMA varag'ida yo'nalish, shakl va tilning har bir variantini alohida qator qiling."]); notes.append([3, "XODIMLAR varag'ida rektor, har fakultetga 1 dekan, 2 zamdekan, 1 ma'naviy mas'ul va har kafedraga 1 mudir yozing."]); notes.append([4, "Faylni saytda Tekshirishdan o'tkazing; xato bo'lmasa import qiling."]); notes.append([5, "Xodimlarning kirish kodlari avtomatik XLSX bo'lib yuklanadi."])
    list_values = [["Daraja", "Ta'lim shakli", "Ta'lim tili", "Lavozim"]]
    max_len = max(len(DARAJALAR), len(TA_LIM_SHAKLLARI), len(TA_LIM_TILLARI), len(ROLE_LABELS) - 2)
    roles = [label for key, label in ROLE_LABELS.items() if key not in {"owner", "talaba"}]
    for index in range(max_len): list_values.append([DARAJALAR[index] if index < len(DARAJALAR) else None, TA_LIM_SHAKLLARI[index] if index < len(TA_LIM_SHAKLLARI) else None, TA_LIM_TILLARI[index] if index < len(TA_LIM_TILLARI) else None, roles[index] if index < len(roles) else None])
    for row in list_values: lists.append(row)
    lists.sheet_state = "hidden"
    validations = [(structure, "E5:E500", f"'ROYXATLAR'!$A$2:$A${len(DARAJALAR)+1}"), (structure, "F5:F500", f"'ROYXATLAR'!$B$2:$B${len(TA_LIM_SHAKLLARI)+1}"), (structure, "G5:G500", f"'ROYXATLAR'!$C$2:$C${len(TA_LIM_TILLARI)+1}"), (staff, "C5:C500", f"'ROYXATLAR'!$D$2:$D${len(roles)+1}")]
    for sheet, cell_range, formula in validations:
        validation = DataValidation(type="list", formula1=formula, allow_blank=True); sheet.add_data_validation(validation); validation.add(cell_range)
    for sheet, widths in [(institute, [25, 55]), (structure, [27, 30, 18, 38, 18, 20, 18]),
                          (staff, [32, 19, 31, 28, 30, 38, 24, 24, 14, 30, 42, 34]),
                          (notes, [10, 110])]:
        sheet.sheet_view.showGridLines = False; sheet.freeze_panes = "A5" if sheet in (structure, staff) else "A2"
        for index, width in enumerate(widths, 1): sheet.column_dimensions[chr(64 + index)].width = width
        for cell in sheet[1]: cell.fill = PatternFill("solid", fgColor=teal); cell.font = Font(bold=True, color="FFFFFF", size=14)
        header_row = 4 if sheet in (structure, staff) else 3 if sheet is institute else 1
        for cell in sheet[header_row]: cell.fill = PatternFill("solid", fgColor=navy); cell.font = Font(bold=True, color="FFFFFF"); cell.alignment = Alignment(vertical="center")
        for row in sheet.iter_rows(min_row=header_row + 1):
            for cell in row: cell.alignment = Alignment(vertical="top", wrap_text=True)
        if sheet in (structure, staff):
            for row in sheet.iter_rows(min_row=5, max_row=sheet.max_row):
                for cell in row: cell.fill = PatternFill("solid", fgColor=cream if sheet is structure else green)
    stream = io.BytesIO(); wb.save(stream); wb.close(); return stream.getvalue()


def _admission_template_xlsx(university: dict[str, Any], programs: list[dict[str, Any]]) -> bytes:
    """Qabul importeri o'qiydigan ustunlar bilan toza XLSX shablon yaratadi."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    headers = [
        "AbiturID", "JSHSHIR", "Familya", "Ism", "Ota ism", "Tug'ilgan sana",
        "Pasport seriya", "Pasport raqam", "Tavsiya turi", "Ta'lim shakli",
        "Ta'lim Tili", "OTM", "Yo'nalish", "Telefon", "Ball", "D y region",
        "D y tuman", "Maktab region", "Maktab tuman", "Maktab turi",
        "Maktab nomi", "Tugatgan yili", "Serya va raqam",
    ]
    wb = Workbook(); sheet = wb.active; sheet.title = "QABUL"
    lists = wb.create_sheet("ROYXATLAR")
    sheet.append(headers)
    # Birinchi bo'sh qatorda OTM nomi tayyor turadi; qolganini foydalanuvchi
    # qo'lda yoki rasmiy qabul ro'yxatidan nusxalab to'ldiradi.
    sheet.append([None] * 11 + [university.get("nomi")] + [None] * 11)

    program_names = sorted({_norm(item.get("nomi")) for item in programs if _norm(item.get("nomi"))})
    list_columns = [
        ("Tavsiya turi", ["Davlat granti", "To'lov-kontrakt"]),
        ("Ta'lim shakli", TA_LIM_SHAKLLARI),
        ("Ta'lim tili", TA_LIM_TILLARI),
        ("Yo'nalish", program_names),
    ]
    max_rows = max([len(values) for _, values in list_columns] + [1])
    lists.append([label for label, _ in list_columns])
    for index in range(max_rows):
        lists.append([values[index] if index < len(values) else None for _, values in list_columns])
    lists.sheet_state = "hidden"

    for column, list_index, value_count in (("I", 1, 2), ("J", 2, len(TA_LIM_SHAKLLARI)),
                                             ("K", 3, len(TA_LIM_TILLARI)), ("M", 4, len(program_names))):
        if value_count < 1:
            continue
        end = value_count + 1
        validation = DataValidation(type="list", formula1=f"'ROYXATLAR'!${get_column_letter(list_index)}$2:${get_column_letter(list_index)}${end}", allow_blank=True)
        sheet.add_data_validation(validation); validation.add(f"{column}2:{column}2000")

    widths = [14, 18, 22, 20, 20, 17, 16, 18, 20, 18, 17, 38, 38, 18, 12,
              20, 20, 20, 20, 18, 34, 15, 20]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="173E5B")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in sheet[2]:
        cell.fill = PatternFill("solid", fgColor="FFF8E8")
    sheet.freeze_panes = "A2"; sheet.auto_filter.ref = "A1:W2000"; sheet.sheet_view.showGridLines = False
    stream = io.BytesIO(); wb.save(stream); wb.close(); return stream.getvalue()


def _parse_admission(content: bytes, filename: str) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    sheets = _workbook_rows(content, filename)
    rows = _active_rows(sheets, ("QABUL", "Лист1", "Sheet1"))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="Qabul faylida ma'lumot qatorlari yo'q")
    headers = _header_map(rows[0])
    required = ["AbiturID", "JSHSHIR", "Familya", "Ism", "Ta'lim shakli", "Ta'lim Tili", "OTM", "Yo'nalish", "Telefon", "Ball"]
    missing = [h for h in required if _key(h) not in headers]
    if missing:
        raise HTTPException(status_code=400, detail="Majburiy ustunlar yetishmaydi: " + ", ".join(missing))
    parsed, errors, warnings = [], [], []
    seen_pin, seen_abitur = set(), set()
    for excel_row, values in enumerate(rows[1:], 2):
        if not any(_norm(v) for v in values):
            continue
        abitur = _text_number(_cell(values, headers, "AbiturID"))
        pin = _digits(_cell(values, headers, "JSHSHIR"))
        family = _norm(_cell(values, headers, "Familya", "Familiya"))
        name = _norm(_cell(values, headers, "Ism"))
        form = _canonical_choice(_cell(values, headers, "Ta'lim shakli"), TA_LIM_SHAKLLARI)
        language = _canonical_choice(_cell(values, headers, "Ta'lim Tili", "Ta'lim tili"), TA_LIM_TILLARI)
        source_direction = _norm(_cell(values, headers, "Yo'nalish", "Ta'lim yo'nalishi"))
        direction = _admission_program_name(source_direction)
        phone = _telefon(_cell(values, headers, "Telefon"))
        score = _float(_cell(values, headers, "Ball", "Kirish bali"))
        row_errors = []
        if not abitur: row_errors.append("AbiturID bo'sh")
        if len(pin) != 14: row_errors.append("JSHSHIR 14 xonali emas")
        if not family or not name: row_errors.append("F.I.Sh. to'liq emas")
        if not form: row_errors.append("Ta'lim shakli tanilmagan")
        if not language: row_errors.append("Ta'lim tili tanilmagan")
        if not source_direction or not direction: row_errors.append("Yo'nalish bo'sh")
        if not phone: row_errors.append("Telefon noto'g'ri")
        if score is None: row_errors.append("Ball son emas")
        if pin in seen_pin: row_errors.append("Fayl ichida JSHSHIR takror")
        if abitur in seen_abitur: row_errors.append("Fayl ichida AbiturID takror")
        if row_errors:
            errors.append({"qator": excel_row, "xatolar": row_errors})
            continue
        seen_pin.add(pin); seen_abitur.add(abitur)
        birth = _iso_date(_cell(values, headers, "Tug'ilgan sana"))
        if not birth:
            warnings.append({"qator": excel_row, "ogohlantirish": "Tug'ilgan sana aniqlanmadi"})
        passport_number = _text_number(_cell(values, headers, "Pasport raqam")) or None
        if passport_number and re.fullmatch(r"\d{6}", passport_number):
            warnings.append({
                "qator": excel_row,
                "ogohlantirish": (
                    "Pasport raqami 6 xonali. Excel boshidagi 0 ni olib tashlagan bo'lishi mumkin; "
                    "tizim avtomatik 0 qo'shmadi, manba faylni tekshiring."
                ),
            })
        parsed.append({
            "excel_row": excel_row, "abitur_id": abitur, "jshshir": pin,
            "jshshir_hash": hashlib.sha256(pin.encode()).hexdigest(),
            "familiya": family, "ism": name, "ota_ism": _norm(_cell(values, headers, "Ota ism")) or None,
            "tugilgan_sana": birth, "pasport_seriya": _norm(_cell(values, headers, "Pasport seriya")) or None,
            "pasport_raqam": passport_number,
            "tavsiya_turi": _norm(_cell(values, headers, "Tavsiya turi")) or None,
            "talim_shakli": form, "talim_tili": language,
            "otm_nomi": _norm(_cell(values, headers, "OTM")),
            "yonalish_manba_nomi": source_direction, "yonalish_nomi": direction,
            "telefon": phone, "ball": score,
            "doimiy_region": _norm(_cell(values, headers, "D y region", "Doimiy region")) or None,
            "doimiy_tuman": _norm(_cell(values, headers, "D y tuman", "Doimiy tuman")) or None,
            "maktab_region": _norm(_cell(values, headers, "Maktab region")) or None,
            "maktab_tuman": _norm(_cell(values, headers, "Maktab tuman")) or None,
            "maktab_turi": _norm(_cell(values, headers, "Maktab turi")) or None,
            "maktab_nomi": _norm(_cell(values, headers, "Maktab nomi")) or None,
            "tugatgan_yili": int(_float(_cell(values, headers, "Tugatgan yili")) or 0) or None,
            "attestat": _norm(_cell(values, headers, "Serya va raqam", "Seriya va raqam")) or None,
        })
    counts = lambda key: {v: sum(1 for r in parsed if r[key] == v) for v in sorted({r[key] for r in parsed})}
    variant_counts: dict[tuple[str, str, str], int] = {}
    for item in parsed:
        variant_key = (item["yonalish_nomi"], item["talim_shakli"], item["talim_tili"])
        variant_counts[variant_key] = variant_counts.get(variant_key, 0) + 1
    summary = {
        "jami_qator": len(rows) - 1, "yaroqli": len(parsed), "xato_soni": len(errors),
        "ogohlantirish_soni": len(warnings), "xatolar": errors[:100], "ogohlantirishlar": warnings[:100],
        "yonalishlar": counts("yonalish_nomi"),
        "manba_yonalishlar": counts("yonalish_manba_nomi"),
        "yonalish_variantlari": [
            {"yonalish": key[0], "talim_shakli": key[1], "talim_tili": key[2], "talaba_soni": count}
            for key, count in sorted(variant_counts.items())
        ],
        "talim_shakllari": counts("talim_shakli"),
        "talim_tillari": counts("talim_tili"),
    }
    return parsed, summary


def _role_key(value: Any) -> Optional[str]:
    k = _key(value)
    aliases = {
        "institutegasi": "owner", "rektor": "rektor", "prorektor": "prorektor",
        "institutadministratori": "institut_admin", "institutadmin": "institut_admin",
        "qabuloperatori": "qabul_operator", "qabuloperator": "qabul_operator",
        "dekan": "dekan", "dekano'rinbosari": "zam_dekan", "dekanorinbosari": "zam_dekan",
        "zamdekan": "zam_dekan", "manaviyatchi": "manaviyatchi",
        "manaviyma'rifiyishlarmas'uli": "manaviyatchi", "manaviymarifiyishlarmasuli": "manaviyatchi",
        "fakultetadministratori": "fakultet_admin", "fakultetadmin": "fakultet_admin",
        "kafedramudiri": "kafedra_mudiri", "professoro'qituvchi": "professor_oqituvchi",
        "professoroqituvchi": "professor_oqituvchi", "tyutor": "tyutor",
    }
    if k in ROLE_LABELS:
        return k
    return aliases.get(k)


def _parse_structure(content: bytes, filename: str) -> tuple[dict[str, Any], dict[str, Any]]:
    sheets = _workbook_rows(content, filename)
    institute_rows = _named_rows(sheets, "INSTITUT")
    structure_rows = _named_rows(sheets, "TUZILMA")
    staff_rows = _named_rows(sheets, "XODIMLAR")
    if not institute_rows or not structure_rows or not staff_rows:
        raise HTTPException(status_code=400, detail="Shablonda INSTITUT, TUZILMA va XODIMLAR varaqlari bo'lishi shart")

    institute = {}
    for row in institute_rows:
        if len(row) >= 2 and _norm(row[0]):
            institute[_key(row[0])] = _norm(row[1])
    name = institute.get(_key("Institut nomi")) or institute.get("institutnomi")
    if not name:
        raise HTTPException(status_code=400, detail="INSTITUT varag'ida 'Institut nomi' to'ldirilmagan")

    required = ["Fakultet", "Kafedra", "Yo'nalish nomi", "Daraja", "Ta'lim shakli", "Ta'lim tili"]
    structure_header_index, sh = _find_header_row(structure_rows, required, "TUZILMA")
    structures, errors, warnings = [], [], []
    seen_variants = set()
    for row_no, row in enumerate(structure_rows[structure_header_index + 1:], structure_header_index + 2):
        if not any(_norm(v) for v in row):
            continue
        faculty = _norm(_cell(row, sh, "Fakultet"))
        department = _norm(_cell(row, sh, "Kafedra"))
        program = _norm(_cell(row, sh, "Yo'nalish nomi"))
        code = _norm(_cell(row, sh, "Yo'nalish kodi")) or None
        degree = _canonical_choice(_cell(row, sh, "Daraja"), DARAJALAR)
        form = _canonical_choice(_cell(row, sh, "Ta'lim shakli"), TA_LIM_SHAKLLARI)
        language = _canonical_choice(_cell(row, sh, "Ta'lim tili"), TA_LIM_TILLARI)
        row_errors = []
        if not faculty: row_errors.append("Fakultet bo'sh")
        if not department: row_errors.append("Kafedra bo'sh")
        if not program: row_errors.append("Yo'nalish bo'sh")
        if not degree: row_errors.append("Daraja noto'g'ri")
        if not form: row_errors.append("Ta'lim shakli noto'g'ri")
        if not language: row_errors.append("Ta'lim tili noto'g'ri")
        variant_key = (_key(faculty), _key(department), _key(program), degree, form, language)
        if variant_key in seen_variants: row_errors.append("Aynan shu yo'nalish/shakl/til takrorlangan")
        if row_errors:
            errors.append({"varaq": "TUZILMA", "qator": row_no, "xatolar": row_errors}); continue
        seen_variants.add(variant_key)
        structures.append({"fakultet": faculty, "kafedra": department, "yonalish": program,
                           "yonalish_kodi": code, "daraja": degree, "talim_shakli": form, "talim_tili": language})

    staff_required = ["F.I.Sh.", "Lavozim"]
    staff_header_index, xh = _find_header_row(staff_rows, staff_required, "XODIMLAR")
    staff, seen_people = [], set()
    for row_no, row in enumerate(staff_rows[staff_header_index + 1:], staff_header_index + 2):
        if not any(_norm(v) for v in row):
            continue
        fish = _norm(_cell(row, xh, "F.I.Sh.", "FISH"))
        role = _role_key(_cell(row, xh, "Lavozim"))
        phone_raw = _cell(row, xh, "Telefon")
        phone = _telefon(phone_raw) if phone_raw not in (None, "") else None
        faculty = _norm(_cell(row, xh, "Fakultet")) or None
        department = _norm(_cell(row, xh, "Kafedra")) or None
        program = _norm(_cell(row, xh, "Yo'nalish")) or None
        experience_raw = _cell(row, xh, "Staj (yil)", "Staj", "Mehnat staji")
        experience = _float(experience_raw) if experience_raw not in (None, "") else None
        row_errors = []
        if not fish: row_errors.append("F.I.Sh. bo'sh")
        if not role: row_errors.append("Lavozim tanilmadi")
        if phone_raw not in (None, "") and not phone: row_errors.append("Telefon noto'g'ri")
        if role in FACULTY_WIDE and not faculty: row_errors.append("Bu lavozim uchun fakultet shart")
        if role in DEPARTMENT_WIDE and not department: row_errors.append("Kafedra mudiri uchun kafedra shart")
        if experience_raw not in (None, "") and experience is None:
            row_errors.append("Staj son emas")
        if experience is not None and not 0 <= experience <= 80:
            row_errors.append("Staj 0–80 yil oralig'ida bo'lishi kerak")
        person_key = (_key(fish), role, _key(faculty), _key(department), _key(program))
        if person_key in seen_people: row_errors.append("Xodim qatori takrorlangan")
        if row_errors:
            errors.append({"varaq": "XODIMLAR", "qator": row_no, "xatolar": row_errors}); continue
        seen_people.add(person_key)
        staff.append({"fish": fish, "telefon": phone, "rol": role, "fakultet": faculty,
                      "kafedra": department, "yonalish": program,
                      "ilmiy_daraja": _norm(_cell(row, xh, "Ilmiy daraja")) or None,
                      "ilmiy_unvon": _norm(_cell(row, xh, "Ilmiy unvon")) or None,
                      "staj_yil": experience,
                      "mutaxassislik": _norm(_cell(row, xh, "Mutaxassislik")) or None,
                      "qisqa_izoh": _norm(_cell(row, xh, "Qisqa izoh"))[:500] or None,
                      "excel_row": row_no})

    faculty_names = sorted({x["fakultet"] for x in structures})
    department_pairs = {(x["fakultet"], x["kafedra"]) for x in structures}
    for item in staff:
        if item["fakultet"] and item["fakultet"] not in faculty_names:
            errors.append({"varaq": "XODIMLAR", "qator": item["excel_row"], "xatolar": ["Fakultet TUZILMA varag'ida yo'q"]})
        if item["kafedra"] and (item["fakultet"], item["kafedra"]) not in department_pairs:
            errors.append({"varaq": "XODIMLAR", "qator": item["excel_row"], "xatolar": ["Kafedra va fakultet mos emas"]})
    completeness = {}
    for faculty in faculty_names:
        members = [x for x in staff if x["fakultet"] == faculty]
        counts = {r: sum(1 for x in members if x["rol"] == r) for r in ("dekan", "zam_dekan", "manaviyatchi", "fakultet_admin")}
        completeness[faculty] = counts
        if counts["dekan"] != 1:
            errors.append({"varaq": "XODIMLAR", "qator": None, "xatolar": [f"{faculty}: aynan 1 ta dekan bo'lishi kerak"]})
        if counts["zam_dekan"] != 2:
            errors.append({"varaq": "XODIMLAR", "qator": None, "xatolar": [f"{faculty}: aynan 2 ta dekan o'rinbosari bo'lishi kerak"]})
        if counts["manaviyatchi"] != 1:
            errors.append({"varaq": "XODIMLAR", "qator": None, "xatolar": [f"{faculty}: aynan 1 ta ma'naviyatchi bo'lishi kerak"]})
        if counts["fakultet_admin"] == 0:
            warnings.append(f"{faculty}: alohida admin kiritilmagan; importni bajargan admin avtomatik biriktiriladi")

    payload = {"institut": {"nomi": name, "viloyat": institute.get(_key("Viloyat")) or None,
                             "tuman": institute.get(_key("Tuman")) or None},
               "tuzilma": structures, "xodimlar": staff}
    summary = {"institut_nomi": name, "fakultet_soni": len(faculty_names),
               "kafedra_soni": len(department_pairs), "yonalish_variant_soni": len(structures),
               "xodim_soni": len(staff), "fakultet_toldirilishi": completeness,
               "xato_soni": len(errors), "xatolar": errors[:150], "ogohlantirishlar": warnings}
    return payload, summary


def _parse_staff_template(content: bytes, filename: str) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """Alohida XODIMLAR shablonini tuzilmani o'zgartirmasdan tekshiradi."""
    sheets = _workbook_rows(content, filename)
    rows = _named_rows(sheets, "XODIMLAR") or _active_rows(sheets)
    header_index, headers = _find_header_row(rows, ["F.I.Sh.", "Lavozim"], "XODIMLAR")
    staff: list[dict[str, Any]] = []; errors: list[dict[str, Any]] = []; seen = set()
    for row_no, row in enumerate(rows[header_index + 1:], header_index + 2):
        fish = _norm(_cell(row, headers, "F.I.Sh.", "FISH"))
        # Yuklangan tayyor shablondagi namunaviy/majburiy bo'sh qatorlar xato emas.
        if not fish:
            continue
        role = _role_key(_cell(row, headers, "Lavozim"))
        phone_raw = _cell(row, headers, "Telefon")
        phone = _telefon(phone_raw) if phone_raw not in (None, "") else None
        faculty = _norm(_cell(row, headers, "Fakultet")) or None
        department = _norm(_cell(row, headers, "Kafedra")) or None
        program = _norm(_cell(row, headers, "Yo'nalish")) or None
        experience_raw = _cell(row, headers, "Staj (yil)", "Staj", "Mehnat staji")
        experience = _float(experience_raw) if experience_raw not in (None, "") else None
        row_errors = []
        if not role: row_errors.append("Lavozim tanilmadi")
        if phone_raw not in (None, "") and not phone: row_errors.append("Telefon noto'g'ri")
        if role in FACULTY_WIDE and not faculty: row_errors.append("Bu lavozim uchun fakultet shart")
        if role in DEPARTMENT_WIDE and not department: row_errors.append("Kafedra mudiri uchun kafedra shart")
        if experience_raw not in (None, "") and experience is None: row_errors.append("Staj son emas")
        if experience is not None and not 0 <= experience <= 80: row_errors.append("Staj 0–80 yil oralig'ida bo'lishi kerak")
        unique_key = (_key(fish), role, _key(faculty), _key(department), _key(program))
        if unique_key in seen: row_errors.append("Xodim qatori takrorlangan")
        if row_errors:
            errors.append({"varaq": "XODIMLAR", "qator": row_no, "xatolar": row_errors}); continue
        seen.add(unique_key)
        staff.append({
            "fish": fish, "telefon": phone, "rol": role,
            "fakultet": faculty, "kafedra": department, "yonalish": program,
            "ilmiy_daraja": _norm(_cell(row, headers, "Ilmiy daraja")) or None,
            "ilmiy_unvon": _norm(_cell(row, headers, "Ilmiy unvon")) or None,
            "staj_yil": experience,
            "mutaxassislik": _norm(_cell(row, headers, "Mutaxassislik")) or None,
            "qisqa_izoh": _norm(_cell(row, headers, "Qisqa izoh"))[:500] or None,
            "excel_row": row_no,
        })
    if not staff and not errors:
        errors.append({"varaq": "XODIMLAR", "qator": None, "xatolar": ["Kamida 1 ta F.I.Sh. kiriting"]})
    return staff, {"xodim_soni": len(staff), "xato_soni": len(errors), "xatolar": errors[:150]}


def _store_batch(cur, university_id: Optional[int], kind: str, filename: str, content: bytes, payload: Any, summary: dict, user_id: int) -> int:
    cur.execute("""INSERT INTO universitet_import_batchlar(
        universitet_id,import_turi,fayl_nomi,fayl_sha256,payload,xulosa,yaratilgan_by)
        VALUES(%s,%s,%s,%s,%s::jsonb,%s::jsonb,%s) RETURNING id""",
        (university_id, kind, filename, hashlib.sha256(content).hexdigest(), json.dumps(payload, ensure_ascii=False), json.dumps(summary, ensure_ascii=False), user_id))
    return int(cur.fetchone()["id"])


def _resolve_staff_named_scope(cur, university_id: int, item: dict[str, Any]) -> tuple[Optional[int], Optional[int], Optional[int]]:
    """Xodim shablonidagi nomlarni aynan bitta haqiqiy tuzilma zanjiriga yechadi."""
    faculty_id = department_id = program_id = None
    if item.get("fakultet"):
        cur.execute("SELECT id,nomi FROM fakultetlar WHERE universitet_id=%s AND faol=TRUE", (university_id,))
        matches = [row for row in cur.fetchall() if _key(row["nomi"]) == _key(item["fakultet"])]
        if len(matches) != 1:
            raise HTTPException(status_code=400, detail=f"Fakultet aniq topilmadi: {item['fakultet']}")
        faculty_id = int(matches[0]["id"])
    if item.get("kafedra"):
        cur.execute("""SELECT k.id,k.nomi,k.fakultet_id FROM kafedralar k
            JOIN fakultetlar f ON f.id=k.fakultet_id
            WHERE f.universitet_id=%s AND k.faol=TRUE""", (university_id,))
        matches = [row for row in cur.fetchall()
                   if _key(row["nomi"]) == _key(item["kafedra"])
                   and (faculty_id is None or int(row["fakultet_id"]) == faculty_id)]
        if len(matches) != 1:
            raise HTTPException(status_code=400, detail=f"Kafedra aniq topilmadi: {item['kafedra']}")
        department_id = int(matches[0]["id"])
        faculty_id = faculty_id or int(matches[0]["fakultet_id"])
    if item.get("yonalish"):
        cur.execute("""SELECT id,nomi,fakultet_id,kafedra_id FROM universitet_yonalishlari
            WHERE universitet_id=%s AND faol=TRUE""", (university_id,))
        matches = [row for row in cur.fetchall()
                   if _key(row["nomi"]) == _key(item["yonalish"])
                   and (faculty_id is None or int(row["fakultet_id"]) == faculty_id)
                   and (department_id is None or int(row["kafedra_id"]) == department_id)]
        if len(matches) != 1:
            raise HTTPException(status_code=400, detail=f"Yo'nalish aniq topilmadi: {item['yonalish']}")
        program_id = int(matches[0]["id"])
        faculty_id = faculty_id or int(matches[0]["fakultet_id"])
        department_id = department_id or int(matches[0]["kafedra_id"])
    return _normalized_role_scope(item["rol"], faculty_id, department_id, program_id)


class InstituteCreate(BaseModel):
    token: str
    nomi: str
    viloyat: Optional[str] = None
    tuman: Optional[str] = None


class DepartmentInput(BaseModel):
    nomi: str
    yonalishlar: list[str] = Field(default_factory=list)


class FacultyInput(BaseModel):
    nomi: str
    kafedralar: list[DepartmentInput] = Field(default_factory=list)


class StructureManual(BaseModel):
    token: str
    universitet_id: int
    fakultetlar: list[FacultyInput]


class StaffCreate(BaseModel):
    token: str
    universitet_id: int
    fish: str
    telefon: Optional[str] = None
    rol: str
    fakultet_id: Optional[int] = None
    kafedra_id: Optional[int] = None
    yonalish_id: Optional[int] = None
    ilmiy_daraja: Optional[str] = None
    ilmiy_unvon: Optional[str] = None
    staj_yil: Optional[float] = Field(default=None, ge=0, le=80)
    mutaxassislik: Optional[str] = None
    qisqa_izoh: Optional[str] = Field(default=None, max_length=500)


class StaffUpdate(BaseModel):
    token: str
    universitet_id: int
    fish: str
    rol: str
    fakultet_id: Optional[int] = None
    kafedra_id: Optional[int] = None
    yonalish_id: Optional[int] = None
    faol: bool = True
    ilmiy_daraja: Optional[str] = None
    ilmiy_unvon: Optional[str] = None
    staj_yil: Optional[float] = Field(default=None, ge=0, le=80)
    mutaxassislik: Optional[str] = None
    qisqa_izoh: Optional[str] = Field(default=None, max_length=500)


class TutorAssign(BaseModel):
    token: str
    universitet_id: int
    tyutor_user_id: int
    fakultet_id: Optional[int] = None
    yonalish_id: Optional[int] = None
    talim_shakli: Optional[str] = None
    talim_tili: Optional[str] = None
    qabul_turi: Optional[str] = None


class TutorAssignmentUpdate(BaseModel):
    token: str
    faol: bool = False


class InstituteToken(BaseModel):
    token: str


class BatchCommit(BaseModel):
    token: str
    batch_id: int
    default_kafedra_id: Optional[int] = None
    auto_create_yonalishlar: bool = False
    yonalish_mosliklari: dict[str, int] = Field(default_factory=dict)
    yangi_yonalish_kafedralari: dict[str, int] = Field(default_factory=dict)
    otm_nomi_farqini_tasdiqlash: bool = False


class StageUpdate(BaseModel):
    token: str
    bosqich: int


class ContactNoteUpdate(BaseModel):
    token: str
    izoh: str = ""


class StructureArchiveCommit(BaseModel):
    token: str
    universitet_id: int
    obyekt_turi: str
    obyekt_id: int
    tasdiq: bool = False


class StructureRestore(BaseModel):
    token: str
    universitet_id: int


class InviteSend(BaseModel):
    token: str
    kanal: str = "copy"  # copy | sms


class RedeemCode(BaseModel):
    token: str
    kirish_kodi: str


@router.post("/institut_yarat")
def create_institute(req: InstituteCreate):
    p = _p(); user_id = p._admin_tekshir(req.token)
    if not _norm(req.nomi):
        raise HTTPException(status_code=400, detail="Institut nomini kiriting")
    conn = p._db(); cur = conn.cursor()
    try:
        _ensure_schema(cur)
        cur.execute("INSERT INTO universitetlar(nomi,viloyat,tuman) VALUES(%s,%s,%s) RETURNING id", (_norm(req.nomi), _norm(req.viloyat) or None, _norm(req.tuman) or None))
        university_id = int(cur.fetchone()["id"])
        _assign_role(cur, university_id, user_id, "institut_admin", created_by=user_id)
        cur.execute("UPDATE users SET universitet_id=COALESCE(universitet_id,%s),lavozim=COALESCE(lavozim,'institut_admin') WHERE user_id=%s", (university_id, user_id))
        _audit(cur, university_id, user_id, "institut_yaratildi", "universitet", university_id)
        conn.commit()
        return {"holat": "yaratildi", "universitet_id": university_id}
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


@router.get("/super_admin/institutlar")
def super_admin_institutes(token: Optional[str] = Query(None, include_in_schema=False), authorization: Optional[str] = Header(None)):
    """Super-admin institut parolisiz mavjud institutni tanlaydi.

    Taklif kodi faqat oddiy xodim va talaba uchun. Global administratorga
    institutlar ro'yxati va yangi institut yaratish huquqi to'g'ridan-to'g'ri
    beriladi.
    """
    user_id = _uid(token, authorization); p = _p(); conn = p._db(); cur = conn.cursor()
    try:
        _ensure_schema(cur)
        if not _is_global_admin(cur, user_id):
            raise HTTPException(status_code=403, detail="Faqat super administrator uchun")
        cur.execute("""SELECT u.id,u.nomi,u.viloyat,u.tuman,
                       (SELECT COUNT(*) FROM fakultetlar f
                         WHERE f.universitet_id=u.id AND f.faol=TRUE) fakultet_soni,
                       (SELECT COUNT(*) FROM universitet_qabul_talabalari qt
                         WHERE qt.universitet_id=u.id) talaba_soni
                  FROM universitetlar u
                  LEFT JOIN universitet_workspace_map uwm ON uwm.universitet_id=u.id
                  LEFT JOIN organization_trials o ON o.context_id=uwm.context_id
                  LEFT JOIN learning_contexts c ON c.id=uwm.context_id
                 WHERE uwm.context_id IS NULL OR (
                       o.organization_type='institute'
                       AND c.context_type='university'
                       AND c.active=TRUE
                       AND LOWER(COALESCE(o.lifecycle_status,''))
                           IN ('trial','read_only','active')
                 )
                 ORDER BY u.nomi""")
        rows = cur.fetchall()
        conn.commit()
        return {"super_admin": True, "institutlar": rows}
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


@router.get("/bootstrap")
def bootstrap(workspace_id: Optional[int] = None, universitet_id: Optional[int] = None, token: Optional[str] = Query(None, include_in_schema=False), authorization: Optional[str] = Header(None)):
    user_id = _uid(token, authorization); p = _p(); conn = p._db(); cur = conn.cursor()
    try:
        _ensure_schema(cur)
        uid = int(universitet_id) if universitet_id else _resolve_university(cur, user_id, workspace_id, create=True)
        _require_active_university_source(cur, uid)
        cur.execute("SELECT id,nomi,viloyat,tuman FROM universitetlar WHERE id=%s", (uid,))
        university = cur.fetchone()
        if not university:
            raise HTTPException(status_code=404, detail="Institut topilmadi")
        roles = _require_member(cur, user_id, uid)
        cur.execute("SELECT COUNT(*) n FROM fakultetlar WHERE universitet_id=%s AND faol=TRUE", (uid,)); faculties = int(cur.fetchone()["n"])
        cur.execute("SELECT COUNT(*) n FROM universitet_yonalishlari WHERE universitet_id=%s AND faol=TRUE", (uid,)); programs = int(cur.fetchone()["n"])
        cur.execute("SELECT COUNT(*) n FROM universitet_qabul_talabalari WHERE universitet_id=%s", (uid,)); students = int(cur.fetchone()["n"])
        cur.execute("""SELECT
            COUNT(DISTINCT user_id) FILTER(WHERE rol NOT IN ('owner','institut_admin','fakultet_admin')) xodim,
            COUNT(DISTINCT user_id) FILTER(WHERE rol='professor_oqituvchi') oqituvchi,
            COUNT(DISTINCT user_id) FILTER(WHERE rol='tyutor') tyutor
            FROM universitet_xodim_rollari
            WHERE universitet_id=%s AND faol=TRUE""", (uid,))
        staff_counts = cur.fetchone()
        names = _role_names(roles); global_admin = _is_global_admin(cur, user_id)
        permissions = {
            "tuzilma_korish": True,
            "tuzilma_boshqarish": bool(names & MANAGE_STRUCTURE_ROLES),
            "xodim_korish": bool(names & (INSTITUTE_WIDE | FACULTY_WIDE | DEPARTMENT_WIDE)),
            "xodim_boshqarish": bool(names & MANAGE_STAFF_ROLES),
            "admin_boshqarish": global_admin,
            "qabul_korish": bool(names & PRIVATE_ROLES),
            "qabul_hisobot_korish": bool(names & REPORT_VIEW_ROLES),
            "hujjat_belgilash": bool(names & MARK_DOCUMENT_ROLES),
            "hemis_belgilash": bool(names & MARK_HEMIS_ROLES),
            "bazaga_belgilash": bool(names & MARK_DATABASE_ROLES),
            "saytga_kiritish": bool(names & MARK_DATABASE_ROLES),
            # V2256: rahbariyat (rektor, dekan, mudir, ma'naviyatchi...) hujjat holatini
            # to'liq ko'radi (topshirgan/topshirmagan, izoh, hisobot); belgilash huquqi alohida.
            "qabul_holatlari_toliq": bool(names & (MARK_DOCUMENT_ROLES | INSTITUTE_WIDE | FACULTY_WIDE | DEPARTMENT_WIDE)),
            "sayt_holati_korish": bool(names & (MARK_DOCUMENT_ROLES | INSTITUTE_WIDE | FACULTY_WIDE | DEPARTMENT_WIDE)),
            "parol_korish": bool(names & PASSWORD_VIEW_ROLES),
            "tyutor_korish": bool(names & (INSTITUTE_WIDE | FACULTY_WIDE | {"tyutor"})),
            "tyutor_boshqarish": bool(names & (INSTITUTE_WIDE | {"dekan", "fakultet_admin"})),
            "maxfiy_malumot": bool(names & PRIVATE_ROLES),
            "super_admin": global_admin,
        }
        conn.commit()
        return {"release": SAMTM_INSTITUTE_RELEASE, "universitet": university, "rollar": roles,
                "asosiy_rol": roles[0]["rol"], "ruxsatlar": permissions,
                "sonlar": {"fakultet": faculties, "yonalish": programs, "talaba": students,
                           "xodim": int(staff_counts["xodim"] or 0),
                           "oqituvchi": int(staff_counts["oqituvchi"] or 0),
                           "tyutor": int(staff_counts["tyutor"] or 0)}}
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


@router.get("/tuzilma")
def structure(universitet_id: int, token: Optional[str] = Query(None, include_in_schema=False), authorization: Optional[str] = Header(None)):
    user_id = _uid(token, authorization); p = _p(); conn = p._db(); cur = conn.cursor()
    try:
        _ensure_schema(cur); _require_member(cur, user_id, universitet_id)
        cur.execute("""SELECT f.id,f.nomi,
            (SELECT COUNT(*) FROM kafedralar k WHERE k.fakultet_id=f.id AND k.faol=TRUE) kafedra_soni,
            (SELECT COUNT(*) FROM universitet_yonalishlari y WHERE y.fakultet_id=f.id AND y.faol=TRUE) yonalish_soni
            FROM fakultetlar f WHERE f.universitet_id=%s AND f.faol=TRUE ORDER BY f.nomi""", (universitet_id,))
        faculties = [dict(r) for r in cur.fetchall()]
        for f in faculties:
            cur.execute("SELECT id,nomi FROM kafedralar WHERE fakultet_id=%s AND faol=TRUE ORDER BY nomi", (f["id"],))
            f["kafedralar"] = [dict(r) for r in cur.fetchall()]
            for d in f["kafedralar"]:
                cur.execute("""SELECT y.id,y.kodi,y.nomi,y.daraja,
                    COUNT(qt.id) AS talaba_soni,
                    COALESCE(ARRAY_AGG(DISTINCT qt.talim_shakli) FILTER(WHERE qt.talim_shakli IS NOT NULL), ARRAY[]::TEXT[]) AS talim_shakllari,
                    COALESCE(ARRAY_AGG(DISTINCT qt.talim_tili) FILTER(WHERE qt.talim_tili IS NOT NULL), ARRAY[]::TEXT[]) AS talim_tillari
                    FROM universitet_yonalishlari y
                    LEFT JOIN universitet_qabul_talabalari qt ON qt.yonalish_id=y.id
                    WHERE y.kafedra_id=%s AND y.faol=TRUE
                    GROUP BY y.id ORDER BY y.nomi""", (d["id"],))
                d["yonalishlar"] = cur.fetchall()
                cur.execute("""SELECT xr.user_id,u.full_name FROM universitet_xodim_rollari xr
                    JOIN users u ON u.user_id=xr.user_id
                    WHERE xr.kafedra_id=%s AND xr.rol='kafedra_mudiri' AND xr.faol=TRUE LIMIT 1""", (d["id"],))
                d["mudir"] = cur.fetchone()
            cur.execute("""SELECT xr.id,xr.user_id,xr.rol,u.full_name
                FROM universitet_xodim_rollari xr JOIN users u ON u.user_id=xr.user_id
                WHERE xr.fakultet_id=%s AND xr.faol=TRUE ORDER BY xr.rol,u.full_name""", (f["id"],))
            f["rahbariyat"] = cur.fetchall()
            counts = {role: sum(1 for x in f["rahbariyat"] if x["rol"] == role) for role in ("dekan", "zam_dekan", "manaviyatchi", "fakultet_admin")}
            f["toldirilish"] = {"dekan": counts["dekan"], "zam_dekan": counts["zam_dekan"], "manaviyatchi": counts["manaviyatchi"], "admin": counts["fakultet_admin"],
                                "tayyor": counts["dekan"] == 1 and counts["zam_dekan"] == 2 and counts["manaviyatchi"] == 1 and counts["fakultet_admin"] >= 1}
            # Kafedra va yo'nalishlar DBdagi haqiqiy ota-ona IDsi bilan
            # qaytariladi. Nom bir xil bo'lsa ham turli kafedralardagi obyektlar
            # birlashtirilmaydi va boshqa kafedraga ko'chirib ko'rsatilmaydi.
            f["kafedralar"] = sorted(f["kafedralar"], key=lambda item: _key(item["nomi"]))
            for department in f["kafedralar"]:
                department["yonalishlar"] = sorted(
                    [dict(program) for program in department["yonalishlar"]],
                    key=lambda item: (_key(item["nomi"]), int(item["id"])),
                )
            f["kafedra_soni"] = len(f["kafedralar"])
            f["yonalish_soni"] = sum(
                len(department["yonalishlar"])
                for department in f["kafedralar"]
            )
        cur.execute("""SELECT xr.rol,xr.user_id,u.full_name FROM universitet_xodim_rollari xr
            JOIN users u ON u.user_id=xr.user_id
            WHERE xr.universitet_id=%s AND xr.fakultet_id IS NULL AND xr.faol=TRUE
              AND xr.rol IN ('owner','rektor','prorektor','institut_admin')
            ORDER BY xr.rol,u.full_name""", (universitet_id,))
        return {"fakultetlar": faculties, "institut_rahbariyat": cur.fetchall(),
                "talim_shakllari": TA_LIM_SHAKLLARI, "talim_tillari": TA_LIM_TILLARI, "darajalar": DARAJALAR}
    finally:
        cur.close(); conn.close()


@router.get("/fakultet/{faculty_id}/yonalishlar")
def faculty_programs(faculty_id: int, token: Optional[str] = Query(None, include_in_schema=False), authorization: Optional[str] = Header(None)):
    """Muassasalar tashqi oynasi uchun haqiqiy yo'nalishlar va talaba soni."""
    actor = _uid(token, authorization); p = _p(); conn = p._db(); cur = conn.cursor()
    try:
        _ensure_schema(cur)
        cur.execute("SELECT universitet_id,nomi FROM fakultetlar WHERE id=%s", (faculty_id,)); faculty = cur.fetchone()
        if not faculty: raise HTTPException(status_code=404, detail="Fakultet topilmadi")
        _require_member(cur, actor, faculty["universitet_id"])
        cur.execute("""SELECT y.id,y.nomi,y.daraja,y.kodi,y.kafedra_id,k.nomi kafedra_nomi,
                   COUNT(qt.id) AS talaba_soni,
                   COUNT(qt.id) FILTER(WHERE qt.hujjat_topshirgan_at IS NOT NULL) AS hujjat_soni,
                   COUNT(qt.id) FILTER(WHERE qt.bazaga_kiritilgan_at IS NOT NULL) AS baza_soni,
                   COALESCE(STRING_AGG(DISTINCT qt.talim_shakli, ', '), '') AS talim_shakllari,
                   COALESCE(STRING_AGG(DISTINCT qt.talim_tili, ', '), '') AS talim_tillari
            FROM universitet_yonalishlari y
            JOIN kafedralar k ON k.id=y.kafedra_id
            LEFT JOIN universitet_qabul_talabalari qt ON qt.yonalish_id=y.id
            WHERE y.fakultet_id=%s AND y.faol=TRUE
            GROUP BY y.id,k.nomi ORDER BY y.nomi""", (faculty_id,))
        programs = [dict(row) for row in cur.fetchall()]
        programs.sort(key=lambda item: (_key(item["kafedra_nomi"]), _key(item["nomi"]), int(item["id"])))
        return {"fakultet_id": faculty_id, "fakultet_nomi": faculty["nomi"],
                "yonalishlar": programs}
    finally:
        cur.close(); conn.close()


@router.get("/tuzilma/shablon")
def structure_template(universitet_id: int, token: Optional[str] = Query(None, include_in_schema=False), authorization: Optional[str] = Header(None)):
    user_id = _uid(token, authorization); p = _p(); conn = p._db(); cur = conn.cursor()
    try:
        _ensure_schema(cur); roles = _require_member(cur, user_id, universitet_id)
        if not _has_any(roles, MANAGE_STRUCTURE_ROLES): raise HTTPException(status_code=403, detail="Shablon yuklash huquqi yo'q")
        cur.execute("SELECT id,nomi,viloyat,tuman FROM universitetlar WHERE id=%s", (universitet_id,)); university = cur.fetchone()
        if not university: raise HTTPException(status_code=404, detail="Institut topilmadi")
        cur.execute("SELECT id,nomi FROM fakultetlar WHERE universitet_id=%s ORDER BY nomi", (universitet_id,)); faculties = [dict(row) for row in cur.fetchall()]
        cur.execute("""SELECT k.id,k.nomi kafedra,f.nomi fakultet FROM kafedralar k
            JOIN fakultetlar f ON f.id=k.fakultet_id WHERE f.universitet_id=%s ORDER BY f.nomi,k.nomi""", (universitet_id,)); departments = [dict(row) for row in cur.fetchall()]
        cur.execute("""SELECT f.nomi fakultet,k.nomi kafedra,y.kodi,y.nomi yonalish,y.daraja,
                   v.talim_shakli,v.talim_tili
            FROM universitet_yonalishlari y JOIN fakultetlar f ON f.id=y.fakultet_id
            JOIN kafedralar k ON k.id=y.kafedra_id
            LEFT JOIN universitet_yonalish_variantlari v ON v.yonalish_id=y.id AND v.faol=TRUE
            WHERE y.universitet_id=%s AND y.faol=TRUE ORDER BY f.nomi,k.nomi,y.nomi,v.talim_shakli,v.talim_tili""", (universitet_id,))
        rows = [dict(row) for row in cur.fetchall()]
        used_departments = {(_key(row["fakultet"]), _key(row["kafedra"])) for row in rows}
        for department in departments:
            if (_key(department["fakultet"]), _key(department["kafedra"])) not in used_departments:
                rows.append({"fakultet": department["fakultet"], "kafedra": department["kafedra"], "kodi": None,
                             "yonalish": None, "daraja": "Bakalavriat", "talim_shakli": "Kunduzgi", "talim_tili": "O‘zbekcha"})
        content = _structure_template_xlsx(dict(university), rows, faculties, departments)
        safe_name = re.sub(r"[^A-Za-z0-9_-]+", "_", _norm(university["nomi"]))[:60] or "institut"
        return Response(content=content, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition": f'attachment; filename="{safe_name}_tuzilma_xodimlar.xlsx"'})
    finally:
        cur.close(); conn.close()


@router.get("/xodim/shablon")
def staff_template(universitet_id: int,
                   token: Optional[str] = Query(None, include_in_schema=False),
                   authorization: Optional[str] = Header(None)):
    """Admin uchun faqat o'z qamrovidagi tuzilma bilan XODIMLAR XLSX shabloni."""
    actor = _uid(token, authorization); p = _p(); conn = p._db(); cur = conn.cursor()
    try:
        _ensure_schema(cur); roles = _require_member(cur, actor, universitet_id)
        if not _has_any(roles, MANAGE_STAFF_ROLES):
            raise HTTPException(status_code=403, detail="Xodim shablonini yuklash huquqi yo'q")
        cur.execute("SELECT id,nomi,viloyat,tuman FROM universitetlar WHERE id=%s", (universitet_id,))
        university = cur.fetchone()
        if not university: raise HTTPException(status_code=404, detail="Institut topilmadi")
        names = _role_names(roles)
        faculty_scope = None if names & INSTITUTE_WIDE else sorted({
            int(role["fakultet_id"]) for role in roles
            if role.get("fakultet_id") and role["rol"] in {"dekan", "fakultet_admin"}
        })
        if faculty_scope is not None and not faculty_scope:
            raise HTTPException(status_code=403, detail="Xodim shabloni uchun fakultet qamrovi topilmadi")
        faculty_where = "universitet_id=%s"; faculty_params: list[Any] = [universitet_id]
        if faculty_scope is not None:
            faculty_where += " AND id=ANY(%s)"; faculty_params.append(faculty_scope)
        cur.execute(f"SELECT id,nomi FROM fakultetlar WHERE {faculty_where} AND faol=TRUE ORDER BY nomi", faculty_params)
        faculties = [dict(row) for row in cur.fetchall()]
        faculty_ids = [int(item["id"]) for item in faculties]
        cur.execute("""SELECT k.id,k.nomi kafedra,f.nomi fakultet
            FROM kafedralar k JOIN fakultetlar f ON f.id=k.fakultet_id
            WHERE f.universitet_id=%s AND f.id=ANY(%s) AND k.faol=TRUE
            ORDER BY f.nomi,k.nomi""", (universitet_id, faculty_ids or [-1]))
        departments = [dict(row) for row in cur.fetchall()]
        cur.execute("""SELECT f.nomi fakultet,k.nomi kafedra,y.kodi,y.nomi yonalish,y.daraja,
                   v.talim_shakli,v.talim_tili
            FROM universitet_yonalishlari y
            JOIN fakultetlar f ON f.id=y.fakultet_id JOIN kafedralar k ON k.id=y.kafedra_id
            LEFT JOIN universitet_yonalish_variantlari v ON v.yonalish_id=y.id AND v.faol=TRUE
            WHERE y.universitet_id=%s AND y.fakultet_id=ANY(%s) AND y.faol=TRUE
            ORDER BY f.nomi,k.nomi,y.nomi""", (universitet_id, faculty_ids or [-1]))
        rows = [dict(row) for row in cur.fetchall()]
        content = _structure_template_xlsx(dict(university), rows, faculties, departments)
        return Response(content=content,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition": 'attachment; filename="institut_xodimlari_shabloni.xlsx"'})
    finally:
        cur.close(); conn.close()


@router.post("/xodim/import_preview")
async def staff_import_preview(universitet_id: int, fayl: UploadFile = File(...),
                               token: Optional[str] = Query(None, include_in_schema=False),
                               authorization: Optional[str] = Header(None)):
    actor = _uid(token, authorization); content = await fayl.read(); filename = fayl.filename or "xodimlar.xlsx"
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="Xodimlar fayli 10 MB dan katta")
    if not filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Xodimlar uchun XLSX yoki XLS fayl yuklang")
    staff, summary = _parse_staff_template(content, filename)
    p = _p(); conn = p._db(); cur = conn.cursor()
    try:
        _ensure_schema(cur); roles = _require_member(cur, actor, universitet_id)
        if not _has_any(roles, MANAGE_STAFF_ROLES):
            raise HTTPException(status_code=403, detail="Xodim importiga ruxsat yo'q")
        errors = list(summary["xatolar"]); resolved = []
        for item in staff:
            try:
                if item["rol"] in ADMIN_ROLES and not _is_global_admin(cur, actor):
                    raise HTTPException(status_code=403, detail="Administratorni faqat super administrator import qiladi")
                if not (_role_names(roles) & INSTITUTE_WIDE) and item["rol"] in INSTITUTE_WIDE:
                    raise HTTPException(status_code=403, detail="Institut rahbariyatini fakultet administratori import qilmaydi")
                faculty_id, department_id, program_id = _resolve_staff_named_scope(cur, universitet_id, item)
                _validate_assignment_scope(cur, universitet_id, roles, faculty_id, department_id, program_id)
                resolved.append({**item, "fakultet_id": faculty_id,
                                 "kafedra_id": department_id, "yonalish_id": program_id})
            except HTTPException as exc:
                errors.append({"varaq": "XODIMLAR", "qator": item["excel_row"],
                               "xatolar": [str(exc.detail)]})
        summary.update({"xato_soni": len(errors), "xatolar": errors[:150],
                        "xodim_soni": len(resolved)})
        batch_id = _store_batch(cur, universitet_id, "xodim", filename, content,
                                {"xodimlar": resolved}, summary, actor)
        conn.commit()
        return {"batch_id": batch_id, "xulosa": summary,
                "commit_mumkin": summary["xato_soni"] == 0}
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


@router.post("/xodim/import_commit")
def staff_import_commit(req: BatchCommit):
    p = _p(); actor = p._jwt_tekshir(req.token); conn = p._db(); cur = conn.cursor()
    try:
        _ensure_schema(cur)
        cur.execute("SELECT * FROM universitet_import_batchlar WHERE id=%s FOR UPDATE", (req.batch_id,))
        batch = cur.fetchone()
        if not batch or batch["import_turi"] != "xodim":
            raise HTTPException(status_code=404, detail="Xodim preview topilmadi")
        if batch["holat"] != "preview":
            raise HTTPException(status_code=409, detail="Bu preview avval ishlatilgan")
        if int(batch["yaratilgan_by"]) != actor and not _is_global_admin(cur, actor):
            raise HTTPException(status_code=403, detail="Bu preview boshqa foydalanuvchiga tegishli")
        university_id = int(batch["universitet_id"])
        roles = _require_member(cur, actor, university_id)
        if not _has_any(roles, MANAGE_STAFF_ROLES):
            raise HTTPException(status_code=403, detail="Xodim importiga ruxsat yo'q")
        summary = batch["xulosa"] if isinstance(batch["xulosa"], dict) else json.loads(batch["xulosa"])
        if summary.get("xato_soni"):
            raise HTTPException(status_code=400, detail="Xatoli xodimlar shabloni import qilinmaydi")
        payload = batch["payload"] if isinstance(batch["payload"], dict) else json.loads(batch["payload"])
        credentials = []
        for item in payload["xodimlar"]:
            if item["rol"] in ADMIN_ROLES and not _is_global_admin(cur, actor):
                raise HTTPException(status_code=403, detail="Administratorni faqat super administrator import qiladi")
            faculty_id, department_id, program_id = _resolve_staff_named_scope(cur, university_id, item)
            _validate_assignment_scope(cur, university_id, roles, faculty_id, department_id, program_id)
            user_id, role_id, code = _new_placeholder(
                cur, item["fish"], university_id, item["rol"], item.get("telefon"), actor,
                faculty_id, department_id, program_id,
            )
            _save_staff_profile(
                cur, role_id, ilmiy_daraja=item.get("ilmiy_daraja"),
                ilmiy_unvon=item.get("ilmiy_unvon"), staj_yil=item.get("staj_yil"),
                mutaxassislik=item.get("mutaxassislik"), qisqa_izoh=item.get("qisqa_izoh"),
            )
            if item["rol"] == "tyutor" and program_id:
                cur.execute("""INSERT INTO universitet_tyutor_yonalishlari(
                    universitet_id,tyutor_user_id,fakultet_id,yonalish_id,yaratilgan_by)
                    VALUES(%s,%s,%s,%s,%s) ON CONFLICT DO NOTHING""",
                    (university_id, user_id, faculty_id, program_id, actor))
            credentials.append({"fish": item["fish"], "lavozim": ROLE_LABELS[item["rol"]],
                                "fakultet": item.get("fakultet"), "kafedra": item.get("kafedra"),
                                "yonalish": item.get("yonalish"), "telefon": item.get("telefon"),
                                "kirish_kodi": code, "kod_muddati": "2 oy"})
        cur.execute("UPDATE universitet_import_batchlar SET holat='committed',commit_at=NOW() WHERE id=%s",
                    (req.batch_id,))
        _audit(cur, university_id, actor, "xodim_import_commit", "import_batch",
               req.batch_id, {"xodim": len(credentials)})
        credential_file = _credentials_xlsx(
            credentials, "institut_xodimlari_kirish_kodlari.xlsx",
            "INSTITUT XODIMLARI — BIR MARTALIK KIRISH KODLARI",
        ) if credentials else None
        conn.commit()
        return {"holat": "import_qilindi", "xodim": len(credentials),
                "kirish_kodlari": credentials, "kirish_kodlari_fayli": credential_file}
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


@router.post("/tuzilma/manual")
def manual_structure(req: StructureManual):
    p = _p(); user_id = p._jwt_tekshir(req.token); conn = p._db(); cur = conn.cursor()
    try:
        _ensure_schema(cur); roles = _require_member(cur, user_id, req.universitet_id)
        if not _has_any(roles, MANAGE_STRUCTURE_ROLES): raise HTTPException(status_code=403, detail="Tuzilmani boshqarish huquqi yo'q")
        if not req.fakultetlar: raise HTTPException(status_code=400, detail="Kamida 1 ta fakultet kiriting")
        created = {"fakultet": 0, "kafedra": 0, "yonalish": 0}
        for f in req.fakultetlar:
            fname = _norm(f.nomi)
            if not fname: raise HTTPException(status_code=400, detail="Fakultet nomi bo'sh")
            cur.execute("SELECT id FROM fakultetlar WHERE universitet_id=%s AND LOWER(nomi)=LOWER(%s)", (req.universitet_id, fname)); fr = cur.fetchone()
            if fr: faculty_id = int(fr["id"])
            else:
                cur.execute("INSERT INTO fakultetlar(universitet_id,nomi) VALUES(%s,%s) RETURNING id", (req.universitet_id, fname)); faculty_id = int(cur.fetchone()["id"]); created["fakultet"] += 1
                if _is_global_admin(cur, user_id):
                    _assign_role(cur, req.universitet_id, user_id, "fakultet_admin", faculty_id=faculty_id, created_by=user_id)
            for d in f.kafedralar:
                dname = _norm(d.nomi)
                if not dname: raise HTTPException(status_code=400, detail=f"{fname}: kafedra nomi bo'sh")
                cur.execute("SELECT id FROM kafedralar WHERE fakultet_id=%s AND LOWER(nomi)=LOWER(%s)", (faculty_id, dname)); dr = cur.fetchone()
                if dr: department_id = int(dr["id"])
                else:
                    cur.execute("INSERT INTO kafedralar(fakultet_id,nomi) VALUES(%s,%s) RETURNING id", (faculty_id, dname)); department_id = int(cur.fetchone()["id"]); created["kafedra"] += 1
                for program in d.yonalishlar:
                    pname = _norm(program)
                    if not pname: continue
                    cur.execute("""INSERT INTO universitet_yonalishlari(universitet_id,fakultet_id,kafedra_id,nomi)
                        VALUES(%s,%s,%s,%s) ON CONFLICT DO NOTHING RETURNING id""", (req.universitet_id, faculty_id, department_id, pname))
                    if cur.fetchone(): created["yonalish"] += 1
        _audit(cur, req.universitet_id, user_id, "tuzilma_qolda_saqlandi", detail=created)
        conn.commit(); return {"holat": "saqlandi", **created}
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


@router.post("/xodim/manual")
def manual_staff(req: StaffCreate):
    p = _p(); user_id = p._jwt_tekshir(req.token); conn = p._db(); cur = conn.cursor()
    try:
        _ensure_schema(cur); roles = _require_member(cur, user_id, req.universitet_id)
        if not _has_any(roles, MANAGE_STAFF_ROLES): raise HTTPException(status_code=403, detail="Xodim biriktirish huquqi yo'q")
        if req.rol in ADMIN_ROLES and not _is_global_admin(cur, user_id):
            raise HTTPException(status_code=403, detail="Institut va fakultet administratorini faqat super administrator qo'shadi")
        if not (_role_names(roles) & INSTITUTE_WIDE) and req.rol in INSTITUTE_WIDE:
            raise HTTPException(status_code=403, detail="Institut rahbari yoki adminini faqat institut administratori qo'shadi")
        faculty_id, department_id, program_id = _normalized_role_scope(
            req.rol, req.fakultet_id, req.kafedra_id, req.yonalish_id
        )
        _validate_assignment_scope(cur, req.universitet_id, roles, faculty_id, department_id, program_id)
        fish = _norm(req.fish); phone = _telefon(req.telefon) if req.telefon else None
        if not fish: raise HTTPException(status_code=400, detail="F.I.Sh. kiriting")
        if req.telefon and not phone: raise HTTPException(status_code=400, detail="Telefon +998 bilan to'g'ri yozilsin")
        placeholder, role_id, code = _new_placeholder(
            cur, fish, req.universitet_id, req.rol, phone, user_id,
            faculty_id, department_id, program_id
        )
        _save_staff_profile(
            cur, role_id, ilmiy_daraja=req.ilmiy_daraja,
            ilmiy_unvon=req.ilmiy_unvon, staj_yil=req.staj_yil,
            mutaxassislik=req.mutaxassislik, qisqa_izoh=req.qisqa_izoh,
        )
        _audit(cur, req.universitet_id, user_id, "xodim_qoshildi", "xodim_rol", role_id, {"rol": req.rol})
        conn.commit()
        return {"holat": "yaratildi", "user_id": placeholder, "rol_id": role_id, "fish": fish,
                "lavozim": ROLE_LABELS[req.rol], "kirish_kodi": code, "kod_muddati": "2 oy"}
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


@router.patch("/xodim/{role_id}")
def update_staff(role_id: int, req: StaffUpdate):
    """Rahbar/xodim F.I.Sh., lavozimi va qamrovini haqiqiy bazada tahrirlaydi."""
    p = _p(); actor = p._jwt_tekshir(req.token); conn = p._db(); cur = conn.cursor()
    try:
        _ensure_schema(cur); roles = _require_member(cur, actor, req.universitet_id)
        if not _has_any(roles, MANAGE_STAFF_ROLES):
            raise HTTPException(status_code=403, detail="Xodimni tahrirlash huquqi yo'q")
        cur.execute("SELECT * FROM universitet_xodim_rollari WHERE id=%s AND universitet_id=%s FOR UPDATE", (role_id, req.universitet_id))
        current = cur.fetchone()
        if not current: raise HTTPException(status_code=404, detail="Xodim biriktirishi topilmadi")
        if req.rol not in ROLE_LABELS or req.rol == "talaba":
            raise HTTPException(status_code=400, detail=f"Noto'g'ri lavozim: {req.rol}")
        if (req.rol in ADMIN_ROLES or current["rol"] in ADMIN_ROLES) and not _is_global_admin(cur, actor):
            raise HTTPException(status_code=403, detail="Administratorni faqat super administrator tahrirlaydi")
        if not (_role_names(roles) & INSTITUTE_WIDE) and req.rol in INSTITUTE_WIDE:
            raise HTTPException(status_code=403, detail="Institut rahbariyatini faqat institut administratori tahrirlaydi")
        faculty_id, department_id, program_id = _normalized_role_scope(
            req.rol, req.fakultet_id, req.kafedra_id, req.yonalish_id
        )
        if req.rol in FACULTY_WIDE and not faculty_id:
            raise HTTPException(status_code=400, detail="Bu lavozim uchun fakultet tanlanishi shart")
        if req.rol in DEPARTMENT_WIDE and not department_id:
            raise HTTPException(status_code=400, detail="Kafedra mudiri uchun kafedra tanlanishi shart")
        _validate_assignment_scope(cur, req.universitet_id, roles, faculty_id, department_id, program_id)
        if req.faol:
            singleton_sql, singleton_params = None, []
            if req.rol == "rektor":
                singleton_sql, singleton_params = "universitet_id=%s AND rol='rektor'", [req.universitet_id]
            elif req.rol in {"dekan", "manaviyatchi"}:
                singleton_sql, singleton_params = "fakultet_id=%s AND rol=%s", [faculty_id, req.rol]
            elif req.rol == "kafedra_mudiri":
                singleton_sql, singleton_params = "kafedra_id=%s AND rol='kafedra_mudiri'", [department_id]
            if singleton_sql:
                cur.execute(f"""SELECT 1 FROM universitet_xodim_rollari
                    WHERE {singleton_sql} AND faol=TRUE AND id<>%s LIMIT 1""",
                    singleton_params + [role_id])
                if cur.fetchone():
                    raise HTTPException(status_code=409, detail="Bu lavozimga boshqa xodim allaqachon biriktirilgan")
            if req.rol == "zam_dekan":
                cur.execute("""SELECT COUNT(*) n FROM universitet_xodim_rollari
                    WHERE fakultet_id=%s AND rol='zam_dekan' AND faol=TRUE AND id<>%s""",
                    (faculty_id, role_id))
                if int(cur.fetchone()["n"] or 0) >= 2:
                    raise HTTPException(status_code=409, detail="Bu fakultetda 2 ta dekan o'rinbosari allaqachon bor")
        fish = _norm(req.fish)
        if not fish: raise HTTPException(status_code=400, detail="F.I.Sh. kiriting")
        cur.execute("UPDATE users SET full_name=%s WHERE user_id=%s", (fish, current["user_id"]))
        # Eski rahbar qamrovi o'zgarsa legacy Muassasalar kartasida arvoh
        # rahbar qolmasin; yangi qamrov quyidagi _sync bilan qayta yoziladi.
        if current["rol"] == "rektor" and (not req.faol or req.rol != "rektor"):
            cur.execute("UPDATE universitetlar SET rektor_user_id=NULL WHERE id=%s AND rektor_user_id=%s",
                        (req.universitet_id, current["user_id"]))
        elif current["rol"] == "dekan" and current.get("fakultet_id") and (
                not req.faol or req.rol != "dekan" or current["fakultet_id"] != faculty_id):
            cur.execute("UPDATE fakultetlar SET dekan_user_id=NULL WHERE id=%s AND dekan_user_id=%s",
                        (current["fakultet_id"], current["user_id"]))
        elif current["rol"] == "kafedra_mudiri" and current.get("kafedra_id") and (
                not req.faol or req.rol != "kafedra_mudiri" or current["kafedra_id"] != department_id):
            cur.execute("UPDATE kafedralar SET mudir_user_id=NULL WHERE id=%s AND mudir_user_id=%s",
                        (current["kafedra_id"], current["user_id"]))
        cur.execute("""UPDATE universitet_xodim_rollari
            SET rol=%s,fakultet_id=%s,kafedra_id=%s,yonalish_id=%s,faol=%s
            WHERE id=%s""", (req.rol, faculty_id, department_id, program_id, req.faol, role_id))
        if req.faol:
            _sync_legacy_leader(cur, req.universitet_id, int(current["user_id"]), req.rol,
                                faculty_id, department_id)
        profile_fields = {"ilmiy_daraja", "ilmiy_unvon", "staj_yil", "mutaxassislik", "qisqa_izoh"}
        supplied_fields = set(getattr(req, "model_fields_set", getattr(req, "__fields_set__", set())))
        if supplied_fields & profile_fields:
            cur.execute("SELECT * FROM universitet_xodim_profili WHERE xodim_rol_id=%s", (role_id,))
            old_profile = dict(cur.fetchone() or {})
            profile_values = {
                field: getattr(req, field) if field in supplied_fields else old_profile.get(field)
                for field in profile_fields
            }
            _save_staff_profile(cur, role_id, force=True, **profile_values)
        _audit(cur, req.universitet_id, actor, "xodim_tahrirlandi", "xodim_rol", role_id,
               {"rol": req.rol, "faol": req.faol})
        conn.commit(); return {"holat": "saqlandi", "rol_id": role_id, "fish": fish}
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


@router.post("/tuzilma/import_preview")
async def structure_preview(universitet_id: int, fayl: UploadFile = File(...), token: Optional[str] = Query(None, include_in_schema=False), authorization: Optional[str] = Header(None)):
    user_id = _uid(token, authorization); p = _p()
    content = await fayl.read()
    if len(content) > 15 * 1024 * 1024: raise HTTPException(status_code=413, detail="Tuzilma fayli 15 MB dan katta")
    if not (fayl.filename or "").lower().endswith(".xlsx"): raise HTTPException(status_code=400, detail="Institut tuzilmasi uchun .xlsx shablondan foydalaning")
    payload, summary = _parse_structure(content, fayl.filename or "institut_tuzilma.xlsx")
    conn = p._db(); cur = conn.cursor()
    try:
        _ensure_schema(cur); roles = _require_member(cur, user_id, universitet_id)
        if not _has_any(roles, MANAGE_STRUCTURE_ROLES): raise HTTPException(status_code=403, detail="Tuzilma importiga ruxsat yo'q")
        cur.execute("SELECT nomi FROM universitetlar WHERE id=%s", (universitet_id,)); university = cur.fetchone()
        if not university: raise HTTPException(status_code=404, detail="Institut topilmadi")
        summary["nom_mosligi"] = _key(university["nomi"]) == _key(payload["institut"]["nomi"])
        if not summary["nom_mosligi"]:
            summary["xatolar"].append({"varaq": "INSTITUT", "qator": 2, "xatolar": [f"Institut nomi mos emas: saytda '{university['nomi']}'"]})
            summary["xato_soni"] += 1
        batch_id = _store_batch(cur, universitet_id, "tuzilma", fayl.filename or "institut_tuzilma.xlsx", content, payload, summary, user_id)
        conn.commit(); return {"batch_id": batch_id, "xulosa": summary, "commit_mumkin": summary["xato_soni"] == 0}
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


@router.post("/tuzilma/import_commit")
def structure_commit(req: BatchCommit):
    p = _p(); actor = p._jwt_tekshir(req.token); conn = p._db(); cur = conn.cursor()
    try:
        _ensure_schema(cur)
        cur.execute("SELECT * FROM universitet_import_batchlar WHERE id=%s FOR UPDATE", (req.batch_id,)); batch = cur.fetchone()
        if not batch or batch["import_turi"] != "tuzilma": raise HTTPException(status_code=404, detail="Tuzilma preview topilmadi")
        if batch["holat"] != "preview": raise HTTPException(status_code=409, detail="Bu preview avval ishlatilgan")
        if batch["yaratilgan_by"] != actor and not _is_global_admin(cur, actor): raise HTTPException(status_code=403, detail="Bu preview boshqa foydalanuvchiga tegishli")
        roles = _require_member(cur, actor, batch["universitet_id"])
        if not _has_any(roles, MANAGE_STRUCTURE_ROLES): raise HTTPException(status_code=403, detail="Tuzilma importiga ruxsat yo'q")
        summary = batch["xulosa"] if isinstance(batch["xulosa"], dict) else json.loads(batch["xulosa"])
        if summary.get("xato_soni"): raise HTTPException(status_code=400, detail="Xatoli shablon commit qilinmaydi")
        payload = batch["payload"] if isinstance(batch["payload"], dict) else json.loads(batch["payload"])
        university_id = int(batch["universitet_id"])
        cur.execute("UPDATE universitetlar SET viloyat=COALESCE(%s,viloyat),tuman=COALESCE(%s,tuman) WHERE id=%s",
                    (payload["institut"].get("viloyat"), payload["institut"].get("tuman"), university_id))
        faculty_map, department_map, program_map = {}, {}, {}
        counts = {"fakultet": 0, "kafedra": 0, "yonalish": 0, "variant": 0, "xodim": 0}
        for item in payload["tuzilma"]:
            fk = _key(item["fakultet"])
            if fk not in faculty_map:
                cur.execute("SELECT id FROM fakultetlar WHERE universitet_id=%s AND LOWER(nomi)=LOWER(%s)", (university_id, item["fakultet"])); row = cur.fetchone()
                if row: faculty_id = int(row["id"])
                else:
                    cur.execute("INSERT INTO fakultetlar(universitet_id,nomi) VALUES(%s,%s) RETURNING id", (university_id, item["fakultet"])); faculty_id = int(cur.fetchone()["id"]); counts["fakultet"] += 1
                faculty_map[fk] = faculty_id
                # Administrator rolini faqat global super admin beradi.
                if _is_global_admin(cur, actor):
                    _assign_role(cur, university_id, actor, "fakultet_admin", faculty_id=faculty_id, created_by=actor)
            faculty_id = faculty_map[fk]
            dk = (fk, _key(item["kafedra"]))
            if dk not in department_map:
                cur.execute("SELECT id FROM kafedralar WHERE fakultet_id=%s AND LOWER(nomi)=LOWER(%s)", (faculty_id, item["kafedra"])); row = cur.fetchone()
                if row: department_id = int(row["id"])
                else:
                    cur.execute("INSERT INTO kafedralar(fakultet_id,nomi) VALUES(%s,%s) RETURNING id", (faculty_id, item["kafedra"])); department_id = int(cur.fetchone()["id"]); counts["kafedra"] += 1
                department_map[dk] = department_id
            department_id = department_map[dk]
            pk = (dk, _key(item["yonalish"]), item["daraja"])
            if pk not in program_map:
                cur.execute("""SELECT id FROM universitet_yonalishlari WHERE universitet_id=%s AND kafedra_id=%s AND LOWER(nomi)=LOWER(%s) AND daraja=%s""",
                            (university_id, department_id, item["yonalish"], item["daraja"])); row = cur.fetchone()
                if row: program_id = int(row["id"])
                else:
                    cur.execute("""INSERT INTO universitet_yonalishlari(universitet_id,fakultet_id,kafedra_id,kodi,nomi,daraja)
                        VALUES(%s,%s,%s,%s,%s,%s) RETURNING id""", (university_id, faculty_id, department_id, item.get("yonalish_kodi"), item["yonalish"], item["daraja"])); program_id = int(cur.fetchone()["id"]); counts["yonalish"] += 1
                program_map[pk] = program_id
            program_id = program_map[pk]
            cur.execute("""INSERT INTO universitet_yonalish_variantlari(yonalish_id,talim_shakli,talim_tili)
                VALUES(%s,%s,%s) ON CONFLICT DO NOTHING RETURNING id""", (program_id, item["talim_shakli"], item["talim_tili"]))
            if cur.fetchone(): counts["variant"] += 1

        credentials = []
        for person in payload["xodimlar"]:
            role = person["rol"]
            if role in ADMIN_ROLES and not _is_global_admin(cur, actor):
                raise HTTPException(status_code=403, detail="Shablondagi administratorlarni faqat super administrator import qiladi")
            faculty_id = faculty_map.get(_key(person.get("fakultet"))) if person.get("fakultet") else None
            department_id = department_map.get((_key(person.get("fakultet")), _key(person.get("kafedra")))) if person.get("kafedra") else None
            program_id = None
            if person.get("yonalish"):
                candidates = [pid for key, pid in program_map.items()
                              if key[0][0] == _key(person.get("fakultet"))
                              and (not person.get("kafedra") or key[0][1] == _key(person.get("kafedra")))
                              and key[1] == _key(person["yonalish"])]
                if len(candidates) == 1: program_id = candidates[0]
            placeholder, role_id, code = _new_placeholder(cur, person["fish"], university_id, role, person.get("telefon"), actor, faculty_id, department_id, program_id)
            _save_staff_profile(
                cur, role_id, ilmiy_daraja=person.get("ilmiy_daraja"),
                ilmiy_unvon=person.get("ilmiy_unvon"), staj_yil=person.get("staj_yil"),
                mutaxassislik=person.get("mutaxassislik"),
                qisqa_izoh=person.get("qisqa_izoh"),
            )
            if role == "tyutor" and program_id:
                cur.execute("""INSERT INTO universitet_tyutor_yonalishlari(
                    universitet_id,tyutor_user_id,yonalish_id,yaratilgan_by)
                    VALUES(%s,%s,%s,%s) ON CONFLICT DO NOTHING""", (university_id, placeholder, program_id, actor))
            counts["xodim"] += 1
            credentials.append({"fish": person["fish"], "lavozim": ROLE_LABELS[role],
                                "fakultet": person.get("fakultet"), "kafedra": person.get("kafedra"),
                                "yonalish": person.get("yonalish"), "telefon": person.get("telefon"),
                                "kirish_kodi": code, "kod_muddati": "2 oy"})
        cur.execute("UPDATE universitet_import_batchlar SET holat='committed',commit_at=NOW() WHERE id=%s", (req.batch_id,))
        _audit(cur, university_id, actor, "tuzilma_import_commit", "import_batch", req.batch_id, counts)
        credential_file = _credentials_xlsx(credentials, "institut_xodimlari_kirish_kodlari.xlsx", "INSTITUT XODIMLARI — BIR MARTALIK KIRISH KODLARI") if credentials else None
        conn.commit(); return {"holat": "import_qilindi", "sonlar": counts, "kirish_kodlari": credentials,
                               "kirish_kodlari_fayli": credential_file, "keyingi_bosqich": "qabul_importi",
                               "eslatma": "Kirish kodlari faqat shu javobda bir marta ko'rsatiladi"}
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


@router.get("/xodimlar")
def staff_list(universitet_id: int, arxiv: bool = False, token: Optional[str] = Query(None, include_in_schema=False), authorization: Optional[str] = Header(None)):
    user_id = _uid(token, authorization); p = _p(); conn = p._db(); cur = conn.cursor()
    try:
        _ensure_schema(cur); roles = _require_member(cur, user_id, universitet_id); names = _role_names(roles)
        if not (names & (INSTITUTE_WIDE | FACULTY_WIDE | DEPARTMENT_WIDE)):
            raise HTTPException(status_code=403, detail="Xodimlar ro'yxatini ko'rish huquqi yo'q")
        where = ["xr.universitet_id=%s", "xr.faol=%s"]; params: list[Any] = [universitet_id, not arxiv]
        if not (names & INSTITUTE_WIDE):
            faculty_ids = sorted({int(r["fakultet_id"]) for r in roles if r.get("fakultet_id") and r["rol"] in FACULTY_WIDE})
            department_ids = sorted({int(r["kafedra_id"]) for r in roles if r.get("kafedra_id") and r["rol"] in DEPARTMENT_WIDE})
            clauses = ["xr.user_id=%s"]; scoped: list[Any] = [user_id]
            if faculty_ids: clauses.append("xr.fakultet_id=ANY(%s)"); scoped.append(faculty_ids)
            if department_ids: clauses.append("xr.kafedra_id=ANY(%s)"); scoped.append(department_ids)
            where.append("(" + " OR ".join(clauses) + ")"); params.extend(scoped)
        cur.execute(f"""SELECT xr.id,xr.user_id,xr.rol,xr.fakultet_id,xr.kafedra_id,xr.yonalish_id,u.full_name,f.nomi fakultet_nomi,k.nomi kafedra_nomi,y.nomi yonalish_nomi,
            xp.ilmiy_daraja,xp.ilmiy_unvon,xp.staj_yil,xp.mutaxassislik,xp.qisqa_izoh,
            CASE WHEN tk.id IS NULL THEN NULL WHEN xk.ishlatildi THEN 'ulangan' ELSE 'taklif_yuborilgan' END kirish_holati
            FROM universitet_xodim_rollari xr JOIN users u ON u.user_id=xr.user_id
            LEFT JOIN fakultetlar f ON f.id=xr.fakultet_id LEFT JOIN kafedralar k ON k.id=xr.kafedra_id
            LEFT JOIN universitet_yonalishlari y ON y.id=xr.yonalish_id
            LEFT JOIN universitet_xodim_profili xp ON xp.xodim_rol_id=xr.id
            LEFT JOIN LATERAL (SELECT * FROM universitet_taklif_kodlari t WHERE t.xodim_rol_id=xr.id ORDER BY t.id DESC LIMIT 1) tk ON TRUE
            LEFT JOIN xodim_kod xk ON xk.kod=tk.kod_hash
            WHERE {' AND '.join(where)} ORDER BY f.nomi NULLS FIRST,xr.rol,u.full_name""", params)
        rows = [dict(r) for r in cur.fetchall()]
        for r in rows: r["lavozim_nomi"] = ROLE_LABELS.get(r["rol"], r["rol"])
        return {"xodimlar": rows, "lavozimlar": ROLE_LABELS, "arxiv": arxiv}
    finally:
        cur.close(); conn.close()


def _program_visible_to_member(cur, university_id: int, user_id: int,
                               roles: list[dict[str, Any]], program: dict[str, Any]) -> bool:
    """Yo'nalish jamoasi va lokal ro'yxatlari uchun bitta qamrov tekshiruvi."""
    names = _role_names(roles)
    if names & INSTITUTE_WIDE:
        return True
    program_id = int(program["id"])
    faculty_id = int(program["fakultet_id"])
    department_id = int(program["kafedra_id"])
    for role in roles:
        if role["rol"] == "talaba" and role.get("yonalish_id") == program_id:
            return True
        if role.get("yonalish_id") and int(role["yonalish_id"]) == program_id:
            return True
        if role.get("kafedra_id") and int(role["kafedra_id"]) == department_id:
            return True
        if role.get("fakultet_id") and int(role["fakultet_id"]) == faculty_id:
            return True
    if "tyutor" in names:
        cur.execute("""SELECT 1 FROM universitet_tyutor_yonalishlari ty
            WHERE ty.universitet_id=%s AND ty.tyutor_user_id=%s AND ty.faol=TRUE
              AND (ty.fakultet_id IS NULL OR ty.fakultet_id=%s)
              AND (ty.yonalish_id IS NULL OR ty.yonalish_id=%s)
            LIMIT 1""", (university_id, user_id, faculty_id, program_id))
        return cur.fetchone() is not None
    return False


@router.get("/xodimlar/team")
def direction_team(universitet_id: int, yonalish_id: int,
                   yonalish_ids: Optional[str] = None,
                   token: Optional[str] = Query(None, include_in_schema=False),
                   authorization: Optional[str] = Header(None)):
    """Tanlangan yo'nalishning rahbariyat, mudir, o'qituvchi va tyutorlari.

    Kasbiy profil eski bazada bo'lmasa ``null`` qaytadi. Shaxsiy telefon,
    login yoki taklif kodi bu endpoint orqali hech qachon berilmaydi.
    """
    actor = _uid(token, authorization); p = _p(); conn = p._db(); cur = conn.cursor()
    try:
        _ensure_schema(cur); roles = _require_member(cur, actor, universitet_id)
        cur.execute("""SELECT y.id,y.nomi,y.daraja,y.fakultet_id,y.kafedra_id,
                   f.nomi fakultet_nomi,k.nomi kafedra_nomi
            FROM universitet_yonalishlari y
            JOIN fakultetlar f ON f.id=y.fakultet_id
            JOIN kafedralar k ON k.id=y.kafedra_id
            WHERE y.id=%s AND y.universitet_id=%s AND y.faol=TRUE""",
            (yonalish_id, universitet_id))
        program = cur.fetchone()
        if not program:
            raise HTTPException(status_code=404, detail="Yo'nalish topilmadi")
        program_ids = {int(yonalish_id)}
        aliases: list[dict[str, Any]] = []
        if yonalish_ids:
            try:
                program_ids.update(int(value.strip()) for value in yonalish_ids.split(",") if value.strip())
            except ValueError:
                raise HTTPException(status_code=400, detail="Yo'nalish identifikatorlari noto'g'ri")
            cur.execute("""SELECT id,nomi,daraja,fakultet_id,kafedra_id FROM universitet_yonalishlari
                WHERE universitet_id=%s AND faol=TRUE AND id=ANY(%s)""",
                (universitet_id, sorted(program_ids)))
            aliases = [dict(item) for item in cur.fetchall()]
            if len(aliases) != len(program_ids):
                raise HTTPException(status_code=400, detail="Alias yo'nalishlardan biri topilmadi")
            if any(int(item["fakultet_id"]) != int(program["fakultet_id"])
                   or _key(item["nomi"]) != _key(program["nomi"])
                   or _key(item["daraja"]) != _key(program["daraja"])
                   for item in aliases):
                raise HTTPException(status_code=400, detail="Faqat bir xil nom va darajadagi yo'nalish aliaslari birlashtiriladi")
        visible_programs = [dict(program)]
        if yonalish_ids:
            visible_programs.extend(item for item in aliases if int(item["id"]) != int(yonalish_id))
        if not any(_program_visible_to_member(cur, universitet_id, actor, roles, item)
                   for item in visible_programs):
            raise HTTPException(status_code=403, detail="Bu yo'nalish jamoasini ko'rish huquqi yo'q")
        department_ids = sorted({int(item["kafedra_id"]) for item in visible_programs})
        cur.execute("""SELECT DISTINCT ON (xr.id)
                   xr.id,xr.user_id,xr.rol,u.full_name,
                   xr.fakultet_id,xr.kafedra_id,xr.yonalish_id,
                   f.nomi fakultet_nomi,k.nomi kafedra_nomi,y.nomi yonalish_nomi,
                   xp.ilmiy_daraja,xp.ilmiy_unvon,xp.staj_yil,
                   xp.mutaxassislik,xp.qisqa_izoh
            FROM universitet_xodim_rollari xr
            JOIN users u ON u.user_id=xr.user_id
            LEFT JOIN fakultetlar f ON f.id=xr.fakultet_id
            LEFT JOIN kafedralar k ON k.id=xr.kafedra_id
            LEFT JOIN universitet_yonalishlari y ON y.id=xr.yonalish_id
            LEFT JOIN universitet_xodim_profili xp ON xp.xodim_rol_id=xr.id
            WHERE xr.universitet_id=%s AND xr.faol=TRUE AND (
                 xr.rol IN ('rektor','prorektor')
              OR (xr.rol IN ('dekan','zam_dekan','manaviyatchi','fakultet_admin')
                  AND xr.fakultet_id=%s)
              OR (xr.rol='kafedra_mudiri' AND xr.kafedra_id=ANY(%s))
              OR (xr.rol='professor_oqituvchi' AND (
                    xr.yonalish_id=ANY(%s)
                 OR (xr.yonalish_id IS NULL AND xr.kafedra_id=ANY(%s))
                 OR (xr.yonalish_id IS NULL AND xr.kafedra_id IS NULL AND xr.fakultet_id=%s)))
              OR (xr.rol='tyutor' AND (
                    xr.yonalish_id=ANY(%s)
                 OR (xr.yonalish_id IS NULL AND xr.fakultet_id=%s)
                 OR EXISTS(SELECT 1 FROM universitet_tyutor_yonalishlari ty
                     WHERE ty.universitet_id=xr.universitet_id
                       AND ty.tyutor_user_id=xr.user_id AND ty.faol=TRUE
                       AND (ty.fakultet_id IS NULL OR ty.fakultet_id=%s)
                       AND (ty.yonalish_id IS NULL OR ty.yonalish_id=ANY(%s)))))
            )
            ORDER BY xr.id,u.full_name""",
            (universitet_id, program["fakultet_id"], department_ids,
             sorted(program_ids), department_ids, program["fakultet_id"],
             sorted(program_ids), program["fakultet_id"], program["fakultet_id"], sorted(program_ids)))
        category_for_role = {
            "rektor": "institut_rahbariyati", "prorektor": "institut_rahbariyati",
            "dekan": "fakultet_rahbariyati", "zam_dekan": "fakultet_rahbariyati",
            "manaviyatchi": "fakultet_rahbariyati", "fakultet_admin": "fakultet_rahbariyati",
            "kafedra_mudiri": "kafedra_rahbariyati",
            "professor_oqituvchi": "oqituvchilar", "tyutor": "tyutorlar",
        }
        groups = {key: [] for key in (
            "institut_rahbariyati", "fakultet_rahbariyati", "kafedra_rahbariyati",
            "oqituvchilar", "tyutorlar",
        )}
        staff = []
        for raw in cur.fetchall():
            item = dict(raw)
            item["fish"] = item.pop("full_name")
            item["lavozim_nomi"] = ROLE_LABELS.get(item["rol"], item["rol"])
            item["guruh"] = category_for_role.get(item["rol"], "oqituvchilar")
            groups[item["guruh"]].append(item); staff.append(item)
        for items in groups.values():
            items.sort(key=lambda item: (_key(item["lavozim_nomi"]), _key(item["fish"])))
        program_payload = dict(program); program_payload["alias_ids"] = sorted(program_ids)
        return {"yonalish": program_payload, "guruhlar": groups, "xodimlar": staff,
                "jami": len(staff)}
    finally:
        cur.close(); conn.close()


@router.get("/qidiruv/odamlar")
def global_people_search(universitet_id: int, q: str = "", limit: int = 20,
                         token: Optional[str] = Query(None, include_in_schema=False),
                         authorization: Optional[str] = Header(None)):
    """Institut bo'ylab ism/ID bo'yicha xodim va talabani ruxsat doirasida topadi."""
    actor = _uid(token, authorization); p = _p(); conn = p._db(); cur = conn.cursor()
    try:
        _ensure_schema(cur); roles = _require_member(cur, actor, universitet_id)
        names = _role_names(roles)
        if not (names & PRIVATE_ROLES):
            raise HTTPException(status_code=403, detail="Odamlar qidiruviga ruxsat yo'q")
        query = _norm(q)
        if len(query) < 2:
            return {"q": query, "natijalar": [], "jami": 0, "minimal_belgi": 2}
        limit = max(5, min(50, limit)); term = f"%{query}%"
        results: list[dict[str, Any]] = []

        staff_where = ["xr.universitet_id=%s", "xr.faol=TRUE"]
        staff_params: list[Any] = [universitet_id]
        if not (names & INSTITUTE_WIDE):
            faculty_id_set = {int(r["fakultet_id"]) for r in roles if r.get("fakultet_id")}
            department_id_set = {int(r["kafedra_id"]) for r in roles if r.get("kafedra_id")}
            program_id_set = {int(r["yonalish_id"]) for r in roles if r.get("yonalish_id")}
            if "tyutor" in names:
                cur.execute("""SELECT COALESCE(ty.fakultet_id,y.fakultet_id) fakultet_id,
                           y.kafedra_id,ty.yonalish_id
                    FROM universitet_tyutor_yonalishlari ty
                    LEFT JOIN universitet_yonalishlari y ON y.id=ty.yonalish_id
                    WHERE ty.universitet_id=%s AND ty.tyutor_user_id=%s AND ty.faol=TRUE""",
                    (universitet_id, actor))
                for assignment in cur.fetchall():
                    if assignment.get("fakultet_id"): faculty_id_set.add(int(assignment["fakultet_id"]))
                    if assignment.get("kafedra_id"): department_id_set.add(int(assignment["kafedra_id"]))
                    if assignment.get("yonalish_id"): program_id_set.add(int(assignment["yonalish_id"]))
            faculty_ids = sorted(faculty_id_set)
            department_ids = sorted(department_id_set)
            program_ids = sorted(program_id_set)
            staff_scope = ["xr.user_id=%s"]; staff_scope_params: list[Any] = [actor]
            if faculty_ids:
                staff_scope.append("xr.fakultet_id=ANY(%s)"); staff_scope_params.append(faculty_ids)
            if department_ids:
                staff_scope.append("xr.kafedra_id=ANY(%s)"); staff_scope_params.append(department_ids)
            if program_ids:
                staff_scope.append("xr.yonalish_id=ANY(%s)"); staff_scope_params.append(program_ids)
            staff_where.append("(" + " OR ".join(staff_scope) + ")")
            staff_params.extend(staff_scope_params)
        staff_where.append("u.full_name ILIKE %s"); staff_params.append(term)
        cur.execute(f"""SELECT xr.id,xr.user_id,xr.rol,u.full_name,
                   f.nomi fakultet_nomi,k.nomi kafedra_nomi,y.nomi yonalish_nomi
            FROM universitet_xodim_rollari xr JOIN users u ON u.user_id=xr.user_id
            LEFT JOIN fakultetlar f ON f.id=xr.fakultet_id
            LEFT JOIN kafedralar k ON k.id=xr.kafedra_id
            LEFT JOIN universitet_yonalishlari y ON y.id=xr.yonalish_id
            WHERE {' AND '.join(staff_where)}
            ORDER BY u.full_name LIMIT %s""", staff_params + [limit])
        for raw in cur.fetchall():
            item = dict(raw)
            item.update({"turi": "xodim", "fish": item.pop("full_name"),
                         "lavozim_nomi": ROLE_LABELS.get(item["rol"], item["rol"])})
            results.append(item)

        student_scope_sql, student_scope_params = _student_scope_clause(
            cur, universitet_id, actor, roles
        )
        student_search = """(qt.familiya ILIKE %s OR qt.ism ILIKE %s
            OR qt.ota_ism ILIKE %s OR CONCAT_WS(' ',qt.familiya,qt.ism,qt.ota_ism) ILIKE %s
            OR qt.abitur_id ILIKE %s)"""
        cur.execute(f"""SELECT qt.id,qt.abitur_id,qt.familiya,qt.ism,qt.ota_ism,
                   qt.talim_shakli,qt.talim_tili,y.id yonalish_id,y.nomi yonalish_nomi,
                   f.id fakultet_id,f.nomi fakultet_nomi
            FROM universitet_qabul_talabalari qt
            JOIN universitet_yonalishlari y ON y.id=qt.yonalish_id
            JOIN fakultetlar f ON f.id=y.fakultet_id
            WHERE qt.universitet_id=%s AND {student_scope_sql} AND {student_search}
            ORDER BY qt.familiya,qt.ism LIMIT %s""",
            [universitet_id, *student_scope_params, term, term, term, term, term, limit])
        for raw in cur.fetchall():
            item = dict(raw)
            item["fish"] = " ".join(value for value in (
                item.pop("familiya"), item.pop("ism"), item.pop("ota_ism")
            ) if value)
            item["turi"] = "talaba"; results.append(item)
        results.sort(key=lambda item: (_key(item["fish"]), item["turi"]))
        results = results[:limit]
        return {"q": query, "natijalar": results, "jami": len(results),
                "chegara": limit}
    finally:
        cur.close(); conn.close()


@router.post("/xodim/{role_id}/kirish_kodi")
def staff_invite_code(role_id: int, req: InstituteToken):
    """Super-admin/rahbar xodimning ishlatilmagan kodini ko'radi; bo'lmasa yangisini yaratadi."""
    p = _p(); actor = p._jwt_tekshir(req.token); conn = p._db(); cur = conn.cursor()
    try:
        _ensure_schema(cur)
        cur.execute("""SELECT xr.*,u.full_name FROM universitet_xodim_rollari xr
            JOIN users u ON u.user_id=xr.user_id WHERE xr.id=%s FOR UPDATE""", (role_id,))
        row = cur.fetchone()
        if not row: raise HTTPException(status_code=404, detail="Xodim topilmadi")
        roles = _require_member(cur, actor, row["universitet_id"])
        if not _has_any(roles, MANAGE_STAFF_ROLES):
            raise HTTPException(status_code=403, detail="Kirish kodini boshqarish huquqi yo'q")
        if row["user_id"] is None or int(row["user_id"]) >= 0:
            raise HTTPException(
                status_code=409,
                detail="Xodim haqiqiy akkauntga ulangan; unga yangi kirish kodi yaratib yoki ko'rsatib bo'lmaydi",
            )
        cur.execute("""SELECT tk.kod_shifr FROM universitet_taklif_kodlari tk
            JOIN xodim_kod xk ON xk.kod=tk.kod_hash
            WHERE tk.xodim_rol_id=%s AND xk.ishlatildi=FALSE
              AND tk.placeholder_user_id=%s
              AND xk.yaratildi>NOW()-INTERVAL '2 months'
            ORDER BY tk.id DESC LIMIT 1""", (role_id, row["user_id"]))
        found = cur.fetchone(); code = _open_invite_code(found["kod_shifr"]) if found else None
        if not code:
            _, _, code = _new_placeholder(cur, row["full_name"], row["universitet_id"], row["rol"], None,
                                           actor, row["fakultet_id"], row["kafedra_id"], row["yonalish_id"])
        _audit(cur, row["universitet_id"], actor, "xodim_kirish_kodi_korildi", "xodim_rol", role_id)
        conn.commit(); return {"fish": row["full_name"], "lavozim": ROLE_LABELS.get(row["rol"], row["rol"]),
                               "kirish_kodi": code, "kod_muddati": "2 oy"}
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


@router.get("/audit")
def audit_list(universitet_id: int, q: str = "", page: int = 1, token: Optional[str] = Query(None, include_in_schema=False), authorization: Optional[str] = Header(None)):
    actor = _uid(token, authorization); p = _p(); conn = p._db(); cur = conn.cursor()
    try:
        _ensure_schema(cur); roles = _require_member(cur, actor, universitet_id)
        if not (_role_names(roles) & INSTITUTE_WIDE):
            raise HTTPException(status_code=403, detail="Institut logini faqat institut rahbariyati ko'radi")
        term = f"%{_norm(q)}%"; offset = (max(1, page)-1)*100
        cur.execute("""SELECT l.id,l.amal,l.obyekt_turi,l.obyekt_id,l.tafsilot,l.yaratilgan_at,
                   u.full_name actor_fish
            FROM universitet_audit_log l LEFT JOIN users u ON u.user_id=l.actor_user_id
            WHERE l.universitet_id=%s AND (%s='' OR l.amal ILIKE %s OR COALESCE(u.full_name,'') ILIKE %s)
            ORDER BY l.id DESC LIMIT 100 OFFSET %s""", (universitet_id, _norm(q), term, term, offset))
        return {"loglar": [dict(x) for x in cur.fetchall()], "page": max(1, page)}
    finally:
        cur.close(); conn.close()


@router.post("/tyutor/biriktir")
def assign_tutor(req: TutorAssign):
    p = _p(); actor = p._jwt_tekshir(req.token); conn = p._db(); cur = conn.cursor()
    try:
        _ensure_schema(cur); roles = _require_member(cur, actor, req.universitet_id)
        if not _has_any(roles, MANAGE_STAFF_ROLES): raise HTTPException(status_code=403, detail="Tyutor biriktirish huquqi yo'q")
        faculty_id, program_id = req.fakultet_id, req.yonalish_id
        if program_id:
            cur.execute("SELECT fakultet_id FROM universitet_yonalishlari WHERE id=%s AND universitet_id=%s AND faol=TRUE", (program_id, req.universitet_id))
            program = cur.fetchone()
            if not program: raise HTTPException(status_code=400, detail="Yo'nalish bu institutga tegishli emas")
            if faculty_id and int(faculty_id) != int(program["fakultet_id"]):
                raise HTTPException(status_code=400, detail="Fakultet va yo'nalish bir-biriga mos emas")
            faculty_id = int(program["fakultet_id"])
        if faculty_id:
            _validate_assignment_scope(cur, req.universitet_id, roles, int(faculty_id), None, program_id)
        elif not (_role_names(roles) & INSTITUTE_WIDE):
            raise HTTPException(status_code=403, detail="Barcha institut qamrovini faqat institut rahbariyati biriktiradi")
        form = _canonical_choice(req.talim_shakli, TA_LIM_SHAKLLARI) if req.talim_shakli else None
        language = _canonical_choice(req.talim_tili, TA_LIM_TILLARI) if req.talim_tili else None
        admission_type = _norm(req.qabul_turi).lower() or None
        if admission_type not in {None, "grant", "kontrakt"}:
            raise HTTPException(status_code=400, detail="Qabul turi grant yoki kontrakt bo'lishi kerak")
        cur.execute("SELECT 1 FROM universitet_xodim_rollari WHERE universitet_id=%s AND user_id=%s AND rol='tyutor' AND faol=TRUE", (req.universitet_id, req.tyutor_user_id))
        if not cur.fetchone(): raise HTTPException(status_code=400, detail="Tanlangan xodim tyutor emas")
        cur.execute("""SELECT id,faol FROM universitet_tyutor_yonalishlari
            WHERE universitet_id=%s AND tyutor_user_id=%s
              AND fakultet_id IS NOT DISTINCT FROM %s AND yonalish_id IS NOT DISTINCT FROM %s
              AND talim_shakli IS NOT DISTINCT FROM %s AND talim_tili IS NOT DISTINCT FROM %s
              AND qabul_turi IS NOT DISTINCT FROM %s
            LIMIT 1""", (req.universitet_id, req.tyutor_user_id, faculty_id, program_id, form, language, admission_type))
        existing = cur.fetchone()
        if existing:
            cur.execute("UPDATE universitet_tyutor_yonalishlari SET faol=TRUE,yaratilgan_by=%s WHERE id=%s", (actor, existing["id"]))
            assignment_id, status = int(existing["id"]), "qayta_faollashtirildi" if not existing["faol"] else "avval_biriktirilgan"
        else:
            cur.execute("""INSERT INTO universitet_tyutor_yonalishlari(
                universitet_id,tyutor_user_id,fakultet_id,yonalish_id,talim_shakli,talim_tili,qabul_turi,yaratilgan_by)
                VALUES(%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
                (req.universitet_id, req.tyutor_user_id, faculty_id, program_id, form, language, admission_type, actor))
            assignment_id, status = int(cur.fetchone()["id"]), "biriktirildi"
        _audit(cur, req.universitet_id, actor, "tyutor_qamrovi_biriktirildi", "tyutor_qamrovi", assignment_id,
               {"fakultet_id": faculty_id, "yonalish_id": program_id, "talim_shakli": form, "talim_tili": language, "qabul_turi": admission_type})
        conn.commit()
        return {"holat": status, "biriktirish_id": assignment_id}
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


@router.get("/tyutor/biriktirishlar")
def tutor_assignments(universitet_id: int, token: Optional[str] = Query(None, include_in_schema=False), authorization: Optional[str] = Header(None)):
    user_id = _uid(token, authorization); p = _p(); conn = p._db(); cur = conn.cursor()
    try:
        _ensure_schema(cur); roles = _require_member(cur, user_id, universitet_id); names = _role_names(roles)
        if not (names & (INSTITUTE_WIDE | FACULTY_WIDE | {"tyutor"})):
            raise HTTPException(status_code=403, detail="Tyutor biriktirishlarini ko'rish huquqi yo'q")
        where = ["ty.universitet_id=%s", "ty.faol=TRUE"]; params: list[Any] = [universitet_id]
        if not (names & INSTITUTE_WIDE):
            faculty_ids = sorted({int(r["fakultet_id"]) for r in roles if r.get("fakultet_id") and r["rol"] in FACULTY_WIDE})
            clauses = ["ty.tyutor_user_id=%s"]; scoped: list[Any] = [user_id]
            if faculty_ids:
                clauses.append("COALESCE(ty.fakultet_id,y.fakultet_id)=ANY(%s)"); scoped.append(faculty_ids)
            where.append("(" + " OR ".join(clauses) + ")"); params.extend(scoped)
        cur.execute(f"""SELECT ty.id,ty.tyutor_user_id,u.full_name,ty.fakultet_id,f.nomi fakultet_nomi,
            ty.yonalish_id,y.nomi yonalish_nomi,ty.talim_shakli,ty.talim_tili,ty.qabul_turi
            FROM universitet_tyutor_yonalishlari ty JOIN users u ON u.user_id=ty.tyutor_user_id
            LEFT JOIN universitet_yonalishlari y ON y.id=ty.yonalish_id
            LEFT JOIN fakultetlar f ON f.id=COALESCE(ty.fakultet_id,y.fakultet_id)
            WHERE {' AND '.join(where)} ORDER BY u.full_name,f.nomi NULLS FIRST,y.nomi NULLS FIRST,ty.talim_shakli,ty.talim_tili""", params)
        return {"biriktirishlar": [dict(row) for row in cur.fetchall()]}
    finally:
        cur.close(); conn.close()


@router.patch("/tyutor/biriktirish/{assignment_id}")
def update_tutor_assignment(assignment_id: int, req: TutorAssignmentUpdate):
    p = _p(); actor = p._jwt_tekshir(req.token); conn = p._db(); cur = conn.cursor()
    try:
        _ensure_schema(cur)
        cur.execute("SELECT * FROM universitet_tyutor_yonalishlari WHERE id=%s FOR UPDATE", (assignment_id,)); row = cur.fetchone()
        if not row: raise HTTPException(status_code=404, detail="Tyutor biriktirishi topilmadi")
        roles = _require_member(cur, actor, row["universitet_id"])
        if not _has_any(roles, MANAGE_STAFF_ROLES): raise HTTPException(status_code=403, detail="Tyutor biriktirishini o'zgartirish huquqi yo'q")
        target_faculty = row["fakultet_id"]
        if not target_faculty and row["yonalish_id"]:
            cur.execute("SELECT fakultet_id FROM universitet_yonalishlari WHERE id=%s", (row["yonalish_id"],)); program = cur.fetchone()
            target_faculty = program["fakultet_id"] if program else None
        if target_faculty:
            _validate_assignment_scope(cur, row["universitet_id"], roles, target_faculty, None, row["yonalish_id"])
        elif not (_role_names(roles) & INSTITUTE_WIDE):
            raise HTTPException(status_code=403, detail="Institut qamrovini faqat institut rahbariyati o'zgartiradi")
        cur.execute("UPDATE universitet_tyutor_yonalishlari SET faol=%s WHERE id=%s", (req.faol, assignment_id))
        _audit(cur, row["universitet_id"], actor, "tyutor_qamrovi_holati", "tyutor_qamrovi", assignment_id, {"faol": req.faol})
        conn.commit(); return {"holat": "faol" if req.faol else "o'chirildi"}
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


@router.get("/qabul/shablon")
def admission_template(
    universitet_id: int,
    fakultet_id: Optional[int] = None,
    token: Optional[str] = Query(None, include_in_schema=False),
    authorization: Optional[str] = Header(None),
):
    """Tanlangan institut/fakultet yo'nalishlari bilan qabul XLSX shabloni."""
    user_id = _uid(token, authorization); p = _p(); conn = p._db(); cur = conn.cursor()
    try:
        _ensure_schema(cur); roles = _require_member(cur, user_id, universitet_id)
        if not _has_any(roles, ADMISSION_IMPORT_ROLES):
            raise HTTPException(status_code=403, detail="Qabul shablonini yuklash huquqi yo'q")
        if fakultet_id is not None:
            _validate_assignment_scope(cur, universitet_id, roles, fakultet_id, None, None)
        cur.execute("SELECT id,nomi FROM universitetlar WHERE id=%s", (universitet_id,))
        university = cur.fetchone()
        if not university:
            raise HTTPException(status_code=404, detail="Institut topilmadi")
        cur.execute("""SELECT y.id,y.nomi
            FROM universitet_yonalishlari y
            JOIN fakultetlar f ON f.id=y.fakultet_id AND f.faol=TRUE
            JOIN kafedralar k ON k.id=y.kafedra_id AND k.faol=TRUE
            WHERE y.universitet_id=%s AND y.faol=TRUE
              AND (%s IS NULL OR y.fakultet_id=%s)
            ORDER BY y.nomi""", (universitet_id, fakultet_id, fakultet_id))
        content = _admission_template_xlsx(dict(university), [dict(row) for row in cur.fetchall()])
        return _xlsx_download(content, "institut_qabul_shabloni.xlsx")
    finally:
        cur.close(); conn.close()


@router.post("/qabul/import_preview")
async def admission_preview(universitet_id: int, fakultet_id: Optional[int] = None, fayl: UploadFile = File(...), token: Optional[str] = Query(None, include_in_schema=False), authorization: Optional[str] = Header(None)):
    user_id = _uid(token, authorization); p = _p()
    content = await fayl.read()
    if len(content) > 25 * 1024 * 1024: raise HTTPException(status_code=413, detail="Qabul fayli 25 MB dan katta")
    if not (fayl.filename or "").lower().endswith((".xls", ".xlsx")): raise HTTPException(status_code=400, detail="Faqat .xls yoki .xlsx fayl qabul qilinadi")
    parsed, summary = await run_in_threadpool(
        _parse_admission,
        content,
        fayl.filename or "qabul.xls",
    )
    conn = p._db(); cur = conn.cursor()
    try:
        _ensure_schema(cur); roles = _require_member(cur, user_id, universitet_id)
        if not _has_any(roles, MARK_DOCUMENT_ROLES): raise HTTPException(status_code=403, detail="Qabul importiga ruxsat yo'q")
        if fakultet_id is not None:
            cur.execute("SELECT id,nomi FROM fakultetlar WHERE id=%s AND universitet_id=%s AND faol=TRUE", (fakultet_id, universitet_id))
            selected_faculty = cur.fetchone()
            if not selected_faculty: raise HTTPException(status_code=400, detail="Tanlangan fakultet bu institutga tegishli emas")
            _validate_assignment_scope(cur, universitet_id, roles, fakultet_id, None, None)
        else:
            selected_faculty = None
        cur.execute("""SELECT k.id,k.nomi,k.fakultet_id,f.nomi fakultet_nomi FROM kafedralar k
            JOIN fakultetlar f ON f.id=k.fakultet_id
            WHERE f.universitet_id=%s AND f.faol=TRUE AND k.faol=TRUE
              AND (%s IS NULL OR f.id=%s) ORDER BY f.nomi,k.nomi""", (universitet_id, fakultet_id, fakultet_id))
        department_rows = [dict(row) for row in cur.fetchall()]
        cur.execute("""SELECT y.id,y.nomi,y.kafedra_id,y.fakultet_id
            FROM universitet_yonalishlari y
            JOIN fakultetlar f ON f.id=y.fakultet_id AND f.faol=TRUE
            JOIN kafedralar k ON k.id=y.kafedra_id AND k.faol=TRUE
            WHERE y.universitet_id=%s AND y.faol=TRUE
              AND (%s IS NULL OR y.fakultet_id=%s)
            ORDER BY y.nomi""", (universitet_id, fakultet_id, fakultet_id))
        program_rows = [dict(row) for row in cur.fetchall()]
        matching, exact, ambiguous, unknown = {}, {}, {}, {}
        for name, count in summary["yonalishlar"].items():
            same_programs = [row for row in program_rows if _key(row["nomi"]) == _key(name)]
            same = [row for row in department_rows if _key(row["nomi"]) == _key(name)]
            departments = _rank_departments(name, department_rows)
            selected_program_id, selected_department_id, status = None, None, "yangi_yonalish"
            if len(same_programs) == 1:
                selected_program_id = int(same_programs[0]["id"])
                selected_department_id = int(same_programs[0]["kafedra_id"])
                status = "yonalish_aniq"
                exact[name] = selected_program_id
            elif len(same_programs) > 1:
                ambiguous[name] = len(same_programs)
            elif len(same) == 1:
                selected_department_id, status = int(same[0]["id"]), "aniq"
            elif len(same) > 1:
                ambiguous[name] = len(same)
            elif departments:
                top = departments[0]; gap = top["moslik_foizi"] - (departments[1]["moslik_foizi"] if len(departments) > 1 else 0)
                if top["moslik_foizi"] >= 88 and gap >= 5:
                    selected_department_id, status = top["id"], "yaqin_topildi"
            create_name = None if selected_department_id else _base_program_name(name)
            if not selected_program_id and not selected_department_id: unknown[name] = count
            variants = [item for item in summary.get("yonalish_variantlari", []) if item["yonalish"] == name]
            matching[name] = {
                "talaba_soni": count, "holat": status,
                "tanlangan_yonalish_id": selected_program_id,
                "tanlangan_kafedra_id": selected_department_id,
                "yaratiladigan_kafedra_nomi": create_name,
                "yaratiladigan_yonalish_nomi": create_name,
                "variantlar": variants,
                "yonalish_variantlari": [],
                "kafedra_variantlari": departments,
            }
        summary["tanlangan_fakultet_id"] = int(fakultet_id) if fakultet_id is not None else None
        summary["tanlangan_fakultet_nomi"] = selected_faculty["nomi"] if selected_faculty else None
        summary["noma_lum_yonalishlar"] = unknown
        summary["noaniq_yonalishlar"] = ambiguous
        summary["mos_yonalishlar"] = exact
        summary["yonalish_moslashtirish"] = matching
        cur.execute("SELECT nomi FROM universitetlar WHERE id=%s", (universitet_id,)); university = cur.fetchone()
        # Saytda nom qo'lda kiritilganda bitta harf xatosi (masalan,
        # "insitituti") qabulni bekordan-bekor to'xtatmasin. Jiddiy farq
        # bo'lsa esa foydalanuvchi baribir alohida tasdiqlaydi.
        foreign_names = sorted({
            r["otm_nomi"] for r in parsed
            if r["otm_nomi"] and _name_similarity(r["otm_nomi"], university["nomi"]) < 90
        })
        summary["otm_nomi_farqi"] = foreign_names
        batch_id = _store_batch(cur, universitet_id, "qabul", fayl.filename or "qabul.xls", content, parsed, summary, user_id)
        conn.commit(); return {"batch_id": batch_id, "xulosa": summary,
                               "commit_mumkin": summary["xato_soni"] == 0 and fakultet_id is not None}
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


@router.post("/qabul/import_commit")
def admission_commit(req: BatchCommit):
    p = _p(); actor = p._jwt_tekshir(req.token); conn = p._db(); cur = conn.cursor()
    try:
        _ensure_schema(cur)
        cur.execute("SELECT * FROM universitet_import_batchlar WHERE id=%s FOR UPDATE", (req.batch_id,)); batch = cur.fetchone()
        if not batch or batch["import_turi"] != "qabul": raise HTTPException(status_code=404, detail="Qabul preview topilmadi")
        if batch["yaratilgan_by"] != actor and not _is_global_admin(cur, actor): raise HTTPException(status_code=403, detail="Bu preview boshqa foydalanuvchiga tegishli")
        if batch["holat"] != "preview": raise HTTPException(status_code=409, detail="Bu preview avval ishlatilgan")
        roles = _require_member(cur, actor, batch["universitet_id"])
        if not _has_any(roles, MARK_DOCUMENT_ROLES): raise HTTPException(status_code=403, detail="Qabul importiga ruxsat yo'q")
        payload = batch["payload"] if isinstance(batch["payload"], list) else json.loads(batch["payload"])
        summary = batch["xulosa"] if isinstance(batch["xulosa"], dict) else json.loads(batch["xulosa"])
        if summary.get("xato_soni"): raise HTTPException(status_code=400, detail="Xatoli fayl commit qilinmaydi; preview xatolarini tuzating")
        if summary.get("otm_nomi_farqi") and not req.otm_nomi_farqini_tasdiqlash:
            raise HTTPException(status_code=400, detail="Fayldagi OTM nomi tanlangan institutga mos emas; avval farqni tasdiqlang")
        if req.auto_create_yonalishlar:
            raise HTTPException(
                status_code=400,
                detail=(
                    "Qabul importi tuzilma yaratmaydi. Har bir Excel "
                    "yo'nalishini mavjud kafedra ichidagi yo'nalishga moslang"
                ),
            )
        summary_matching = summary.get("yonalish_moslashtirish") or {}
        selected_faculty_id = summary.get("tanlangan_fakultet_id")
        if not selected_faculty_id:
            raise HTTPException(status_code=400, detail="Talabalar importi uchun fakultet tanlanmagan")
        direction_names = sorted({r["yonalish_nomi"] for r in payload})
        resolved: dict[str, dict[str, Any]] = {}
        for name in direction_names:
            suggested = summary_matching.get(name) or {}
            program_id = req.yonalish_mosliklari.get(name) or suggested.get("tanlangan_yonalish_id")
            if program_id:
                cur.execute("""SELECT y.id,y.nomi,y.fakultet_id,y.kafedra_id
                    FROM universitet_yonalishlari y
                    JOIN fakultetlar f ON f.id=y.fakultet_id AND f.faol=TRUE
                    JOIN kafedralar k ON k.id=y.kafedra_id AND k.faol=TRUE
                    WHERE y.id=%s AND y.universitet_id=%s
                      AND y.fakultet_id=%s AND y.faol=TRUE""",
                    (program_id, batch["universitet_id"], selected_faculty_id))
                existing_program = cur.fetchone()
                if not existing_program:
                    raise HTTPException(status_code=400, detail=f"{name}: tanlangan yo‘nalish bu fakultetga tegishli emas")
                resolved[name] = dict(existing_program)
                continue
            raise HTTPException(
                status_code=400,
                detail=(
                    f"{name}: mavjud yo'nalish tanlanmagan. "
                    "Qabul importi kafedra yoki yo'nalish yaratmaydi"
                ),
            )
        for variant in {(r["yonalish_nomi"], r["talim_shakli"], r["talim_tili"]) for r in payload}:
            program_id = resolved[variant[0]]["id"]
            cur.execute("""INSERT INTO universitet_yonalish_variantlari(yonalish_id,talim_shakli,talim_tili)
                VALUES(%s,%s,%s) ON CONFLICT DO NOTHING""", (program_id, variant[1], variant[2]))
        inserted = updated = 0
        sql = """INSERT INTO universitet_qabul_talabalari(
            universitet_id,yonalish_id,abitur_id,jshshir,jshshir_hash,familiya,ism,ota_ism,tugilgan_sana,
            pasport_seriya,pasport_raqam,tavsiya_turi,talim_shakli,talim_tili,telefon,ball,doimiy_region,
            doimiy_tuman,maktab_region,maktab_tuman,maktab_turi,maktab_nomi,tugatgan_yili,attestat,otm_nomi,import_batch_id)
            VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            ON CONFLICT(universitet_id,jshshir_hash) DO UPDATE SET
              yonalish_id=EXCLUDED.yonalish_id,abitur_id=EXCLUDED.abitur_id,familiya=EXCLUDED.familiya,
              ism=EXCLUDED.ism,ota_ism=EXCLUDED.ota_ism,tugilgan_sana=EXCLUDED.tugilgan_sana,
              pasport_seriya=EXCLUDED.pasport_seriya,pasport_raqam=EXCLUDED.pasport_raqam,
              tavsiya_turi=EXCLUDED.tavsiya_turi,talim_shakli=EXCLUDED.talim_shakli,talim_tili=EXCLUDED.talim_tili,
              telefon=EXCLUDED.telefon,ball=EXCLUDED.ball,doimiy_region=EXCLUDED.doimiy_region,
              doimiy_tuman=EXCLUDED.doimiy_tuman,import_batch_id=EXCLUDED.import_batch_id,yangilangan_at=NOW()
            RETURNING id,(xmax=0) AS inserted"""
        imported_students: list[tuple[int, dict[str, Any]]] = []
        for r in payload:
            program_id = resolved[r["yonalish_nomi"]]["id"]
            cur.execute(sql, (batch["universitet_id"], program_id, r["abitur_id"], r["jshshir"], r["jshshir_hash"],
                r["familiya"], r["ism"], r.get("ota_ism"), r.get("tugilgan_sana"), r.get("pasport_seriya"), r.get("pasport_raqam"),
                r.get("tavsiya_turi"), r["talim_shakli"], r["talim_tili"], r.get("telefon"), r.get("ball"),
                r.get("doimiy_region"), r.get("doimiy_tuman"), r.get("maktab_region"), r.get("maktab_tuman"),
                r.get("maktab_turi"), r.get("maktab_nomi"), r.get("tugatgan_yili"), r.get("attestat"), r.get("otm_nomi"), req.batch_id))
            saved = cur.fetchone(); imported_students.append((int(saved["id"]), r))
            if saved["inserted"]: inserted += 1
            else: updated += 1
        # Importning asosiy vazifasi — talabalarni tez va ishonchli bazaga
        # saqlash. Yuzlab kirish kodlarini shu tranzaksiyada yaratish Railway
        # timeoutiga olib kelib, butun importni rollback qilardi. Kod talaba
        # kartasidagi “Kirish kodi” amali bilan keyin, kerak bo‘lganda yaratiladi.
        credentials = []
        connected = 0
        cur.execute("UPDATE universitet_import_batchlar SET holat='committed',commit_at=NOW() WHERE id=%s", (req.batch_id,))
        audit_detail = {"yangi": inserted, "yangilangan": updated, "kod_yaratildi": len(credentials), "avval_ulangan": connected}
        _audit(cur, batch["universitet_id"], actor, "qabul_import_commit", "import_batch", req.batch_id, audit_detail)
        credential_file = _credentials_xlsx(credentials, "1_kurs_talabalari_kirish_kodlari.xlsx", "1-KURS TALABALARI — BIR MARTALIK KIRISH KODLARI") if credentials else None
        conn.commit(); return {"holat": "import_qilindi", "yangi": inserted, "yangilangan": updated,
                               "jami": inserted + updated, "kod_yaratildi": len(credentials),
                               "avval_ulangan": connected, "kirish_kodlari_fayli": credential_file,
                               "keyingi_bosqich": "talabalar_saqlanib_boldi_kod_keyin_beriladi"}
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


@router.get("/qabul/talabalar")
def admission_students(universitet_id: int, q: str = "", fakultet_id: Optional[int] = None, yonalish_id: Optional[int] = None,
                       yonalish_ids: Optional[str] = None,
                       bosqich: Optional[int] = None, bosqich_min: Optional[int] = None,
                       holat: str = "all",
                       talim_shakli: Optional[str] = None, talim_tili: Optional[str] = None,
                       region: Optional[str] = None, qabul_turi: Optional[str] = None,
                       sort: str = "ball_desc", page: int = 1, page_size: int = 50,
                       token: Optional[str] = Query(None, include_in_schema=False), authorization: Optional[str] = Header(None)):
    user_id = _uid(token, authorization); p = _p(); conn = p._db(); cur = conn.cursor()
    try:
        _ensure_schema(cur); roles = _require_member(cur, user_id, universitet_id); names = _role_names(roles)
        if not _has_any(roles, PRIVATE_ROLES): raise HTTPException(status_code=403, detail="Qabul ro'yxatini ko'rish huquqi yo'q")
        tutor_only = "tyutor" in names and not (names & MARK_DOCUMENT_ROLES)
        scope_sql, scope_params = _student_scope_clause(cur, universitet_id, user_id, roles)
        where = ["qt.universitet_id=%s", "y.faol=TRUE", "f.faol=TRUE", "k.faol=TRUE", scope_sql]; params: list[Any] = [universitet_id, *scope_params]
        # Hisob kartalari va filtr variantlari ham aynan ochilgan
        # fakultet/yo'nalish doirasida hisoblanadi. Qidiruv yoki bosqich kabi
        # vaqtinchalik UI filtrlari bu kontekstni o'zgartirmaydi.
        context_where = ["qt.universitet_id=%s", "y.faol=TRUE", "f.faol=TRUE", "k.faol=TRUE", scope_sql]
        context_params: list[Any] = [universitet_id, *scope_params]
        site_entered_sql = "(qt.saytga_kiritilgan_at IS NOT NULL OR qt.birinchi_kirish_at IS NOT NULL OR (qt.user_id IS NOT NULL AND qt.user_id>=0))"
        status_aliases = {
            "all": "all", "barchasi": "all", "jami": "all",
            "hujjat": "hujjat", "hujjattopshirgan": "hujjat",
            "topshirmagan": "topshirmagan", "hujjattopshirmagan": "topshirmagan",
            "hujjatqolgan": "topshirmagan", "qolgan": "topshirmagan",
            "baza": "baza", "bazagakiritilgan": "baza",
            "saytgakirmagan": "saytga_kirmagan",
            "saytgakirgan": "saytga_kirgan",
            "hemiskutilmoqda": "hemis_kutilmoqda", "bazakutilmoqda": "hemis_kutilmoqda",
            "suhbat": "suhbat", "suhbatlashgan": "suhbat", "izohli": "suhbat",
        }
        selected_status = status_aliases.get(_key(holat or "all"))
        if selected_status is None:
            raise HTTPException(
                status_code=400,
                detail="Holat all, hujjat, topshirmagan, baza, hemis_kutilmoqda, saytga_kirmagan yoki saytga_kirgan bo'lishi kerak",
            )
        if selected_status == "hujjat":
            where.append("qt.hujjat_topshirgan_at IS NOT NULL")
        elif selected_status == "topshirmagan":
            where.append("qt.hujjat_topshirgan_at IS NULL")
        elif selected_status == "baza":
            where.append("qt.bazaga_kiritilgan_at IS NOT NULL")
        elif selected_status == "hemis_kutilmoqda":
            where.append("qt.hujjat_topshirgan_at IS NOT NULL AND qt.bazaga_kiritilgan_at IS NULL")
        elif selected_status == "suhbat":
            where.append("NULLIF(BTRIM(qt.aloqa_izohi),'') IS NOT NULL")
        elif selected_status == "saytga_kirmagan":
            where.append(f"NOT {site_entered_sql}")
        elif selected_status == "saytga_kirgan":
            where.append(site_entered_sql)
        if bosqich is not None:
            if bosqich not in (1, 2, 3, 4): raise HTTPException(status_code=400, detail="Bosqich 1–4 oralig'ida bo'lishi kerak")
            if bosqich == 2: where.append("qt.hujjat_topshirgan_at IS NOT NULL")
            if bosqich == 3: where.append("qt.bazaga_kiritilgan_at IS NOT NULL")
            if bosqich == 4: where.append(site_entered_sql)
        elif bosqich_min is not None:
            requested_stage = max(1, min(4, bosqich_min))
            if requested_stage == 2: where.append("qt.hujjat_topshirgan_at IS NOT NULL")
            if requested_stage == 3: where.append("qt.bazaga_kiritilgan_at IS NOT NULL")
            if requested_stage == 4: where.append(site_entered_sql)
        if fakultet_id:
            where.append("y.fakultet_id=%s"); params.append(fakultet_id)
            context_where.append("y.fakultet_id=%s"); context_params.append(fakultet_id)
        if yonalish_ids:
            try:
                requested_program_ids = sorted({int(value.strip()) for value in yonalish_ids.split(",") if value.strip()})
            except ValueError:
                raise HTTPException(status_code=400, detail="Yo‘nalish identifikatorlari noto‘g‘ri")
            if not requested_program_ids:
                raise HTTPException(status_code=400, detail="Yo‘nalish tanlanmagan")
            where.append("qt.yonalish_id=ANY(%s)"); params.append(requested_program_ids)
            context_where.append("qt.yonalish_id=ANY(%s)"); context_params.append(requested_program_ids)
        elif yonalish_id:
            where.append("qt.yonalish_id=%s"); params.append(yonalish_id)
            context_where.append("qt.yonalish_id=%s"); context_params.append(yonalish_id)

        # Dropdown variantlari boshqa dropdown tanlovi sabab yo'qolib ketmasin.
        # Fakultet/yo'nalish qamrovi saqlanadi, shakl/til/hudud esa mustaqil qoladi.
        option_where = list(context_where)
        option_params = list(context_params)
        if talim_shakli:
            where.append("qt.talim_shakli=%s"); params.append(talim_shakli)
            context_where.append("qt.talim_shakli=%s"); context_params.append(talim_shakli)
        if talim_tili:
            where.append("qt.talim_tili=%s"); params.append(talim_tili)
            context_where.append("qt.talim_tili=%s"); context_params.append(talim_tili)
        if region:
            where.append("qt.doimiy_region=%s"); params.append(region)
            context_where.append("qt.doimiy_region=%s"); context_params.append(region)
        admission_kind = _key(qabul_turi)
        if admission_kind in {"grant", "budjet", "budjetgrant"}:
            admission_filter = "(qt.tavsiya_turi ILIKE %s OR qt.tavsiya_turi ILIKE %s)"
            admission_params = ["%grant%", "%budjet%"]
            where.append(admission_filter); params.extend(admission_params)
            context_where.append(admission_filter); context_params.extend(admission_params)
        elif admission_kind in {"kontrakt", "shartnoma", "tolovkontrakt"}:
            admission_filter = "(qt.tavsiya_turi ILIKE %s OR qt.tavsiya_turi ILIKE %s)"
            admission_params = ["%kontrakt%", "%shartnoma%"]
            where.append(admission_filter); params.extend(admission_params)
            context_where.append(admission_filter); context_params.extend(admission_params)
        elif admission_kind:
            raise HTTPException(status_code=400, detail="Qabul turi grant yoki kontrakt bo'lishi kerak")
        if _norm(q):
            term = "%" + _norm(q) + "%"; digits = _digits(q)
            search_parts = [
                "qt.familiya ILIKE %s", "qt.ism ILIKE %s", "qt.ota_ism ILIKE %s",
                "CONCAT_WS(' ',qt.familiya,qt.ism,qt.ota_ism) ILIKE %s",
                "qt.abitur_id ILIKE %s", "qt.telefon ILIKE %s",
            ]
            params += [term, term, term, term, term, term]
            if len(digits) == 14:
                search_parts.append("qt.jshshir_hash=%s"); params.append(hashlib.sha256(digits.encode()).hexdigest())
            where.append("(" + " OR ".join(search_parts) + ")")
        order = {"ball_desc": "qt.ball DESC NULLS LAST,qt.familiya", "ball_asc": "qt.ball ASC NULLS LAST,qt.familiya", "name": "qt.familiya,qt.ism", "newest": "qt.id DESC"}.get(sort, "qt.ball DESC NULLS LAST")
        clause = " AND ".join(where)
        cur.execute(f"""SELECT COUNT(*) n FROM universitet_qabul_talabalari qt
            JOIN universitet_yonalishlari y ON y.id=qt.yonalish_id
            JOIN fakultetlar f ON f.id=y.fakultet_id
            JOIN kafedralar k ON k.id=y.kafedra_id
            WHERE {clause}""", params); total = int(cur.fetchone()["n"])
        page_size = max(10, min(100, page_size)); page = max(1, page); offset = (page - 1) * page_size
        cur.execute(f"""SELECT qt.id,qt.familiya,qt.ism,qt.ota_ism,qt.ball,qt.talim_shakli,qt.talim_tili,qt.tavsiya_turi,
            qt.doimiy_region,qt.doimiy_tuman,qt.qabul_bosqichi,qt.hujjat_topshirgan_at,
            qt.bazaga_kiritilgan_at,qt.birinchi_kirish_at,qt.telefon,qt.aloqa_izohi,qt.aloqa_izohi_at,y.id yonalish_id,y.nomi yonalish_nomi,
            (qt.hujjat_topshirgan_at IS NOT NULL) hujjat_topshirgan,
            (qt.bazaga_kiritilgan_at IS NOT NULL) bazaga_kiritilgan,
            {site_entered_sql} saytga_kirgan,
            CASE WHEN {site_entered_sql} THEN 'saytga_kirgan'
                 WHEN tk.kod_hash IS NOT NULL THEN 'kirish_kodi_tayyor'
                 ELSE 'bazaga_kiritilmagan' END sayt_holati
            FROM universitet_qabul_talabalari qt JOIN universitet_yonalishlari y ON y.id=qt.yonalish_id
            JOIN fakultetlar f ON f.id=y.fakultet_id
            JOIN kafedralar k ON k.id=y.kafedra_id
            LEFT JOIN LATERAL (SELECT tk.kod_hash FROM universitet_taklif_kodlari tk WHERE tk.qabul_talaba_id=qt.id ORDER BY tk.id DESC LIMIT 1) tk ON TRUE
            LEFT JOIN xodim_kod xk ON xk.kod=tk.kod_hash
            WHERE {clause} ORDER BY {order} LIMIT %s OFFSET %s""", params + [page_size, offset])
        rows = []
        for row_index, r in enumerate(cur.fetchall(), start=offset + 1):
            item = dict(r); item["fish"] = " ".join(x for x in [r["familiya"], r["ism"], r["ota_ism"]] if x); item["telefon_mask"] = _mask_phone(r["telefon"]); item["tartib_raqami"] = row_index
            if tutor_only:
                item.pop("telefon", None)
                item.pop("qabul_bosqichi", None)
                item["bazaga_belgilash_mumkin"] = bool(item["hujjat_topshirgan"] and not item["bazaga_kiritilgan"])
                item.pop("hujjat_topshirgan", None)
                for hidden in ("hujjat_topshirgan_at", "bazaga_kiritilgan_at", "birinchi_kirish_at"):
                    item.pop(hidden, None)
            rows.append(item)
        context_clause = " AND ".join(context_where)
        cur.execute(f"""SELECT COUNT(*) jami,
            COUNT(*) FILTER(WHERE qt.hujjat_topshirgan_at IS NOT NULL) hujjat,
            COUNT(*) FILTER(WHERE qt.hujjat_topshirgan_at IS NULL) topshirmagan,
            COUNT(*) FILTER(WHERE qt.tavsiya_turi ILIKE '%%grant%%'
                                  OR qt.tavsiya_turi ILIKE '%%budjet%%') AS "grant",
            COUNT(*) FILTER(WHERE qt.tavsiya_turi ILIKE '%%kontrakt%%'
                                  OR qt.tavsiya_turi ILIKE '%%shartnoma%%') AS kontrakt,
            COUNT(*) FILTER(WHERE qt.bazaga_kiritilgan_at IS NOT NULL) baza,
            COUNT(*) FILTER(WHERE {site_entered_sql}) sayt,
            COUNT(*) FILTER(WHERE NOT {site_entered_sql}) saytga_kirmagan,
            COUNT(*) FILTER(WHERE NULLIF(BTRIM(qt.aloqa_izohi),'') IS NOT NULL) suhbat
            FROM universitet_qabul_talabalari qt JOIN universitet_yonalishlari y ON y.id=qt.yonalish_id
            JOIN fakultetlar f ON f.id=y.fakultet_id
            JOIN kafedralar k ON k.id=y.kafedra_id
            WHERE {context_clause}""", context_params)
        counts = cur.fetchone()
        option_clause = " AND ".join(option_where)
        cur.execute(f"""SELECT
            ARRAY_REMOVE(ARRAY_AGG(DISTINCT qt.talim_shakli ORDER BY qt.talim_shakli),NULL) shakllar,
            ARRAY_REMOVE(ARRAY_AGG(DISTINCT qt.talim_tili ORDER BY qt.talim_tili),NULL) tillar,
            ARRAY_REMOVE(ARRAY_AGG(DISTINCT qt.doimiy_region ORDER BY qt.doimiy_region),NULL) hududlar
            FROM universitet_qabul_talabalari qt JOIN universitet_yonalishlari y ON y.id=qt.yonalish_id
            JOIN fakultetlar f ON f.id=y.fakultet_id
            JOIN kafedralar k ON k.id=y.kafedra_id
            WHERE {option_clause}""", option_params)
        filter_options = dict(cur.fetchone())
        filter_options["qabul_turlari"] = ["grant", "kontrakt"]
        safe_counts = {"jami": int(counts["jami"] or 0), "baza": int(counts["baza"] or 0),
                       "sayt": int(counts["sayt"] or 0),
                       "saytga_kirmagan": int(counts["saytga_kirmagan"] or 0),
                       "suhbat": int(counts["suhbat"] or 0)} if tutor_only else counts
        return {"talabalar": rows, "jami": total, "sahifa": page, "sahifa_soni": math.ceil(total/page_size) if total else 0,
                "hisoblar": safe_counts, "filtrlar": filter_options, "holat": selected_status,
                "tyutor_rejimi": tutor_only}
    finally:
        cur.close(); conn.close()


def _admission_report_range(
    boshlanish: Optional[date],
    tugash: Optional[date],
    kun: Optional[date] = None,
) -> tuple[date, date]:
    """Yangi sana oralig'i va eski bitta-kun kontraktini birga qo'llaydi."""
    if kun is not None and boshlanish is None and tugash is None:
        boshlanish = tugash = kun
    elif boshlanish is None and tugash is None:
        boshlanish = tugash = date.today()
    elif boshlanish is None:
        boshlanish = tugash
    elif tugash is None:
        tugash = boshlanish
    if boshlanish is None or tugash is None or boshlanish > tugash:
        raise HTTPException(status_code=400, detail="Hisobot sana oralig'i noto'g'ri")
    if (tugash - boshlanish).days > 365:
        raise HTTPException(status_code=400, detail="Bir so'rovda ko'pi bilan 366 kun tanlanadi")
    return boshlanish, tugash


def _validate_admission_report_dimensions(
    cur,
    university_id: int,
    faculty_id: Optional[int] = None,
    department_id: Optional[int] = None,
    program_id: Optional[int] = None,
) -> None:
    """Hisobot filtrlari bir faol institut zanjiriga tegishli ekanini tekshiradi."""
    if program_id is not None:
        cur.execute("""SELECT y.fakultet_id,y.kafedra_id
            FROM universitet_yonalishlari y
            JOIN fakultetlar f ON f.id=y.fakultet_id AND f.faol=TRUE
            JOIN kafedralar k ON k.id=y.kafedra_id AND k.faol=TRUE
            WHERE y.id=%s AND y.universitet_id=%s AND y.faol=TRUE""",
                    (program_id, university_id))
        row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=400, detail="Yo'nalish bu institutga tegishli emas")
        if faculty_id is not None and int(row["fakultet_id"]) != int(faculty_id):
            raise HTTPException(status_code=400, detail="Yo'nalish tanlangan fakultetga tegishli emas")
        if department_id is not None and int(row["kafedra_id"]) != int(department_id):
            raise HTTPException(status_code=400, detail="Yo'nalish tanlangan kafedraga tegishli emas")
        return
    if department_id is not None:
        cur.execute("""SELECT k.fakultet_id
            FROM kafedralar k JOIN fakultetlar f ON f.id=k.fakultet_id
            WHERE k.id=%s AND f.universitet_id=%s
              AND k.faol=TRUE AND f.faol=TRUE""", (department_id, university_id))
        row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=400, detail="Kafedra bu institutga tegishli emas")
        if faculty_id is not None and int(row["fakultet_id"]) != int(faculty_id):
            raise HTTPException(status_code=400, detail="Kafedra tanlangan fakultetga tegishli emas")
        return
    if faculty_id is not None:
        cur.execute("SELECT 1 FROM fakultetlar WHERE id=%s AND universitet_id=%s AND faol=TRUE",
                    (faculty_id, university_id))
        if not cur.fetchone():
            raise HTTPException(status_code=400, detail="Fakultet bu institutga tegishli emas")


def _admission_report_access(cur, university_id: int, user_id: int) -> list[dict[str, Any]]:
    roles = _require_member(cur, user_id, university_id)
    if not _has_any(roles, REPORT_VIEW_ROLES):
        raise HTTPException(status_code=403, detail="Qabul hisobotini ko'rish huquqi yo'q")
    return roles


def _admission_daily_report_data(
    cur,
    university_id: int,
    user_id: int,
    boshlanish: date,
    tugash: date,
    faculty_id: Optional[int] = None,
) -> dict[str, Any]:
    roles = _admission_report_access(cur, university_id, user_id)
    _validate_admission_report_dimensions(cur, university_id, faculty_id=faculty_id)
    scope_sql, scope_params = _student_scope_clause(cur, university_id, user_id, roles)
    where = [
        "qt.universitet_id=%s",
        "f.faol=TRUE",
        "k.faol=TRUE",
        "y.faol=TRUE",
        scope_sql,
        "event.event_at >= (%s::date AT TIME ZONE 'Asia/Tashkent')",
        "event.event_at < ((%s::date + 1) AT TIME ZONE 'Asia/Tashkent')",
    ]
    params: list[Any] = [university_id, *scope_params, boshlanish, tugash]
    if faculty_id is not None:
        where.append("y.fakultet_id=%s")
        params.append(faculty_id)
    cur.execute(f"""SELECT
            f.id fakultet_id,f.nomi fakultet_nomi,
            k.id kafedra_id,k.nomi kafedra_nomi,
            y.id yonalish_id,y.nomi yonalish_nomi,
            qt.talim_shakli,qt.talim_tili,
            event.event_type,
            (event.event_at AT TIME ZONE 'Asia/Tashkent')::date event_day,
            COUNT(*)::INTEGER soni
        FROM universitet_qabul_talabalari qt
        JOIN universitet_yonalishlari y ON y.id=qt.yonalish_id
        JOIN fakultetlar f ON f.id=y.fakultet_id
        JOIN kafedralar k ON k.id=y.kafedra_id
        CROSS JOIN LATERAL (VALUES
            ('hujjat',qt.hujjat_topshirgan_at),
            ('hemis',qt.bazaga_kiritilgan_at),
            ('sayt',COALESCE(qt.birinchi_kirish_at,qt.saytga_kiritilgan_at))
        ) AS event(event_type,event_at)
        WHERE {' AND '.join(where)}
        GROUP BY f.id,f.nomi,k.id,k.nomi,y.id,y.nomi,
                 qt.talim_shakli,qt.talim_tili,event.event_type,event_day
        ORDER BY f.nomi,k.nomi,y.nomi,qt.talim_shakli,qt.talim_tili,event_day""", params)

    days = [
        (boshlanish + timedelta(days=offset)).isoformat()
        for offset in range((tugash - boshlanish).days + 1)
    ]
    grouped: dict[tuple[Any, ...], dict[str, Any]] = {}
    totals = {"hujjat": 0, "baza": 0, "hemis": 0, "sayt": 0}
    day_totals = {day: 0 for day in days}
    for raw in cur.fetchall():
        item = dict(raw)
        key = (
            item["fakultet_id"], item["kafedra_id"], item["yonalish_id"],
            item.get("talim_shakli"), item.get("talim_tili"),
        )
        report_row = grouped.setdefault(key, {
            "fakultet_id": int(item["fakultet_id"]),
            "fakultet_nomi": item["fakultet_nomi"],
            "kafedra_id": int(item["kafedra_id"]),
            "kafedra_nomi": item["kafedra_nomi"],
            "yonalish_id": int(item["yonalish_id"]),
            "asosiy_yonalish_nomi": item["yonalish_nomi"],
            "yonalish_nomi": " · ".join(filter(None, (
                item["yonalish_nomi"], item.get("talim_shakli"), item.get("talim_tili"),
            ))),
            "talim_shakli": item.get("talim_shakli"),
            "talim_tili": item.get("talim_tili"),
            "kunlar": {day: 0 for day in days},
            "hemis_kunlar": {day: 0 for day in days},
            "sayt_kunlar": {day: 0 for day in days},
            "hujjat": 0,
            "baza": 0,
            "hemis": 0,
            "sayt": 0,
            "jami": 0,
        })
        event_type = item["event_type"]
        event_day = item["event_day"].isoformat()
        count = int(item["soni"] or 0)
        if event_type == "hujjat":
            report_row["kunlar"][event_day] = count
            report_row["hujjat"] += count
            report_row["jami"] += count
            totals["hujjat"] += count
            day_totals[event_day] = day_totals.get(event_day, 0) + count
        elif event_type == "hemis":
            report_row["hemis_kunlar"][event_day] = count
            report_row["baza"] += count
            report_row["hemis"] += count
            totals["baza"] += count
            totals["hemis"] += count
        elif event_type == "sayt":
            report_row["sayt_kunlar"][event_day] = count
            report_row["sayt"] += count
            totals["sayt"] += count

    rows = list(grouped.values())
    document_rows = [row for row in rows if row["hujjat"] > 0]
    return {
        "boshlanish": boshlanish.isoformat(),
        "tugash": tugash.isoformat(),
        "kun": boshlanish.isoformat() if boshlanish == tugash else None,
        "kunlar": days,
        "sanalar": days,
        "hisoblar": totals,
        "jami_kunlar": day_totals,
        "kun_jami": day_totals,
        "jami": totals["hujjat"],
        "qatorlar": document_rows,
        # Eski bir-kunlik frontend baza/sayt ustunlarini shu aliasdan o'qiydi.
        "yonalishlar": rows,
    }


def _admission_general_report_data(
    cur,
    university_id: int,
    user_id: int,
    *,
    q: str = "",
    faculty_id: Optional[int] = None,
    department_id: Optional[int] = None,
    program_id: Optional[int] = None,
    talim_shakli: Optional[str] = None,
    talim_tili: Optional[str] = None,
    qabul_turi: Optional[str] = None,
) -> dict[str, Any]:
    roles = _admission_report_access(cur, university_id, user_id)
    _validate_admission_report_dimensions(
        cur, university_id, faculty_id=faculty_id,
        department_id=department_id, program_id=program_id,
    )
    scope_sql, scope_params = _student_scope_clause(cur, university_id, user_id, roles)
    where = [
        "qt.universitet_id=%s", "f.faol=TRUE", "k.faol=TRUE", "y.faol=TRUE", scope_sql,
    ]
    params: list[Any] = [university_id, *scope_params]
    if faculty_id is not None:
        where.append("y.fakultet_id=%s"); params.append(faculty_id)
    if department_id is not None:
        where.append("y.kafedra_id=%s"); params.append(department_id)
    if program_id is not None:
        where.append("qt.yonalish_id=%s"); params.append(program_id)

    # Dropdown variantlari qidiruv yoki boshqa dropdown tanlovi sabab yo'qolmasin.
    option_clause = " AND ".join(where)
    option_params = list(params)
    if talim_shakli:
        where.append("qt.talim_shakli=%s"); params.append(talim_shakli)
    if talim_tili:
        where.append("qt.talim_tili=%s"); params.append(talim_tili)
    admission_kind = _key(qabul_turi)
    if admission_kind in {"grant", "budjet", "budjetgrant"}:
        where.append("(qt.tavsiya_turi ILIKE %s OR qt.tavsiya_turi ILIKE %s)")
        params.extend(["%grant%", "%budjet%"])
    elif admission_kind in {"kontrakt", "shartnoma", "tolovkontrakt"}:
        where.append("(qt.tavsiya_turi ILIKE %s OR qt.tavsiya_turi ILIKE %s)")
        params.extend(["%kontrakt%", "%shartnoma%"])
    elif admission_kind:
        raise HTTPException(status_code=400, detail="Qabul turi grant yoki kontrakt bo'lishi kerak")
    if _norm(q):
        term = "%" + _norm(q) + "%"
        where.append("""(qt.familiya ILIKE %s OR qt.ism ILIKE %s OR qt.ota_ism ILIKE %s
            OR CONCAT_WS(' ',qt.familiya,qt.ism,qt.ota_ism) ILIKE %s
            OR qt.abitur_id ILIKE %s)""")
        params.extend([term, term, term, term, term])

    cur.execute(f"""SELECT
            f.id fakultet_id,f.nomi fakultet_nomi,
            k.id kafedra_id,k.nomi kafedra_nomi,
            y.id yonalish_id,y.nomi yonalish_nomi,
            COUNT(*)::INTEGER jami,
            (COUNT(*) FILTER(WHERE qt.hujjat_topshirgan_at IS NOT NULL))::INTEGER hujjat,
            (COUNT(*) FILTER(WHERE qt.bazaga_kiritilgan_at IS NOT NULL))::INTEGER hemis
        FROM universitet_qabul_talabalari qt
        JOIN universitet_yonalishlari y ON y.id=qt.yonalish_id
        JOIN fakultetlar f ON f.id=y.fakultet_id
        JOIN kafedralar k ON k.id=y.kafedra_id
        WHERE {' AND '.join(where)}
        GROUP BY f.id,f.nomi,k.id,k.nomi,y.id,y.nomi
        ORDER BY f.nomi,k.nomi,y.nomi""", params)
    rows: list[dict[str, Any]] = []
    totals = {"jami": 0, "hujjat": 0, "qolgan": 0, "hemis": 0,
              "baza": 0, "hemis_kutilmoqda": 0}
    for raw in cur.fetchall():
        row = dict(raw)
        row["jami"] = int(row["jami"] or 0)
        row["hujjat"] = int(row["hujjat"] or 0)
        row["hemis"] = int(row["hemis"] or 0)
        row["baza"] = row["hemis"]
        row["qolgan"] = max(0, row["jami"] - row["hujjat"])
        row["hemis_kutilmoqda"] = max(0, row["hujjat"] - row["hemis"])
        for key in totals:
            totals[key] += int(row[key] or 0)
        rows.append(row)

    cur.execute(f"""SELECT
            ARRAY_REMOVE(ARRAY_AGG(DISTINCT qt.talim_shakli ORDER BY qt.talim_shakli),NULL) shakllar,
            ARRAY_REMOVE(ARRAY_AGG(DISTINCT qt.talim_tili ORDER BY qt.talim_tili),NULL) tillar
        FROM universitet_qabul_talabalari qt
        JOIN universitet_yonalishlari y ON y.id=qt.yonalish_id
        JOIN fakultetlar f ON f.id=y.fakultet_id
        JOIN kafedralar k ON k.id=y.kafedra_id
        WHERE {option_clause}""", option_params)
    options = dict(cur.fetchone())
    options["shakllar"] = options.get("shakllar") or []
    options["tillar"] = options.get("tillar") or []
    options["qabul_turlari"] = ["grant", "kontrakt"]
    return {"hisoblar": totals, "qatorlar": rows, "yonalishlar": rows, "filtrlar": options}


def _admission_report_xlsx(
    title: str,
    subtitle: str,
    headers: list[str],
    rows: list[list[Any]],
    widths: list[int],
    freeze_panes: str = "A4",
) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="Hisobot XLSX uchun openpyxl o'rnatilmagan") from exc
    wb = Workbook(); ws = wb.active; ws.title = "QABUL HISOBOTI"
    ws.append([title] + [None] * (len(headers) - 1))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.append([subtitle] + [None] * (len(headers) - 1))
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws.append(headers)
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=15)
    ws["A1"].fill = PatternFill("solid", fgColor="0D7A77")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A2"].font = Font(italic=True, color="5A6A72")
    ws["A2"].alignment = Alignment(horizontal="center")
    for cell in ws[3]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="173E5B")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in rows:
        ws.append(row)
    for column, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(column)].width = width
    ws.freeze_panes = freeze_panes
    ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{max(3, ws.max_row)}"
    ws.sheet_view.showGridLines = False
    for row in ws.iter_rows(min_row=4):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    stream = io.BytesIO(); wb.save(stream); wb.close(); return stream.getvalue()


def _xlsx_download(content: bytes, filename: str) -> Response:
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/qabul/kunlik_hisobot")
def admission_daily_report(
    universitet_id: int,
    boshlanish: Optional[date] = None,
    tugash: Optional[date] = None,
    kun: Optional[date] = None,
    fakultet_id: Optional[int] = None,
    token: Optional[str] = Query(None, include_in_schema=False),
    authorization: Optional[str] = Header(None),
):
    """Hujjat qabuli va keyingi bosqichlarni sana oralig'ida jamlaydi."""
    start, end = _admission_report_range(boshlanish, tugash, kun)
    user_id = _uid(token, authorization); p = _p(); conn = p._db(); cur = conn.cursor()
    try:
        _ensure_schema(cur)
        return _admission_daily_report_data(cur, universitet_id, user_id, start, end, fakultet_id)
    finally:
        cur.close(); conn.close()


@router.get("/qabul/kunlik_hisobot.xlsx")
def admission_daily_report_xlsx(
    universitet_id: int,
    boshlanish: Optional[date] = None,
    tugash: Optional[date] = None,
    kun: Optional[date] = None,
    fakultet_id: Optional[int] = None,
    token: Optional[str] = Query(None, include_in_schema=False),
    authorization: Optional[str] = Header(None),
):
    start, end = _admission_report_range(boshlanish, tugash, kun)
    user_id = _uid(token, authorization); p = _p(); conn = p._db(); cur = conn.cursor()
    try:
        _ensure_schema(cur)
        report = _admission_daily_report_data(cur, universitet_id, user_id, start, end, fakultet_id)
        day_headers = [datetime.fromisoformat(day).strftime("%d.%m") for day in report["kunlar"]]
        body = []
        for index, row in enumerate(report["qatorlar"], 1):
            body.append([
                index, row["fakultet_nomi"], row["kafedra_nomi"],
                row["asosiy_yonalish_nomi"], row.get("talim_shakli"), row.get("talim_tili"),
                *[row["kunlar"].get(day, 0) for day in report["kunlar"]], row["jami"],
            ])
        body.append([None, None, None, "Jami", None, None,
                     *[report["jami_kunlar"].get(day, 0) for day in report["kunlar"]], report["jami"]])
        headers = ["№", "Fakultet", "Kafedra", "Yo'nalish", "Ta'lim shakli", "Ta'lim tili",
                   *day_headers, "Jami"]
        content = _admission_report_xlsx(
            "HUJJAT QABULI — KUNLIK HISOBOT",
            f"{start.isoformat()} — {end.isoformat()}",
            headers, body, [7, 26, 28, 38, 18, 18, *([11] * len(day_headers)), 12], "G4",
        )
        return _xlsx_download(content, "qabul_kunlik_hisobot.xlsx")
    finally:
        cur.close(); conn.close()


@router.get("/qabul/umumiy_hisobot")
def admission_general_report(
    universitet_id: int,
    q: str = "",
    fakultet_id: Optional[int] = None,
    kafedra_id: Optional[int] = None,
    yonalish_id: Optional[int] = None,
    talim_shakli: Optional[str] = None,
    talim_tili: Optional[str] = None,
    qabul_turi: Optional[str] = None,
    token: Optional[str] = Query(None, include_in_schema=False),
    authorization: Optional[str] = Header(None),
):
    user_id = _uid(token, authorization); p = _p(); conn = p._db(); cur = conn.cursor()
    try:
        _ensure_schema(cur)
        return _admission_general_report_data(
            cur, universitet_id, user_id, q=q, faculty_id=fakultet_id,
            department_id=kafedra_id, program_id=yonalish_id,
            talim_shakli=talim_shakli, talim_tili=talim_tili, qabul_turi=qabul_turi,
        )
    finally:
        cur.close(); conn.close()


@router.get("/qabul/umumiy_hisobot.xlsx")
def admission_general_report_xlsx(
    universitet_id: int,
    q: str = "",
    fakultet_id: Optional[int] = None,
    kafedra_id: Optional[int] = None,
    yonalish_id: Optional[int] = None,
    talim_shakli: Optional[str] = None,
    talim_tili: Optional[str] = None,
    qabul_turi: Optional[str] = None,
    token: Optional[str] = Query(None, include_in_schema=False),
    authorization: Optional[str] = Header(None),
):
    user_id = _uid(token, authorization); p = _p(); conn = p._db(); cur = conn.cursor()
    try:
        _ensure_schema(cur)
        report = _admission_general_report_data(
            cur, universitet_id, user_id, q=q, faculty_id=fakultet_id,
            department_id=kafedra_id, program_id=yonalish_id,
            talim_shakli=talim_shakli, talim_tili=talim_tili, qabul_turi=qabul_turi,
        )
        body = [[
            index, row["fakultet_nomi"], row["kafedra_nomi"], row["yonalish_nomi"],
            row["jami"], row["hujjat"], row["qolgan"], row["hemis"], row["hemis_kutilmoqda"],
        ] for index, row in enumerate(report["qatorlar"], 1)]
        totals = report["hisoblar"]
        body.append([None, None, None, "Jami", totals["jami"], totals["hujjat"],
                     totals["qolgan"], totals["hemis"], totals["hemis_kutilmoqda"]])
        content = _admission_report_xlsx(
            "HUJJAT QABULI — UMUMIY HISOBOT",
            "Filtrlangan natija; shaxsiy ma'lumotlar eksport qilinmaydi",
            ["№", "Fakultet", "Kafedra", "Yo'nalish", "Jami", "Hujjat", "Qolgan", "HEMIS", "HEMIS kutilmoqda"],
            body, [7, 26, 28, 40, 12, 12, 12, 12, 18], "A4",
        )
        return _xlsx_download(content, "qabul_umumiy_hisobot.xlsx")
    finally:
        cur.close(); conn.close()


def _admission_students_xlsx_multi(sheets: list[dict[str, Any]]) -> bytes:
    """Bir necha varaqli talabalar ro'yxati (topshirgan / topshirmagan / xulosa)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="Hisobot XLSX uchun openpyxl o'rnatilmagan") from exc
    wb = Workbook(); first = True
    for sheet in sheets:
        ws = wb.active if first else wb.create_sheet(); first = False
        ws.title = str(sheet["title"])[:31]
        headers = sheet["headers"]; n = len(headers)
        ws.append([sheet["heading"]] + [None] * (n - 1)); ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
        ws.append([sheet.get("subtitle", "")] + [None] * (n - 1)); ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n)
        ws.append(headers)
        ws["A1"].font = Font(bold=True, color="FFFFFF", size=14); ws["A1"].fill = PatternFill("solid", fgColor=sheet.get("color", "0D7A77")); ws["A1"].alignment = Alignment(horizontal="center")
        ws["A2"].font = Font(italic=True, color="5A6A72"); ws["A2"].alignment = Alignment(horizontal="center")
        for cell in ws[3]:
            cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor="173E5B")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in sheet["rows"]:
            ws.append(row)
        for column, width in enumerate(sheet["widths"], 1):
            ws.column_dimensions[get_column_letter(column)].width = width
        ws.freeze_panes = "A4"
        ws.auto_filter.ref = f"A3:{get_column_letter(n)}{max(3, ws.max_row)}"
        ws.sheet_view.showGridLines = False
        for row in ws.iter_rows(min_row=4):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    stream = io.BytesIO(); wb.save(stream); wb.close(); return stream.getvalue()


@router.get("/qabul/talabalar_royxati.xlsx")
def admission_students_xlsx(
    universitet_id: int,
    q: str = "",
    fakultet_id: Optional[int] = None,
    yonalish_id: Optional[int] = None,
    yonalish_ids: Optional[str] = None,
    talim_shakli: Optional[str] = None,
    talim_tili: Optional[str] = None,
    qabul_turi: Optional[str] = None,
    token: Optional[str] = Query(None, include_in_schema=False),
    authorization: Optional[str] = Header(None),
):
    """Talabalar ro'yxati Excel: 1-varaq hujjat topshirganlar, 2-varaq topshirmaganlar (izoh bilan), 3-varaq xulosa."""
    user_id = _uid(token, authorization); p = _p(); conn = p._db(); cur = conn.cursor()
    try:
        _ensure_schema(cur); roles = _require_member(cur, user_id, universitet_id); names = _role_names(roles)
        if not (names & MARK_DOCUMENT_ROLES):
            raise HTTPException(status_code=403, detail="Talabalar ro'yxatini eksport qilish huquqi yo'q")
        scope_sql, scope_params = _student_scope_clause(cur, universitet_id, user_id, roles)
        where = ["qt.universitet_id=%s", "y.faol=TRUE", "f.faol=TRUE", "k.faol=TRUE", scope_sql]; params: list[Any] = [universitet_id, *scope_params]
        if fakultet_id: where.append("y.fakultet_id=%s"); params.append(fakultet_id)
        if yonalish_ids:
            try: ids = sorted({int(v.strip()) for v in yonalish_ids.split(",") if v.strip()})
            except ValueError: raise HTTPException(status_code=400, detail="Yo‘nalish identifikatorlari noto‘g‘ri")
            if ids: where.append("qt.yonalish_id=ANY(%s)"); params.append(ids)
        elif yonalish_id: where.append("qt.yonalish_id=%s"); params.append(yonalish_id)
        if talim_shakli: where.append("qt.talim_shakli=%s"); params.append(talim_shakli)
        if talim_tili: where.append("qt.talim_tili=%s"); params.append(talim_tili)
        kind = _key(qabul_turi)
        if kind in {"grant", "budjet", "budjetgrant"}: where.append("(qt.tavsiya_turi ILIKE %s OR qt.tavsiya_turi ILIKE %s)"); params += ["%grant%", "%budjet%"]
        elif kind in {"kontrakt", "shartnoma", "tolovkontrakt"}: where.append("(qt.tavsiya_turi ILIKE %s OR qt.tavsiya_turi ILIKE %s)"); params += ["%kontrakt%", "%shartnoma%"]
        if _norm(q):
            term = "%" + _norm(q) + "%"
            where.append("(qt.familiya ILIKE %s OR qt.ism ILIKE %s OR qt.ota_ism ILIKE %s OR CONCAT_WS(' ',qt.familiya,qt.ism,qt.ota_ism) ILIKE %s OR qt.abitur_id ILIKE %s OR qt.telefon ILIKE %s)")
            params += [term] * 6
        cur.execute(f"""SELECT qt.familiya,qt.ism,qt.ota_ism,qt.abitur_id,qt.ball,qt.talim_shakli,qt.talim_tili,qt.tavsiya_turi,
                qt.doimiy_region,qt.doimiy_tuman,qt.telefon,qt.hujjat_topshirgan_at,qt.bazaga_kiritilgan_at,
                qt.aloqa_izohi,qt.aloqa_izohi_at,y.nomi yonalish_nomi,k.nomi kafedra_nomi,f.nomi fakultet_nomi
            FROM universitet_qabul_talabalari qt JOIN universitet_yonalishlari y ON y.id=qt.yonalish_id
            JOIN fakultetlar f ON f.id=y.fakultet_id JOIN kafedralar k ON k.id=y.kafedra_id
            WHERE {" AND ".join(where)}
            ORDER BY f.nomi,k.nomi,y.nomi,qt.familiya,qt.ism""", params)
        rows = [dict(r) for r in cur.fetchall()]
        _audit(cur, universitet_id, user_id, "talabalar_royxati_xlsx", "universitet", universitet_id, {"soni": len(rows)})
        conn.commit()

        def fish(r): return " ".join(x for x in [r["familiya"], r["ism"], r["ota_ism"]] if x)
        def dt(v): return v.strftime("%d.%m.%Y") if v else None
        def region(r): return ", ".join(x for x in [r["doimiy_region"], r["doimiy_tuman"]] if x) or None
        submitted = [r for r in rows if r["hujjat_topshirgan_at"] is not None]
        pending = [r for r in rows if r["hujjat_topshirgan_at"] is None]
        with_note = sum(1 for r in pending if (r["aloqa_izohi"] or "").strip())
        submitted_rows = [[i, fish(r), r["yonalish_nomi"], r["kafedra_nomi"], r["talim_shakli"], r["talim_tili"], r["tavsiya_turi"], r["ball"], r["telefon"], region(r), dt(r["hujjat_topshirgan_at"]), "Ha" if r["bazaga_kiritilgan_at"] else "Yo'q", r["aloqa_izohi"]] for i, r in enumerate(submitted, 1)]
        pending_rows = [[i, fish(r), r["yonalish_nomi"], r["kafedra_nomi"], r["talim_shakli"], r["talim_tili"], r["tavsiya_turi"], r["ball"], r["telefon"], region(r), r["aloqa_izohi"], dt(r["aloqa_izohi_at"])] for i, r in enumerate(pending, 1)]
        # Xulosa: yo'nalish kesimida
        summary: dict[str, dict[str, Any]] = {}
        for r in rows:
            key = (r["fakultet_nomi"], r["kafedra_nomi"], r["yonalish_nomi"])
            item = summary.setdefault(key, {"jami": 0, "hujjat": 0, "qolgan": 0, "hemis": 0, "izoh": 0})
            item["jami"] += 1
            if r["hujjat_topshirgan_at"]: item["hujjat"] += 1
            else:
                item["qolgan"] += 1
                if (r["aloqa_izohi"] or "").strip(): item["izoh"] += 1
            if r["bazaga_kiritilgan_at"]: item["hemis"] += 1
        summary_rows = [[i, k[0], k[1], k[2], v["jami"], v["hujjat"], v["qolgan"], v["izoh"], v["hemis"]] for i, (k, v) in enumerate(sorted(summary.items()), 1)]
        summary_rows.append([None, None, None, "Jami", len(rows), len(submitted), len(pending), with_note, sum(1 for r in rows if r["bazaga_kiritilgan_at"])])
        content = _admission_students_xlsx_multi([
            {"title": "Hujjat topshirgan", "heading": f"HUJJAT TOPSHIRGANLAR — {len(submitted)} ta", "subtitle": "Fakultet → kafedra → yo'nalish → familiya tartibida", "color": "8A5A1C",
             "headers": ["№", "F.I.Sh.", "Yo'nalish", "Kafedra", "Ta'lim shakli", "Ta'lim tili", "Qabul turi", "Ball", "Telefon", "Hudud", "Topshirgan sana", "HEMIS", "Izoh"],
             "widths": [6, 34, 28, 22, 12, 12, 16, 8, 16, 26, 14, 8, 36], "rows": submitted_rows},
            {"title": "Hujjat topshirmagan", "heading": f"HUJJAT TOPSHIRMAGANLAR — {len(pending)} ta (izohli: {with_note})", "subtitle": "Telefon suhbati izohi bilan: nega topshirmadi, qachon keladi", "color": "A84444",
             "headers": ["№", "F.I.Sh.", "Yo'nalish", "Kafedra", "Ta'lim shakli", "Ta'lim tili", "Qabul turi", "Ball", "Telefon", "Hudud", "Suhbat izohi", "Izoh sanasi"],
             "widths": [6, 34, 28, 22, 12, 12, 16, 8, 16, 26, 44, 12], "rows": pending_rows},
            {"title": "Xulosa", "heading": "YO'NALISHLAR KESIMIDA XULOSA", "subtitle": "Filtrlangan natija", "color": "0D7A77",
             "headers": ["№", "Fakultet", "Kafedra", "Yo'nalish", "Jami", "Hujjat topshirgan", "Topshirmagan", "Izohli (suhbat)", "HEMIS"],
             "widths": [6, 26, 26, 36, 10, 16, 14, 14, 10], "rows": summary_rows},
        ])
        return _xlsx_download(content, "qabul_talabalar_royxati.xlsx")
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


def _abitur_id_key(value: Any) -> str:
    """Excel'dan kelgan AbiturID ni normallashtiradi: 12345.0 -> 12345, bo'shliqlar olib tashlanadi."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return re.sub(r"\s+", "", text).upper()


@router.get("/qabul/hujjat_import_shablon.xlsx")
def admission_document_import_template(universitet_id: int, token: Optional[str] = Query(None, include_in_schema=False), authorization: Optional[str] = Header(None)):
    """Bitta ustunli shablon: AbiturID. Xodim ID larni qo'yib import qiladi — hammasi “Hujjat topshirgan” bo'ladi."""
    user_id = _uid(token, authorization); p = _p(); conn = p._db(); cur = conn.cursor()
    try:
        _ensure_schema(cur); roles = _require_member(cur, user_id, universitet_id)
        if not _has_any(roles, MARK_DOCUMENT_ROLES): raise HTTPException(status_code=403, detail="Hujjat belgilash huquqi yo'q")
    finally:
        cur.close(); conn.close()
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="openpyxl o'rnatilmagan") from exc
    wb = Workbook(); ws = wb.active; ws.title = "HUJJAT"
    ws.append(["AbiturID"]); ws.append(["1234567"]); ws.append(["1234568"])
    ws["A1"].font = Font(bold=True, color="FFFFFF"); ws["A1"].fill = PatternFill("solid", fgColor="173E5B"); ws["A1"].alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 18
    ws["C1"] = "Yo'riqnoma"; ws["C1"].font = Font(bold=True)
    ws["C2"] = "A ustuniga hujjat topshirgan abituriyentlarning AbiturID raqamlarini yozing (har qatorga bittadan)."
    ws["C3"] = "Namuna qatorlarni (1234567, 1234568) o'chirib tashlang. Boshqa ustunlar o'qilmaydi."
    ws["C4"] = "Faylni “Hujjat topshirganlarni import qilish” tugmasi orqali yuklang — ID lar bazadan topilib, “Hujjat topshirgan” qilib belgilanadi."
    ws.column_dimensions["C"].width = 110
    ws.freeze_panes = "A2"
    stream = io.BytesIO(); wb.save(stream); wb.close()
    return _xlsx_download(stream.getvalue(), "hujjat_topshirganlar_shablon.xlsx")


_DOC_IMPORT_ACTIONS = {
    # amal: (ustun, yangi qiymat SQL, "allaqachon shunday" sharti, huquq to'plami nomi, tavsif)
    "hujjat_belgilash": ("hujjat_topshirgan_at", "NOW()", "IS NOT NULL", "hujjat", "Hujjat topshirgan qilish"),
    "hujjat_bekor": ("hujjat_topshirgan_at", "NULL", "IS NULL", "hujjat", "Hujjat topshirgan belgisini olib tashlash"),
    "baza_belgilash": ("bazaga_kiritilgan_at", "NOW()", "IS NOT NULL", "baza", "Bazaga (HEMIS) kiritilgan qilish"),
    "baza_bekor": ("bazaga_kiritilgan_at", "NULL", "IS NULL", "baza", "Baza belgisini olib tashlash"),
}


@router.post("/qabul/hujjat_import")
async def admission_document_import(universitet_id: int, fayl: UploadFile = File(...), tasdiqlash: int = 0, amal: str = "hujjat_belgilash", token: Optional[str] = Query(None, include_in_schema=False), authorization: Optional[str] = Header(None)):
    """Excel'dagi AbiturID lar bo'yicha ommaviy amal: hujjat belgilash / bekor, baza belgilash / bekor.

    tasdiqlash=0 — faqat tekshirish; tasdiqlash=1 — bajarish.
    """
    if amal not in _DOC_IMPORT_ACTIONS: raise HTTPException(status_code=400, detail="Amal noto'g'ri")
    column, new_value_sql, already_sql, right, action_label = _DOC_IMPORT_ACTIONS[amal]
    user_id = _uid(token, authorization); p = _p()
    content = await fayl.read()
    if len(content) > 10 * 1024 * 1024: raise HTTPException(status_code=413, detail="Fayl 10 MB dan katta")
    if not (fayl.filename or "").lower().endswith((".xls", ".xlsx")): raise HTTPException(status_code=400, detail="Faqat .xls yoki .xlsx fayl qabul qilinadi")
    sheets = await run_in_threadpool(_workbook_rows, content, fayl.filename or "hujjat.xlsx")
    rows = _active_rows(sheets, ("HUJJAT", "AbiturID", "Sheet1"))
    ids: list[str] = []
    seen: set[str] = set()
    for index, row in enumerate(rows):
        value = row[0] if row else None
        key = _abitur_id_key(value)
        if not key: continue
        if index == 0 and _key(key) in {"abiturid", "abituriyentid", "id"}: continue
        if key in {"1234567", "1234568"}: continue  # shablon namunasi
        if key in seen: continue
        seen.add(key); ids.append(key)
    if not ids: raise HTTPException(status_code=400, detail="Faylda AbiturID topilmadi — A ustunini tekshiring")
    if len(ids) > 20000: raise HTTPException(status_code=400, detail="Bir importda ko'pi bilan 20 000 ta ID")
    conn = p._db(); cur = conn.cursor()
    try:
        _ensure_schema(cur); roles = _require_member(cur, user_id, universitet_id)
        if right == "hujjat" and not _has_any(roles, MARK_DOCUMENT_ROLES): raise HTTPException(status_code=403, detail="Hujjat belgilash huquqi yo'q")
        if right == "baza" and not _has_any(roles, MARK_HEMIS_ROLES): raise HTTPException(status_code=403, detail="Baza belgilash huquqi yo'q")
        scope_sql, scope_params = _student_scope_clause(cur, universitet_id, user_id, roles)
        cur.execute(f"""SELECT qt.id,qt.abitur_id,qt.familiya,qt.ism,qt.ota_ism,qt.hujjat_topshirgan_at,qt.bazaga_kiritilgan_at,
                               qt.saytga_kiritilgan_at,qt.birinchi_kirish_at,qt.user_id,y.nomi yonalish_nomi
                        FROM universitet_qabul_talabalari qt
                        JOIN universitet_yonalishlari y ON y.id=qt.yonalish_id
                        JOIN fakultetlar f ON f.id=y.fakultet_id
                        JOIN kafedralar k ON k.id=y.kafedra_id
                        WHERE qt.universitet_id=%s AND UPPER(REGEXP_REPLACE(COALESCE(qt.abitur_id,''),'\\s','','g'))=ANY(%s) AND {scope_sql}""",
                    [universitet_id, ids, *scope_params])
        found = {_abitur_id_key(r["abitur_id"]): dict(r) for r in cur.fetchall()}
        def already_done(r):
            return (r[column] is not None) if already_sql == "IS NOT NULL" else (r[column] is None)
        def blocked_reason(r):
            # Tartib buzilmasin: hujjat belgisi bazadan oldin; baza belgisi saytga kirgandan keyin olinmaydi.
            if amal == "hujjat_bekor" and r["bazaga_kiritilgan_at"] is not None: return "avval baza belgisini olib tashlang"
            if amal == "baza_belgilash" and r["hujjat_topshirgan_at"] is None: return "hujjat topshirmagan"
            site_entered = r["saytga_kiritilgan_at"] is not None or r["birinchi_kirish_at"] is not None or (r["user_id"] is not None and int(r["user_id"]) >= 0)
            if amal == "baza_bekor" and site_entered: return "talaba saytga kirgan"
            return None
        to_mark, already, blocked = [], [], []
        for k in ids:
            if k not in found: continue
            r = found[k]
            if already_done(r): already.append(r); continue
            reason = blocked_reason(r)
            if reason: blocked.append({**r, "sabab": reason}); continue
            to_mark.append(r)
        missing = [k for k in ids if k not in found]
        def fish(r): return " ".join(x for x in [r["familiya"], r["ism"], r["ota_ism"]] if x)
        marked = 0
        if int(tasdiqlash or 0) == 1 and to_mark:
            stage_sql = {"hujjat_belgilash": "GREATEST(qabul_bosqichi,2)", "hujjat_bekor": "1", "baza_belgilash": "GREATEST(qabul_bosqichi,3)", "baza_bekor": "2"}[amal]
            cur.execute(f"""UPDATE universitet_qabul_talabalari
                              SET {column}={new_value_sql},qabul_bosqichi={stage_sql},yangilangan_at=NOW()
                            WHERE id=ANY(%s)""", ([int(r["id"]) for r in to_mark],))
            marked = cur.rowcount
            _audit(cur, universitet_id, user_id, f"ommaviy_{amal}", "universitet", universitet_id, {"belgilandi": marked, "allaqachon": len(already), "topilmadi": len(missing), "bloklandi": len(blocked)})
        conn.commit()
        return {
            "tasdiqlandi": int(tasdiqlash or 0) == 1,
            "amal": amal, "amal_nomi": action_label,
            "bloklandi": len(blocked),
            "bloklandi_royxat": [{"abitur_id": r["abitur_id"], "fish": fish(r), "sabab": r["sabab"]} for r in blocked[:100]],
            "jami_id": len(ids),
            "belgilanadi": len(to_mark), "belgilandi": marked,
            "allaqachon": len(already), "topilmadi": len(missing),
            "belgilanadi_royxat": [{"abitur_id": r["abitur_id"], "fish": fish(r), "yonalish": r["yonalish_nomi"]} for r in to_mark[:300]],
            "allaqachon_royxat": [{"abitur_id": r["abitur_id"], "fish": fish(r)} for r in already[:100]],
            "topilmadi_royxat": missing[:300],
        }
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


@router.get("/qabul/talaba/{student_id}")
def admission_detail(student_id: int, universitet_id: int, token: Optional[str] = Query(None, include_in_schema=False), authorization: Optional[str] = Header(None)):
    user_id = _uid(token, authorization); p = _p(); conn = p._db(); cur = conn.cursor()
    try:
        _ensure_schema(cur); roles = _require_member(cur, user_id, universitet_id)
        if not _has_any(roles, PRIVATE_ROLES): raise HTTPException(status_code=403, detail="Shaxsiy ma'lumotni ko'rish huquqi yo'q")
        cur.execute("""SELECT qt.*,y.nomi yonalish_nomi,y.fakultet_id,f.nomi fakultet_nomi,k.nomi kafedra_nomi,
            (qt.hujjat_topshirgan_at IS NOT NULL) hujjat_topshirgan,
            (qt.bazaga_kiritilgan_at IS NOT NULL) bazaga_kiritilgan,
            (qt.saytga_kiritilgan_at IS NOT NULL OR qt.birinchi_kirish_at IS NOT NULL OR
             (qt.user_id IS NOT NULL AND qt.user_id>=0)) saytga_kirgan,
            CASE WHEN qt.saytga_kiritilgan_at IS NOT NULL OR qt.birinchi_kirish_at IS NOT NULL OR
                           (qt.user_id IS NOT NULL AND qt.user_id>=0) THEN 'saytga_kirgan'
                 WHEN qt.user_id<0 THEN 'kirish_kodi_tayyor'
                 ELSE 'bazaga_kiritilmagan' END sayt_holati
            FROM universitet_qabul_talabalari qt JOIN universitet_yonalishlari y ON y.id=qt.yonalish_id
            JOIN fakultetlar f ON f.id=y.fakultet_id JOIN kafedralar k ON k.id=y.kafedra_id
            WHERE qt.id=%s AND qt.universitet_id=%s""", (student_id, universitet_id)); row = cur.fetchone()
        if not row: raise HTTPException(status_code=404, detail="Talaba topilmadi")
        if not _student_access_allowed(cur, universitet_id, user_id, roles, dict(row)):
            raise HTTPException(status_code=403, detail="Bu talaba sizning biriktirilgan qamrovingizga tegishli emas")
        _audit(cur, universitet_id, user_id, "talaba_maxfiy_malumoti_korildi", "qabul_talaba", student_id)
        conn.commit(); result = dict(row); result["fish"] = " ".join(x for x in [row["familiya"],row["ism"],row["ota_ism"]] if x)
        names = _role_names(roles)
        if "tyutor" in names and not (names & MARK_DOCUMENT_ROLES):
            result["telefon_mask"] = _mask_phone(result.get("telefon"))
            for hidden in (
                "jshshir", "jshshir_hash", "tugilgan_sana", "pasport_seriya", "pasport_raqam",
                "telefon", "telegram_username", "max_username", "doimiy_region", "doimiy_tuman",
                "maktab_region", "maktab_tuman", "maktab_turi", "maktab_nomi", "attestat",
                "import_batch_id", "qabul_bosqichi", "hujjat_topshirgan_at",
                "bazaga_kiritilgan_at", "saytga_kiritilgan_at", "birinchi_kirish_at", "user_id",
            ):
                result.pop(hidden, None)
            result.pop("hujjat_topshirgan", None)
        return result
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


@router.patch("/qabul/talaba/{student_id}/bosqich")
def update_stage(student_id: int, req: StageUpdate):
    if req.bosqich not in (2, 3, 4): raise HTTPException(status_code=400, detail="Qo'lda faqat hujjat, baza yoki saytga kirish bosqichi belgilanadi")
    p = _p(); actor = p._jwt_tekshir(req.token); conn = p._db(); cur = conn.cursor()
    try:
        _ensure_schema(cur); cur.execute("SELECT * FROM universitet_qabul_talabalari WHERE id=%s FOR UPDATE", (student_id,)); row = cur.fetchone()
        if not row: raise HTTPException(status_code=404, detail="Talaba topilmadi")
        roles = _require_member(cur, actor, row["universitet_id"]); names = _role_names(roles)
        if not _student_access_allowed(cur, row["universitet_id"], actor, roles, dict(row)):
            raise HTTPException(status_code=403, detail="Bu talaba sizga biriktirilmagan")
        if req.bosqich == 2:
            if not (names & MARK_DOCUMENT_ROLES):
                raise HTTPException(status_code=403, detail="Hujjat qabulini belgilash huquqi yo'q")
        else:
            # HEMIS/baza va sayt holatini faqat kafedra mudiri yoki aynan shu
            # qamrovga biriktirilgan tyutor belgilaydi. Admin hujjat qabulini
            # tasdiqlaydi, ammo keyingi bosqichlarni uning nomidan bosmaydi.
            if not (names & MARK_HEMIS_ROLES):
                raise HTTPException(status_code=403, detail="HEMIS va sayt holatini belgilash huquqi yo'q")
            if row["hujjat_topshirgan_at"] is None:
                raise HTTPException(status_code=409, detail="Avval admin hujjat topshirilganini tasdiqlashi kerak")
        if req.bosqich == 2 and row["hujjat_topshirgan_at"] is not None:
            conn.commit(); return {"holat": "avval_belgilangan", "bosqich": 2}
        if req.bosqich == 3 and row["bazaga_kiritilgan_at"] is not None:
            conn.commit(); return {"holat": "avval_belgilangan", "bosqich": 3}
        if req.bosqich == 4 and row["saytga_kiritilgan_at"] is not None:
            conn.commit(); return {"holat": "avval_belgilangan", "bosqich": 4}
        code = None
        if req.bosqich == 4:
            if row["bazaga_kiritilgan_at"] is None:
                raise HTTPException(status_code=409, detail="Avval talabaning bazaga kiritilganini belgilang")
            cur.execute("""UPDATE universitet_qabul_talabalari SET qabul_bosqichi=GREATEST(qabul_bosqichi,4),
                saytga_kiritilgan_at=COALESCE(saytga_kiritilgan_at,NOW()),yangilangan_at=NOW() WHERE id=%s""", (student_id,))
        elif req.bosqich == 3:
            if row["hujjat_topshirgan_at"] is None:
                raise HTTPException(status_code=409, detail="Avval hujjat topshirilganini belgilang")
            code = _create_student_invite(cur, dict(row), actor)
            cur.execute("""UPDATE universitet_qabul_talabalari SET qabul_bosqichi=GREATEST(qabul_bosqichi,3),
                bazaga_kiritilgan_at=COALESCE(bazaga_kiritilgan_at,NOW()),yangilangan_at=NOW() WHERE id=%s""", (student_id,))
        else:
            cur.execute("""UPDATE universitet_qabul_talabalari SET qabul_bosqichi=GREATEST(qabul_bosqichi,2),
                hujjat_topshirgan_at=COALESCE(hujjat_topshirgan_at,NOW()),yangilangan_at=NOW() WHERE id=%s""", (student_id,))
        _audit(cur, row["universitet_id"], actor, "qabul_bosqichi", "qabul_talaba", student_id, {"bosqich": req.bosqich})
        visible_code = code if names & PASSWORD_VIEW_ROLES else None
        conn.commit(); return {"holat": "yangilandi", "bosqich": req.bosqich, "kirish_kodi": visible_code, "kod_muddati": "2 oy" if visible_code else None}
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


@router.patch("/qabul/talaba/{student_id}/bosqich_bekor")
def revert_stage(student_id: int, req: StageUpdate):
    """Xato belgilangan bosqichni qaytaradi: 2 — hujjat, 3 — baza/HEMIS.

    Hujjat belgisi faqat baza belgisi yo'q bo'lsa qaytariladi; baza belgisi
    talaba saytga kirmagan bo'lsa qaytariladi. Saytga kirgan talaba qaytarilmaydi.
    """
    if req.bosqich not in (2, 3): raise HTTPException(status_code=400, detail="Faqat hujjat (2) yoki baza (3) belgisi bekor qilinadi")
    p = _p(); actor = p._jwt_tekshir(req.token); conn = p._db(); cur = conn.cursor()
    try:
        _ensure_schema(cur); cur.execute("SELECT * FROM universitet_qabul_talabalari WHERE id=%s FOR UPDATE", (student_id,)); row = cur.fetchone()
        if not row: raise HTTPException(status_code=404, detail="Talaba topilmadi")
        roles = _require_member(cur, actor, row["universitet_id"]); names = _role_names(roles)
        if not _student_access_allowed(cur, row["universitet_id"], actor, roles, dict(row)):
            raise HTTPException(status_code=403, detail="Bu talaba sizga biriktirilmagan")
        site_entered = row["saytga_kiritilgan_at"] is not None or row["birinchi_kirish_at"] is not None or (row["user_id"] is not None and int(row["user_id"]) >= 0)
        if req.bosqich == 2:
            if not (names & MARK_DOCUMENT_ROLES):
                raise HTTPException(status_code=403, detail="Hujjat belgisini bekor qilish huquqi yo'q")
            if row["bazaga_kiritilgan_at"] is not None:
                raise HTTPException(status_code=409, detail="Avval “Bazaga kiritilgan” belgisini bekor qiling")
            if row["hujjat_topshirgan_at"] is None:
                conn.commit(); return {"holat": "belgilanmagan", "bosqich": 2}
            cur.execute("""UPDATE universitet_qabul_talabalari
                              SET hujjat_topshirgan_at=NULL,qabul_bosqichi=1,yangilangan_at=NOW()
                            WHERE id=%s""", (student_id,))
        else:
            if not (names & MARK_HEMIS_ROLES):
                raise HTTPException(status_code=403, detail="Baza belgisini bekor qilish huquqi yo'q")
            if site_entered:
                raise HTTPException(status_code=409, detail="Talaba saytga kirgan; baza belgisi qaytarilmaydi")
            if row["bazaga_kiritilgan_at"] is None:
                conn.commit(); return {"holat": "belgilanmagan", "bosqich": 3}
            cur.execute("""UPDATE universitet_qabul_talabalari
                              SET bazaga_kiritilgan_at=NULL,qabul_bosqichi=2,yangilangan_at=NOW()
                            WHERE id=%s""", (student_id,))
        _audit(cur, row["universitet_id"], actor, "qabul_bosqichi_bekor", "qabul_talaba", student_id, {"bosqich": req.bosqich})
        conn.commit(); return {"holat": "bekor_qilindi", "bosqich": req.bosqich}
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


@router.patch("/qabul/talaba/{student_id}/izoh")
def update_contact_note(student_id: int, req: ContactNoteUpdate):
    """Telefon suhbati izohi. Bo'sh matn yuborilsa izoh o'chiriladi."""
    note = re.sub(r"\s+", " ", str(req.izoh or "")).strip()[:600]
    p = _p(); actor = p._jwt_tekshir(req.token); conn = p._db(); cur = conn.cursor()
    try:
        _ensure_schema(cur); cur.execute("SELECT * FROM universitet_qabul_talabalari WHERE id=%s FOR UPDATE", (student_id,)); row = cur.fetchone()
        if not row: raise HTTPException(status_code=404, detail="Talaba topilmadi")
        roles = _require_member(cur, actor, row["universitet_id"])
        if not _has_any(roles, PRIVATE_ROLES) or not _student_access_allowed(cur, row["universitet_id"], actor, roles, dict(row)):
            raise HTTPException(status_code=403, detail="Bu talabaga izoh yozish huquqi yo'q")
        if note:
            cur.execute("""UPDATE universitet_qabul_talabalari
                              SET aloqa_izohi=%s,aloqa_izohi_at=NOW(),aloqa_izohi_by=%s,yangilangan_at=NOW()
                            WHERE id=%s""", (note, actor, student_id))
        else:
            cur.execute("""UPDATE universitet_qabul_talabalari
                              SET aloqa_izohi=NULL,aloqa_izohi_at=NULL,aloqa_izohi_by=NULL,yangilangan_at=NOW()
                            WHERE id=%s""", (student_id,))
        _audit(cur, row["universitet_id"], actor, "qabul_aloqa_izohi", "qabul_talaba", student_id, {"izoh": note})
        conn.commit(); return {"holat": "saqlandi" if note else "ochirildi", "aloqa_izohi": note or None}
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


@router.post("/qabul/talaba/{student_id}/taklif")
def invite_student(student_id: int, req: InviteSend):
    p = _p(); actor = p._jwt_tekshir(req.token); conn = p._db(); cur = conn.cursor()
    try:
        _ensure_schema(cur); cur.execute("SELECT * FROM universitet_qabul_talabalari WHERE id=%s FOR UPDATE", (student_id,)); row = cur.fetchone()
        if not row: raise HTTPException(status_code=404, detail="Talaba topilmadi")
        roles = _require_member(cur, actor, row["universitet_id"])
        if not _has_any(roles, PASSWORD_VIEW_ROLES) or not _student_access_allowed(cur, row["universitet_id"], actor, roles, dict(row)):
            raise HTTPException(status_code=403, detail="Taklif yuborish huquqi yo'q")
        if row["bazaga_kiritilgan_at"] is None: raise HTTPException(status_code=409, detail="Talaba hali bazaga kiritilgan deb belgilanmagan")
        if row["saytga_kiritilgan_at"] is not None or row["birinchi_kirish_at"] is not None or (row["user_id"] is not None and int(row["user_id"]) >= 0):
            raise HTTPException(status_code=409, detail="Talaba saytga kirgan; kirish kodi qayta berilmaydi")
        code = _create_student_invite(cur, dict(row), actor)
        base = getattr(p, "FRONTEND_URL", "").rstrip("/")
        link = f"{base}/?kirish_kodi={code}"
        text = f"Institut ta'lim platformasiga kirish kodi: {code}. Kod 2 oy amal qiladi. Kirish: {link}"
        sent = False
        if req.kanal == "sms":
            if not row["telefon"]: raise HTTPException(status_code=400, detail="Talabada telefon raqami yo'q")
            sent = bool(p._sms_yubor(row["telefon"], text))
        _audit(cur, row["universitet_id"], actor, "talaba_taklif_yaratildi", "qabul_talaba", student_id, {"kanal": req.kanal, "sent": sent})
        conn.commit(); return {"holat": "yuborildi" if sent else "tayyor", "kirish_kodi": code, "havola": link, "sms_matni": text, "sms_yuborildi": sent}
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


@router.get("/qabul/talaba/{student_id}/kirish_kodi")
def reveal_student_password(student_id: int, token: Optional[str] = Query(None, include_in_schema=False), authorization: Optional[str] = Header(None)):
    actor = _uid(token, authorization); p = _p(); conn = p._db(); cur = conn.cursor()
    try:
        _ensure_schema(cur); cur.execute("SELECT * FROM universitet_qabul_talabalari WHERE id=%s FOR UPDATE", (student_id,)); row = cur.fetchone()
        if not row: raise HTTPException(status_code=404, detail="Talaba topilmadi")
        roles = _require_member(cur, actor, row["universitet_id"])
        if not _has_any(roles, PASSWORD_VIEW_ROLES):
            raise HTTPException(status_code=403, detail="Talaba kirish kodini ko'rish huquqi yo'q")
        if not _student_access_allowed(cur, row["universitet_id"], actor, roles, dict(row)):
            raise HTTPException(status_code=403, detail="Bu talaba sizning qamrovingizga tegishli emas")
        if row["saytga_kiritilgan_at"] is not None or row["birinchi_kirish_at"] is not None or (row["user_id"] is not None and int(row["user_id"]) >= 0):
            raise HTTPException(status_code=409, detail="Talaba saytga birinchi marta kirgan; xavfsizlik uchun parol endi ko'rsatilmaydi")
        code = _create_student_invite(cur, dict(row), actor)
        _audit(cur, row["universitet_id"], actor, "talaba_kirish_kodi_korildi", "qabul_talaba", student_id)
        conn.commit(); return {"fish": " ".join(x for x in [row["familiya"], row["ism"], row["ota_ism"]] if x),
                               "kirish_kodi": code, "kod_muddati": "2 oy", "birinchi_kirishgacha_korinadi": True}
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


@router.post("/kirish_kodi_qabul")
def redeem_code(req: RedeemCode):
    p = _p(); user_id = p._jwt_tekshir(req.token); plain, stored = p._xodim_kod_variantlari(req.kirish_kodi)
    conn = p._db(); cur = conn.cursor()
    try:
        _ensure_schema(cur); p._xodim_kod_jadvali(cur)
        cur.execute("""SELECT tk.*,xk.ishlatildi,(xk.yaratildi>NOW()-INTERVAL '2 months') hali_yangi
            FROM universitet_taklif_kodlari tk JOIN xodim_kod xk ON xk.kod=tk.kod_hash
            WHERE tk.kod_hash IN (%s,%s) FOR UPDATE""", (stored, plain)); invite = cur.fetchone()
        if not invite: raise HTTPException(status_code=400, detail="Kirish kodi noto'g'ri")
        if invite["ishlatildi"]: raise HTTPException(status_code=409, detail="Kirish kodi ishlatilgan")
        if not invite["hali_yangi"]: raise HTTPException(status_code=410, detail="Kirish kodi muddati tugagan")
        placeholder = invite["placeholder_user_id"]
        cur.execute("SELECT universitet_id FROM users WHERE user_id=%s FOR UPDATE", (user_id,))
        account = cur.fetchone()
        if not account: raise HTTPException(status_code=404, detail="Foydalanuvchi akkaunti topilmadi")
        if invite["xodim_rol_id"]:
            cur.execute("UPDATE universitet_xodim_rollari SET user_id=%s WHERE id=%s", (user_id, invite["xodim_rol_id"]))
        if invite["qabul_talaba_id"]:
            cur.execute("""UPDATE universitet_qabul_talabalari SET user_id=%s,
                bazaga_kiritilgan_at=COALESCE(bazaga_kiritilgan_at,NOW()),
                saytga_kiritilgan_at=NOW(),birinchi_kirish_at=NOW(),qabul_bosqichi=4,
                yangilangan_at=NOW() WHERE id=%s""", (user_id, invite["qabul_talaba_id"]))
        p._telefon_jadvallari(cur)
        cur.execute("UPDATE telefon_hisob SET user_id=%s WHERE user_id=%s", (user_id, placeholder))
        cur.execute("""UPDATE users SET
            universitet_id=COALESCE(universitet_id,%s),
            lavozim=COALESCE(lavozim,%s)
            WHERE user_id=%s""",
            (invite["universitet_id"], "talaba" if invite["turi"] == "talaba" else "institut_xodimi", user_id))
        cur.execute("UPDATE xodim_kod SET ishlatildi=TRUE WHERE kod=%s", (invite["kod_hash"],))
        cur.execute("UPDATE universitet_taklif_kodlari SET ishlatildi_at=NOW(),kod_shifr=NULL WHERE id=%s", (invite["id"],))
        _audit(cur, invite["universitet_id"], user_id, "kirish_kodi_qabul", invite["turi"], invite["qabul_talaba_id"] or invite["xodim_rol_id"])
        conn.commit(); return {"holat": "ulandi", "universitet_id": invite["universitet_id"], "turi": invite["turi"]}
    except Exception:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


def _structure_archive_summary(cur, universitet_id: int, kind: str, object_id: int) -> dict[str, Any]:
    if kind == "fakultet":
        cur.execute("SELECT id,nomi FROM fakultetlar WHERE id=%s AND universitet_id=%s AND faol=TRUE", (object_id, universitet_id)); row = cur.fetchone()
        if not row: raise HTTPException(status_code=404, detail="Fakultet topilmadi")
        cur.execute("""SELECT COUNT(DISTINCT k.id) kafedra,COUNT(DISTINCT y.id) yonalish,
            COUNT(DISTINCT qt.id) talaba,COUNT(DISTINCT xr.id) xodim
            FROM fakultetlar f LEFT JOIN kafedralar k ON k.fakultet_id=f.id
            LEFT JOIN universitet_yonalishlari y ON y.fakultet_id=f.id
            LEFT JOIN universitet_qabul_talabalari qt ON qt.yonalish_id=y.id
            LEFT JOIN universitet_xodim_rollari xr ON xr.fakultet_id=f.id AND xr.faol=TRUE WHERE f.id=%s""", (object_id,))
    elif kind == "kafedra":
        cur.execute("""SELECT k.id,k.nomi FROM kafedralar k JOIN fakultetlar f ON f.id=k.fakultet_id
            WHERE k.id=%s AND f.universitet_id=%s AND k.faol=TRUE""", (object_id, universitet_id)); row = cur.fetchone()
        if not row: raise HTTPException(status_code=404, detail="Kafedra topilmadi")
        cur.execute("""SELECT 1 kafedra,COUNT(DISTINCT y.id) yonalish,COUNT(DISTINCT qt.id) talaba,
            COUNT(DISTINCT xr.id) xodim FROM kafedralar k
            LEFT JOIN universitet_yonalishlari y ON y.kafedra_id=k.id
            LEFT JOIN universitet_qabul_talabalari qt ON qt.yonalish_id=y.id
            LEFT JOIN universitet_xodim_rollari xr ON xr.kafedra_id=k.id AND xr.faol=TRUE WHERE k.id=%s""", (object_id,))
    elif kind == "yonalish":
        cur.execute("SELECT id,nomi FROM universitet_yonalishlari WHERE id=%s AND universitet_id=%s AND faol=TRUE", (object_id, universitet_id)); row = cur.fetchone()
        if not row: raise HTTPException(status_code=404, detail="Yo‘nalish topilmadi")
        cur.execute("""SELECT 0 kafedra,1 yonalish,COUNT(DISTINCT qt.id) talaba,
            COUNT(DISTINCT xr.id) xodim FROM universitet_yonalishlari y
            LEFT JOIN universitet_qabul_talabalari qt ON qt.yonalish_id=y.id
            LEFT JOIN universitet_xodim_rollari xr ON xr.yonalish_id=y.id AND xr.faol=TRUE WHERE y.id=%s""", (object_id,))
    else:
        raise HTTPException(status_code=400, detail="Obyekt turi noto‘g‘ri")
    counts = dict(cur.fetchone()); return {"obyekt_turi": kind, "obyekt_id": object_id, "nomi": row["nomi"], "hisoblar": counts, "saqlanish_muddati": "1 yil"}


@router.get("/tuzilma/arxiv_preview")
def structure_archive_preview(universitet_id: int, obyekt_turi: str, obyekt_id: int,
                              token: Optional[str] = Query(None, include_in_schema=False), authorization: Optional[str] = Header(None)):
    actor = _uid(token, authorization); p = _p(); conn = p._db(); cur = conn.cursor()
    try:
        _ensure_schema(cur); roles = _require_member(cur, actor, universitet_id)
        if not _has_any(roles, MANAGE_STRUCTURE_ROLES): raise HTTPException(status_code=403, detail="Tuzilmani arxivlash huquqi yo‘q")
        return _structure_archive_summary(cur, universitet_id, obyekt_turi, obyekt_id)
    finally: cur.close(); conn.close()


@router.post("/tuzilma/arxivlash")
def structure_archive_commit(req: StructureArchiveCommit):
    if not req.tasdiq: raise HTTPException(status_code=400, detail="Arxivlash tasdiqlanmadi")
    p = _p(); actor = p._jwt_tekshir(req.token); conn = p._db(); cur = conn.cursor()
    try:
        _ensure_schema(cur); roles = _require_member(cur, actor, req.universitet_id)
        if not _has_any(roles, MANAGE_STRUCTURE_ROLES): raise HTTPException(status_code=403, detail="Tuzilmani arxivlash huquqi yo‘q")
        summary = _structure_archive_summary(cur, req.universitet_id, req.obyekt_turi, req.obyekt_id)
        if req.obyekt_turi == "fakultet":
            cur.execute("UPDATE fakultetlar SET faol=FALSE,arxiv_at=NOW(),arxiv_until=NOW()+INTERVAL '1 year' WHERE id=%s", (req.obyekt_id,))
        elif req.obyekt_turi == "kafedra":
            cur.execute("UPDATE kafedralar SET faol=FALSE,arxiv_at=NOW(),arxiv_until=NOW()+INTERVAL '1 year' WHERE id=%s", (req.obyekt_id,))
        else:
            cur.execute("UPDATE universitet_yonalishlari SET faol=FALSE,arxiv_at=NOW(),arxiv_until=NOW()+INTERVAL '1 year' WHERE id=%s", (req.obyekt_id,))
        cur.execute("""INSERT INTO universitet_tuzilma_arxivi(universitet_id,obyekt_turi,obyekt_id,nomi,hisoblar,arxivlagan_by)
            VALUES(%s,%s,%s,%s,%s::jsonb,%s)""", (req.universitet_id,req.obyekt_turi,req.obyekt_id,summary["nomi"],json.dumps(summary["hisoblar"]),actor))
        _audit(cur, req.universitet_id, actor, "tuzilma_arxivlandi", req.obyekt_turi, req.obyekt_id, summary["hisoblar"])
        conn.commit(); return {"holat": "arxivlandi", **summary}
    except Exception: conn.rollback(); raise
    finally: cur.close(); conn.close()


@router.get("/tuzilma/arxiv")
def structure_archive_list(universitet_id: int, token: Optional[str] = Query(None, include_in_schema=False), authorization: Optional[str] = Header(None)):
    actor = _uid(token, authorization); p = _p(); conn = p._db(); cur = conn.cursor()
    try:
        _ensure_schema(cur); roles = _require_member(cur, actor, universitet_id)
        if not _has_any(roles, MANAGE_STRUCTURE_ROLES): raise HTTPException(status_code=403, detail="Arxivni ko‘rish huquqi yo‘q")
        cur.execute("SELECT id,obyekt_turi,obyekt_id,nomi,hisoblar,arxiv_at,arxiv_until FROM universitet_tuzilma_arxivi WHERE universitet_id=%s AND faol=TRUE ORDER BY arxiv_at DESC", (universitet_id,))
        return {"arxiv": [dict(row) for row in cur.fetchall()]}
    finally: cur.close(); conn.close()


@router.post("/tuzilma/arxiv/{archive_id}/tiklash")
def structure_archive_restore(archive_id: int, req: StructureRestore):
    p = _p(); actor = p._jwt_tekshir(req.token); conn = p._db(); cur = conn.cursor()
    try:
        _ensure_schema(cur); roles = _require_member(cur, actor, req.universitet_id)
        if not _has_any(roles, MANAGE_STRUCTURE_ROLES): raise HTTPException(status_code=403, detail="Arxivdan tiklash huquqi yo‘q")
        cur.execute("SELECT * FROM universitet_tuzilma_arxivi WHERE id=%s AND universitet_id=%s AND faol=TRUE FOR UPDATE", (archive_id,req.universitet_id)); item=cur.fetchone()
        if not item: raise HTTPException(status_code=404, detail="Arxiv yozuvi topilmadi")
        table={"fakultet":"fakultetlar","kafedra":"kafedralar","yonalish":"universitet_yonalishlari"}[item["obyekt_turi"]]
        cur.execute(f"UPDATE {table} SET faol=TRUE,arxiv_at=NULL,arxiv_until=NULL WHERE id=%s", (item["obyekt_id"],))
        # Bola obyekt tiklansa, ko‘rinishi uchun uning ota zanjiri ham ochiladi.
        if item["obyekt_turi"] == "kafedra":
            cur.execute("""UPDATE fakultetlar f SET faol=TRUE,arxiv_at=NULL,arxiv_until=NULL
                FROM kafedralar k WHERE k.id=%s AND f.id=k.fakultet_id""", (item["obyekt_id"],))
        elif item["obyekt_turi"] == "yonalish":
            cur.execute("""UPDATE kafedralar k SET faol=TRUE,arxiv_at=NULL,arxiv_until=NULL
                FROM universitet_yonalishlari y WHERE y.id=%s AND k.id=y.kafedra_id""", (item["obyekt_id"],))
            cur.execute("""UPDATE fakultetlar f SET faol=TRUE,arxiv_at=NULL,arxiv_until=NULL
                FROM universitet_yonalishlari y WHERE y.id=%s AND f.id=y.fakultet_id""", (item["obyekt_id"],))
        cur.execute("UPDATE universitet_tuzilma_arxivi SET faol=FALSE,tiklangan_at=NOW() WHERE id=%s", (archive_id,))
        _audit(cur,req.universitet_id,actor,"tuzilma_arxivdan_tiklandi",item["obyekt_turi"],item["obyekt_id"])
        conn.commit(); return {"holat":"tiklandi","nomi":item["nomi"]}
    except Exception: conn.rollback(); raise
    finally: cur.close(); conn.close()


@router.get("/talaba/yonalish_katalogi")
def student_directory(universitet_id: int, token: Optional[str] = Query(None, include_in_schema=False), authorization: Optional[str] = Header(None)):
    user_id = _uid(token, authorization); p = _p(); conn = p._db(); cur = conn.cursor()
    try:
        _ensure_schema(cur); roles = _require_member(cur, user_id, universitet_id)
        student_role = next((r for r in roles if r["rol"] == "talaba"), None)
        if not student_role: raise HTTPException(status_code=403, detail="Bu bo'lim talaba uchun")
        program_id = int(student_role["yonalish_id"])
        cur.execute("""SELECT y.id,y.nomi,y.daraja,f.id fakultet_id,f.nomi fakultet_nomi,k.nomi kafedra_nomi
            FROM universitet_yonalishlari y JOIN fakultetlar f ON f.id=y.fakultet_id JOIN kafedralar k ON k.id=y.kafedra_id WHERE y.id=%s""", (program_id,)); program = cur.fetchone()
        cur.execute("""SELECT qt.id,qt.familiya,qt.ism,qt.ota_ism FROM universitet_qabul_talabalari qt
            WHERE qt.universitet_id=%s AND qt.yonalish_id=%s
              AND (qt.birinchi_kirish_at IS NOT NULL OR (qt.user_id IS NOT NULL AND qt.user_id>=0))
            ORDER BY qt.familiya,qt.ism""", (universitet_id, program_id))
        students = [{"id": r["id"], "fish": " ".join(x for x in [r["familiya"],r["ism"],r["ota_ism"]] if x)} for r in cur.fetchall()]
        cur.execute("""SELECT DISTINCT xr.rol,u.full_name FROM universitet_xodim_rollari xr JOIN users u ON u.user_id=xr.user_id
            WHERE xr.universitet_id=%s AND xr.faol=TRUE AND (
              xr.rol IN ('rektor','prorektor','institut_admin') OR xr.fakultet_id=%s OR
              (xr.rol='tyutor' AND EXISTS(SELECT 1 FROM universitet_tyutor_yonalishlari ty WHERE ty.tyutor_user_id=xr.user_id AND ty.yonalish_id=%s AND ty.faol=TRUE)))
            ORDER BY xr.rol,u.full_name""", (universitet_id, program["fakultet_id"], program_id))
        staff = [{"rol": r["rol"], "lavozim_nomi": ROLE_LABELS.get(r["rol"],r["rol"]), "fish": r["full_name"]} for r in cur.fetchall()]
        return {"yonalish": program, "talabalar": students, "masullar": staff}
    finally:
        cur.close(); conn.close()


@router.get("/tyutor/yetarlilik")
def tutor_capacity(universitet_id: int, token: Optional[str] = Query(None, include_in_schema=False), authorization: Optional[str] = Header(None)):
    user_id = _uid(token, authorization); p = _p(); conn = p._db(); cur = conn.cursor()
    try:
        _ensure_schema(cur); roles = _require_member(cur, user_id, universitet_id)
        scope = _scope_program_ids_for_user(cur, universitet_id, user_id, roles)
        where = "y.universitet_id=%s AND y.faol=TRUE"; params: list[Any] = [universitet_id]
        if scope is not None:
            where += " AND y.id=ANY(%s)"; params.append(list(scope))
        cur.execute(f"""SELECT y.id,y.nomi,COUNT(DISTINCT qt.id) FILTER(WHERE qt.talim_shakli='Kunduzgi') kunduzgi_1kurs,
            COUNT(DISTINCT ty.tyutor_user_id) FILTER(WHERE ty.faol=TRUE) tyutor_soni
            FROM universitet_yonalishlari y
            LEFT JOIN universitet_qabul_talabalari qt ON qt.yonalish_id=y.id
            LEFT JOIN universitet_tyutor_yonalishlari ty ON ty.yonalish_id=y.id
            WHERE {where} GROUP BY y.id,y.nomi ORDER BY y.nomi""", params)
        rows=[]
        for r in cur.fetchall():
            needed = math.ceil(int(r["kunduzgi_1kurs"] or 0)/150) if r["kunduzgi_1kurs"] else 0
            item=dict(r); item["tavsiya_etilgan_minimum"] = needed; item["yetarli"] = int(r["tyutor_soni"] or 0) >= needed; rows.append(item)
        return {"yonalishlar": rows, "mezon": "Kunduzgi 1–3-kursning har 120–150 talabasi uchun 1 tyutor"}
    finally:
        cur.close(); conn.close()
